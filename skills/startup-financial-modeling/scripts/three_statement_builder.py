"""
three_statement_builder.py — SaaS Series A scaffold for 17-sheet xlsx 財務モデル

Source of truth:
  - references/06_three_statement.md §2-4 (IS/BS/CFS/WC/Debt full row layouts)
  - references/06_three_statement.md §12 (Sanity Checks)
  - references/02_saas_metrics.md §2-3 (ARR / MRR / NRR canonical definitions)
  - references/15_input_schema.md §4-5 (input schema, SaaS Series A defaults)
  - references/_terminology.md §1 (IB Color), §3 (Sheet naming), §6 (SaaS canonical)

Scope:
  SaaS Series A monthly model (36-month forecast). Phase 1 of Phase 6 build_model.
  - Implements:    00_Cover, 01_Assumptions, 02_Revenue, 03_OpEx,
                   04_IS, 05_BS (incl. WC sub-section), 06_CFS, 07_Debt,
                   11_KPI_Dashboard
  - Skeleton-only: 08_CapTable, 09_DCF (incl. Sensitivity sub-section),
                   10_Comps, 12_SanityChecks, 13_IC_Memo

Other 業態 (Marketplace / Hardware / Bio / Fintech etc.) は Phase 6 phase 2 で拡張。

Conventions (advisor-confirmed):
  * Initial BS at t=0: Cash = starting_cash; Debt = starting_debt; AR/AP/DR = 0.
    APIC plug = Cash − Debt. RE_0 = 0. Operating accounts build from month-1 driver.
  * Monthly WC formulas use `× days / 30` (per 06 §3.1 line 266).
  * Term-loan interest uses Beginning Balance only (06 §4.2 "no-iteration" alt).
  * Tax = MAX(0, PBT) × ETR (NOL not modeled in scaffold).
  * Revolver as Cash/Debt plug; min cash = 0 in scaffold.
  * ARR_t = starting_arr × (1 + growth)^t ; MRR_t = ARR_t / 12 ; Revenue_t = MRR_t.

License: internal (Act / startup-financial-modeling skill)
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import Iterable

# Ensure ib_format.py (sibling) is importable when run as CLI.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

import ib_format as ib

# ============================================================================
# Constants
# ============================================================================

DEFAULT_FORECAST_MONTHS = 36
DEFAULT_START_DATE = "2026-04"  # Japan FY start

# Column layout for time-series sheets (D..D+35 = months 1..36)
LABEL_COL = 2          # B
UNIT_COL = 3           # C
PERIOD_START_COL = 4   # D


# ============================================================================
# Input dataclass
# ============================================================================


@dataclass
class SaaSInput:
    """SaaS Series A model input. canonical: 15_input_schema §4-5."""

    company_name: str
    reporting_currency: str = "JPY"          # JPY / USD
    fiscal_year_end_month: int = 3           # Japan default
    accounting_basis: str = "j_gaap"         # j_gaap / ifrs / us_gaap
    forecast_period_months: int = DEFAULT_FORECAST_MONTHS
    model_start_date: str = DEFAULT_START_DATE  # YYYY-MM (month 1)

    # Revenue drivers (canonical: 02 §2)
    starting_arr_money_m: float = 240.0      # ¥M or $M (ARR at t=0)
    monthly_growth_rate: float = 0.08        # 8% MoM
    nrr: float = 1.15                        # Net Revenue Retention (display only)
    grr: float = 0.92                        # Gross Revenue Retention (display only)
    gross_margin_pct: float = 0.78

    # Cost drivers
    starting_employees: int = 50
    fte_breakdown: dict[str, int] = field(
        default_factory=lambda: {
            "engineering": 18,
            "sales": 12,
            "customer_success": 8,
            "g_and_a": 12,
        }
    )
    flc_per_fte_money_m: float = 14.0        # Fully-Loaded Cost (年間 ¥M / FTE)
    sm_pct_of_revenue: float = 0.40
    rd_pct_of_revenue: float = 0.25
    ga_pct_of_revenue: float = 0.15

    # Working capital (06 §3)
    dso_days: float = 45.0
    dpo_days: float = 30.0
    deferred_rev_pct_of_arr: float = 0.10    # Annual prepay component → DR

    # Cap structure (簡易 — 04b 詳細は cap_table_builder で別途)
    starting_cash_money_m: float = 500.0
    starting_debt_money_m: float = 0.0

    # Tax (07 §2.1)
    effective_tax_rate: float = 0.3152       # 日本大企業 2026

    # Stage / business
    stage: str = "series_a"
    business_model: str = "saas"

    # CapEx & D&A (small for SaaS scaffold)
    capex_pct_revenue: float = 0.02
    depreciation_useful_life_months: int = 60   # 5 yr

    # Debt (term loan beginning balance interest)
    debt_interest_rate_annual: float = 0.04


# ============================================================================
# Cell-reference helpers
# ============================================================================


def sref(sheet: str, cell: str) -> str:
    """Return a single-quoted sheet-cell reference safe for digit-prefix sheets.

    Example:
        sref('01_Assumptions', '$C$5')  ->  "'01_Assumptions'!$C$5"
    """
    return f"'{sheet}'!{cell}"


def srange(sheet: str, start: str, end: str) -> str:
    """Return a single-quoted sheet-range reference."""
    return f"'{sheet}'!{start}:{end}"


def col_letter(month_index: int) -> str:
    """0-based month index → column letter (D for month 1, E for month 2, ...).

    Note: month_index here is 1-based externally; pass `m=1` for month 1 (col D).
    """
    return get_column_letter(PERIOD_START_COL + month_index - 1)


def period_cell(month: int, row: int) -> str:
    """A1-style cell address for (month, row). month=1 → col D."""
    return f"{col_letter(month)}{row}"


# ============================================================================
# Public API
# ============================================================================


def get_default_input() -> SaaSInput:
    """Demo input matching evals.json case 1 (saas_series_a_3yr_model)."""
    return SaaSInput(
        company_name="Demo SaaS Co.",
        reporting_currency="JPY",
        fiscal_year_end_month=3,
        accounting_basis="j_gaap",
        forecast_period_months=36,
        model_start_date="2026-04",
        starting_arr_money_m=240.0,    # MRR ¥20M × 12
        monthly_growth_rate=0.08,
        nrr=1.15,
        grr=0.92,
        gross_margin_pct=0.78,
        starting_employees=50,
        fte_breakdown={
            "engineering": 18,
            "sales": 12,
            "customer_success": 8,
            "g_and_a": 12,
        },
        flc_per_fte_money_m=14.0,
        sm_pct_of_revenue=0.40,
        rd_pct_of_revenue=0.25,
        ga_pct_of_revenue=0.15,
        dso_days=45.0,
        dpo_days=30.0,
        deferred_rev_pct_of_arr=0.10,
        starting_cash_money_m=500.0,
        starting_debt_money_m=0.0,
        effective_tax_rate=0.3152,
        stage="series_a",
        business_model="saas",
        capex_pct_revenue=0.02,
        depreciation_useful_life_months=60,
        debt_interest_rate_annual=0.04,
    )


def build_saas_series_a_model(
    input_data: SaaSInput,
    output_path: str | None = None,
) -> Workbook:
    """Build SaaS Series A 14-sheet xlsx (Phase 6 Stage A — collapsed from 17).

    Args:
        input_data: SaaSInput with hard-input drivers.
        output_path: If provided, save() the workbook to this path.

    Returns:
        openpyxl Workbook ready for further edits or save().
    """
    wb = build_workbook(input_data)
    if output_path is not None:
        wb.save(output_path)
    return wb


def build_workbook(input_data: SaaSInput) -> Workbook:
    """Generate full 14-sheet workbook in canonical order (Phase 6 Stage A).

    Implemented sheets: 00..07, 11.
    Skeleton sheets:    08..10, 12..13 (header-only placeholders).
    """
    wb = Workbook()
    # Default workbook ships with one "Sheet" — replace with 00_Cover.
    default = wb.active
    wb.remove(default)

    # Create all 17 sheets in canonical order.
    sheets: dict[str, Worksheet] = {}
    for name in ib.CANONICAL_SHEET_ORDER:
        sheets[name] = wb.create_sheet(name)

    inp = input_data

    # Implemented sheets (order matters: drivers/assumptions populated first,
    # so downstream sheets can reference cells via $C$N anchors).
    _build_cover(sheets["00_Cover"], inp)
    _build_assumptions(sheets["01_Assumptions"], inp)
    _build_kpi_dashboard(sheets["11_KPI_Dashboard"], inp)
    _build_revenue(sheets["02_Revenue"], inp)
    _build_opex(sheets["03_OpEx"], inp)
    _build_debt(sheets["07_Debt"], inp)
    _build_is(sheets["04_IS"], inp)
    _build_bs(sheets["05_BS"], inp)
    _build_wc(sheets["05_BS"], inp)  # WC schedule embeds at row 60+ on 05_BS (after BS rows)
    _build_cfs(sheets["06_CFS"], inp)

    # Skeleton sheets (Phase 6 phase 2)
    _build_skeleton_sheets(wb, inp)

    # Named ranges (top input cells → friendly names for formula authoring).
    _register_named_ranges(wb)

    # Phase 6 Wave 3: canonical cross-sheet named ranges (references/_named_ranges.md §2).
    # Registers row-range names like Rev_Total / IS_NI / BS_Cash for cross-sheet
    # `=INDEX(<Name>, m)` references — 行/列 insertion 耐性を担保する。
    _register_canonical_named_ranges(wb, inp.forecast_period_months)

    # Phase 6 補強 §2.X: canonical 6-role × 17-sheet tab color (one pass).
    # Overrides any role-based set_tab_color calls inside individual builders
    # so the workbook is guaranteed D10-compliant at save time.
    ib.apply_canonical_tab_colors(wb)

    # First-sheet selection
    wb.active = wb.index(sheets["00_Cover"])

    return wb


# ============================================================================
# Sheet builders — to be filled in by subsequent passes
# ============================================================================


def _build_cover(ws: Worksheet, inp: SaaSInput) -> None:
    """00_Cover — title + project meta. Uses ib.write_cover."""
    title = inp.company_name
    subtitle = f"SaaS Series A — {inp.forecast_period_months}-month Model"
    ib.write_cover(
        ws,
        title=title,
        subtitle=subtitle,
        project_code="STARTUP-FIN-MODEL",
        confidential=True,
        author="Act / startup-financial-modeling skill",
        version="1.0",
        date_str=inp.model_start_date,
    )
    ib.set_tab_color(ws, "cover")


def _build_assumptions(ws: Worksheet, inp: SaaSInput) -> None:
    """01_Assumptions — hard-input cells (blue). All downstream formulas
    reference $C$N anchors here. Unit column (C) holds the value, with
    label (B) and unit hint (D).
    """
    ib.set_tab_color(ws, "assumptions")
    ib.setup_sheet_layout(ws, n_periods=4, has_unit_col=False, freeze_at="A4")

    # Column layout for this sheet only:
    # A: margin   B: label   C: value   D: unit
    # Phase 6 Fix #2: A 列幅は ib.COL_MARGIN_WIDTH (=20.0) を使う。元 2.0 → 20.0。
    ws.column_dimensions["A"].width = ib.COL_MARGIN_WIDTH
    ws.column_dimensions["B"].width = 44.0
    ws.column_dimensions["C"].width = 16.0
    ws.column_dimensions["D"].width = 14.0
    ws.column_dimensions["E"].width = 60.0  # comment column

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=5,
        label="01 — Assumptions (Hard Inputs)",
    )

    rows: list[tuple[str, object, str, str, str]] = [
        # (label, value, unit, fmt, comment)  fmt ∈ money/pct/int/text
        ("Company name", inp.company_name, "", "text", "Cover で参照"),
        ("Reporting currency", inp.reporting_currency, "", "text", "JPY / USD"),
        ("Fiscal year-end month", inp.fiscal_year_end_month, "month", "int", "Japan default = 3"),
        ("Accounting basis", inp.accounting_basis, "", "text", "j_gaap / ifrs / us_gaap"),
        ("Forecast period (months)", inp.forecast_period_months, "months", "int", "Series A IB 標準 36"),
        ("Model start (month 1)", inp.model_start_date, "YYYY-MM", "text", "06 §13 monthly grid"),
        ("__SECTION__", "Revenue Drivers", "", "section", "02_saas_metrics §2"),
        ("Starting ARR", inp.starting_arr_money_m, "¥M", "money", "ARR at t=0 (pre-forecast)"),
        ("Monthly growth rate (MoM)", inp.monthly_growth_rate, "%", "pct", "ARR_t = ARR_{t-1}×(1+g)"),
        ("NRR (Net Revenue Retention)", inp.nrr, "x", "pct", "Display only (display)"),
        ("GRR (Gross Revenue Retention)", inp.grr, "x", "pct", "Display only"),
        ("Gross margin", inp.gross_margin_pct, "%", "pct", "Series A SaaS 中央値 70-80%"),
        ("__SECTION__", "Cost Drivers", "", "section", "06 §2.1 OpEx Build"),
        ("Starting employees", inp.starting_employees, "FTE", "int", "Headcount at t=0"),
        ("FTE — Engineering", inp.fte_breakdown.get("engineering", 0), "FTE", "int", ""),
        ("FTE — Sales", inp.fte_breakdown.get("sales", 0), "FTE", "int", ""),
        ("FTE — Customer Success", inp.fte_breakdown.get("customer_success", 0), "FTE", "int", ""),
        ("FTE — G&A", inp.fte_breakdown.get("g_and_a", 0), "FTE", "int", ""),
        ("FLC per FTE (annual)", inp.flc_per_fte_money_m, "¥M", "money", "Fully-Loaded Cost (給与+社保+benefits)"),
        ("S&M % of revenue", inp.sm_pct_of_revenue, "%", "pct", "scaffold simple driver"),
        ("R&D % of revenue", inp.rd_pct_of_revenue, "%", "pct", ""),
        ("G&A % of revenue", inp.ga_pct_of_revenue, "%", "pct", ""),
        ("__SECTION__", "Working Capital", "", "section", "06 §3"),
        ("DSO (days)", inp.dso_days, "days", "int", "AR = Rev × DSO/30 (monthly)"),
        ("DPO (days)", inp.dpo_days, "days", "int", "AP = COGS × DPO/30 (monthly)"),
        ("Deferred Rev % of ARR", inp.deferred_rev_pct_of_arr, "%", "pct", "Annual prepay 部分"),
        ("__SECTION__", "Cap Structure (Initial BS)", "", "section", "04b §1 / 06 §B"),
        ("Starting cash", inp.starting_cash_money_m, "¥M", "money", "BS Cash at t=0"),
        ("Starting debt", inp.starting_debt_money_m, "¥M", "money", "BS LT Debt at t=0"),
        ("__SECTION__", "Tax / Capex / Debt", "", "section", "07 §2 / 06 §B / 06 §4"),
        ("Effective tax rate", inp.effective_tax_rate, "%", "pct", "日本大企業 2026: 31.52%"),
        ("CapEx % of revenue", inp.capex_pct_revenue, "%", "pct", "SaaS 軽 CapEx"),
        ("Depreciation useful life", inp.depreciation_useful_life_months, "months", "int", "5 yr 直線"),
        ("Debt interest rate (annual)", inp.debt_interest_rate_annual, "%", "pct", "Term loan, begin-bal interest"),
        ("__SECTION__", "Stage / Business Model", "", "section", "15 §2-3"),
        ("Stage", inp.stage, "", "text", "auto-judge enum"),
        ("Business model", inp.business_model, "", "text", "SaaS only in Phase 1"),
    ]

    r = 3
    for label, value, unit, fmt, comment in rows:
        if label == "__SECTION__":
            cell = ws.cell(row=r, column=2)
            cell.value = value
            cell.font = ib.FONT_SUBSECTION
            from openpyxl.styles import Alignment as _Align
            cell.alignment = _Align(horizontal="left", vertical="center", indent=0)
            r += 1
            continue
        ws.cell(row=r, column=2).value = label
        ib.apply_label(ws.cell(row=r, column=2), indent=1)
        c = ws.cell(row=r, column=3)
        c.value = value
        if fmt == "money":
            ib.apply_hard_input(c, ib.FMT_MONEY)
        elif fmt == "pct":
            ib.apply_hard_input(c, ib.FMT_PERCENT)
        elif fmt == "int":
            ib.apply_hard_input(c, ib.FMT_INTEGER)
        else:  # text
            from openpyxl.styles import Alignment as _Align
            c.font = ib.FONT_HARD_INPUT
            c.alignment = _Align(horizontal="right", vertical="center")
        ws.cell(row=r, column=4).value = unit
        ib.apply_comment(ws.cell(row=r, column=4))
        if comment:
            ws.cell(row=r, column=5).value = comment
            ib.apply_comment(ws.cell(row=r, column=5))
        r += 1

    # Cache key cell anchors for downstream sheets.
    # We resolve these by re-walking rows; simplest is to record per-label row.
    # Stash on workbook.defined_names later (named ranges). For now, also write
    # explicit anchor labels in column F for self-documentation.
    _ASSUMPTION_ROWS.clear()
    rr = 3
    for label, _v, _u, fmt, _c in rows:
        if label == "__SECTION__":
            rr += 1
            continue
        _ASSUMPTION_ROWS[label] = rr
        rr += 1


# Module-level cache: label → row in 01_Assumptions (column C holds the value).
_ASSUMPTION_ROWS: dict[str, int] = {}


def _assum(label: str) -> str:
    """Return the absolute reference to an assumption cell by its label."""
    row = _ASSUMPTION_ROWS[label]
    return sref("01_Assumptions", f"$C${row}")


# ============================================================================
# Drivers, Revenue, OpEx, Schedules — placeholders to be filled
# ============================================================================


def _build_kpi_dashboard(ws: Worksheet, inp: SaaSInput) -> None:
    """11_KPI_Dashboard — period header + ARR / MRR / Revenue / FTE drivers (monthly).

    Phase 6 Stage A: this function absorbs the content of the deprecated
    _build_drivers (was 02_Drivers). All row anchors (ARR=row 8, Revenue=row 10,
    Total FTE=row 13) are preserved so existing cross-sheet refs continue to work
    after the simple sheet-name rename.
    """
    n = inp.forecast_period_months
    ib.set_tab_color(ws, "driver")
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="11 — KPI Dashboard / Driver Build (Monthly)",
    )

    # Period header row 3: month index + month label (M01, M02, ... M36)
    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    ib.apply_label(ws.cell(row=3, column=3), bold=True)
    for m in range(1, n + 1):
        cell = ws.cell(row=3, column=PERIOD_START_COL + m - 1)
        ib.apply_year_header(cell, f"M{m:02d}")

    # Row 5: Month index numeric
    ws.cell(row=5, column=2).value = "Month index (t)"
    ib.apply_label(ws.cell(row=5, column=2))
    ws.cell(row=5, column=3).value = "n"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        c.value = m
        ib.apply_formula(c, ib.FMT_INTEGER)

    # Row 6: Days in period (scaffold = 30)
    ws.cell(row=6, column=2).value = "Days in period"
    ib.apply_label(ws.cell(row=6, column=2))
    ws.cell(row=6, column=3).value = "days"
    for m in range(1, n + 1):
        c = ws.cell(row=6, column=PERIOD_START_COL + m - 1)
        c.value = 30
        ib.apply_formula(c, ib.FMT_INTEGER)

    # Row 8: ARR (¥M) — geometric MoM growth
    ws.cell(row=8, column=2).value = "ARR (end of month)"
    ib.apply_label(ws.cell(row=8, column=2), bold=True)
    ws.cell(row=8, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=8, column=PERIOD_START_COL + m - 1)
        if m == 1:
            c.value = f"={_assum('Starting ARR')}*(1+{_assum('Monthly growth rate (MoM)')})"
        else:
            prev = period_cell(m - 1, 8)
            c.value = f"={prev}*(1+{_assum('Monthly growth rate (MoM)')})"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 9: MRR = ARR / 12
    ws.cell(row=9, column=2).value = "MRR (= ARR / 12)"
    ib.apply_label(ws.cell(row=9, column=2))
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 8)}/12"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 10: Revenue (recognized) = MRR
    ws.cell(row=10, column=2).value = "Revenue (recognized, this month)"
    ib.apply_label(ws.cell(row=10, column=2), bold=True)
    ws.cell(row=10, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=10, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 9)}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 11: YoY growth (display, computed for m>=13)
    ws.cell(row=11, column=2).value = "Revenue YoY growth"
    ib.apply_label(ws.cell(row=11, column=2))
    ws.cell(row=11, column=3).value = "%"
    for m in range(1, n + 1):
        c = ws.cell(row=11, column=PERIOD_START_COL + m - 1)
        if m >= 13:
            this_cell = period_cell(m, 10)
            yago = period_cell(m - 12, 10)
            c.value = f"={this_cell}/{yago}-1"
        else:
            c.value = "=NA()"
        ib.apply_formula(c, ib.FMT_PERCENT)

    # Row 13: Total FTE — flat (scaffold). Sum of fte breakdown lookups.
    ws.cell(row=13, column=2).value = "Total FTE"
    ib.apply_label(ws.cell(row=13, column=2), bold=True)
    ws.cell(row=13, column=3).value = "FTE"
    fte_total_formula = (
        f"={_assum('FTE — Engineering')}+{_assum('FTE — Sales')}"
        f"+{_assum('FTE — Customer Success')}+{_assum('FTE — G&A')}"
    )
    for m in range(1, n + 1):
        c = ws.cell(row=13, column=PERIOD_START_COL + m - 1)
        c.value = fte_total_formula
        ib.apply_formula(c, ib.FMT_INTEGER)

    # Row 14: ARR per FTE (annualized) = ARR / FTE
    ws.cell(row=14, column=2).value = "ARR per FTE"
    ib.apply_label(ws.cell(row=14, column=2))
    ws.cell(row=14, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=14, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 8)}/{period_cell(m, 13)}"
        ib.apply_formula(c, ib.FMT_MONEY_DECIMAL)

    # Row 16: NRR / GRR display (constants from assumptions)
    ws.cell(row=16, column=2).value = "NRR (display)"
    ib.apply_label(ws.cell(row=16, column=2))
    ws.cell(row=16, column=3).value = "x"
    for m in range(1, n + 1):
        c = ws.cell(row=16, column=PERIOD_START_COL + m - 1)
        c.value = f"={_assum('NRR (Net Revenue Retention)')}"
        ib.apply_formula(c, ib.FMT_MULTIPLE)

    ws.cell(row=17, column=2).value = "GRR (display)"
    ib.apply_label(ws.cell(row=17, column=2))
    ws.cell(row=17, column=3).value = "x"
    for m in range(1, n + 1):
        c = ws.cell(row=17, column=PERIOD_START_COL + m - 1)
        c.value = f"={_assum('GRR (Gross Revenue Retention)')}"
        ib.apply_formula(c, ib.FMT_MULTIPLE)

    # Comment footer
    ws.cell(row=20, column=2).value = (
        "Conventions: ARR_t = ARR_{t-1}×(1+g). MRR=ARR/12. Revenue_t=MRR_t. "
        "ARR_0 = '01_Assumptions'!Starting ARR. (canonical: 02 §2.1, 06 §10.2)"
    )
    ib.apply_comment(ws.cell(row=20, column=2))


def _build_revenue(ws: Worksheet, inp: SaaSInput) -> None:
    """02_Revenue — cohort/build view, links to drivers, gross profit."""
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="03 — Revenue Build",
    )

    # Period header
    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    ib.apply_label(ws.cell(row=3, column=3), bold=True)
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # Phase 6 Fix #3: 階層 line item を「列シフト方式」で表現する。
    #   Row 5: Revenue (total)         ← parent label on B (col 2), unit on C (col 3)
    #   Row 6:   Subscription          ← child label on C (col 3), no unit cell on this row
    #   Row 7:   Services / Other      ← child label on C
    # period 列 (D 以降) の数値配置は変更しない。
    # ※ unit セル (C 列) は親行のみ保持し、子行は省略する (advisor 案 b)。

    # Row 5: Revenue (total) — parent header (label on B; total formula in periods)
    ws.cell(row=5, column=2).value = "Revenue (total)"
    ib.apply_label(ws.cell(row=5, column=2), bold=True)
    ws.cell(row=5, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        c.value = f"=SUM({period_cell(m, 6)}:{period_cell(m, 7)})"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 6: Subscription — child line item (label on C, indented by column shift)
    # Phase 6 Wave 3: cross-sheet ref via named range (=INDEX(KPI_Revenue, m))
    ws.cell(row=6, column=3).value = "Subscription"
    ib.apply_label(ws.cell(row=6, column=3))
    for m in range(1, n + 1):
        c = ws.cell(row=6, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(KPI_Revenue,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 7: Services / Other — child line item (label on C)
    ws.cell(row=7, column=3).value = "Services / Other"
    ib.apply_label(ws.cell(row=7, column=3))
    for m in range(1, n + 1):
        c = ws.cell(row=7, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Phase 6 Fix #3: Revenue (total) は row 5 に移動した。
    # COGS/Gross Profit は新しい total 行 (5) を参照する。
    # COGS は row 8 へ、Gross Profit は row 9 へ、Margin は row 10 へずらす
    # (以前 9/10/11 だったが、行詰めで 1 行ずつ前に)。

    # Row 8: COGS = Revenue × (1 − GM)
    ws.cell(row=8, column=2).value = "COGS (= Revenue × (1 − GM))"
    ib.apply_label(ws.cell(row=8, column=2))
    ws.cell(row=8, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=8, column=PERIOD_START_COL + m - 1)
        c.value = f"=-{period_cell(m, 5)}*(1-{_assum('Gross margin')})"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 9: Gross Profit
    ws.cell(row=9, column=2).value = "Gross Profit"
    ib.apply_label(ws.cell(row=9, column=2), bold=True)
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 5)}+{period_cell(m, 8)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 10: Gross margin (display)
    ws.cell(row=10, column=2).value = "Gross margin %"
    ib.apply_label(ws.cell(row=10, column=2))
    ws.cell(row=10, column=3).value = "%"
    for m in range(1, n + 1):
        c = ws.cell(row=10, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 9)}/{period_cell(m, 5)}"
        ib.apply_formula(c, ib.FMT_PERCENT_BPS)


def _build_opex(ws: Worksheet, inp: SaaSInput) -> None:
    """03_OpEx — S&M / R&D / G&A breakdown + headcount-based personnel.

    In scaffold we use %-of-revenue drivers (canonical fallback when bottom-up
    headcount × FLC unavailable). Headcount × FLC is shown for memo only.
    """
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="04 — OpEx Build",
    )

    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    ib.apply_label(ws.cell(row=3, column=3), bold=True)
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # Row 5: Memo — Total FTE × FLC / 12 (annual / 12 = monthly cost)
    # Phase 6 Wave 3: cross-sheet via FTE_Total named range
    ws.cell(row=5, column=2).value = "Memo: Personnel cost (HC × FLC / 12)"
    ib.apply_label(ws.cell(row=5, column=2))
    ws.cell(row=5, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(FTE_Total,{m})*{_assum('FLC per FTE (annual)')}/12"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 7: S&M = Revenue × S&M %
    # Phase 6 Wave 3: cross-sheet ref via Rev_Total named range
    ws.cell(row=7, column=2).value = "S&M (% of revenue)"
    ib.apply_label(ws.cell(row=7, column=2))
    ws.cell(row=7, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=7, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(Rev_Total,{m})*{_assum('S&M % of revenue')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 8: R&D
    ws.cell(row=8, column=2).value = "R&D (% of revenue)"
    ib.apply_label(ws.cell(row=8, column=2))
    ws.cell(row=8, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=8, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(Rev_Total,{m})*{_assum('R&D % of revenue')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 9: G&A
    ws.cell(row=9, column=2).value = "G&A (% of revenue)"
    ib.apply_label(ws.cell(row=9, column=2))
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(Rev_Total,{m})*{_assum('G&A % of revenue')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 11: Total OpEx (negative numbers, sum)
    ws.cell(row=11, column=2).value = "Total OpEx"
    ib.apply_label(ws.cell(row=11, column=2), bold=True)
    ws.cell(row=11, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=11, column=PERIOD_START_COL + m - 1)
        c.value = f"=SUM({period_cell(m, 7)}:{period_cell(m, 9)})"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 13: D&A (memo) — straight line on cumulative CapEx
    # Phase 6 Wave 3: cross-sheet ref via BS_DA_Memo named range
    ws.cell(row=13, column=2).value = "D&A (allocated to OpEx)"
    ib.apply_label(ws.cell(row=13, column=2))
    ws.cell(row=13, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=13, column=PERIOD_START_COL + m - 1)
        # D&A is computed on the BS sheet roll; here we link via named range.
        c.value = f"=INDEX(BS_DA_Memo,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)


def _build_wc(ws: Worksheet, inp: SaaSInput, row_offset: int = 55) -> None:
    """Working Capital Schedule sub-section, embedded into 05_BS at row 60+.

    Phase 6 Stage A: 08_WC was merged into 05_BS as a sub-section. The schedule
    content keeps its original row layout but is shifted by ``row_offset`` (default
    55, so old WC row 5 lands at BS row 60). All formulas that previously referenced
    08_WC row N now reference 05_BS row (N + row_offset).

    Layout rows (after offset, default offset = 55):
      60: DSO (days)         61: DPO (days)
      63: AR_t = Rev_t × DSO/30
      64: AP_t = (-COGS_t) × DPO/30
      65: Deferred Rev_t = ARR_t × DR%
      67: Δ AR              68: Δ AP             69: Δ DR
    """
    n = inp.forecast_period_months
    R = row_offset

    # Section header band that opens the WC sub-section on 05_BS.
    # Place at row 1+R (= 56 when R=55) so it sits above the period header at row 3+R.
    ib.apply_section_header_band(
        ws, row=1 + R, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="Working Capital Schedule (was 08_WC pre-Phase 6 Stage A)",
    )

    ws.cell(row=3 + R, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3 + R, column=2), bold=True)
    ws.cell(row=3 + R, column=3).value = "Unit"
    for m_ in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3 + R, column=PERIOD_START_COL + m_ - 1), f"M{m_:02d}")

    # Row 5+R: DSO days (constant)
    ws.cell(row=5 + R, column=2).value = "DSO (days)"
    ib.apply_label(ws.cell(row=5 + R, column=2))
    ws.cell(row=5 + R, column=3).value = "days"
    for m_ in range(1, n + 1):
        c = ws.cell(row=5 + R, column=PERIOD_START_COL + m_ - 1)
        c.value = f"={_assum('DSO (days)')}"
        ib.apply_formula(c, ib.FMT_INTEGER)

    # Row 6+R: DPO days
    ws.cell(row=6 + R, column=2).value = "DPO (days)"
    ib.apply_label(ws.cell(row=6 + R, column=2))
    ws.cell(row=6 + R, column=3).value = "days"
    for m_ in range(1, n + 1):
        c = ws.cell(row=6 + R, column=PERIOD_START_COL + m_ - 1)
        c.value = f"={_assum('DPO (days)')}"
        ib.apply_formula(c, ib.FMT_INTEGER)

    # Row 8+R: AR  = Rev × DSO / 30
    # Phase 6 Wave 3: cross-sheet ref via Rev_Total named range
    ws.cell(row=8 + R, column=2).value = "AR (Accounts Receivable)"
    ib.apply_label(ws.cell(row=8 + R, column=2), bold=True)
    ws.cell(row=8 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=8 + R, column=PERIOD_START_COL + m_ - 1)
        c.value = f"=INDEX(Rev_Total,{m_})*{period_cell(m_, 5 + R)}/30"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 9+R: AP  = (-COGS) × DPO / 30
    # Phase 6 Wave 3: COGS は 02_Revenue row 8 (Rev sheet 内負数表示)。canonical name
    # は Profit_Gross が row 9 に登録されているため row 8 の COGS は 直接 sref を保持する。
    # → Profit_Gross は GP (=Rev+COGS) なので, COGS = Profit_Gross - Rev_Total という展開も可能。
    # ただし row 8 を named range 化していないため、sref を維持し最低 disruption を取る。
    ws.cell(row=9 + R, column=2).value = "AP (Accounts Payable)"
    ib.apply_label(ws.cell(row=9 + R, column=2), bold=True)
    ws.cell(row=9 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=9 + R, column=PERIOD_START_COL + m_ - 1)
        cogs_neg = sref("02_Revenue", period_cell(m_, 8))
        c.value = f"=-{cogs_neg}*{period_cell(m_, 6 + R)}/30"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 10+R: Deferred Rev = ARR × DR%
    # Phase 6 Wave 3: cross-sheet ref via KPI_ARR named range
    ws.cell(row=10 + R, column=2).value = "Deferred Revenue"
    ib.apply_label(ws.cell(row=10 + R, column=2), bold=True)
    ws.cell(row=10 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=10 + R, column=PERIOD_START_COL + m_ - 1)
        c.value = f"=INDEX(KPI_ARR,{m_})*{_assum('Deferred Rev % of ARR')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 12+R: Δ AR
    ws.cell(row=12 + R, column=2).value = "Δ AR"
    ib.apply_label(ws.cell(row=12 + R, column=2))
    ws.cell(row=12 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=12 + R, column=PERIOD_START_COL + m_ - 1)
        cur = period_cell(m_, 8 + R)
        prev = "0" if m_ == 1 else period_cell(m_ - 1, 8 + R)
        c.value = f"={cur}-{prev}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 13+R: Δ AP
    ws.cell(row=13 + R, column=2).value = "Δ AP"
    ib.apply_label(ws.cell(row=13 + R, column=2))
    ws.cell(row=13 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=13 + R, column=PERIOD_START_COL + m_ - 1)
        cur = period_cell(m_, 9 + R)
        prev = "0" if m_ == 1 else period_cell(m_ - 1, 9 + R)
        c.value = f"={cur}-{prev}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 14+R: Δ Deferred Rev
    ws.cell(row=14 + R, column=2).value = "Δ Deferred Revenue"
    ib.apply_label(ws.cell(row=14 + R, column=2))
    ws.cell(row=14 + R, column=3).value = "¥M"
    for m_ in range(1, n + 1):
        c = ws.cell(row=14 + R, column=PERIOD_START_COL + m_ - 1)
        cur = period_cell(m_, 10 + R)
        prev = "0" if m_ == 1 else period_cell(m_ - 1, 10 + R)
        c.value = f"={cur}-{prev}"
        ib.apply_formula(c, ib.FMT_MONEY)

def _build_debt(ws: Worksheet, inp: SaaSInput) -> None:
    """07_Debt — Term loan begin-balance interest, no revolver in scaffold.

    Layout rows:
      5: Beginning balance
      6: (+) Issuance
      7: (−) Mandatory amortization (zero in scaffold)
      8: (−) Voluntary repayment
      9: Ending balance
     11: Begin-balance interest expense (Begin × annual rate / 12)
    """
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="09 — Debt Schedule (Term Loan)",
    )

    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # Row 5: Beginning balance
    ws.cell(row=5, column=2).value = "Beginning balance"
    ib.apply_label(ws.cell(row=5, column=2))
    ws.cell(row=5, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        if m == 1:
            c.value = f"={_assum('Starting debt')}"
        else:
            c.value = f"={period_cell(m - 1, 9)}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 6: Issuance (zero scaffold)
    ws.cell(row=6, column=2).value = "(+) New issuance"
    ib.apply_label(ws.cell(row=6, column=2))
    ws.cell(row=6, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=6, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Row 7: Mandatory amort (zero scaffold)
    ws.cell(row=7, column=2).value = "(−) Mandatory amortization"
    ib.apply_label(ws.cell(row=7, column=2))
    ws.cell(row=7, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=7, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Row 8: Voluntary (zero scaffold)
    ws.cell(row=8, column=2).value = "(−) Voluntary repayment"
    ib.apply_label(ws.cell(row=8, column=2))
    ws.cell(row=8, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=8, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Row 9: Ending balance = Begin + Issue − Mand − Vol
    ws.cell(row=9, column=2).value = "Ending balance"
    ib.apply_label(ws.cell(row=9, column=2), bold=True)
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = (
            f"={period_cell(m, 5)}+{period_cell(m, 6)}"
            f"-{period_cell(m, 7)}-{period_cell(m, 8)}"
        )
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 11: Interest expense (Begin × rate / 12)
    ws.cell(row=11, column=2).value = "Interest expense (begin-bal × rate / 12)"
    ib.apply_label(ws.cell(row=11, column=2))
    ws.cell(row=11, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=11, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 5)}*{_assum('Debt interest rate (annual)')}/12"
        ib.apply_formula(c, ib.FMT_MONEY)


def _build_is(ws: Worksheet, inp: SaaSInput) -> None:
    """04_IS — Income Statement (P&L). 06 §2.1.

    Row map (key lines):
       5: Total Revenue              (link 02_Revenue row 7)
       7: COGS                       (link 02_Revenue row 9)
       9: Gross Profit
      11: S&M  (link 03_OpEx row 7)
      12: R&D  (link 03_OpEx row 8)
      13: G&A  (link 03_OpEx row 9)
      14: D&A allocation             (link 05_BS row 50; same as 04 row 13)
      16: Total OpEx (incl. D&A)
      18: EBITDA
      19: EBIT
      21: Interest expense           (link 07_Debt row 11)
      22: Other income / expense     (zero in scaffold)
      23: Pre-tax income (PBT)
      24: Tax provision = MAX(0, PBT) × ETR
      26: Net Income (NI)
    """
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="05 — Income Statement (P&L)",
    )

    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # Row 5: Total Revenue
    # Phase 6 Wave 3: cross-sheet ref via Rev_Total named range
    ws.cell(row=5, column=2).value = "Total Revenue"
    ib.apply_label(ws.cell(row=5, column=2), bold=True)
    ws.cell(row=5, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(Rev_Total,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 7: COGS
    # Phase 6 Wave 3: 02_Revenue row 8 (COGS) は本 builder では named range 化していない
    # (canonical reference の Cost_COGS は 03_OpEx 配置だが、本 builder では COGS は
    # 02_Revenue 内の行)。row 8 は既に Profit_Gross に隣接する低リスク位置のため
    # cell anchor を維持する (rewriting risk > insertion robustness benefit)。
    ws.cell(row=7, column=2).value = "COGS"
    ib.apply_label(ws.cell(row=7, column=2))
    ws.cell(row=7, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=7, column=PERIOD_START_COL + m - 1)
        c.value = f"={sref('02_Revenue', period_cell(m, 8))}"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 9: Gross Profit
    ws.cell(row=9, column=2).value = "Gross Profit"
    ib.apply_label(ws.cell(row=9, column=2), bold=True)
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 5)}+{period_cell(m, 7)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 11-13: OpEx breakdown (all negative values from 03_OpEx)
    # Phase 6 Wave 3: cross-sheet refs via OpEx_SM / OpEx_RD / OpEx_GA named ranges
    for r, opex_name, label in [(11, "OpEx_SM", "S&M"), (12, "OpEx_RD", "R&D"), (13, "OpEx_GA", "G&A")]:
        ws.cell(row=r, column=2).value = label
        ib.apply_label(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3).value = "¥M"
        for m in range(1, n + 1):
            c = ws.cell(row=r, column=PERIOD_START_COL + m - 1)
            c.value = f"=INDEX({opex_name},{m})"
            ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 14: D&A allocation (negative)
    # Phase 6 Wave 3: cross-sheet ref via BS_DA_Memo named range
    ws.cell(row=14, column=2).value = "D&A (allocated)"
    ib.apply_label(ws.cell(row=14, column=2))
    ws.cell(row=14, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=14, column=PERIOD_START_COL + m - 1)
        # D&A on BS memo row is a positive expense; here we negate.
        c.value = f"=-INDEX(BS_DA_Memo,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 16: Total OpEx (incl D&A)
    ws.cell(row=16, column=2).value = "Total OpEx (incl. D&A)"
    ib.apply_label(ws.cell(row=16, column=2), bold=True)
    ws.cell(row=16, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=16, column=PERIOD_START_COL + m - 1)
        c.value = f"=SUM({period_cell(m, 11)}:{period_cell(m, 14)})"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 18: EBITDA = GP + (S&M+R&D+G&A) — i.e. GP plus negative cash opex (excluding D&A)
    ws.cell(row=18, column=2).value = "EBITDA"
    ib.apply_label(ws.cell(row=18, column=2), bold=True)
    ws.cell(row=18, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=18, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 9)}+SUM({period_cell(m, 11)}:{period_cell(m, 13)})"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 19: EBIT = GP + Total OpEx (incl D&A)
    ws.cell(row=19, column=2).value = "EBIT (Operating Income)"
    ib.apply_label(ws.cell(row=19, column=2), bold=True)
    ws.cell(row=19, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=19, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 9)}+{period_cell(m, 16)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 21: Interest expense (negative)
    # Phase 6 Wave 3: cross-sheet ref via Debt_InterestExpense named range
    ws.cell(row=21, column=2).value = "Interest expense"
    ib.apply_label(ws.cell(row=21, column=2))
    ws.cell(row=21, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=21, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(Debt_InterestExpense,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 22: Other income / expense (zero scaffold)
    ws.cell(row=22, column=2).value = "Other income / (expense)"
    ib.apply_label(ws.cell(row=22, column=2))
    ws.cell(row=22, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=22, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Row 23: Pre-tax income (PBT) = EBIT + Interest + Other
    ws.cell(row=23, column=2).value = "Pre-tax Income (PBT)"
    ib.apply_label(ws.cell(row=23, column=2), bold=True)
    ws.cell(row=23, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=23, column=PERIOD_START_COL + m - 1)
        c.value = (
            f"={period_cell(m, 19)}+{period_cell(m, 21)}+{period_cell(m, 22)}"
        )
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 24: Tax provision = MAX(0, PBT) × ETR (NOL not modeled)
    ws.cell(row=24, column=2).value = "Tax provision (NOL not modeled)"
    ib.apply_label(ws.cell(row=24, column=2))
    ws.cell(row=24, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=24, column=PERIOD_START_COL + m - 1)
        c.value = f"=-MAX(0,{period_cell(m, 23)})*{_assum('Effective tax rate')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 26: Net Income
    ws.cell(row=26, column=2).value = "Net Income"
    ib.apply_label(ws.cell(row=26, column=2), bold=True)
    ws.cell(row=26, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=26, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 23)}+{period_cell(m, 24)}"
        ib.apply_grand_total(c, ib.FMT_MONEY)

    # Row 28: NI margin (display)
    ws.cell(row=28, column=2).value = "NI margin %"
    ib.apply_label(ws.cell(row=28, column=2))
    ws.cell(row=28, column=3).value = "%"
    for m in range(1, n + 1):
        c = ws.cell(row=28, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 26)}/{period_cell(m, 5)}"
        ib.apply_formula(c, ib.FMT_PERCENT_BPS)


def _build_bs(ws: Worksheet, inp: SaaSInput) -> None:
    """05_BS — Balance Sheet, monthly. 06 §2.2 + Cash plug.

    Convention (advisor): t=0 in column C ("Beg.") holds initial BS:
        Cash_0 = starting_cash, Debt_0 = starting_debt, APIC_0 = Cash_0 − Debt_0,
        all operating items = 0, RE_0 = 0.

    Row map:
      5: Cash
      6: AR (link 08_WC row 8)
      7: Total Current Assets
      9: Gross PP&E (cumulative CapEx)
     10: Accum Depreciation (cumulative D&A, negative)
     11: Net PP&E
     13: Total Assets
     16: AP (link 08_WC row 9)
     17: Deferred Revenue (link 08_WC row 10)
     18: LT Debt (link 07_Debt row 9)
     20: Total Liabilities
     22: APIC
     23: Retained Earnings (RE_t = RE_{t-1} + NI_t)
     24: Total Equity
     26: Total L+E
     28: BS check (must = 0)

     Memo rows:
     45: CapEx (this month)  = Revenue × CapEx %
     46: Cumulative CapEx (links to row 9)
     50: D&A (this month)    = Cumulative gross PP&E / useful life
    """
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="06 — Balance Sheet",
    )

    # Row 3 header: column C = "Beg." (initial BS at t=0), then M01..M36
    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ib.apply_year_header(ws.cell(row=3, column=3), "Beg.")
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # ---- Memo: CapEx, Cumulative CapEx, D&A ----
    # Row 45: CapEx
    ws.cell(row=45, column=2).value = "Memo: CapEx (this month)"
    ib.apply_label(ws.cell(row=45, column=2))
    # Beg. = 0
    c = ws.cell(row=45, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=45, column=PERIOD_START_COL + m - 1)
        rev = sref("02_Revenue", period_cell(m, 5))
        c.value = f"={rev}*{_assum('CapEx % of revenue')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 46: Cumulative CapEx (= Gross PP&E at this month)
    ws.cell(row=46, column=2).value = "Memo: Cumulative CapEx (= Gross PP&E)"
    ib.apply_label(ws.cell(row=46, column=2))
    c = ws.cell(row=46, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=46, column=PERIOD_START_COL + m - 1)
        if m == 1:
            prev_col = "C"
        else:
            prev_col = col_letter(m - 1)
        c.value = f"={prev_col}46+{period_cell(m, 45)}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 50: D&A this month = Cumulative CapEx / useful life
    ws.cell(row=50, column=2).value = "Memo: D&A (this month, straight-line)"
    ib.apply_label(ws.cell(row=50, column=2))
    c = ws.cell(row=50, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=50, column=PERIOD_START_COL + m - 1)
        c.value = (
            f"=MIN({period_cell(m, 46)}/{_assum('Depreciation useful life')},"
            f"{period_cell(m, 46)}-IF({col_letter(m - 1) if m > 1 else 'C'}51>0,"
            f"{col_letter(m - 1) if m > 1 else 'C'}51,0))"
        )
        # Simpler & safer: SL = Cumulative Gross PP&E / life ; the floor will be
        # capped at gross PP&E via accum dep ≤ gross. Override with simple form:
        c.value = f"={period_cell(m, 46)}/{_assum('Depreciation useful life')}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 51: Cumulative D&A (= Accum Dep magnitude)
    ws.cell(row=51, column=2).value = "Memo: Cumulative D&A"
    ib.apply_label(ws.cell(row=51, column=2))
    c = ws.cell(row=51, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=51, column=PERIOD_START_COL + m - 1)
        prev_col = "C" if m == 1 else col_letter(m - 1)
        c.value = f"={prev_col}51+{period_cell(m, 50)}"
        ib.apply_formula(c, ib.FMT_MONEY)

    # ---- BS rows ----

    # Row 5: Cash — at t=0 = starting; at t≥1 = plug s.t. BS balances.
    ws.cell(row=5, column=2).value = "Cash & Cash Equivalents"
    ib.apply_label(ws.cell(row=5, column=2), bold=True)
    c = ws.cell(row=5, column=3)
    c.value = f"={_assum('Starting cash')}"
    ib.apply_link_intra(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        # Cash plug: prev Cash + CFS Net Change
        # Phase 6 Wave 3: cross-sheet ref via CF_NetChange named range
        prev_col = "C" if m == 1 else col_letter(m - 1)
        c.value = f"={prev_col}5+INDEX(CF_NetChange,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 6: AR
    # Phase 6 Wave 3: intra-sheet ref via WC_AR sheet-scoped name
    ws.cell(row=6, column=2).value = "Accounts Receivable"
    ib.apply_label(ws.cell(row=6, column=2))
    c = ws.cell(row=6, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=6, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(WC_AR,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 7: Total Current Assets
    ws.cell(row=7, column=2).value = "Total Current Assets"
    ib.apply_label(ws.cell(row=7, column=2), bold=True)
    for m in range(0, n + 1):  # include col C (Beg.)
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=7, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}5+{col_letter_}6"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 9: Gross PP&E (= cumulative CapEx)
    ws.cell(row=9, column=2).value = "Gross PP&E"
    ib.apply_label(ws.cell(row=9, column=2))
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=9, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}46"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 10: Accum Depreciation (negative)
    ws.cell(row=10, column=2).value = "Accumulated Depreciation"
    ib.apply_label(ws.cell(row=10, column=2))
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=10, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"=-{col_letter_}51"
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 11: Net PP&E
    ws.cell(row=11, column=2).value = "Net PP&E"
    ib.apply_label(ws.cell(row=11, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=11, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}9+{col_letter_}10"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 13: Total Assets
    ws.cell(row=13, column=2).value = "TOTAL ASSETS"
    ib.apply_label(ws.cell(row=13, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=13, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}7+{col_letter_}11"
        ib.apply_grand_total(c, ib.FMT_MONEY)

    # Row 16: AP (link from WC)
    ws.cell(row=16, column=2).value = "Accounts Payable"
    ib.apply_label(ws.cell(row=16, column=2))
    c = ws.cell(row=16, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=16, column=PERIOD_START_COL + m - 1)
        # Phase 6 Wave 3: intra-sheet ref via WC_AP sheet-scoped name
        c.value = f"=INDEX(WC_AP,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 17: Deferred Revenue
    # Phase 6 Wave 3: intra-sheet ref via WC_DefRev sheet-scoped name
    ws.cell(row=17, column=2).value = "Deferred Revenue"
    ib.apply_label(ws.cell(row=17, column=2))
    c = ws.cell(row=17, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=17, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(WC_DefRev,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 18: LT Debt
    # Phase 6 Wave 3: cross-sheet ref via Debt_Ending named range
    ws.cell(row=18, column=2).value = "Long-term Debt"
    ib.apply_label(ws.cell(row=18, column=2))
    c = ws.cell(row=18, column=3)
    c.value = f"={_assum('Starting debt')}"
    ib.apply_link_intra(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=18, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(Debt_Ending,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 20: Total Liabilities
    ws.cell(row=20, column=2).value = "TOTAL LIABILITIES"
    ib.apply_label(ws.cell(row=20, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=20, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}16+{col_letter_}17+{col_letter_}18"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 22: APIC = Cash_0 − Debt_0 (constant; fixed at Beg.)
    ws.cell(row=22, column=2).value = "Additional Paid-in Capital (APIC)"
    ib.apply_label(ws.cell(row=22, column=2))
    c = ws.cell(row=22, column=3)
    c.value = f"={_assum('Starting cash')}-{_assum('Starting debt')}"
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=22, column=PERIOD_START_COL + m - 1)
        prev_col = "C" if m == 1 else col_letter(m - 1)
        c.value = f"={prev_col}22"  # constant in scaffold (no equity events)
        ib.apply_formula(c, ib.FMT_MONEY)

    # Row 23: Retained Earnings (RE_t = RE_{t-1} + NI_t)
    ws.cell(row=23, column=2).value = "Retained Earnings"
    ib.apply_label(ws.cell(row=23, column=2))
    c = ws.cell(row=23, column=3)
    c.value = 0
    ib.apply_formula(c, ib.FMT_MONEY)
    for m in range(1, n + 1):
        c = ws.cell(row=23, column=PERIOD_START_COL + m - 1)
        # Phase 6 Wave 3: cross-sheet ref via IS_NI named range
        prev_col = "C" if m == 1 else col_letter(m - 1)
        c.value = f"={prev_col}23+INDEX(IS_NI,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 24: Total Equity
    ws.cell(row=24, column=2).value = "TOTAL EQUITY"
    ib.apply_label(ws.cell(row=24, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=24, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}22+{col_letter_}23"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 26: Total L+E
    ws.cell(row=26, column=2).value = "TOTAL LIABILITIES + EQUITY"
    ib.apply_label(ws.cell(row=26, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=26, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}20+{col_letter_}24"
        ib.apply_grand_total(c, ib.FMT_MONEY)

    # Row 28: BS check
    ws.cell(row=28, column=2).value = "BS CHECK (Assets − L+E)"
    ib.apply_label(ws.cell(row=28, column=2), bold=True)
    for m in range(0, n + 1):
        col = 3 if m == 0 else PERIOD_START_COL + m - 1
        c = ws.cell(row=28, column=col)
        col_letter_ = "C" if m == 0 else col_letter(m)
        c.value = f"={col_letter_}13-{col_letter_}26"
        ib.apply_formula(c, ib.FMT_MONEY_DECIMAL)


def _build_cfs(ws: Worksheet, inp: SaaSInput) -> None:
    """06_CFS — Cash Flow Statement (indirect method). 06 §2.3.

    Row map:
      5: Net Income            (link 04_IS row 26)
      6: (+) D&A               (link 05_BS row 50)
      7: (+) SBC               (zero scaffold)
      8: (-) Δ AR              (-Δ from 08_WC row 12)
      9: (+) Δ AP              (+Δ from 08_WC row 13)
     10: (+) Δ Deferred Rev    (+Δ from 08_WC row 14)
     14: CFO subtotal
     17: (-) CapEx             (-row 45 of 05_BS)
     20: CFI subtotal
     23: Debt issuance / repayment (zero scaffold)
     30: CFF subtotal
     32: Net Change in Cash = CFO + CFI + CFF
     33: Cash Beg.
     34: Cash End.
     35: CASH CHECK            (= 0)
    """
    n = inp.forecast_period_months
    ib.setup_sheet_layout(ws, n_periods=n, has_unit_col=True, freeze_at="D5")

    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    ib.apply_section_header_band(
        ws, row=1, start_col=2, end_col=PERIOD_START_COL + n - 1,
        label="07 — Cash Flow Statement (Indirect)",
    )

    ws.cell(row=3, column=2).value = "Period"
    ib.apply_label(ws.cell(row=3, column=2), bold=True)
    ws.cell(row=3, column=3).value = "Unit"
    for m in range(1, n + 1):
        ib.apply_year_header(ws.cell(row=3, column=PERIOD_START_COL + m - 1), f"M{m:02d}")

    # CFO --- Row 5: Net Income
    # Phase 6 Wave 3: cross-sheet ref via IS_NI named range
    ws.cell(row=5, column=2).value = "Net Income"
    ib.apply_label(ws.cell(row=5, column=2))
    ws.cell(row=5, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=5, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(IS_NI,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 6: + D&A
    # Phase 6 Wave 3: cross-sheet ref via BS_DA_Memo named range
    ws.cell(row=6, column=2).value = "(+) Depreciation & Amortization"
    ib.apply_label(ws.cell(row=6, column=2))
    ws.cell(row=6, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=6, column=PERIOD_START_COL + m - 1)
        c.value = f"=INDEX(BS_DA_Memo,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 7: + SBC (zero scaffold)
    ws.cell(row=7, column=2).value = "(+) Stock-Based Compensation"
    ib.apply_label(ws.cell(row=7, column=2))
    ws.cell(row=7, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=7, column=PERIOD_START_COL + m - 1)
        c.value = 0
        ib.apply_hard_input(c, ib.FMT_MONEY)

    # Row 8: − Δ AR
    # Phase 6 Wave 3: cross-sheet ref via WC_dAR named range
    ws.cell(row=8, column=2).value = "(−) Increase in AR"
    ib.apply_label(ws.cell(row=8, column=2))
    ws.cell(row=8, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=8, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(WC_dAR,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 9: + Δ AP
    # Phase 6 Wave 3: cross-sheet ref via WC_dAP named range
    ws.cell(row=9, column=2).value = "(+) Increase in AP"
    ib.apply_label(ws.cell(row=9, column=2))
    ws.cell(row=9, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=9, column=PERIOD_START_COL + m - 1)
        c.value = f"=+INDEX(WC_dAP,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 10: + Δ Deferred Rev
    # Phase 6 Wave 3: cross-sheet ref via WC_dDefRev named range
    ws.cell(row=10, column=2).value = "(+) Increase in Deferred Revenue"
    ib.apply_label(ws.cell(row=10, column=2))
    ws.cell(row=10, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=10, column=PERIOD_START_COL + m - 1)
        c.value = f"=+INDEX(WC_dDefRev,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 14: CFO subtotal
    ws.cell(row=14, column=2).value = "Cash Flow from Operating (CFO)"
    ib.apply_label(ws.cell(row=14, column=2), bold=True)
    ws.cell(row=14, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=14, column=PERIOD_START_COL + m - 1)
        c.value = f"=SUM({period_cell(m, 5)}:{period_cell(m, 10)})"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # CFI Row 17: − CapEx
    # Phase 6 Wave 3: cross-sheet ref via BS_CapEx_Memo named range
    ws.cell(row=17, column=2).value = "(−) CapEx"
    ib.apply_label(ws.cell(row=17, column=2))
    ws.cell(row=17, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=17, column=PERIOD_START_COL + m - 1)
        c.value = f"=-INDEX(BS_CapEx_Memo,{m})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 20: CFI subtotal
    ws.cell(row=20, column=2).value = "Cash Flow from Investing (CFI)"
    ib.apply_label(ws.cell(row=20, column=2), bold=True)
    ws.cell(row=20, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=20, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 17)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # CFF Row 23: Debt issuance (zero)
    ws.cell(row=23, column=2).value = "(+) Debt issuance / (−) repayment, net"
    ib.apply_label(ws.cell(row=23, column=2))
    ws.cell(row=23, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=23, column=PERIOD_START_COL + m - 1)
        # Debt change: (Debt_t − Debt_{t-1}) from 07_Debt
        # Phase 6 Wave 3: cross-sheet ref via Debt_Ending named range
        if m == 1:
            c.value = f"=INDEX(Debt_Ending,1)-{_assum('Starting debt')}"
        else:
            c.value = f"=INDEX(Debt_Ending,{m})-INDEX(Debt_Ending,{m - 1})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 30: CFF subtotal
    ws.cell(row=30, column=2).value = "Cash Flow from Financing (CFF)"
    ib.apply_label(ws.cell(row=30, column=2), bold=True)
    ws.cell(row=30, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=30, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 23)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 32: Net Change in Cash
    ws.cell(row=32, column=2).value = "Net Change in Cash"
    ib.apply_label(ws.cell(row=32, column=2), bold=True)
    ws.cell(row=32, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=32, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 14)}+{period_cell(m, 20)}+{period_cell(m, 30)}"
        ib.apply_subtotal(c, ib.FMT_MONEY)

    # Row 33: Cash Beg.
    # Phase 6 Wave 3: m=1 は BS 'Beg.' col C を参照 (named range は D:AA を覆うので INDEX 不可)
    # — sref で残し、m≥2 のみ BS_Cash 経由に変換。
    ws.cell(row=33, column=2).value = "Cash — Beginning"
    ib.apply_label(ws.cell(row=33, column=2))
    ws.cell(row=33, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=33, column=PERIOD_START_COL + m - 1)
        if m == 1:
            c.value = f"={sref('05_BS', '$C$5')}"
        else:
            c.value = f"=INDEX(BS_Cash,{m - 1})"
        ib.apply_link_intra(c, ib.FMT_MONEY)

    # Row 34: Cash End.
    ws.cell(row=34, column=2).value = "Cash — Ending"
    ib.apply_label(ws.cell(row=34, column=2), bold=True)
    ws.cell(row=34, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=34, column=PERIOD_START_COL + m - 1)
        c.value = f"={period_cell(m, 33)}+{period_cell(m, 32)}"
        ib.apply_grand_total(c, ib.FMT_MONEY)

    # Row 35: CASH CHECK = CFS Cash End − BS Cash
    ws.cell(row=35, column=2).value = "CASH CHECK (CFS End − BS Cash)"
    ib.apply_label(ws.cell(row=35, column=2), bold=True)
    ws.cell(row=35, column=3).value = "¥M"
    for m in range(1, n + 1):
        c = ws.cell(row=35, column=PERIOD_START_COL + m - 1)
        # Phase 6 Wave 3: cross-sheet ref via BS_Cash named range
        c.value = f"={period_cell(m, 34)}-INDEX(BS_Cash,{m})"
        ib.apply_formula(c, ib.FMT_MONEY_DECIMAL)


def _build_skeleton_sheets(wb: Workbook, inp: SaaSInput) -> None:
    """Skeleton (placeholder) for sheets 10..16. Header + TODO note."""
    # Phase 6 Stage A: 11_KPI_Dashboard is now built fully via _build_kpi_dashboard
    # before this skeleton pass; we do NOT re-skeleton it here. 13_Sensitivity is
    # gone (now embedded as sub-section on 09_DCF), so it is removed from this list.
    skeleton_specs = [
        ("08_CapTable", "Cap Table — see scripts/cap_table_builder.py"),
        ("09_DCF", "DCF + Sensitivity (sub-section) — see references/05_valuation_wacc.md §1-§13"),
        ("10_Comps", "Trading & Transaction Comps — references/05 §10"),
        ("12_SanityChecks", "Sanity Checks — references/06_three_statement.md §12 (uses scripts/sanity_checks.py)"),
        ("13_IC_Memo", "IC Memo — references/08_investment_thesis.md §17 template"),
    ]
    for name, note in skeleton_specs:
        ws = wb[name]
        ib.setup_sheet_layout(ws, n_periods=4, has_unit_col=False, freeze_at="A1")
        # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
        ib.apply_section_header_band(
            ws, row=1, start_col=2, end_col=8,
            label=f"{name} — Skeleton (Phase 6 phase 2)",
        )
        ws.cell(row=3, column=2).value = note
        ib.apply_comment(ws.cell(row=3, column=2))
        ib.apply_working_highlight(ws.cell(row=3, column=2))
        if name == "12_SanityChecks":
            ib.set_tab_color(ws, "check")
        elif name == "13_IC_Memo":
            ib.set_tab_color(ws, "memo")


# ============================================================================
# Named ranges
# ============================================================================


def _register_canonical_named_ranges(wb: Workbook, n_periods: int) -> None:
    """Phase 6 Wave 3: canonical named range registration (references/_named_ranges.md §2).

    Cross-sheet 参照を行/列 insertion 耐性のある named range で集約管理する。
    canonical naming は references/_named_ranges.md §2 を参照。

    実装方針:
      - **Workbook-scoped row-range names** (Rev_Total / IS_NI / BS_Cash 等):
        time-series row 全体を `$D$row:$<lastcol>$row` で登録。
        cross-sheet 参照側は `=INDEX(<Name>, m)` で個別期取得 (§7.1, §A.1)。
      - **Sheet-scoped row-range names** (BS_AR / WC_dAR 等): 同 sheet 内の
        intermediate line item 用。
      - 行は **builder の actual layout** (本ファイルの _build_* 関数) を SSoT とし、
        canonical reference の "推定 cell" 列とは独立に決定する。

    旧 `_register_named_ranges` (assumption alias) は別関数で並存。本関数は
    builder の cross-sheet linkage 用 canonical name のみ扱う。
    """
    last_col = get_column_letter(PERIOD_START_COL + n_periods - 1)

    # ----- Workbook-scoped row-range names (cross-sheet 参照される基幹値) -----
    workbook_ranges: dict[str, tuple[str, str]] = {
        # 02_Revenue
        "Rev_Total":         ("02_Revenue", f"$D$5:${last_col}$5"),
        "Profit_Gross":      ("02_Revenue", f"$D$9:${last_col}$9"),
        # 03_OpEx
        "OpEx_SM":           ("03_OpEx",    f"$D$7:${last_col}$7"),
        "OpEx_RD":           ("03_OpEx",    f"$D$8:${last_col}$8"),
        "OpEx_GA":           ("03_OpEx",    f"$D$9:${last_col}$9"),
        "OpEx_Total":        ("03_OpEx",    f"$D$11:${last_col}$11"),
        "Cost_DA":           ("03_OpEx",    f"$D$13:${last_col}$13"),
        # 04_IS
        "IS_Revenue":        ("04_IS",      f"$D$5:${last_col}$5"),
        "IS_COGS":           ("04_IS",      f"$D$7:${last_col}$7"),
        "IS_GrossProfit":    ("04_IS",      f"$D$9:${last_col}$9"),
        "IS_EBITDA":         ("04_IS",      f"$D$18:${last_col}$18"),
        "IS_EBIT":           ("04_IS",      f"$D$19:${last_col}$19"),
        "IS_Interest":       ("04_IS",      f"$D$21:${last_col}$21"),
        "IS_PBT":            ("04_IS",      f"$D$23:${last_col}$23"),
        "IS_Tax":            ("04_IS",      f"$D$24:${last_col}$24"),
        "IS_NI":             ("04_IS",      f"$D$26:${last_col}$26"),
        # 05_BS (本 builder では Total Assets 等が canonical reference の row と
        # 異なる位置にある — actual row が SSoT)
        "BS_Cash":           ("05_BS",      f"$D$5:${last_col}$5"),
        "BS_CurrentAssets":  ("05_BS",      f"$D$7:${last_col}$7"),
        "BS_PPE_Net":        ("05_BS",      f"$D$11:${last_col}$11"),
        "BS_TotalAssets":    ("05_BS",      f"$D$13:${last_col}$13"),
        "BS_LTDebt":         ("05_BS",      f"$D$18:${last_col}$18"),
        "BS_TotalLiab":      ("05_BS",      f"$D$20:${last_col}$20"),
        "BS_RE":             ("05_BS",      f"$D$23:${last_col}$23"),
        "BS_TotalEquity":    ("05_BS",      f"$D$24:${last_col}$24"),
        "BS_LE_Total":       ("05_BS",      f"$D$26:${last_col}$26"),
        # 06_CFS
        "CF_Operating":      ("06_CFS",     f"$D$14:${last_col}$14"),
        "CF_CapEx":          ("06_CFS",     f"$D$17:${last_col}$17"),
        "CF_Investing":      ("06_CFS",     f"$D$20:${last_col}$20"),
        "CF_Financing":      ("06_CFS",     f"$D$30:${last_col}$30"),
        "CF_NetChange":      ("06_CFS",     f"$D$32:${last_col}$32"),
        "CF_EndCash":        ("06_CFS",     f"$D$34:${last_col}$34"),
        # 07_Debt
        "Debt_Ending":         ("07_Debt", f"$D$9:${last_col}$9"),
        "Debt_InterestExpense": ("07_Debt", f"$D$11:${last_col}$11"),
        # 11_KPI_Dashboard (本 builder では ARR / Revenue / FTE_Total が driver row)
        "KPI_ARR":           ("11_KPI_Dashboard", f"$D$8:${last_col}$8"),
        "KPI_Revenue":       ("11_KPI_Dashboard", f"$D$10:${last_col}$10"),
        "FTE_Total":         ("11_KPI_Dashboard", f"$D$13:${last_col}$13"),
        # BS memo rows (CapEx / D&A) — cross-sheet referenced by 03_OpEx, 06_CFS, 04_IS
        # 本 builder では memo を 05_BS に置いている (canonical では 03_OpEx の Cost_DA に
        # 相当)。cross-sheet access のため workbook-scoped で登録する。
        "BS_CapEx_Memo":     ("05_BS",  f"$D$45:${last_col}$45"),
        "BS_DA_Memo":        ("05_BS",  f"$D$50:${last_col}$50"),
        # WC Δ rows (cross-sheet referenced by 06_CFS rows 8/9/10)
        "WC_dAR":            ("05_BS",  f"$D$67:${last_col}$67"),
        "WC_dAP":            ("05_BS",  f"$D$68:${last_col}$68"),
        "WC_dDefRev":        ("05_BS",  f"$D$69:${last_col}$69"),
    }
    ib.bulk_register_workbook_names(wb, workbook_ranges)

    # ----- Sheet-scoped row-range names (05_BS WC sub-section, intra-sheet only) -----
    ws_bs = wb["05_BS"]
    bs_sheet_ranges: dict[str, str] = {
        # WC sub-section (row_offset = 55, so AR=63, AP=64, DR=65)
        "BS_AR":             f"$D$6:${last_col}$6",
        "BS_AP":             f"$D$16:${last_col}$16",
        "BS_DefRev_Current": f"$D$17:${last_col}$17",
        "WC_AR":             f"$D$63:${last_col}$63",
        "WC_AP":             f"$D$64:${last_col}$64",
        "WC_DefRev":         f"$D$65:${last_col}$65",
    }
    ib.bulk_register_sheet_names(ws_bs, bs_sheet_ranges)


def _register_named_ranges(wb: Workbook) -> None:
    """Register friendly defined names for key assumption cells.

    Only the most-referenced inputs get named ranges. Keeps SSoT in
    01_Assumptions but lets future formulas use `gross_margin_pct` etc.
    """
    name_map = {
        "starting_arr_money_m": "Starting ARR",
        "monthly_growth_rate": "Monthly growth rate (MoM)",
        "nrr": "NRR (Net Revenue Retention)",
        "grr": "GRR (Gross Revenue Retention)",
        "gross_margin_pct": "Gross margin",
        "sm_pct_of_revenue": "S&M % of revenue",
        "rd_pct_of_revenue": "R&D % of revenue",
        "ga_pct_of_revenue": "G&A % of revenue",
        "dso_days": "DSO (days)",
        "dpo_days": "DPO (days)",
        "deferred_rev_pct_of_arr": "Deferred Rev % of ARR",
        "starting_cash_money_m": "Starting cash",
        "starting_debt_money_m": "Starting debt",
        "effective_tax_rate": "Effective tax rate",
        "capex_pct_revenue": "CapEx % of revenue",
        "depreciation_useful_life_months": "Depreciation useful life",
        "debt_interest_rate_annual": "Debt interest rate (annual)",
    }
    for friendly, label in name_map.items():
        if label not in _ASSUMPTION_ROWS:
            continue
        row = _ASSUMPTION_ROWS[label]
        ref = f"'01_Assumptions'!$C${row}"
        try:
            dn = DefinedName(name=friendly, attr_text=ref)
            wb.defined_names[friendly] = dn
        except Exception:
            # If openpyxl version differs, fall back to legacy API.
            try:
                wb.defined_names.append(DefinedName(name=friendly, attr_text=ref))
            except Exception:
                pass


# ============================================================================
# CLI
# ============================================================================


def _cli() -> None:
    import argparse

    ap = argparse.ArgumentParser(
        description="SaaS Series A 14-sheet xlsx builder (Phase 6 Stage A)."
    )
    ap.add_argument("--output", default="model.xlsx", help="Output xlsx path")
    ap.add_argument(
        "--demo",
        action="store_true",
        help="Use evals.json case 1 demo input (saas_series_a_3yr_model)",
    )
    args = ap.parse_args()

    if args.demo:
        inp = get_default_input()
    else:
        # Phase 6 phase 2: load from yaml/json config. For now demo only.
        inp = get_default_input()
    wb = build_saas_series_a_model(inp)
    wb.save(args.output)
    print(f"Saved: {args.output} ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    _cli()
