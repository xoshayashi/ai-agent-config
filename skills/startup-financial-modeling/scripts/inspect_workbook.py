#!/usr/bin/env python3
"""IBフォーマット作法の機械検査。references/ib_format_spec.md の実装契約を検証する。

使い方: python3 inspect_workbook.py <xlsx> [--recalc]
  --recalc: LibreOffice(soffice)で再計算し、数式エラーゼロ・総合判定OKまで検証する
終了コード 0=全パス / 1=違反あり / 2=recalc検証失敗
"""
from __future__ import annotations

import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

ALLOWED_FMTS = {
    "General",
    '#,##0,,_);(#,##0,,);"-"_)',
    '#,##0_);(#,##0);"-"_)',
    "0.0%_);(0.0%)",
    '#,##0.0_);(#,##0.0);"-"_)',
    "0.0x_);(0.0x)",
    '0.0"x"_);(0.0"x");"-"_)',
    "0.0000%_);(0.0000%)",
}
VOLATILE = re.compile(r"\b(OFFSET|INDIRECT|NOW|TODAY|RAND|RANDBETWEEN)\s*\(",
                      re.IGNORECASE)
PURE_LINK = re.compile(r"^='[^'!]+'!\$?[A-Z]{1,3}\$?\d+$")
BANNED_HEADERS = ("source", "driver", "ソース列")

BLUE = "FF0000FF"
BLACK = "FF000000"
GREEN = "FF008000"
GRAY = "FF808080"
WHITE = "FFFFFFFF"
NAVY = "FF1F3864"


def color_of(cell):
    f = cell.font
    if f is None or f.color is None or f.color.rgb is None:
        return BLACK
    return str(f.color.rgb)


def _print_end(ws):
    pa = str(ws.print_area or "").replace("$", "")
    m = re.search(r":([A-Z]{1,2})(\d+)$", pa)
    return int(m.group(2)) if m else ws.max_row


def main(path):
    wb = openpyxl.load_workbook(path)
    fails = []
    ok = []

    def gate(cond, label, detail=""):
        (ok if cond else fails).append(f"{label}" + (f" — {detail}" if detail and not cond else ""))

    gate(len(wb.sheetnames) <= 10, f"シート数≦10（{len(wb.sheetnames)}）")

    for ws in wb.worksheets:
        t = ws.title
        # A列スペーサー・B列起点
        a_vals = [c.value for row in ws.iter_rows(min_col=1, max_col=1)
                  for c in row if c.value not in (None, "")]
        gate(not a_vals, f"[{t}] A列が空", str(a_vals[:3]))
        gate((ws.column_dimensions["A"].width or 8.43) <= 2,
             f"[{t}] A列幅≦2")
        for ic in ("B", "C"):
            gate((ws.column_dimensions[ic].width or 8.43) <= 2.5,
                 f"[{t}] インデント列{ic}の幅≦2.5")
        gate((ws.column_dimensions["E"].width or 8.43) <= 12,
             f"[{t}] 単位列Eの幅≦12")
        indented = [c.coordinate for row in ws.iter_rows() for c in row
                    if c.alignment is not None and c.alignment.indent]
        gate(not indented, f"[{t}] ネイティブインデント不使用", str(indented[:3]))
        row1 = [c.value for row in ws.iter_rows(min_row=1, max_row=1)
                for c in row if c.value not in (None, "")]
        gate(not row1, f"[{t}] 1行目が空", str(row1[:3]))
        gate(ws.cell(row=2, column=2).value not in (None, ""),
             f"[{t}] タイトルがB2")
        # 結合セル・行高・グリッド線
        gate(len(ws.merged_cells.ranges) == 0, f"[{t}] 結合セルなし",
             str(len(ws.merged_cells.ranges)))
        # 見出し行(5)だけは折り返しのため高さを持てる。数値行の行高いじりは禁止。
        heights = [f"{r}:{rd.height}" for r, rd in ws.row_dimensions.items()
                   if rd.height is not None and r != 5]
        gate(not heights, f"[{t}] 行高はデフォルトのまま（見出し行を除く）",
             str(heights[:3]))
        gate(ws.sheet_view.showGridLines is False, f"[{t}] グリッド線オフ")
        # 禁止ヘッダー（Source/driver列）
        banned = [c.value for row in ws.iter_rows(max_row=8)
                  for c in row if isinstance(c.value, str)
                  and c.value.strip().lower() in BANNED_HEADERS]
        gate(not banned, f"[{t}] Source/driver列なし", str(banned))

        fills = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is None and (c.fill is None or
                                        c.fill.patternType is None):
                    continue
                # フィルはタイトル帯（1行目ネイビー）のみ
                if c.fill is not None and c.fill.patternType == "solid":
                    fg = str(c.fill.fgColor.rgb)
                    if not (c.row == 2 and fg == NAVY):
                        fills += 1
                if c.value is None:
                    continue
                # フォント統一
                f = c.font
                if f.name != "Arial" or (f.size or 10) != 10:
                    fails.append(f"[{t}]{c.coordinate} フォント Arial10 以外: "
                                 f"{f.name}{f.size}")
                    continue
                # 数値書式ホワイトリスト
                if c.number_format not in ALLOWED_FMTS:
                    fails.append(f"[{t}]{c.coordinate} 書式外: {c.number_format}")
                # カラーコード検査（数値・数式セルのみ。C列以降）
                if c.column >= 6 and c.row > 5:
                    col = color_of(c)
                    if isinstance(c.value, str) and c.value.startswith("="):
                        gate(not VOLATILE.search(c.value),
                             f"[{t}]{c.coordinate} 揮発関数なし", c.value)
                        expect = GREEN if PURE_LINK.match(c.value) else BLACK
                        if col not in (expect,):
                            fails.append(
                                f"[{t}]{c.coordinate} 数式の色: {col} "
                                f"期待{expect} {c.value[:40]}")
                    elif isinstance(c.value, (int, float)):
                        if col != BLUE:
                            fails.append(
                                f"[{t}]{c.coordinate} 入力値が青字でない: {col}")
        gate(fills == 0, f"[{t}] フィルはタイトル帯のみ", f"{fills}件の逸脱")

    # 入力ゾーン規則: 青字入力は前提条件とバリュエーション（コンプ表）のみ
    INPUT_ZONES = {"前提条件", "バリュエーション"}
    for ws in wb.worksheets:
        if ws.title in INPUT_ZONES:
            continue
        stray = [c.coordinate for row in ws.iter_rows() for c in row
                 if isinstance(c.value, (int, float))
                 and not isinstance(c.value, bool)
                 and c.font and c.font.color and c.font.color.rgb == "FF0000FF"]
        gate(not stray, f"[{ws.title}] 入力ゾーン外に青字入力なし", str(stray[:4]))

    # --- 数珠つなぎ（純リンクの純リンク参照）禁止 ---
    pure = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and PURE_LINK.match(c.value):
                    m = re.match(r"^='([^'!]+)'!\$?([A-Z]{1,3})\$?(\d+)$",
                                 c.value)
                    pure[(ws.title, c.coordinate)] = (m.group(1),
                                                      f"{m.group(2)}{m.group(3)}")
    chains = [f"{sh}!{co}→{t[0]}!{t[1]}" for (sh, co), t in pure.items()
              if t in pure]
    gate(not chains, f"数珠つなぎリンクなし（{len(chains)}件）", str(chains[:4]))

    # --- 単位列Eの網羅（期間列に値を持つラベル行） ---
    NON_PERIOD = {"ベンチマーク", "評価", "根拠", "単位",
                  "根拠・出典・備考（印刷範囲外）"}

    def period_cols(ws):
        return [c for c in range(6, ws.max_column + 1)
                if ws.cell(row=5, column=c).value not in NON_PERIOD
                and ws.cell(row=5, column=c).value is not None]
    nounit = []
    for ws in wb.worksheets:
        pcols = period_cols(ws)
        for r in range(6, ws.max_row + 1):
            lab = (ws.cell(row=r, column=2).value
                   or ws.cell(row=r, column=3).value
                   or ws.cell(row=r, column=4).value)
            has = any(ws.cell(row=r, column=c).value is not None
                      for c in pcols)
            if lab and has and not ws.cell(row=r, column=5).value:
                nounit.append(f"{ws.title}!{r}")
    gate(not nounit, f"単位列Eの網羅（欠落{len(nounit)}件）", str(nounit[:4]))

    # --- ラベルが単位列にはみ出さない ---
    def vw(s):
        return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)
    clipped = []
    for ws in wb.worksheets:
        dw = ws.column_dimensions["D"].width or 32
        for r in range(6, ws.max_row + 1):
            # 数値のある行だけを検査する。説明行（数値なし）は右のセルが空なので
            # 文章がにじみ出て読める＝切れない。ここを一律に縛ると、
            # 平易な説明そのものが書けなくなる。
            has_num = any(ws.cell(row=r, column=c).value is not None
                          for c in range(6, 13))
            if not has_num:
                continue
            for c in (2, 3, 4):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and not v.startswith("="):
                    avail = dw + (4 - c) * 2.14
                    if vw(v) > avail:
                        clipped.append(f"{ws.title}!{ws.cell(row=r, column=c).coordinate}")
    gate(not clipped, f"ラベルが列幅に収まる（超過{len(clipped)}件）",
         str(clipped[:5]))

    # --- ラベル・ヘッダーの括弧整合（機械的短縮による破損検知） ---
    PAIRS = (("（", "）"), ("(", ")"), ("「", "」"), ("【", "】"))
    broken = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                if c.column > 5 and c.row != 5 and not (
                        ws.cell(row=5, column=c.column).value
                        == "根拠・出典・備考（印刷範囲外）"):
                    continue
                for a, b in PAIRS:
                    if v.count(a) != v.count(b):
                        broken.append(f"{ws.title}!{c.coordinate}")
                        break
    gate(not broken, f"ラベルの括弧が閉じている（破損{len(broken)}件）",
         str(broken[:5]))

    # --- 期間ヘッダーが列幅に収まる（折り返しがあれば最大3行まで許容） ---
    hdr_over = []
    for ws in wb.worksheets:
        for c in range(6, ws.max_column + 1):
            cell = ws.cell(row=5, column=c)
            v = cell.value
            if not isinstance(v, str) or v.startswith("根拠"):
                continue
            w = ws.column_dimensions[
                openpyxl.utils.get_column_letter(c)].width or 11.5
            lines = 3 if (cell.alignment and cell.alignment.wrap_text) else 1
            if vw(v) > w * lines:
                hdr_over.append(f"{ws.title}!{cell.coordinate}")
    gate(not hdr_over, f"期間ヘッダーが列幅に収まる（超過{len(hdr_over)}件）",
         str(hdr_over[:5]))

    # --- 数値列が見出しに引きずられて広がっていない（構造ゲート） ---
    # 印刷は fitToWidth なので、1列でも標準より広いと**全シートが縮小されて**
    # 数字が読めなくなる。幅の合計を定数で縛るのではなく、「列は役割ごとの
    # 標準幅である」ことを問う（期間列の本数が違うシートにもそのまま効く）。
    STD_PERIOD_W = 11.5
    wide = []
    for ws in wb.worksheets:
        for c in range(6, ws.max_column + 1):
            v = ws.cell(row=5, column=c).value
            if not isinstance(v, str) or v.startswith("根拠"):
                continue
            L = openpyxl.utils.get_column_letter(c)
            w = ws.column_dimensions[L].width or STD_PERIOD_W
            # 「数値列か」は**セルの数値書式**で決める（ラベルでも位置でもない）。
            # ベンチマーク・評価のような文章を持つ補助列は対象外（幅が要る）。
            numeric = any(
                ws.cell(row=r, column=c).number_format != "General"
                and ws.cell(row=r, column=c).value is not None
                for r in range(6, min(ws.max_row, 240) + 1))
            if numeric and w > STD_PERIOD_W + 0.6:
                wide.append(f"{ws.title}!{L}={w:.0f}")
    gate(not wide, f"数値列が標準幅（{STD_PERIOD_W}）を超えない（超過{len(wide)}件）",
         str(wide[:6]))

    # --- 備考の可視幅・印刷範囲の終端 ---
    def vwidth(s):
        return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)
    overflow, prints = [], []
    for ws in wb.worksheets:
        ncol = next((c for c in range(6, ws.max_column + 1)
                     if ws.cell(row=5, column=c).value
                     == "根拠・出典・備考（印刷範囲外）"),
                    None)
        if not ncol:
            continue
        limit = ws.column_dimensions[
            openpyxl.utils.get_column_letter(ncol)].width or 46
        # 備考の幅は「印刷面の行」だけを縛る。印刷範囲外（監査証跡）の備考は、
        # 読み手ではなく検証者のためのものなので、長くてよい。
        for r in range(6, _print_end(ws) + 1):
            v = ws.cell(row=r, column=ncol).value
            if isinstance(v, str) and not v.startswith("=") \
                    and vwidth(v) > limit:
                overflow.append(f"{ws.title}!{r}")
        # 印刷範囲: 列は備考列で終端。行は途中で切ってよいが（チェック明細を
        # 監査証跡として印刷面から外す）、総合判定は必ず印刷面に載ること。
        bcol = next((c for c in range(6, ws.max_column + 1)
                     if ws.cell(row=5, column=c).value == "根拠"), ncol)
        end = openpyxl.utils.get_column_letter(bcol)
        pa = str(ws.print_area or "").replace("$", "")
        if pa:
            m = re.search(r":([A-Z]{1,2})(\d+)$", pa)
            last = int(m.group(2)) if m else 0
            bad = (not m) or m.group(1) != end or last > ws.max_row
            if not bad and ws.title == "サマリー":
                vr = next((r for r in range(6, ws.max_row + 1)
                           if any(isinstance(ws.cell(row=r, column=c).value,
                                             str)
                                  and str(ws.cell(row=r,
                                                  column=c).value).startswith(
                                      "総合判定")
                                  for c in (2, 3, 4))), None)
                bad = bool(vr and vr > last)
            if bad:
                prints.append(f"{ws.title}:{ws.print_area}")
    gate(not overflow, f"備考が列幅内に収まる（超過{len(overflow)}件）",
         str(overflow[:4]))
    gate(not prints, f"印刷範囲は根拠列で終端（逸脱{len(prints)}件）",
         str(prints[:3]))

    # --- 導出行の備考が「検証が通った」と名乗らない（逃げ道の封じ込め） ---
    # 実例: 出荷台数を期末の目標台数から逆算しておきながら、備考に
    #   「ロールフォワード閉合: 期首＋新規−退役＝期末」
    # と書いていた。この等式は**逆算した時点で定義上必ず成り立つ**。
    # 読み手は「検算が通った行」と受け取るが、実際には何も検証していない。
    # 検証を名乗れるのはチェック行だけ。導出行が権威を借りるのを構造的に禁じる。
    # --- 「検証」を名乗る語が、見出し・説明行に逃げていないか ---
    # 導出行の備考だけを禁じたら、同じ主張がセクション見出しとブロック注記に
    # 移って生き延びた（監査ジャッジが実証）:
    #   「ソースの看板主張（18か月回収）を、モデルの数字で検証する行」
    # 実際は製造原価が「18ヶ月回収から逆算」した値なので、回収期間は恒等的に
    # 18.0にしかならない。**前提から解いた数字で前提を検証している。**
    # 逃げ道は語ではなく構造で塞ぐ: 印刷面のあらゆるテキストで「検証」「一致」を
    # 名乗らせない（チェック行とその集約行だけが名乗ってよい）。
    VERIFY_WORDS = ("を検証", "の検証", "検証する", "と一致", "に一致",
                    "整合を確認", "検証点")
    boast = []
    for ws in wb.worksheets:
        last = _print_end(ws)
        for r in range(6, last + 1):
            for c in (2, 3, 4):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                if "■" in v or "□" in v:      # チェック行は名乗ってよい
                    continue
                if any(w in v for w in VERIFY_WORDS):
                    boast.append(f"{ws.title}!{r}{ws.cell(row=r, column=c).column_letter} {v[:24]}")
    gate(not boast,
         f"印刷面の見出し・説明が検証を名乗らない（違反{len(boast)}件）",
         str(boast[:4]))

    # 対象は**数式を持つ導出行**だけ。節見出し（数式なし）から検算行を
    # 「そこで確かめている」と指し示すのは正当なので対象外。
    CLAIMS = ("閉合", "検算", "整合確認", "一致確認", "検証済")
    borrowed = []
    for ws in wb.worksheets:
        for r in range(6, ws.max_row + 1):
            lab = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                        if isinstance(ws.cell(row=r, column=c).value, str)
                        and ws.cell(row=r, column=c).value.strip()), "")
            if "■" in lab or "□" in lab:      # チェック行・その集約行は名乗ってよい
                continue
            f = ws.cell(row=r, column=6).value
            if not (isinstance(f, str) and f.startswith("=")):
                continue
            note = ws.cell(row=r, column=ncol).value
            if isinstance(note, str) and any(k in note for k in CLAIMS):
                borrowed.append(f"{ws.title}!{r} {lab[:16]}")
    gate(not borrowed,
         f"導出行の備考が検証を名乗らない（違反{len(borrowed)}件）",
         str(borrowed[:4]))

    # --- 中身のない見出し（ダングリング見出し）を禁じる ---
    # 「入力から計算し直した検算」の見出しだけが印刷面に残り、明細は印刷範囲外
    # にあったため、PDFに空の見出しページが出ていた。見出しは中身の予告であって、
    # 中身が付いてこないなら紙面に出す理由がない。
    dangling = []
    for ws in wb.worksheets:
        last = _print_end(ws)
        for r in range(6, last + 1):
            head = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                         if isinstance(ws.cell(row=r, column=c).value, str)
                         and ws.cell(row=r, column=c).value.strip()), None)
            if not head:
                continue
            if head.startswith(("（注記）", "（免責）")):
                continue                                # 注記・免責は本文
            if any(ws.cell(row=r, column=c).value is not None
                   for c in range(6, 13)):
                continue                                # 値を持つ行は見出しでない
            # **単位だけを持つ行**も空行（ケース候補を折りたたんだ親ラベルが
            # 12本残り、前提条件が8ページに膨らんだ）。見出しは中身の予告であって、
            # 中身が付いてこないなら紙面に出す理由がない。
            if ws.cell(row=r, column=5).value is not None:
                dangling.append(f"{ws.title}!{r} {head[:20]}（単位のみ）")
                continue
            has_body = False
            for rr in range(r + 1, min(last, r + 40) + 1):
                nxt = next((ws.cell(row=rr, column=c).value for c in (2, 3, 4)
                            if isinstance(ws.cell(row=rr, column=c).value, str)
                            and ws.cell(row=rr, column=c).value.strip()), None)
                if any(ws.cell(row=rr, column=c).value is not None
                       for c in range(6, 13)):
                    has_body = True
                    break
                if nxt and ws.cell(row=rr, column=2).value:
                    break                               # 次の節見出しに当たった
            if not has_body:
                dangling.append(f"{ws.title}!{r} {head[:20]}")
    gate(not dangling, f"印刷面に中身のない見出しがない（{len(dangling)}件）",
         str(dangling[:4]))

    # --- 孤児の青字入力（印刷面にあるのに誰も読まないセル）を禁じる ---
    # 期初現金のJ列を990億円に書き換えても総合判定はOKのままだった。
    # 参照が $F$ 固定なのに、値をF〜J全列に青字で複製していたため。
    # **編集できるのに何も起きないセルを紙面に出してはならない**（読み手を欺く）。
    refd = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                for m in re.finditer(
                        r"(?:'([^']+)'!)?\$?([A-Z]{1,2})\$?(\d+)"
                        r"(?::\$?([A-Z]{1,2})\$?(\d+))?", c.value):
                    sh = m.group(1) or ws.title
                    c1 = openpyxl.utils.column_index_from_string(m.group(2))
                    r1 = int(m.group(3))
                    c2 = (openpyxl.utils.column_index_from_string(m.group(4))
                          if m.group(4) else c1)
                    r2 = int(m.group(5)) if m.group(5) else r1
                    for rr in range(r1, r2 + 1):
                        for cc in range(c1, c2 + 1):
                            refd.add((sh, rr, cc))
    orphan = []
    for ws in wb.worksheets:
        last = _print_end(ws)
        for r in range(6, last + 1):
            for c in range(6, 13):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, (int, float)):
                    continue
                if isinstance(cell.value, bool):
                    continue
                if color_of(cell) != BLUE:
                    continue
                if (ws.title, r, c) not in refd:
                    orphan.append(f"{ws.title}!{cell.coordinate}")
    gate(not orphan,
         f"印刷面の青字入力はすべて参照されている（孤児{len(orphan)}件）",
         str(orphan[:6]))

    # --- ケース依存のアラートは、全ケース列を走査していること ---
    # 「いずれかの手法で株式価値が0以下」が中位列しか見ておらず、低位で
    # 負の株式価値を印刷しながら警告が消灯していた（3名がcritical）。
    # 1本直しても、同型が3本残っていた（次の巡でIBが実演）。**構造で塞ぐ。**
    #
    # 判定: ケース列（低位/中位/高位）を持つシートのアラート行が、
    # ケース列を持つ行を参照しているのに、参照が中位列（G）だけなら FAIL。
    CASE_HEADS = ("低位", "中位", "高位")
    narrow = []
    for ws in wb.worksheets:
        heads = [ws.cell(row=5, column=c).value for c in range(6, 9)]
        if list(heads) != list(CASE_HEADS):
            continue
        for r in range(6, ws.max_row + 1):
            lab = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                        if isinstance(ws.cell(row=r, column=c).value, str)
                        and ws.cell(row=r, column=c).value.strip()), "")
            if not lab.startswith("□要説明"):
                continue
            f = ws.cell(row=r, column=6).value
            if not isinstance(f, str):
                continue
            # 参照している自シートの行を集める
            rows = {int(m.group(2)) for m in
                    re.finditer(r"(?<!!)(?<!\$)\b([A-H])(\d+)\b", f)}
            # そのうち「ケース列を3つとも持つ行」（＝ケース依存の行）
            cased = [rr for rr in rows
                     if all(ws.cell(row=rr, column=c).value is not None
                            for c in (6, 7, 8))]
            if not cased:
                continue
            # ケース依存の行を読んでいるのに、F列・H列を一度も参照していない
            reads = {m.group(1) for m in
                     re.finditer(r"(?<!!)(?<!\$)\b([A-H])\d+\b", f)}
            if "F" not in reads and "H" not in reads:
                narrow.append(f"{ws.title}!{r} {lab[:22]}")
    gate(not narrow,
         f"ケース依存の警告が全ケース列を走査（違反{len(narrow)}件）",
         str(narrow[:4]))

    # --- 印刷面の文字列が列幅で切り詰められていない（原則19） ---
    # 位置表示の値は正しく「FY2027 FY2028 FY2029 FY2030」（4年）だったのに、
    # 右寄せのまま列幅を超えたため、PDFには「/2029 FY2030」と印字された。
    # **4年の問題が2年に見え、最も逼迫した年が消えた。**
    # セルの値が正しいことは、紙面が正しいことを意味しない。
    def _col_w(ws, c):
        return ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width \
            or 11.5
    truncated = []
    for ws in wb.worksheets:
        last = _print_end(ws)
        pe_col = 12  # 印刷面の右端（根拠列まで）
        m = re.search(r":([A-Z]{1,2})\d+$", str(ws.print_area or "").replace("$", ""))
        if m:
            pe_col = openpyxl.utils.column_index_from_string(m.group(1))
        for r in range(6, last + 1):
            for c in range(6, pe_col + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                need = vw(v)
                if need <= _col_w(ws, c):
                    continue
                # **溢れは連続する空セルの幅を合計して吸収できるかで判定する**
                # （右隣1セルだけを見ると、幅69の文字がH/I/J(空)を越えてK(タグ)に
                #  衝突しても見逃す＝ゲートがCJK幅盲点で嘘をつく）。
                left = (cell.alignment and
                        cell.alignment.horizontal == "left")
                avail = _col_w(ws, c)
                step = 1 if left else -1
                nc = c + step
                while 6 <= nc <= pe_col and \
                        ws.cell(row=r, column=nc).value is None:
                    avail += _col_w(ws, nc)
                    nc += step
                if need > avail:
                    truncated.append(f"{ws.title}!{cell.coordinate} "
                                     f"幅{need:.0f}>{avail:.0f} {v[:16]}")
    gate(not truncated,
         f"印刷面の文字列が切り詰められない（超過{len(truncated)}件）",
         str(truncated[:4]))

    # --- 軸の語彙を混ぜない（原則17）: 列軸の値にシナリオ名を付けない ---
    # 「（参考）Downsideケースの回収倍率 1.9倍」は、実体がバリュエーションの
    # **低位列**（保守的倍率）のMOICだった。実際にDownsideシナリオを回すと3.59倍。
    # 1ページ目に、同じ名前で約半分の数字が印刷されていた。
    SCEN = ("Base", "Upside", "Downside")
    CASE_COL = ("低位", "中位", "高位")
    mixed = []
    for ws in wb.worksheets:
        last = _print_end(ws)
        heads = [ws.cell(row=5, column=c).value for c in range(6, 9)]
        is_case_sheet = list(heads) == list(CASE_COL)
        for r in range(6, last + 1):
            lab = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                        if isinstance(ws.cell(row=r, column=c).value, str)
                        and ws.cell(row=r, column=c).value.strip()), "")
            if not lab or not any(k in lab for k in SCEN):
                continue
            f = ws.cell(row=r, column=6).value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            # シナリオ名を名乗る行が、ケース列を持つシートの単一列を引いている
            m = re.search(r"'([^']+)'!\$?([A-H])\$?(\d+)", f)
            if m and m.group(1) in wb.sheetnames:
                src = wb[m.group(1)]
                if [src.cell(row=5, column=c).value
                        for c in range(6, 9)] == list(CASE_COL):
                    mixed.append(f"{ws.title}!{r} {lab[:24]}")
            elif is_case_sheet:
                mixed.append(f"{ws.title}!{r} {lab[:24]}")
    gate(not mixed,
         f"列軸の値にシナリオ名を付けない（違反{len(mixed)}件）",
         str(mixed[:4]))

    # --- チェック行は必ずクラス表示（■必達／□要説明）を持つ ---
    # 「（0=OK）」で終わる行は検算行。マーカーが無いと読み手が導出行と区別できず、
    # 「検算が通っている」ように見える導出行が生まれる。
    unmarked = []
    for ws in wb.worksheets:
        for r in range(6, ws.max_row + 1):
            lab = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                        if isinstance(ws.cell(row=r, column=c).value, str)
                        and ws.cell(row=r, column=c).value.strip()), "")
            if lab.endswith("（0=OK）") and not ("■" in lab or "□" in lab):
                unmarked.append(f"{ws.title}!{r} {lab[:20]}")
    gate(not unmarked,
         f"チェック行にクラス表示がある（欠落{len(unmarked)}件）",
         str(unmarked[:4]))
    # 改ページは印刷範囲の中にだけ置く（範囲外の改ページは死んだ設定）。
    # 結論（総合判定）が2頁目に孤立する事故を防ぐ。
    dead = []
    for ws in wb.worksheets:
        pa = str(ws.print_area or "").replace("$", "")
        m = re.search(r":([A-Z]{1,2})(\d+)$", pa) if pa else None
        last = int(m.group(2)) if m else ws.max_row
        for b in (ws.row_breaks.brk if ws.row_breaks else []):
            if b.id >= last:
                dead.append(f"{ws.title}!行{b.id}")
    gate(not dead, f"改ページは印刷範囲内（範囲外{len(dead)}件）", str(dead[:4]))


    # --- 提示レイヤー（読み手のための構造）のゲート ---
    # 契約は references/presentation_layer_spec.md。
    # 1) 全シートが「STEP n｜何を決めるか」を名乗る（どこから来た数字か分かる導線）
    # 2) 印刷面に業界の隠語を出さない（1語分からないと、その行の保証ごと失われる）
    # 3) 画面固定（237行の入力シートで見出しが消えない）
    JARGON = ("コークスクリュー", "ロールフォワード", "タイアウト", "FCFF",
              "DLOM", "アセットF", "Jカーブ", "Sim2Real", "ARPA", "CAC",
              "MOIC", "Rule of 40", "Magic Number", "Burn Multiple")
    missing_step, jargon_hits, nofreeze = [], [], []
    for ws in wb.worksheets:
        v3 = ws.cell(row=3, column=2).value
        if not (isinstance(v3, str) and v3.startswith("STEP ")):
            missing_step.append(ws.title)
        if ws.freeze_panes != "F6":
            nofreeze.append(ws.title)
        last = _print_end(ws)
        for r in range(6, last + 1):
            for c in (2, 3, 4):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                for w in JARGON:
                    # 「日本語（略号）」の形なら可（投資家の共通語は残す）。
                    # 裸で出したら不可（1語分からないと行の意味ごと失われる）。
                    if w in v and f"（{w}" not in v and f"({w}" not in v:
                        jargon_hits.append(f"{ws.title}!{r}:{w}")
    gate(not missing_step, f"全シートにSTEP行（欠落{len(missing_step)}）",
         str(missing_step[:4]))
    gate(not nofreeze, f"全シートで画面固定F6（欠落{len(nofreeze)}）",
         str(nofreeze[:4]))
    gate(not jargon_hits, f"印刷面に隠語なし（{len(jargon_hits)}件）",
         str(sorted(set(jargon_hits))[:5]))

    # 印刷面の青字入力行には根拠タグ（【記載】【仮置き】等）が必ず出ていること。
    # 出どころの分からない数字が、投資家の読む面に残ってはいけない。
    notag = []
    for ws in wb.worksheets:
        bcol = next((c for c in range(6, ws.max_column + 1)
                     if ws.cell(row=5, column=c).value == "根拠"), None)
        if not bcol:
            continue
        for r in range(6, _print_end(ws) + 1):
            blue = any(isinstance(ws.cell(row=r, column=c).value, (int, float))
                       and color_of(ws.cell(row=r, column=c)) == BLUE
                       for c in range(6, bcol))
            if blue and not ws.cell(row=r, column=bcol).value:
                notag.append(f"{ws.title}!{r}")
    gate(not notag, f"印刷面の入力に根拠タグ（欠落{len(notag)}件）",
         str(notag[:5]))

    print(f"PASS {sum(1 for _ in ok)} 項目")
    if fails:
        print(f"FAIL {len(fails)} 件:")
        for x in fails[:40]:
            print("  -", x)
        return 1
    print("全ゲート通過")
    return 0


def recalc_gate(path):
    """LibreOfficeで再計算し、#エラーゼロと総合判定OKを強制する。"""
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    if not Path(soffice).exists():
        print("FAIL recalc: sofficeが見つかりません（再計算ゲートは必須）")
        return 2
    with tempfile.TemporaryDirectory() as td:
        prof = Path(tempfile.mkdtemp(prefix="lo_")).as_uri()
        subprocess.run([soffice, f"-env:UserInstallation={prof}",
                        "--headless", "--calc", "--convert-to",
                        "xlsx", "--outdir", td, path],
                       check=True, capture_output=True, timeout=300)
        out = Path(td) / Path(path).name
        wb = openpyxl.load_workbook(out, data_only=True)
        errs = [f"{ws.title}!{c.coordinate}:{c.value}"
                for ws in wb.worksheets for row in ws.iter_rows()
                for c in row if isinstance(c.value, str)
                and c.value.startswith("#")]
        verdict = None
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                label = next((c.value for c in row[1:4]
                              if isinstance(c.value, str) and c.value.strip()),
                             None)
                if isinstance(label, str) and label.startswith("総合判定"):
                    verdict = next((c.value for c in row[5:]
                                    if c.value is not None), None)
        if errs:
            print(f"RECALC FAIL: 数式エラー {len(errs)}件: {errs[:5]}")
            return 2
        if verdict != "OK":
            print(f"RECALC FAIL: 総合判定 = {verdict or '検出不能'}")
            return 2
        print("recalcゲート通過（数式エラーゼロ・総合判定OK）")
        return 0


if __name__ == "__main__":
    code = main(sys.argv[1])
    if code == 0 and "--recalc" in sys.argv[2:]:
        code = recalc_gate(sys.argv[1])
    sys.exit(code)
