"""
ib_format.py — IB Design Language を openpyxl で適用するヘルパー集

Source of truth: references/00_design_guidelines.md §付録 B
Related references:
  - 00_design_guidelines.md (visual design canonical)
  - 01a_modeling_standards.md §4-5 (規範)
  - _terminology.md §1-3 (functional color / brand color / sheet naming SSoT)

Usage:
    from ib_format import apply_hard_input, apply_formula, IB_PALETTE, ...
    apply_hard_input(ws['B5'])
    write_section_header(ws, 'A1', 'Revenue Build')

License: internal (Act / startup-financial-modeling skill)
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Literal

from openpyxl.cell.cell import Cell
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================================
# 1. Functional Color Palette  (canonical = _terminology.md §1)
# ============================================================================

IB_HARD_INPUT = "0000FF"      # Blue — ユーザー入力セル
IB_FORMULA = "000000"         # Black — 数式セル
IB_LINK_INTRA = "008000"      # Green — シート間参照
IB_LINK_EXTERNAL = "FF0000"   # Red — 外部 file 参照
IB_COMMENT = "666666"         # Gray — 注記 (italic)
IB_INK = "2D332E"             # Body ink

# Convenience dict for routing
IB_PALETTE: dict[str, str] = {
    "hard_input": IB_HARD_INPUT,
    "formula": IB_FORMULA,
    "link_intra": IB_LINK_INTRA,
    "link_external": IB_LINK_EXTERNAL,
    "comment": IB_COMMENT,
    "ink": IB_INK,
}

# ============================================================================
# 2. Brand Colors  (canonical = _terminology.md §2)
# ============================================================================

BRAND_PRIMARY_DEEP = "004F49"   # Act Primary deep — Input タブ色 (§2.X canonical)
BRAND_PRIMARY = "008A80"        # Act Primary — Driver タブ色 (§2.X canonical)
BRAND_INK = "2D332E"            # Cover タイトル
BRAND_SURFACE = "ECE9E1"        # Cover 背景
BRAND_ACCENT = "ECC85A"         # Memo タブ色 (§2.X canonical) — 1 sheet 1 か所まで

# Status colors (DESIGN.md §Colors)
BRAND_WARNING = "D6913D"        # Warning orange — Check タブ色 (§2.X canonical)
BRAND_DANGER = "C04A4A"         # 既存 waterfall_neg と同色
BRAND_SUCCESS = "3F8F5E"

# Bank-specific (optional override)
BRAND_NAVY = "1F3A66"            # Cover タブ色 (§2.X canonical)
BRAND_GS_NAVY = "7399C6"
BRAND_MS_BLUE = "015DAA"
BRAND_LAZARD_ORANGE = "FF6E00"
BRAND_EVERCORE_NAVY = "001F3F"

# Output role color (slate, neutral)
BRAND_SLATE = "666666"           # Output タブ色 (§2.X canonical)

# ============================================================================
# 3. Backgrounds
# ============================================================================

BG_WHITE = "FFFFFF"
BG_TOTAL_BAND = "F2F2F2"        # 合計行の薄バンディング
BG_HEADER_BAND = "1F3A66"       # Section header の濃ネイビーバンド
BG_WORKING = "FFF9C4"           # WIP / TODO セル (yellow highlight)

# Hyperlink (Excel default Office Theme link color = #0563C1)
LINK_COLOR = "0563C1"           # internal/external hyperlink (matches Excel default)

# ----------------------------------------------------------------------------
# Chart palettes (Phase 6 Fix #8 — IB-compliant palette for openpyxl charts)
# Source of truth: _terminology.md §1-2 (functional + brand colors)
# ----------------------------------------------------------------------------

# Football field (Low / Mid / High) — navy / teal / accent
IB_CHART_COLORS_FOOTBALL: list[str] = ["1F3A66", "008A80", "ECC85A"]

# Bar chart default palette — navy / gray / accent
IB_CHART_COLORS_BAR: list[str] = ["1F3A66", "666666", "ECC85A"]

# Line chart default palette — teal / navy / accent / gray
IB_CHART_COLORS_LINE: list[str] = ["008A80", "1F3A66", "ECC85A", "666666"]

# Waterfall chart specific colors
IB_CHART_COLORS_WATERFALL_POS: str = "008A80"   # teal — positive contribution
IB_CHART_COLORS_WATERFALL_NEG: str = "C04A4A"   # danger — negative contribution
IB_CHART_COLORS_WATERFALL_TOTAL: str = "666666" # gray — totals/subtotals

# ============================================================================
# 4. Fonts  (Phase 6 補強: Calibri 11pt 基準 — Google Sheets default 互換)
# ============================================================================
# Rule (per user spec):
#   - 全 cell font は Calibri 11pt 基準。Google Sheets の default が Calibri な
#     ので手修正時に違和感がないこと。
#   - section header = 12pt bold
#   - cover title    = 14pt bold
#   - comment / footnote = 10pt italic
# Chart 内部 font (axis tick / title) は本 file の責務外。`apply_chart_palette`
# は色のみ操作するため影響なし。

FONT_FAMILY = "Calibri"
FONT_SIZE_BASE = 11        # 本文セルの基準
FONT_SIZE_SMALL = 10       # comment / footnote / unit label
FONT_SIZE_LARGE = 12       # section header
FONT_SIZE_TITLE = 14       # cover title
FONT_SIZE_TINY = 9         # chart axis tick (charts only — cells should not use this)

FONT_BODY = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_INK)
FONT_BODY_BOLD = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)
FONT_HARD_INPUT = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_HARD_INPUT)
FONT_FORMULA = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_FORMULA)
FONT_LINK_INTRA = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_LINK_INTRA)
FONT_LINK_EXTERNAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_LINK_EXTERNAL)
FONT_COMMENT = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
FONT_FOOTNOTE = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
FONT_SECTION_HEADER = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color=IB_INK)
FONT_TITLE = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, bold=True, color=IB_INK)
FONT_YEAR_HEADER = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)
FONT_SUBTOTAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)
FONT_GRAND_TOTAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)
FONT_COVER_TITLE = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, bold=True, color="FFFFFF")

# --- Backwards-compat aliases (旧名 — 1 release 維持) -----------------------
# external builders (three_statement_builder, valuation_builder, cap_table_builder,
# sanity_checks) が import 済の旧名を破壊しないため alias を提供。
FONT_SECTION = FONT_TITLE                # alias (旧 14pt → 新 14pt; sheet-top title 用途で使われている)
FONT_SUBSECTION = FONT_BODY_BOLD         # alias (旧 11pt bold → 新 11pt bold)
FONT_TOTAL = FONT_SUBTOTAL               # alias (旧 10pt bold → 新 11pt bold)

# ============================================================================
# 5. Number Formats  (Phase 6 補強: 厳格な単位/通貨/負数表示)
# ============================================================================
# User spec rules:
#   - 基本単位 = 円 (cell value は必ず円単位の生数値で保存)
#   - 千円 / 百万円 / 億円は number_format で表現 (cell value は変えない)
#   - 負数は赤文字  ([Red] section)
#   - 通貨記号 (¥, $) を表示形式に含める
#
# Excel section syntax: positive;negative;zero[;text]
#   - "_)" は positive 後の右余白 (右括弧分の幅を空けて、negative の括弧囲みと
#     右端揃えになるようにする — IB convention)
#   - 数値の "," は 3 桁区切り、フォーマット末尾の "," は表示倍率 ÷1,000
#     (二重の ",," は ÷1,000,000 = 百万円表示)
# ----------------------------------------------------------------------------

# --- JPY (¥) -----------------------------------------------------------------
FMT_JPY_YEN              = '¥#,##0_);[Red](¥#,##0);"-"_)'
FMT_JPY_THOUSAND         = '¥#,##0,_);[Red](¥#,##0,);"-"_)'      # 千円表示
FMT_JPY_MILLION          = '¥#,##0,,_);[Red](¥#,##0,,);"-"_)'    # 百万円表示
# 億円: ",," = ÷1M (百万単位の表示) + "億" 接尾。厳密な ÷1億 を Excel の数値
# format で実現する手段は存在しない (comma scale は 3 桁単位)。よって
# 「百万円のまま億接尾」となる近似である点に caller は注意。
# 用途: ざっくり大まかな表示用 (例: "¥1,234,億" → 実数は 1,234M = 1,234百万円)
FMT_JPY_HUNDRED_MILLION  = '¥#,##0,,_)"億";[Red](¥#,##0,,)"億";"-"_)'

# --- USD ($) -----------------------------------------------------------------
FMT_USD_DOLLAR           = '"$"#,##0_);[Red]("$"#,##0);"-"_)'
FMT_USD_THOUSAND         = '"$"#,##0,_);[Red]("$"#,##0,);"-"_)'
FMT_USD_MILLION          = '"$"#,##0,,_);[Red]("$"#,##0,,);"-"_)'

# --- 通貨記号なし (純粋数値、3 桁区切り、負数 赤) ---------------------------
FMT_NUM                  = '#,##0_);[Red](#,##0);"-"_)'
FMT_NUM_THOUSAND         = '#,##0,_);[Red](#,##0,);"-"_)'
FMT_NUM_MILLION          = '#,##0,,_);[Red](#,##0,,);"-"_)'

# --- 比率 / パーセント (負数 赤) ---------------------------------------------
FMT_PCT_0                = '0%_);[Red](0%);"-"_)'
FMT_PCT_1                = '0.0%_);[Red](0.0%);"-"_)'
FMT_PCT_2                = '0.00%_);[Red](0.00%);"-"_)'

# --- 倍数 (x 表示、負数 赤) --------------------------------------------------
FMT_MULTIPLE_1           = '0.0"x"_);[Red](0.0"x");"-"_)'
FMT_MULTIPLE_2           = '0.00"x"_);[Red](0.00"x");"-"_)'

# --- Date --------------------------------------------------------------------
FMT_DATE_YMD             = "yyyy-mm-dd"
FMT_DATE_YM              = "yyyy-mm"

# --- 整数 (株式数等、3 桁区切り、負数 赤) ------------------------------------
FMT_INTEGER              = '#,##0_);[Red](#,##0);"-"_)'

# --- 株価 (小数 2 位) --------------------------------------------------------
FMT_PER_SHARE            = '#,##0.00_);[Red](#,##0.00);"-"_)'

# --- Backwards-compat aliases (旧名) -----------------------------------------
# 旧 builder の import 済名前を破壊しないため alias を提供。
# 既定通貨/scale が JPY × million なので、旧 FMT_MONEY 系は新 FMT_JPY_MILLION
# を指す。USD ベースのモデルは呼び出し側で `fmt=FMT_USD_MILLION` を明示指定。
FMT_MONEY                = FMT_JPY_MILLION              # alias (¥M ≒ 百万円表示)
FMT_MONEY_DECIMAL        = '¥#,##0.0,,_);[Red](¥#,##0.0,,);"-"_)'  # 百万円 1 桁
FMT_PERCENT              = FMT_PCT_1                    # alias (1 桁%)
FMT_PERCENT_BPS          = FMT_PCT_2                    # alias (2 桁%)
FMT_MULTIPLE             = FMT_MULTIPLE_1               # alias (1 桁 x)
FMT_SHARES               = FMT_INTEGER                  # alias
FMT_DATE_SHORT           = "mmm-yy"
FMT_DATE_LONG            = FMT_DATE_YMD
FMT_DEFAULT              = FMT_NUM_MILLION              # DEPRECATED: prefer FMT_NUM_MILLION 直接

# ============================================================================
# 6. Layout Constants  (Web Design 由来 token system: T-shirt size + modular)
# ============================================================================
# Spacing scale: Material Design-style T-shirt sizing applied to Excel row/col.
# Ratio ≒ 1.25 (Major Third) on row heights — 15 / 18 / 22 / 32 ≈ 1.0 / 1.2 /
# 1.47 / 2.13 (relaxed). column widths follow {tiny, small, base, large, xl}
# ramp consistent with B(label)=40 / C(unit)=8 / D-(period)=16 baseline.

# Column widths
COL_MARGIN_WIDTH = 20.0         # 左端マージン (Phase 6 Fix #2: was 2.0; 全シート見やすさ向上)
COL_LABEL_WIDTH = 40.0          # 行ラベル
COL_UNIT_WIDTH = 8.0            # Unit (¥M / %)
COL_PERIOD_WIDTH = 16.0         # 各期 (Y1 / Q1 / Jan-26)  (Phase 6 Fix #4: was 12.0)

# Column width T-shirt scale (auxiliary; for non-canonical layouts)
COL_WIDTH_TINY = 6.0
COL_WIDTH_SMALL = 12.0
COL_WIDTH_BASE = 16.0           # = COL_PERIOD_WIDTH
COL_WIDTH_LARGE = 20.0          # = COL_MARGIN_WIDTH
COL_WIDTH_XLARGE = 28.0

# Row heights (canonical — 既存 builder import 互換)
ROW_BODY_HEIGHT = 15.0
ROW_SPACER_HEIGHT = 8.0
ROW_SECTION_HEIGHT = 22.0

# Row height T-shirt scale (補助 token. 既存 ROW_*_HEIGHT を踏襲)
ROW_HEIGHT_TIGHT = 14.0         # 詰め (sub-section row)
ROW_HEIGHT_BASE = 15.0          # = ROW_BODY_HEIGHT
ROW_HEIGHT_RELAXED = 22.0       # = ROW_SECTION_HEIGHT
ROW_HEIGHT_HERO = 32.0          # cover title 等 (modular ratio ~2.13x)

# ============================================================================
# 7. Border styles  (subtotal / grand total / section divider / box / table)
# ============================================================================
# IB Border convention (canonical = Macabacus / WSP / IB pitchbook):
#   - 小計 (subtotal)         : top = thin black single
#   - 合計 (grand total)      : top = thin black single, bottom = thin double
#   - Section 区切り          : bottom = thin black
#   - Section 末 medium       : bottom = medium black (sub-section ⇒ section の境界)
#   - Sensitivity / call-out  : 4 辺 = thin black (box)
#   - Heavy emphasis (base case): 4 辺 = medium black

# Sides — line weight × color (gray は補助、IB convention は黒基本)
THIN_LINE = Side(border_style="thin", color=IB_FORMULA)
MEDIUM_LINE = Side(border_style="medium", color=IB_FORMULA)
THICK_LINE = Side(border_style="thick", color=IB_FORMULA)
DOUBLE_LINE = Side(border_style="double", color=IB_FORMULA)
HAIRLINE_GRAY = Side(border_style="hair", color=IB_COMMENT)       # data table 内部 hairline
THIN_GRAY = Side(border_style="thin", color=IB_COMMENT)           # 補助仕切り

BORDER_SUBTOTAL = Border(top=THIN_LINE)                           # 小計: 上 single
BORDER_GRAND_TOTAL = Border(top=THIN_LINE, bottom=DOUBLE_LINE)    # 合計: 上 single + 下 double
BORDER_SECTION_DIVIDER = Border(bottom=THIN_LINE)                 # Section 区切り
BORDER_SECTION_END_MEDIUM = Border(bottom=MEDIUM_LINE)            # Section 末 (medium)
BORDER_TOP_THIN = Border(top=THIN_LINE)
BORDER_BOTTOM_THIN = Border(bottom=THIN_LINE)
BORDER_BOTTOM_DOUBLE = Border(bottom=DOUBLE_LINE)
BORDER_BOX_THIN = Border(top=THIN_LINE, bottom=THIN_LINE, left=THIN_LINE, right=THIN_LINE)
BORDER_BOX_MEDIUM = Border(top=MEDIUM_LINE, bottom=MEDIUM_LINE, left=MEDIUM_LINE, right=MEDIUM_LINE)
BORDER_BOX_THICK = Border(top=THICK_LINE, bottom=THICK_LINE, left=THICK_LINE, right=THICK_LINE)

# ============================================================================
# 8. Page Setup
# ============================================================================

PRINT_MARGIN_INCH = 0.5
PRINT_HEADER_INCH = 0.3
PRINT_ORIENTATION_LANDSCAPE = "landscape"
PRINT_ORIENTATION_PORTRAIT = "portrait"

# ============================================================================
# 9. Sheet Naming  (canonical = _terminology.md §3)
# ============================================================================

# Phase 6 Stage A (17 → 14 sheet refactor):
#   Removed: 02_Drivers (→ content embedded in 11_KPI_Dashboard / 02_Revenue / 03_OpEx)
#            08_WC (→ embedded as sub-section "Working Capital Schedule" in 05_BS)
#            13_Sensitivity (→ embedded as sub-section "Sensitivity Analysis" in 09_DCF)
#   Renumbered: 03_Revenue→02_Revenue, 04_OpEx→03_OpEx, 05_IS→04_IS, 06_BS→05_BS,
#               07_CFS→06_CFS, 09_Debt→07_Debt, 10_CapTable→08_CapTable,
#               11_DCF→09_DCF, 12_Comps→10_Comps, 14_KPI_Dashboard→11_KPI_Dashboard,
#               15_SanityChecks→12_SanityChecks, 16_IC_Memo→13_IC_Memo
CANONICAL_SHEET_ORDER: tuple[str, ...] = (
    "00_Cover",
    "01_Assumptions",
    "02_Revenue",
    "03_OpEx",
    "04_IS",
    "05_BS",
    "06_CFS",
    "07_Debt",
    "08_CapTable",
    "09_DCF",
    "10_Comps",
    "11_KPI_Dashboard",
    "12_SanityChecks",
    "13_IC_Memo",
)

OPTIONAL_SHEETS: tuple[str, ...] = ("99_Glossary",)

# Sheets removed in Phase 6 Stage A. Loaders that encounter these names in
# legacy xlsx files should warn and ignore (or migrate via _OLD_TO_NEW_SHEET_MAP).
DEPRECATED_SHEETS: frozenset[str] = frozenset({
    "02_Drivers",
    "08_WC",
    "13_Sensitivity",
})

# Mapping from legacy 17-sheet names to the new 14-sheet canonical names.
# Used by sanity_checks._resolve_sheet() for migration warnings.
_OLD_TO_NEW_SHEET_MAP: dict[str, str] = {
    "03_Revenue":       "02_Revenue",
    "04_OpEx":          "03_OpEx",
    "05_IS":            "04_IS",
    "06_BS":            "05_BS",
    "07_CFS":           "06_CFS",
    "09_Debt":          "07_Debt",
    "10_CapTable":      "08_CapTable",
    "11_DCF":           "09_DCF",
    "12_Comps":         "10_Comps",
    "14_KPI_Dashboard": "11_KPI_Dashboard",
    "15_SanityChecks":  "12_SanityChecks",
    "16_IC_Memo":       "13_IC_Memo",
    # Removed sheets — content moved into the listed target sheet.
    "02_Drivers":       "11_KPI_Dashboard",
    "08_WC":            "05_BS",
    "13_Sensitivity":   "09_DCF",
}

# ============================================================================
# 10. Cell-level helper functions
# ============================================================================


def apply_hard_input(
    cell,
    fmt: str = FMT_JPY_MILLION,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """Hard input セルに IB blue + 数値書式を適用 (wrap_text=False, Phase 6 Fix #1).

    Phase 6 補強: default を `FMT_JPY_MILLION` (¥ 百万円表示、円単位生 value、負数 赤)
    に変更。USD ベースのモデルでは呼び出し側で fmt=FMT_USD_MILLION を明示指定する。

    Phase 6 Wave 2.1 拡張: `named_range` kwarg を渡すと自動で named range を登録する
    (canonical 名前は references/_named_ranges.md §2)。`fmt` は positional 互換のため
    cell 直後に維持し、`named_range` / `scope` は keyword-only。

    Args:
        cell: openpyxl Cell
        fmt: number_format (positional 互換 — 既存 caller の互換性のため)
        named_range: §2 canonical name (None でスキップ、kwarg-only)
        scope: "workbook" (default) / "sheet" (kwarg-only)
    """
    cell.font = FONT_HARD_INPUT
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_formula(
    cell,
    fmt: str = FMT_JPY_MILLION,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """Formula セルに black + 数値書式を適用 (wrap_text=False).

    Phase 6 補強: default を `FMT_JPY_MILLION` に変更。
    Phase 6 Wave 2.1: `named_range` kwarg で自動 named range 登録。
    """
    cell.font = FONT_FORMULA
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_link_intra(
    cell,
    fmt: str = FMT_JPY_MILLION,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """Cross-sheet link セル (=03_Revenue!B5 等) に green を適用 (wrap_text=False).

    Phase 6 補強: default を `FMT_JPY_MILLION` に変更。
    Phase 6 Wave 2.1: `named_range` kwarg で自動 named range 登録。
    """
    cell.font = FONT_LINK_INTRA
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_link_external(
    cell,
    fmt: str = FMT_JPY_MILLION,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """External link (別 file 参照) に red を適用 (wrap_text=False).

    Phase 6 補強: default を `FMT_JPY_MILLION` に変更。
    Phase 6 Wave 2.1: `named_range` kwarg で自動 named range 登録。
    """
    cell.font = FONT_LINK_EXTERNAL
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_label(cell, indent: int = 0, bold: bool = False, wrap_text: bool = False) -> None:
    """行ラベル (左寄せ、indent あり、Calibri 11pt).

    Phase 6 Fix #1: wrap_text=False が default。長文は隣接セルにはみ出して表示する
    (IB 慣習)。明示的に wrap_text=True を指定したいときのみ折り返す。
    Phase 6 補強: font は Calibri 11pt 固定 (Google Sheets default 互換)。
    """
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=bold, color=IB_INK)
    cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=indent, wrap_text=wrap_text
    )


def apply_label_indent(cell, level: int = 0, bold: bool = False) -> None:
    """階層 line item の indent (Excel native indent). Phase 6 Fix #3.

    level 0 → 通常ラベル。level 1+ → Excel の indent で字下げ表現。
    列シフト方式と併用可能 (apply_label_indent は label セル自身に適用)。
    Phase 6 補強: font は Calibri 11pt 固定。
    """
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=bold, color=IB_INK)
    cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=level, wrap_text=False
    )


def write_hierarchical_line_item(
    ws: Worksheet,
    row: int,
    level: int,
    label: str,
    value_start_col: int = 4,
    bold: bool = False,
) -> int:
    """階層 line item を書き込む helper. Phase 6 Fix #3.

    親項目 (level=0) → B 列 (col 2)
    子項目 (level=1) → C 列 (col 3)
    孫項目 (level=2) → D 列 (col 4)  (注: 既定 PERIOD_START_COL=4 と衝突するため、
                                       value_start_col を上げる builder 側で対応)

    Args:
        ws: Worksheet
        row: 1-based row number
        level: 0/1/2  階層深さ
        label: ラベル文字列
        value_start_col: 数値が始まる列 (default=4=D)。level=2 を使うときは builder
            側で value_start_col を 1 つ右にずらすこと。
        bold: True で太字 (親項目で使う)

    Returns:
        実際にラベルを書いた列 (col_index)
    """
    label_col = 2 + level  # B(2), C(3), D(4)
    cell = ws.cell(row=row, column=label_col)
    cell.value = label
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=bold, color=IB_INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    return label_col


def set_uniform_column_width(
    ws: Worksheet, columns: Iterable[str], width: float
) -> None:
    """指定列を一律 width に設定. Phase 6 Fix #4.

    Args:
        ws: Worksheet
        columns: 列文字 iterable, e.g. ['C', 'D', 'E', 'F']
        width: openpyxl 列幅 (≈ 文字数)
    """
    for col_letter in columns:
        ws.column_dimensions[col_letter].width = width


def apply_section_header(cell, label: str) -> None:
    """Section header (濃ネイビー bg + 白文字 bold).

    Phase 6 補強: Calibri 12pt bold (FONT_SIZE_LARGE)。
    """
    cell.value = label
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BG_HEADER_BAND)
    cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=1, wrap_text=False
    )


def apply_section_header_band(
    ws: Worksheet,
    *,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
) -> None:
    """Section header band (No-Merge variant, Phase 6 補強).

    `apply_section_header` を leftmost cell に適用し、navy fill を
    end_col まで propagate する (merge_cells を使わずに 1 行分の band を表現)。

    Why no merge:
        merge_cells は sort / filter / range select / formula reference /
        named range / fill drag を破壊するため (IB best practice / Macabacus・
        WSP 推奨). 同等の見た目は fill propagation で実現する。

    See: references/_design_consistency_rules.md §4.7 No-Merge Rule

    Args:
        ws: Worksheet
        row: header row (1-indexed)
        start_col: leftmost column index (1-indexed, value はここに置く)
        end_col: rightmost column index (1-indexed, fill はここまで)
        label: section header 文字列
    """
    if end_col < start_col:
        raise ValueError(
            f"apply_section_header_band: end_col ({end_col}) must be >= "
            f"start_col ({start_col})"
        )
    head = ws.cell(row=row, column=start_col)
    apply_section_header(head, label)
    if end_col == start_col:
        return
    band_fill = PatternFill("solid", fgColor=BG_HEADER_BAND)
    band_font = Font(
        name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color="FFFFFF"
    )
    band_align = Alignment(
        horizontal="left", vertical="center", wrap_text=False
    )
    for col in range(start_col + 1, end_col + 1):
        c = ws.cell(row=row, column=col)
        # value は左端のみ (No-Merge 原則)
        c.fill = band_fill
        c.font = band_font
        c.alignment = band_align


def apply_year_header(cell, label: str) -> None:
    """期ヘッダ (Y1 / Q1 / Jan-26 等、italic + bold)."""
    cell.value = label
    cell.font = FONT_YEAR_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def apply_subtotal(
    cell,
    fmt: str = FMT_MONEY,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """小計セル (top single border + bold).

    Phase 6 Wave 2.1: `named_range` kwarg で自動 named range 登録。
    """
    cell.font = FONT_TOTAL
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    cell.border = BORDER_SUBTOTAL
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_grand_total(
    cell,
    fmt: str = FMT_MONEY,
    *,
    named_range: str | None = None,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> None:
    """合計セル (top single + bottom double + bold + total band).

    Phase 6 Wave 2.1: `named_range` kwarg で自動 named range 登録。
    """
    cell.font = FONT_TOTAL
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    cell.border = BORDER_GRAND_TOTAL
    cell.fill = PatternFill("solid", fgColor=BG_TOTAL_BAND)
    if named_range is not None:
        register_named_range(cell, named_range, scope=scope)


def apply_comment(cell, wrap_text: bool = False) -> None:
    """注記セル (gray + italic).

    Phase 6 Fix #1: wrap_text=False が default。長文は隣接の空きセルにはみ出して
    表示するのが IB 慣習。明示的に折り返したい場合は wrap_text=True を渡す。
    """
    cell.font = FONT_COMMENT
    cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=wrap_text
    )


def apply_working_highlight(cell) -> None:
    """WIP / TODO セル (yellow highlight)."""
    cell.fill = PatternFill("solid", fgColor=BG_WORKING)


def apply_chart_palette(chart, palette: list[str]) -> None:
    """各 series に palette の色を順に循環適用. Phase 6 Fix #8.

    palette 例: IB_CHART_COLORS_BAR, IB_CHART_COLORS_LINE 等。
    series 数 > palette 長さなら modulo で循環。

    Args:
        chart: openpyxl chart instance (BarChart / LineChart / etc.)
        palette: hex color string list (例 ["1F3A66", "008A80", ...])
    """
    for i, series in enumerate(chart.series):
        color = palette[i % len(palette)]
        try:
            series.graphicalProperties = GraphicalProperties(solidFill=color)
        except Exception:
            # Fallback: 一部 chart 種では graphicalProperties 経路で fail するため、
            # silent fallthrough (caller 側で別経路適用するときの邪魔をしない).
            pass


# ----------------------------------------------------------------------------
# 10b. Unit label / currency helpers  (Phase 6 補強)
# ----------------------------------------------------------------------------


def fmt_for_currency(currency: str = "JPY", scale: str = "million") -> str:
    """指定通貨 × scale の number_format を返す.

    cell value は常に基本単位 (円 / dollar) の生数値で保存し、表示は
    number_format で切り替える設計。

    Args:
        currency: "JPY" / "USD"
        scale: "actual" (生値) / "thousand" (千) / "million" (百万) /
               "hundred_million" (億, JPY のみ; arithmetic は百万単位)

    Returns:
        Excel number_format 文字列。マッチしない場合は ``FMT_NUM_MILLION``。
    """
    table: dict[tuple[str, str], str] = {
        ("JPY", "actual"): FMT_JPY_YEN,
        ("JPY", "thousand"): FMT_JPY_THOUSAND,
        ("JPY", "million"): FMT_JPY_MILLION,
        ("JPY", "hundred_million"): FMT_JPY_HUNDRED_MILLION,
        ("USD", "actual"): FMT_USD_DOLLAR,
        ("USD", "thousand"): FMT_USD_THOUSAND,
        ("USD", "million"): FMT_USD_MILLION,
    }
    return table.get((currency, scale), FMT_NUM_MILLION)


_UNIT_LABEL_MAP: dict[tuple[str, str], str] = {
    ("JPY", "actual"): "(単位: 円)",
    ("JPY", "thousand"): "(単位: 千円)",
    ("JPY", "million"): "(単位: 百万円)",
    ("JPY", "hundred_million"): "(単位: 億円)",
    ("USD", "actual"): "(Unit: $)",
    ("USD", "thousand"): "(Unit: $K)",
    ("USD", "million"): "(Unit: $M)",
}


def apply_unit_label(cell, currency: str = "JPY", scale: str = "million") -> None:
    """シート右上等に「(単位: 百万円)」「(Unit: $M)」のような単位明示ラベルを書く.

    cell.value をこの関数が直接設定する点に注意 (caller は cell の選択のみ)。
    font は Calibri 10pt italic gray (FONT_COMMENT 同等)。

    Args:
        cell: 単位ラベルを表示するセル (sheet header 右上等)
        currency: "JPY" / "USD"
        scale: "actual" / "thousand" / "million" / "hundred_million"
    """
    cell.value = _UNIT_LABEL_MAP.get((currency, scale), "(単位: 不明)")
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)


# ----------------------------------------------------------------------------
# 10c. Border helpers  (Phase 6 補強: section / box / table 共通枠線)
# ----------------------------------------------------------------------------


def _merge_border(existing: Border | None, *, top=None, bottom=None, left=None, right=None) -> Border:
    """既存 Border に新しい side を上書きする (None の side は保持)."""
    if existing is None:
        return Border(top=top, bottom=bottom, left=left, right=right)
    return Border(
        top=top if top is not None else existing.top,
        bottom=bottom if bottom is not None else existing.bottom,
        left=left if left is not None else existing.left,
        right=right if right is not None else existing.right,
    )


def apply_section_bottom_border(cell, *, double: bool = False, weight: str = "medium") -> None:
    """セクション末 (subtotal / grand total / sub-section 区切り) の bottom border.

    Args:
        cell: 適用先 cell
        double: True で grand total の二重線 (medium/thick より優先)
        weight: "thin" / "medium" / "thick" (double=False のとき有効)
    """
    if double:
        side = DOUBLE_LINE
    else:
        side = {"thin": THIN_LINE, "medium": MEDIUM_LINE, "thick": THICK_LINE}.get(weight, MEDIUM_LINE)
    cell.border = _merge_border(cell.border, bottom=side)


def apply_box_border(
    ws: Worksheet,
    range_str: str,
    *,
    weight: str = "thin",
    inner_hairline: bool = False,
) -> None:
    """範囲全体に枠線 (sensitivity matrix / call-out box / data table 用).

    Args:
        ws: Worksheet
        range_str: "B5:G12" 形式 or "B5" (単一 cell)
        weight: "thin" / "medium" / "thick" — 外枠の線種
        inner_hairline: True で内部の cell 間に gray hairline を引く (data table 用)

    Note:
        既存 cell.border は上書きされる外枠 side のみ更新し、他の side は保持される。
    """
    side_map = {"thin": THIN_LINE, "medium": MEDIUM_LINE, "thick": THICK_LINE}
    outer = side_map.get(weight, THIN_LINE)
    min_col, min_row, max_col, max_row = range_boundaries(range_str)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            top = outer if r == min_row else (HAIRLINE_GRAY if inner_hairline else None)
            bottom = outer if r == max_row else (HAIRLINE_GRAY if inner_hairline else None)
            left = outer if c == min_col else (HAIRLINE_GRAY if inner_hairline else None)
            right = outer if c == max_col else (HAIRLINE_GRAY if inner_hairline else None)
            cell.border = _merge_border(cell.border, top=top, bottom=bottom, left=left, right=right)


# ----------------------------------------------------------------------------
# 10d. Conditional formatting helpers  (Phase 6 補強)
# ----------------------------------------------------------------------------
# 用途: sensitivity matrix / cohort retention / KPI heatmap 等
# 色は IB convention に寄せる (red→yellow→green or 中立 grayscale)
# ※ caller は range_str に sheet 内の data range のみを指定 (header 行/列は除く)
# ----------------------------------------------------------------------------

# Heatmap colors — color-blind safer ramp (青→黄→赤 は deuteranopia でも判別可)
HEATMAP_LOW_COOL = "5A8FBF"     # cool blue (低)
HEATMAP_MID_NEUTRAL = "F5F5F5"  # neutral light gray (中)
HEATMAP_HIGH_WARM = "C04A4A"    # warm red (高)

# Excel-classic green-yellow-red (IB sensitivity 表で使われる伝統的 palette)
HEATMAP_GREEN = "63BE7B"
HEATMAP_YELLOW = "FFEB84"
HEATMAP_RED = "F8696B"


def apply_heatmap_3color(
    ws: Worksheet,
    range_str: str,
    *,
    low_color: str = HEATMAP_GREEN,
    mid_color: str = HEATMAP_YELLOW,
    high_color: str = HEATMAP_RED,
    invert: bool = False,
) -> None:
    """sensitivity / cohort retention / KPI matrix 用の 3 色 heatmap.

    Args:
        ws: Worksheet
        range_str: "B5:G12" 形式の data range (header 行/列は含めない)
        low_color/mid_color/high_color: 6-hex (#は不要)
        invert: True で low/high を逆転 (低が悪い指標で使う; e.g. churn は低=良 だが NPS は高=良)
    """
    if invert:
        low_color, high_color = high_color, low_color
    rule = ColorScaleRule(
        start_type="min", start_color=low_color,
        mid_type="percentile", mid_value=50, mid_color=mid_color,
        end_type="max", end_color=high_color,
    )
    ws.conditional_formatting.add(range_str, rule)


def apply_data_bar(ws: Worksheet, range_str: str, *, color: str = BRAND_PRIMARY) -> None:
    """data bar conditional format (KPI dashboard で進捗 bar として使う).

    Args:
        ws: Worksheet
        range_str: data range
        color: 6-hex bar color (default = Act Primary teal)
    """
    rule = DataBarRule(
        start_type="min", end_type="max", color=color,
        showValue=True,
    )
    ws.conditional_formatting.add(range_str, rule)


# ----------------------------------------------------------------------------
# 10e. Data validation  (Phase 6 補強: dropdown / numeric range)
# ----------------------------------------------------------------------------


def apply_dropdown(
    ws: Worksheet,
    range_str: str,
    options: list[str],
    *,
    allow_blank: bool = True,
    prompt_title: str | None = None,
    prompt: str | None = None,
) -> None:
    """セル範囲に dropdown を付ける (Mode / Currency / Scale / Case 選択用).

    Args:
        ws: Worksheet
        range_str: "C3" or "C3:C20" 等
        options: dropdown に表示する選択肢 (例 ["Base", "Bull", "Bear"])
        allow_blank: 空白を許容するか
        prompt_title: cell 選択時 tooltip タイトル
        prompt: cell 選択時 tooltip 本文

    Note:
        Excel の list-formula は 255 文字制限あり。長い list は別 sheet に書き、
        "=Lists!$A$2:$A$N" 形式で参照すること (本 helper は in-line list のみ対応)。
    """
    # double quote escape (Excel formula は ' ではなく " が引用符)
    escaped = [opt.replace('"', '""') for opt in options]
    formula = f'"{",".join(escaped)}"'
    if len(formula) > 255:
        raise ValueError(
            f"dropdown options too long ({len(formula)} > 255 chars). "
            "Use external list reference instead."
        )
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False,
    )
    if prompt_title:
        dv.promptTitle = prompt_title
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    dv.add(range_str)
    ws.add_data_validation(dv)


def apply_numeric_validation(
    ws: Worksheet,
    range_str: str,
    *,
    minimum: float | None = None,
    maximum: float | None = None,
    integer: bool = False,
    error_message: str | None = None,
) -> None:
    """数値範囲制約を設定 (例: % は 0..1, 月数は 0..120, 株数は 0+ 整数).

    Args:
        ws: Worksheet
        range_str: 対象範囲
        minimum / maximum: 上下限 (None で片側制約)
        integer: True で整数のみ
        error_message: 違反時のエラー
    """
    if minimum is not None and maximum is not None:
        operator, formula1, formula2 = "between", str(minimum), str(maximum)
    elif minimum is not None:
        operator, formula1, formula2 = "greaterThanOrEqual", str(minimum), None
    elif maximum is not None:
        operator, formula1, formula2 = "lessThanOrEqual", str(maximum), None
    else:
        return  # nothing to validate

    dv = DataValidation(
        type="whole" if integer else "decimal",
        operator=operator,
        formula1=formula1,
        formula2=formula2,
        allow_blank=True,
    )
    if error_message:
        dv.error = error_message
        dv.errorTitle = "Invalid input"
        dv.showErrorMessage = True
    dv.add(range_str)
    ws.add_data_validation(dv)


# ----------------------------------------------------------------------------
# 10f. Hyperlink helpers  (Phase 6 補強: Cover sheet ToC navigation)
# ----------------------------------------------------------------------------


def apply_internal_link(
    cell,
    *,
    target_sheet: str,
    target_cell: str = "A1",
    display_text: str | None = None,
) -> None:
    """セルを内部 sheet jump link 化 (Cover の ToC や cross-sheet ナビ用).

    Args:
        cell: 適用先 cell
        target_sheet: 飛び先 sheet 名 (例 "03_Revenue")
        target_cell: 飛び先 cell (default "A1")
        display_text: 表示テキスト (省略時は target_sheet 名)

    Excel hyperlink 形式: ``#'SheetName'!A1`` (シート名は単引用符)。
    """
    cell.value = display_text if display_text is not None else target_sheet
    cell.hyperlink = f"#'{target_sheet}'!{target_cell}"
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=LINK_COLOR, underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def write_toc(
    ws: Worksheet,
    sheet_names: Iterable[str],
    *,
    start_row: int = 10,
    label_col: str = "B",
    title: str = "Contents",
) -> int:
    """Cover シートに目次 (Table of Contents) を書く.

    Args:
        ws: Cover sheet
        sheet_names: 飛び先 sheet 名の iterable (e.g. CANONICAL_SHEET_ORDER[1:])
        start_row: 目次の開始 row
        label_col: 目次 label 列 (default "B")
        title: 目次の見出し

    Returns:
        目次最後の row (caller が次行を続けて書ける)
    """
    title_cell = ws[f"{label_col}{start_row}"]
    title_cell.value = title
    title_cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color=IB_INK)

    row = start_row + 1
    for name in sheet_names:
        cell = ws[f"{label_col}{row}"]
        apply_internal_link(cell, target_sheet=name, target_cell="A1", display_text=name)
        row += 1
    return row - 1


# ----------------------------------------------------------------------------
# 10g. Row-height helper  (Phase 6 補強: T-shirt size mapping)
# ----------------------------------------------------------------------------


def apply_row_heights(ws: Worksheet, row_kind_map: dict[int, str]) -> None:
    """row → kind ("tight" / "base" / "relaxed" / "hero") を一括設定.

    Args:
        ws: Worksheet
        row_kind_map: {row_index: kind_name}
    """
    height_map = {
        "tight": ROW_HEIGHT_TIGHT,
        "base": ROW_HEIGHT_BASE,
        "relaxed": ROW_HEIGHT_RELAXED,
        "hero": ROW_HEIGHT_HERO,
    }
    for row, kind in row_kind_map.items():
        ws.row_dimensions[row].height = height_map.get(kind, ROW_HEIGHT_BASE)


# ----------------------------------------------------------------------------
# 10h. Named range helpers  (Phase 6 Wave 2.1 — references/_named_ranges.md)
# ----------------------------------------------------------------------------
# 行/列 insertion 耐性のため、cell anchor (`$C$5`) に代えて named range を
# 採用する。canonical 名前は references/_named_ranges.md §2 を参照。
#
# 設計原則 (詳細は §1):
#   - Workbook-scoped: 2+ sheet から参照される基幹値 (例 Rev_Total, IS_NI)
#   - Sheet-scoped:    1 sheet 内に閉じる中間値 (例 Rev_Subscription)
#   - 命名: PascalCase + Underscore section prefix (Rev_, IS_, BS_, Cap_, ...)
#   - 25 字以内、Excel 予約名 / cell address / 単一語 / 言語混在 を回避
#
# openpyxl 3.1.x canonical API (§5):
#   wb.defined_names.append(DefinedName(name=..., attr_text=..., localSheetId=...))
# ----------------------------------------------------------------------------

# Excel 予約名 (§3.1) — 登録は受け付けるが Excel 機能と混線するため厳禁。
_EXCEL_RESERVED_NAMES: frozenset[str] = frozenset({
    "Print_Area", "Print_Titles",
    "Auto_Open", "Auto_Close", "Auto_Activate", "Auto_Deactivate",
    "Database", "Criteria", "Extract",
    "Sheet_Title", "Consolidate_Area",
})

# 推奨 max length (§1.5)。Excel 公式は 255 字だが、IB / Macabacus 慣習 30 字、
# 本 skill は 25 字を canonical とする。
_NAMED_RANGE_MAX_LEN = 25


def _validate_named_range(name: str) -> None:
    """References §1.4 / §3 の NG パターン違反を検知して ValueError を raise.

    検証項目:
      - 1..25 文字 (空 / 25 字超 NG)
      - 先頭文字: letter or underscore (数字始まり / 記号始まり NG)
      - 全文字: alphanumeric or underscore のみ (hyphen / 空白 / 多言語 NG)
      - Excel 予約名 (§3.1) NG
      - Cell address 衝突の簡易検出 (§3.2; e.g. ``A1`` / ``Q1`` / ``FY24``)

    Note:
        case-insensitive な内部比較で "_var" のような hidden name 始まり
        パターンも検出する。ただし途中 underscore (``Cap_PPS_PostMoney``)
        は許容。連続 underscore (``Rev__Total``) は §3.4 で NG だが、
        パターン false-positive を避けるため明示的にはチェックしない
        (canonical table を逸脱した名前を呼び出し側で書かない前提)。
    """
    if not isinstance(name, str) or not name:
        raise ValueError("Named range must be a non-empty string")
    if len(name) > _NAMED_RANGE_MAX_LEN:
        raise ValueError(
            f"Named range '{name}' is {len(name)} chars; max {_NAMED_RANGE_MAX_LEN} "
            "(IB convention; references/_named_ranges.md §1.5)"
        )
    first = name[0]
    # ASCII-strict: Python's str.isalpha() / .isalnum() accept many non-ASCII
    # letters (e.g. 売上). _named_ranges.md §3.6 prohibits language mixing,
    # so restrict to ASCII letter / digit / underscore.
    if not ((first.isascii() and first.isalpha()) or first == "_"):
        raise ValueError(
            f"Named range '{name}' must start with an ASCII letter or underscore "
            "(Excel identifier rule; references/_named_ranges.md §1.3)"
        )
    if not all(ch == "_" or (ch.isascii() and ch.isalnum()) for ch in name):
        raise ValueError(
            f"Named range '{name}' must contain only ASCII alphanumeric and underscore "
            "(no hyphen/space/non-ASCII; references/_named_ranges.md §1.3, §3.6)"
        )
    if name in _EXCEL_RESERVED_NAMES:
        raise ValueError(
            f"Named range '{name}' is an Excel reserved name "
            "(references/_named_ranges.md §3.1)"
        )
    # Cell address 簡易衝突検出 (§3.2):
    #   pattern = "<letters><digits>" かつ全 ASCII (e.g. A1, Q1, FY24, IRR1)
    #   Excel 列は最大 XFD なので letters 1..3 字、digits 1..7 字を粗くチェック。
    if "_" not in name and len(name) <= 10:
        i = 0
        while i < len(name) and name[i].isalpha():
            i += 1
        if 1 <= i <= 3 and i < len(name) and name[i:].isdigit():
            raise ValueError(
                f"Named range '{name}' may collide with a cell address "
                f"(letters+digits pattern; references/_named_ranges.md §3.2). "
                "Add a section prefix (e.g. 'Period_Q1', 'Cell_A1')."
            )


def _abs_coord(cell_ref: str) -> str:
    """単 cell の coordinate (例 'B5' / '$B$5' / 'B5:G12') を absolute 表記に変換.

    range (':' を含む) は呼び出し側で完全な absolute ref を渡している前提なので
    そのまま返す。単 cell の場合は ``$<col>$<row>`` 形式を保証する。
    """
    if ":" in cell_ref or "$" in cell_ref:
        return cell_ref
    col = "".join(ch for ch in cell_ref if ch.isalpha())
    row = "".join(ch for ch in cell_ref if ch.isdigit())
    if not col or not row:
        # 既に絶対化されていない予期せぬ形式 — caller の責任。そのまま返す。
        return cell_ref
    return f"${col}${row}"


# NOTE on openpyxl 3.1.x API: `wb.defined_names` and `ws.defined_names` are
# `DefinedNameDict` (dict subclass) — they do NOT expose `.append()`. The
# canonical pattern is dict-like assignment:
#   - workbook-scoped: ``wb.defined_names[name] = DefinedName(name=name, attr_text=...)``
#   - sheet-scoped:    ``ws.defined_names[name] = DefinedName(name=name, attr_text=..., localSheetId=<idx>)``
# (`_named_ranges.md` §5 documents `.append()` which is for the older
# `DefinedNameList` Sequence; verified empirically that 3.1.5 raises
# ``AttributeError: 'DefinedNameDict' object has no attribute 'append'``.)


def register_workbook_name(
    wb: Workbook,
    name: str,
    sheet_title: str | None = None,
    cell_ref: str | None = None,
    *,
    cell: Cell | None = None,
) -> DefinedName:
    """Workbook-scoped named range を登録 (どの sheet からでも bare name 参照可).

    cross-sheet で参照される基幹値 (Rev_Total / IS_NI / BS_Cash 等) に使う。
    canonical API (openpyxl 3.1.x):
        wb.defined_names[name] = DefinedName(name=name, attr_text=...)

    Args:
        wb: openpyxl Workbook
        name: §2 canonical table から取った name
        sheet_title: 定義 cell の sheet 名 (cell 引数を渡す場合は不要)
        cell_ref: 絶対参照表記 ($B$5 等) — 未絶対化形式 (B5) も自動で $ 化
        cell: openpyxl Cell オブジェクト (sheet_title / cell_ref を自動推定)

    Returns:
        登録された DefinedName

    Raises:
        ValueError: name が canonical 規約違反 / cell + (sheet_title|cell_ref) の両方未指定

    Examples:
        >>> register_workbook_name(wb, "Rev_Total", "03_Revenue", "$C$5")
        >>> register_workbook_name(wb, "Rev_Total", cell=ws["C5"])
    """
    _validate_named_range(name)
    if cell is not None:
        sheet_title = cell.parent.title
        cell_ref = cell.coordinate
    if sheet_title is None or cell_ref is None:
        raise ValueError(
            "register_workbook_name requires either `cell` or both "
            "`sheet_title` and `cell_ref`"
        )
    full_ref = f"'{sheet_title}'!{_abs_coord(cell_ref)}"
    dn = DefinedName(name=name, attr_text=full_ref)
    wb.defined_names[name] = dn
    return dn


def register_sheet_name(
    ws: Worksheet,
    name: str,
    cell_ref: str | None = None,
    *,
    cell: Cell | None = None,
) -> DefinedName:
    """Sheet-scoped named range を登録 (同 sheet 内で bare name、他 sheet では qualified).

    sheet 内に閉じる中間値 (Rev_Subscription / ARR_New / WC_DSO 等) に使う。
    canonical API (openpyxl 3.1.x):
        ws.defined_names[name] = DefinedName(name=name, attr_text=..., localSheetId=ws_index)

    `localSheetId` を明示しなくても save 時に openpyxl が ws 位置から自動付与
    するが、明示することで in-memory inspection (loadworkbook 後と同等) でも
    一貫した値が得られる。

    Args:
        ws: openpyxl Worksheet (定義先 sheet)
        name: §2 canonical table から取った name
        cell_ref: 絶対参照表記 ($D$6 等) — 未絶対化形式も自動で $ 化
        cell: Cell オブジェクト (cell_ref を自動推定; ws と一致する前提)

    Returns:
        登録された DefinedName

    Raises:
        ValueError: name が canonical 規約違反 / cell も cell_ref も未指定
    """
    _validate_named_range(name)
    if cell is not None:
        cell_ref = cell.coordinate
    if cell_ref is None:
        raise ValueError(
            "register_sheet_name requires either `cell` or `cell_ref`"
        )
    wb = ws.parent
    full_ref = f"'{ws.title}'!{_abs_coord(cell_ref)}"
    sheet_id = wb.index(ws)
    dn = DefinedName(name=name, attr_text=full_ref, localSheetId=sheet_id)
    ws.defined_names[name] = dn
    return dn


def register_range_name(
    ws: Worksheet,
    name: str,
    range_str: str,
    *,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> DefinedName:
    """範囲 (multi-cell range) を name に bind して登録.

    Year_Headers / Rev_Total (range 形式) / DCF_FCFF (D11:AA11) 等に使う。

    Args:
        ws: openpyxl Worksheet (定義先 sheet)
        name: §2 canonical table から取った name
        range_str: 絶対参照表記 ($D$5:$AA$5 等) — caller が完全な absolute ref を用意
        scope: "workbook" (default) / "sheet"

    Returns:
        登録された DefinedName
    """
    _validate_named_range(name)
    wb = ws.parent
    full_ref = f"'{ws.title}'!{range_str}"
    if scope == "workbook":
        dn = DefinedName(name=name, attr_text=full_ref)
        wb.defined_names[name] = dn
    elif scope == "sheet":
        sheet_id = wb.index(ws)
        dn = DefinedName(name=name, attr_text=full_ref, localSheetId=sheet_id)
        ws.defined_names[name] = dn
    else:
        raise ValueError(f"Invalid scope: {scope!r} (expected 'workbook' or 'sheet')")
    return dn


def register_named_range(
    cell: Cell,
    name: str,
    *,
    scope: Literal["workbook", "sheet"] = "workbook",
) -> DefinedName:
    """Cell から named range を登録する thin wrapper (§5.3 canonical API).

    apply_* helper 内部で使う一段薄い entry point。caller code は通常
    `apply_hard_input(cell, named_range="Rev_Total")` のように apply_* 経由で
    呼び出すため直接は使わないが、定型 cell に formatting なしで name だけ
    付けたい場合に利用可。

    Args:
        cell: openpyxl Cell
        name: §2 canonical name
        scope: "workbook" / "sheet"
    """
    if scope == "workbook":
        return register_workbook_name(cell=cell, wb=cell.parent.parent, name=name)
    elif scope == "sheet":
        return register_sheet_name(ws=cell.parent, name=name, cell=cell)
    else:
        raise ValueError(f"Invalid scope: {scope!r} (expected 'workbook' or 'sheet')")


def bulk_register_workbook_names(
    wb: Workbook, mapping: dict[str, tuple[str, str]]
) -> list[DefinedName]:
    """Workbook-scoped named range を bulk 登録 (§A.2 patterns).

    Builder の冒頭で _named_ranges.md §2 全 entry を一度に流し込む想定。

    Args:
        wb: openpyxl Workbook
        mapping: {name: (sheet_title, cell_ref)} の dict
            cell_ref は単 cell ($C$5) または range ($D$5:$AA$5) どちらでも可

    Returns:
        登録された DefinedName のリスト

    Example:
        >>> bulk_register_workbook_names(wb, {
        ...     "Rev_Total":      ("03_Revenue", "$D$5:$AA$5"),
        ...     "Cost_COGS":      ("04_OpEx",    "$D$10:$AA$10"),
        ...     "BS_TotalAssets": ("06_BS",      "$D$22:$AA$22"),
        ... })
    """
    out: list[DefinedName] = []
    for name, (sheet, ref) in mapping.items():
        out.append(register_workbook_name(wb, name, sheet, ref))
    return out


def bulk_register_sheet_names(
    ws: Worksheet, mapping: dict[str, str]
) -> list[DefinedName]:
    """Sheet-scoped named range を bulk 登録 (1 sheet 内の line item を一括登録).

    Args:
        ws: 定義先 Worksheet
        mapping: {name: cell_ref} の dict (cell_ref は単 cell or range)

    Returns:
        登録された DefinedName のリスト

    Example:
        >>> bulk_register_sheet_names(ws_rev, {
        ...     "Rev_Subscription": "$D$6:$AA$6",
        ...     "Rev_Services":     "$D$7:$AA$7",
        ... })
    """
    out: list[DefinedName] = []
    for name, ref in mapping.items():
        if ":" in ref:
            out.append(register_range_name(ws, name, ref, scope="sheet"))
        else:
            out.append(register_sheet_name(ws, name, ref))
    return out


def list_workbook_names(wb: Workbook) -> list[tuple[str, str, str]]:
    """全 named range を ``(name, scope_label, attr_text)`` の tuple list で返す.

    debug / D12 sanity check coverage 計測用。openpyxl 3.1.x では workbook-scoped
    と sheet-scoped が別々の DefinedNameDict に格納されるため両方を walk する。

    Returns:
        sorted list of tuples; scope_label は ``"workbook"`` または
        ``"sheet[<idx>]:<title>"``.
    """
    out: list[tuple[str, str, str]] = []
    for nm, dn in wb.defined_names.items():
        out.append((nm, "workbook", dn.attr_text or ""))
    for idx, ws in enumerate(wb.worksheets):
        for nm, dn in ws.defined_names.items():
            out.append((nm, f"sheet[{idx}]:{ws.title}", dn.attr_text or ""))
    return sorted(out)


# ============================================================================
# 11. Sheet-level helper functions
# ============================================================================


def setup_sheet_layout(
    ws: Worksheet,
    *,
    n_periods: int = 36,
    has_unit_col: bool = True,
    freeze_at: str = "C5",
) -> None:
    """シート全体のレイアウト初期化.

    Column structure:
        A: margin (narrow)
        B: label
        C: unit (optional)
        D..D+n-1: periods
    """
    ws.column_dimensions["A"].width = COL_MARGIN_WIDTH
    ws.column_dimensions["B"].width = COL_LABEL_WIDTH
    if has_unit_col:
        ws.column_dimensions["C"].width = COL_UNIT_WIDTH
        period_start = 4  # D
    else:
        period_start = 3  # C
    for i in range(n_periods):
        col_letter = get_column_letter(period_start + i)
        ws.column_dimensions[col_letter].width = COL_PERIOD_WIDTH

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_at


def setup_print_layout(
    ws: Worksheet,
    *,
    orientation: Literal["landscape", "portrait"] = "landscape",
    fit_to_width: int = 1,
    print_title_rows: str | None = None,
    print_title_cols: str | None = None,
    header_left: str | None = None,
    header_right: str | None = None,
    footer_left: str | None = None,
    footer_center: str | None = None,
    footer_right: str | None = None,
) -> None:
    """印刷設定 (landscape default + fit-to-width).

    Phase 6 補強: print title rows / cols (各ページに repeat する header) と、
    header / footer (page number, document title, date) を任意で受け取る。

    Args:
        ws: Worksheet
        orientation: "landscape" / "portrait"
        fit_to_width: fit-to-width pages (1 = 1 ページ幅に収める)
        print_title_rows: 各ページで repeat する行 (例: "1:4")
        print_title_cols: 各ページで repeat する列 (例: "A:B")
        header_left/right: page header の左/右 (Excel header コード使用可: "&A"=sheet 名, "&D"=日付)
        footer_left/center/right: page footer の各位置 ("&P"=page #, "&N"=total pages)

    Excel format codes (Microsoft 仕様):
        &P = current page #     &N = total pages    &D = date    &T = time
        &A = sheet name         &F = file name      &Z = path    &G = picture
    """
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = PRINT_MARGIN_INCH
    ws.page_margins.right = PRINT_MARGIN_INCH
    ws.page_margins.top = PRINT_MARGIN_INCH
    ws.page_margins.bottom = PRINT_MARGIN_INCH
    ws.page_margins.header = PRINT_HEADER_INCH
    ws.page_margins.footer = PRINT_HEADER_INCH

    # Print title (各ページで repeat される行/列)
    if print_title_rows:
        ws.print_title_rows = print_title_rows
    if print_title_cols:
        ws.print_title_cols = print_title_cols

    # Header / Footer (オプション)
    if header_left is not None:
        ws.oddHeader.left.text = header_left
    if header_right is not None:
        ws.oddHeader.right.text = header_right
    if footer_left is not None:
        ws.oddFooter.left.text = footer_left
    if footer_center is not None:
        ws.oddFooter.center.text = footer_center
    if footer_right is not None:
        ws.oddFooter.right.text = footer_right


def write_cover(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str = "",
    project_code: str = "",
    confidential: bool = True,
    author: str = "",
    version: str = "1.0",
    date_str: str = "",
) -> None:
    """Cover sheet に必須情報を書き込む.

    Layout (B-row 中心):
        B2: CONFIDENTIAL stamp (top-right)
        B6: title (大)
        B8: subtitle (中)
        B14: project_code / version / date (footer)
    """
    if confidential:
        ws["F2"] = "CONFIDENTIAL"
        ws["F2"].font = Font(name=FONT_FAMILY, size=FONT_SIZE_TINY, bold=True, color=BRAND_PRIMARY_DEEP)
        ws["F2"].alignment = Alignment(horizontal="right")

    ws["B6"] = title
    ws["B6"].font = FONT_COVER_TITLE
    ws["B6"].fill = PatternFill("solid", fgColor=BRAND_PRIMARY_DEEP)
    ws["B6"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[6].height = 60

    if subtitle:
        ws["B8"] = subtitle
        ws["B8"].font = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, color=BRAND_INK)
        ws["B8"].alignment = Alignment(horizontal="left")

    footer_lines = []
    if project_code:
        footer_lines.append(f"Project: {project_code}")
    if version:
        footer_lines.append(f"Version: {version}")
    if date_str:
        footer_lines.append(f"Date: {date_str}")
    if author:
        footer_lines.append(f"Author: {author}")
    if footer_lines:
        ws["B14"] = " | ".join(footer_lines)
        apply_comment(ws["B14"])

    setup_sheet_layout(ws, n_periods=4, has_unit_col=False, freeze_at="A1")
    ws.sheet_properties.tabColor = BRAND_NAVY  # §2.X canonical: Cover = Navy


# ----------------------------------------------------------------------------
# Sheet Tab Color — canonical 6 role × 17 sheet mapping
# Source of truth: _design_consistency_rules.md §2.X (Phase 6 補強)
# ----------------------------------------------------------------------------

# 6 canonical role → tab color (hex, no leading #)
TAB_COLOR_BY_ROLE: dict[str, str] = {
    "cover":  BRAND_NAVY,           # 1F3A66 — document title page
    "input":  BRAND_PRIMARY_DEEP,   # 004F49 — user-editable hard inputs
    "driver": BRAND_PRIMARY,        # 008A80 — calculated drivers / KPI
    "output": BRAND_SLATE,          # 666666 — calculated 3-statement / valuation outputs
    "check":  BRAND_WARNING,        # D6913D — sanity check status
    "memo":   BRAND_ACCENT,         # ECC85A — IC thesis prose
}

# Legacy role aliases (旧名 — 1 release 維持). External builders may still pass
# the pre-Phase 6 role strings; these resolve to the new canonical role/color.
_LEGACY_ROLE_ALIASES: dict[str, str] = {
    "assumptions": "input",
    "drivers":     "driver",
    "financials":  "output",
    "checks":      "check",
    "ic_memo":     "memo",
}

# 14-sheet canonical role assignment (_design_consistency_rules.md §2.X.2)
# Phase 6 Stage A: collapsed from 17 sheets — driver content lives on
# 11_KPI_Dashboard (and inline on 02_Revenue / 03_OpEx); WC schedule is a
# sub-section of 05_BS; Sensitivity is a sub-section of 09_DCF.
SHEET_ROLE_MAPPING: dict[str, str] = {
    "00_Cover":          "cover",
    "01_Assumptions":    "input",
    "02_Revenue":        "driver",
    "03_OpEx":           "driver",
    "04_IS":             "output",
    "05_BS":             "output",
    "06_CFS":            "output",
    "07_Debt":           "output",
    "08_CapTable":       "output",
    "09_DCF":            "output",
    "10_Comps":          "output",
    "11_KPI_Dashboard":  "driver",
    "12_SanityChecks":   "check",
    "13_IC_Memo":        "memo",
}


def get_sheet_role(sheet_name: str) -> str | None:
    """Return canonical role for a sheet name, or None if not in 17-sheet set.

    Optional sheets (e.g. ``17_MA_Exit_Analysis``, ``20a_…`` segment splits,
    ``99_Glossary``) return ``None``; callers may extend ``SHEET_ROLE_MAPPING``
    in-place to assign a role to such optional sheets.
    """
    return SHEET_ROLE_MAPPING.get(sheet_name)


def set_tab_color(ws: Worksheet, role: str | None = None) -> None:
    """Sheet のタブ色を canonical mapping に基づき設定.

    canonical 6 role × 14 sheet (_design_consistency_rules.md §2.X):
      - cover  → BRAND_NAVY          (#1F3A66)  00_Cover
      - input  → BRAND_PRIMARY_DEEP  (#004F49)  01_Assumptions
      - driver → BRAND_PRIMARY       (#008A80)  02_Revenue, 03_OpEx, 11_KPI_Dashboard
      - output → BRAND_SLATE         (#666666)  04_IS, 05_BS, 06_CFS, 07_Debt,
                                                 08_CapTable, 09_DCF, 10_Comps
      - check  → BRAND_WARNING       (#D6913D)  12_SanityChecks
      - memo   → BRAND_ACCENT        (#ECC85A)  13_IC_Memo

    Args:
        ws: Worksheet
        role: 明示 role (cover/input/driver/output/check/memo or 旧名 alias)。
            ``None`` のとき ``ws.title`` から ``SHEET_ROLE_MAPPING`` で推論。

    sheet name から推論できないとき (= 17 canonical 外の optional sheet で
    role 未指定) は no-op。
    """
    resolved = role
    if resolved is None:
        resolved = SHEET_ROLE_MAPPING.get(ws.title)
    else:
        # 旧名 alias を canonical role に正規化
        resolved = _LEGACY_ROLE_ALIASES.get(resolved, resolved)
    if resolved is None:
        return
    color = TAB_COLOR_BY_ROLE.get(resolved)
    if color:
        ws.sheet_properties.tabColor = color


def apply_canonical_tab_colors(wb) -> int:
    """ブック内全 sheet を走査し、SHEET_ROLE_MAPPING にある sheet にのみ
    canonical tab color を一括適用. 適用件数を返す.

    用途: builder の保存直前に呼び出すと、output / driver 系を含めた
    17 sheet 全体に canonical tab color が確実に行き渡る。
    """
    applied = 0
    for ws in wb.worksheets:
        role = SHEET_ROLE_MAPPING.get(ws.title)
        if role is None:
            continue
        color = TAB_COLOR_BY_ROLE[role]
        ws.sheet_properties.tabColor = color
        applied += 1
    return applied


# ============================================================================
# 12. Validation utilities
# ============================================================================


@dataclass(frozen=True)
class CellRoleViolation:
    sheet: str
    cell: str
    expected_role: str
    actual_color: str
    detail: str


def validate_sheet_naming(workbook_sheetnames: Iterable[str]) -> list[str]:
    """ブックの sheet 名が canonical 順か検証. 違反一覧を返す.

    canonical: 00_Cover ... 13_IC_Memo (14 sheets, + optional 99_Glossary).
    Phase 6 Stage A: collapsed from 17 sheets — DEPRECATED_SHEETS detected
    here are flagged for migration.
    """
    expected = list(CANONICAL_SHEET_ORDER)
    actual = list(workbook_sheetnames)
    violations: list[str] = []

    # Flag any deprecated sheet names explicitly so the user knows they need
    # to regenerate via the new builders.
    for name in actual:
        if name in DEPRECATED_SHEETS:
            new_loc = _OLD_TO_NEW_SHEET_MAP.get(name, "n/a")
            violations.append(
                f"Deprecated sheet '{name}' detected — content now lives on '{new_loc}' "
                f"(Phase 6 Stage A 17→14 sheet refactor)."
            )

    actual_filtered = [s for s in actual if s not in OPTIONAL_SHEETS and s not in DEPRECATED_SHEETS]
    if actual_filtered != expected[: len(actual_filtered)]:
        for i, name in enumerate(actual_filtered):
            if i >= len(expected):
                violations.append(f"Extra sheet '{name}' (only 00-13 + 99_Glossary allowed)")
                continue
            if name != expected[i]:
                violations.append(f"Position {i}: got '{name}', expected '{expected[i]}'")

    return violations


# ============================================================================
# 13. Convenience constants for build_model.py / sanity_checks.py
# ============================================================================

# Money unit hints (input_schema reporting_currency と対応)
MONEY_UNIT_BY_CCY: dict[str, str] = {
    "JPY": "百万円",
    "USD": "$M",
    "EUR": "€M",
    "GBP": "£M",
    "SGD": "S$M",
}


__all__ = [
    # Colors
    "IB_HARD_INPUT", "IB_FORMULA", "IB_LINK_INTRA", "IB_LINK_EXTERNAL", "IB_COMMENT",
    "IB_INK", "IB_PALETTE",
    "BRAND_PRIMARY_DEEP", "BRAND_PRIMARY", "BRAND_INK", "BRAND_SURFACE", "BRAND_ACCENT",
    "BRAND_NAVY", "BRAND_GS_NAVY", "BRAND_MS_BLUE", "BRAND_LAZARD_ORANGE",
    "BRAND_EVERCORE_NAVY",
    "BRAND_WARNING", "BRAND_DANGER", "BRAND_SUCCESS", "BRAND_SLATE",
    # Sheet tab color canonical mapping (Phase 6 補強 §2.X)
    "TAB_COLOR_BY_ROLE", "SHEET_ROLE_MAPPING", "get_sheet_role",
    "apply_canonical_tab_colors",
    "BG_WHITE", "BG_TOTAL_BAND", "BG_HEADER_BAND", "BG_WORKING",
    "LINK_COLOR",
    # Heatmap palettes
    "HEATMAP_LOW_COOL", "HEATMAP_MID_NEUTRAL", "HEATMAP_HIGH_WARM",
    "HEATMAP_GREEN", "HEATMAP_YELLOW", "HEATMAP_RED",
    # Chart palettes (Phase 6 Fix #8)
    "IB_CHART_COLORS_FOOTBALL", "IB_CHART_COLORS_BAR", "IB_CHART_COLORS_LINE",
    "IB_CHART_COLORS_WATERFALL_POS", "IB_CHART_COLORS_WATERFALL_NEG",
    "IB_CHART_COLORS_WATERFALL_TOTAL",
    # Fonts (Phase 6 補強: Calibri 11pt 基準)
    "FONT_FAMILY", "FONT_SIZE_BASE", "FONT_SIZE_SMALL", "FONT_SIZE_LARGE",
    "FONT_SIZE_TITLE", "FONT_SIZE_TINY",
    "FONT_BODY", "FONT_BODY_BOLD", "FONT_HARD_INPUT", "FONT_FORMULA",
    "FONT_LINK_INTRA", "FONT_LINK_EXTERNAL",
    "FONT_COMMENT", "FONT_FOOTNOTE", "FONT_SECTION_HEADER", "FONT_TITLE",
    "FONT_YEAR_HEADER", "FONT_SUBTOTAL", "FONT_GRAND_TOTAL", "FONT_COVER_TITLE",
    # Font legacy aliases (旧名)
    "FONT_SECTION", "FONT_SUBSECTION", "FONT_TOTAL",
    # Number formats (Phase 6 補強: 通貨×scale 厳格化、負数 赤)
    "FMT_JPY_YEN", "FMT_JPY_THOUSAND", "FMT_JPY_MILLION", "FMT_JPY_HUNDRED_MILLION",
    "FMT_USD_DOLLAR", "FMT_USD_THOUSAND", "FMT_USD_MILLION",
    "FMT_NUM", "FMT_NUM_THOUSAND", "FMT_NUM_MILLION",
    "FMT_PCT_0", "FMT_PCT_1", "FMT_PCT_2",
    "FMT_MULTIPLE_1", "FMT_MULTIPLE_2",
    "FMT_DATE_YMD", "FMT_DATE_YM",
    "FMT_INTEGER", "FMT_PER_SHARE",
    # Number format legacy aliases
    "FMT_MONEY", "FMT_MONEY_DECIMAL", "FMT_PERCENT", "FMT_PERCENT_BPS",
    "FMT_MULTIPLE", "FMT_SHARES", "FMT_DATE_SHORT", "FMT_DATE_LONG",
    "FMT_DEFAULT",
    # Layout
    "COL_MARGIN_WIDTH", "COL_LABEL_WIDTH", "COL_UNIT_WIDTH", "COL_PERIOD_WIDTH",
    "COL_WIDTH_TINY", "COL_WIDTH_SMALL", "COL_WIDTH_BASE", "COL_WIDTH_LARGE", "COL_WIDTH_XLARGE",
    "ROW_BODY_HEIGHT", "ROW_SPACER_HEIGHT", "ROW_SECTION_HEIGHT",
    "ROW_HEIGHT_TIGHT", "ROW_HEIGHT_BASE", "ROW_HEIGHT_RELAXED", "ROW_HEIGHT_HERO",
    # Borders
    "BORDER_SUBTOTAL", "BORDER_GRAND_TOTAL", "BORDER_SECTION_DIVIDER",
    "BORDER_SECTION_END_MEDIUM", "BORDER_TOP_THIN", "BORDER_BOTTOM_THIN",
    "BORDER_BOTTOM_DOUBLE", "BORDER_BOX_THIN", "BORDER_BOX_MEDIUM", "BORDER_BOX_THICK",
    "THIN_LINE", "MEDIUM_LINE", "THICK_LINE", "DOUBLE_LINE", "HAIRLINE_GRAY", "THIN_GRAY",
    # Sheet naming (Phase 6 Stage A 14-sheet canonical)
    "CANONICAL_SHEET_ORDER", "OPTIONAL_SHEETS",
    "DEPRECATED_SHEETS",
    # Cell-level helpers
    "apply_hard_input", "apply_formula", "apply_link_intra", "apply_link_external",
    "apply_label", "apply_label_indent", "write_hierarchical_line_item",
    "set_uniform_column_width", "apply_chart_palette",
    "apply_section_header", "apply_year_header",
    "apply_subtotal", "apply_grand_total", "apply_comment", "apply_working_highlight",
    # Unit / currency helpers (Phase 6 補強)
    "apply_unit_label", "fmt_for_currency",
    # Border helpers (Phase 6 補強)
    "apply_section_bottom_border", "apply_box_border",
    # Conditional formatting (Phase 6 補強)
    "apply_heatmap_3color", "apply_data_bar",
    # Data validation (Phase 6 補強)
    "apply_dropdown", "apply_numeric_validation",
    # Hyperlink (Phase 6 補強)
    "apply_internal_link", "write_toc",
    # Row height (Phase 6 補強)
    "apply_row_heights",
    # Named range helpers (Phase 6 Wave 2.1 — references/_named_ranges.md §5)
    "register_workbook_name", "register_sheet_name", "register_range_name",
    "register_named_range",
    "bulk_register_workbook_names", "bulk_register_sheet_names",
    "list_workbook_names",
    # Sheet-level helpers
    "setup_sheet_layout", "setup_print_layout", "write_cover", "set_tab_color",
    "validate_sheet_naming",
    # Currency
    "MONEY_UNIT_BY_CCY",
    # Models
    "CellRoleViolation",
]


# ============================================================================
# 14. Smoke test (run via `python ib_format.py`)
# ============================================================================
# Intended only as a self-test. Builders import named members; this block is
# not executed at import time.

if __name__ == "__main__":
    # ---- Sub-test: _validate_named_range -----------------------------------
    # Valid names (from references/_named_ranges.md §2 canonical table)
    for ok_name in [
        "Rev_Total", "ARR_EOY", "Cap_PPS_PostMoney",
        "KPI_LTV_CAC", "Year_Headers", "_Hidden",
    ]:
        _validate_named_range(ok_name)

    # NG names — each should raise ValueError
    ng_cases = [
        "",                                    # empty
        "Print_Area",                          # Excel reserved (§3.1)
        "Auto_Open",                           # Excel reserved
        "A1",                                  # cell address collision (§3.2)
        "Q1",                                  # cell address collision
        "FY24",                                # cell address collision
        "Rev-Total",                           # hyphen
        "Rev Total",                           # space
        "1Rev",                                # digit start
        "x" * 26,                              # too long (>25)
        "売上_Total",                          # non-ASCII (§3.6)
    ]
    for bad in ng_cases:
        try:
            _validate_named_range(bad)
        except ValueError:
            pass
        else:
            raise AssertionError(f"_validate_named_range should reject {bad!r}")

    # ---- Sub-test: register_workbook_name / register_sheet_name ------------
    from openpyxl import Workbook as _Wb
    wb = _Wb()
    ws = wb.active
    ws.title = "03_Revenue"
    ws2 = wb.create_sheet("05_IS")

    # Workbook scope via apply_hard_input(named_range=)
    # Note: apply_hard_input does NOT take a value arg — set cell.value separately.
    ws["C5"].value = 1_000_000_000
    apply_hard_input(ws["C5"], named_range="Rev_Total", scope="workbook")

    assert "Rev_Total" in wb.defined_names, "Rev_Total should be registered (workbook)"
    rt = wb.defined_names["Rev_Total"]
    assert rt.attr_text == "'03_Revenue'!$C$5", f"unexpected attr_text: {rt.attr_text!r}"
    assert rt.localSheetId is None, "workbook scope should have localSheetId=None"

    # Sheet scope via apply_hard_input(named_range=, scope="sheet")
    ws["C6"].value = 500_000_000
    apply_hard_input(ws["C6"], named_range="Rev_Subscription", scope="sheet")
    assert "Rev_Subscription" in ws.defined_names, "Rev_Subscription should be in ws.defined_names"
    sub_dn = ws.defined_names["Rev_Subscription"]
    assert sub_dn.attr_text == "'03_Revenue'!$C$6"
    assert sub_dn.localSheetId == 0, f"sheet-scoped localSheetId expected 0; got {sub_dn.localSheetId}"
    # Sheet-scoped name should NOT appear in workbook-level dict
    assert "Rev_Subscription" not in wb.defined_names

    # ---- Sub-test: register_range_name -------------------------------------
    register_range_name(ws, "Rev_TotalRange", "$D$5:$AA$5", scope="workbook")
    assert wb.defined_names["Rev_TotalRange"].attr_text == "'03_Revenue'!$D$5:$AA$5"

    register_range_name(ws, "ARR_BoP", "$D$15:$AA$15", scope="sheet")
    assert ws.defined_names["ARR_BoP"].localSheetId == 0
    assert "ARR_BoP" not in wb.defined_names

    # ---- Sub-test: register_workbook_name with explicit sheet_title --------
    register_workbook_name(wb, "IS_NI", "05_IS", "$D$24")
    assert wb.defined_names["IS_NI"].attr_text == "'05_IS'!$D$24"

    # ---- Sub-test: validation rejection at register_workbook_name ----------
    try:
        register_workbook_name(wb, "Print_Area", "03_Revenue", "$A$1")
    except ValueError:
        pass
    else:
        raise AssertionError("register_workbook_name should reject Excel reserved name")

    try:
        register_workbook_name(wb, "A1", "03_Revenue", "$A$1")
    except ValueError:
        pass
    else:
        raise AssertionError("register_workbook_name should reject cell-address collision")

    # ---- Sub-test: bulk_register_workbook_names ----------------------------
    bulk_register_workbook_names(wb, {
        "Cost_COGS":     ("05_IS", "$D$6"),
        "BS_Cash":       ("05_IS", "$D$10"),
    })
    assert "Cost_COGS" in wb.defined_names and "BS_Cash" in wb.defined_names

    # ---- Sub-test: bulk_register_sheet_names -------------------------------
    bulk_register_sheet_names(ws, {
        "Rev_Services":  "$D$7:$AA$7",
        "Rev_Initial":   "$D$8",
    })
    assert "Rev_Services" in ws.defined_names
    assert "Rev_Initial" in ws.defined_names

    # ---- Sub-test: list_workbook_names -------------------------------------
    listed = list_workbook_names(wb)
    listed_names = {n for n, _, _ in listed}
    assert {"Rev_Total", "Rev_Subscription", "IS_NI", "Cost_COGS"}.issubset(listed_names), (
        f"list_workbook_names missing entries; got {sorted(listed_names)}"
    )
    # workbook vs sheet labelling
    scope_by_name = {n: s for n, s, _ in listed}
    assert scope_by_name["Rev_Total"] == "workbook"
    assert scope_by_name["Rev_Subscription"].startswith("sheet[0]:")

    # ---- Sub-test: round-trip via save / load_workbook ---------------------
    import os, tempfile
    from openpyxl import load_workbook
    tmp = os.path.join(tempfile.gettempdir(), "ib_format_named_range_test.xlsx")
    wb.save(tmp)
    wb2 = load_workbook(tmp)
    names2_global = set(wb2.defined_names.keys())
    names2_sheet0 = set(wb2.worksheets[0].defined_names.keys())
    assert {"Rev_Total", "IS_NI", "Cost_COGS"}.issubset(names2_global), (
        f"round-trip lost workbook-scoped names; got {sorted(names2_global)}"
    )
    assert {"Rev_Subscription", "Rev_Services"}.issubset(names2_sheet0), (
        f"round-trip lost sheet-scoped names; got {sorted(names2_sheet0)}"
    )

    # ---- Sub-test: backward compat (positional fmt) ------------------------
    wb3 = _Wb()
    ws3 = wb3.active
    apply_hard_input(ws3["B5"], FMT_USD_MILLION)         # positional fmt — must work
    apply_formula(ws3["B6"], FMT_PERCENT)                # legacy alias positional
    apply_link_intra(ws3["B7"], FMT_JPY_MILLION)
    apply_link_external(ws3["B8"])                       # default fmt
    apply_subtotal(ws3["B9"], FMT_MONEY)
    apply_grand_total(ws3["B10"], FMT_MONEY)
    assert ws3["B5"].number_format == FMT_USD_MILLION
    assert ws3["B6"].number_format == FMT_PERCENT

    print("ib_format.py smoke test: PASS")
    print(f"  - {len(listed)} named ranges registered in primary test workbook")
    print(f"  - round-trip xlsx written to {tmp}")
