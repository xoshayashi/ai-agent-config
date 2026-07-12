"""Run the startup-financial-modeling skill quality gate.

This script gives future agents one executable closeout path instead of making
them rediscover the domain rubric, strict-audit sample builds, and pytest scope
from separate reference files.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


def skill_root() -> Path:
    return Path(__file__).resolve().parents[1]


def rubric_command(root: Path) -> list[str]:
    return [
        "node",
        str(root / "build" / "evals" / "startup-finance-rubric" / "emit-startup-finance-rubric.js"),
        str(root),
    ]


def strict_audit_commands(root: Path, tmp_dir: Path) -> list[list[str]]:
    build_model = root / "scripts" / "build_model.py"
    return [
        [
            sys.executable,
            str(build_model),
            "--mode",
            "pricing",
            "--output",
            str(tmp_dir / "pricing_quality_gate.xlsx"),
            "--strict-audit",
        ],
        [
            sys.executable,
            str(build_model),
            "--mode",
            "cap_table",
            "--output",
            str(tmp_dir / "cap_table_quality_gate.xlsx"),
            "--strict-audit",
        ],
    ]


def architecture_command(root: Path) -> list[str]:
    return [sys.executable, str(root / "scripts" / "architecture_gate.py")]


def scenario_eval_command(root: Path, *, full: bool, min_score: float) -> list[str]:
    command = [
        sys.executable,
        str(root / "scripts" / "scenario_eval.py"),
        "--min-score",
        str(min_score),
    ]
    if full:
        command.append("--full")
    return command


def pytest_command(root: Path, full: bool) -> list[str]:
    if full:
        targets = [
            "build/tests/test_build_model.py",
            "build/tests/test_economic_quality.py",
            "tests/test_build_model_wrapper.py",
        ]
    else:
        targets = [
            "tests/test_build_model_wrapper.py",
            "build/tests/test_build_model.py",
        ]
    return [sys.executable, "-m", "pytest", *targets]


def run_command(command: list[str], root: Path) -> subprocess.CompletedProcess[str]:
    print(f"[gate] {' '.join(command)}")
    return subprocess.run(
        command,
        cwd=root,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )


def spreadsheet_engine() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def recalculated_workbook_path(output_path: Path, recalc_dir: Path) -> Path:
    return recalc_dir / output_path.name


def formula_error_cells(xlsx_path: Path) -> list[str]:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(xlsx_path, data_only=True)
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    return errors


def recalc_and_inspect(xlsx_path: Path, root: Path, tmp_dir: Path) -> int:
    engine = spreadsheet_engine()
    if engine is None:
        print("[gate] soffice/libreoffice unavailable; skipping spreadsheet-engine recalc")
        return 0
    recalc_dir = tmp_dir / "recalc"
    recalc_dir.mkdir(exist_ok=True)
    command = [
        engine,
        "--headless",
        "--calc",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(recalc_dir),
        str(xlsx_path),
    ]
    result = run_command(command, root)
    print(result.stdout, end="")
    if result.returncode != 0:
        return result.returncode
    recalculated = recalculated_workbook_path(xlsx_path, recalc_dir)
    if not recalculated.exists():
        print(f"[gate] recalculated workbook not found: {recalculated}", file=sys.stderr)
        return 2
    errors = formula_error_cells(recalculated)
    if errors:
        print("[gate] recalculated workbook contains spreadsheet errors:", file=sys.stderr)
        for error in errors[:50]:
            print(f"[gate] {error}", file=sys.stderr)
        if len(errors) > 50:
            print(f"[gate] ... {len(errors) - 50} more", file=sys.stderr)
        return 2
    runtime_dir = root / "build" / "runtime"
    sys.path.insert(0, str(runtime_dir))
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
        from build_model import audit_recalculated_financial_model  # noqa: PLC0415

        finance_issues = audit_recalculated_financial_model(
            load_workbook(recalculated, data_only=True)
        )
    finally:
        try:
            sys.path.remove(str(runtime_dir))
        except ValueError:
            pass
    if finance_issues:
        print("[gate] recalculated workbook failed financial integrity audit:", file=sys.stderr)
        for issue in finance_issues[:50]:
            print(f"[gate] {issue}", file=sys.stderr)
        if len(finance_issues) > 50:
            print(f"[gate] ... {len(finance_issues) - 50} more", file=sys.stderr)
        return 2
    print(f"[gate] recalculated workbook has no cached formula errors: {recalculated}")
    print("[gate] recalculated workbook passed financial integrity audit")
    return 0


def run_rubric(command: list[str], root: Path) -> subprocess.CompletedProcess[str]:
    print(f"[gate] {' '.join(command)}")
    return subprocess.run(
        command,
        cwd=root,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )


def domain_score(output: str) -> float:
    payload = json.loads(output)
    for metric in payload.get("metrics", []):
        if metric.get("id") == "startup-finance-domain-rubric-score":
            return float(metric["value"])
    raise ValueError("startup-finance-domain-rubric-score metric not found")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--min-score",
        type=float,
        default=95.0,
        help="Minimum startup-finance domain rubric score required for pass.",
    )
    parser.add_argument(
        "--full",
        action="store_true",
        help="Run the full pytest suite, including economic-quality tests.",
    )
    parser.add_argument(
        "--skip-pytest",
        action="store_true",
        help="Run rubric and strict-audit sample builds only.",
    )
    parser.add_argument(
        "--skip-recalc",
        action="store_true",
        help="Skip optional LibreOffice/soffice recalculation even if available.",
    )
    parser.add_argument(
        "--scenario-min-score",
        type=float,
        default=95.0,
        help="Minimum scenario eval score required for pass.",
    )
    args = parser.parse_args(argv)

    root = skill_root()
    if shutil.which("node") is None:
        print("[gate] node is required to run startup-finance-rubric", file=sys.stderr)
        return 2

    rubric = run_rubric(rubric_command(root), root)
    print(rubric.stdout, end="")
    if rubric.stderr:
        print(rubric.stderr, end="", file=sys.stderr)
    if rubric.returncode != 0:
        return rubric.returncode

    try:
        score = domain_score(rubric.stdout)
    except (json.JSONDecodeError, KeyError, TypeError, ValueError) as exc:
        print(f"[gate] could not parse startup-finance-rubric output: {exc}", file=sys.stderr)
        return 2
    if score < args.min_score:
        print(
            f"[gate] domain rubric score {score:g} is below required {args.min_score:g}",
            file=sys.stderr,
        )
        return 2
    print(f"[gate] domain rubric score {score:g}/{args.min_score:g}+")

    result = run_command(architecture_command(root), root)
    print(result.stdout, end="")
    if result.returncode != 0:
        return result.returncode

    result = run_command(
        scenario_eval_command(root, full=args.full, min_score=args.scenario_min_score),
        root,
    )
    print(result.stdout, end="")
    if result.returncode != 0:
        return result.returncode

    with tempfile.TemporaryDirectory(prefix="sfm-quality-gate-") as tmp:
        tmp_dir = Path(tmp)
        for command in strict_audit_commands(root, tmp_dir):
            result = run_command(command, root)
            print(result.stdout, end="")
            if result.returncode != 0:
                return result.returncode
            if not args.skip_recalc:
                output_path = Path(command[command.index("--output") + 1])
                rc = recalc_and_inspect(output_path, root, tmp_dir)
                if rc != 0:
                    return rc

    if not args.skip_pytest:
        result = run_command(pytest_command(root, full=args.full), root)
        print(result.stdout, end="")
        if result.returncode != 0:
            return result.returncode

    print("[gate] startup-financial-modeling quality gate passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
