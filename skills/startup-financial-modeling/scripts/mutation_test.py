#!/usr/bin/env python3
"""変異テスト: チェック体系の「実効性」を機械検証する（動的ゲート）。

契約（`references/check_design_spec.md` 原則8）:

  **印刷面に出る導出セルをゼロ化して数字が動いたなら、総合判定は必ず鳴ること。**

背景（このスキルで最も高くついた失敗）:
  エラー級チェックは存在するだけでは無意味で、恒等式（両辺が同じ経路を読む式）だと
  壊れたモデルを素通りさせる。さらに悪いことに、**ハーネス自身の対象選定**が
  偽の保証になっていた——「（参考）」を含むラベルを検査対象から外していたため、
  DCFが参照する唯一のFCF行が無検査のまま出荷された。
  対象は `model_surface.py` が**構造だけ**で決める（ラベル列挙・シンク列挙をしない）。

変異オペレータ: **導出セルのゼロ化**（＝リンク脱落そのもの）
  印刷面の導出セルを1つずつゼロにして再計算する。全期間列を対象にする
  （中央列だけだと MIN/MAX/IF の飽和で特定の列でしか露見しない欠陥を見逃す）。

判定（オラクル）: 印刷面の数値が動いたか。
  ゼロ化して印刷面の数値が動いたのに総合判定が「OK」のままなら盲点。
  動かないセル（もともと0など）は「不活性」として必ず一覧出力する（黙って捨てない）。

このハーネスが見ないもの（免責）:
  - 前提条件（入力）の値そのものの誤り。チェックでは原理的に検知できない
  - 文字列を返す表示行（「期間内なし」等）。算術タイで守れない
  → いずれもソース照合ブロックと人間のレビューで担保する

使い方: python3 mutation_test.py <xlsx> [--max N] [--workers N] [--verbose]
終了コード 0=盲点なし / 1=盲点あり・検出力不足 / 2=環境エラー
"""
from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from model_surface import surface_cells

REL_TOL = 1e-6   # 同一数式の再計算は決定的なので、相対1e-6で十分に安全


def _recalc(path: Path, outdir: Path) -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    if not Path(soffice).exists():
        print("FAIL: sofficeが見つかりません（変異テストは必須ゲート）")
        sys.exit(2)
    outdir.mkdir(parents=True, exist_ok=True)   # soffice --outdir は存在しない出力先を作らない
    prof = Path(tempfile.mkdtemp(prefix="lo_")).as_uri()
    subprocess.run([soffice, f"-env:UserInstallation={prof}",
                    "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(outdir), str(path)],
                   check=True, capture_output=True, timeout=300)
    return outdir / path.name


def _state(path: Path, surface):
    """(総合判定, 印刷面の数値ベクトル) を返す。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    verdict = None
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            lab = next((c.value for c in row[1:4]
                        if isinstance(c.value, str) and c.value.strip()), None)
            if isinstance(lab, str) and lab.startswith("総合判定"):
                verdict = next((c.value for c in row[5:]
                                if c.value is not None), None)
    vec = []
    for sheet, r, c, _ in surface:
        v = wb[sheet].cell(row=r, column=c).value
        vec.append(v if isinstance(v, (int, float)) else 0)
    return verdict, vec


def _moved(a, b):
    return any(abs(x - y) > max(abs(y) * REL_TOL, 1e-9) for x, y in zip(a, b))


def main(path: str, max_mut: int, min_detect: int, workers: int,
         verbose: bool) -> int:
    src = Path(path)
    surface = surface_cells(openpyxl.load_workbook(src))   # 構造で決まる対象
    total = len(surface)
    if not total:
        print("FAIL: 印刷面の導出セルが見つかりません（構造が想定外）")
        return 2

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        b_v, b_vec = _state(_recalc(src, td / "base"), surface)
        if b_v != "OK":
            print(f"FAIL: ベースラインの総合判定が OK でない（{b_v}）")
            return 2

        if max_mut and max_mut < total:
            step = max(1, total // max_mut)
            picked = surface[::step][:max_mut]
        else:
            picked = surface                    # 既定は全数

        def run_one(job):
            i, (sheet, r, c, label) = job
            m = openpyxl.load_workbook(src)
            m[sheet].cell(row=r, column=c).value = 0
            mp = td / f"m{i}.xlsx"
            m.save(mp)
            v, vec = _state(_recalc(mp, td / f"rc{i}"), surface)
            return sheet, r, c, label, v, vec

        blind, detected, inert = [], 0, []
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for sheet, r, c, label, v, vec in ex.map(run_one,
                                                     enumerate(picked)):
                cell = f"{sheet}!{get_column_letter(c)}{r}"
                if not _moved(vec, b_vec):
                    inert.append(f"{cell} {label[:24]}")
                    continue
                if v == "OK":
                    blind.append(f"{cell} {label[:30]}")
                else:
                    detected += 1
                    if verbose:
                        print(f"  ✓ 検知: {cell} {label[:24]}")

        cov = 100.0 * len(picked) / total if total else 0.0
        print(f"印刷面セルのゼロ化変異 {len(picked)}/{total}セル"
              f"（カバレッジ {cov:.0f}%）: 検知 {detected} / "
              f"盲点 {len(blind)} / 不活性 {len(inert)}")
        if inert:
            # 「不活性」＝壊しても印刷面の数字が動かないセル（もともと0など）。
            # 黙って捨てず必ず出す（意味のある行が並んでいたら設計が疑わしい）。
            print(f"  不活性セル {len(inert)}件（壊しても数字が動かない）:")
            for x in inert[:12]:
                print("   -", x)
            if len(inert) > 12:
                print(f"   … 他 {len(inert) - 12}件")
        if blind:
            print(f"FAIL: 盲点 {len(blind)}件"
                  "（セルを落として印刷面の数字が動いたのに総合判定OK）:")
            for b in blind:
                print("  -", b)
            return 1
        if cov < 100.0:
            print(f"NOTE: サンプリング実行（カバレッジ {cov:.0f}%）。"
                  "出荷認証は全数（--max を外す）で通すこと")
        if detected < min_detect:
            print(f"FAIL: 検出力不足（数字が動いた変異が {detected}件 < "
                  f"下限 {min_detect}件）。空虚な合格の疑い")
            return 1
        print("PASS 変異テスト: 印刷面に効く全セルの脱落を総合判定が検知する")
        return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--max", type=int, default=0, dest="max_mut",
                    help="間引き実行（既定0＝全数）")
    ap.add_argument("--min-detect", type=int, default=8, dest="min_detect")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--verbose", action="store_true")
    a = ap.parse_args()
    sys.exit(main(a.xlsx, a.max_mut, a.min_detect, a.workers, a.verbose))
