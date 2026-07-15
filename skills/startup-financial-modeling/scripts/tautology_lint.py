#!/usr/bin/env python3
"""静的恒等式リンター: チェックが「偽の保証」でないかを再計算なしで検証する。

契約（`references/check_design_spec.md` 原則1・5・6・8の機械化）:

  **印刷面に出る導出セルは、すべてエラー級チェックで守られていること。**

判定原理:
  エラー級チェックは `対象 −（独立経路での再導出）` の形をしている。
  あるセル X が**両辺の依存錐**に入っているなら、X が壊れると両辺が同じ向きに
  動き、差が相殺されうる。そのチェックは X を守っていない。
  X が**片側の錐だけ**にあるなら、X の破損は必ず差として現れる（＝独立経路）。

  → X が「守られている」とは、あるエラー級チェックの列 c について
     X ∈ 錐(左辺_c) △ 錐(右辺_c)（対称差）が成り立つこと。

  錐は**セル単位**で取る（行単位だと `期首(c) = 期初 + Σ_{i<c} 増減(i)` のような
  累計経路が自己参照に見え、正しい設計を誤検知する）。

検査対象は `model_surface.py` が**構造だけ**で決める（行ラベルにも事業の形にも
依存しない）。保護を与えないチェック（スケール不変・減算形でない比率）は無視する。

mutation_test.py との関係:
  静的（数秒・全列・線形の依存関係のみ）と動的（実際に壊す・MIN/MAXの飽和も見る）。
  静的が「守られている」と言っても動的が嘘だと言うことがある（MAXの非採用側等）。
  両方を verify.sh のゲートに置く。

使い方: python3 tautology_lint.py <xlsx> [--verbose]
終了コード 0=無防備セルなし / 1=無防備あり / 2=構造エラー
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from model_surface import INPUT_SHEET, SUMMARY_SHEET, label_of, surface_cells

RANGE = re.compile(
    r"(?:'([^']+)'!|([^\W\d_]+)!)?"
    r"(\$?[A-Z]{1,2}\$?\d+(?::\$?[A-Z]{1,2}\$?\d+)?)")
FUNC = re.compile(r"[A-Z]+\(")


def cells_of(formula: str, cur: str):
    """数式が参照するセルを (sheet, row, col) で列挙。範囲は展開する。"""
    out = set()
    body = FUNC.sub("(", formula)          # 関数名が列参照に化けるのを防ぐ
    for m in RANGE.finditer(body):
        sheet = m.group(1) or m.group(2) or cur
        try:
            c1, r1, c2, r2 = range_boundaries(m.group(3).replace("$", ""))
        except ValueError:
            continue
        for r in range(r1, (r2 or r1) + 1):
            for c in range(c1, (c2 or c1) + 1):
                out.add((sheet, r, c))
    return out


def split_sides(f: str):
    """`=左辺 −（右辺）` を深さ0の最初の減算で二分する。ABS()は剥がす。"""
    s = f.lstrip("=").strip()
    if s.upper().startswith("ABS(") and s.endswith(")"):
        s = s[4:-1].strip()
    while True:
        cut = _split(s)
        if cut:
            return cut
        if s.startswith("(") and s.endswith(")"):
            s = s[1:-1].strip()
            continue
        return None


def _split(s: str):
    d = 0
    for i, ch in enumerate(s):
        d += (ch == "(") - (ch == ")")
        if ch == "-" and d == 0 and i > 0:
            return s[:i], s[i + 1:]
    return None


def main(path: str, verbose: bool) -> int:
    wb = openpyxl.load_workbook(Path(path))

    prec: dict = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    prec[(ws.title, c.row, c.column)] = cells_of(c.value,
                                                                 ws.title)

    def cone(seeds):
        """先祖セルの推移閉包。"""
        seen, stack = set(), list(seeds)
        while stack:
            n = stack.pop()
            if n in seen:
                continue
            seen.add(n)
            stack.extend(prec.get(n, ()))
        return seen

    # --- エラー級チェックの実体を総合判定から辿る（構造のみ） ---
    if SUMMARY_SHEET not in wb.sheetnames:
        print("FAIL: サマリーがありません")
        return 2
    smry = wb[SUMMARY_SHEET]
    vrow = next((r for r in range(1, smry.max_row + 1)
                 if label_of(smry, r).startswith("総合判定")), None)
    if not vrow:
        print("FAIL: 総合判定行が見つかりません")
        return 2
    vf = smry.cell(row=vrow, column=6).value or ""
    agg = {int(m) for m in re.findall(r"IFERROR\(F(\d+)", vf)}
    if not agg:
        print("FAIL: 総合判定がエラー級チェックを集約していません")
        return 2

    checks = []
    for r in sorted(agg):
        f = smry.cell(row=r, column=6).value
        if not isinstance(f, str):
            continue
        if "SUMPRODUCT(--(ABS(" not in f:
            # サマリー上に直接置かれたチェック行（転記タイ等）はその行が実体
            checks.append((SUMMARY_SHEET, r, label_of(smry, r)))
            continue
        m = re.search(r"'([^']+)'!\$?([A-Z]{1,2})\$?(\d+)", f)
        if m:
            # 前提条件上のチェック（採用値の行タイ等）も必ず数える。
            # ここで入力シートを除外していたため、採用値60セルが
            # 「守られていない」と誤判定され、逆に本物の穴も見えなくなっていた。
            checks.append((m.group(1), int(m.group(3)), label_of(smry, r)))

    protected: dict = defaultdict(list)
    scale_free: set = set()
    unsplit: set = set()
    for sheet, row, clabel in checks:
        ws = wb[sheet]
        ok = False
        for col in range(6, 13):
            f = ws.cell(row=row, column=col).value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            sides = split_sides(f)
            if not sides:
                continue
            lref, rref = (cells_of(x, sheet) for x in sides)
            if not rref:
                # 再導出側にセル参照がない（例: 持分合計−1）＝スケール不変。
                # 水準を固定しないので、どのセルにも保護を与えない。
                scale_free.add(clabel)
                continue
            ok = True
            lhs, rhs = cone(lref), cone(rref)
            for cell in lhs ^ rhs:
                protected[cell].append(clabel)
        if not ok:
            unsplit.add(clabel)

    # --- 検査対象＝印刷面の導出セル（構造で決まる。ラベル列挙なし） ---
    holes = defaultdict(list)
    universe = 0
    cover = defaultdict(lambda: [0, 0])
    for sheet, r, c, lab in surface_cells(wb):
        universe += 1
        owners = protected.get((sheet, r, c))
        if not owners:
            holes[(sheet, r, lab)].append(get_column_letter(c))
            continue
        for o in set(owners):
            cover[o][0] += 1
        if len(set(owners)) == 1:
            cover[owners[0]][1] += 1

    for c in sorted(scale_free | unsplit):
        why = ("スケール不変（再導出側が定数）" if c in scale_free
               else "`対象−再導出` の減算形でない")
        print(f"NOTE: {c[:34]} は保護を与えない（{why}）")
    if verbose:
        print("チェック別の守備範囲（独占0＝行脱落は他チェックでも検知される）:")
        for label, (tot, uniq) in sorted(cover.items(), key=lambda x: x[1][1]):
            print(f"  {label[:38]:40s} 守備 {tot:4d} / 独占 {uniq:4d}")
    print(f"静的恒等式リンター: エラー級チェック {len(checks)}本 / "
          f"印刷面の導出セル {universe}個 / 無防備 "
          f"{sum(len(v) for v in holes.values())}個（{len(holes)}行）")
    if holes:
        print("FAIL: 依存錐の両側に現れる（＝壊しても差が相殺されうる）セル:")
        for (sheet, r, lab), cols in sorted(holes.items()):
            print(f"  - {sheet}!{r} {lab[:30]} 列{','.join(cols)}")
        print("→ その行を対象（左辺）に、原本ドライバー／入力プリミティブまで"
              "展開した再計算チェック（原則6）か、累計経路のロール整合（原則5）"
              "を足すこと")
        return 1
    print("PASS 静的恒等式リンター: 印刷面の全導出セルが"
          "非対称な独立経路チェックで守られている")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1], "--verbose" in sys.argv))
