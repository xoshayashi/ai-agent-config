"""Scenario evals for startup-financial-modeling workbook generation.

The quality gate checks deterministic docs/runtime contracts. This script adds
end-to-end model-quality scenarios: build representative startup archetypes,
reopen the workbook, optionally recalculate it, and score finance integrity.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class Scenario:
    id: str
    story: str
    expected_sheets: tuple[str, ...]
    expected_markers: tuple[tuple[str, str], ...]


SCENARIOS: tuple[Scenario, ...] = (
    Scenario(
        id="recurring-software",
        story=(
            "# Recurring software plan\n"
            "B2B SaaS company with ARR, churn, expansion, sales headcount, "
            "gross margin 78%, and a five-year fundraising model. Source: "
            "management memo."
        ),
        expected_sheets=("P&L", "BS", "CF", "Capital Stack", "Ownership", "Valuation", "IC Memo"),
        expected_markers=(("KPI", "Rule of 40"), ("Valuation", "MOIC at selected EV")),
    ),
    Scenario(
        id="marketplace",
        story=(
            "# Marketplace plan\n"
            "Two-sided marketplace with GMV, take rate, payment fees, buyer "
            "and seller liquidity, repeat behavior, incentives, working "
            "capital, and funding need. Source: marketplace operating memo."
        ),
        expected_sheets=("Revenue Build", "Cost Build", "KPI", "Scenarios", "Valuation"),
        expected_markers=(("KPI", "Contribution margin"), ("Scenarios", "GMV / liquidity scale")),
    ),
    Scenario(
        id="hardware-asset-heavy",
        story=(
            "# Robotics deployment plan\n"
            "Hardware and robot deployment startup selling units, attaching "
            "service revenue, carrying BOM, warranty, field service, capex, "
            "leases, inventory, and debt capacity. Source: board plan."
        ),
        expected_sheets=("Cost Build", "People Plan", "BS", "CF", "Capital Stack", "Valuation"),
        expected_markers=(("KPI", "Asset payback"), ("Scenarios", "Deployment capacity scale")),
    ),
    Scenario(
        id="pre-revenue-milestone",
        story=(
            "# Deeptech milestone plan\n"
            "Pre-revenue R&D company with PoC, prototype, regulatory, hiring, "
            "grant, convertible, and next-round milestone financing. Source: "
            "seed IC memo."
        ),
        expected_sheets=("Assumptions", "CF", "Capital Stack", "Ownership", "IC Memo"),
        expected_markers=(("KPI", "Milestone runway"), ("Scenarios", "Prototype / program cost factor")),
    ),
    Scenario(
        id="fintech-balance-sheet",
        story=(
            "# Fintech lending plan\n"
            "Fintech balance-sheet startup with origination, spreads, credit "
            "losses, collections, warehouse line, covenants, and financing "
            "capacity. Source: lender diligence memo."
        ),
        expected_sheets=("BS", "CF", "Capital Stack", "KPI", "Scenarios", "Valuation"),
        expected_markers=(("KPI", "Loss / collection quality"), ("Scenarios", "Warehouse / debt headroom")),
    ),
)


def skill_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _load_runtime(root: Path) -> Any:
    runtime_dir = root / "build" / "runtime"
    sys.path.insert(0, str(runtime_dir))
    import build_model  # noqa: PLC0415
    import live_comps  # noqa: PLC0415

    def fake_public_comps(tickers: list[str], *, timeout: float = 8.0) -> Any:
        comps = [
            live_comps.PublicComp(
                ticker=ticker.strip().upper(),
                name=f"{ticker.strip().upper()} Scenario Comparable",
                currency="USD",
                market_cap=100_000_000_000.0,
                enterprise_value=110_000_000_000.0,
                revenue_multiple=8.0 + idx,
                ebitda_multiple=20.0 + idx,
                source_url=f"https://example.test/{ticker.strip().upper()}",
                as_of_date="2026-07-08",
                status="current",
            )
            for idx, ticker in enumerate(tickers)
            if ticker.strip()
        ]
        return live_comps.PublicCompsResult(
            comps=comps,
            revenue_multiple_median=9.0 if comps else None,
            ebitda_multiple_median=21.0 if comps else None,
            source_url="https://example.test/scenario-comps",
            as_of_date="2026-07-08",
        )

    build_model.lc.fetch_public_comps = fake_public_comps
    return build_model


def _spreadsheet_engine() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def _formula_error_cells(wb: Any) -> list[str]:
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    return errors


def _sheet_contains(wb: Any, sheet: str, marker: str) -> bool:
    if sheet not in wb.sheetnames:
        return False
    ws = wb[sheet]
    return any(cell.value == marker for row in ws.iter_rows() for cell in row)


def _recalculate(path: Path, tmp_dir: Path) -> Path | None:
    engine = _spreadsheet_engine()
    if engine is None:
        return None
    recalc_dir = tmp_dir / "recalc"
    recalc_dir.mkdir(exist_ok=True)
    subprocess.run(
        [engine, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", str(recalc_dir), str(path)],
        check=True,
        capture_output=True,
        timeout=120,
    )
    recalculated = recalc_dir / path.name
    return recalculated if recalculated.exists() else None


def run_scenario_eval(*, full: bool) -> dict[str, Any]:
    root = skill_root()
    build_model = _load_runtime(root)
    from openpyxl import load_workbook  # noqa: PLC0415

    scenarios = SCENARIOS if full else SCENARIOS[:2]
    checks: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="sfm-scenario-eval-") as tmp:
        tmp_dir = Path(tmp)
        for scenario in scenarios:
            src = tmp_dir / f"{scenario.id}.md"
            out = tmp_dir / f"{scenario.id}.xlsx"
            src.write_text(scenario.story, encoding="utf-8")
            try:
                build_model.build_model(None, out, mode="full", source_md=src)
                wb = load_workbook(out, data_only=False)
                workbook_issues = build_model.audit_workbook(wb)
                facts, _ = build_model._facts_for_inputs(None, src)
                economic_issues = build_model.ek.audit_economic_coherence(facts)
                recalculated = _recalculate(out, tmp_dir)
                recalc_issues: list[str] = []
                recalc_status = "skipped"
                if recalculated is not None:
                    data_wb = load_workbook(recalculated, data_only=True)
                    recalc_issues = [
                        *_formula_error_cells(data_wb),
                        *build_model.audit_recalculated_financial_model(data_wb),
                    ]
                    recalc_status = "pass" if not recalc_issues else "fail"
                missing_sheets = [sheet for sheet in scenario.expected_sheets if sheet not in wb.sheetnames]
                missing_markers = [
                    f"{sheet}:{marker}"
                    for sheet, marker in scenario.expected_markers
                    if not _sheet_contains(wb, sheet, marker)
                ]
                issues = [*workbook_issues, *economic_issues, *recalc_issues]
                issues.extend(f"missing sheet {sheet}" for sheet in missing_sheets)
                issues.extend(f"missing marker {marker}" for marker in missing_markers)
                passed = not issues
            except Exception as exc:  # noqa: BLE001 - eval should report the failure.
                issues = [f"scenario raised {type(exc).__name__}: {exc}"]
                recalc_status = "not-run"
                passed = False
            checks.append(
                {
                    "id": f"scenario-{scenario.id}",
                    "category": "scenario-eval",
                    "status": "pass" if passed else "fail",
                    "severity": "info" if passed else "error",
                    "message": f"{scenario.id} {'passed' if passed else 'failed'}",
                    "recalc": recalc_status,
                    "issues": issues,
                }
            )
    passed_count = sum(1 for check in checks if check["status"] == "pass")
    score = round(100 * passed_count / max(len(checks), 1), 1)
    return {
        "checks": checks,
        "metrics": [
            {
                "id": "startup-finance-scenario-eval-score",
                "category": "scenario-eval",
                "value": score,
                "unit": "points",
                "band": "good" if score >= 95 else "poor",
            }
        ],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--full", action="store_true", help="Run all scenario evals instead of the quick subset.")
    parser.add_argument("--min-score", type=float, default=95.0)
    args = parser.parse_args(argv)
    result = run_scenario_eval(full=args.full)
    print(json.dumps(result, indent=2))
    score = next(metric["value"] for metric in result["metrics"] if metric["id"] == "startup-finance-scenario-eval-score")
    return 0 if score >= args.min_score else 2


if __name__ == "__main__":
    sys.exit(main())
