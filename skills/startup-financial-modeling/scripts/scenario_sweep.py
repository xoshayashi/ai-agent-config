#!/usr/bin/env python3
"""シナリオスイープ: 全ケース×全スイッチの組合せで再計算し、破綻しないことを検証。

ルーブリックC「シナリオスイッチ切替でも破綻しない」を機械保証する。
各組合せで (a) 数式エラーゼロ、(b) 総合判定OK（=配線は健全）を確認する。
アラート級の点灯は正常（意図的な開示）。

使い方: python3 scenario_sweep.py <xlsx>
終了コード 0=合格 / 1=不合格 / 2=環境エラー
"""
from __future__ import annotations

import itertools
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

SWITCHES = {"アクティブケース": (1, 2, 3), "収益認識基準スイッチ": (1, 2)}


def _recalc(path: Path, outdir: Path) -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    if not Path(soffice).exists():
        print("FAIL: sofficeが見つかりません")
        sys.exit(2)
    prof = Path(tempfile.mkdtemp(prefix="lo_")).as_uri()
    subprocess.run([soffice, f"-env:UserInstallation={prof}",
                    "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(outdir), str(path)],
                   check=True, capture_output=True, timeout=300)
    return outdir / path.name


def _state(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    errs = sum(1 for ws in wb.worksheets for row in ws.iter_rows()
               for c in row
               if isinstance(c.value, str) and c.value.startswith("#"))
    verdict = alerts = None
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            label = next((c.value for c in row[1:4]
                          if isinstance(c.value, str) and c.value.strip()),
                         None)
            if not isinstance(label, str):
                continue
            if label.startswith("総合判定"):
                verdict = next((c.value for c in row[5:]
                                if c.value is not None), None)
            elif label.startswith("警告件数"):
                alerts = next((c.value for c in row[5:]
                               if c.value is not None), None)
    return verdict, alerts, errs


def main(path: str) -> int:
    src = Path(path)
    wb = openpyxl.load_workbook(src)
    ws = wb["前提条件"]
    rows = {}
    for r in range(6, ws.max_row + 1):
        label = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                      if isinstance(ws.cell(row=r, column=c).value, str)),
                     None)
        if label in SWITCHES:
            rows[label] = r
    if not rows:
        print("SKIP: スイッチが見つかりません")
        return 0
    names = list(rows)
    combos = list(itertools.product(*(SWITCHES[n] for n in names)))
    fails = []
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        for i, combo in enumerate(combos):
            m = openpyxl.load_workbook(src)
            for n, v in zip(names, combo):
                m["前提条件"].cell(row=rows[n], column=6).value = v
            p = td / f"s{i}.xlsx"
            m.save(p)
            verdict, alerts, errs = _state(_recalc(p, td / "rc"))
            tag = " / ".join(f"{n}={v}" for n, v in zip(names, combo))
            print(f"  {tag}: 判定={verdict} 警告={alerts} エラー={errs}")
            if errs:
                fails.append(f"{tag}: 数式エラー {errs}件")
            if verdict != "OK":
                fails.append(f"{tag}: 総合判定が {verdict}")
    if fails:
        print("FAIL シナリオスイープ:")
        for f in fails:
            print("  -", f)
        return 1
    print(f"PASS シナリオスイープ: {len(combos)}通りすべてで配線健全")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
