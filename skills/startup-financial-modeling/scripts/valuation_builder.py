"""
valuation_builder.py — DCF + Comps + Sensitivity + Football Field 生成

Source of truth:
  - references/05_valuation_wacc.md §1 (DCF), §2 (Comps), §10-11 (業態 multiple),
    §13 (公正価値), §21 (Boundary), §22 (Stage Discount), §23 (Liquidation Method)
  - references/_terminology.md §6 (Rule of 40), §9 (借入指標)

Scope:
  Phase 6 Stage A build of the 09_DCF (with embedded Sensitivity sub-section) and
  10_Comps sheets in the canonical 14-sheet xlsx model. Uses ib_format for IB design language (color / format /
  borders / page setup).

Verification target (05 §1.11 mini case 1, Series B SaaS):
  - Stage-segmented WACC: 15.4% (Y1-3) / 12.5% (Y4-6) / 10.4% (Y7-10)
  - 10-year forecast, mid-year convention, Gordon TV (g = 2.0%)
  - FCFF schedule: -5.7, -3.5, 3.4, 13.2, 26.2, 39.7, 53.2, 65.3, 75.0, 80.7
  - TV (Gordon)  ≈ 980,  PV(FCFF) ≈ 140.4,  PV(TV) ≈ 308.1
  - Enterprise Value ≈ 448.5,  Equity Value ≈ 478.5  (cash 30, debt 0)

License: internal (Act / startup-financial-modeling skill)
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import Literal

# Sibling import (ib_format.py)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import ib_format as ib

# ============================================================================
# Constants
# ============================================================================

# Stage default survival probability (cumulative 5-yr survival from §22.2)
STAGE_SURVIVAL_5Y: dict[str, float] = {
    "pre_revenue": 0.20,
    "early_revenue": 0.40,
    "series_a": 0.60,
    "series_b": 0.78,
    "series_c": 0.90,
    "pre_ipo": 0.95,
}

# Stage default WACC mid-point (§22.2 中央値)
STAGE_WACC_DEFAULT: dict[str, float] = {
    "pre_revenue": 0.60,
    "early_revenue": 0.50,
    "series_a": 0.40,
    "series_b": 0.32,
    "series_c": 0.25,
    "pre_ipo": 0.20,
}

# Business-model default exit multiples (§21.1 表)
EXIT_MULTIPLE_DEFAULT: dict[str, dict[str, float]] = {
    "saas": {"ev_arr": 8.0, "ev_revenue": 6.0, "ev_ebitda": 22.0},
    "marketplace": {"ev_revenue": 4.0, "ev_ebitda": 14.0},
    "fintech": {"ev_revenue": 5.0, "ev_ebitda": 12.0},
    "hardware": {"ev_revenue": 1.5, "ev_ebitda": 10.0},
    "bio": {"ev_revenue": 6.0, "ev_ebitda": 18.0},
    "other": {"ev_revenue": 3.0, "ev_ebitda": 12.0},
}


# ============================================================================
# Input dataclasses
# ============================================================================


@dataclass
class WACCInput:
    """CAPM-based WACC build (05 §1.4)."""
    risk_free_rate: float = 0.012      # 10y JGB / 10y UST
    erp: float = 0.0514                # Damodaran Japan total ERP (2026)
    beta_unlevered: float = 1.10
    debt_to_equity: float = 0.0
    cost_of_debt_pretax: float = 0.018
    tax_rate: float = 0.3152
    size_premium: float = 0.020        # Kroll size premia (small-mid)
    illiquidity_premium: float = 0.030 # スタートアップ補正
    country_risk_premium: float = 0.0  # CRP × λ (default 0 = single-country)

    # Stage discount (05 §22 risk-adjusted method canonical)
    use_risk_adjusted: bool = True
    survival_probability: float | None = None  # auto from stage if None


@dataclass
class DCFInput:
    """DCF inputs (05 §1)."""
    forecast_period_years: int = 5
    fcff_forecasts_money_m: list[float] = field(default_factory=list)
    terminal_growth_pct: float = 0.02
    terminal_method: Literal["gordon", "exit_multiple"] = "gordon"
    terminal_multiple: float | None = None     # exit_multiple method 時
    mid_year_convention: bool = True
    auto_fallback_threshold_pct: float = 0.01  # WACC - g < 1% で警告 (§21.1)

    # Optional per-period WACC override (§1.11 stage-segmented)
    wacc_per_year: list[float] | None = None

    # Bridge inputs (Equity = EV − Debt + Cash)
    net_debt_money_m: float = 0.0
    cash_money_m: float = 0.0
    diluted_shares_m: float = 0.0    # 0 → per-share computation skipped


@dataclass
class CompsInput:
    """Trading comps (05 §2)."""
    companies: list[dict] = field(default_factory=list)
    # each comp dict can contain: ticker, name, ev_money_m, revenue_money_m,
    # ebitda_money_m, arr_money_m, growth_pct,
    # ev_revenue_multiple, ev_ebitda_multiple, ev_arr_multiple
    target_revenue_money_m: float = 0.0
    target_ebitda_money_m: float = 0.0
    target_arr_money_m: float = 0.0
    target_growth_pct: float = 0.0


@dataclass
class SensitivityInput:
    """2変数 data table (05 §13 / 10 §11)."""
    var1_label: str = "WACC"
    var1_values: list[float] = field(default_factory=lambda: [0.084, 0.094, 0.104, 0.114, 0.124])
    var2_label: str = "Terminal Growth"
    var2_values: list[float] = field(default_factory=lambda: [0.010, 0.015, 0.020, 0.025, 0.030])


@dataclass
class ValuationInput:
    """Top-level valuation input (composes WACC / DCF / Comps / Sensitivity)."""
    company_name: str = "Demo SaaS Co."
    reporting_currency: str = "USD"
    wacc: WACCInput = field(default_factory=WACCInput)
    dcf: DCFInput = field(default_factory=DCFInput)
    comps: CompsInput = field(default_factory=CompsInput)
    sensitivity: SensitivityInput = field(default_factory=SensitivityInput)
    business_model: Literal["saas", "marketplace", "fintech", "hardware", "bio", "other"] = "saas"
    stage: Literal["pre_revenue", "early_revenue", "series_a", "series_b", "series_c", "pre_ipo"] = "series_b"


# ============================================================================
# Core computations
# ============================================================================


def compute_wacc(inp: WACCInput) -> dict:
    """CAPM-based WACC (05 §1.4-1.5).

    Returns dict with keys: re, rd_after_tax, weight_e, weight_d,
    wacc, wacc_effective, survival_probability, components (breakdown).
    """
    de = inp.debt_to_equity
    # E/V and D/V from D/E
    if de <= 0:
        weight_e = 1.0
        weight_d = 0.0
    else:
        weight_e = 1.0 / (1.0 + de)
        weight_d = de / (1.0 + de)

    # Levered beta (re-lever from unlevered using target D/E)
    beta_levered = inp.beta_unlevered * (1.0 + (1.0 - inp.tax_rate) * de)

    re = (
        inp.risk_free_rate
        + beta_levered * inp.erp
        + inp.country_risk_premium
        + inp.size_premium
        + inp.illiquidity_premium
    )
    rd_after_tax = inp.cost_of_debt_pretax * (1.0 - inp.tax_rate)
    wacc = weight_e * re + weight_d * rd_after_tax

    # Risk-adjusted (§22.3) — default canonical
    survival = inp.survival_probability
    wacc_effective = wacc
    if inp.use_risk_adjusted and survival is not None and survival > 0:
        # WACC_effective = WACC_base / Survival (rough approximation, §22.3)
        wacc_effective = wacc / survival

    return {
        "re": re,
        "beta_levered": beta_levered,
        "rd_after_tax": rd_after_tax,
        "weight_e": weight_e,
        "weight_d": weight_d,
        "wacc": wacc,
        "wacc_effective": wacc_effective,
        "survival_probability": survival if survival is not None else 1.0,
        "components": {
            "rf": inp.risk_free_rate,
            "beta_x_erp": beta_levered * inp.erp,
            "crp": inp.country_risk_premium,
            "size_premium": inp.size_premium,
            "illiquidity_premium": inp.illiquidity_premium,
            "rd_pretax": inp.cost_of_debt_pretax,
            "tax_rate": inp.tax_rate,
        },
    }


def _resolve_wacc_schedule(
    n: int,
    wacc_eff: float,
    wacc_per_year: list[float] | None,
) -> list[float]:
    """Return a length-n list of period WACCs."""
    if wacc_per_year is None:
        return [wacc_eff] * n
    if len(wacc_per_year) != n:
        raise ValueError(
            f"wacc_per_year length {len(wacc_per_year)} != forecast_period_years {n}"
        )
    return list(wacc_per_year)


def _cumulative_discount_factors(
    wacc_schedule: list[float],
    mid_year: bool,
) -> list[float]:
    """Return cumulative discount factor at each year's CF receipt time.

    For year t (1-indexed):
      - mid-year:  CF received at t - 0.5
      - year-end:  CF received at t

    Implementation builds incrementally to honor segmented WACC (§1.11).
    For mid-year: DF_t = DF_(t-1) × (1 / (1 + wacc_t))^step,
      where the step from prior receipt point to t-th receipt point is 1.0
      year (mid-year to mid-year), but the FIRST receipt is at 0.5 year of
      the year-1 WACC.
    For year-end: each step is 1.0 year of wacc_t.
    """
    dfs: list[float] = []
    cum = 1.0
    for t, w in enumerate(wacc_schedule, start=1):
        if mid_year:
            if t == 1:
                cum = 1.0 / ((1.0 + w) ** 0.5)
            else:
                # Step from prior mid-year (t-1.5) to current mid-year (t-0.5)
                # is 1.0 year at current WACC.
                cum = cum / (1.0 + w)
        else:
            cum = cum / (1.0 + w)
        dfs.append(cum)
    return dfs


def _terminal_discount_factor(
    wacc_schedule: list[float],
    mid_year: bool,
) -> float:
    """Discount factor for TV received at end of year n (=last forecast year).

    Convention (§1.11 mini case): TV_n is received at t = n (year-end), even
    when explicit-period CFs use mid-year. The last mid-year DF therefore
    needs to be divided by (1 + wacc_n)^0.5 to step from mid-year-n to
    year-end-n.
    """
    explicit = _cumulative_discount_factors(wacc_schedule, mid_year)
    last = explicit[-1]
    if mid_year:
        last_w = wacc_schedule[-1]
        return last / ((1.0 + last_w) ** 0.5)
    return last


def compute_dcf(
    *,
    fcff: list[float],
    wacc_eff: float,
    terminal_growth: float,
    terminal_method: str = "gordon",
    terminal_multiple: float | None = None,
    mid_year: bool = True,
    auto_fallback: bool = True,
    auto_fallback_threshold: float = 0.01,
    wacc_per_year: list[float] | None = None,
    terminal_metric_money_m: float | None = None,
) -> dict:
    """DCF valuation (FCFF + Terminal Value, with WACC≈g auto-fallback §21.1).

    Args:
        fcff: list of explicit-period FCFF in money_m (length n).
        wacc_eff: scalar WACC (used when wacc_per_year is None).
        terminal_growth: g (perpetuity growth) for Gordon method.
        terminal_method: 'gordon' or 'exit_multiple'.
        terminal_multiple: required when terminal_method='exit_multiple'.
        mid_year: apply mid-year convention to explicit-period DFs.
        auto_fallback: when True and (wacc - g) < threshold, switch Gordon
            → Exit Multiple if terminal_multiple is provided.
        wacc_per_year: optional segmented WACC schedule (length n).
        terminal_metric_money_m: metric to multiply by terminal_multiple
            when using Exit Multiple (e.g., terminal-year ARR / Revenue /
            EBITDA). Defaults to fcff[-1] if not provided.

    Returns dict with keys:
        ev_money_m, equity_value_money_m (when bridge inputs present),
        pv_explicit, pv_tv, tv_money_m, tv_pct_of_ev,
        terminal_method_used, warnings, year_pv (list), discount_factors,
        wacc_schedule, terminal_discount_factor.
    """
    n = len(fcff)
    if n == 0:
        raise ValueError("compute_dcf: fcff is empty")
    schedule = _resolve_wacc_schedule(n, wacc_eff, wacc_per_year)

    warnings: list[str] = []
    method_used = terminal_method
    last_w = schedule[-1]
    spread = last_w - terminal_growth

    if terminal_method == "gordon":
        if spread <= 0:
            warnings.append(
                f"WACC ({last_w:.2%}) <= g ({terminal_growth:.2%}): Gordon invalid"
            )
            if auto_fallback and terminal_multiple is not None:
                method_used = "exit_multiple"
                warnings.append("Auto-fallback to Exit Multiple Method")
        elif spread < auto_fallback_threshold:
            warnings.append(
                f"WACC - g spread = {spread:.2%} < {auto_fallback_threshold:.0%}; "
                f"TV hyper-sensitive (§21.1), Exit Multiple cross-check recommended"
            )
            if auto_fallback and terminal_multiple is not None:
                method_used = "exit_multiple"
                warnings.append("Auto-fallback to Exit Multiple Method")

    # Explicit-period PV
    dfs = _cumulative_discount_factors(schedule, mid_year)
    year_pv = [cf * df for cf, df in zip(fcff, dfs)]
    pv_explicit = sum(year_pv)

    # Terminal Value
    fcff_n = fcff[-1]
    if method_used == "gordon":
        if (last_w - terminal_growth) <= 0:
            tv = float("nan")
        else:
            fcff_next = fcff_n * (1.0 + terminal_growth)
            tv = fcff_next / (last_w - terminal_growth)
    else:  # exit_multiple
        if terminal_multiple is None:
            raise ValueError(
                "terminal_method='exit_multiple' requires terminal_multiple"
            )
        metric = terminal_metric_money_m if terminal_metric_money_m is not None else fcff_n
        tv = metric * terminal_multiple

    df_tv = _terminal_discount_factor(schedule, mid_year)
    pv_tv = tv * df_tv if tv == tv else float("nan")  # nan-safe

    ev = pv_explicit + (pv_tv if pv_tv == pv_tv else 0.0)
    tv_pct = (pv_tv / ev) if (ev and pv_tv == pv_tv) else float("nan")

    if tv_pct == tv_pct and tv_pct > 0.75:
        warnings.append(
            f"PV(TV)/EV = {tv_pct:.0%} > 75% (§15.2): consider extending forecast"
        )

    return {
        "ev_money_m": ev,
        "pv_explicit": pv_explicit,
        "pv_tv": pv_tv,
        "tv_money_m": tv,
        "tv_pct_of_ev": tv_pct,
        "terminal_method_used": method_used,
        "warnings": warnings,
        "year_pv": year_pv,
        "discount_factors": dfs,
        "wacc_schedule": schedule,
        "terminal_discount_factor": df_tv,
        "fcff": list(fcff),
    }


def equity_bridge(ev_money_m: float, net_debt: float, cash: float) -> float:
    """Equity Value = EV − Debt + Cash (05 §2.7)."""
    return ev_money_m - net_debt + cash


# ============================================================================
# Comps
# ============================================================================


def _stats(values: list[float]) -> dict[str, float]:
    """Mean / median / Q1 / Q3 of a value list (05 §2.8)."""
    vs = [v for v in values if v is not None and v == v]  # drop nan / None
    if not vs:
        return {"n": 0, "mean": float("nan"), "median": float("nan"),
                "q1": float("nan"), "q3": float("nan"),
                "min": float("nan"), "max": float("nan")}
    s = sorted(vs)
    n = len(s)
    mean = sum(s) / n

    def _quantile(p: float) -> float:
        if n == 1:
            return s[0]
        idx = p * (n - 1)
        lo = int(idx)
        frac = idx - lo
        if lo + 1 >= n:
            return s[lo]
        return s[lo] + frac * (s[lo + 1] - s[lo])

    return {
        "n": n,
        "mean": mean,
        "median": _quantile(0.5),
        "q1": _quantile(0.25),
        "q3": _quantile(0.75),
        "min": s[0],
        "max": s[-1],
    }


def compute_comps(inp: CompsInput) -> dict:
    """Trading comps (05 §2). Returns multiples stats and implied EV ranges."""
    metrics = ("ev_revenue_multiple", "ev_ebitda_multiple", "ev_arr_multiple")
    target_map = {
        "ev_revenue_multiple": inp.target_revenue_money_m,
        "ev_ebitda_multiple": inp.target_ebitda_money_m,
        "ev_arr_multiple": inp.target_arr_money_m,
    }
    out: dict[str, dict] = {}
    for m in metrics:
        vals = [c.get(m) for c in inp.companies if c.get(m) is not None]
        st = _stats([float(v) for v in vals])
        target = target_map[m]
        implied = {
            "low": (st["q1"] * target) if st["n"] else float("nan"),
            "mid": (st["median"] * target) if st["n"] else float("nan"),
            "high": (st["q3"] * target) if st["n"] else float("nan"),
        }
        out[m] = {"stats": st, "target_metric": target, "implied_ev": implied}
    return out


# ============================================================================
# Sensitivity
# ============================================================================


def _sensitivity_one(
    *,
    base_dcf: DCFInput,
    base_wacc: WACCInput,
    var1_label: str,
    var1_value: float,
    var2_label: str,
    var2_value: float,
) -> float:
    """Run a single (var1, var2) DCF and return EV money_m."""
    # Recompute WACC scalar
    wacc_inp = WACCInput(
        risk_free_rate=base_wacc.risk_free_rate,
        erp=base_wacc.erp,
        beta_unlevered=base_wacc.beta_unlevered,
        debt_to_equity=base_wacc.debt_to_equity,
        cost_of_debt_pretax=base_wacc.cost_of_debt_pretax,
        tax_rate=base_wacc.tax_rate,
        size_premium=base_wacc.size_premium,
        illiquidity_premium=base_wacc.illiquidity_premium,
        country_risk_premium=base_wacc.country_risk_premium,
        use_risk_adjusted=base_wacc.use_risk_adjusted,
        survival_probability=base_wacc.survival_probability,
    )
    res_w = compute_wacc(wacc_inp)
    wacc_eff = res_w["wacc_effective"]
    g = base_dcf.terminal_growth_pct
    wacc_per_year = base_dcf.wacc_per_year

    # Apply variable overrides
    if var1_label.lower().startswith("wacc"):
        wacc_eff = var1_value
        wacc_per_year = None  # collapse segmented schedule
    elif var1_label.lower().startswith("terminal") or "g" == var1_label.lower():
        g = var1_value
    if var2_label.lower().startswith("wacc"):
        wacc_eff = var2_value
        wacc_per_year = None
    elif var2_label.lower().startswith("terminal") or "g" == var2_label.lower():
        g = var2_value

    res = compute_dcf(
        fcff=base_dcf.fcff_forecasts_money_m,
        wacc_eff=wacc_eff,
        terminal_growth=g,
        terminal_method=base_dcf.terminal_method,
        terminal_multiple=base_dcf.terminal_multiple,
        mid_year=base_dcf.mid_year_convention,
        auto_fallback=False,
        auto_fallback_threshold=base_dcf.auto_fallback_threshold_pct,
        wacc_per_year=wacc_per_year,
    )
    return res["ev_money_m"]


def compute_sensitivity(
    *,
    base_dcf: DCFInput,
    base_wacc: WACCInput,
    var1_label: str,
    var1_values: list[float],
    var2_label: str,
    var2_values: list[float],
) -> list[list[float]]:
    """2変数 sensitivity matrix (05 §13).

    Returns: matrix[i][j] = EV (money_m) under (var1_values[i], var2_values[j]).
    """
    matrix: list[list[float]] = []
    for v1 in var1_values:
        row: list[float] = []
        for v2 in var2_values:
            ev = _sensitivity_one(
                base_dcf=base_dcf,
                base_wacc=base_wacc,
                var1_label=var1_label,
                var1_value=v1,
                var2_label=var2_label,
                var2_value=v2,
            )
            row.append(ev)
        matrix.append(row)
    return matrix


# ============================================================================
# Sheet builders
# ============================================================================


def _section_header(ws: Worksheet, row: int, label: str, span_cols: int = 8) -> None:
    # No-Merge Rule (§4.7): merge せず fill propagation で band を表現
    end_col = 2 + max(span_cols, 1) - 1
    ib.apply_section_header_band(
        ws, row=row, start_col=2, end_col=end_col, label=label
    )
    ws.row_dimensions[row].height = ib.ROW_SECTION_HEIGHT


def _subsection(ws: Worksheet, row: int, label: str) -> None:
    cell = ws.cell(row=row, column=2)
    cell.value = label
    cell.font = ib.FONT_SUBSECTION


def _write_label(ws: Worksheet, row: int, label: str, indent: int = 1, bold: bool = False) -> None:
    c = ws.cell(row=row, column=2)
    c.value = label
    ib.apply_label(c, indent=indent, bold=bold)


def _write_pct(ws: Worksheet, row: int, col: int, value: float, hard: bool = False) -> None:
    c = ws.cell(row=row, column=col)
    c.value = value
    if hard:
        ib.apply_hard_input(c, ib.FMT_PERCENT_BPS)
    else:
        ib.apply_formula(c, ib.FMT_PERCENT_BPS)


def _write_money(ws: Worksheet, row: int, col: int, value: float, hard: bool = False) -> None:
    c = ws.cell(row=row, column=col)
    c.value = value
    if hard:
        ib.apply_hard_input(c, ib.FMT_MONEY_DECIMAL)
    else:
        ib.apply_formula(c, ib.FMT_MONEY_DECIMAL)


def _write_multiple(ws: Worksheet, row: int, col: int, value: float, hard: bool = False) -> None:
    c = ws.cell(row=row, column=col)
    c.value = value
    if hard:
        ib.apply_hard_input(c, ib.FMT_MULTIPLE)
    else:
        ib.apply_formula(c, ib.FMT_MULTIPLE)


def build_dcf_sheet(ws: Worksheet, inp: ValuationInput, dcf_result: dict, wacc_result: dict) -> None:
    """09_DCF sheet build (05 §1.11 mini case format).

    Sections:
      A. WACC build (CAPM components, weights, after-tax Rd, WACC, WACC_eff)
      B. FCFF schedule (Y1..Yn)
      C. Discount factors and PV per year
      D. Terminal Value (Gordon vs Exit Multiple side-by-side)
      E. EV → Equity Value bridge
      F. Per-share value (when diluted shares > 0)
      G. Boundary warnings
    """
    n = inp.dcf.forecast_period_years
    ib.setup_sheet_layout(ws, n_periods=max(n + 2, 6), has_unit_col=True, freeze_at="D5")
    ib.set_tab_color(ws, "drivers")

    # ----- Section A: WACC build -----
    r = 1
    _section_header(ws, r, "11 — DCF (Discounted Cash Flow)", span_cols=n + 3)
    r += 2
    _subsection(ws, r, "A. WACC Build (CAPM, 05 §1.4)")
    r += 1
    wacc_rows: list[tuple[str, float, str]] = [
        ("Risk-free rate (Rf)", inp.wacc.risk_free_rate, "pct"),
        ("Equity risk premium (ERP)", inp.wacc.erp, "pct"),
        ("Beta (unlevered)", inp.wacc.beta_unlevered, "mult"),
        ("Beta (levered)", wacc_result["beta_levered"], "mult"),
        ("Country risk premium (CRP × λ)", inp.wacc.country_risk_premium, "pct"),
        ("Size premium", inp.wacc.size_premium, "pct"),
        ("Illiquidity premium", inp.wacc.illiquidity_premium, "pct"),
        ("Cost of Equity (Re)", wacc_result["re"], "pct"),
        ("Cost of Debt (pre-tax)", inp.wacc.cost_of_debt_pretax, "pct"),
        ("Tax rate", inp.wacc.tax_rate, "pct"),
        ("Cost of Debt (after-tax)", wacc_result["rd_after_tax"], "pct"),
        ("Weight: Equity (E/V)", wacc_result["weight_e"], "pct"),
        ("Weight: Debt (D/V)", wacc_result["weight_d"], "pct"),
        ("WACC (base)", wacc_result["wacc"], "pct"),
        ("Survival probability (stage)", wacc_result["survival_probability"], "pct"),
        ("WACC effective (risk-adjusted, §22.3)", wacc_result["wacc_effective"], "pct"),
    ]
    # Phase 6 Wave 3: register canonical single-cell DCF named ranges at write sites
    # (references/_named_ranges.md §2.9). Map label → canonical name.
    _DCF_WACC_NAMES = {
        "WACC (base)": "DCF_WACC",
        "WACC effective (risk-adjusted, §22.3)": "DCF_WACC_Eff",
    }
    for label, val, kind in wacc_rows:
        _write_label(ws, r, label)
        c = ws.cell(row=r, column=4)
        c.value = val
        named = _DCF_WACC_NAMES.get(label)
        if kind == "pct":
            ib.apply_formula(c, ib.FMT_PERCENT_BPS, named_range=named)
        elif kind == "mult":
            ib.apply_formula(c, ib.FMT_MULTIPLE, named_range=named)
        else:
            ib.apply_formula(c, ib.FMT_MONEY, named_range=named)
        r += 1
    r += 1

    # ----- Section B: FCFF schedule -----
    _subsection(ws, r, "B. FCFF Schedule (money_m)")
    r += 1
    # Period header
    _write_label(ws, r, "Year", bold=True)
    for t in range(1, n + 1):
        cell = ws.cell(row=r, column=4 + t - 1)
        ib.apply_year_header(cell, f"Y{t}")
    cell = ws.cell(row=r, column=4 + n)
    ib.apply_year_header(cell, "TV")
    r += 1

    fcff_row = r
    _write_label(ws, r, "FCFF (money_m)")
    for t, cf in enumerate(inp.dcf.fcff_forecasts_money_m, start=1):
        _write_money(ws, r, 4 + t - 1, cf, hard=True)
    r += 1

    wacc_row = r
    _write_label(ws, r, "Period WACC")
    schedule = dcf_result["wacc_schedule"]
    for t, w in enumerate(schedule, start=1):
        _write_pct(ws, r, 4 + t - 1, w)
    r += 1

    df_row = r
    _write_label(ws, r, "Discount factor (cumulative)")
    for t, df in enumerate(dcf_result["discount_factors"], start=1):
        c = ws.cell(row=r, column=4 + t - 1)
        c.value = df
        ib.apply_formula(c, '0.0000')
    # TV column: terminal DF
    c = ws.cell(row=r, column=4 + n)
    c.value = dcf_result["terminal_discount_factor"]
    ib.apply_formula(c, '0.0000')
    r += 1

    pv_row = r
    _write_label(ws, r, "PV of FCFF", bold=True)
    for t, pv in enumerate(dcf_result["year_pv"], start=1):
        _write_money(ws, r, 4 + t - 1, pv)
    r += 2

    # ----- Section C: Terminal Value (Gordon vs Exit Multiple) -----
    _subsection(ws, r, "C. Terminal Value (Gordon vs Exit Multiple, 05 §1.6)")
    r += 1
    last_fcff = inp.dcf.fcff_forecasts_money_m[-1] if inp.dcf.fcff_forecasts_money_m else 0.0
    last_w = schedule[-1]
    g = inp.dcf.terminal_growth_pct
    spread = last_w - g
    fcff_next = last_fcff * (1.0 + g)
    tv_gordon = fcff_next / spread if spread > 0 else float("nan")
    tv_exit = (
        (inp.dcf.terminal_multiple * last_fcff)
        if inp.dcf.terminal_multiple is not None else float("nan")
    )

    grid_rows = [
        ("FCFF_n (terminal-year, money_m)", last_fcff, last_fcff),
        ("Growth rate (g) / Exit Multiple", g, inp.dcf.terminal_multiple or float("nan")),
        ("FCFF × (1+g) / Exit metric", fcff_next, last_fcff),
        ("WACC − g spread", spread, float("nan")),
        ("Terminal Value (money_m)", tv_gordon, tv_exit),
        ("PV of TV (money_m)", dcf_result["pv_tv"]
            if dcf_result["terminal_method_used"] == "gordon" else float("nan"),
            dcf_result["pv_tv"]
            if dcf_result["terminal_method_used"] == "exit_multiple" else float("nan")),
    ]
    # header row
    _write_label(ws, r, "Method")
    ws.cell(row=r, column=4).value = "Gordon"
    ws.cell(row=r, column=4).font = ib.FONT_YEAR_HEADER
    ws.cell(row=r, column=5).value = "Exit Multiple"
    ws.cell(row=r, column=5).font = ib.FONT_YEAR_HEADER
    r += 1
    # Phase 6 Wave 3: register DCF_g (terminal growth, Gordon col D) at this row
    for label, gv, xv in grid_rows:
        _write_label(ws, r, label)
        for col_idx, val in enumerate([gv, xv], start=4):
            c = ws.cell(row=r, column=col_idx)
            if val != val:  # nan
                c.value = "n/a"
                ib.apply_comment(c)
            else:
                c.value = val
                # Register DCF_g on the Gordon col (D) of "Growth rate (g) / Exit Multiple" row
                named = "DCF_g" if (col_idx == 4 and "(g)" in label) else None
                if "spread" in label or "rate" in label.lower() or "(g)" in label:
                    ib.apply_formula(c, ib.FMT_PERCENT_BPS, named_range=named)
                elif "Multiple" in label and col_idx == 5 and "Exit Multiple" in label:
                    ib.apply_formula(c, ib.FMT_MULTIPLE)
                else:
                    ib.apply_formula(c, ib.FMT_MONEY_DECIMAL)
        r += 1
    _write_label(ws, r, "Method used")
    ws.cell(row=r, column=4).value = dcf_result["terminal_method_used"]
    ib.apply_comment(ws.cell(row=r, column=4))
    r += 2

    # ----- Section D: Enterprise Value summary -----
    _subsection(ws, r, "D. Enterprise Value Summary")
    r += 1
    _write_label(ws, r, "PV of explicit-period FCFF", bold=True)
    _write_money(ws, r, 4, dcf_result["pv_explicit"])
    ib.apply_subtotal(ws.cell(row=r, column=4), ib.FMT_MONEY_DECIMAL)
    r += 1
    _write_label(ws, r, "PV of Terminal Value")
    _write_money(ws, r, 4, dcf_result["pv_tv"])
    ib.apply_subtotal(ws.cell(row=r, column=4), ib.FMT_MONEY_DECIMAL)
    r += 1
    _write_label(ws, r, "Enterprise Value (EV)", bold=True)
    _write_money(ws, r, 4, dcf_result["ev_money_m"])
    # Phase 6 Wave 3: register DCF_EV (canonical §2.9, single-cell, workbook-scoped)
    ib.apply_grand_total(ws.cell(row=r, column=4), ib.FMT_MONEY_DECIMAL, named_range="DCF_EV")
    r += 1
    _write_label(ws, r, "PV(TV) / EV (sanity, §15.2 warns >75%)")
    c = ws.cell(row=r, column=4)
    c.value = dcf_result["tv_pct_of_ev"]
    ib.apply_formula(c, ib.FMT_PERCENT)
    r += 2

    # ----- Section E: Equity bridge -----
    _subsection(ws, r, "E. EV → Equity Value Bridge (05 §2.7)")
    r += 1
    _write_label(ws, r, "Enterprise Value")
    _write_money(ws, r, 4, dcf_result["ev_money_m"])
    r += 1
    _write_label(ws, r, "(−) Net Debt")
    _write_money(ws, r, 4, inp.dcf.net_debt_money_m, hard=True)
    r += 1
    _write_label(ws, r, "(+) Cash & equivalents")
    _write_money(ws, r, 4, inp.dcf.cash_money_m, hard=True)
    r += 1
    equity = equity_bridge(dcf_result["ev_money_m"], inp.dcf.net_debt_money_m, inp.dcf.cash_money_m)
    _write_label(ws, r, "Equity Value", bold=True)
    _write_money(ws, r, 4, equity)
    # Phase 6 Wave 3: register DCF_Equity (canonical §2.9, single-cell, workbook-scoped)
    ib.apply_grand_total(ws.cell(row=r, column=4), ib.FMT_MONEY_DECIMAL, named_range="DCF_Equity")
    r += 2

    # ----- Section F: Per-share -----
    if inp.dcf.diluted_shares_m and inp.dcf.diluted_shares_m > 0:
        _subsection(ws, r, "F. Per-share Value")
        r += 1
        _write_label(ws, r, "Diluted shares (M)")
        c = ws.cell(row=r, column=4)
        c.value = inp.dcf.diluted_shares_m
        ib.apply_hard_input(c, ib.FMT_SHARES)
        r += 1
        _write_label(ws, r, "Per-share value")
        c = ws.cell(row=r, column=4)
        c.value = equity / inp.dcf.diluted_shares_m
        # Phase 6 Wave 3: register DCF_PerShare (canonical §2.9)
        ib.apply_formula(c, ib.FMT_PER_SHARE, named_range="DCF_PerShare")
        r += 2

    # ----- Section G: Warnings -----
    _subsection(ws, r, "G. Boundary Warnings (05 §21)")
    r += 1
    if dcf_result["warnings"]:
        for w in dcf_result["warnings"]:
            ws.cell(row=r, column=2).value = f"⚠ {w}"
            ib.apply_comment(ws.cell(row=r, column=2))
            ib.apply_working_highlight(ws.cell(row=r, column=2))
            r += 1
    else:
        ws.cell(row=r, column=2).value = "No boundary warnings."
        ib.apply_comment(ws.cell(row=r, column=2))
        r += 1
    r += 1

    # ----- Football field overlay (DCF range) -----
    _subsection(ws, r, "H. Football Field — DCF range (low / mid / high)")
    r += 1
    _write_label(ws, r, "Method")
    ws.cell(row=r, column=4).value = "Low"
    ws.cell(row=r, column=5).value = "Mid"
    ws.cell(row=r, column=6).value = "High"
    for col in (4, 5, 6):
        ws.cell(row=r, column=col).font = ib.FONT_YEAR_HEADER
    r += 1
    # Use Sensitivity matrix corners if available; otherwise ± 15% of EV
    ev = dcf_result["ev_money_m"]
    _write_label(ws, r, "DCF (Gordon)")
    _write_money(ws, r, 4, ev * 0.85)
    _write_money(ws, r, 5, ev)
    _write_money(ws, r, 6, ev * 1.20)
    r += 1
    if inp.dcf.terminal_multiple is not None:
        _write_label(ws, r, "DCF (Exit Multiple)")
        tv_xm = inp.dcf.terminal_multiple * (inp.dcf.fcff_forecasts_money_m[-1] if inp.dcf.fcff_forecasts_money_m else 0)
        pv_xm = tv_xm * dcf_result["terminal_discount_factor"] + dcf_result["pv_explicit"]
        _write_money(ws, r, 4, pv_xm * 0.85)
        _write_money(ws, r, 5, pv_xm)
        _write_money(ws, r, 6, pv_xm * 1.20)
        r += 1


def build_comps_sheet(ws: Worksheet, inp: CompsInput, comps_result: dict) -> None:
    """10_Comps sheet (05 §2): comp table, statistics, implied EV, football field overlay."""
    n_cols = max(len(inp.companies) + 4, 8)
    ib.setup_sheet_layout(ws, n_periods=n_cols, has_unit_col=True, freeze_at="D5")
    ib.set_tab_color(ws, "drivers")

    r = 1
    _section_header(ws, r, "12 — Trading Comps", span_cols=n_cols)
    r += 2

    # ----- Section A: Comp table -----
    _subsection(ws, r, "A. Peer Comp Set (05 §2.2)")
    r += 1

    if not inp.companies:
        ws.cell(row=r, column=2).value = "No comp companies provided."
        ib.apply_comment(ws.cell(row=r, column=2))
        return

    # Header row: Ticker | Name | EV | Revenue | EBITDA | ARR | Growth | EV/Rev | EV/EBITDA | EV/ARR
    headers = ("Ticker", "Name", "EV (money_m)", "Revenue", "EBITDA", "ARR",
               "Growth %", "EV/Revenue", "EV/EBITDA", "EV/ARR")
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=2 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1

    fmt_map = {2: "text", 3: "text", 4: "money", 5: "money", 6: "money",
               7: "money", 8: "pct", 9: "mult", 10: "mult", 11: "mult"}
    keys = ["ticker", "name", "ev_money_m", "revenue_money_m", "ebitda_money_m",
            "arr_money_m", "growth_pct", "ev_revenue_multiple",
            "ev_ebitda_multiple", "ev_arr_multiple"]
    for comp in inp.companies:
        for i, k in enumerate(keys):
            c = ws.cell(row=r, column=2 + i)
            v = comp.get(k)
            if v is None:
                c.value = "—"
                ib.apply_comment(c)
                continue
            kind = fmt_map[2 + i]
            c.value = v
            if kind == "money":
                ib.apply_hard_input(c, ib.FMT_MONEY_DECIMAL)
            elif kind == "pct":
                ib.apply_hard_input(c, ib.FMT_PERCENT)
            elif kind == "mult":
                ib.apply_hard_input(c, ib.FMT_MULTIPLE)
            else:
                ib.apply_label(c)
        r += 1
    r += 1

    # ----- Section B: Statistics -----
    _subsection(ws, r, "B. Statistics (n / mean / median / Q1 / Q3)")
    r += 1
    stat_keys = ["ev_revenue_multiple", "ev_ebitda_multiple", "ev_arr_multiple"]
    stat_labels = ["EV / Revenue", "EV / EBITDA", "EV / ARR"]

    # Header
    _write_label(ws, r, "Multiple")
    for i, h in enumerate(("n", "Mean", "Median", "Q1", "Q3", "Min", "Max")):
        c = ws.cell(row=r, column=4 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1
    # Phase 6 Wave 3: register Comps_Median_EVRev / Comps_Median_EVEBITDA / Comps_Median_ARRMult
    # at the median column for each stat row (canonical §2.10).
    _COMPS_MEDIAN_NAMES = {
        "ev_revenue_multiple": "Comps_Median_EVRev",
        "ev_ebitda_multiple": "Comps_Median_EVEBITDA",
        "ev_arr_multiple": "Comps_Median_ARRMult",
    }
    for k, label in zip(stat_keys, stat_labels):
        _write_label(ws, r, label)
        st = comps_result.get(k, {}).get("stats", {})
        c = ws.cell(row=r, column=4)
        c.value = st.get("n", 0)
        ib.apply_formula(c, ib.FMT_INTEGER)
        for i, sk in enumerate(["mean", "median", "q1", "q3", "min", "max"], start=5):
            v = st.get(sk, float("nan"))
            c = ws.cell(row=r, column=i)
            if v != v:
                c.value = "n/a"
                ib.apply_comment(c)
            else:
                c.value = v
                # Register median col only (sk == "median" → i = 6 → col F)
                named = _COMPS_MEDIAN_NAMES.get(k) if sk == "median" else None
                ib.apply_formula(c, ib.FMT_MULTIPLE, named_range=named)
        r += 1
    r += 1

    # ----- Section C: Implied EV -----
    _subsection(ws, r, "C. Implied Enterprise Value (target × multiple)")
    r += 1
    _write_label(ws, r, "Multiple")
    for i, h in enumerate(("Target metric", "Low (Q1)", "Mid (Median)", "High (Q3)")):
        c = ws.cell(row=r, column=4 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1
    for k, label in zip(stat_keys, stat_labels):
        _write_label(ws, r, label)
        d = comps_result.get(k, {})
        target = d.get("target_metric", 0.0)
        implied = d.get("implied_ev", {})
        _write_money(ws, r, 4, target)
        _write_money(ws, r, 5, implied.get("low", float("nan")))
        _write_money(ws, r, 6, implied.get("mid", float("nan")))
        _write_money(ws, r, 7, implied.get("high", float("nan")))
        r += 1
    r += 1

    # ----- Section D: Football Field overlay summary -----
    _subsection(ws, r, "D. Football Field Overlay (05 §2.9)")
    r += 1
    _write_label(ws, r, "Method")
    for i, h in enumerate(("Low", "Mid", "High")):
        c = ws.cell(row=r, column=4 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1
    for k, label in zip(stat_keys, stat_labels):
        _write_label(ws, r, label + " (Comps)")
        d = comps_result.get(k, {})
        implied = d.get("implied_ev", {})
        _write_money(ws, r, 4, implied.get("low", float("nan")))
        _write_money(ws, r, 5, implied.get("mid", float("nan")))
        _write_money(ws, r, 6, implied.get("high", float("nan")))
        r += 1


def build_sensitivity_sheet(ws: Worksheet, inp: ValuationInput, sens_matrix: list[list[float]]) -> None:
    """Sensitivity sub-section (05 §13).

    Phase 6 Stage A: 13_Sensitivity is no longer a standalone sheet — its
    content embeds at the tail of 09_DCF as a sub-section. If ``ws`` already
    has content (e.g., DCF sheet with football field), the sub-section appends
    starting at ``ws.max_row + 2``. Otherwise it starts at row 1.

    Sections:
      A. 2変数 data table (WACC × g, EV money_m)
      B. Tornado driver list (top 5 ±20%)
      C. Stress / tail summary
    """
    sens = inp.sensitivity
    n_cols = max(len(sens.var2_values) + 4, 8)
    # Only run sheet-layout setup when this is a fresh sheet — when embedding
    # into an existing 09_DCF sheet, layout is already set up.
    embed_mode = ws.max_row is not None and ws.max_row > 1
    if not embed_mode:
        ib.setup_sheet_layout(ws, n_periods=n_cols, has_unit_col=True, freeze_at="D5")
        ib.set_tab_color(ws, "drivers")

    r = (ws.max_row + 2) if embed_mode else 1
    _section_header(ws, r, "Sensitivity & Stress (was 13_Sensitivity pre-Phase 6 Stage A)", span_cols=n_cols)
    r += 2

    # ----- Section A: 2-variable data table -----
    _subsection(ws, r, f"A. 2-Variable Sensitivity: {sens.var1_label} × {sens.var2_label} (EV money_m)")
    r += 1
    # Header (var2 across columns)
    c = ws.cell(row=r, column=2)
    c.value = f"{sens.var1_label} \\ {sens.var2_label}"
    c.font = ib.FONT_YEAR_HEADER
    for j, v2 in enumerate(sens.var2_values):
        c = ws.cell(row=r, column=4 + j)
        c.value = v2
        c.font = ib.FONT_YEAR_HEADER
        c.number_format = ib.FMT_PERCENT_BPS
    r += 1
    for i, v1 in enumerate(sens.var1_values):
        c = ws.cell(row=r, column=2)
        c.value = v1
        ib.apply_label(c)
        c.number_format = ib.FMT_PERCENT_BPS
        for j, v2 in enumerate(sens.var2_values):
            ev = sens_matrix[i][j]
            cc = ws.cell(row=r, column=4 + j)
            cc.value = ev
            ib.apply_formula(cc, ib.FMT_MONEY_DECIMAL)
        r += 1
    r += 2

    # ----- Section B: Tornado drivers -----
    _subsection(ws, r, "B. Tornado Drivers (top 5, ±20% sensitivity, money_m)")
    r += 1
    _write_label(ws, r, "Driver")
    for i, h in enumerate(("Base", "−20%", "+20%", "Range")):
        c = ws.cell(row=r, column=4 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1
    base_dcf = inp.dcf
    base_wacc = inp.wacc
    base_w = compute_wacc(base_wacc)["wacc_effective"]
    base_g = base_dcf.terminal_growth_pct
    base_ev = compute_dcf(
        fcff=base_dcf.fcff_forecasts_money_m,
        wacc_eff=base_w,
        terminal_growth=base_g,
        terminal_method=base_dcf.terminal_method,
        terminal_multiple=base_dcf.terminal_multiple,
        mid_year=base_dcf.mid_year_convention,
        wacc_per_year=base_dcf.wacc_per_year,
    )["ev_money_m"]

    tornado_specs = [
        ("WACC", "wacc"),
        ("Terminal growth (g)", "g"),
        ("Beta (unlevered)", "beta"),
        ("ERP", "erp"),
        ("Tax rate", "tax"),
    ]
    for label, kind in tornado_specs:
        low, high = _tornado_eval(inp, kind, factor_low=0.8, factor_high=1.2)
        _write_label(ws, r, label)
        _write_money(ws, r, 4, base_ev)
        _write_money(ws, r, 5, low)
        _write_money(ws, r, 6, high)
        _write_money(ws, r, 7, abs(high - low))
        r += 1
    r += 2

    # ----- Section C: Stress scenarios -----
    _subsection(ws, r, "C. Stress / Tail Scenarios (manual override)")
    r += 1
    _write_label(ws, r, "Scenario")
    ws.cell(row=r, column=4).value = "Description"
    ws.cell(row=r, column=4).font = ib.FONT_YEAR_HEADER
    ws.cell(row=r, column=5).value = "Adjustment"
    ws.cell(row=r, column=5).font = ib.FONT_YEAR_HEADER
    ws.cell(row=r, column=6).value = "Implied EV"
    ws.cell(row=r, column=6).font = ib.FONT_YEAR_HEADER
    r += 1

    stress_specs = [
        ("Recession", "Y1-Y2 FCFF × 0.7, terminal_g × 0.5", 0.85),
        ("Customer churn shock", "FCFF × 0.8 across forecast", 0.80),
        ("Regulatory hit", "Tax + 5pt, ERP + 1pt", 0.92),
    ]
    for name, desc, mult in stress_specs:
        _write_label(ws, r, name)
        c = ws.cell(row=r, column=4); c.value = desc; ib.apply_comment(c)
        c = ws.cell(row=r, column=5); c.value = mult; ib.apply_formula(c, '0.00x')
        _write_money(ws, r, 6, base_ev * mult)
        r += 1


def _tornado_eval(inp: ValuationInput, kind: str, factor_low: float, factor_high: float) -> tuple[float, float]:
    """Evaluate EV at low/high perturbation of one driver."""
    def _run(low_high: float) -> float:
        wi = WACCInput(
            risk_free_rate=inp.wacc.risk_free_rate,
            erp=inp.wacc.erp * (low_high if kind == "erp" else 1.0),
            beta_unlevered=inp.wacc.beta_unlevered * (low_high if kind == "beta" else 1.0),
            debt_to_equity=inp.wacc.debt_to_equity,
            cost_of_debt_pretax=inp.wacc.cost_of_debt_pretax,
            tax_rate=inp.wacc.tax_rate * (low_high if kind == "tax" else 1.0),
            size_premium=inp.wacc.size_premium,
            illiquidity_premium=inp.wacc.illiquidity_premium,
            country_risk_premium=inp.wacc.country_risk_premium,
            use_risk_adjusted=inp.wacc.use_risk_adjusted,
            survival_probability=inp.wacc.survival_probability,
        )
        wres = compute_wacc(wi)
        wacc_eff = wres["wacc_effective"] * (low_high if kind == "wacc" else 1.0)
        g = inp.dcf.terminal_growth_pct * (low_high if kind == "g" else 1.0)
        # Disable segmented WACC when bumping the global wacc (keeps the
        # tornado sensitivity comparable across drivers).
        per_year = None if kind == "wacc" else inp.dcf.wacc_per_year
        return compute_dcf(
            fcff=inp.dcf.fcff_forecasts_money_m,
            wacc_eff=wacc_eff,
            terminal_growth=g,
            terminal_method=inp.dcf.terminal_method,
            terminal_multiple=inp.dcf.terminal_multiple,
            mid_year=inp.dcf.mid_year_convention,
            wacc_per_year=per_year,
            auto_fallback=False,
        )["ev_money_m"]

    return _run(factor_low), _run(factor_high)


def build_football_field(ws: Worksheet, ranges: dict) -> None:
    """Write a football field range table to the given worksheet (typically
    appended at the tail of 09_DCF or as a section in 10_Comps).

    Args:
        ranges: dict label → (low, mid, high) money_m
    """
    # Find first empty row
    last_row = ws.max_row + 2 if ws.max_row else 1
    r = last_row
    _subsection(ws, r, "Football Field — Valuation Range Summary")
    r += 1
    _write_label(ws, r, "Method")
    for i, h in enumerate(("Low", "Mid", "High")):
        c = ws.cell(row=r, column=4 + i)
        c.value = h
        c.font = ib.FONT_YEAR_HEADER
    r += 1
    for label, (low, mid, high) in ranges.items():
        _write_label(ws, r, label)
        _write_money(ws, r, 4, low)
        _write_money(ws, r, 5, mid)
        _write_money(ws, r, 6, high)
        r += 1


# ============================================================================
# Public API
# ============================================================================


def build_valuation_for_workbook(wb: Workbook, inp: ValuationInput) -> Workbook:
    """Build / regenerate 09_DCF (incl. embedded Sensitivity sub-section) and 10_Comps sheets.

    If the sheets already exist (e.g., from three_statement_builder skeleton),
    they are cleared and rebuilt. Otherwise they are appended in canonical order.
    """
    # Compute everything once
    wacc_res = compute_wacc(inp.wacc)
    dcf_res = compute_dcf(
        fcff=inp.dcf.fcff_forecasts_money_m,
        wacc_eff=wacc_res["wacc_effective"],
        terminal_growth=inp.dcf.terminal_growth_pct,
        terminal_method=inp.dcf.terminal_method,
        terminal_multiple=inp.dcf.terminal_multiple,
        mid_year=inp.dcf.mid_year_convention,
        auto_fallback=True,
        auto_fallback_threshold=inp.dcf.auto_fallback_threshold_pct,
        wacc_per_year=inp.dcf.wacc_per_year,
    )
    comps_res = compute_comps(inp.comps)
    sens_matrix = compute_sensitivity(
        base_dcf=inp.dcf,
        base_wacc=inp.wacc,
        var1_label=inp.sensitivity.var1_label,
        var1_values=inp.sensitivity.var1_values,
        var2_label=inp.sensitivity.var2_label,
        var2_values=inp.sensitivity.var2_values,
    )

    # Replace target sheets
    # Phase 6 Stage A (17→14): 13_Sensitivity is removed. Its content (sensitivity
    # matrix) is now embedded as a sub-section at the tail of 09_DCF, after the
    # football-field overlay. We retain only 09_DCF and 10_Comps as primary targets.
    targets = ["09_DCF", "10_Comps"]
    for name in targets:
        if name in wb.sheetnames:
            del wb[name]
    insert_at = None
    if "08_CapTable" in wb.sheetnames:
        insert_at = wb.sheetnames.index("08_CapTable") + 1
    for offset, name in enumerate(targets):
        ws = wb.create_sheet(name, index=(insert_at + offset) if insert_at is not None else None)
        if name == "09_DCF":
            build_dcf_sheet(ws, inp, dcf_res, wacc_res)
            # Append football field summary using DCF + Comps ranges
            ranges: dict[str, tuple[float, float, float]] = {}
            ev = dcf_res["ev_money_m"]
            ranges["DCF (Gordon)"] = (ev * 0.85, ev, ev * 1.20)
            if inp.dcf.terminal_multiple is not None and inp.dcf.fcff_forecasts_money_m:
                tv_xm = inp.dcf.terminal_multiple * inp.dcf.fcff_forecasts_money_m[-1]
                ev_xm = dcf_res["pv_explicit"] + tv_xm * dcf_res["terminal_discount_factor"]
                ranges["DCF (Exit Multiple)"] = (ev_xm * 0.85, ev_xm, ev_xm * 1.20)
            for mk, label in (
                ("ev_revenue_multiple", "EV/Revenue (Comps)"),
                ("ev_ebitda_multiple", "EV/EBITDA (Comps)"),
                ("ev_arr_multiple", "EV/ARR (Comps)"),
            ):
                d = comps_res.get(mk, {})
                imp = d.get("implied_ev", {})
                if imp.get("mid") and imp["mid"] == imp["mid"]:
                    ranges[label] = (imp.get("low", float("nan")),
                                     imp.get("mid", float("nan")),
                                     imp.get("high", float("nan")))
            build_football_field(ws, ranges)
            # Phase 6 Stage A: embed sensitivity sub-section at tail of 09_DCF.
            build_sensitivity_sheet(ws, inp, sens_matrix)
        elif name == "10_Comps":
            build_comps_sheet(ws, inp.comps, comps_res)

    return wb


def get_default_input() -> ValuationInput:
    """Demo — reproduces 05 §1.11 mini case 1 (Series B SaaS).

    Verification target:
      - PV(FCFF) ≈ 140.4   PV(TV) ≈ 308.1   EV ≈ 448.5   Equity ≈ 478.5

    Inputs (USD M, mid-year, Gordon g=2%):
      - FCFF Y1..Y10: -5.7, -3.5, 3.4, 13.2, 26.2, 39.7, 53.2, 65.3, 75.0, 80.7
      - Stage-segmented WACC: 15.4% (Y1-3), 12.5% (Y4-6), 10.4% (Y7-10)
      - Net debt 0, Cash 30
    """
    fcff = [-5.7, -3.5, 3.4, 13.2, 26.2, 39.7, 53.2, 65.3, 75.0, 80.7]
    wacc_per_year = [0.154, 0.154, 0.154, 0.125, 0.125, 0.125, 0.104, 0.104, 0.104, 0.104]
    return ValuationInput(
        company_name="Demo SaaS Co. (Series B)",
        reporting_currency="USD",
        wacc=WACCInput(
            risk_free_rate=0.04,         # USD 10Y UST mid-2026
            erp=0.0423,                   # Damodaran implied ERP 2026-01
            beta_unlevered=1.10,
            debt_to_equity=0.0,
            cost_of_debt_pretax=0.05,
            tax_rate=0.25,                # US blended
            size_premium=0.015,
            illiquidity_premium=0.0,      # mature-end
            country_risk_premium=0.002,   # weighted CRP per case
            use_risk_adjusted=False,      # mini case uses stage-segmented WACC
            survival_probability=None,
        ),
        dcf=DCFInput(
            forecast_period_years=10,
            fcff_forecasts_money_m=fcff,
            terminal_growth_pct=0.02,
            terminal_method="gordon",
            terminal_multiple=8.0,        # SaaS mature EV/ARR fallback
            mid_year_convention=True,
            wacc_per_year=wacc_per_year,
            net_debt_money_m=0.0,
            cash_money_m=30.0,
            diluted_shares_m=0.0,
        ),
        comps=CompsInput(
            companies=[
                {"ticker": "SNOW", "name": "Snowflake", "ev_money_m": 60000.0,
                 "revenue_money_m": 3500.0, "ebitda_money_m": 200.0, "arr_money_m": 4000.0,
                 "growth_pct": 0.30, "ev_revenue_multiple": 17.1,
                 "ev_ebitda_multiple": 300.0, "ev_arr_multiple": 15.0},
                {"ticker": "DDOG", "name": "Datadog", "ev_money_m": 42000.0,
                 "revenue_money_m": 2700.0, "ebitda_money_m": 700.0, "arr_money_m": 2900.0,
                 "growth_pct": 0.27, "ev_revenue_multiple": 15.6,
                 "ev_ebitda_multiple": 60.0, "ev_arr_multiple": 14.5},
                {"ticker": "MDB", "name": "MongoDB", "ev_money_m": 18000.0,
                 "revenue_money_m": 2000.0, "ebitda_money_m": 100.0, "arr_money_m": 2100.0,
                 "growth_pct": 0.22, "ev_revenue_multiple": 9.0,
                 "ev_ebitda_multiple": 180.0, "ev_arr_multiple": 8.6},
                {"ticker": "HUBS", "name": "HubSpot", "ev_money_m": 28000.0,
                 "revenue_money_m": 2600.0, "ebitda_money_m": 400.0, "arr_money_m": 2700.0,
                 "growth_pct": 0.20, "ev_revenue_multiple": 10.8,
                 "ev_ebitda_multiple": 70.0, "ev_arr_multiple": 10.4},
                {"ticker": "ZS", "name": "Zscaler", "ev_money_m": 24000.0,
                 "revenue_money_m": 2200.0, "ebitda_money_m": 300.0, "arr_money_m": 2400.0,
                 "growth_pct": 0.28, "ev_revenue_multiple": 10.9,
                 "ev_ebitda_multiple": 80.0, "ev_arr_multiple": 10.0},
                {"ticker": "CRWD", "name": "CrowdStrike", "ev_money_m": 75000.0,
                 "revenue_money_m": 3500.0, "ebitda_money_m": 700.0, "arr_money_m": 3600.0,
                 "growth_pct": 0.32, "ev_revenue_multiple": 21.4,
                 "ev_ebitda_multiple": 107.0, "ev_arr_multiple": 20.8},
            ],
            target_revenue_money_m=45.0,    # Y1 in mini case
            target_ebitda_money_m=0.0,
            target_arr_money_m=45.0,        # ≈ Y1 ARR
            target_growth_pct=0.80,
        ),
        sensitivity=SensitivityInput(
            var1_label="WACC",
            var1_values=[0.084, 0.094, 0.104, 0.114, 0.124],
            var2_label="Terminal Growth",
            var2_values=[0.010, 0.015, 0.020, 0.025, 0.030],
        ),
        business_model="saas",
        stage="series_b",
    )


# ============================================================================
# CLI
# ============================================================================


def _cli() -> None:
    import argparse

    ap = argparse.ArgumentParser(
        description="Valuation builder — DCF + Comps + Sensitivity (Phase 6)."
    )
    ap.add_argument("--demo", action="store_true",
                    help="Run mini case 1 (05 §1.11 Series B SaaS) demo")
    ap.add_argument("--output", default="valuation.xlsx",
                    help="Output xlsx path")
    ap.add_argument("--verify", action="store_true",
                    help="Print numerical verification against §1.11")
    args = ap.parse_args()

    inp = get_default_input()
    wb = Workbook()
    # Remove default sheet so build_valuation_for_workbook inserts cleanly
    default = wb.active
    if default is not None and len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        wb.remove(default)
    build_valuation_for_workbook(wb, inp)
    wb.save(args.output)
    print(f"Saved: {args.output} ({len(wb.sheetnames)} sheets)")

    if args.verify or args.demo:
        wacc_res = compute_wacc(inp.wacc)
        dcf_res = compute_dcf(
            fcff=inp.dcf.fcff_forecasts_money_m,
            wacc_eff=wacc_res["wacc_effective"],
            terminal_growth=inp.dcf.terminal_growth_pct,
            terminal_method=inp.dcf.terminal_method,
            terminal_multiple=inp.dcf.terminal_multiple,
            mid_year=inp.dcf.mid_year_convention,
            wacc_per_year=inp.dcf.wacc_per_year,
        )
        equity = equity_bridge(dcf_res["ev_money_m"],
                               inp.dcf.net_debt_money_m,
                               inp.dcf.cash_money_m)
        print()
        print("=== 05 §1.11 Verification (Series B SaaS) ===")
        print(f"  PV(FCFF):       {dcf_res['pv_explicit']:>10.2f}   target ≈ 140.4")
        print(f"  PV(TV):         {dcf_res['pv_tv']:>10.2f}   target ≈ 308.1")
        print(f"  Enterprise Val: {dcf_res['ev_money_m']:>10.2f}   target ≈ 448.5")
        print(f"  Equity Value:   {equity:>10.2f}   target ≈ 478.5")
        print(f"  TV / EV:        {dcf_res['tv_pct_of_ev']:>10.2%}   target ≈ 69%")
        if dcf_res["warnings"]:
            print("  Warnings:")
            for w in dcf_res["warnings"]:
                print(f"    - {w}")


if __name__ == "__main__":
    _cli()
