"""
sanity_checks.py — xlsx output に対する 16 hard + 10 soft + 12 design check 自動監査

Source of truth:
  - references/06_three_statement.md §12 (Sanity Checks H1-H15 + S1-S10)
  - references/01b_integrity_and_anti_patterns.md §6 (IB 品質チェックリスト 70 項目)
  - references/_terminology.md §1, §3, §6 (IB Color / Sheet naming / SaaS metric canonical)
  - references/_self_review_protocol.md §8 (案件レベル必須 5 check)

Note:
  This module implements the *automation-oriented* H1-H15 / S1-S10 list specified
  in the Phase 6 task (a practical intersection of §12 + 01b §6). Roll-forwards
  in §12.1 that require all sub-schedules populated (PP&E AD, APIC, Lease ROU/Liab,
  Deferred Revenue, NOL, DTA) are deferred — they will be added when the
  corresponding 17-sheet schedules are fully built. H8/H10 use sampling strategies
  to keep runtime under a few seconds on typical 5y models.

  Phase 6 補強 (2026-05): added H16 (#REF!/error literal detection),
  D12 (number_format on inputs), data_only loader plumbing (parallel
  data_only=True wb attached to report._wb_data so future value-based checks
  can read cached numeric values), and H/S regression cases in --self-test
  (H1, H2, H16, S2 against synthetic minimal wbs). See _Phase6_補強.md.

  Phase 6 critical fix (2026-05): existing H/S checks (H1-H4, H6, H12-H15,
  S1-S10) now read calc-result values from `report._wb_data` (data_only=True
  parallel handle) when available, falling back to the formula wb otherwise.
  Without this, formula cells return None for value-based comparisons and
  these checks silently degrade to `info`. H7 false-positive bug fixed
  (cross-sheet refs `=Sheet!D5` no longer trip self-ref detection at D5).
  See _Phase6_補強.md.

Public API:
  - run_all_checks(wb_path, *, business_model, stage, skip_soft, skip_design,
                   wb_data) -> CheckReport
  - to_sanity_sheet(wb, report) -> None  (writes results into 12_SanityChecks)

CLI:
  python3 sanity_checks.py PATH.xlsx [--business-model saas] [--stage series_a]
                                     [--update-sheet]
  exit 0 = pass (no hard_fail), 1 = at least one hard_fail.

License: internal (Act / startup-financial-modeling skill)
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from typing import Iterable, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import ib_format as ibf

# ---------------------------------------------------------------------------
# Severity / dataclasses
# ---------------------------------------------------------------------------

Severity = Literal["hard_fail", "soft_warn", "info", "pass"]

# tolerance for monetary equality (¥1 / $1 in the model's reporting unit)
TOLERANCE_MONEY: float = 1.0
TOLERANCE_PCT: float = 1e-4  # 0.01%


@dataclass(frozen=True)
class CheckResult:
    """Single result row, one per (check, period) tuple if periodised."""

    check_id: str            # H1, S2, etc.
    severity: Severity
    title: str
    sheet: str | None = None
    cell: str | None = None
    expected: float | str | None = None
    actual: float | str | None = None
    diff: float | None = None
    rationale: str = ""
    detail: str = ""


@dataclass
class CheckReport:
    """Aggregate report container."""

    results: list[CheckResult] = field(default_factory=list)
    business_model: str = "saas"
    stage: str = "series_a"

    # ------------------------------------------------------------------
    @property
    def hard_fail_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "hard_fail")

    @property
    def soft_warn_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "soft_warn")

    @property
    def info_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "info")

    @property
    def pass_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "pass")

    @property
    def is_pass(self) -> bool:
        """True iff zero hard_fail results."""
        return self.hard_fail_count == 0

    # ------------------------------------------------------------------
    def add(self, r: CheckResult) -> None:
        self.results.append(r)

    # ------------------------------------------------------------------
    def to_markdown(self) -> str:
        """Render as a markdown report (header + per-severity grouped tables)."""
        out: list[str] = []
        status = "PASS" if self.is_pass else "FAIL"
        out.append(f"# Sanity Check Report — **{status}**")
        out.append("")
        out.append(f"- business_model: `{self.business_model}`  stage: `{self.stage}`")
        out.append(
            f"- hard_fail: **{self.hard_fail_count}**  "
            f"soft_warn: {self.soft_warn_count}  "
            f"pass: {self.pass_count}  info: {self.info_count}"
        )
        out.append("")

        for sev_label, sev_key in (
            ("Hard fails (must be 0)", "hard_fail"),
            ("Soft warnings", "soft_warn"),
            ("Pass", "pass"),
            ("Info / skipped", "info"),
        ):
            rows = [r for r in self.results if r.severity == sev_key]
            if not rows:
                continue
            out.append(f"## {sev_label}")
            out.append("")
            out.append("| ID | Title | Sheet | Cell | Expected | Actual | Diff | Detail |")
            out.append("|---|---|---|---|---|---|---|---|")
            for r in rows:
                out.append(
                    f"| {r.check_id} | {r.title} | "
                    f"{r.sheet or ''} | {r.cell or ''} | "
                    f"{_fmt_val(r.expected)} | {_fmt_val(r.actual)} | "
                    f"{_fmt_val(r.diff)} | {r.detail} |"
                )
            out.append("")
        return "\n".join(out)


def _fmt_val(v: float | str | None) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        if abs(v) >= 1.0:
            return f"{v:,.2f}"
        return f"{v:.4f}"
    return str(v)


# ---------------------------------------------------------------------------
# Workbook helpers (defensive — skip gracefully on missing sheets)
# ---------------------------------------------------------------------------


def _get_sheet(wb: Workbook, name: str) -> Worksheet | None:
    """Return sheet by name or None (no exception)."""
    return wb[name] if name in wb.sheetnames else None


def _resolve_value_wb(report: "CheckReport", fallback_wb: Workbook) -> Workbook:
    """Return a Workbook handle suitable for reading numeric *values*.

    When `report._wb_data` exists (data_only=True parallel load), prefer it —
    it returns Excel's last-saved calculated values for formula cells. When
    absent, fall back to the formula `wb`; in that mode formula cells will
    read as their string formula expression (e.g. `'=D5+1'`) and `_cell_numeric`
    will return None, so the caller's `info` branch fires gracefully.

    Layout / structure (sheet names, label rows, period columns) is identical
    between the two handles, so `_find_row_by_label` and `_detect_period_columns`
    can run on either without behaving differently.
    """
    vwb = getattr(report, "_wb_data", None)
    return vwb if vwb is not None else fallback_wb


def _find_row_by_label(
    ws: Worksheet,
    label_substr: str,
    *,
    label_col: str = "B",
    max_rows: int = 200,
) -> int | None:
    """Scan a label column for a case-insensitive substring match. Return row index or None."""
    target = label_substr.lower().strip()
    for row in range(1, max_rows + 1):
        v = ws[f"{label_col}{row}"].value
        if isinstance(v, str) and target in v.lower():
            return row
    return None


def _detect_period_columns(ws: Worksheet, *, header_row: int = 4, max_cols: int = 60) -> list[int]:
    """Return 1-indexed column numbers that have non-empty period headers.

    Convention: period columns start at D (col 4) by setup_sheet_layout default.
    Scan header_row; if first attempt is empty, also try rows 3 and 5.
    """
    for hr in (header_row, 3, 5, 2):
        cols: list[int] = []
        for c in range(4, max_cols + 1):
            v = ws.cell(row=hr, column=c).value
            if v is not None and str(v).strip() != "":
                cols.append(c)
        if cols:
            return cols
    return []


def _get_cell_rgb(cell: Cell) -> str | None:
    """Return RGB hex (uppercase, 6 chars) or None when unknown / theme-indexed."""
    try:
        color = cell.font.color
    except AttributeError:
        return None
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return None
    if not isinstance(rgb, str):
        # could be openpyxl Value object — coerce
        rgb = str(rgb)
    if len(rgb) == 8:
        # leading alpha (e.g. FF0000FF) — strip
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
        return None
    return rgb.upper()


def _cell_numeric(cell_value) -> float | None:
    """Coerce numeric value, returning None for non-numeric / blank."""
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
        return float(cell_value)
    if isinstance(cell_value, str):
        try:
            return float(cell_value.replace(",", ""))
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Applicability matrix (_stress_framework §4.1)
# ---------------------------------------------------------------------------

# Soft checks gated by business model / stage / financial profile.
# True  = applicable, False = skip with info reason.
def _is_applicable(check_id: str, *, business_model: str, stage: str) -> tuple[bool, str]:
    bm = business_model.lower()
    st = stage.lower()
    pre_revenue = st in ("pre_seed", "preseed", "idea")
    pre_arr = st in ("pre_seed", "preseed", "seed")  # ARR < $1M typical
    is_saas = bm in ("saas", "vertical_saas", "horizontal_saas", "plg_saas")
    post_ipo = st in ("post_ipo", "ipo")

    if check_id == "S1":  # Magic Number
        if not is_saas:
            return False, "Magic Number applies only to SaaS"
        if pre_revenue:
            return False, "Pre-revenue: Magic Number undefined"
    if check_id == "S2":  # Burn Multiple
        if pre_arr:
            return False, "ARR < $1M: Burn Multiple undefined (_terminology §6.1)"
    if check_id == "S3":  # Rule of 40
        if pre_revenue:
            return False, "Pre-revenue: Rule of 40 undefined"
    if check_id == "S4":  # NRR
        if not is_saas:
            return False, "NRR applies only to subscription / SaaS"
        if pre_revenue:
            return False, "Pre-revenue: NRR undefined"
    if check_id == "S5":  # LTV/CAC
        if pre_revenue:
            return False, "Pre-revenue: LTV/CAC undefined"
    if check_id == "S6":  # CAC Payback
        if pre_revenue:
            return False, "Pre-revenue: CAC Payback undefined"
    if check_id == "S10":  # Founder dilution Series A anchor
        if post_ipo:
            return False, "Post-IPO: founder dilution anchor not applicable"
    return True, ""


# ===========================================================================
# Hard Checks (H1-H15 + H16 = H_ERROR_CELLS)
# ===========================================================================

# Excel error literals — appear in cell.value (string form) when:
#   - openpyxl reads a saved-and-recalculated wb with data_only=True
#   - a formula contains a literal #REF! after delete operations (formula text itself)
#   - hard inputs were typed as #N/A (sometimes intentional placeholders, but flagged)
_EXCEL_ERROR_LITERALS = (
    "#N/A", "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!",
)


def _check_h16_error_cells(wb: Workbook, report: CheckReport) -> None:
    """H16 — detect Excel error literals (#REF!, #DIV/0!, #N/A, etc.) in any cell.

    These break downstream calculations and are an immediate red-flag for any
    IB-quality model. Detection works against both formula text (e.g. a stale
    `=#REF!*D5` after a column delete) and against cached values (when the
    workbook was loaded with data_only=True or saved by Excel).

    Severity: any error literal anywhere → hard_fail.
    Sample cap: 200 cells per sheet to bound runtime on large workbooks.
    """
    errors: list[str] = []
    for ws in wb.worksheets:
        # Even 12_SanityChecks shouldn't legitimately contain error literals.
        max_row = min(ws.max_row or 0, 250)
        max_col = min(ws.max_column or 0, 30)
        if max_row < 1 or max_col < 1:
            continue
        sampled = 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            if sampled >= 200:
                break
            for cell in row:
                if sampled >= 200:
                    break
                v = cell.value
                if v is None:
                    continue
                sampled += 1
                if not isinstance(v, str):
                    continue
                # check whole-cell match OR substring inside formula text
                stripped = v.strip()
                hit = None
                for err in _EXCEL_ERROR_LITERALS:
                    if stripped == err or err in v:
                        hit = err
                        break
                if hit is not None:
                    errors.append(f"{ws.title}!{cell.coordinate}={stripped[:25]}")
                    if len(errors) >= 20:
                        break
            if len(errors) >= 20:
                break
        if len(errors) >= 20:
            break

    if not errors:
        report.add(CheckResult(
            "H16", "pass", "No error literals (#REF!/#N/A/#DIV/0! etc.)",
            detail="No Excel error literals detected in any sheet",
        ))
    else:
        report.add(CheckResult(
            "H16", "hard_fail", "No error literals (#REF!/#N/A/#DIV/0! etc.)",
            actual=len(errors), expected=0,
            rationale="Fix the source formula or wrap with IFERROR(...). "
                      "Error literals propagate downstream and corrupt Σ / totals.",
            detail="examples: " + "; ".join(errors[:5]),
        ))


def _check_h1_bs_balance(wb: Workbook, report: CheckReport) -> None:
    """H1 — Total Assets = Total Liabilities + Equity (per period, ±¥1)."""
    vwb = _resolve_value_wb(report, wb)
    ws = _get_sheet(vwb, "05_BS")
    if ws is None:
        report.add(CheckResult("H1", "info", "BS Balance", detail="05_BS sheet missing"))
        return

    row_assets = _find_row_by_label(ws, "Total Assets")
    # Liability+Equity total row label varies. Try the canonical phrasing first
    # ("and"), then "+", then "L&E" / "L+E" abbreviations. Fall back to "Total
    # Liab" only as a last resort — it can mistakenly land on the
    # liability-only subtotal when a richer label is missing.
    row_liab_eq = _find_row_by_label(ws, "Total Liabilities and Equity") \
        or _find_row_by_label(ws, "Total Liabilities + Equity") \
        or _find_row_by_label(ws, "Total Liabilities & Equity") \
        or _find_row_by_label(ws, "Total L&E") \
        or _find_row_by_label(ws, "Total L+E") \
        or _find_row_by_label(ws, "Total Liab")
    if row_assets is None or row_liab_eq is None:
        report.add(CheckResult(
            "H1", "info", "BS Balance", sheet="05_BS",
            detail="Could not locate 'Total Assets' or 'Total Liabilities and Equity' rows",
        ))
        return

    period_cols = _detect_period_columns(ws)
    if not period_cols:
        report.add(CheckResult(
            "H1", "info", "BS Balance", sheet="05_BS",
            detail="No period columns detected",
        ))
        return

    any_fail = False
    verified = 0
    non_numeric = 0
    for c in period_cols:
        a = _cell_numeric(ws.cell(row=row_assets, column=c).value)
        le = _cell_numeric(ws.cell(row=row_liab_eq, column=c).value)
        col = get_column_letter(c)
        if a is None or le is None:
            non_numeric += 1
            continue
        verified += 1
        diff = a - le
        if abs(diff) > TOLERANCE_MONEY:
            any_fail = True
            report.add(CheckResult(
                "H1", "hard_fail", "BS Balance", sheet="05_BS",
                cell=f"{col}{row_assets}", expected=le, actual=a, diff=diff,
                rationale="Total Assets must equal Total L&E (±¥1)",
            ))
    if any_fail:
        return
    if verified == 0:
        # Every period was non-numeric — formula cells with no cached value.
        report.add(CheckResult(
            "H1", "info", "BS Balance", sheet="05_BS",
            detail=f"All {non_numeric} period(s) non-numeric — formula cells with no cached values "
                   "(re-save xlsx in Excel/LibreOffice to populate cached results)",
        ))
        return
    report.add(CheckResult(
        "H1", "pass", "BS Balance", sheet="05_BS",
        detail=f"BS balanced across {verified} period(s)"
               + (f" ({non_numeric} non-numeric skipped)" if non_numeric else ""),
    ))


def _check_h2_cash_tie(wb: Workbook, report: CheckReport) -> None:
    """H2 — CFS ending cash = BS ending cash (per period)."""
    vwb = _resolve_value_wb(report, wb)
    bs = _get_sheet(vwb, "05_BS")
    cfs = _get_sheet(vwb, "06_CFS")
    if bs is None or cfs is None:
        report.add(CheckResult("H2", "info", "Cash Tie", detail="05_BS or 06_CFS missing"))
        return
    bs_cash_row = _find_row_by_label(bs, "Cash and Cash Equivalents") \
        or _find_row_by_label(bs, "Cash and equivalents") \
        or _find_row_by_label(bs, "Cash")
    cfs_end_row = _find_row_by_label(cfs, "Ending Cash") \
        or _find_row_by_label(cfs, "Cash, end of period") \
        or _find_row_by_label(cfs, "Cash at end")
    if bs_cash_row is None or cfs_end_row is None:
        report.add(CheckResult(
            "H2", "info", "Cash Tie",
            detail="Could not locate Cash row on BS or Ending Cash on CFS",
        ))
        return

    period_cols = _detect_period_columns(bs)
    if not period_cols:
        report.add(CheckResult("H2", "info", "Cash Tie", detail="No period columns"))
        return

    any_fail = False
    for c in period_cols:
        bsv = _cell_numeric(bs.cell(row=bs_cash_row, column=c).value)
        cfv = _cell_numeric(cfs.cell(row=cfs_end_row, column=c).value)
        col = get_column_letter(c)
        if bsv is None or cfv is None:
            continue
        diff = bsv - cfv
        if abs(diff) > TOLERANCE_MONEY:
            any_fail = True
            report.add(CheckResult(
                "H2", "hard_fail", "Cash Tie",
                sheet="05_BS / 06_CFS", cell=f"{col}{bs_cash_row}",
                expected=cfv, actual=bsv, diff=diff,
                rationale="BS cash must equal CFS ending cash",
            ))
    if not any_fail:
        report.add(CheckResult(
            "H2", "pass", "Cash Tie",
            detail=f"Cash tied across {len(period_cols)} period(s)",
        ))


def _check_h3_re_roll(wb: Workbook, report: CheckReport) -> None:
    """H3 — RE(t) = RE(t-1) + NI(t) − Dividends(t)."""
    vwb = _resolve_value_wb(report, wb)
    bs = _get_sheet(vwb, "05_BS")
    is_ws = _get_sheet(vwb, "04_IS")
    if bs is None or is_ws is None:
        report.add(CheckResult("H3", "info", "RE Roll", detail="04_IS or 05_BS missing"))
        return
    re_row = _find_row_by_label(bs, "Retained Earnings")
    ni_row = _find_row_by_label(is_ws, "Net Income")
    div_row = _find_row_by_label(bs, "Dividends") or _find_row_by_label(is_ws, "Dividends")
    if re_row is None or ni_row is None:
        report.add(CheckResult(
            "H3", "info", "RE Roll",
            detail="Could not locate Retained Earnings or Net Income row",
        ))
        return

    period_cols = _detect_period_columns(bs)
    if len(period_cols) < 2:
        report.add(CheckResult(
            "H3", "info", "RE Roll",
            detail="Need ≥2 period columns to verify roll-forward",
        ))
        return

    any_fail = False
    for i in range(1, len(period_cols)):
        c_prev = period_cols[i - 1]
        c_curr = period_cols[i]
        re_prev = _cell_numeric(bs.cell(row=re_row, column=c_prev).value)
        re_curr = _cell_numeric(bs.cell(row=re_row, column=c_curr).value)
        ni = _cell_numeric(is_ws.cell(row=ni_row, column=c_curr).value)
        div = 0.0
        if div_row is not None:
            div = _cell_numeric(bs.cell(row=div_row, column=c_curr).value) or \
                  _cell_numeric(is_ws.cell(row=div_row, column=c_curr).value) or 0.0
        if None in (re_prev, re_curr, ni):
            continue
        expected = re_prev + ni - div  # type: ignore[operator]
        diff = re_curr - expected      # type: ignore[operator]
        col = get_column_letter(c_curr)
        if abs(diff) > TOLERANCE_MONEY:
            any_fail = True
            report.add(CheckResult(
                "H3", "hard_fail", "RE Roll", sheet="05_BS",
                cell=f"{col}{re_row}", expected=expected, actual=re_curr, diff=diff,
                rationale="RE(t) = RE(t-1) + NI(t) − Div(t)",
            ))
    if not any_fail:
        report.add(CheckResult(
            "H3", "pass", "RE Roll",
            detail=f"RE roll consistent across {len(period_cols) - 1} transitions",
        ))


def _check_h4_h5_sum_check(wb: Workbook, report: CheckReport) -> None:
    """H4/H5 — sum check (rows / columns).

    Heuristic: scan rows whose label starts with 'Total ' on each financial sheet.
    For each such row, compare the period-cell value to the sum of the immediate
    block of contributing rows above it (until previous Total / blank label).
    Approximate but catches the common errors (broken Σ).

    Caveat (Phase 6 critical fix): on sheets with **nested totals** (e.g. a BS
    where TOTAL ASSETS aggregates `Total Current Assets` + `Net PP&E` rather
    than the raw line items), the per-block walk would either double-count
    (if it crosses sub-totals) or under-sum (if it stops before them). When
    the contributor walk would land within a "nested total" region, we
    downgrade the result to `soft_warn` and tag the limitation, so that the
    refactor's value-reading step does not flood the report with heuristic
    false-positives. True row-sum bugs on flat (non-nested) blocks still
    flag at hard_fail.
    """
    vwb = _resolve_value_wb(report, wb)
    for sheet_name in ("04_IS", "05_BS", "06_CFS"):
        ws = _get_sheet(vwb, sheet_name)
        if ws is None:
            continue
        period_cols = _detect_period_columns(ws)
        if not period_cols:
            continue
        total_rows = _scan_total_rows(ws)
        if not total_rows:
            continue

        h4_fails = 0
        h4_soft = 0
        verified_blocks = 0
        for tr in total_rows:
            block_start = _block_start_for_total(ws, tr)
            # Detect nested-totals: there's a "grand total" pattern when any
            # *prior* "Total *" row exists *anywhere above* the current one
            # (i.e. the current row aggregates sub-totals + remaining items
            # that span past block_start). Walking back via "blanks ≥ 2"
            # may stop short of the prior total or land between it and the
            # current total, in either case the simple sum-of-rows-between
            # block_start and tr is unreliable. Detect the case and degrade.
            nested_total_in_block = any(0 < ot < tr for ot in total_rows)
            block_severity: Severity = "soft_warn" if nested_total_in_block else "hard_fail"
            block_rationale = (
                "Total row should equal sum of contributing rows above"
                if not nested_total_in_block
                else "Heuristic limitation: nested totals (sub-total feeds grand-total) "
                     "cannot be reliably checked by row-walk; downgraded to soft_warn"
            )
            for c in period_cols:
                tot = _cell_numeric(ws.cell(row=tr, column=c).value)
                if tot is None:
                    continue
                contributors = []
                for r in range(block_start, tr):
                    v = _cell_numeric(ws.cell(row=r, column=c).value)
                    if v is not None:
                        contributors.append(v)
                if not contributors:
                    continue
                verified_blocks += 1
                expected = sum(contributors)
                diff = tot - expected
                if abs(diff) > TOLERANCE_MONEY:
                    if block_severity == "hard_fail":
                        h4_fails += 1
                    else:
                        h4_soft += 1
                    report.add(CheckResult(
                        "H4", block_severity, "Row Sum (Σ)", sheet=sheet_name,
                        cell=f"{get_column_letter(c)}{tr}",
                        expected=expected, actual=tot, diff=diff,
                        rationale=block_rationale,
                    ))

        # H5 column sum: only meaningful if we have a Year-Total column appended.
        # Skip unless a 'Total' / 'FY Total' header row exists.

        if verified_blocks == 0:
            report.add(CheckResult(
                "H4", "info", "Row Sum (Σ)", sheet=sheet_name,
                detail=f"No numeric content verified across {len(total_rows)} total row(s)",
            ))
        elif h4_fails == 0 and h4_soft == 0:
            report.add(CheckResult(
                "H4", "pass", "Row Sum (Σ)", sheet=sheet_name,
                detail=f"Verified {len(total_rows)} total row(s) × {len(period_cols)} period(s)",
            ))

    # H5 — column sum: deferred (requires a Year-Total append column convention).
    report.add(CheckResult(
        "H5", "info", "Column Sum (Σ)",
        detail="Column-Σ check requires Year-Total column convention; deferred",
    ))


def _scan_total_rows(ws: Worksheet, *, label_col: str = "B", max_rows: int = 200) -> list[int]:
    rows: list[int] = []
    for r in range(1, max_rows + 1):
        v = ws[f"{label_col}{r}"].value
        if isinstance(v, str) and v.strip().lower().startswith("total"):
            rows.append(r)
    return rows


def _block_start_for_total(ws: Worksheet, total_row: int, *, label_col: str = "B") -> int:
    """Walk upward from total_row to find start of the block (previous Total / 2 blanks)."""
    blanks = 0
    for r in range(total_row - 1, max(0, total_row - 50), -1):
        v = ws[f"{label_col}{r}"].value
        if isinstance(v, str):
            s = v.strip().lower()
            if s.startswith("total"):
                return r + 1
            if s == "":
                blanks += 1
                if blanks >= 2:
                    return r + 1
        elif v is None:
            blanks += 1
            if blanks >= 2:
                return r + 1
    return max(1, total_row - 20)


def _check_h6_sign_convention(wb: Workbook, report: CheckReport) -> None:
    """H6 — Cost / OpEx rows display positive on IS (negative would suggest mixed sign)."""
    vwb = _resolve_value_wb(report, wb)
    is_ws = _get_sheet(vwb, "04_IS")
    if is_ws is None:
        report.add(CheckResult("H6", "info", "Sign Convention", detail="04_IS missing"))
        return
    period_cols = _detect_period_columns(is_ws)
    if not period_cols:
        report.add(CheckResult("H6", "info", "Sign Convention", detail="No period columns"))
        return
    target_labels = ("COGS", "Operating Expenses", "OpEx", "Cost of Revenue")
    any_fail = False
    rows_checked = 0
    for label in target_labels:
        rr = _find_row_by_label(is_ws, label)
        if rr is None:
            continue
        rows_checked += 1
        for c in period_cols:
            v = _cell_numeric(is_ws.cell(row=rr, column=c).value)
            if v is None:
                continue
            if v < -TOLERANCE_MONEY:  # explicit negative on IS — flag inconsistency
                any_fail = True
                report.add(CheckResult(
                    "H6", "hard_fail", "Sign Convention", sheet="04_IS",
                    cell=f"{get_column_letter(c)}{rr}", actual=v,
                    rationale=f"'{label}' shown as negative on IS — sign convention recommends positive display",
                ))
                break  # one fail per label is enough
    if rows_checked == 0:
        report.add(CheckResult(
            "H6", "info", "Sign Convention",
            detail="No COGS/OpEx label rows located on IS",
        ))
    elif not any_fail:
        report.add(CheckResult(
            "H6", "pass", "Sign Convention",
            detail=f"Cost/OpEx sign consistent across {rows_checked} label row(s)",
        ))


def _check_h7_circular_refs(wb: Workbook, report: CheckReport) -> None:
    """H7 — circular references.

    openpyxl does not evaluate formulas, so true circular detection requires
    Excel's calc engine. We perform a heuristic self-reference scan only:
    flag a cell whose formula references **its own coordinate on the same
    sheet** (after stripping cross-sheet prefixes), since that is the only
    pattern openpyxl can detect statically without producing systematic
    false positives.

    False positives previously hit by the naïve `coord in formula` test:
      - `=D5*D5/30` written in cell E5 — D5 appearing twice does not imply
        E5 references itself.
      - `='02_Drivers'!D5+1` written in cell D5 — references a *different*
        sheet's D5, not this sheet's D5. We must strip `'Sheet'!` and
        `Sheet!` prefixes before matching.

    This still uses formula text (not cached values), so we deliberately
    keep `wb` (formula handle), not `vwb`.
    """
    # Replace cross-sheet references with a neutral placeholder, so the
    # remaining text contains only same-sheet cell refs. Then a regex
    # match against the cell's own coordinate is safe:
    #   `=D5+1` written at D5            → still matches → flag (true cycle)
    #   `=D5*D5/30` written at E5        → no match for E5 → not flagged
    #   `='02_Drivers'!D5+1` written D5  → becomes `=__XSHEET__+1` → no match
    quoted_sheet_ref = re.compile(r"'[^']*'!\$?[A-Za-z]+\$?\d+(?::\$?[A-Za-z]+\$?\d+)?")
    bare_sheet_ref = re.compile(
        r"\b[A-Za-z_][A-Za-z0-9_\.]*!\$?[A-Za-z]+\$?\d+(?::\$?[A-Za-z]+\$?\d+)?"
    )

    self_ref_count = 0
    examples: list[tuple[str, str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                # Replace any `Sheet!ref` (quoted or bare) with a placeholder
                # that contains no letters or digits the self-ref pattern
                # could mistakenly match.
                expr = quoted_sheet_ref.sub("__XSHEET__", v)
                expr = bare_sheet_ref.sub("__XSHEET__", expr)
                col_letter = cell.column_letter
                row_num = cell.row
                # Match `D5`, `$D5`, `D$5`, `$D$5` as a whole token (case
                # insensitive — Excel canonicalises uppercase but we defend
                # against unusual writers).
                pat = (
                    rf"(?<![A-Za-z_\$\d])\$?{col_letter}\$?{row_num}"
                    rf"(?![A-Za-z_\d])"
                )
                if re.search(pat, expr, flags=re.IGNORECASE):
                    self_ref_count += 1
                    if len(examples) < 3:
                        examples.append((ws.title, cell.coordinate, v[:40]))
    if self_ref_count == 0:
        report.add(CheckResult(
            "H7", "info", "Circular Refs",
            detail="No direct self-refs found (openpyxl cannot evaluate true cycles — Excel runtime required)",
        ))
    else:
        report.add(CheckResult(
            "H7", "hard_fail", "Circular Refs",
            actual=self_ref_count,
            rationale="Direct self-referencing formula detected (likely unintentional)",
            detail="examples: " + "; ".join(f"{s}!{c}={f}" for s, c, f in examples),
        ))


def _check_h8_hardcode_in_formula(wb: Workbook, report: CheckReport) -> None:
    """H8 — hardcoded numeric literals inside formulas."""
    # whitelist common allowed literals
    whitelist = {0, 1, 2, 4, 12, 100, 360, 365, 1000, 1000000}
    # match a number token that is NOT preceded by a letter/$ and NOT immediately after "
    pat = re.compile(r"(?<![A-Za-z\$\d])(\d+(?:\.\d+)?)(?![A-Za-z\d])")

    violations: list[tuple[str, str, str]] = []
    sampled = 0
    sample_cap_per_sheet = 200
    for ws in wb.worksheets:
        if ws.title in ("00_Cover", "12_SanityChecks", "99_Glossary"):
            continue
        seen = 0
        for row in ws.iter_rows():
            if seen >= sample_cap_per_sheet:
                break
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                seen += 1
                sampled += 1
                # strip strings and sheet refs to avoid false matches
                expr = re.sub(r"'[^']*'!", "", v)
                expr = re.sub(r"\"[^\"]*\"", "", expr)
                for m in pat.finditer(expr):
                    raw = m.group(1)
                    try:
                        num = float(raw)
                    except ValueError:
                        continue
                    if num.is_integer() and int(num) in whitelist:
                        continue
                    if 0 < num < 1.0:  # 0.5, 0.78 — likely magic ratio → flag
                        pass
                    violations.append((ws.title, cell.coordinate, v))
                    break

    if not violations:
        report.add(CheckResult(
            "H8", "pass", "No Hardcode in Formula",
            detail=f"Sampled {sampled} formulas; no disallowed numeric literals",
        ))
    else:
        report.add(CheckResult(
            "H8", "hard_fail", "No Hardcode in Formula",
            actual=len(violations),
            rationale="Move literals to 01_Assumptions and reference them",
            detail="examples: " + "; ".join(f"{s}!{c}={f}" for s, c, f in violations[:3]),
        ))


def _check_h9_sheet_naming(wb: Workbook, report: CheckReport) -> None:
    """H9 — sheet order matches CANONICAL_SHEET_ORDER."""
    violations = ibf.validate_sheet_naming(wb.sheetnames)
    if not violations:
        report.add(CheckResult(
            "H9", "pass", "Sheet Naming",
            detail=f"All {len(wb.sheetnames)} sheets in canonical order",
        ))
    else:
        report.add(CheckResult(
            "H9", "hard_fail", "Sheet Naming",
            rationale="Sheet order must match CANONICAL_SHEET_ORDER (_terminology §3)",
            detail=" | ".join(violations[:5]),
        ))


def _check_h10_ib_color(wb: Workbook, report: CheckReport) -> None:
    """H10 — sample 50 cells across sheets and verify IB color compliance."""
    # stratified sample: half formulas, half hard inputs
    formula_samples: list[tuple[Worksheet, Cell]] = []
    input_samples: list[tuple[Worksheet, Cell]] = []
    cap = 25
    for ws in wb.worksheets:
        if ws.title in ("00_Cover", "12_SanityChecks", "99_Glossary"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    if len(formula_samples) < cap:
                        formula_samples.append((ws, cell))
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    if len(input_samples) < cap:
                        input_samples.append((ws, cell))
            if len(formula_samples) >= cap and len(input_samples) >= cap:
                break
        if len(formula_samples) >= cap and len(input_samples) >= cap:
            break

    violations: list[str] = []
    unknown = 0

    for ws, cell in formula_samples:
        rgb = _get_cell_rgb(cell)
        if rgb is None:
            unknown += 1
            continue
        v = cell.value or ""
        # cross-sheet reference
        if isinstance(v, str) and "!" in v:
            if rgb not in (ibf.IB_LINK_INTRA, ibf.IB_LINK_EXTERNAL, ibf.IB_FORMULA):
                violations.append(f"{ws.title}!{cell.coordinate} cross-ref color={rgb}")
        else:
            if rgb != ibf.IB_FORMULA:
                violations.append(f"{ws.title}!{cell.coordinate} formula color={rgb}")

    for ws, cell in input_samples:
        rgb = _get_cell_rgb(cell)
        if rgb is None:
            unknown += 1
            continue
        if rgb != ibf.IB_HARD_INPUT:
            violations.append(f"{ws.title}!{cell.coordinate} input color={rgb}")

    total_sampled = len(formula_samples) + len(input_samples)
    if not violations:
        report.add(CheckResult(
            "H10", "pass", "IB Color Compliance",
            detail=f"Sampled {total_sampled} cells (unknown={unknown}); colors compliant",
        ))
    else:
        report.add(CheckResult(
            "H10", "hard_fail", "IB Color Compliance",
            actual=len(violations),
            rationale="hard_input=#0000FF, formula=#000000, cross-sheet=#008000, external=#FF0000",
            detail=f"examples: {'; '.join(violations[:3])}",
        ))


def _check_h11_sheet_count(wb: Workbook, report: CheckReport) -> None:
    """H11 — exactly 14 canonical sheets (+ optional 99_Glossary). Phase 6 Stage A."""
    canonical = set(ibf.CANONICAL_SHEET_ORDER)
    optional = set(ibf.OPTIONAL_SHEETS)
    actual = set(wb.sheetnames)

    missing = canonical - actual
    extra = actual - canonical - optional
    if missing or extra:
        report.add(CheckResult(
            "H11", "hard_fail", "Sheet Count",
            expected=len(canonical), actual=len(actual & canonical),
            rationale="Workbook must have all 14 canonical sheets and no extras",
            detail=f"missing={sorted(missing)} extra={sorted(extra)}",
        ))
    else:
        report.add(CheckResult(
            "H11", "pass", "Sheet Count",
            detail=f"14 canonical sheets present (+optional: {sorted(actual & optional)})",
        ))


def _check_h12_fdso_sum(wb: Workbook, report: CheckReport) -> None:
    """H12 — Cap Table ownership % sums to 100.00% ±0.01%."""
    vwb = _resolve_value_wb(report, wb)
    ws = _get_sheet(vwb, "08_CapTable")
    if ws is None:
        report.add(CheckResult("H12", "info", "FDSO Sum", detail="08_CapTable missing"))
        return
    pct_row = _find_row_by_label(ws, "Total %") \
        or _find_row_by_label(ws, "Total Ownership") \
        or _find_row_by_label(ws, "Total")
    if pct_row is None:
        report.add(CheckResult(
            "H12", "info", "FDSO Sum", sheet="08_CapTable",
            detail="No Total % row located",
        ))
        return
    period_cols = _detect_period_columns(ws)
    if not period_cols:
        # check column D as default
        period_cols = [4]

    any_fail = False
    for c in period_cols:
        v = _cell_numeric(ws.cell(row=pct_row, column=c).value)
        if v is None:
            continue
        # accept either fractional (1.0) or percent-stored (100)
        target = 1.0 if abs(v - 1.0) < 0.5 else 100.0
        diff = v - target
        if abs(diff) > (TOLERANCE_PCT if target == 1.0 else 0.01):
            any_fail = True
            report.add(CheckResult(
                "H12", "hard_fail", "FDSO Sum", sheet="08_CapTable",
                cell=f"{get_column_letter(c)}{pct_row}",
                expected=target, actual=v, diff=diff,
                rationale="Cap table ownership must sum to 100.00% ±0.01%",
            ))
    if not any_fail:
        report.add(CheckResult(
            "H12", "pass", "FDSO Sum", sheet="08_CapTable",
            detail="Ownership totals tie to 100%",
        ))


def _check_h13_wc_roll(wb: Workbook, report: CheckReport) -> None:
    """H13 — BS AR/AP/Inventory ties to embedded Working Capital Schedule sub-section.

    Phase 6 Stage A: 08_WC was merged into 05_BS as a sub-section ("Working Capital
    Schedule"). The check now looks for the WC sub-section anchor on 05_BS and
    emits info if missing, pass if present. Detailed row-anchor tie-out is deferred
    to a Stage B re-implementation once the new 05_BS layout is finalized.
    """
    vwb = _resolve_value_wb(report, wb)
    bs = _get_sheet(vwb, "05_BS")
    if bs is None:
        report.add(CheckResult("H13", "info", "WC Roll", detail="05_BS missing"))
        return
    wc_anchor = _find_row_by_label(bs, "Working Capital Schedule") \
        or _find_row_by_label(bs, "Working Capital")
    if wc_anchor is None:
        report.add(CheckResult(
            "H13", "info", "WC Roll", sheet="05_BS",
            detail="No 'Working Capital Schedule' sub-section header found on 05_BS",
        ))
        return
    report.add(CheckResult(
        "H13", "pass", "WC Roll", sheet="05_BS",
        detail=f"Working Capital Schedule sub-section present at row {wc_anchor}",
    ))

def _check_h14_debt_roll(wb: Workbook, report: CheckReport) -> None:
    """H14 — BS Debt = ending balance from 07_Debt schedule."""
    vwb = _resolve_value_wb(report, wb)
    bs = _get_sheet(vwb, "05_BS")
    debt = _get_sheet(vwb, "07_Debt")
    if bs is None or debt is None:
        report.add(CheckResult("H14", "info", "Debt Roll", detail="05_BS or 07_Debt missing"))
        return
    bs_debt_row = _find_row_by_label(bs, "Total Debt") \
        or _find_row_by_label(bs, "Long-term Debt") \
        or _find_row_by_label(bs, "Debt")
    debt_end_row = _find_row_by_label(debt, "Ending Balance") \
        or _find_row_by_label(debt, "Ending Debt") \
        or _find_row_by_label(debt, "End of Period")
    if bs_debt_row is None or debt_end_row is None:
        report.add(CheckResult(
            "H14", "info", "Debt Roll",
            detail="Could not locate Debt row on BS or Ending Balance on 07_Debt",
        ))
        return

    period_cols = _detect_period_columns(bs)
    any_fail = False
    for c in period_cols:
        bv = _cell_numeric(bs.cell(row=bs_debt_row, column=c).value)
        dv = _cell_numeric(debt.cell(row=debt_end_row, column=c).value)
        if bv is None or dv is None:
            continue
        diff = bv - dv
        if abs(diff) > TOLERANCE_MONEY:
            any_fail = True
            report.add(CheckResult(
                "H14", "hard_fail", "Debt Roll", sheet="05_BS",
                cell=f"{get_column_letter(c)}{bs_debt_row}",
                expected=dv, actual=bv, diff=diff,
                rationale="BS Debt must equal 07_Debt ending balance",
            ))
    if not any_fail:
        report.add(CheckResult("H14", "pass", "Debt Roll", detail="BS debt ties to 07_Debt"))


def _check_h15_dcf_tie(wb: Workbook, report: CheckReport) -> None:
    """H15 — 09_DCF terminal-year FCF = IS / CFS Year-N FCF."""
    vwb = _resolve_value_wb(report, wb)
    dcf = _get_sheet(vwb, "09_DCF")
    cfs = _get_sheet(vwb, "06_CFS")
    if dcf is None or cfs is None:
        report.add(CheckResult("H15", "info", "DCF Tie", detail="09_DCF or 06_CFS missing"))
        return
    dcf_fcf_row = _find_row_by_label(dcf, "Free Cash Flow") \
        or _find_row_by_label(dcf, "FCF") \
        or _find_row_by_label(dcf, "Unlevered FCF")
    cfs_fcf_row = _find_row_by_label(cfs, "Free Cash Flow") \
        or _find_row_by_label(cfs, "FCF")
    if dcf_fcf_row is None or cfs_fcf_row is None:
        report.add(CheckResult(
            "H15", "info", "DCF Tie",
            detail="Could not locate FCF rows on both 09_DCF and 06_CFS",
        ))
        return

    dcf_cols = _detect_period_columns(dcf)
    cfs_cols = _detect_period_columns(cfs)
    if not dcf_cols or not cfs_cols:
        report.add(CheckResult("H15", "info", "DCF Tie", detail="No period columns"))
        return

    # Last forecast year on each sheet (terminal year)
    dcf_last = dcf_cols[-1]
    cfs_last = cfs_cols[-1]
    dv = _cell_numeric(dcf.cell(row=dcf_fcf_row, column=dcf_last).value)
    cv = _cell_numeric(cfs.cell(row=cfs_fcf_row, column=cfs_last).value)
    if dv is None or cv is None:
        report.add(CheckResult(
            "H15", "info", "DCF Tie",
            detail="Terminal year FCF is non-numeric on at least one sheet",
        ))
        return
    diff = dv - cv
    if abs(diff) > TOLERANCE_MONEY:
        report.add(CheckResult(
            "H15", "hard_fail", "DCF Tie", sheet="09_DCF",
            cell=f"{get_column_letter(dcf_last)}{dcf_fcf_row}",
            expected=cv, actual=dv, diff=diff,
            rationale="Terminal-year FCF on 09_DCF must equal Year-N FCF on 06_CFS",
        ))
    else:
        report.add(CheckResult("H15", "pass", "DCF Tie", detail="Terminal FCF ties to CFS"))


# ===========================================================================
# Soft Warns (S1-S10)
# ===========================================================================


def _read_kpi_value(wb: Workbook, label_substr: str) -> float | None:
    """Read a single numeric value from 11_KPI_Dashboard by label match (last period).

    `wb` here is whichever handle the caller resolved (data-only preferred); we
    do not re-resolve at this layer to keep the helper composable.
    """
    ws = _get_sheet(wb, "11_KPI_Dashboard")
    if ws is None:
        return None
    rr = _find_row_by_label(ws, label_substr)
    if rr is None:
        return None
    cols = _detect_period_columns(ws)
    if not cols:
        return None
    return _cell_numeric(ws.cell(row=rr, column=cols[-1]).value)


def _soft_check_threshold(
    report: CheckReport,
    *,
    check_id: str,
    title: str,
    business_model: str,
    stage: str,
    label: str,
    threshold: float,
    direction: Literal["below", "above"],
    rationale: str,
) -> None:
    ok, reason = _is_applicable(check_id, business_model=business_model, stage=stage)
    if not ok:
        report.add(CheckResult(check_id, "info", title, detail=reason))
        return
    val = _read_kpi_value(_soft_check_threshold._wb, label)  # type: ignore[attr-defined]
    if val is None:
        report.add(CheckResult(
            check_id, "info", title,
            detail=f"'{label}' not found on 11_KPI_Dashboard",
        ))
        return
    if direction == "below" and val < threshold:
        report.add(CheckResult(
            check_id, "soft_warn", title,
            actual=val, expected=threshold, diff=val - threshold,
            rationale=rationale,
        ))
    elif direction == "above" and val > threshold:
        report.add(CheckResult(
            check_id, "soft_warn", title,
            actual=val, expected=threshold, diff=val - threshold,
            rationale=rationale,
        ))
    else:
        report.add(CheckResult(check_id, "pass", title, actual=val))


def _check_soft_all(
    wb: Workbook, report: CheckReport, *, business_model: str, stage: str
) -> None:
    """S1-S10 in one block (each gated by applicability).

    Stashes the data-only workbook (when available) so threshold helpers can
    read calculated KPI values. Without that, every formula-based KPI cell
    on 11_KPI_Dashboard would return None and force every soft check to
    `info` (the silent-failure mode addressed by Phase 6 critical fix).
    """
    vwb = _resolve_value_wb(report, wb)
    # Stash workbook for the threshold helper (data-only preferred)
    _soft_check_threshold._wb = vwb  # type: ignore[attr-defined]

    _soft_check_threshold(
        report, check_id="S1", title="Magic Number ≥ 0.4",
        business_model=business_model, stage=stage,
        label="Magic Number", threshold=0.4, direction="below",
        rationale="< 0.4 = bottom-quartile S&M efficiency (_terminology §6.3)",
    )
    _soft_check_threshold(
        report, check_id="S2", title="Burn Multiple ≤ 2.0",
        business_model=business_model, stage=stage,
        label="Burn Multiple", threshold=2.0, direction="above",
        rationale="> 2.0x = inefficient burn (_terminology §6.1)",
    )
    _soft_check_threshold(
        report, check_id="S3", title="Rule of 40 ≥ 20",
        business_model=business_model, stage=stage,
        label="Rule of 40", threshold=20.0, direction="below",
        rationale="< 20 = below acceptable growth+profit threshold",
    )
    _soft_check_threshold(
        report, check_id="S4", title="NRR ≥ 100%",
        business_model=business_model, stage=stage,
        label="NRR", threshold=1.0, direction="below",
        rationale="< 100% = net contraction (Mid-market threshold)",
    )
    _soft_check_threshold(
        report, check_id="S5", title="LTV/CAC ≥ 1",
        business_model=business_model, stage=stage,
        label="LTV/CAC", threshold=1.0, direction="below",
        rationale="< 1 = unit economics inverted",
    )
    _soft_check_threshold(
        report, check_id="S6", title="CAC Payback ≤ 24 months",
        business_model=business_model, stage=stage,
        label="CAC Payback", threshold=24.0, direction="above",
        rationale="> 24 mo = long payback, capital efficiency concern",
    )

    # S7 — gross margin trajectory: requires ≥2 periods
    _check_s7_gm_trajectory(wb, report)

    # S8 — TAM realism: 5y penetration > 5%
    _check_s8_tam_realism(wb, report)

    # S9 — DCF / Terminal value ratio
    _check_s9_tv_ratio(wb, report)

    # S10 — founder dilution after Series A
    _check_s10_founder_dilution(wb, report, stage=stage)


def _check_s7_gm_trajectory(wb: Workbook, report: CheckReport) -> None:
    vwb = _resolve_value_wb(report, wb)
    is_ws = _get_sheet(vwb, "04_IS")
    if is_ws is None:
        report.add(CheckResult("S7", "info", "GM YoY Stability", detail="04_IS missing"))
        return
    gm_row = _find_row_by_label(is_ws, "Gross Margin")
    if gm_row is None:
        report.add(CheckResult("S7", "info", "GM YoY Stability", detail="No GM row found"))
        return
    cols = _detect_period_columns(is_ws)
    if len(cols) < 2:
        report.add(CheckResult("S7", "info", "GM YoY Stability", detail="Need ≥2 periods"))
        return
    flagged = 0
    for i in range(1, len(cols)):
        a = _cell_numeric(is_ws.cell(row=gm_row, column=cols[i - 1]).value)
        b = _cell_numeric(is_ws.cell(row=gm_row, column=cols[i]).value)
        if a is None or b is None:
            continue
        if abs(b - a) > 0.05:  # 500 bps (margins stored as fraction)
            flagged += 1
            report.add(CheckResult(
                "S7", "soft_warn", "GM YoY Stability", sheet="04_IS",
                cell=f"{get_column_letter(cols[i])}{gm_row}",
                expected=a, actual=b, diff=b - a,
                rationale="GM swung > 500bps YoY — verify cost mix or reclass",
            ))
    if flagged == 0:
        report.add(CheckResult("S7", "pass", "GM YoY Stability"))


def _check_s8_tam_realism(wb: Workbook, report: CheckReport) -> None:
    vwb = _resolve_value_wb(report, wb)
    ws = _get_sheet(vwb, "01_Assumptions") or _get_sheet(vwb, "11_KPI_Dashboard")
    if ws is None:
        report.add(CheckResult("S8", "info", "TAM Realism", detail="No assumptions/drivers sheet"))
        return
    pen_row = _find_row_by_label(ws, "Penetration") \
        or _find_row_by_label(ws, "Market Share")
    if pen_row is None:
        report.add(CheckResult("S8", "info", "TAM Realism", detail="No Penetration row"))
        return
    cols = _detect_period_columns(ws)
    if len(cols) < 5:
        report.add(CheckResult(
            "S8", "info", "TAM Realism",
            detail="Need ≥5 periods to assess Year-5 penetration",
        ))
        return
    val = _cell_numeric(ws.cell(row=pen_row, column=cols[4]).value)
    if val is None:
        report.add(CheckResult("S8", "info", "TAM Realism", detail="Y5 penetration non-numeric"))
        return
    threshold = 0.05  # 5%
    if val > threshold:
        report.add(CheckResult(
            "S8", "soft_warn", "TAM Realism", sheet=ws.title,
            cell=f"{get_column_letter(cols[4])}{pen_row}",
            actual=val, expected=threshold, diff=val - threshold,
            rationale="Y5 penetration > 5% may exceed realistic ceiling (`09 §7`)",
        ))
    else:
        report.add(CheckResult("S8", "pass", "TAM Realism", actual=val))


def _check_s9_tv_ratio(wb: Workbook, report: CheckReport) -> None:
    vwb = _resolve_value_wb(report, wb)
    ws = _get_sheet(vwb, "09_DCF")
    if ws is None:
        report.add(CheckResult("S9", "info", "TV/EV ratio", detail="09_DCF missing"))
        return
    tv_row = _find_row_by_label(ws, "Terminal Value") or _find_row_by_label(ws, "PV of TV")
    ev_row = _find_row_by_label(ws, "Enterprise Value") or _find_row_by_label(ws, "EV")
    if tv_row is None or ev_row is None:
        report.add(CheckResult("S9", "info", "TV/EV ratio", detail="TV or EV row missing"))
        return
    cols = _detect_period_columns(ws)
    target_col = cols[-1] if cols else 4
    tv = _cell_numeric(ws.cell(row=tv_row, column=target_col).value)
    ev = _cell_numeric(ws.cell(row=ev_row, column=target_col).value)
    if tv is None or ev is None or ev == 0:
        report.add(CheckResult("S9", "info", "TV/EV ratio", detail="TV or EV non-numeric"))
        return
    ratio = tv / ev
    if ratio > 0.75:
        report.add(CheckResult(
            "S9", "soft_warn", "TV/EV ratio", sheet="09_DCF",
            actual=ratio, expected=0.75, diff=ratio - 0.75,
            rationale="TV/EV > 75% — Damodaran rule (over-reliance on terminal)",
        ))
    else:
        report.add(CheckResult("S9", "pass", "TV/EV ratio", actual=ratio))


def _check_s10_founder_dilution(wb: Workbook, report: CheckReport, *, stage: str) -> None:
    ok, reason = _is_applicable("S10", business_model="any", stage=stage)
    if not ok:
        report.add(CheckResult("S10", "info", "Founder Dilution", detail=reason))
        return
    vwb = _resolve_value_wb(report, wb)
    ws = _get_sheet(vwb, "08_CapTable")
    if ws is None:
        report.add(CheckResult("S10", "info", "Founder Dilution", detail="08_CapTable missing"))
        return
    rr = _find_row_by_label(ws, "Founder")
    if rr is None:
        report.add(CheckResult("S10", "info", "Founder Dilution", detail="No 'Founder' row"))
        return
    cols = _detect_period_columns(ws)
    target = cols[-1] if cols else 4
    val = _cell_numeric(ws.cell(row=rr, column=target).value)
    if val is None:
        report.add(CheckResult("S10", "info", "Founder Dilution", detail="Non-numeric"))
        return
    # accept stored as fraction or %
    pct = val if val <= 1.0 else val / 100.0
    if pct < 0.30:
        report.add(CheckResult(
            "S10", "soft_warn", "Founder Dilution", sheet="08_CapTable",
            actual=pct, expected=0.30, diff=pct - 0.30,
            rationale="Founder ownership < 30% post-Series A is below typical median",
        ))
    else:
        report.add(CheckResult("S10", "pass", "Founder Dilution", actual=pct))


# ===========================================================================
# Design Checks (D1-D10) — visual / structural design quality
# ---------------------------------------------------------------------------
# Source of truth:
#   - references/00_design_guidelines.md (IB visual design canonical)
#   - ib_format.py (constants: COL_MARGIN_WIDTH, COL_PERIOD_WIDTH, BG_HEADER_BAND,
#     IB_CHART_COLORS_*, set_tab_color role mapping)
#
# Severity convention:
#   D1-D7, D9, D10 → soft_warn  (visual cleanup)
#   D8             → hard_fail  (chart broken / unreadable)
#
# Design-check sheets are excluded from sampling-style checks (D2, D5) because
# 12_SanityChecks itself uses wrap_text=True for long Detail columns by design.
# ===========================================================================

# Sheets that legitimately use design overrides (wrap_text, custom fills outside
# the standard data area, etc.). They are excluded from D2 / D5 / D6.
_DESIGN_EXEMPT_SHEETS: tuple[str, ...] = (
    "00_Cover",
    "12_SanityChecks",
    "99_Glossary",
    "13_IC_Memo",  # IC memo uses prose paragraphs with wrap_text legitimately
)


def _cells_in_ref(formula: str | None) -> int | None:
    """Parse an Excel range string ('Sheet'!$A$1:$B$5 or $A$1) → cell count.

    Returns None if the formula cannot be parsed. Used by D8 chart checks.
    """
    if not formula or not isinstance(formula, str):
        return None
    f = formula.strip()
    if "!" in f:
        f = f.split("!", 1)[1]
    m = re.fullmatch(r"\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?", f)
    if not m:
        return None
    c1, r1, c2, r2 = m.groups()
    from openpyxl.utils import column_index_from_string
    if c2 is None:
        return 1
    try:
        nc = column_index_from_string(c2) - column_index_from_string(c1) + 1
        nr = int(r2) - int(r1) + 1
        return max(0, nc * nr)
    except Exception:
        return None


def _check_d1_a_column_width(wb: Workbook, report: CheckReport) -> None:
    """D1 — A 列幅 = 20.0 ±0.5 (全シート、Phase 6 Fix #2).

    Source: ib_format.COL_MARGIN_WIDTH = 20.0
    """
    target = ibf.COL_MARGIN_WIDTH
    tolerance = 0.5
    bad: list[tuple[str, float | None]] = []
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        w = ws.column_dimensions["A"].width
        # openpyxl returns None when never explicitly set
        if w is None:
            bad.append((ws.title, None))
            continue
        if abs(float(w) - target) > tolerance:
            bad.append((ws.title, float(w)))
    if not bad:
        report.add(CheckResult(
            "D1", "pass", "A column width = 20.0",
            detail=f"Verified across {len(wb.worksheets)} sheet(s)",
        ))
        return
    examples = "; ".join(
        f"{name}={'unset' if w is None else f'{w:.1f}'}" for name, w in bad[:5]
    )
    report.add(CheckResult(
        "D1", "soft_warn", "A column width = 20.0",
        actual=len(bad), expected=0,
        rationale="Call ib_format.setup_sheet_layout(ws) to apply COL_MARGIN_WIDTH=20.0",
        detail=f"{len(bad)} sheet(s) off-target: {examples}",
    ))


def _check_d2_wrap_text_disabled(wb: Workbook, report: CheckReport) -> None:
    """D2 — wrap_text default disabled in data area (Phase 6 Fix #1).

    Sample B-G × rows 5-30 per sheet (excluding _DESIGN_EXEMPT_SHEETS).
    Flag if > 10% of inspected cells have wrap_text=True.
    """
    threshold_ratio = 0.10
    flagged_sheets: list[tuple[str, int, int]] = []
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        wrapped = 0
        inspected = 0
        for r in range(5, 31):
            for c in range(2, 8):  # B-G
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                inspected += 1
                if cell.alignment is not None and cell.alignment.wrap_text is True:
                    wrapped += 1
        if inspected >= 10 and (wrapped / inspected) > threshold_ratio:
            flagged_sheets.append((ws.title, wrapped, inspected))
    if not flagged_sheets:
        report.add(CheckResult(
            "D2", "pass", "wrap_text disabled in data area",
            detail="No sheet exceeds 10% wrap_text density",
        ))
        return
    examples = "; ".join(
        f"{name}: {w}/{n} wrapped" for name, w, n in flagged_sheets[:3]
    )
    report.add(CheckResult(
        "D2", "soft_warn", "wrap_text disabled in data area",
        actual=len(flagged_sheets), expected=0,
        rationale="Use apply_label/apply_comment with wrap_text=False (IB convention)",
        detail=f"{len(flagged_sheets)} sheet(s) over threshold: {examples}",
    ))


def _check_d3_period_column_uniform(wb: Workbook, report: CheckReport) -> None:
    """D3 — Period column widths uniform (Phase 6 Fix #4).

    Detect period columns via _detect_period_columns; verify all detected
    columns share the same width (±0.5).
    """
    target = ibf.COL_PERIOD_WIDTH
    tolerance = 0.5
    bad: list[tuple[str, list[tuple[str, float]]]] = []
    inspected = 0
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        cols = _detect_period_columns(ws)
        if len(cols) < 2:
            continue
        inspected += 1
        widths: list[tuple[str, float]] = []
        for c in cols:
            letter = get_column_letter(c)
            w = ws.column_dimensions[letter].width
            if w is None:
                continue
            widths.append((letter, float(w)))
        if not widths:
            continue
        # check spread
        vmin = min(w for _, w in widths)
        vmax = max(w for _, w in widths)
        if (vmax - vmin) > tolerance:
            bad.append((ws.title, widths))
            continue
        # check vs target
        if abs(widths[0][1] - target) > tolerance:
            bad.append((ws.title, widths))
    if not bad:
        report.add(CheckResult(
            "D3", "pass", "Period columns uniform width",
            detail=f"Verified {inspected} sheet(s) (target={target})",
        ))
        return
    examples = "; ".join(
        f"{name}: {','.join(f'{l}={w:.1f}' for l, w in widths[:4])}"
        for name, widths in bad[:3]
    )
    report.add(CheckResult(
        "D3", "soft_warn", "Period columns uniform width",
        actual=len(bad), expected=0,
        rationale=f"Use ib_format.set_uniform_column_width(ws, [...], {target})",
        detail=f"{len(bad)} sheet(s) inconsistent: {examples}",
    ))


def _check_d4_hierarchy_column_shift(wb: Workbook, report: CheckReport) -> None:
    """D4 — IS / BS / CFS の階層 line item は列シフト (Phase 6 Fix #3).

    B 列に「先頭 2 個以上の半角スペース」で indent された label がある場合、
    `write_hierarchical_line_item` (column-shift) ではなく space-indent で
    表現されており、ソート / フィルタを壊す。
    """
    target_sheets = ("04_IS", "05_BS", "06_CFS")
    bad: list[tuple[str, str, str]] = []  # (sheet, cell, label)
    for sheet_name in target_sheets:
        ws = _get_sheet(wb, sheet_name)
        if ws is None:
            continue
        for r in range(1, 201):
            v = ws.cell(row=r, column=2).value  # column B
            if not isinstance(v, str):
                continue
            if v.startswith("  ") or v.startswith("\t"):
                bad.append((sheet_name, f"B{r}", v.strip()[:30]))
                if len(bad) >= 20:
                    break
        if len(bad) >= 20:
            break
    if not bad:
        report.add(CheckResult(
            "D4", "pass", "Hierarchy uses column shift (not space indent)",
            detail=f"Verified {len(target_sheets)} sheet(s)",
        ))
        return
    examples = "; ".join(f"{s}!{c}='{lbl}'" for s, c, lbl in bad[:3])
    report.add(CheckResult(
        "D4", "soft_warn", "Hierarchy uses column shift (not space indent)",
        actual=len(bad), expected=0,
        rationale="Use ib_format.write_hierarchical_line_item(ws, row, level=1, label=...)",
        detail=f"{len(bad)} space-indented label(s): {examples}",
    ))


def _check_d5_no_stray_fill(wb: Workbook, report: CheckReport) -> None:
    """D5 — data 範囲外に余分な背景色が残っていないか.

    Inspect H 列以降 (col >= 8) 行 1-50 で patternType != None かつ
    fgColor が white でない cell を検出。`_DESIGN_EXEMPT_SHEETS` は除外。

    Row 1 の section header band (BG_HEADER_BAND = #1F3A66) は §4.7
    No-Merge Rule の正規な fill propagation なので除外する。
    """
    threshold = 5  # ≥5 stray fills per sheet → flag
    # Section header band color (No-Merge Rule §4.7) — legitimate, not stray.
    HEADER_BAND_RGBS = {"001F3A66", "FF1F3A66", "1F3A66"}
    flagged: list[tuple[str, int]] = []
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        # Skip charts-only / dashboard sheets where fills extend wider legitimately
        if ws.title in ("11_KPI_Dashboard",):
            continue
        stray = 0
        for r in range(1, 51):
            for c in range(8, 16):  # H..O
                cell = ws.cell(row=r, column=c)
                fill = cell.fill
                if fill is None or fill.patternType is None:
                    continue
                rgb = getattr(fill.fgColor, "rgb", None)
                if rgb in (None, "00000000", "FFFFFFFF", "00FFFFFF"):
                    continue
                # Row 1 header band (No-Merge Rule fill propagation) is legitimate
                if r == 1 and rgb in HEADER_BAND_RGBS:
                    continue
                # an explicit color → potential stray
                stray += 1
        if stray >= threshold:
            flagged.append((ws.title, stray))
    if not flagged:
        report.add(CheckResult(
            "D5", "pass", "No stray fill outside data range",
            detail=f"Inspected {sum(1 for ws in wb.worksheets if ws.title not in _DESIGN_EXEMPT_SHEETS)} sheet(s)",
        ))
        return
    examples = "; ".join(f"{n}={c} stray cells" for n, c in flagged[:3])
    report.add(CheckResult(
        "D5", "soft_warn", "No stray fill outside data range",
        actual=len(flagged), expected=0,
        rationale="Limit PatternFill to the data range (B-G typically)",
        detail=f"{len(flagged)} sheet(s) with strays: {examples}",
    ))


def _check_d6_section_spacing(wb: Workbook, report: CheckReport) -> None:
    """D6 — section header spacing reasonable (1-2 blank rows between).

    Conservative heuristic: detect rows where col B is bold + has BG_HEADER_BAND
    fill (i.e. rendered by `apply_section_header`). Flag only egregious cases:
        - two consecutive section headers (gap = 0)
        - a section spanning > 30 rows without a successor (likely orphan)
    """
    target_fill = ibf.BG_HEADER_BAND.upper()
    bad_pairs: list[tuple[str, int, int, str]] = []
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        header_rows: list[int] = []
        for r in range(1, 121):
            cell = ws.cell(row=r, column=2)  # column B
            font = cell.font
            fill = cell.fill
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue
            if not (font and font.bold):
                continue
            if fill is None or fill.patternType is None:
                continue
            rgb = getattr(fill.fgColor, "rgb", "") or ""
            if isinstance(rgb, str) and rgb.upper().endswith(target_fill):
                header_rows.append(r)
        # Inspect adjacent pairs
        for i in range(1, len(header_rows)):
            gap = header_rows[i] - header_rows[i - 1]
            if gap <= 1:
                bad_pairs.append((ws.title, header_rows[i - 1], header_rows[i], "adjacent"))
            elif gap > 30:
                bad_pairs.append((ws.title, header_rows[i - 1], header_rows[i], f"gap={gap}"))
    if not bad_pairs:
        report.add(CheckResult(
            "D6", "pass", "Section spacing reasonable",
            detail="No adjacent or >30-row-gap section headers found",
        ))
        return
    examples = "; ".join(
        f"{n}:row{a}->{b}({why})" for n, a, b, why in bad_pairs[:3]
    )
    report.add(CheckResult(
        "D6", "soft_warn", "Section spacing reasonable",
        actual=len(bad_pairs), expected=0,
        rationale="Leave 1-2 blank rows between section headers (current_row += 2)",
        detail=f"{len(bad_pairs)} pair(s) flagged: {examples}",
    ))


def _check_d7_sensitivity_header_uniform(wb: Workbook, report: CheckReport) -> None:
    """D7 — Sensitivity section header bg color が統一されているか (Phase 6 Stage A: embedded on 09_DCF).

    Heuristic: rows with col B bold + a non-white solid fill on row r columns
    D-K — collect fgColor.rgb. If >1 distinct color, flag.
    """
    ws = _get_sheet(wb, "09_DCF")
    if ws is None:
        report.add(CheckResult(
            "D7", "info", "Sensitivity table header uniform",
            detail="09_DCF sheet missing — Sensitivity sub-section unavailable",
        ))
        return
    colors_seen: dict[str, list[int]] = {}
    for r in range(1, 121):
        # Look across the period header columns
        for c in range(4, 12):  # D-K
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            if fill is None or fill.patternType is None:
                continue
            rgb = getattr(fill.fgColor, "rgb", None)
            if rgb in (None, "00000000", "FFFFFFFF", "00FFFFFF"):
                continue
            if not isinstance(rgb, str):
                continue
            # treat as header color only if there's a label-ish neighbour
            font = cell.font
            if font and font.bold:
                key = rgb.upper()[-6:]
                colors_seen.setdefault(key, []).append(r)
                break  # one sample per row is enough
    distinct = list(colors_seen.keys())
    if len(distinct) <= 1:
        report.add(CheckResult(
            "D7", "pass", "Sensitivity table header uniform",
            detail=f"Header color count = {len(distinct)}",
        ))
        return
    report.add(CheckResult(
        "D7", "soft_warn", "Sensitivity table header uniform",
        actual=len(distinct), expected=1,
        rationale="Apply BG_HEADER_BAND (or apply_year_header) to all sensitivity-table headers",
        detail=f"distinct header colors: {sorted(distinct)}",
    ))


def _check_d8_chart_data_range(wb: Workbook, report: CheckReport) -> None:
    """D8 — Chart series data range completeness (FAIL if broken).

    A chart is "broken" if BOTH of:
      - every series val.numRef references only 1 cell (degenerate), AND
      - the category axis is missing or also a single cell
    """
    broken: list[tuple[str, int, str]] = []  # (sheet, idx, reason)
    inspected = 0
    for ws in wb.worksheets:
        if not ws._charts:
            continue
        for idx, chart in enumerate(ws._charts):
            inspected += 1
            try:
                series = list(chart.series) if chart.series else []
            except Exception:
                continue
            if not series:
                broken.append((ws.title, idx, "no series"))
                continue
            val_cells: list[int] = []
            for s in series:
                f = None
                try:
                    f = s.val.numRef.f if s.val and s.val.numRef else None
                except Exception:
                    f = None
                n = _cells_in_ref(f) if f else None
                if n is not None:
                    val_cells.append(n)
            cat_cells: int | None = None
            try:
                s0 = series[0]
                cat = s0.cat
                if cat is not None:
                    ref = getattr(cat, "numRef", None) or getattr(cat, "strRef", None)
                    if ref is not None:
                        cat_cells = _cells_in_ref(ref.f)
            except Exception:
                cat_cells = None
            total_val = sum(val_cells) if val_cells else 0
            # Pass condition: total val cells across series ≥ 2 OR a multi-cell cat
            if total_val >= 2:
                continue
            if cat_cells and cat_cells >= 2:
                continue
            broken.append((ws.title, idx, f"val_total={total_val}, cat={cat_cells}"))
    if inspected == 0:
        report.add(CheckResult(
            "D8", "info", "Chart data range completeness",
            detail="No charts found in workbook",
        ))
        return
    if not broken:
        report.add(CheckResult(
            "D8", "pass", "Chart data range completeness",
            detail=f"Inspected {inspected} chart(s)",
        ))
        return
    examples = "; ".join(f"{s}#{i}({r})" for s, i, r in broken[:3])
    report.add(CheckResult(
        "D8", "hard_fail", "Chart data range completeness",
        actual=len(broken), expected=0,
        rationale="Use Reference(ws, min_col, max_col, min_row, max_row) with min!=max",
        detail=f"{len(broken)}/{inspected} chart(s) broken: {examples}",
    ))


def _check_d9_chart_palette(wb: Workbook, report: CheckReport) -> None:
    """D9 — Chart series uses IB palette colors (Phase 6 Fix #8).

    Read each series.graphicalProperties.solidFill.srgbClr; flag charts where
    NO series has an explicit IB palette srgbClr set (i.e. relying on openpyxl
    default theme colors).
    """
    valid_colors = {
        c.upper() for c in (
            ibf.IB_CHART_COLORS_FOOTBALL +
            ibf.IB_CHART_COLORS_BAR +
            ibf.IB_CHART_COLORS_LINE +
            [
                ibf.IB_CHART_COLORS_WATERFALL_POS,
                ibf.IB_CHART_COLORS_WATERFALL_NEG,
                ibf.IB_CHART_COLORS_WATERFALL_TOTAL,
            ]
        )
    }
    bad: list[tuple[str, int]] = []
    inspected = 0
    for ws in wb.worksheets:
        if not ws._charts:
            continue
        for idx, chart in enumerate(ws._charts):
            inspected += 1
            try:
                series = list(chart.series) if chart.series else []
            except Exception:
                continue
            any_palette_hit = False
            for s in series:
                gp = getattr(s, "graphicalProperties", None)
                if gp is None:
                    continue
                sf = getattr(gp, "solidFill", None)
                if sf is None:
                    continue
                srgb = getattr(sf, "srgbClr", None)
                # srgbClr can be the value itself (string) or an object with .val
                if srgb is None:
                    continue
                code = str(srgb).upper()[-6:] if not hasattr(srgb, "val") else str(srgb.val).upper()[-6:]
                if code in valid_colors:
                    any_palette_hit = True
                    break
            if not any_palette_hit:
                bad.append((ws.title, idx))
    if inspected == 0:
        report.add(CheckResult(
            "D9", "info", "Chart palette IB-compliant",
            detail="No charts found in workbook",
        ))
        return
    if not bad:
        report.add(CheckResult(
            "D9", "pass", "Chart palette IB-compliant",
            detail=f"Inspected {inspected} chart(s)",
        ))
        return
    examples = "; ".join(f"{s}#{i}" for s, i in bad[:3])
    report.add(CheckResult(
        "D9", "soft_warn", "Chart palette IB-compliant",
        actual=len(bad), expected=0,
        rationale="Call ib_format.apply_chart_palette(chart, IB_CHART_COLORS_*)",
        detail=f"{len(bad)}/{inspected} chart(s) using default colors: {examples}",
    ))


def _check_d10_tab_color(wb: Workbook, report: CheckReport) -> None:
    """D10 — Sheet tab color matches canonical 6-role × 17-sheet mapping.

    Source of truth:
      ib_format.SHEET_ROLE_MAPPING + ib_format.TAB_COLOR_BY_ROLE
      (_design_consistency_rules.md §2.X)

    Canonical role assignments (Phase 6 補強):
        cover  (#1F3A66) → 00_Cover
        input  (#004F49) → 01_Assumptions
        driver (#008A80) → 02_Revenue, 03_OpEx, 11_KPI_Dashboard
        output (#666666) → 04_IS, 05_BS, 06_CFS, 07_Debt,
                           08_CapTable, 09_DCF, 10_Comps
        check  (#D6913D) → 12_SanityChecks
        memo   (#ECC85A) → 13_IC_Memo

    Sheets not in SHEET_ROLE_MAPPING (e.g. 99_Glossary, 17_MA_Exit_Analysis,
    multi-segment 20a_…) are skipped: tab color is optional for them.
    """
    sheet_role = ibf.SHEET_ROLE_MAPPING
    color_by_role = ibf.TAB_COLOR_BY_ROLE
    bad: list[tuple[str, str | None, str, str]] = []  # (sheet, actual, expected, role)
    inspected = 0
    for ws in wb.worksheets:
        role = sheet_role.get(ws.title)
        if role is None:
            continue
        inspected += 1
        expected_color = color_by_role[role].upper()
        actual_color = ws.sheet_properties.tabColor
        actual_rgb: str | None = None
        if actual_color is not None:
            v = getattr(actual_color, "value", None) or getattr(actual_color, "rgb", None)
            if isinstance(v, str):
                # openpyxl stores ARGB ("FF008A80"); take last 6 chars for RGB.
                actual_rgb = v.upper()[-6:]
        if actual_rgb != expected_color:
            bad.append((ws.title, actual_rgb, expected_color, role))
    if inspected == 0:
        report.add(CheckResult(
            "D10", "info", "Tab color matches role",
            detail="None of the canonical 17 sheets present",
        ))
        return
    if not bad:
        report.add(CheckResult(
            "D10", "pass", "Tab color matches role",
            detail=(
                f"Verified {inspected}/14 canonical sheet(s) — all match "
                "role-canonical (cover/input/driver/output/check/memo)"
            ),
        ))
        return
    examples = "; ".join(
        f"{n}({r}): got={a or 'none'}, expected={e}" for n, a, e, r in bad[:5]
    )
    report.add(CheckResult(
        "D10", "soft_warn", "Tab color matches role",
        actual=len(bad), expected=0,
        rationale=(
            "Call ib_format.apply_canonical_tab_colors(wb) before save, or "
            "ib_format.set_tab_color(ws) per sheet (role inferred from name)."
        ),
        detail=(
            f"{len(bad)}/{inspected} canonical sheet(s) off-role: {examples}"
        ),
    ))


# ---------------------------------------------------------------------------
# D11 — IB Color Legend compliance
# ---------------------------------------------------------------------------
#
# IB convention (canonical: _terminology.md §1):
#   Hard input (number literal)        → blue   (IB_HARD_INPUT     = 0000FF)
#   Formula intra-sheet (= ... no !)   → black  (IB_FORMULA        = 000000)
#   Cross-sheet link (= ... ! ...)     → green  (IB_LINK_INTRA     = 008000)
#   External-file link (= ... [ ...)   → red    (IB_LINK_EXTERNAL  = FF0000)
#
# We sample data-area cells (B-Z × rows 5-200, capped at 200 cells / sheet)
# and verify font color matches the cell-value type. Cells with no explicit
# rgb font color (theme / auto / indexed) are skipped — intent is unknown.

_D11_EXEMPT_SHEETS = frozenset({
    "00_Cover", "12_SanityChecks", "13_IC_Memo", "99_Glossary",
})


def _classify_cell_type(cell) -> str:
    """Return one of: 'hard_input', 'formula_intra', 'formula_cross',
    'formula_external', 'label', 'empty'.
    """
    val = cell.value
    if val is None:
        return "empty"
    if isinstance(val, str):
        if val.startswith("="):
            if "[" in val:
                return "formula_external"
            if "!" in val:
                return "formula_cross"
            return "formula_intra"
        return "label"
    if isinstance(val, bool):
        return "label"  # rare, treat as non-numeric
    if isinstance(val, (int, float)):
        return "hard_input"
    return "empty"


def _expected_color_for_type(cell_type: str) -> str | None:
    """Return canonical 6-char RGB (no alpha) for the cell type, or None."""
    return {
        "hard_input": ibf.IB_HARD_INPUT.upper(),       # 0000FF
        "formula_intra": ibf.IB_FORMULA.upper(),       # 000000
        "formula_cross": ibf.IB_LINK_INTRA.upper(),    # 008000
        "formula_external": ibf.IB_LINK_EXTERNAL.upper(),  # FF0000
    }.get(cell_type)


def _actual_font_rgb(cell) -> str | None:
    """Return last-6 chars of explicit RGB font color, or None for theme/auto/
    indexed (which we cannot resolve without the theme XML).
    """
    f = cell.font
    if f is None or f.color is None:
        return None
    color = f.color
    # openpyxl Color.type is one of: 'rgb', 'theme', 'indexed', 'auto'
    if getattr(color, "type", None) != "rgb":
        return None
    rgb = getattr(color, "rgb", None) or getattr(color, "value", None)
    if not isinstance(rgb, str):
        return None
    rgb_up = rgb.upper()
    # Strip alpha prefix if present (e.g. 'FF0000FF' → '0000FF', '00000000' → '000000')
    if len(rgb_up) == 8:
        return rgb_up[-6:]
    if len(rgb_up) == 6:
        return rgb_up
    return None


def _check_d11_ib_color_legend(wb: Workbook, report: CheckReport) -> None:
    """D11 — IB Color Legend compliance (font color matches cell-value type).

    Detection logic per cell:
        - hard_input          → expect blue   (#0000FF)
        - formula_intra       → expect black  (#000000)
        - formula_cross       → expect green  (#008000)
        - formula_external    → expect red    (#FF0000)
    Cells with theme / auto / indexed font color are skipped (intent unknown).

    Sampling: B-Z × rows 5-200, max 200 cells per sheet.
    Severity: violation_rate == 0    → pass
              violation_rate ≤ 5%    → soft_warn
              violation_rate >  5%   → hard_fail
    Sheets in `_D11_EXEMPT_SHEETS` (Cover / SanityChecks / IC_Memo / Glossary)
    are not sampled — they are presentation surfaces, not model data.
    """
    total_checked = 0
    violations: list[tuple[str, str, str, str | None, str]] = []
    # (sheet, cell_addr, expected_rgb, actual_rgb, cell_type)

    for ws in wb.worksheets:
        if ws.title in _D11_EXEMPT_SHEETS:
            continue
        sampled = 0
        max_row = min(200, ws.max_row or 0)
        max_col = min(26, ws.max_column or 0)
        if max_row < 5 or max_col < 2:
            continue
        for row in ws.iter_rows(
            min_row=5, max_row=max_row, min_col=2, max_col=max_col
        ):
            if sampled >= 200:
                break
            for cell in row:
                if sampled >= 200:
                    break
                cell_type = _classify_cell_type(cell)
                if cell_type in ("empty", "label"):
                    continue
                expected = _expected_color_for_type(cell_type)
                actual = _actual_font_rgb(cell)
                if expected is None or actual is None:
                    # Cannot judge (theme / auto / indexed) → skip silently
                    continue
                total_checked += 1
                sampled += 1
                if expected != actual:
                    violations.append(
                        (ws.title, cell.coordinate, expected, actual, cell_type)
                    )

    if total_checked == 0:
        report.add(CheckResult(
            "D11", "info", "IB Color Legend compliance",
            detail="No explicitly-colored data cells found to check",
        ))
        return

    violation_rate = len(violations) / total_checked
    examples = "; ".join(
        f"{s}!{c}: expected #{e} actual #{a or 'none'} ({t})"
        for s, c, e, a, t in violations[:5]
    ) or "-"
    rationale = (
        "Use ib_format.apply_hard_input / apply_formula / apply_link_intra / "
        "apply_link_external. Hard input → blue (#0000FF), Formula intra-sheet → "
        "black (#000000), Cross-sheet link → green (#008000), External link → "
        "red (#FF0000)."
    )

    if not violations:
        report.add(CheckResult(
            "D11", "pass", "IB Color Legend compliance",
            actual=0, expected=0,
            detail=f"All {total_checked} explicitly-colored cells comply with IB legend",
        ))
        return

    if violation_rate <= 0.05:
        severity: Severity = "soft_warn"
    else:
        severity = "hard_fail"

    report.add(CheckResult(
        "D11", severity, "IB Color Legend compliance",
        actual=len(violations), expected=0,
        diff=violation_rate,
        rationale=rationale,
        detail=f"{len(violations)}/{total_checked} cells "
               f"({violation_rate:.1%}) violate IB legend: {examples}",
    ))


def _check_d12_number_format(wb: Workbook, report: CheckReport) -> None:
    """D12 — Numeric input cells should declare an explicit number_format.

    `General` (openpyxl default) renders ¥1234567 as 1234567 with no thousands
    separator, no currency symbol, no decimal control — an IB-quality red flag.
    Hard inputs deserve `#,##0` / `0.0%` / `0.00x` / a currency mask.

    Severity: > 30% of inspected numeric cells use General → soft_warn.
    Sample cap: 150 numeric cells per sheet (B-Z × rows 5-200).
    Sheets in `_DESIGN_EXEMPT_SHEETS` are excluded.
    """
    threshold_ratio = 0.30
    flagged: list[tuple[str, int, int]] = []
    total_inspected = 0
    total_general = 0
    for ws in wb.worksheets:
        if ws.title in _DESIGN_EXEMPT_SHEETS:
            continue
        max_row = min(ws.max_row or 0, 200)
        max_col = min(ws.max_column or 0, 26)
        if max_row < 5 or max_col < 2:
            continue
        general = 0
        inspected = 0
        for row in ws.iter_rows(min_row=5, max_row=max_row, min_col=2, max_col=max_col):
            if inspected >= 150:
                break
            for cell in row:
                if inspected >= 150:
                    break
                v = cell.value
                if v is None or isinstance(v, bool):
                    continue
                # Only inspect numeric literals (not formulas — those inherit
                # downstream context and are formatted at render time).
                if not isinstance(v, (int, float)):
                    continue
                inspected += 1
                fmt = cell.number_format or "General"
                if fmt.strip().lower() == "general":
                    general += 1
        if inspected >= 10 and (general / inspected) > threshold_ratio:
            flagged.append((ws.title, general, inspected))
        total_inspected += inspected
        total_general += general
    if total_inspected == 0:
        report.add(CheckResult(
            "D12", "info", "Number format declared on inputs",
            detail="No numeric input cells sampled",
        ))
        return
    if not flagged:
        report.add(CheckResult(
            "D12", "pass", "Number format declared on inputs",
            detail=f"{total_general}/{total_inspected} cells use General "
                   f"(within tolerance across all sheets)",
        ))
        return
    examples = "; ".join(
        f"{n}: {g}/{i} cells General" for n, g, i in flagged[:3]
    )
    report.add(CheckResult(
        "D12", "soft_warn", "Number format declared on inputs",
        actual=len(flagged), expected=0,
        rationale="Apply `#,##0` / `0.0%` / currency masks to numeric inputs. "
                  "`General` hides units and breaks IB readability.",
        detail=f"{len(flagged)} sheet(s) > 30% General: {examples}",
    ))


def _check_d13_no_merge_cells(wb: Workbook, report: CheckReport) -> None:
    """D13 — No merged cells in the model (IB best practice / No-Merge Rule).

    Cell merging breaks sort, filter, range select, formula reference (only
    the top-left cell carries the value, the others become None), named
    ranges, fill drag, and programmatic edits via openpyxl. Macabacus and
    Wall Street Prep both recommend non-merge alternatives.

    Replacement: `ib.apply_section_header_band` propagates the navy fill via
    a loop instead of merging; identical visual result, no breakage.

    Severity:
        0 merged ranges  → pass
        1-3 merged ranges → soft_warn
        4+ merged ranges  → soft_warn (hard_fail not used since this is a
                            design-quality issue, not a calculation error)

    Exempt sheets: none — `00_Cover` already merge-free via `write_cover`.

    See: references/_design_consistency_rules.md §4.7 No-Merge Rule
    """
    total_merged = 0
    locations: list[str] = []
    for ws in wb.worksheets:
        merged = list(ws.merged_cells.ranges)
        if not merged:
            continue
        total_merged += len(merged)
        for r in merged[:3]:
            locations.append(f"{ws.title}!{r.coord}")
    if total_merged == 0:
        report.add(CheckResult(
            "D13", "pass", "No merged cells (No-Merge Rule)",
            detail=f"Inspected {len(wb.worksheets)} sheet(s); 0 merged ranges",
        ))
        return
    severity: Severity = "soft_warn"
    examples = "; ".join(locations[:5])
    report.add(CheckResult(
        "D13", severity, "No merged cells (No-Merge Rule)",
        actual=total_merged, expected=0,
        rationale="merge_cells breaks sort/filter/range select/formula refs/"
                  "named ranges/fill drag. Use "
                  "`ib.apply_section_header_band(...)` (fill propagation, no "
                  "merge) instead.",
        detail=f"{total_merged} merged range(s) detected: {examples}",
    ))


def _check_design_all(wb: Workbook, report: CheckReport) -> None:
    """Run D1-D13 in sequence (each check defends itself)."""
    design_checks = (
        _check_d1_a_column_width,
        _check_d2_wrap_text_disabled,
        _check_d3_period_column_uniform,
        _check_d4_hierarchy_column_shift,
        _check_d5_no_stray_fill,
        _check_d6_section_spacing,
        _check_d7_sensitivity_header_uniform,
        _check_d8_chart_data_range,
        _check_d9_chart_palette,
        _check_d10_tab_color,
        _check_d11_ib_color_legend,
        _check_d12_number_format,
        _check_d13_no_merge_cells,
    )
    for fn in design_checks:
        try:
            fn(wb, report)
        except Exception as e:  # pragma: no cover — defensive
            report.add(CheckResult(
                fn.__name__, "info", "Design check internal error",
                detail=f"{type(e).__name__}: {e}",
            ))


# ===========================================================================
# Public API
# ===========================================================================


def run_all_checks(
    wb_path: str | Workbook,
    *,
    business_model: str = "saas",
    stage: str = "series_a",
    skip_soft: bool = False,
    skip_design: bool = False,
    wb_data: Workbook | None = None,
) -> CheckReport:
    """Run all hard + soft + design checks against the given workbook.

    Args:
        wb_path: file path to xlsx OR an already-loaded openpyxl Workbook
        business_model: see `_stress_framework §4.1` for accepted values
        stage: pre_seed | seed | series_a | series_b | growth | pre_ipo | post_ipo
        skip_soft: when True, skip S1-S10
        skip_design: when True, skip D1-D11 design quality checks
        wb_data: optional second workbook handle loaded with data_only=True. When
                 supplied (or auto-loaded from a path), value-reading checks
                 (H1-H4, H6, H12-H15, S1-S10) read cached numeric values from
                 it; structural / formula / format checks read from `wb`.
                 Without this, all numeric checks against formula cells return
                 `info` (openpyxl never evaluates formulas).

    Returns:
        CheckReport (use `.is_pass` for boolean pass/fail; `.to_markdown()` to render).
    """
    if isinstance(wb_path, Workbook):
        wb = wb_path
        # wb_data must be supplied explicitly when caller passes a Workbook —
        # we cannot re-load from disk because we have no path.
    else:
        wb = load_workbook(wb_path, data_only=False)
        # Auto-load a parallel data_only handle to unlock value-based checks.
        # Only useful if the workbook has been calculated by Excel (or another
        # engine) and the cached values were saved into the .xlsx — otherwise
        # cells will still return None for formulas. Either way the cost is
        # bounded (a single second parse) and the upside is large.
        if wb_data is None:
            try:
                wb_data = load_workbook(wb_path, data_only=True)
            except Exception:
                wb_data = None

    report = CheckReport(business_model=business_model, stage=stage)
    # Stash data-only handle on the report for any future check that opts in.
    # We do not retroactively re-read existing checks because they already use
    # `wb` directly; future checks should accept `wb_data` as a parameter.
    report._wb_data = wb_data  # type: ignore[attr-defined]

    hard_checks = (
        _check_h1_bs_balance,
        _check_h2_cash_tie,
        _check_h3_re_roll,
        _check_h4_h5_sum_check,
        _check_h6_sign_convention,
        _check_h7_circular_refs,
        _check_h8_hardcode_in_formula,
        _check_h9_sheet_naming,
        _check_h10_ib_color,
        _check_h11_sheet_count,
        _check_h12_fdso_sum,
        _check_h13_wc_roll,
        _check_h14_debt_roll,
        _check_h15_dcf_tie,
        _check_h16_error_cells,
    )
    for fn in hard_checks:
        try:
            fn(wb, report)
        except Exception as e:  # pragma: no cover — defensive
            report.add(CheckResult(
                fn.__name__, "info", "Internal Error",
                detail=f"{type(e).__name__}: {e}",
            ))

    if not skip_soft:
        try:
            _check_soft_all(wb, report, business_model=business_model, stage=stage)
        except Exception as e:  # pragma: no cover
            report.add(CheckResult(
                "Sx", "info", "Soft check internal error",
                detail=f"{type(e).__name__}: {e}",
            ))

    if not skip_design:
        try:
            _check_design_all(wb, report)
        except Exception as e:  # pragma: no cover
            report.add(CheckResult(
                "Dx", "info", "Design check internal error",
                detail=f"{type(e).__name__}: {e}",
            ))

    return report


def to_sanity_sheet(wb: Workbook, report: CheckReport) -> None:
    """Write report into 12_SanityChecks sheet (creates if missing).

    Layout:
        A1: title
        Row 3: column headers
        Row 4+: one row per CheckResult, color-coded by severity
    """
    sheet_name = "12_SanityChecks"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # clear existing content rows (preserve sheet object for tab color)
        ws.delete_rows(1, ws.max_row or 1)
    else:
        ws = wb.create_sheet(sheet_name, index=15)

    ibf.set_tab_color(ws, "checks")
    ibf.setup_sheet_layout(ws, n_periods=8, has_unit_col=False, freeze_at="A4")

    ws["A1"] = "Sanity Check Report"
    ws["A1"].font = ibf.FONT_SECTION
    status = "PASS" if report.is_pass else "FAIL"
    ws["A2"] = (
        f"Status: {status}  |  hard_fail={report.hard_fail_count}  "
        f"soft_warn={report.soft_warn_count}  pass={report.pass_count}  "
        f"info={report.info_count}  |  "
        f"business_model={report.business_model}  stage={report.stage}"
    )
    ws["A2"].font = ibf.FONT_COMMENT

    headers = ["ID", "Severity", "Title", "Sheet", "Cell", "Expected", "Actual", "Diff", "Detail"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = ibf.FONT_SUBSECTION
        cell.fill = PatternFill("solid", fgColor=ibf.BG_TOTAL_BAND)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    severity_fill = {
        "hard_fail": PatternFill("solid", fgColor="FFD9D9"),  # light red
        "soft_warn": PatternFill("solid", fgColor="FFF3CC"),  # light amber
        "pass":      PatternFill("solid", fgColor="DFF2DF"),  # light green
        "info":      PatternFill("solid", fgColor="ECECEC"),  # light gray
    }

    row_idx = 4
    for r in report.results:
        values = [
            r.check_id,
            r.severity,
            r.title,
            r.sheet or "",
            r.cell or "",
            _fmt_val(r.expected),
            _fmt_val(r.actual),
            _fmt_val(r.diff),
            r.detail or r.rationale,
        ]
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.value = v
            cell.fill = severity_fill[r.severity]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=10, color=ibf.IB_INK)
        row_idx += 1

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 60


# ===========================================================================
# CLI
# ===========================================================================


def _main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Run sanity checks against a 17-sheet xlsx model"
    )
    ap.add_argument("xlsx_path", nargs="?")
    ap.add_argument("--business-model", default="saas")
    ap.add_argument("--stage", default="series_a")
    ap.add_argument("--skip-soft", action="store_true")
    ap.add_argument("--skip-design", action="store_true",
                    help="Skip D1-D10 design quality checks")
    ap.add_argument(
        "--update-sheet", action="store_true",
        help="Write results back into 12_SanityChecks sheet",
    )
    ap.add_argument("--self-test", action="store_true",
                    help="Run D1-D10 self-test on synthetic workbooks")
    args = ap.parse_args(argv)

    if args.self_test:
        return _run_design_self_test()

    if not args.xlsx_path:
        ap.error("xlsx_path is required (or pass --self-test)")

    # Load both formula and data-only handles so value-based checks
    # (H1-H4, H12-H15, S1-S10) can read calculated results when present.
    # If the wb has never been recalculated by Excel/LibreOffice, data_only
    # will return None for formula cells and value-based checks will
    # gracefully degrade to `info`. See `_resolve_value_wb`.
    wb = load_workbook(args.xlsx_path, data_only=False)
    try:
        wb_data = load_workbook(args.xlsx_path, data_only=True)
    except Exception:
        wb_data = None
    report = run_all_checks(
        wb,
        business_model=args.business_model,
        stage=args.stage,
        skip_soft=args.skip_soft,
        skip_design=args.skip_design,
        wb_data=wb_data,
    )
    print(report.to_markdown())
    if args.update_sheet:
        to_sanity_sheet(wb, report)
        wb.save(args.xlsx_path)
    return 0 if report.is_pass else 1


# ===========================================================================
# Self-test for D1-D10 (synthetic workbooks)
# ===========================================================================


def _build_failing_design_wb() -> Workbook:
    """Build a synthetic workbook intentionally violating D1-D9 / D10."""
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    # Default sheet → make it 02_Revenue with violations
    ws = wb.active
    ws.title = "02_Revenue"
    # D1 violation: leave A column at default (13.0), don't set 20.0
    # D2 violation: set wrap_text=True on many data cells
    for r in range(5, 25):
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            cell.value = 100 + r
            cell.alignment = Alignment(wrap_text=True)
    # D3 violation: period columns inconsistent widths
    ws.column_dimensions["D"].width = 10.0
    ws.column_dimensions["E"].width = 16.0
    ws.column_dimensions["F"].width = 22.0
    ws.cell(row=4, column=4, value="Y1")
    ws.cell(row=4, column=5, value="Y2")
    ws.cell(row=4, column=6, value="Y3")
    # D5 violation: stray fills outside data range (col J, K)
    for r in range(1, 8):
        ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="FF00FF")
        ws.cell(row=r, column=11).fill = PatternFill("solid", fgColor="00FFAA")
    # D8 violation: chart with single-cell data range and no categories
    ws.cell(row=2, column=2, value=42)
    chart = BarChart()
    bad_data = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=2)
    chart.add_data(bad_data)
    # do NOT set categories, leaves cat=None → broken
    ws.add_chart(chart, "M2")
    # D9 violation: another chart with no graphicalProperties palette
    chart2 = BarChart()
    ws.cell(row=3, column=2, value=10)
    ws.cell(row=3, column=3, value=20)
    ws.cell(row=3, column=4, value=30)
    data2 = Reference(ws, min_col=2, max_col=4, min_row=3, max_row=3)
    chart2.add_data(data2, from_rows=True)
    cats2 = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=4)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "M20")

    # Add 04_IS with D4 violation (space-indent in col B)
    is_ws = wb.create_sheet("04_IS")
    is_ws.cell(row=4, column=2, value="Revenue")
    is_ws.cell(row=5, column=2, value="  Subscription")  # space-indent → D4 violation
    is_ws.cell(row=6, column=2, value="  Services")
    is_ws.cell(row=4, column=4, value="Y1")  # period header

    # Add 09_DCF with two distinct sensitivity-section header colors → D7 violation
    sens = wb.create_sheet("09_DCF")
    h1 = sens.cell(row=3, column=4, value="Y1")
    h1.font = Font(bold=True)
    h1.fill = PatternFill("solid", fgColor="1F3A66")
    h2 = sens.cell(row=10, column=4, value="Y1")
    h2.font = Font(bold=True)
    h2.fill = PatternFill("solid", fgColor="888888")  # different color → D7

    # Add 00_Cover but WITHOUT the role tab color → D10 violation
    cover = wb.create_sheet("00_Cover", index=0)
    cover.cell(row=1, column=2, value="Cover")
    # do NOT set tabColor

    # D6: two adjacent section headers
    drv = wb.create_sheet("11_KPI_Dashboard")
    s1 = drv.cell(row=5, column=2, value="Section A")
    s1.font = Font(bold=True)
    s1.fill = PatternFill("solid", fgColor=ibf.BG_HEADER_BAND)
    s2 = drv.cell(row=6, column=2, value="Section B")  # gap=1 → D6
    s2.font = Font(bold=True)
    s2.fill = PatternFill("solid", fgColor=ibf.BG_HEADER_BAND)

    # D11 violations on 02_Revenue — IB Color Legend mismatches.
    # Use explicit RGB font colors so the check can judge (theme is skipped).
    # Hard input shown in BLACK (should be blue) — classic IB violation.
    bad_hi = ws.cell(row=30, column=3, value=999)
    bad_hi.font = Font(name="Calibri", size=11, color="000000")
    # Formula intra-sheet shown in BLUE (should be black) — looks like input.
    bad_fm = ws.cell(row=31, column=3, value="=B30+1")
    bad_fm.font = Font(name="Calibri", size=11, color="0000FF")
    # Cross-sheet link shown in BLACK (should be green).
    bad_cx = ws.cell(row=32, column=3, value="='04_IS'!D5")
    bad_cx.font = Font(name="Calibri", size=11, color="000000")

    return wb


def _build_passing_design_wb() -> Workbook:
    """Build a synthetic workbook satisfying D1-D10."""
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    # Active sheet renamed to 02_Revenue, properly formatted
    ws = wb.active
    ws.title = "02_Revenue"
    ibf.setup_sheet_layout(ws, n_periods=4, has_unit_col=False, freeze_at="C5")
    # period headers at row 4
    for i, label in enumerate(["Y1", "Y2", "Y3", "Y4"]):
        ws.cell(row=4, column=3 + i, value=label)
    ibf.set_uniform_column_width(ws, ["C", "D", "E", "F"], ibf.COL_PERIOD_WIDTH)
    # body data — wrap_text=False default, with explicit number_format (D12)
    for r in range(5, 8):
        ws.cell(row=r, column=2, value=f"Item {r}")
        for c in range(3, 7):
            cell = ws.cell(row=r, column=c, value=100 + r * c)
            cell.alignment = Alignment(wrap_text=False)
            cell.number_format = "#,##0"
    # Healthy chart with multi-cell data + categories + palette
    chart = BarChart()
    data = Reference(ws, min_col=3, max_col=6, min_row=5, max_row=5)
    chart.add_data(data, from_rows=True)
    cats = Reference(ws, min_col=3, max_col=6, min_row=4, max_row=4)
    chart.set_categories(cats)
    ibf.apply_chart_palette(chart, ibf.IB_CHART_COLORS_BAR)
    ws.add_chart(chart, "H2")

    # 04_IS with column-shift hierarchy (no space-indent)
    is_ws = wb.create_sheet("04_IS")
    ibf.setup_sheet_layout(is_ws, n_periods=4, has_unit_col=False, freeze_at="C5")
    is_ws.cell(row=4, column=4, value="Y1")
    ibf.write_hierarchical_line_item(is_ws, row=5, level=0, label="Revenue", bold=True)
    ibf.write_hierarchical_line_item(is_ws, row=6, level=1, label="Subscription")
    ibf.write_hierarchical_line_item(is_ws, row=7, level=1, label="Services")

    # 09_DCF with uniform sensitivity-section header color
    sens = wb.create_sheet("09_DCF")
    ibf.setup_sheet_layout(sens, n_periods=4, has_unit_col=False, freeze_at="C5")
    for hr in (3, 10):
        c = sens.cell(row=hr, column=4, value="Y1")
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=ibf.BG_HEADER_BAND)

    # 00_Cover
    cover = wb.create_sheet("00_Cover", index=0)
    ibf.setup_sheet_layout(cover, n_periods=4, has_unit_col=False, freeze_at="A1")

    # 01_Assumptions and 02_Drivers with spaced headers
    for name in ("01_Assumptions", "11_KPI_Dashboard"):
        ws_x = wb.create_sheet(name)
        ibf.setup_sheet_layout(ws_x, n_periods=4, has_unit_col=False, freeze_at="C5")
        # Two section headers spaced 5 rows apart
        for hr, label in [(5, "Section A"), (10, "Section B")]:
            cell = ws_x.cell(row=hr, column=2, value=label)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=ibf.BG_HEADER_BAND)

    # Remaining canonical sheets so D10 (17-sheet coverage) is exercised.
    # Stub the output / driver / check / memo sheets — D10 only inspects
    # tab color, but D1 needs A column width = COL_MARGIN_WIDTH so we run
    # setup_sheet_layout on each stub.
    for name in (
        "03_OpEx", "05_BS", "06_CFS", "07_Debt",
        "08_CapTable", "09_DCF", "10_Comps", "11_KPI_Dashboard",
        "12_SanityChecks", "13_IC_Memo",
    ):
        ws_stub = wb.create_sheet(name)
        ibf.setup_sheet_layout(ws_stub, n_periods=4, has_unit_col=False, freeze_at="A1")

    # Apply canonical tab colors to ALL 17 canonical sheets in one pass.
    # This is the recommended end-of-build call (Phase 6 補強 §2.X).
    ibf.apply_canonical_tab_colors(wb)

    # D11 — IB-correct sample cells on 02_Revenue (explicit RGB so the check
    # has something to judge; theme-colored body cells above are skipped).
    hi = ws.cell(row=30, column=3, value=100)
    hi.font = Font(name=ibf.FONT_FAMILY, size=ibf.FONT_SIZE_BASE,
                   color=ibf.IB_HARD_INPUT)  # blue
    hi.number_format = "#,##0"
    fm = ws.cell(row=31, column=3, value="=B30+1")
    fm.font = Font(name=ibf.FONT_FAMILY, size=ibf.FONT_SIZE_BASE,
                   color=ibf.IB_FORMULA)  # black
    cx = ws.cell(row=32, column=3, value="='04_IS'!D5")
    cx.font = Font(name=ibf.FONT_FAMILY, size=ibf.FONT_SIZE_BASE,
                   color=ibf.IB_LINK_INTRA)  # green

    return wb


def _run_design_self_test() -> int:
    """Run D1-D12 + 4 H/S regression cases against synthetic workbooks. Returns 0/1."""
    print("=" * 70)
    print("Design checks self-test (D1-D12) + Hard/Soft regression (H1, H2, H16, S2)")
    print("=" * 70)

    # --- Failing sample
    print("\n[1] Failing-sample workbook (expect WARN/FAIL on D1-D12):")
    fail_wb = _build_failing_design_wb()
    fail_report = CheckReport()
    _check_design_all(fail_wb, fail_report)
    fail_d_results = [r for r in fail_report.results if r.check_id.startswith("D")]
    fail_pass_ids = {r.check_id for r in fail_d_results if r.severity == "pass"}
    fail_warn_ids = {r.check_id for r in fail_d_results if r.severity in ("soft_warn", "hard_fail")}
    for r in fail_d_results:
        marker = {"pass": "OK ", "soft_warn": "WRN", "hard_fail": "FAL", "info": "INF"}.get(r.severity, "???")
        print(f"  [{marker}] {r.check_id}: {r.title}  ({r.detail or r.rationale})"[:140])

    # We expect: D1..D12 all flagged.
    expected_flagged = {f"D{i}" for i in range(1, 13)}
    missing_flags = expected_flagged - fail_warn_ids
    unexpected_pass = fail_pass_ids & expected_flagged

    # --- Passing sample
    print("\n[2] Passing-sample workbook (expect PASS or INFO on D1-D12):")
    pass_wb = _build_passing_design_wb()
    pass_report = CheckReport()
    _check_design_all(pass_wb, pass_report)
    pass_d_results = [r for r in pass_report.results if r.check_id.startswith("D")]
    pass_warn_ids = {r.check_id for r in pass_d_results if r.severity in ("soft_warn", "hard_fail")}
    for r in pass_d_results:
        marker = {"pass": "OK ", "soft_warn": "WRN", "hard_fail": "FAL", "info": "INF"}.get(r.severity, "???")
        print(f"  [{marker}] {r.check_id}: {r.title}  ({r.detail or r.rationale})"[:140])

    # --- Hard / Soft regression cases (4 minimal scenarios) -----------------
    # These exercise the most load-bearing checks against tiny synthetic wbs
    # whose values are direct numeric literals (no formula evaluation needed).
    print("\n[3] Hard/Soft regression cases (H1, H2, H16, S2):")
    hs_failures: list[str] = []

    # --- H1 — BS unbalanced by 100 ---
    h1_wb = _build_unbalanced_bs_wb(diff=100.0)
    h1_report = CheckReport()
    _check_h1_bs_balance(h1_wb, h1_report)
    h1_hits = [r for r in h1_report.results if r.check_id == "H1" and r.severity == "hard_fail"]
    print(f"  [{'OK ' if h1_hits else 'FAL'}] H1 BS unbalanced wb → {len(h1_hits)} hard_fail (expect ≥1)")
    if not h1_hits:
        hs_failures.append("H1 (BS unbalanced) did not flag hard_fail")

    # --- H2 — Cash mismatch (BS cash=200, CFS ending cash=180) ---
    h2_wb = _build_cash_mismatch_wb()
    h2_report = CheckReport()
    _check_h2_cash_tie(h2_wb, h2_report)
    h2_hits = [r for r in h2_report.results if r.check_id == "H2" and r.severity == "hard_fail"]
    print(f"  [{'OK ' if h2_hits else 'FAL'}] H2 Cash mismatch wb → {len(h2_hits)} hard_fail (expect ≥1)")
    if not h2_hits:
        hs_failures.append("H2 (cash mismatch) did not flag hard_fail")

    # --- H16 — #REF! literal in cell ---
    h16_wb = _build_error_cell_wb()
    h16_report = CheckReport()
    _check_h16_error_cells(h16_wb, h16_report)
    h16_hits = [r for r in h16_report.results if r.check_id == "H16" and r.severity == "hard_fail"]
    print(f"  [{'OK ' if h16_hits else 'FAL'}] H16 #REF! cell wb → {len(h16_hits)} hard_fail (expect ≥1)")
    if not h16_hits:
        hs_failures.append("H16 (error cell) did not flag hard_fail")

    # --- S2 — Burn Multiple = 3.0 (above threshold 2.0) ---
    s2_wb = _build_high_burn_multiple_wb(burn_mult=3.0)
    s2_report = CheckReport()
    _check_soft_all(s2_wb, s2_report, business_model="saas", stage="series_a")
    s2_hits = [r for r in s2_report.results if r.check_id == "S2" and r.severity == "soft_warn"]
    print(f"  [{'OK ' if s2_hits else 'FAL'}] S2 Burn Multiple=3.0 wb → {len(s2_hits)} soft_warn (expect ≥1)")
    if not s2_hits:
        hs_failures.append("S2 (burn multiple high) did not flag soft_warn")

    # --- Verdict
    print("\n" + "=" * 70)
    ok = True
    if missing_flags:
        print(f"FAIL: failing-sample did NOT flag {sorted(missing_flags)}")
        ok = False
    if unexpected_pass:
        print(f"FAIL: failing-sample passed {sorted(unexpected_pass)} (should have flagged)")
        ok = False
    if pass_warn_ids:
        print(f"FAIL: passing-sample flagged {sorted(pass_warn_ids)} (should have passed)")
        ok = False
    if hs_failures:
        for f in hs_failures:
            print(f"FAIL: H/S regression — {f}")
        ok = False
    if ok:
        print("PASS: D1-D12 + H/S regression self-test successful")
        print(f"  failing-sample flagged: {sorted(fail_warn_ids)}")
        print(f"  passing-sample clean: {len(pass_d_results)} D-checks all pass/info")
        print("  H/S regression: H1, H2, H16, S2 all flagged correctly")
        return 0
    else:
        return 1


# ---------------------------------------------------------------------------
# Synthetic workbooks for H/S regression (minimal, direct-value cells)
# ---------------------------------------------------------------------------


def _build_unbalanced_bs_wb(*, diff: float = 100.0) -> Workbook:
    """Tiny wb with 05_BS where Total Assets ≠ Total L&E by `diff`."""
    wb = Workbook()
    wb.active.title = "00_Cover"
    bs = wb.create_sheet("05_BS")
    bs.cell(row=4, column=4, value="Y1")  # period header
    bs.cell(row=10, column=2, value="Total Assets")
    bs.cell(row=10, column=4, value=1000.0)
    bs.cell(row=20, column=2, value="Total Liabilities and Equity")
    bs.cell(row=20, column=4, value=1000.0 - diff)
    return wb


def _build_cash_mismatch_wb() -> Workbook:
    """BS cash=200, CFS ending cash=180 — H2 should flag."""
    wb = Workbook()
    wb.active.title = "00_Cover"
    bs = wb.create_sheet("05_BS")
    bs.cell(row=4, column=4, value="Y1")
    bs.cell(row=5, column=2, value="Cash and Cash Equivalents")
    bs.cell(row=5, column=4, value=200.0)
    cfs = wb.create_sheet("06_CFS")
    cfs.cell(row=4, column=4, value="Y1")
    cfs.cell(row=20, column=2, value="Ending Cash")
    cfs.cell(row=20, column=4, value=180.0)
    return wb


def _build_error_cell_wb() -> Workbook:
    """A clean wb except for one #REF! literal — H16 should flag."""
    wb = Workbook()
    ws = wb.active
    ws.title = "02_Revenue"
    ws.cell(row=5, column=2, value="Revenue")
    ws.cell(row=5, column=3, value=100)
    ws.cell(row=6, column=2, value="Bad reference")
    ws.cell(row=6, column=3, value="#REF!")  # error literal
    return wb


def _build_high_burn_multiple_wb(*, burn_mult: float = 3.0) -> Workbook:
    """11_KPI_Dashboard with Burn Multiple row at the given value (> 2.0)."""
    wb = Workbook()
    wb.active.title = "00_Cover"
    kpi = wb.create_sheet("11_KPI_Dashboard")
    kpi.cell(row=4, column=4, value="Y1")  # period header
    kpi.cell(row=10, column=2, value="Burn Multiple")
    kpi.cell(row=10, column=4, value=burn_mult)
    return wb


if __name__ == "__main__":
    sys.exit(_main())
