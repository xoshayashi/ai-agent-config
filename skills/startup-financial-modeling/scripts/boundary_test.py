#!/usr/bin/env python3
"""境界値テスト: 期首残高を非ゼロにしてもモデルが自壊しないことを検証する。

背景（このスキルで繰り返した失敗）:
  期首境界値（期初稼働数・期首簿価・期首人員・期首負債・期首売掛…）を数式内の
  リテラル0で埋めると、既存事業に使った瞬間に初年度の計算が静かに消える。しかも
  ロール整合チェックが両辺で同じゼロを読むため沈黙する。採点では毎回ここを突かれた。

検証内容:
  ラベルが「期首」「期初」で始まる入力行の**F列だけ**に非ゼロを投入して再計算し、
  (a) 数式エラーがゼロ、(b) 総合判定が OK のまま、を確認する。
  期首を入れた結果としてロール整合が壊れるなら、それは境界値が配線されていない証拠。

さらに、スイッチ入力に定義域外の値（例: 9）を入れても #VALUE! を撒かず、
総合判定が「要確認」を返す（自壊せず検知する）ことを確認する。

使い方: python3 boundary_test.py <xlsx>
終了コード 0=合格 / 1=不合格 / 2=環境エラー
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

# 期首入力のラベル種別 → 投入する値（単位は生値）
SEED = {
    "現金": 5_000_000_000, "取得原価": 3_000_000_000,
    "累計償却": 500_000_000, "残高": 2_000_000_000,
    "売掛": 300_000_000, "買掛": 200_000_000,
    "繰越欠損金": 1_000_000_000, "欠損金": 1_000_000_000,
    "人員": 10, "在籍": 10, "顧客": 5, "稼働": 100, "フリート": 100,
}
DEFAULT_SEED = 10


def _recalc(path: Path, outdir: Path) -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    if not Path(soffice).exists():
        print("SKIP: sofficeが見つかりません")
        sys.exit(2)
    prof = Path(tempfile.mkdtemp(prefix="lo_")).as_uri()
    subprocess.run([soffice, f"-env:UserInstallation={prof}",
                    "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(outdir), str(path)],
                   check=True, capture_output=True, timeout=300)
    return outdir / path.name


def _state(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    errs = [f"{ws.title}!{c.coordinate}"
            for ws in wb.worksheets for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("#")]
    verdict, ni, cash = None, [], []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            label = next((c.value for c in row[1:4]
                          if isinstance(c.value, str) and c.value.strip()),
                         None)
            if not isinstance(label, str):
                continue
            if label.startswith("総合判定"):
                verdict = next((c.value for c in row[5:]
                                if c.value is not None), None)
            elif ws.title == "損益計画" and label == "当期純利益":
                ni = [c.value if isinstance(c.value, (int, float)) else 0
                      for c in row[5:10]]
            elif ws.title == "資金繰り" and label == "期末現金":
                cash = [c.value if isinstance(c.value, (int, float)) else 0
                        for c in row[5:10]]
    return verdict, errs, ni, cash


def _seed_for(label: str):
    for key, val in SEED.items():
        if key in label:
            return val
    return DEFAULT_SEED


def main(path: str) -> int:
    src = Path(path)
    fails = []
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        # --- 期首境界値の一括投入 ---
        wb = openpyxl.load_workbook(src)
        ws = wb["前提条件"]
        seeded = []
        for r in range(6, ws.max_row + 1):
            label = next((ws.cell(row=r, column=c).value for c in (2, 3, 4)
                          if isinstance(ws.cell(row=r, column=c).value, str)),
                         None)
            if not isinstance(label, str):
                continue
            if not (label.startswith("期首") or label.startswith("期初")):
                continue
            val = _seed_for(label)
            # F列（初年度）だけに入れる。期首残高は単一値であり、モデル側は
            # $F$固定で読むべき。全列に入れると「同列参照」の欠陥を見逃す。
            ws.cell(row=r, column=6).value = val
            seeded.append(label)
        if not seeded:
            print("SKIP: 期首入力行が見つかりません")
            return 0
        p = td / "opening.xlsx"
        wb.save(p)
        b_v, b_errs, b_ni, b_cash = _state(_recalc(src, td / "base"))
        verdict, errs, ni, cash = _state(_recalc(p, td / "rc"))
        moved = any(abs((x or 0) - (y or 0)) > 1_000_000
                    for x, y in zip(ni, b_ni)) or \
            any(abs((x or 0) - (y or 0)) > 1_000_000
                for x, y in zip(cash, b_cash))
        print(f"期首境界値 {len(seeded)}行に非ゼロを投入: "
              f"総合判定={verdict} 数式エラー={len(errs)}件 "
              f"財務への伝播={'あり' if moved else 'なし'}")
        if errs:
            fails.append(f"期首投入で数式エラー {len(errs)}件: {errs[:4]}")
        if verdict != "OK":
            fails.append(f"期首投入で総合判定が {verdict}"
                         "（期首が配線されていない疑い）")
        if not moved:
            fails.append("期首を入れてもNI・現金が動かない"
                         "（＝入力行が孤児。数式内のリテラル0が残っている疑い）")

        # --- スイッチの定義域外 ---
        wb2 = openpyxl.load_workbook(src)
        ws2 = wb2["前提条件"]
        sw = None
        for r in range(6, ws2.max_row + 1):
            label = next((ws2.cell(row=r, column=c).value for c in (2, 3, 4)
                          if isinstance(ws2.cell(row=r, column=c).value, str)),
                         None)
            if label == "アクティブケース":
                sw = r
        if sw:
            ws2.cell(row=sw, column=6).value = 9
            p2 = td / "oor.xlsx"
            wb2.save(p2)
            v2, e2, _, _ = _state(_recalc(p2, td / "rc2"))
            print(f"スイッチ定義域外(=9): 総合判定={v2} 数式エラー={len(e2)}件")
            if e2:
                fails.append(f"定義域外スイッチで数式エラー {len(e2)}件"
                             "（CHOOSEのクランプが未実装）")
            if v2 != "要確認":
                fails.append("定義域外スイッチを検知できていない"
                             f"（総合判定={v2}）")
    if fails:
        print("FAIL 境界値テスト:")
        for f in fails:
            print("  -", f)
        return 1
    print("PASS 境界値テスト: 期首非ゼロでも自壊せず、定義域外は検知される")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
