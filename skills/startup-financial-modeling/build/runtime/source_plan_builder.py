"""Generic economic-kernel startup financial-plan workbook builder.

The source Markdown route builds an editable investor-grade xlsx from a generic
startup finance kernel. It does not start from a sector template. Source text is
used to infer mechanics such as marketplace, recurring software, hardware /
asset-heavy operations, lending, or pre-revenue R&D, then the same workbook
architecture is composed from drivers, operating logic, capital stack,
ownership, scenarios, valuation, and memo surfaces.
"""

from __future__ import annotations

import sys
import math
import re
import statistics
from copy import copy
from pathlib import Path
from dataclasses import dataclass, field, replace as _dc_replace

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))

import ib_format as ib  # noqa: E402
from economic_kernel import (  # noqa: E402
    PeriodAxis,
    SourceFacts,
    build_period_axis,
    derive_source_facts,
    derive_source_facts_from_mapping,
    driver_surfaces_for,
    expand_annual_series,
    extract_source_facts,
    implied_gross_margin_series,
    mechanic_key,
    months_factor as _months_factor,
    plan_revenue_series,
    project_free_cash_flow,
    project_plan_free_cash_flow,
    scenario_drivers_for,
)



@dataclass(frozen=True)
class LayoutSpec:
    hierarchy_cols: int = 1
    hierarchy_width: float = ib.COL_HIERARCHY_WIDTH
    label_width: float = ib.COL_LABEL_WIDTH
    source_width: float = ib.COL_SOURCE_WIDTH
    unit_width: float = ib.COL_UNIT_WIDTH
    period_width: float = ib.COL_PERIOD_WIDTH
    note_width: float = ib.COL_NOTE_WIDTH

    @property
    def first_hierarchy_col(self) -> int:
        return 2

    @property
    def label_col(self) -> int:
        return self.first_hierarchy_col + self.hierarchy_cols

    @property
    def source_col(self) -> int:
        return self.label_col + 1

    @property
    def unit_col(self) -> int:
        return self.source_col + 1

    @property
    def first_value_col(self) -> int:
        return self.unit_col + 1


LAYOUT = LayoutSpec()
START_PERIOD_COL = LAYOUT.first_value_col


def _start_period_col() -> int:
    return LAYOUT.first_value_col

YEN_INPUT_SCALES = {
    "JPY K": 1_000,
    "JPY M": 1_000_000,
    "JPY B": 1_000_000_000,
    "JPY T": 1_000_000_000_000,
}

USD_INPUT_SCALES = {
    "USD K": 1_000,
    "USD M": 1_000_000,
}

MONEY_INPUT_SCALES = {
    **YEN_INPUT_SCALES,
    **USD_INPUT_SCALES,
}

DISPLAY_UNIT_BY_SCALE = {
    ("JPY", "actual"): "円",
    ("JPY", "thousand"): "千円",
    ("JPY", "million"): "百万円",
    ("JPY", "hundred_million"): "億円",
    ("JPY", "billion"): "十億円",
    ("JPY", "trillion"): "兆円",
    ("USD", "actual"): "$",
    ("USD", "thousand"): "$K",
    ("USD", "million"): "$M",
}

YEN_DISPLAY_UNITS = {
    "JPY": "円",
    "JPY K": "千円",
    "JPY M": "百万円",
    "JPY B": "十億円",
    "JPY T": "兆円",
}

MONEY_INPUT_UNITS = {"JPY", "USD", *MONEY_INPUT_SCALES.keys()}

MONEY_DISPLAY_UNITS = {
    **YEN_DISPLAY_UNITS,
    "USD": "$",
    "USD K": "$K",
    "USD M": "$M",
}

MONEY_DISPLAY_SCALE_FACTORS = (
    1_000_000_000_000,
    1_000_000_000,
    1_000_000,
    1_000,
)


def _display_unit(unit: str, fmt: str | None = None, currency: str = "JPY", scale: str = "million") -> str:
    if unit == "JPY":
        if fmt in {ib.FMT_JPY_YEN, ib.FMT_USD_DOLLAR}:
            return DISPLAY_UNIT_BY_SCALE.get((currency, "actual"), "円")
        if fmt in {ib.FMT_JPY_THOUSAND, ib.FMT_USD_THOUSAND}:
            return DISPLAY_UNIT_BY_SCALE.get((currency, "thousand"), "千円")
        if fmt == ib.FMT_JPY_HUNDRED_MILLION:
            return DISPLAY_UNIT_BY_SCALE.get((currency, "hundred_million"), "億円")
        if fmt == ib.FMT_JPY_BILLION:
            return DISPLAY_UNIT_BY_SCALE.get((currency, "billion"), "十億円")
        if fmt == ib.FMT_JPY_TRILLION:
            return DISPLAY_UNIT_BY_SCALE.get((currency, "trillion"), "兆円")
        if fmt in {ib.FMT_MONEY, ib.FMT_MONEY_DECIMAL, ib.FMT_JPY_MILLION, ib.FMT_USD_MILLION}:
            return DISPLAY_UNIT_BY_SCALE.get((currency, scale), DISPLAY_UNIT_BY_SCALE.get((currency, "million"), "百万円"))
        return DISPLAY_UNIT_BY_SCALE.get((currency, "actual"), YEN_DISPLAY_UNITS["JPY"])
    # Non-JPY currency scale variants are encoded in the unit string
    # (`USD K`, `USD M`) rather than fmt-dispatched from raw `USD`.
    return MONEY_DISPLAY_UNITS.get(unit, unit)


def _normalise_formula_scale(formula: str) -> str:
    # Money source formulas sometimes arrive pre-scaled for presentation.
    # Store raw base-currency formulas and let number_format handle display.
    for factor in MONEY_DISPLAY_SCALE_FACTORS:
        formula = re.sub(rf"\s*/\s*{factor}\b", "", formula)
        formula = re.sub(rf"\s*\*\s*{factor}\b", "", formula)
    return formula


def _model_value(value: object, unit: str) -> object:
    if isinstance(value, str) and value.startswith("="):
        if unit in MONEY_INPUT_UNITS:
            return _normalise_formula_scale(value)
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        scale = MONEY_INPUT_SCALES.get(unit)
        if scale:
            return int(value * scale) if float(value * scale).is_integer() else value * scale
    return value


def _money_format(facts: SourceFacts) -> str:
    return ib.fmt_for_currency(facts.currency, facts.display_scale)


def _money_unit(facts: SourceFacts) -> str:
    return _display_unit("JPY", _money_format(facts), facts.currency, facts.display_scale)


def _facts_for_sheet(ws: Worksheet, facts: SourceFacts | None = None) -> SourceFacts | None:
    if facts is not None:
        return facts
    return getattr(ws, "_startup_facts", None)


def _format_for_unit(unit: str, requested_fmt: str, facts: SourceFacts | None = None) -> str:
    if unit == "%" or unit.startswith("%"):
        return ib.FMT_PERCENT
    if unit == "x":
        return ib.FMT_MULTIPLE
    if unit in {"months", "count"}:
        return ib.FMT_NUM
    if unit in {"units", "customers", "FTE", "days"}:
        return ib.FMT_INTEGER
    if unit == "JPY":
        if facts is not None:
            if requested_fmt == ib.FMT_MONEY:
                return _money_format(facts)
            if requested_fmt == ib.FMT_JPY_YEN:
                return ib.fmt_for_currency(facts.currency, "actual")
            if requested_fmt == ib.FMT_JPY_THOUSAND:
                return ib.fmt_for_currency(facts.currency, "thousand")
        return requested_fmt
    if unit == "JPY K":
        return ib.FMT_JPY_THOUSAND
    if unit == "JPY M":
        return ib.FMT_JPY_MILLION
    if unit == "JPY B":
        return ib.FMT_JPY_BILLION
    if unit == "JPY T":
        return ib.FMT_JPY_TRILLION
    if unit == "USD":
        if requested_fmt in {ib.FMT_MONEY, ib.FMT_USD_MILLION}:
            return ib.FMT_USD_MILLION
        if requested_fmt in {ib.FMT_JPY_YEN, ib.FMT_USD_DOLLAR}:
            return ib.FMT_USD_DOLLAR
        if requested_fmt in {ib.FMT_JPY_THOUSAND, ib.FMT_USD_THOUSAND}:
            return ib.FMT_USD_THOUSAND
        return requested_fmt
    if unit == "USD K":
        return ib.FMT_USD_THOUSAND
    if unit == "USD M":
        return ib.FMT_USD_MILLION
    return requested_fmt


FMT_KEY_MAP = {
    "money": ib.FMT_MONEY,
    "jpy_yen": ib.FMT_JPY_YEN,
    "integer": ib.FMT_INTEGER,
    "percent": ib.FMT_PERCENT,
    "multiple": ib.FMT_MULTIPLE,
    "num": ib.FMT_NUM,
    "text": ib.FMT_NUM,
}


def _apply_unit_cell(cell) -> None:
    cell.font = ib.FONT_COMMENT
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)


def _disable_wrap_text(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.alignment is not None and cell.alignment.wrap_text is True:
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = False
                    cell.alignment = alignment


def _clear_blank_cell_styles(wb: Workbook) -> None:
    ib.clear_blank_cell_styles(wb)


def _autosize_default_layout_columns(wb: Workbook) -> None:
    """Content-drive label / source / note widths on default-layout sheets.

    Runs after all data is written so the measurement reflects the final
    cell content. The width is resolved once per role across the workbook
    and then applied to every default-layout sheet carrying that role, so a
    long source note on one sheet does not leave the same source column
    narrower elsewhere. Per spec.md §2 C1 the period and unit columns stay at
    their canonical fixed widths. Custom-layout sheets (Driver Tree, Cap
    Table, Exit Waterfall, Benchmarks) keep the widths their builders chose.
    """
    sheet_roles: list[tuple[Worksheet, dict[int, str]]] = []
    resolved_widths: dict[str, float] = {}
    for ws in wb.worksheets:
        if not uses_default_layout(ws):
            continue
        facts = _facts_for_sheet(ws)
        if facts is None:
            continue
        role_columns: dict[int, str] = {
            LAYOUT.label_col: "label",
            LAYOUT.source_col: "source",
        }
        note_col = None
        for col in range(_start_period_col() + 1, ws.max_column + 1):
            if ws.cell(row=ib.HEADER_PERIOD_ROW, column=col).value == "Notes":
                note_col = col
                break
        if note_col is not None:
            has_note_content = any(
                ws.cell(row=r, column=note_col).value not in (None, "")
                for r in range(1, ws.max_row + 1)
            )
            if has_note_content:
                role_columns[note_col] = "note"
        sheet_roles.append((ws, role_columns))
        for col, role in role_columns.items():
            width = ib.role_column_width_for_content(ws, col, role)
            if width is None:
                continue
            resolved_widths[role] = max(resolved_widths.get(role, 0), width)
    for ws, role_columns in sheet_roles:
        for col, role in role_columns.items():
            width = resolved_widths.get(role)
            if width is not None:
                ws.column_dimensions[get_column_letter(col)].width = width


def _last_value_bounds(ws: Worksheet) -> tuple[int, int]:
    return ib.last_value_bounds(ws)


def _rendered_bounds(ws: Worksheet) -> tuple[int, int]:
    return ib.rendered_bounds(ws)


def _trim_blank_canvas(wb: Workbook) -> None:
    ib.trim_blank_canvas(wb)


def _set_column_widths(ws: Worksheet, widths: dict[int | str, float]) -> None:
    # Role-min floors follow the v2 grid (_layout_canonical.md): the legacy
    # 54-wide label/source floors would otherwise clobber a register sheet's
    # requested v2 widths (Cap Table / Evidence label col up to 54), the one
    # remaining place the retired v1 layout leaked into generated widths.
    role_min_widths = {
        LAYOUT.label_col: ib.COL_LABEL_WIDTH_V2,
        LAYOUT.source_col: ib.COL_SOURCE_WIDTH_INPUT,
        LAYOUT.unit_col: ib.COL_UNIT_WIDTH_V2,
    }
    indent_cols = set(range(LAYOUT.first_hierarchy_col, LAYOUT.label_col))
    for col, width in widths.items():
        letter = col if isinstance(col, str) else get_column_letter(col)
        col_index = col if isinstance(col, int) else ws[letter + "1"].column
        if col_index in indent_cols:
            # Indent columns are pinned by _setup_sheet to ib.INDENT_COL_WIDTH
            # (Google Sheets 20px). Refuse silently-overriding overrides;
            # callers must shift their first data column past the indent block.
            raise ValueError(
                f"refusing to override indent column {letter} on '{ws.title}'; "
                f"data columns must start at or after column "
                f"{get_column_letter(LAYOUT.label_col)}"
            )
        floor = role_min_widths.get(col_index, 0)
        ws.column_dimensions[letter].width = max(width, floor)


def _apply_text_header(cell, label: str) -> None:
    cell.value = label
    cell.font = ib.FONT_BODY_BOLD
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ib.apply_semantic_fill_span(cell.parent, cell.row, cell.column, cell.column, ib.BG_TABLE_HEADER, bottom=ib.THIN_LINE)


def _compact_status(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value.strip()
    normalized = text.lower()
    if normalized in {"estimate / needs validation", "estimate/needs validation"}:
        return "estimate"
    if "pipeline" in normalized and "estimate" in normalized:
        return "pipeline-backed"
    if "management target" in normalized:
        return "management target"
    if "placeholder" in normalized:
        return "placeholder"
    if "unknown" in normalized:
        return "unknown"
    return text


def _section_band_end_col(ws: Worksheet, explicit_end_col: int | None = None) -> int:
    if explicit_end_col is not None:
        return explicit_end_col
    note_col = getattr(ws, "_startup_note_col", None)
    facts = _facts_for_sheet(ws)
    if facts is not None:
        return max(
            LAYOUT.unit_col,
            _start_period_col() + len(facts.period_labels) - 1,
            note_col or 0,
        )
    return max(LAYOUT.unit_col, ws.max_column)


def _section(ws: Worksheet, row: int, label: str, end_col: int | None = None) -> None:
    band_end_col = _section_band_end_col(ws, end_col)
    cell = ws.cell(row=row, column=LAYOUT.first_hierarchy_col, value=label)
    cell.font = Font(name=ib.FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ib.apply_semantic_fill_span(
        ws,
        row,
        LAYOUT.first_hierarchy_col,
        band_end_col,
        ib.BG_HEADER_BAND,
        bottom=ib.THIN_LINE,
        border_start_col=_row_rule_start_col(ws),
    )
    ws.row_dimensions[row].height = ib.ROW_HEIGHT_RELAXED


def _note(ws: Worksheet, row: int, text: str) -> None:
    note_col = getattr(ws, "_startup_note_col", None)
    if note_col is None:
        note_col = max(ws.max_column + 1, 10)
        ws._startup_note_col = note_col
    ws.column_dimensions[get_column_letter(note_col)].width = LAYOUT.note_width
    cell = ws.cell(row=row, column=note_col, value=text)
    ib.apply_comment(cell, wrap_text=False)


def _highlight_row(ws: Worksheet, row: int, last_col: int | None = None) -> None:
    end_col = last_col if last_col is not None else max(ws.max_column, 9)
    ib.apply_semantic_fill_span(
        ws,
        row,
        LAYOUT.first_hierarchy_col,
        end_col,
        ib.BG_WORKING,
    )
    # Borders go across the detected table block instead of the old
    # `accent_cols = [label_col, *period_cols]` pattern, so the
    # selected-output underline reads as continuous through the source
    # and unit columns. The indent gutter stays borderless via
    # border_start_col = _row_rule_start_col(ws).
    ib.apply_semantic_border_span(
        ws,
        row,
        top=ib.THIN_LINE,
        bottom=ib.THIN_LINE,
        border_start_col=_row_rule_start_col(ws),
    )


def _merge_border(existing: Border | None, *, top=None, bottom=None, left=None, right=None) -> Border:
    if existing is None:
        return Border(top=top, bottom=bottom, left=left, right=right)
    return Border(
        top=top if top is not None else existing.top,
        bottom=bottom if bottom is not None else existing.bottom,
        left=left if left is not None else existing.left,
        right=right if right is not None else existing.right,
    )


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    value = getattr(fill.fgColor, "rgb", None)
    return value[-6:].upper() if isinstance(value, str) else None


def _is_section_row(ws: Worksheet, row: int) -> bool:
    fill = ws.cell(row=row, column=LAYOUT.first_hierarchy_col).fill
    value = getattr(fill.fgColor, "rgb", None)
    if fill.fill_type == "solid" and isinstance(value, str) and value.endswith(ib.BG_HEADER_BAND):
        return True
    color = ws.cell(row=row, column=LAYOUT.first_hierarchy_col).font.color
    font_rgb = getattr(color, "rgb", None)
    return isinstance(font_rgb, str) and font_rgb.endswith(ib.BG_HEADER_BAND)


def _is_highlight_row(ws: Worksheet, row: int) -> bool:
    fill = ws.cell(row=row, column=LAYOUT.first_hierarchy_col).fill
    value = getattr(fill.fgColor, "rgb", None)
    return fill.fill_type == "solid" and isinstance(value, str) and value.endswith(ib.BG_WORKING)


def uses_default_layout(ws: Worksheet) -> bool:
    header_row = ib.HEADER_PERIOD_ROW
    hierarchy_values = [
        ws.cell(row=header_row, column=col).value
        for col in range(LAYOUT.first_hierarchy_col, LAYOUT.label_col)
    ]
    role_values = [
        ws.cell(row=header_row, column=LAYOUT.label_col).value,
        ws.cell(row=header_row, column=LAYOUT.source_col).value,
        ws.cell(row=header_row, column=LAYOUT.unit_col).value,
    ]
    return (
        hierarchy_values == [None] * LAYOUT.hierarchy_cols
        and role_values[0] == "Line item"
        and role_values[1] in {"Driver", "Source / driver"}
        and role_values[2] == "Unit"
    )


def _is_hierarchy_spacer_col(ws: Worksheet, col: int) -> bool:
    width = ws.column_dimensions[get_column_letter(col)].width
    return width is not None and abs(float(width) - float(LAYOUT.hierarchy_width)) < 0.001


def _row_rule_start_col(ws: Worksheet) -> int:
    if LAYOUT.hierarchy_cols > 0:
        return LAYOUT.label_col
    return LAYOUT.label_col if _is_hierarchy_spacer_col(ws, LAYOUT.first_hierarchy_col) else LAYOUT.first_hierarchy_col


def _apply_design_surface_ws(ws: Worksheet) -> None:
    header_row_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    if True:
        default_layout = uses_default_layout(ws)
        max_col = max(ws.max_column, 9)
        max_row = max(ws.max_row, 5)
        ws.freeze_panes = None
        for row in range(1, max_row + 1):
            row_has_value = any(ws.cell(row=row, column=col).value is not None for col in range(1, max_col + 1))
            is_section = _is_section_row(ws, row)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    continue
                if is_section and col == LAYOUT.first_hierarchy_col:
                    if _fill_rgb(cell) == ib.BG_HEADER_BAND:
                        cell.font = Font(name=ib.FONT_FAMILY, size=10, bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(name=ib.FONT_FAMILY, size=10, bold=True, color=ib.BG_HEADER_BAND)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                    continue
                if row == 5 and row_has_value and col >= LAYOUT.first_hierarchy_col and not is_section:
                    cell.fill = header_row_fill
                    if col >= LAYOUT.label_col:
                        cell.border = _merge_border(cell.border, bottom=ib.THIN_LINE)
                if default_layout and col == LAYOUT.source_col and row != 5:
                    ib.apply_comment(cell, wrap_text=False)
                elif default_layout and col == LAYOUT.unit_col and row != 5:
                    _apply_unit_cell(cell)
                elif default_layout and col in (LAYOUT.first_hierarchy_col, LAYOUT.label_col) and row != 5:
                    ib.apply_label(cell, bold=cell.font.bold is True)

def _apply_value_style(cell, fmt: str) -> None:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        if "!" in value:
            ib.apply_link_intra(cell, fmt)
        else:
            ib.apply_formula(cell, fmt)
    else:
        ib.apply_hard_input(cell, fmt)


def _apply_print(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 90
        last_row, last_col = _rendered_bounds(ws)
        ib.setup_print_layout(
            ws,
            orientation="landscape",
            fit_to_width=1,
            print_title_rows="1:6" if getattr(ws, "_v2_period_sheet", False) else "1:5",
            print_title_cols=f"A:{get_column_letter(LAYOUT.unit_col)}",
            footer_right="&P / &N",
        )
        ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
        for row in range(1, last_row + 1):
            if ws.row_dimensions[row].height is None:
                ws.row_dimensions[row].height = ib.ROW_HEIGHT_BASE
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[5].height = 18


def _chart_row_series(data_ref: Reference, series_title: str | None) -> Series:
    """ONE series per charted row.

    openpyxl's ``add_data(..., titles_from_data=True)`` treats a single-ROW
    reference column-wise, emitting one broken 2-cell reversed series per
    column (e.g. ``$G$18:$F$18``). Building the Series explicitly keeps the
    value range exactly the caller's window (min→max, single grain)."""
    return Series(data_ref, title=series_title)


def _add_line_chart(ws: Worksheet, title: str, data_ref: Reference, cats_ref: Reference, anchor: str, y_axis_title: str = "", series_title: str | None = None, x_axis_title: str = "Period") -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.x_axis.delete = False  # openpyxl hides the axis (and its title) unless this is set
    chart.height = 7
    chart.width = 14
    chart.append(_chart_row_series(data_ref, series_title or title))
    chart.set_categories(cats_ref)
    ib.apply_chart_palette(chart, ib.IB_CHART_COLORS_LINE)
    ws.add_chart(chart, anchor)


def _add_bar_chart(ws: Worksheet, title: str, data_ref: Reference, cats_ref: Reference, anchor: str, y_axis_title: str = "", series_title: str | None = None, x_axis_title: str = "Period") -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.x_axis.delete = False  # openpyxl hides the axis (and its title) unless this is set
    chart.height = 7
    chart.width = 14
    chart.append(_chart_row_series(data_ref, series_title or title))
    chart.set_categories(cats_ref)
    ib.apply_chart_palette(chart, ib.IB_CHART_COLORS_BAR)
    ws.add_chart(chart, anchor)


def _scenario_driver_bucket(driver: ScenarioDriver) -> str:
    label = driver.label.lower()
    if any(token in label for token in ("financing", "debt", "lease", "grant", "warehouse", "working-capital")):
        return "financing"
    if any(token in label for token in ("cost", "loss", "incentive", "bom", "service", "prototype", "program")):
        return "cost"
    if any(token in label for token in ("cac", "sales capacity", "hiring", "headcount", "fte", "talent")):
        return "opex"
    if any(token in label for token in ("demand", "gmv", "liquidity", "take-rate", "pricing", "value capture", "utilization", "origination", "spread", "new logo", "conversion", "acv", "expansion", "retention", "churn", "deployment capacity")):
        return "revenue"
    return "opex"


# ============================================================================
# S3 framework v2 — 12-sheet architecture (slice S3-1)
# ============================================================================
# BLUEPRINT_S3.md row-level spec. The v2 sheets share one header contract
# (title row 2 / purpose row 3 / unit caption + FY ruler row 4 / months ruler
# row 5 / period header row 6 / data from row 8, freeze at (first period col,
# row 7)) and one formula contract: every per-period formula references the
# months ruler (F$5) so a single row formula is copy-identical across monthly
# and annual columns (R17).
#
# Check mechanism (documented decision): a check row/cell holds a NUMERIC
# delta that is 0 when the model is coherent (tolerance applied via ROUND /
# explicit tolerance terms using only whitelisted constants). The number
# format '"ERROR";"ERROR";"OK"' renders any non-zero delta as ERROR and zero
# as OK while the cell stays numeric, so the Summary consolidated block can
# aggregate with SUMPRODUCT(ABS(...)) and the master check is a plain SUM.

SOURCE_PLAN_SHEETS_V2 = [
    "Guide",
    "Summary",
    "Assumptions",
    "Revenue Build",
    "Cost Build",
    "People Plan",
    "P&L",
    "BS",
    "CF",
    "Financing",
    "Cap Table",
    "Evidence",
]

# Conditional sheets: outside the full default, entered via mode bundles or
# --additional-sheets (BLUEPRINT_S3 条件付きシート).
CONDITIONAL_SHEETS_V2 = [
    "Valuation & Exit",
    "IC Memo",
    "Pricing",
    "Unit Economics",
    "Segments",
]

# cap_table mode keeps its 3-sheet state-machine surface (Guide + Kernel +
# Ownership); Ownership is rendered by cap_table_builder.
CAP_TABLE_MODE_SHEETS = ["Guide", "Kernel", "Ownership"]

TAB_COLORS_V2 = {
    "Guide": ib.BRAND_SLATE,
    "Summary": ib.BRAND_NAVY,
    "Assumptions": ib.BRAND_PRIMARY_DEEP,
    "Revenue Build": ib.BRAND_PRIMARY,
    "Cost Build": ib.BRAND_PRIMARY,
    "People Plan": ib.BRAND_PRIMARY,
    "P&L": ib.BRAND_SLATE,
    "BS": ib.BRAND_SLATE,
    "CF": ib.BRAND_SLATE,
    "Financing": ib.BRAND_WARNING,
    "Cap Table": ib.BRAND_WARNING,
    "Evidence": ib.BRAND_SLATE,
    "Valuation & Exit": ib.BRAND_WARNING,
    "IC Memo": ib.BRAND_ACCENT,
    "Pricing": ib.BRAND_PRIMARY,
    "Unit Economics": ib.BRAND_PRIMARY,
    "Segments": ib.BRAND_PRIMARY,
    "Kernel": ib.BRAND_PRIMARY_DEEP,
}

V2_FIRST_PERIOD_COL = 6  # column F — canonical across every period-axis sheet
FMT_CHECK_V2 = '"ERROR";"ERROR";"OK"'
FMT_COUNT_V2 = '#,##0;[Red]"▲"#,##0;"-"'

_FONT_TITLE_V2 = Font(name=ib.FONT_FAMILY, size=14, bold=True, color=ib.IB_INK)
_FONT_SNAPSHOT_V2 = Font(name=ib.FONT_FAMILY, size=ib.FONT_SIZE_BASE, italic=True, color=ib.IB_INK)


@dataclass
class CheckEntry:
    """One registered check: `ref` is a sheet-local cell ('F40') or range
    ('F40:AF40') whose numeric value(s) are 0 when the check passes."""

    sheet: str
    ref: str
    description: str


@dataclass
class BuildContext:
    """Shared state for one v2 workbook build."""

    facts: SourceFacts
    axis: PeriodAxis
    bundle: set
    sheet_scale: dict = field(default_factory=dict)
    checks: list = field(default_factory=list)
    rows: dict = field(default_factory=dict)
    snapshots: dict = field(default_factory=dict)
    master_check_cell: str | None = None
    # Driver rows to omit from the Assumptions register (no-dead-driver-rows
    # rule: pass 2 of the build suppresses any pass-1 row with no consumer).
    suppress_assumptions: set = field(default_factory=set)

    @property
    def currency(self) -> str:
        return self.facts.currency

    @property
    def n_cols(self) -> int:
        return len(self.axis.labels)

    @property
    def note_col(self) -> int:
        return V2_FIRST_PERIOD_COL + self.n_cols

    @property
    def period_cols(self) -> list:
        return list(range(V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + self.n_cols))

    @property
    def last_period_letter(self) -> str:
        return get_column_letter(self.period_cols[-1])


def _v2_ref(ctx: BuildContext, sheet: str, cell: str) -> str:
    """Bundle-aware cross-sheet reference: S3-1 sheets only reference sheets
    that exist in the bundle; anything else is a build-time error."""
    if sheet not in ctx.bundle:
        raise KeyError(
            f"bundle-aware reference to {sheet!r} outside bundle "
            f"{sorted(ctx.bundle)} — builders must fall back per blueprint"
        )
    return f"'{sheet}'!{cell}"


def _v2_any(series) -> bool:
    return any(bool(v) for v in (series or []))


def _v2_money_fmt_unit(ctx: BuildContext, scale: str) -> tuple:
    if ctx.currency == "USD":
        return ib.FMT_USD_BY_SCALE[scale], ib.USD_UNIT_BY_SCALE[scale]
    return ib.fmt_jp_for_scale(scale), ib.JP_UNIT_BY_SCALE[scale]


def _v2_pct_fmt(ctx: BuildContext) -> str:
    return ib.FMT_PERCENT if ctx.currency == "USD" else ib.FMT_JP_PERCENT


# Absolute materiality floor for the digit-crush guard: money values below
# this (working-capital / tax-timing rounding noise) do not force the sheet
# scale finer, but every value at or above it must render as a nonzero figure.
# Scaled by the currency's base unit so USD ($100) and JPY (¥10,000) match.
_CRUSH_MATERIALITY_FLOOR_BASE = 10_000.0


def _v2_guard_scale_readable(scale: str, values, currency: str,
                             materiality_floor: float = 0.0) -> str:
    """Digit-crush guard: keep the largest scale where (a) the max |value|
    renders ≥ 3 display units AND (b) the minimum MATERIAL nonzero |value|
    renders ≥ 1 display unit; otherwise step down the ladder. Prevents a
    sheet/row scale where a real figure (an early-month cost, a working-capital
    delta) displays as "0" / "▲0" while larger periods dominate the pick.

    `materiality_floor` (absolute, base-currency) excludes sub-floor rounding /
    tax-timing noise from the min so it cannot force an absurdly fine scale;
    0 keeps the strict per-row behavior used by register / per-unit rows."""
    ladder = list(ib.JP_SCALE_LADDER if currency != "USD" else ib.USD_SCALE_LADDER)
    finite = [abs(float(v)) for v in (values or [])
              if isinstance(v, (int, float)) and not isinstance(v, bool)
              and math.isfinite(float(v))]
    nonzero = [v for v in finite if v > 0]
    if not nonzero or scale not in ladder:
        return scale
    max_abs = max(nonzero)
    material = [v for v in nonzero if v >= materiality_floor] or nonzero
    min_material = min(material)
    idx = ladder.index(scale)
    while idx > 0:
        divisor = ib._SCALE_DIVISOR[ladder[idx]]
        if max_abs >= ib.SCALE_HYSTERESIS_FACTOR * divisor and min_material >= divisor:
            break
        idx -= 1
    return ladder[idx]


def _v2_row_money(ctx: BuildContext, representative_values) -> tuple:
    """Row-scale money format/unit for register + per-unit rows (2-layer rule b)."""
    values = [float(v) for v in (representative_values or []) if isinstance(v, (int, float))]
    scale = ib.pick_row_scale(values, currency=ctx.currency)
    scale = _v2_guard_scale_readable(scale, values, ctx.currency)
    return _v2_money_fmt_unit(ctx, scale)


def _v2_fmt_for(ctx: BuildContext, ukind: str, *, sheet_scale: str | None = None, rep_values=None) -> tuple:
    """Resolve (number_format, unit label) for a semantic unit kind."""
    if ukind == "money":  # engine/statement row on the sheet-wide scale
        return _v2_money_fmt_unit(ctx, sheet_scale or "million")
    if ukind == "money_row":  # register / per-unit row on its own row scale
        return _v2_row_money(ctx, rep_values)
    if ukind == "pct":
        return _v2_pct_fmt(ctx), "%"
    if ukind == "x":
        return ib.FMT_MULTIPLE, "x"
    if ukind == "factor":
        return ib.FMT_FACTOR, "x"
    if ukind == "fte":
        return FMT_COUNT_V2, "FTE"
    if ukind == "units":
        return FMT_COUNT_V2, "units"
    if ukind == "customers":
        return FMT_COUNT_V2, "customers"
    if ukind == "count":
        return FMT_COUNT_V2, "count"
    if ukind == "months":
        return ib.FMT_MONTHS_1DP, "months"
    if ukind == "int":
        return "0", ""
    if ukind == "check":
        return FMT_CHECK_V2, "check"
    if ukind == "status":
        return "General", "status"
    return "General", ""


# Grains whose facts are PER-PERIOD canonical (one value per column, S1): the
# kernel already produces the per-month / per-quarter series, so these pass
# through the expansion helpers unchanged. Annual / hybrid facts are
# annual-canonical (5 fiscal years) and get expanded onto the build axis.
_PER_PERIOD_GRAINS = ("monthly", "quarterly")


def _v2_expand(ctx: BuildContext, annual_series, kind: str) -> list:
    """Expand a facts series onto the build axis.

    Facts are annual-canonical for annual/hybrid grains and PER-PERIOD
    canonical for monthly / quarterly grains (S1), so those pass through."""
    if ctx.axis.grain in _PER_PERIOD_GRAINS:
        return list(annual_series)
    return expand_annual_series(list(annual_series), ctx.axis, kind)


def _v2_expand_event(ctx: BuildContext, annual_series) -> list:
    """Expand a FINANCING series as events: the whole FY amount lands in the
    fiscal year's FIRST monthly column (期首調達仮定); annual columns pass
    through. This keeps financing cash from smearing across months the way a
    flow interpolation would."""
    if ctx.axis.grain in _PER_PERIOD_GRAINS:
        return list(annual_series)
    axis = ctx.axis
    values = list(annual_series)
    fy_order: list[str] = []
    for fy in axis.fy_labels:
        if fy not in fy_order:
            fy_order.append(fy)
    amount_by_fy = {fy: values[idx] for idx, fy in enumerate(fy_order)}
    out: list = []
    seen: set = set()
    for idx, months in enumerate(axis.months_in_period):
        fy = axis.fy_labels[idx]
        if months == 12:
            out.append(amount_by_fy.get(fy, 0))
        elif fy not in seen:
            out.append(amount_by_fy.get(fy, 0))
            seen.add(fy)
        else:
            out.append(0)
    return out


def _v2_annual_rollup_axis(axis: PeriodAxis) -> PeriodAxis:
    """Collapse any axis to one column per fiscal year (Summary roll-up)."""
    labels: list = []
    months: list = []
    ends: list = []
    for idx, fy in enumerate(axis.fy_labels):
        if labels and labels[-1] == fy:
            months[-1] += axis.months_in_period[idx]
            ends[-1] = axis.period_end[idx]
        else:
            labels.append(fy)
            months.append(axis.months_in_period[idx])
            ends.append(axis.period_end[idx])
    return PeriodAxis(
        labels=list(labels),
        months_in_period=months,
        fy_labels=list(labels),
        period_end=ends,
        monthly_count=0,
        grain="annual",
    )


def _v2_fy_flow_rollup(ctx: BuildContext, series) -> list:
    """Aggregate a per-period kernel series (facts grain) to one FY total.

    Only per-period grains (monthly / quarterly) need aggregating; annual and
    hybrid facts are already one value per fiscal year."""
    if ctx.facts.grain not in _PER_PERIOD_GRAINS:
        return list(series)
    out: list = []
    fy_seen: list = []
    axis = build_period_axis(ctx.facts)
    for idx, fy in enumerate(axis.fy_labels[: len(series)]):
        if fy_seen and fy_seen[-1] == fy:
            out[-1] += series[idx]
        else:
            fy_seen.append(fy)
            out.append(series[idx])
    return out


def _v2_pick_sheet_scale(ctx: BuildContext, aggregate_annual_series,
                         detail_annual_series=None) -> str:
    """Sheet-wide scale (2-layer rule a): dominant magnitude of the sheet's
    aggregate money rows, measured over the monthly window when one exists
    (ARCHITECTURE.md §2 — hysteresis handled by ib.pick_sheet_scale).

    Digit-crush guard: the pick then steps down while the smallest MATERIAL
    per-period value — across the aggregate rows AND any `detail_annual_series`
    (the sheet's small line items, e.g. per-unit cost components or
    working-capital deltas), measured over the monthly window so early-month
    figures count — would render below one display unit. This keeps a monthly
    statement whose detail rows are sub-million from crushing them to "0"
    while its aggregates sit comfortably in 百万円. A base-currency materiality
    floor drops tax-timing / rounding noise so it cannot over-shrink the scale.
    """
    max_values = []
    for series in aggregate_annual_series:
        expanded = _v2_expand(ctx, series, "flow")
        window = expanded[: ctx.axis.monthly_count] if ctx.axis.monthly_count else expanded
        max_values.extend(float(v) for v in window)
    scale = ib.pick_sheet_scale(max_values, currency=ctx.currency)
    guard_values = list(max_values)
    for series in (detail_annual_series or []):
        expanded = _v2_expand(ctx, series, "flow")
        window = expanded[: ctx.axis.monthly_count] if ctx.axis.monthly_count else expanded
        guard_values.extend(float(v) for v in window)
    # Floor in the stored currency's own units: ¥10,000 for JPY, ~$100 for USD.
    floor = _CRUSH_MATERIALITY_FLOOR_BASE / 100.0 if ctx.currency == "USD" else _CRUSH_MATERIALITY_FLOOR_BASE
    return _v2_guard_scale_readable(scale, guard_values, ctx.currency,
                                    materiality_floor=floor)


def _v2_caption(ctx: BuildContext, unit_label: str | None) -> str:
    if unit_label is None:  # register sheet — per-row scale
        return "(Unit: see unit column)" if ctx.currency == "USD" else "(単位: 単位列参照)"
    if ctx.currency == "USD":
        return f"(Unit: {unit_label})"
    return f"(単位: {unit_label})"


_FONT_CAPTION_V2 = Font(
    name=ib.FONT_FAMILY, size=ib.FONT_SIZE_SMALL, italic=True, color=ib.IB_COMMENT
)


def _v2_register_caption(ws: Worksheet, caption: str) -> None:
    """C4 unit caption for register sheets that skip the period-ruler header
    (same cell / style contract as write_period_rulers on period sheets)."""
    cell = ws.cell(ib.HEADER_FY_RULER_ROW, ib.HEADER_CAPTION_COL, caption)
    cell.font = _FONT_CAPTION_V2
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _setup_sheet_v2(
    ws: Worksheet,
    ctx: BuildContext,
    title: str,
    purpose: str,
    *,
    period_axis: bool = True,
    unit_caption: str | None = None,
    wide_source: bool = False,
    note_header: str = "Notes",
    axis: PeriodAxis | None = None,
) -> None:
    """v2 header contract (BLUEPRINT_S3 共通ヘッダ契約)."""
    axis = axis or ctx.axis
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS_V2.get(ws.title, ib.BRAND_SLATE)
    ws._startup_facts = ctx.facts
    ws._v2_sheet = True
    ws.column_dimensions["A"].width = ib.COL_GUTTER_WIDTH_V2
    ib.apply_indent_column_widths(ws, [2])
    ws.column_dimensions["C"].width = ib.COL_LABEL_WIDTH_V2
    ws.column_dimensions["D"].width = (
        ib.COL_SOURCE_WIDTH_INPUT if wide_source else ib.COL_DRIVER_TAG_WIDTH
    )
    ws.column_dimensions["E"].width = ib.COL_UNIT_WIDTH_V2
    title_cell = ws.cell(ib.HEADER_TITLE_ROW_V2, ib.HEADER_CAPTION_COL, title)
    title_cell.font = _FONT_TITLE_V2
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    purpose_cell = ws.cell(ib.HEADER_PURPOSE_ROW_V2, ib.HEADER_CAPTION_COL, purpose)
    ib.apply_comment(purpose_cell, wrap_text=False)
    ws._v2_period_sheet = False
    if not period_axis:
        return
    n = len(axis.labels)
    note_col = V2_FIRST_PERIOD_COL + n
    for col in range(V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + n):
        ws.column_dimensions[get_column_letter(col)].width = ib.COL_PERIOD_WIDTH_V2
    ws.column_dimensions[get_column_letter(note_col)].width = ib.COL_NOTE_WIDTH_V2
    ib.write_period_rulers(
        ws,
        V2_FIRST_PERIOD_COL,
        list(axis.fy_labels),
        list(axis.months_in_period),
        list(axis.labels),
        unit_caption,
    )
    ib.apply_semantic_fill_span(
        ws,
        ib.HEADER_PERIOD_ROW,
        2,
        note_col,
        ib.BG_TABLE_HEADER,
        bottom=ib.THIN_LINE,
        border_start_col=3,
    )
    for col, label, align in (
        (3, "Line item", "left"),
        (4, "Driver", "left"),
        (5, "Unit", "right"),
        (note_col, note_header, "left"),
    ):
        cell = ws.cell(ib.HEADER_PERIOD_ROW, col, label)
        cell.font = ib.FONT_BODY_BOLD
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    ib.apply_freeze_pane(ws, V2_FIRST_PERIOD_COL)
    ws._startup_note_col = note_col
    ws._v2_period_sheet = True
    ws._v2_axis = axis
    if axis.monthly_count and axis.monthly_count < n:
        ws._v2_boundary_col = V2_FIRST_PERIOD_COL + axis.monthly_count


def _v2_section(ws: Worksheet, ctx: BuildContext, row: int, label: str, *, end_col: int | None = None) -> None:
    _section(ws, row, label, end_col if end_col is not None else getattr(ws, "_startup_note_col", None))


def _v2_label_cells(
    ws: Worksheet,
    row: int,
    label: str,
    driver: str,
    unit_label: str,
    note_text: str,
    note_col: int | None,
    *,
    bold: bool = False,
) -> None:
    ws.cell(row, 3, label)
    ib.apply_label(ws.cell(row, 3), bold=bold)
    if driver:
        ws.cell(row, 4, driver)
        ib.apply_comment(ws.cell(row, 4), wrap_text=False)
    if unit_label:
        ws.cell(row, 5, unit_label)
        _apply_unit_cell(ws.cell(row, 5))
    if note_text and note_col:
        ws.cell(row, note_col, note_text)
        ib.apply_comment(ws.cell(row, note_col), wrap_text=False)


def _v2_series_row(
    ws: Worksheet,
    ctx: BuildContext,
    row: int,
    label: str,
    ukind: str,
    *,
    values=None,
    formulas=None,
    driver: str = "",
    note: str = "",
    bold: bool = False,
    band: bool = False,
    snapshot: bool = False,
    fmt: str | None = None,
    unit: str | None = None,
    cols=None,
) -> None:
    """One v2 data row: label C / driver D / unit E / period cells F.. .

    `values` (blue hard inputs or generator snapshots) and `formulas`
    (black/green formulas, one string per period column) are exclusive.
    """
    if fmt is None or unit is None:
        auto_fmt, auto_unit = _v2_fmt_for(
            ctx,
            ukind,
            sheet_scale=ctx.sheet_scale.get(ws.title),
            rep_values=values if values is not None else None,
        )
        fmt = fmt if fmt is not None else auto_fmt
        unit = unit if unit is not None else auto_unit
    note_col = getattr(ws, "_startup_note_col", None)
    _v2_label_cells(ws, row, label, driver, unit, note, note_col, bold=bold)
    cols = list(cols) if cols is not None else ctx.period_cols
    payload = values if values is not None else formulas
    for col, value in zip(cols, payload):
        if value is None:
            continue
        cell = ws.cell(row, col, value)
        if snapshot:
            cell.number_format = fmt
            cell.font = _FONT_SNAPSHOT_V2
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        elif ukind == "check":
            ib.apply_formula(cell, fmt)
        elif values is not None:
            ib.apply_hard_input(cell, fmt)
        else:
            _apply_value_style(cell, fmt)
        if bold and not snapshot:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    if band:
        _highlight_row(ws, row, cols[-1])
    elif bold:
        ib.apply_semantic_border_span(ws, row, top=ib.THIN_LINE, border_start_col=3)


def _v2_formulas(ctx: BuildContext, template: str, *, cols=None, per_col=None) -> list:
    """Expand a formula template across period columns.

    {c} = current column letter, {p} = previous column letter. For the FIRST
    period column {p} is column E (the unit-label column), so roll-forward /
    growth formulas must read priors through N(...) — text coerces to 0,
    numbers pass through — keeping one copy-identical row formula (R17). `per_col`
    optionally supplies one extra mapping per column index (dict of extra
    format kwargs)."""
    cols = list(cols) if cols is not None else ctx.period_cols
    out = []
    for idx, col in enumerate(cols):
        extra = per_col[idx] if per_col is not None else {}
        out.append(
            template.format(
                c=get_column_letter(col),
                p=get_column_letter(col - 1),
                **extra,
            )
        )
    return out


def _v2_check_row(
    ws: Worksheet,
    ctx: BuildContext,
    row: int,
    label: str,
    formulas,
    description: str,
    *,
    driver: str = "",
    note: str = "",
    cols=None,
) -> None:
    """Write a check row (0 = OK) and register it for the Summary block."""
    cols = list(cols) if cols is not None else ctx.period_cols
    _v2_series_row(
        ws, ctx, row, label, "check",
        formulas=formulas, driver=driver, note=note, cols=cols,
    )
    first = f"{get_column_letter(cols[0])}{row}"
    last = f"{get_column_letter(cols[len(formulas) - 1])}{row}"
    ref = first if len(formulas) == 1 else f"{first}:{last}"
    ctx.checks.append(CheckEntry(ws.title, ref, description))


def _v2_status_formula(cov_cell: str, ok: str = "ok", review: str = "review") -> str:
    return f'=IF({cov_cell}>=1,"{ok}","{review}")'


def _v2_mark_grain_boundary(ws: Worksheet) -> None:
    """Declare the monthly→annual boundary with a medium left rule (R the
    'declared exception' of the hybrid axis; Guide legend explains it)."""
    boundary_col = getattr(ws, "_v2_boundary_col", None)
    if not boundary_col:
        return
    last_row, _ = ib.last_value_bounds(ws)
    for row in range(ib.HEADER_TITLE_ROW_V2, last_row + 1):
        cell = ws.cell(row, boundary_col)
        cell.border = _merge_border(cell.border, left=ib.MEDIUM_LINE)


# --- scenario mechanism ------------------------------------------------------

_V2_SCENARIO_ROLES = ("demand", "price", "cost", "opex")
_V2_SCENARIO_ROLE_LABELS = {
    "demand": "Demand scale",
    "price": "Price scale",
    "cost": "Variable cost scale",
    "opex": "Opex scale",
}
_V2_SCENARIO_DEFAULTS = {
    "demand": (0.70, 1.00, 1.25, "default demand band"),
    "price": (0.85, 1.00, 1.10, "default pricing band"),
    "cost": (1.25, 1.00, 0.85, "default cost band"),
    "opex": (1.15, 1.00, 0.90, "default opex band"),
}


def _v2_scenario_roles(facts: SourceFacts) -> dict:
    """Map kernel ScenarioDrivers onto the four workbook multiplier roles.

    Buckets come from `_scenario_driver_bucket`; the first revenue-bucket
    driver takes the demand role, the second the price role. A role with no
    matching driver keeps its documented default band."""
    assigned = {}
    for drv in scenario_drivers_for(facts):
        bucket = _scenario_driver_bucket(drv)
        if bucket == "revenue":
            role = "demand" if "demand" not in assigned else ("price" if "price" not in assigned else None)
        elif bucket == "cost":
            role = "cost" if "cost" not in assigned else None
        elif bucket == "opex":
            role = "opex" if "opex" not in assigned else None
        else:  # financing capacity has no S3-1 engine hook — S3-2 (Financing)
            role = None
        if role:
            assigned[role] = (drv.downside, drv.base, drv.upside, drv.label)
    return {role: assigned.get(role, _V2_SCENARIO_DEFAULTS[role]) for role in _V2_SCENARIO_ROLES}


def _v2_case_projection(facts: SourceFacts, d: float, p: float, c: float, o: float) -> dict:
    """Re-derive the kernel projection under one scenario case.

    Mirrors the sheet mechanics: demand scale on units/GMV/customers, price
    scale on price (and take rate via revenue math), variable-cost scale on
    the whole COGS stack (gm' = 1 − (1 − gm)·c), opex scale on the S&M / R&D
    / G&A programs. People comp and headcount stay plan-level, exactly like
    the live sheets under the toggle."""
    periods = len(facts.years)
    new_units = [int(round(u * d)) for u in facts.new_units]
    gmv = [int(round(g * d)) for g in facts.gmv_yen]
    price = [int(round(v * p)) for v in facts.monthly_price_yen]
    customers = [int(round(v * d)) for v in facts.customers]
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(customers)
    revenue = plan_revenue_series(
        new_units,
        gmv,
        price,
        list(facts.take_rate),
        list(facts.other_revenue_share),
        facts.revenue_mode,
        installed_base=customers if pinned else None,
        churn_rate=facts.churn_rate if pinned else None,
        onboarding_months=list(facts.onboarding_months) or None,
        one_time_revenue_per_unit_yen=list(facts.one_time_revenue_per_unit_yen) or None,
        months_per_period=_months_factor(facts.grain),
    )
    tgm = [1.0 - (1.0 - g) * c for g in facts.target_gross_margin]
    total_hc = [
        facts.product_headcount[i] + facts.gtm_headcount[i]
        + facts.operations_headcount[i] + facts.ga_headcount[i]
        for i in range(periods)
    ]
    capex = [
        new_units[i] * facts.capex_per_unit_yen[i] + facts.other_capex_yen[i]
        for i in range(periods)
    ]
    projection = project_free_cash_flow(
        revenue,
        tgm,
        total_hc,
        list(facts.avg_comp_yen),
        list(facts.product_headcount),
        [v * o for v in facts.sm_pct_revenue],
        [int(round(v * o)) for v in facts.rd_program_per_product_fte_yen],
        [int(round(v * o)) for v in facts.rd_program_floor_yen],
        [v * o for v in facts.ga_pct_revenue],
        [int(round(v * o)) for v in facts.fixed_ga_yen],
        capex,
        list(facts.depreciation_life_months),
        list(facts.debt_raise_yen),
        list(facts.debt_interest_rate),
        list(facts.ar_days),
        list(facts.ap_days),
        list(facts.deferred_revenue_share),
        list(facts.inventory_wip_pct_capex),
        list(facts.tax_rate),
        convertibles_yen=list(facts.convertibles_yen),
        lease_financing_yen=list(facts.lease_financing_yen),
        customer_advances_yen=list(facts.customer_advances_yen),
        debt_amortization_yen=list(facts.debt_amortization_yen),
        months_per_period=_months_factor(facts.grain),
    )
    amortization = (list(facts.debt_amortization_yen) + [0] * periods)[:periods]
    cash = float(facts.beginning_cash_yen)
    ending_cash = []
    for idx, period in enumerate(projection):
        cash += (
            period["free_cash_flow"]
            + facts.debt_raise_yen[idx]
            + facts.equity_raise_yen[idx]
            + facts.convertibles_yen[idx]
            + facts.lease_financing_yen[idx]
            + facts.grants_yen[idx]
            - facts.secondary_yen[idx]
            - amortization[idx]
        )
        ending_cash.append(cash)
    return {
        "revenue": revenue,
        "ebitda": [period["ebitda"] for period in projection],
        "ending_cash": ending_cash,
    }


def _v2_base_revenue_annual(facts: SourceFacts) -> list:
    """Annual-canonical total revenue, exactly as the kernel plans it."""
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(facts.customers)
    return plan_revenue_series(
        list(facts.new_units),
        list(facts.gmv_yen),
        list(facts.monthly_price_yen),
        list(facts.take_rate),
        list(facts.other_revenue_share),
        facts.revenue_mode,
        installed_base=list(facts.customers) if pinned else None,
        churn_rate=facts.churn_rate if pinned else None,
        onboarding_months=list(facts.onboarding_months) or None,
        one_time_revenue_per_unit_yen=list(facts.one_time_revenue_per_unit_yen) or None,
        months_per_period=_months_factor(facts.grain),
    )


# --- v2 sheet builders -------------------------------------------------------

_V2_OWNER_MAP = {
    "Assumptions / Revenue Build": "Assumptions / Revenue Build",
    "Cost Build": "Cost Build",
    "People Plan / BS / CF": "People Plan",
    "Capital Stack / Ownership": "Financing / Cap Table",
    "Scenarios / Sensitivity": "Assumptions (scenario block) / Summary",
    "Valuation / IC Memo": "Summary / Valuation & Exit",
}


def _v2_owner_for_bundle(owner: str, bundle: set) -> str:
    """Driver-map owner readout restricted to sheets that exist in the
    bundle — the register must never point the reader at an omitted sheet."""
    kept = []
    for part in (p.strip() for p in owner.split("/")):
        base = part.split(" (")[0].strip()
        if base in bundle:
            kept.append(part)
    return " / ".join(kept) if kept else "Assumptions (register)"


def _build_assumptions_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Assumptions"]
    _setup_sheet_v2(
        ws,
        ctx,
        f"{facts.company} — Assumptions",
        "Driver register: blue cells are the editable input layer; engine sheets reference these rows only.",
        unit_caption=_v2_caption(ctx, None),
        wide_source=True,
        note_header="Evidence status",
    )
    R: dict = {}
    ctx.rows["Assumptions"] = R
    periods = len(facts.years)
    mk = mechanic_key(facts)
    unit_sale = facts.revenue_mode == "unit_sale"
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(facts.customers)
    status_default = str(_compact_status(facts.evidence_status))
    r = ib.DATA_START_ROW_V2

    # Sections / gaps are written lazily so a section whose rows are all
    # suppressed (no-dead-driver-rows pass 2) leaves no orphan header.
    pending = {"section": None, "gap": False}

    def section(label: str) -> None:
        pending["section"] = label

    def gap() -> None:
        pending["gap"] = True

    def flush_pending() -> None:
        nonlocal r
        if pending["gap"]:
            r += 1
            pending["gap"] = False
        if pending["section"]:
            _v2_section(ws, ctx, r, pending["section"])
            pending["section"] = None
            r += 1

    def put(key, label, ukind, series, expand, driver, *, status=None, note="",
            formulas=None, fmt=None, unit=None, cols=None, values_override=None):
        nonlocal r
        if key in ctx.suppress_assumptions:
            return
        flush_pending()
        if values_override is not None:
            values = list(values_override)
        elif formulas is None:
            values = (
                _v2_expand_event(ctx, series) if expand == "event"
                else _v2_expand(ctx, series, expand)
            )
        else:
            values = None
        note_text = "; ".join(t for t in ((status or status_default), note) if t)
        _v2_series_row(
            ws, ctx, r, label, ukind, values=values, formulas=formulas,
            driver=driver, note=note_text, fmt=fmt, unit=unit, cols=cols,
        )
        R[key] = r
        r += 1

    # --- Demand ---------------------------------------------------------
    # B2 units roll-forward (pinned base): the blue new-units series is the
    # exact per-column inversion of the stated ending base WITH the implied
    # churn — the Revenue Build ending row rolls
    # N(prior) + new − N(prior) × churn × months/12 back onto the stated
    # series (FY ends land exactly; the check row on Revenue Build proves it).
    rollforward_pinned = (
        pinned and (not unit_sale) and _v2_any(facts.new_units)
        and _v2_any(facts.customers) and facts.grain not in _PER_PERIOD_GRAINS
    )
    churn_implied: list = []
    new_units_override = None
    if rollforward_pinned:
        prior_end = 0
        for i, end in enumerate(facts.customers):
            new = facts.new_units[i] if i < len(facts.new_units) else 0
            if prior_end > 0:
                churn_implied.append(max(0.0, (prior_end + new - end) / prior_end))
            else:
                churn_implied.append(None)
            prior_end = max(0, int(end))
        first_known = next((v for v in churn_implied if v is not None),
                           float(facts.churn_rate or 0.0))
        churn_implied = [first_known if v is None else v for v in churn_implied]
        # Per-column inversion against the interpolated stated stock, so the
        # roll-forward tracks the stated base INSIDE each FY too (keeps the
        # revenue bridge to the kernel plan intact on hybrid axes).
        stock_cols = _v2_expand(ctx, facts.customers, "stock")
        churn_cols = _v2_expand(ctx, churn_implied, "rate")
        new_units_override = []
        prior_f = 0.0
        for i, months in enumerate(ctx.axis.months_in_period):
            end_f = float(stock_cols[i])
            new_units_override.append(
                max(0.0, end_f - prior_f * (1.0 - float(churn_cols[i]) * months / 12.0)))
            prior_f = end_f
    section("Demand")
    if _v2_any(facts.new_units):
        put("new_units", "New primary units", "units", facts.new_units, "flow",
            "demand ramp", status="stated input" if pinned else None,
            values_override=new_units_override)
    if _v2_any(facts.customers):
        put("customers", "Total customers (ending)", "customers", facts.customers,
            "stock", "installed base", status="stated input" if pinned else None)
    pool = [
        max(int(facts.new_units[i] / max(facts.utilization_conversion[i], 0.01)), facts.customers[i])
        for i in range(periods)
    ]
    if _v2_any(pool) and _v2_any(facts.utilization_conversion):
        put("pool", "Qualified demand pool", "count", pool, "flow", "funnel / addressable accounts")
        put("conversion", "Demand conversion to units", "pct", facts.utilization_conversion,
            "rate", "conversion evidence")
    if (not pinned) and (not unit_sale) and _v2_any(facts.monthly_price_yen):
        churn_schedule = [0.02 + i * 0.005 for i in range(periods)]
        put("churn", "Churn rate (annual)", "pct", churn_schedule, "rate",
            "fleet roll-forward schedule", status="model default",
            note="ramp schedule mirrors the kernel roll-forward")
    elif rollforward_pinned and "new_units" in R and "customers" in R:
        put("churn", "Churn rate (annual)", "pct", churn_implied, "rate",
            "fleet roll-forward (implied by stated base)", status="derived",
            note="inverted from the stated customer base — drives the units roll-forward")
    if (not unit_sale) and _v2_any(facts.net_retention):
        put("net_retention", "Net retention", "pct", facts.net_retention, "rate", "cohort behavior")

    # --- Monetization -----------------------------------------------------
    gap()
    section("Monetization")
    if _v2_any(facts.monthly_price_yen):
        put("price", "Unit sale price" if unit_sale else "Monthly price / unit",
            "money_row", facts.monthly_price_yen, "rate",
            "one-time sale price" if unit_sale else "pricing anchor")
    one_time_stated = bool(facts.one_time_revenue_per_unit_yen)
    if "price" in R:
        price_row = R["price"]
        if unit_sale:
            fee_fmt, fee_unit = _v2_row_money(ctx, facts.monthly_price_yen)
            put("one_time_fee", "One-time revenue / new unit", "money_row", None, None,
                "one-time sale price",
                formulas=_v2_formulas(ctx, f"={{c}}{price_row}"), fmt=fee_fmt, unit=fee_unit)
        elif one_time_stated:
            put("one_time_fee", "One-time revenue / new unit", "money_row",
                facts.one_time_revenue_per_unit_yen, "rate",
                "stated one-time / onboarding fee", status="stated input")
        elif getattr(facts, "onboarding_pinned", False):
            onboarding = list(facts.onboarding_months) or [3.0] * periods
            months = _v2_expand(ctx, [float(m) for m in onboarding], "rate")
            fee_fmt, fee_unit = _v2_row_money(
                ctx, [facts.monthly_price_yen[i] * onboarding[min(i, len(onboarding) - 1)] for i in range(periods)]
            )
            put("one_time_fee", "One-time revenue / new unit", "money_row", None, None,
                "stated onboarding months", status="stated input",
                formulas=[
                    f"={get_column_letter(col)}{price_row}*{months[idx]:g}"
                    for idx, col in enumerate(ctx.period_cols)
                ],
                fmt=fee_fmt, unit=fee_unit)
        else:
            fee_fmt, fee_unit = _v2_row_money(ctx, [v * 3 for v in facts.monthly_price_yen])
            put("one_time_fee", "One-time revenue / new unit", "money_row", None, None,
                "placeholder: 3-month onboarding default", status="placeholder",
                formulas=_v2_formulas(ctx, f"={{c}}{price_row}*3"), fmt=fee_fmt, unit=fee_unit)
    if mk == "marketplace" and _v2_any(facts.gmv_yen):
        put("gmv", "Gross merchandise value", "money_row", facts.gmv_yen, "flow",
            "GMV / economic volume")
        if _v2_any(facts.take_rate):
            put("take", "Take rate", "pct", facts.take_rate, "rate", "transaction monetization")
    if _v2_any(facts.other_revenue_share):
        put("other_share", "Other revenue / total revenue", "pct", facts.other_revenue_share,
            "rate", "services / add-ons")
    if ("price" in R or "gmv" in R) and facts.customer_roi_yen:
        put("roi", "Customer annual value / ROI", "money_row",
            [facts.customer_roi_yen] * periods, "rate", "customer value proof")
        if _v2_any(facts.value_capture_share):
            put("value_capture", "Value capture share", "pct", facts.value_capture_share,
                "rate", "pricing power evidence")

    # --- Cost-to-serve ----------------------------------------------------
    gap()
    section("Cost-to-serve drivers")
    if _v2_any(facts.variable_cogs_pct):
        put("vc_pct", "Variable COGS / revenue", "pct", facts.variable_cogs_pct, "rate",
            "cost-to-serve curve")
    if _v2_any(facts.delivery_cost_yen):
        put("delivery", "Delivery cost / unit / month", "money_row", facts.delivery_cost_yen,
            "rate", "implementation / service")
    if _v2_any(facts.cloud_cost_yen):
        put("cloud", "Cloud / platform cost / unit / month", "money_row", facts.cloud_cost_yen,
            "rate", "infrastructure")
    if _v2_any(facts.support_cost_yen):
        put("support", "Support cost / customer / month", "money_row", facts.support_cost_yen,
            "rate", "support operations")
    if _v2_any(facts.target_gross_margin):
        put("tgm", "Target gross margin", "pct", facts.target_gross_margin, "rate",
            "margin policy / benchmark")
    if _v2_any(facts.support_tickets_per_customer):
        put("tickets", "Support tickets / customer / year", "count",
            facts.support_tickets_per_customer, "rate", "support workload")
    if _v2_any(facts.support_fte_capacity_tickets):
        put("ticket_capacity", "Support FTE ticket capacity / year", "count",
            facts.support_fte_capacity_tickets, "rate", "support capacity")

    # --- People -----------------------------------------------------------
    gap()
    section("People & compensation")
    welfare = float(getattr(facts, "statutory_welfare_rate", 0.0) or 0.0)
    if welfare > 0:
        base_salary = [int(round(v / (1.0 + welfare))) for v in facts.avg_comp_yen]
        put("base_salary", "Avg base salary / FTE (annual)", "money_row", base_salary,
            "rate", "talent cost", note="base salary before statutory welfare")
        put("welfare_rate", "Statutory welfare rate", "pct", [welfare] * periods, "rate",
            "statutory welfare (法定福利費)", status="stated input")
    else:
        put("loaded_comp", "Avg loaded comp / FTE (annual)", "money_row", facts.avg_comp_yen,
            "rate", "talent cost", note="treated as fully loaded (incl. statutory welfare)")
    if _v2_any(facts.sm_pct_revenue):
        put("sm_pct", "S&M programs / revenue", "pct", facts.sm_pct_revenue, "rate",
            "go-to-market spend policy")
    if _v2_any(facts.rd_program_per_product_fte_yen):
        put("rd_per_fte", "R&D program / product FTE (annual)", "money_row",
            facts.rd_program_per_product_fte_yen, "rate", "product roadmap spend")
    if _v2_any(facts.rd_program_floor_yen):
        put("rd_floor", "R&D program floor (annual)", "money_row", facts.rd_program_floor_yen,
            "rate", "minimum roadmap spend")
    if _v2_any(facts.ga_pct_revenue):
        put("ga_pct", "G&A programs / revenue", "pct", facts.ga_pct_revenue, "rate",
            "company infrastructure")
    if _v2_any(facts.fixed_ga_yen):
        put("fixed_ga", "Fixed G&A / systems (annual)", "money_row", facts.fixed_ga_yen,
            "rate", "systems and admin base")
    if "People Plan" not in ctx.bundle and ({"P&L", "CF", "Valuation & Exit"} & ctx.bundle):
        # Compact bundles (three_statement / burn_runway / ma_exit /
        # dcf_only): the statement sheets and the Valuation & Exit compact
        # engine read headcount straight off this register instead of a
        # People Plan engine sheet.
        total_hc = [
            facts.product_headcount[i] + facts.gtm_headcount[i]
            + facts.operations_headcount[i] + facts.ga_headcount[i]
            for i in range(periods)
        ]
        put("total_hc", "Total headcount (plan)", "fte", total_hc, "stock",
            "hiring plan", note="compact bundle — People Plan sheet not included")
        if _v2_any(facts.product_headcount):
            put("product_hc", "Product / R&D FTE (plan)", "fte",
                facts.product_headcount, "stock", "R&D program driver")

    # --- Capital & assets ---------------------------------------------------
    gap()
    section("Capital & assets")
    if _v2_any(facts.capex_per_unit_yen):
        put("capex_unit", "CapEx / primary unit", "money_row", facts.capex_per_unit_yen,
            "rate", "asset / setup investment")
    if _v2_any(facts.other_capex_yen):
        put("other_capex", "Other CapEx (annual)", "money_row", facts.other_capex_yen,
            "rate", "labs / systems / tooling")
    if _v2_any(facts.depreciation_life_months):
        put("dep_life", "Depreciation life", "count", facts.depreciation_life_months,
            "rate", "asset policy", fmt=FMT_COUNT_V2, unit="months")

    # --- Working capital & tax ----------------------------------------------
    gap()
    section("Working capital & tax")
    if ctx.currency != "USD":
        put("consumption_tax", "Consumption tax rate", "pct",
            [facts.consumption_tax_rate] * periods, "rate", "消費税",
            note="tax-inclusive CF uses a balance-method accrual")
    ar_site = float(getattr(facts, "ar_site_months", 0.0) or 0.0)
    ap_site = float(getattr(facts, "ap_site_months", 0.0) or 0.0)
    if ar_site > 0:
        put("ar_site", "AR collection site", "months", [ar_site] * periods, "rate",
            "回収サイト", status="stated input")
    elif _v2_any(facts.ar_days):
        put("ar_days", "AR days", "count", facts.ar_days, "rate", "collection terms",
            fmt=FMT_COUNT_V2, unit="days")
    if ap_site > 0:
        put("ap_site", "AP payment site", "months", [ap_site] * periods, "rate",
            "支払サイト", status="stated input")
    elif _v2_any(facts.ap_days):
        put("ap_days", "AP days", "count", facts.ap_days, "rate", "supplier terms",
            fmt=FMT_COUNT_V2, unit="days")
    if _v2_any(facts.deferred_revenue_share):
        put("deferred_share", "Deferred revenue / revenue", "pct",
            facts.deferred_revenue_share, "rate", "billing terms / prepayment")
    if _v2_any(facts.inventory_wip_pct_capex):
        put("inv_wip", "Inventory / WIP share of CapEx", "pct",
            facts.inventory_wip_pct_capex, "rate", "build pipeline")
    if _v2_any(facts.tax_rate):
        put("tax_rate", "Corporate tax rate", "pct", facts.tax_rate, "rate", "NOL / tax timing")
    if "fye" not in ctx.suppress_assumptions:
        flush_pending()
        _v2_series_row(ws, ctx, r, "Fiscal year end month", "int",
                       values=[int(facts.fiscal_year_end_month)], cols=[V2_FIRST_PERIOD_COL],
                       driver="決算月 (info)", unit="month",
                       note=f"{status_default}; FY labels end in month {facts.fiscal_year_end_month}")
        R["fye"] = r
        r += 1
    if "beginning_cash" not in ctx.suppress_assumptions:
        flush_pending()
        _v2_series_row(ws, ctx, r, "Beginning cash (model start)", "money_row",
                       values=[facts.beginning_cash_yen], cols=[V2_FIRST_PERIOD_COL],
                       driver="opening balance",
                       note=f"{status_default}; single opening value — first period column only")
        R["beginning_cash"] = r
        r += 1

    # --- Financing policy -----------------------------------------------------
    gap()
    section("Financing policy")
    for key, label, series in (
        ("equity_raise", "Equity financing", facts.equity_raise_yen),
        ("debt_raise", "Debt financing", facts.debt_raise_yen),
        ("grants", "Grants / subsidies", facts.grants_yen),
        ("convertibles", "Convertible instruments", facts.convertibles_yen),
        ("lease", "Lease financing", facts.lease_financing_yen),
        ("advances", "Customer advances", facts.customer_advances_yen),
        ("secondary", "Founder / investor secondary", facts.secondary_yen),
        ("amortization", "Debt amortization", facts.debt_amortization_yen),
    ):
        if _v2_any(series):
            put(key, label, "money_row", series, "event", "funding plan",
                note="期首調達仮定 — 各会計年度の調達額はそのFYの第1月列に一括計上")
    if _v2_any(facts.debt_interest_rate):
        put("debt_rate", "Debt interest rate (annual)", "pct", facts.debt_interest_rate,
            "rate", "debt terms")
    if _v2_any(facts.target_min_runway_months):
        put("min_runway", "Target minimum runway", "months",
            [float(v) for v in facts.target_min_runway_months], "rate", "financing policy")

    # --- Driver map (compact ex-Driver Tree) ----------------------------------
    gap()
    section("Driver map")
    flush_pending()
    note_col = ws._startup_note_col
    for surface in driver_surfaces_for(facts):
        owner = _v2_owner_for_bundle(
            _V2_OWNER_MAP.get(surface.workbook_owner, "Summary"), ctx.bundle)
        _v2_label_cells(ws, r, surface.driver, f"{surface.layer} → {owner}", "",
                        surface.decision_relevance, note_col)
        r += 1

    # --- Scenario settings ------------------------------------------------------
    gap()
    section("Scenario settings")
    flush_pending()
    ib.apply_semantic_fill_span(ws, r, 2, 8, ib.BG_TABLE_HEADER, bottom=ib.THIN_LINE,
                                border_start_col=3)
    case_label = ws.cell(r, 3, "Case")
    case_label.font = ib.FONT_BODY_BOLD
    case_label.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for col, label in zip((6, 7, 8), ("Downside", "Base", "Upside")):
        ib.apply_year_header(ws.cell(r, col, label), label)
    R["case_header"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Scenario toggle", "int", values=[2],
                   driver="1=Downside / 2=Base / 3=Upside",
                   note="blue input — drives every Effective scale row below",
                   cols=[V2_FIRST_PERIOD_COL], unit="case")
    R["scenario_toggle"] = r
    toggle_row = r
    r += 1
    roles = _v2_scenario_roles(facts)
    for role in _V2_SCENARIO_ROLES:
        down, base_v, up, source_label = roles[role]
        _v2_series_row(ws, ctx, r, _V2_SCENARIO_ROLE_LABELS[role], "factor",
                       values=[down, base_v, up], driver=source_label, cols=[6, 7, 8])
        R[f"case_{role}"] = r
        r += 1
    for role in _V2_SCENARIO_ROLES:
        case_row = R[f"case_{role}"]
        _v2_series_row(ws, ctx, r, f"Effective {role} scale", "factor",
                       formulas=[f"=INDEX($F${case_row}:$H${case_row},$F${toggle_row})"],
                       driver="toggle-selected", cols=[V2_FIRST_PERIOD_COL])
        R[f"eff_{role}"] = r
        R[f"eff_{role}_cell"] = f"$F${r}"
        r += 1


def _build_revenue_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Revenue Build"]
    A = ctx.rows["Assumptions"]
    base_rev = ctx.snapshots.setdefault("base_revenue_annual", _v2_base_revenue_annual(facts))
    scale = _v2_pick_sheet_scale(ctx, [base_rev])
    ctx.sheet_scale["Revenue Build"] = scale
    _fmt_money, unit_money = _v2_money_fmt_unit(ctx, scale)
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Revenue Build",
        "Bottom-up revenue engine: demand → monetization → total revenue. "
        "Every period formula references the months ruler (row 5), so one row formula spans monthly and annual columns.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["Revenue Build"] = R
    unit_sale = facts.revenue_mode == "unit_sale"
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(facts.customers)
    effd = f"'Assumptions'!{A['eff_demand_cell']}"
    effp = f"'Assumptions'!{A['eff_price_cell']}"
    r = ib.DATA_START_ROW_V2

    _v2_section(ws, ctx, r, "Revenue drivers")
    r += 1
    if "new_units" in A:
        _v2_series_row(ws, ctx, r, "New primary units", "units",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['new_units']}*{effd}"),
                       driver="× demand")
        R["new"] = r
        r += 1
    ending_template = None
    ending_driver = ""
    ending_note = "first column rolls from N(prior)=0 — the unit column acts as the beginning base"
    units_rollforward_pinned = pinned and "customers" in A and "churn" in A and "new" in R
    if units_rollforward_pinned:
        # B2: even a stated (pinned) base is a live roll-forward — new units
        # and the implied churn drive the ending fleet; the FY-end
        # consistency check below proves it still lands on the stated series.
        ending_template = (
            f"=N({{p}}{r})*(1-'Assumptions'!{{c}}{A['churn']}*{{c}}$5/12)+{{c}}{R['new']}"
        )
        ending_driver = "roll-fwd"
        ending_note = "roll-forward off new units and churn — FY ends tie to the stated base (check below)"
    elif pinned and "customers" in A:
        ending_template = f"='Assumptions'!{{c}}{A['customers']}*{effd}"
        ending_driver = "stated base"
        ending_note = ""
    elif "new" in R and "churn" in A:
        ending_template = (
            f"=N({{p}}{r})*(1-'Assumptions'!{{c}}{A['churn']}*{{c}}$5/12)+{{c}}{R['new']}"
        )
        ending_driver = "roll-fwd"
    elif "new" in R:
        ending_template = f"=N({{p}}{r})+{{c}}{R['new']}"
        ending_driver = "roll-fwd"
    if ending_template:
        _v2_series_row(ws, ctx, r, "Ending primary units", "units",
                       formulas=_v2_formulas(ctx, ending_template), driver=ending_driver,
                       note=ending_note)
        R["ending"] = r
        r += 1
        _v2_series_row(ws, ctx, r, "Average primary units", "units",
                       formulas=_v2_formulas(ctx, f"=(N({{p}}{R['ending']})+{{c}}{R['ending']})/2"),
                       driver="period avg")
        R["avg"] = r
        r += 1
    if "price" in A:
        price_fmt, price_unit = _v2_row_money(ctx, facts.monthly_price_yen)
        _v2_series_row(ws, ctx, r, "Unit sale price" if unit_sale else "Monthly price / unit",
                       "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['price']}*{effp}"),
                       driver="× price", fmt=price_fmt, unit=price_unit)
        R["price"] = r
        r += 1
    if "gmv" in A:
        _v2_series_row(ws, ctx, r, "Gross merchandise value", "money",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['gmv']}*{effd}"),
                       driver="× demand")
        R["gmv"] = r
        r += 1
    if "take" in A:
        _v2_series_row(ws, ctx, r, "Take rate", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['take']}*{effp}"),
                       driver="× price")
        R["take"] = r
        r += 1

    r += 1
    _v2_section(ws, ctx, r, "Revenue streams")
    r += 1
    stream_rows = []
    if (not unit_sale) and "avg" in R and "price" in R:
        _v2_series_row(ws, ctx, r, "Recurring revenue", "money",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['avg']}*{{c}}{R['price']}*{{c}}$5"),
                       driver="avg × price × months")
        R["recurring"] = r
        stream_rows.append(r)
        r += 1
    if unit_sale and "new" in R and "price" in R:
        _v2_series_row(ws, ctx, r, "One-time revenue", "money",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['new']}*{{c}}{R['price']}"),
                       driver="new × price")
        R["one_time"] = r
        stream_rows.append(r)
        r += 1
    elif "one_time_fee" in A and "new" in R:
        _v2_series_row(ws, ctx, r, "One-time revenue", "money",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['new']}*'Assumptions'!{{c}}{A['one_time_fee']}"),
                       driver="new × fee")
        R["one_time"] = r
        stream_rows.append(r)
        r += 1
    if "gmv" in R and "take" in R:
        _v2_series_row(ws, ctx, r, "Transaction revenue", "money",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['gmv']}*{{c}}{R['take']}"),
                       driver="GMV × take")
        R["transaction"] = r
        stream_rows.append(r)
        r += 1
    if stream_rows and "other_share" in A:
        first, last = stream_rows[0], stream_rows[-1]
        _v2_series_row(ws, ctx, r, "Other revenue", "money",
                       formulas=_v2_formulas(
                           ctx, f"=SUM({{c}}{first}:{{c}}{last})*'Assumptions'!{{c}}{A['other_share']}"),
                       driver="share of subtotal")
        R["other"] = r
        stream_rows.append(r)
        r += 1
    if stream_rows:
        total_formulas = _v2_formulas(ctx, f"=SUM({{c}}{stream_rows[0]}:{{c}}{stream_rows[-1]})")
    else:
        total_formulas = _v2_formulas(ctx, "=0")
    _v2_series_row(ws, ctx, r, "Total revenue", "money", formulas=total_formulas,
                   bold=True, band=True)
    R["total"] = r
    r += 1
    _v2_series_row(
        ws, ctx, r, "Revenue growth", "pct",
        formulas=_v2_formulas(
            ctx,
            f'=IF({{c}}$5<>{{p}}$5,"-",IF({{p}}{R["total"]}=0,"-",{{c}}{R["total"]}/{{p}}{R["total"]}-1))',
        ),
        driver="same grain", note='"-" where the neighboring column is a different grain')
    R["growth"] = r
    r += 1
    if "ending" in R:
        rep = [base_rev[-1] / max(facts.customers[-1] if _v2_any(facts.customers) else 1, 1)]
        rpc_fmt, rpc_unit = _v2_row_money(ctx, rep)
        _v2_series_row(ws, ctx, r, "Revenue / customer (annualized)", "money_row",
                       formulas=_v2_formulas(
                           ctx,
                           f"=IF({{c}}{R['ending']}=0,0,{{c}}{R['total']}*12/{{c}}$5/{{c}}{R['ending']})"),
                       driver="run rate", fmt=rpc_fmt, unit=rpc_unit)
        R["rev_per_customer"] = r
        r += 1

    if not ("pool" in A and "conversion" in A and "new" in R):
        r += 1
        _v2_section(ws, ctx, r, "Demand support")
        r += 1
        _v2_label_cells(ws, r, "No demand-funnel evidence provided — DD gap", "", "",
                        "state a qualified pool and conversion evidence so the selected units are demand-backed",
                        ws._startup_note_col)
        r += 1
    if "pool" in A and "conversion" in A and "new" in R:
        r += 1
        _v2_section(ws, ctx, r, "Demand support")
        r += 1
        _v2_series_row(ws, ctx, r, "Qualified demand pool", "count",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['pool']}"),
                       driver="funnel")
        pool_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Demand conversion to units", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['conversion']}"),
                       driver="conversion")
        conv_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Implied new units from funnel", "units",
                       formulas=_v2_formulas(ctx, f"={{c}}{pool_row}*{{c}}{conv_row}"),
                       driver="pool × conv")
        implied_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Demand support coverage", "x",
                       formulas=_v2_formulas(
                           ctx, f"=IF({{c}}{R['new']}=0,0,{{c}}{implied_row}/{{c}}{R['new']})"),
                       driver="fill", bold=True)
        coverage_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Demand support status", "status",
                       formulas=_v2_formulas(ctx, f'=IF({{c}}{coverage_row}>=1,"ok","review")'),
                       note="coverage < 1x means the funnel does not support the selected units")
        r += 1

    if not ("price" in R and "roi" in A and "value_capture" in A and "tgm" in A):
        r += 1
        _v2_section(ws, ctx, r, "Price support")
        r += 1
        _v2_label_cells(ws, r, "No pricing evidence provided — DD gap", "", "",
                        "state a price anchor with ROI / value-capture evidence before treating monetization as proven",
                        ws._startup_note_col)
        r += 1
    if "price" in R and "roi" in A and "value_capture" in A and "tgm" in A:
        r += 1
        _v2_section(ws, ctx, r, "Price support")
        r += 1
        roi_fmt, roi_unit = _v2_row_money(ctx, [facts.customer_roi_yen])
        _v2_series_row(ws, ctx, r, "Customer annual value / ROI", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['roi']}"),
                       driver="value proof", fmt=roi_fmt, unit=roi_unit)
        roi_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Value capture share", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['value_capture']}"),
                       driver="capture")
        vc_row = r
        r += 1
        vp_fmt, vp_unit = _v2_row_money(
            ctx, [facts.customer_roi_yen * max(facts.value_capture_share[-1], 0.01) / 12])
        _v2_series_row(ws, ctx, r, "Value-based monthly price", "money_row",
                       formulas=_v2_formulas(ctx, f"={{c}}{roi_row}*{{c}}{vc_row}/12"),
                       driver="ROI share", fmt=vp_fmt, unit=vp_unit)
        value_price_row = r
        r += 1
        serve_terms = "+".join(
            f"'Assumptions'!{{c}}{A[key]}" for key in ("delivery", "cloud", "support") if key in A
        )
        floor_row = None
        if serve_terms:
            floor_fmt, floor_unit = _v2_row_money(
                ctx,
                [facts.delivery_cost_yen[-1] + facts.cloud_cost_yen[-1] + facts.support_cost_yen[-1]])
            _v2_series_row(ws, ctx, r, "Cost-plus monthly floor", "money_row",
                           formulas=_v2_formulas(
                               ctx,
                               f"=({serve_terms})/(1-MAX(1/100,'Assumptions'!{{c}}{A['tgm']}))"),
                           driver="cost floor", fmt=floor_fmt, unit=floor_unit)
            floor_row = r
            r += 1
        anchor = f"MAX({{c}}{value_price_row},{{c}}{floor_row})" if floor_row else f"{{c}}{value_price_row}"
        _v2_series_row(ws, ctx, r, "Price support ratio", "x",
                       formulas=_v2_formulas(
                           ctx, f"=IF({{c}}{R['price']}=0,0,{anchor}/{{c}}{R['price']})"),
                       driver="WTP / price", bold=True,
                       note="supported price anchor ÷ selected price — ≥1x means headroom")
        support_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Price support status", "status",
                       formulas=_v2_formulas(ctx, f'=IF({{c}}{support_row}>=1,"ok","review")'))
        r += 1

    if _v2_any(base_rev):
        r += 1
        fy_year_idx = {f"FY{year}": idx for idx, year in enumerate(facts.years)}
        snap_values = [
            base_rev[fy_year_idx[ctx.axis.fy_labels[i]]] for i in range(ctx.n_cols)
        ]
        _v2_series_row(ws, ctx, r, "FY total revenue (generator snapshot)", "money",
                       values=snap_values, snapshot=True, driver="kernel plan",
                       note="generated snapshot — rerun the generator after changing inputs")
        R["snap"] = r
        r += 1
        last = ctx.last_period_letter
        _v2_check_row(
            ws, ctx, r, "Revenue bridge check",
            _v2_formulas(
                ctx,
                f"=ROUND(MAX(0,ABS(SUMIF($F$4:${last}$4,{{c}}$4,$F${R['total']}:${last}${R['total']})"
                f"-{{c}}${R['snap']}*{effd}*{effp})-12*{{c}}${R['snap']}/100),0)",
            ),
            "Revenue Build: FY sum of period revenue vs kernel plan (±12%, scenario-scaled)",
            driver="vs kernel",
        )
        R["bridge_check"] = r
        r += 1

    if units_rollforward_pinned and "ending" in R:
        last = ctx.last_period_letter
        stated = f"'Assumptions'!{{c}}{A['customers']}*{effd}"
        _v2_check_row(
            ws, ctx, r, "Units roll-forward check",
            _v2_formulas(
                ctx,
                f"=IF(COUNTIF($F$4:{{c}}$4,{{c}}$4)<COUNTIF($F$4:${last}$4,{{c}}$4),0,"
                f"ROUND(MAX(0,ABS({{c}}{R['ending']}-{stated})-MAX(3,({stated})/100)),0))",
            ),
            "Revenue Build: FY-end units roll-forward lands on the stated customer base "
            "(±1% / ±3 units; mid-FY columns are not compared)",
            driver="vs stated base",
        )
        R["rollforward_check"] = r
        r += 1

    # Chart: single-grain window only (monthly window when hybrid/monthly).
    chart_cols = (
        (V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + ctx.axis.monthly_count - 1)
        if ctx.axis.monthly_count
        else (V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + ctx.n_cols - 1)
    )
    cats = Reference(ws, min_col=chart_cols[0], max_col=chart_cols[1], min_row=ib.HEADER_PERIOD_ROW)
    data = Reference(ws, min_col=chart_cols[0], max_col=chart_cols[1],
                     min_row=R["total"], max_row=R["total"])
    _add_line_chart(
        ws,
        "Total revenue (monthly window)" if ctx.axis.monthly_count else "Total revenue",
        data, cats, f"B{r + 2}", unit_money,
    )


def _build_people_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["People Plan"]
    A = ctx.rows["Assumptions"]
    RB = ctx.rows["Revenue Build"]
    periods = len(facts.years)
    total_hc_annual = [
        facts.product_headcount[i] + facts.gtm_headcount[i]
        + facts.operations_headcount[i] + facts.ga_headcount[i]
        for i in range(periods)
    ]
    people_cost_annual = [total_hc_annual[i] * facts.avg_comp_yen[i] for i in range(periods)]
    scale = _v2_pick_sheet_scale(ctx, [people_cost_annual])
    ctx.sheet_scale["People Plan"] = scale
    _fmt_money, unit_money = _v2_money_fmt_unit(ctx, scale)
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — People Plan",
        "Headcount by function, fully loaded compensation, people cost, and capacity checks.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["People Plan"] = R
    r = ib.DATA_START_ROW_V2
    _v2_section(ws, ctx, r, "Headcount")
    r += 1
    dept_rows = []
    for key, label, series in (
        ("product", "Product / R&D FTE", facts.product_headcount),
        ("gtm", "GTM FTE", facts.gtm_headcount),
        ("ops", "Operations / CS FTE", facts.operations_headcount),
        ("ga", "G&A FTE", facts.ga_headcount),
    ):
        if _v2_any(series):
            _v2_series_row(ws, ctx, r, label, "fte",
                           values=_v2_expand(ctx, series, "stock"), driver="hiring plan")
            R[key] = r
            dept_rows.append(r)
            r += 1
    if dept_rows:
        _v2_series_row(ws, ctx, r, "Total headcount", "fte",
                       formulas=_v2_formulas(ctx, f"=SUM({{c}}{dept_rows[0]}:{{c}}{dept_rows[-1]})"),
                       bold=True, band=True)
    else:
        _v2_series_row(ws, ctx, r, "Total headcount", "fte",
                       formulas=_v2_formulas(ctx, "=0"), bold=True, band=True)
    R["total"] = r
    r += 2

    _v2_section(ws, ctx, r, "Compensation")
    r += 1
    comp_fmt, comp_unit = _v2_row_money(ctx, facts.avg_comp_yen)
    welfare = float(getattr(facts, "statutory_welfare_rate", 0.0) or 0.0)
    if welfare > 0 and "base_salary" in A and "welfare_rate" in A:
        _v2_series_row(ws, ctx, r, "Avg base salary / FTE (annual)", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['base_salary']}"),
                       driver="base", fmt=comp_fmt, unit=comp_unit)
        base_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Statutory welfare rate", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['welfare_rate']}"),
                       driver="法定福利費")
        rate_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Fully loaded comp / FTE (annual)", "money_row",
                       formulas=_v2_formulas(ctx, f"={{c}}{base_row}*(1+{{c}}{rate_row})"),
                       driver="base × (1+rate)", bold=True, fmt=comp_fmt, unit=comp_unit)
        R["loaded"] = r
        r += 1
    else:
        _v2_series_row(ws, ctx, r, "Fully loaded comp / FTE (annual)", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['loaded_comp']}"),
                       driver="loaded", fmt=comp_fmt, unit=comp_unit,
                       note="no statutory welfare rate stated — comp treated as fully loaded")
        R["loaded"] = r
        r += 1
    _v2_series_row(ws, ctx, r, "People cost", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['total']}*{{c}}{R['loaded']}*{{c}}$5/12"),
                   driver="HC × comp", bold=True, band=True)
    R["people_cost"] = r
    r += 2

    _v2_section(ws, ctx, r, "Capacity")
    r += 1
    rev_fte_rep = [ctx.snapshots["base_revenue_annual"][-1] / max(total_hc_annual[-1], 1)]
    rf_fmt, rf_unit = _v2_row_money(ctx, rev_fte_rep)
    _v2_series_row(ws, ctx, r, "Revenue / FTE (annualized)", "money_row",
                   formulas=_v2_formulas(
                       ctx,
                       f"=IF({{c}}{R['total']}=0,0,'Revenue Build'!{{c}}{RB['total']}*12/{{c}}$5/{{c}}{R['total']})"),
                   driver="run rate", fmt=rf_fmt, unit=rf_unit)
    R["rev_per_fte"] = r
    r += 1
    if "ops" in R and "ending" in RB:
        _v2_series_row(ws, ctx, r, "Customers / CS FTE", "count",
                       formulas=_v2_formulas(
                           ctx, f"=IF({{c}}{R['ops']}=0,0,'Revenue Build'!{{c}}{RB['ending']}/{{c}}{R['ops']})"),
                       driver="load")
        r += 1
        if "tickets" in A and "ticket_capacity" in A:
            _v2_series_row(ws, ctx, r, "Required CS FTE from ticket load", "fte",
                           formulas=_v2_formulas(
                               ctx,
                               f"='Revenue Build'!{{c}}{RB['ending']}*'Assumptions'!{{c}}{A['tickets']}"
                               f"/'Assumptions'!{{c}}{A['ticket_capacity']}"),
                           driver="tickets", fmt=ib.FMT_MONTHS_1DP, unit="FTE")
            required_row = r
            r += 1
            _v2_series_row(ws, ctx, r, "CS capacity coverage", "x",
                           formulas=_v2_formulas(
                               ctx, f"=IF({{c}}{required_row}=0,0,{{c}}{R['ops']}/{{c}}{required_row})"),
                           driver="fill", bold=True)
            coverage_row = r
            r += 1
            _v2_series_row(ws, ctx, r, "CS capacity status", "status",
                           formulas=_v2_formulas(ctx, f'=IF({{c}}{coverage_row}>=1,"ok","review")'))
            r += 1
    _v2_series_row(
        ws, ctx, r, "Headcount growth (FY)", "pct",
        formulas=_v2_formulas(
            ctx,
            f'=IF(OR({{c}}$5<12,{{p}}$5<>{{c}}$5),"-",IF({{p}}{R["total"]}=0,"-",{{c}}{R["total"]}/{{p}}{R["total"]}-1))',
        ),
        driver="annual only", note="FY-basis growth — meaningful on annual columns only")
    hc_growth_row = r
    r += 1
    _v2_series_row(
        ws, ctx, r, "HC growth vs revenue growth flag", "status",
        formulas=_v2_formulas(
            ctx,
            f'=IF({{c}}{hc_growth_row}="-","-",IF(\'Revenue Build\'!{{c}}{RB["growth"]}="-","-",'
            f'IF({{c}}{hc_growth_row}>\'Revenue Build\'!{{c}}{RB["growth"]},"review","ok")))',
        ),
        note='"review" = headcount grows faster than revenue on an FY basis')
    r += 1


def _build_cost_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Cost Build"]
    A = ctx.rows["Assumptions"]
    RB = ctx.rows["Revenue Build"]
    PP = ctx.rows["People Plan"]
    periods = len(facts.years)
    base_rev = ctx.snapshots["base_revenue_annual"]
    cogs_annual = [base_rev[i] * (1.0 - facts.target_gross_margin[i]) for i in range(periods)]
    capex_annual = [
        facts.new_units[i] * facts.capex_per_unit_yen[i] + facts.other_capex_yen[i]
        for i in range(periods)
    ]
    scale = _v2_pick_sheet_scale(ctx, [cogs_annual, capex_annual])
    ctx.sheet_scale["Cost Build"] = scale
    _fmt_money, unit_money = _v2_money_fmt_unit(ctx, scale)
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Cost Build",
        "Cost-to-serve, gross profit, OpEx programs, and CapEx. COGS components are governor-calibrated to the target gross margin.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["Cost Build"] = R
    effc = f"'Assumptions'!{A['eff_cost_cell']}"
    effo = f"'Assumptions'!{A['eff_opex_cell']}"
    rev = f"'Revenue Build'!{{c}}{RB['total']}"
    r = ib.DATA_START_ROW_V2

    _v2_section(ws, ctx, r, "COGS")
    r += 1
    cogs_rows = []
    if "vc_pct" in A:
        _v2_series_row(ws, ctx, r, "Variable COGS", "money",
                       formulas=_v2_formulas(ctx, f"={rev}*'Assumptions'!{{c}}{A['vc_pct']}*{effc}"),
                       driver="% of rev")
        R["variable"] = r
        cogs_rows.append(r)
        r += 1
    if "delivery" in A and "avg" in RB:
        _v2_series_row(ws, ctx, r, "Delivery cost", "money",
                       formulas=_v2_formulas(
                           ctx,
                           f"='Revenue Build'!{{c}}{RB['avg']}*'Assumptions'!{{c}}{A['delivery']}*{{c}}$5*{effc}"),
                       driver="per unit")
        cogs_rows.append(r)
        r += 1
    if "cloud" in A and "avg" in RB:
        _v2_series_row(ws, ctx, r, "Cloud / platform cost", "money",
                       formulas=_v2_formulas(
                           ctx,
                           f"='Revenue Build'!{{c}}{RB['avg']}*'Assumptions'!{{c}}{A['cloud']}*{{c}}$5*{effc}"),
                       driver="per unit")
        cogs_rows.append(r)
        r += 1
    if "support" in A and "customers" in A:
        _v2_series_row(ws, ctx, r, "Support cost", "money",
                       formulas=_v2_formulas(
                           ctx,
                           f"='Assumptions'!{{c}}{A['customers']}*'Assumptions'!{{c}}{A['support']}*{{c}}$5*{effc}"),
                       driver="per customer")
        cogs_rows.append(r)
        r += 1
    if cogs_rows:
        total_formulas = _v2_formulas(ctx, f"=SUM({{c}}{cogs_rows[0]}:{{c}}{cogs_rows[-1]})")
    else:
        total_formulas = _v2_formulas(ctx, "=0")
    _v2_series_row(ws, ctx, r, "Total COGS", "money", formulas=total_formulas, bold=True, band=True)
    R["cogs"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Gross profit", "money",
                   formulas=_v2_formulas(ctx, f"={rev}-{{c}}{R['cogs']}"), bold=True)
    R["gp"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Gross margin", "pct",
                   formulas=_v2_formulas(ctx, f"=IF({rev}=0,0,{{c}}{R['gp']}/{rev})"))
    R["gm"] = r
    r += 2

    _v2_section(ws, ctx, r, "OpEx programs")
    r += 1
    if "sm_pct" in A:
        _v2_series_row(ws, ctx, r, "S&M programs", "money",
                       formulas=_v2_formulas(ctx, f"={rev}*'Assumptions'!{{c}}{A['sm_pct']}*{effo}"),
                       driver="% of rev")
        R["sm"] = r
        r += 1
    rd_terms = []
    if "rd_floor" in A:
        rd_terms.append(f"'Assumptions'!{{c}}{A['rd_floor']}*{{c}}$5/12")
    if "rd_per_fte" in A and "product" in PP:
        rd_terms.append(
            f"'People Plan'!{{c}}{PP['product']}*'Assumptions'!{{c}}{A['rd_per_fte']}*{{c}}$5/12")
    if rd_terms:
        rd_body = f"MAX({','.join(rd_terms)})" if len(rd_terms) > 1 else rd_terms[0]
        _v2_series_row(ws, ctx, r, "R&D programs", "money",
                       formulas=_v2_formulas(ctx, f"=({rd_body})*{effo}"),
                       driver="floor vs FTE")
        R["rd"] = r
        r += 1
    ga_terms = []
    if "ga_pct" in A:
        ga_terms.append(f"{rev}*'Assumptions'!{{c}}{A['ga_pct']}")
    if "fixed_ga" in A:
        ga_terms.append(f"'Assumptions'!{{c}}{A['fixed_ga']}*{{c}}$5/12")
    if ga_terms:
        _v2_series_row(ws, ctx, r, "G&A programs", "money",
                       formulas=_v2_formulas(ctx, f"=({'+'.join(ga_terms)})*{effo}"),
                       driver="% + fixed")
        R["ga"] = r
        r += 1
    r += 1

    _v2_section(ws, ctx, r, "CapEx")
    r += 1
    capex_terms = []
    if "capex_unit" in A and "new" in RB:
        capex_terms.append(f"'Revenue Build'!{{c}}{RB['new']}*'Assumptions'!{{c}}{A['capex_unit']}")
    if "other_capex" in A:
        capex_terms.append(f"'Assumptions'!{{c}}{A['other_capex']}*{{c}}$5/12")
    _v2_series_row(ws, ctx, r, "CapEx", "money",
                   formulas=_v2_formulas(ctx, f"={'+'.join(capex_terms)}" if capex_terms else "=0"),
                   driver="unit + other", bold=True)
    R["capex"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Capital intensity", "pct",
                   formulas=_v2_formulas(ctx, f"=IF({rev}=0,0,{{c}}{R['capex']}/{rev})"))
    R["capital_intensity"] = r
    r += 2

    _v2_section(ws, ctx, r, "Cost-to-serve")
    r += 1
    _v2_series_row(ws, ctx, r, "Implied COGS / revenue", "pct",
                   formulas=_v2_formulas(ctx, f"=IF({rev}=0,0,{{c}}{R['cogs']}/{rev})"))
    implied_row = r
    r += 1
    if "tgm" in A:
        _v2_series_row(ws, ctx, r, "Target COGS / revenue", "pct",
                       formulas=_v2_formulas(ctx, f"=1-'Assumptions'!{{c}}{A['tgm']}"),
                       driver="1 − target GM")
        target_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Cost-to-serve support ratio", "x",
                       formulas=_v2_formulas(
                           ctx, f"=IF({{c}}{target_row}=0,0,{{c}}{implied_row}/{{c}}{target_row})"),
                       bold=True)
        ratio_row = r
        r += 1
        _v2_series_row(ws, ctx, r, "Cost-to-serve status", "status",
                       formulas=_v2_formulas(
                           ctx,
                           f'=IF({{c}}{ratio_row}=0,"n/a",IF(ABS({{c}}{ratio_row}-1)<=1/3,"ok","review"))'))
        r += 1
    ws.cell(r, 3, "COGS calibration note")
    ib.apply_label(ws.cell(r, 3))
    governor_note = ws.cell(
        r, ws._startup_note_col,
        "Note: variable COGS, delivery, cloud, and support are governor-rescaled by one per-period "
        "factor so total COGS lands on the target gross margin — cost mix is profile-shaped, "
        "cost level is calibrated.",
    )
    ib.apply_comment(governor_note, wrap_text=False)
    r += 1
    if "tgm" in A and _v2_any(base_rev):
        _v2_check_row(
            ws, ctx, r, "Gross margin vs target check",
            _v2_formulas(
                ctx,
                f"=IF(AND({{c}}$5=12,'Assumptions'!$F${A['scenario_toggle']}=2),"
                f"ROUND(MAX(0,ABS({{c}}{R['gm']}-'Assumptions'!{{c}}{A['tgm']})*100-3),0),0)",
            ),
            "Cost Build: gross margin vs target on annual columns at Base (±3pp)",
            driver="vs target",
            note="evaluated on annual columns in the Base case only",
        )
        R["gm_check"] = r
        r += 1


# --- statement layer (P&L / BS / CF) -----------------------------------------


def _v2_compact_ops(ws: Worksheet, ctx: BuildContext, r: int) -> tuple:
    """Compact operating block for bundles without the engine sheets.

    Writes revenue / COGS / people / OpEx / CapEx rows straight off the
    Assumptions register (BLUEPRINT_S3 bundle-aware 参照). Returns
    (next_row, row registry)."""
    facts = ctx.facts
    A = ctx.rows["Assumptions"]
    C: dict = {}
    unit_sale = facts.revenue_mode == "unit_sale"
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(facts.customers)
    effd = f"'Assumptions'!{A['eff_demand_cell']}"
    effp = f"'Assumptions'!{A['eff_price_cell']}"
    effc = f"'Assumptions'!{A['eff_cost_cell']}"
    effo = f"'Assumptions'!{A['eff_opex_cell']}"
    _v2_section(ws, ctx, r, "Compact operating build")
    r += 1
    if "new_units" in A:
        _v2_series_row(ws, ctx, r, "New primary units", "units",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['new_units']}*{effd}"),
                       driver="× demand")
        C["new"] = r
        r += 1
    if pinned and "customers" in A:
        _v2_series_row(ws, ctx, r, "Ending units / customers", "units",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['customers']}*{effd}"),
                       driver="stated base",
                       note="compact block — engine sheets are not in this bundle")
        C["ending"] = r
        r += 1
    elif "new_units" in A:
        if "churn" in A:
            tmpl = (f"=N({{p}}{r})*(1-'Assumptions'!{{c}}{A['churn']}*{{c}}$5/12)"
                    f"+'Assumptions'!{{c}}{A['new_units']}*{effd}")
        else:
            tmpl = f"=N({{p}}{r})+'Assumptions'!{{c}}{A['new_units']}*{effd}"
        _v2_series_row(ws, ctx, r, "Ending units / customers", "units",
                       formulas=_v2_formulas(ctx, tmpl), driver="roll-fwd",
                       note="compact block — engine sheets are not in this bundle")
        C["ending"] = r
        r += 1
    if "ending" in C:
        _v2_series_row(ws, ctx, r, "Average units", "units",
                       formulas=_v2_formulas(ctx, f"=(N({{p}}{C['ending']})+{{c}}{C['ending']})/2"),
                       driver="period avg")
        C["avg"] = r
        r += 1
    rev_terms = []
    if unit_sale and "new_units" in A and "price" in A:
        rev_terms.append(
            f"'Assumptions'!{{c}}{A['new_units']}*{effd}*'Assumptions'!{{c}}{A['price']}*{effp}")
    elif "avg" in C and "price" in A:
        rev_terms.append(f"{{c}}{C['avg']}*'Assumptions'!{{c}}{A['price']}*{effp}*{{c}}$5")
        if "one_time_fee" in A and "new_units" in A:
            rev_terms.append(
                f"'Assumptions'!{{c}}{A['new_units']}*{effd}*'Assumptions'!{{c}}{A['one_time_fee']}")
    if "gmv" in A and "take" in A:
        rev_terms.append(
            f"'Assumptions'!{{c}}{A['gmv']}*{effd}*'Assumptions'!{{c}}{A['take']}*{effp}")
    if rev_terms and "other_share" in A:
        rev_formula = f"=({'+'.join(rev_terms)})*(1+'Assumptions'!{{c}}{A['other_share']})"
    elif rev_terms:
        rev_formula = f"={'+'.join(rev_terms)}"
    else:
        rev_formula = "=0"
    _v2_series_row(ws, ctx, r, "Revenue (compact)", "money",
                   formulas=_v2_formulas(ctx, rev_formula), driver="Assumptions direct",
                   bold=True)
    C["rev"] = r
    r += 1
    if "tgm" in A:
        cogs_formula = f"={{c}}{C['rev']}*(1-'Assumptions'!{{c}}{A['tgm']})*{effc}"
        cogs_driver = "1 − target GM"
    elif "vc_pct" in A:
        cogs_formula = f"={{c}}{C['rev']}*'Assumptions'!{{c}}{A['vc_pct']}*{effc}"
        cogs_driver = "% of revenue"
    else:
        cogs_formula = "=0"
        cogs_driver = ""
    _v2_series_row(ws, ctx, r, "COGS (compact)", "money",
                   formulas=_v2_formulas(ctx, cogs_formula), driver=cogs_driver,
                   note="target-margin basis — Cost Build is not in this bundle")
    C["cogs"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Gross profit", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{C['rev']}-{{c}}{C['cogs']}"), bold=True)
    C["gp"] = r
    r += 1
    if "base_salary" in A and "welfare_rate" in A:
        loaded = f"'Assumptions'!{{c}}{A['base_salary']}*(1+'Assumptions'!{{c}}{A['welfare_rate']})"
    elif "loaded_comp" in A:
        loaded = f"'Assumptions'!{{c}}{A['loaded_comp']}"
    else:
        loaded = None
    if "total_hc" in A and loaded is not None:
        _v2_series_row(ws, ctx, r, "People cost", "money",
                       formulas=_v2_formulas(
                           ctx, f"='Assumptions'!{{c}}{A['total_hc']}*{loaded}*{{c}}$5/12"),
                       driver="HC × comp")
        C["people"] = r
        r += 1
    opex_terms = []
    if "sm_pct" in A:
        opex_terms.append(f"{{c}}{C['rev']}*'Assumptions'!{{c}}{A['sm_pct']}")
    if "ga_pct" in A:
        opex_terms.append(f"{{c}}{C['rev']}*'Assumptions'!{{c}}{A['ga_pct']}")
    if "fixed_ga" in A:
        opex_terms.append(f"'Assumptions'!{{c}}{A['fixed_ga']}*{{c}}$5/12")
    rd_terms = []
    if "rd_floor" in A:
        rd_terms.append(f"'Assumptions'!{{c}}{A['rd_floor']}*{{c}}$5/12")
    if "rd_per_fte" in A and "product_hc" in A:
        rd_terms.append(
            f"'Assumptions'!{{c}}{A['product_hc']}*'Assumptions'!{{c}}{A['rd_per_fte']}*{{c}}$5/12")
    if rd_terms:
        opex_terms.append(f"MAX({','.join(rd_terms)})" if len(rd_terms) > 1 else rd_terms[0])
    if opex_terms:
        _v2_series_row(ws, ctx, r, "Program OpEx (S&M / R&D / G&A)", "money",
                       formulas=_v2_formulas(ctx, f"=({'+'.join(opex_terms)})*{effo}"),
                       driver="% + programs")
        C["programs"] = r
        r += 1
    opex_refs = "+".join(f"{{c}}{C[key]}" for key in ("people", "programs") if key in C)
    _v2_series_row(ws, ctx, r, "EBITDA (compact)", "money",
                   formulas=_v2_formulas(
                       ctx, f"={{c}}{C['gp']}-({opex_refs})" if opex_refs else f"={{c}}{C['gp']}"),
                   bold=True, band=True)
    C["ebitda"] = r
    r += 1
    capex_terms = []
    if "capex_unit" in A and "new_units" in A:
        capex_terms.append(
            f"'Assumptions'!{{c}}{A['new_units']}*{effd}*'Assumptions'!{{c}}{A['capex_unit']}")
    if "other_capex" in A:
        capex_terms.append(f"'Assumptions'!{{c}}{A['other_capex']}*{{c}}$5/12")
    _v2_series_row(ws, ctx, r, "CapEx (compact)", "money",
                   formulas=_v2_formulas(ctx, f"={'+'.join(capex_terms)}" if capex_terms else "=0"),
                   driver="unit + other")
    C["capex"] = r
    r += 2
    return r, C


def _v2_statement_scale(ctx: BuildContext) -> tuple:
    """Shared sheet scale for the statement sheets (P&L / BS / CF).

    The statements carry small JP working-capital / tax-timing balance rows
    (未払消費税等, 預り金・未払社会保険料) that are ~2 orders below revenue; on a
    monthly grain they would crush to "0" at the revenue-driven 百万円 scale.
    Feed annual proxies of those rows to the digit-crush guard so the sheet
    scale drops to 千円 whenever they are material and present."""
    facts = ctx.facts
    base_rev = ctx.snapshots.setdefault(
        "base_revenue_annual", _v2_base_revenue_annual(facts))
    projection = ctx.snapshots.setdefault(
        "base_projection", project_plan_free_cash_flow(facts))
    cash = [abs(p["ending_cash"]) for p in projection]
    detail: list = []
    ctax_rate = float(getattr(facts, "consumption_tax_rate", 0.0) or 0.0)
    if ctax_rate > 0:  # 未払消費税等 balance ≈ (rev − cogs) × rate × quarter share
        detail.append([max(0.0, p["revenue"] - p["cogs"]) * ctax_rate * 0.25
                       for p in projection])
    welfare = float(getattr(facts, "statutory_welfare_rate", 0.0) or 0.0)
    if welfare > 0:  # 預り金・未払社会保険料 balance ≈ monthly people cost × rate
        periods = len(facts.years)
        people_cost = [
            (facts.product_headcount[i] + facts.gtm_headcount[i]
             + facts.operations_headcount[i] + facts.ga_headcount[i])
            * facts.avg_comp_yen[i] for i in range(periods)
        ]
        detail.append([pc / 12.0 * welfare for pc in people_cost])
    scale = _v2_pick_sheet_scale(ctx, [base_rev], detail) if _v2_any(base_rev) else ib.pick_sheet_scale(
        cash, currency=ctx.currency)
    return scale, _v2_money_fmt_unit(ctx, scale)


def _build_pl_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["P&L"]
    A = ctx.rows["Assumptions"]
    scale, (fmt_money, unit_money) = _v2_statement_scale(ctx)
    ctx.sheet_scale["P&L"] = scale
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — P&L",
        "Reference-only statement (tax-exclusive). Operating rows reference the engine sheets; "
        "no new operating logic is computed here.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["P&L"] = R
    engines = "Revenue Build" in ctx.bundle and "Cost Build" in ctx.bundle
    r = ib.DATA_START_ROW_V2
    if engines:
        RB = ctx.rows["Revenue Build"]
        CB = ctx.rows["Cost Build"]
        PP = ctx.rows.get("People Plan", {})
        rev_src = f"'Revenue Build'!{{c}}{RB['total']}"
        cogs_src = f"'Cost Build'!{{c}}{CB['cogs']}"
        capex_src = f"'Cost Build'!{{c}}{CB['capex']}"
        people_src = f"'People Plan'!{{c}}{PP['people_cost']}" if "people_cost" in PP else None
        program_rows = [
            (label, f"'Cost Build'!{{c}}{CB[key]}")
            for key, label in (("sm", "S&M programs"), ("rd", "R&D programs"), ("ga", "G&A programs"))
            if key in CB
        ]
    else:
        r, C = _v2_compact_ops(ws, ctx, r)
        rev_src = f"{{c}}{C['rev']}"
        cogs_src = f"{{c}}{C['cogs']}"
        capex_src = f"{{c}}{C['capex']}"
        people_src = f"{{c}}{C['people']}" if "people" in C else None
        program_rows = (
            [("Program OpEx (S&M / R&D / G&A)", f"{{c}}{C['programs']}")] if "programs" in C else []
        )
        R["capex_compact"] = C["capex"]
        R["people_compact"] = C.get("people")

    _v2_section(ws, ctx, r, "Operating performance")
    r += 1
    _v2_series_row(ws, ctx, r, "Total revenue", "money",
                   formulas=_v2_formulas(ctx, f"={rev_src}"), driver="engine ref",
                   bold=True, band=True)
    R["rev"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Total COGS", "money",
                   formulas=_v2_formulas(ctx, f"={cogs_src}"), driver="engine ref")
    R["cogs"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Gross profit", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['rev']}-{{c}}{R['cogs']}"), bold=True)
    R["gp"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Gross margin", "pct",
                   formulas=_v2_formulas(
                       ctx, f"=IF({{c}}{R['rev']}=0,0,{{c}}{R['gp']}/{{c}}{R['rev']})"))
    R["gm"] = r
    r += 2

    _v2_section(ws, ctx, r, "Operating expenses")
    r += 1
    opex_first = r
    if people_src:
        _v2_series_row(ws, ctx, r, "People cost", "money",
                       formulas=_v2_formulas(ctx, f"={people_src}"), driver="engine ref")
        R["people"] = r
        r += 1
    for label, src in program_rows:
        _v2_series_row(ws, ctx, r, label, "money",
                       formulas=_v2_formulas(ctx, f"={src}"), driver="engine ref")
        R.setdefault("programs_rows", []).append(r)
        r += 1
    opex_last = r - 1
    if opex_last >= opex_first:
        total_opex = f"=SUM({{c}}{opex_first}:{{c}}{opex_last})"
    else:
        total_opex = "=0"
    _v2_series_row(ws, ctx, r, "Total OpEx", "money",
                   formulas=_v2_formulas(ctx, total_opex), bold=True)
    R["opex"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "EBITDA", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['gp']}-{{c}}{R['opex']}"),
                   bold=True, band=True)
    R["ebitda"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "EBITDA margin", "pct",
                   formulas=_v2_formulas(
                       ctx, f"=IF({{c}}{R['rev']}=0,0,{{c}}{R['ebitda']}/{{c}}{R['rev']})"))
    R["ebitda_margin"] = r
    r += 2

    _v2_section(ws, ctx, r, "Below EBITDA (tax-exclusive)")
    r += 1
    # Reference balance rows land in a section below; formulas may reference
    # forward rows, so reserve their positions relative to the layout order.
    has_da = "dep_life" in A
    debt_terms = [key for key in ("debt_raise", "convertibles", "lease") if key in A]
    has_interest = bool(debt_terms) and "debt_rate" in A
    if has_da:
        # Half-year convention on annual columns (kernel Task 3.3 G): assets
        # acquired in the period take half a period's charge; monthly columns
        # charge from the acquisition month.
        _v2_series_row(
            ws, ctx, r, "D&A", "money",
            formulas=_v2_formulas(
                ctx,
                "=(N({p}%(cum)s)+IF({c}$5=12,%(capex)s/2,%(capex)s))*{c}$5"
                "/'Assumptions'!{c}%(life)s" % {
                    "cum": "%(cum)s", "capex": capex_src, "life": A["dep_life"]},
            ),
            driver="cum CapEx / life",
            note="半年規約は年次列のみ適用; 月次列は取得月から全額償却",
        )
        R["da"] = r
        r += 1
    _v2_series_row(ws, ctx, r, "EBIT", "money",
                   formulas=_v2_formulas(
                       ctx,
                       f"={{c}}{R['ebitda']}-{{c}}{R['da']}" if has_da else f"={{c}}{R['ebitda']}"),
                   bold=True)
    R["ebit"] = r
    r += 1
    if has_interest:
        _v2_series_row(ws, ctx, r, "Interest expense", "money",
                       formulas=_v2_formulas(
                           ctx,
                           "=N({p}%(bal)s)*'Assumptions'!{c}%(rate)s*{c}$5/12" % {
                               "bal": "%(bal)s", "rate": A["debt_rate"]}),
                       driver="opening balance × rate",
                       note="期首残高ベース(循環参照回避)")
        R["interest"] = r
        r += 1
    _v2_series_row(ws, ctx, r, "EBT", "money",
                   formulas=_v2_formulas(
                       ctx,
                       f"={{c}}{R['ebit']}-{{c}}{R['interest']}" if has_interest
                       else f"={{c}}{R['ebit']}"))
    R["ebt"] = r
    r += 1
    if "tax_rate" in A:
        _v2_series_row(ws, ctx, r, "Corporate tax", "money",
                       formulas=_v2_formulas(
                           ctx, f"=MAX(0,{{c}}{R['ebt']}*'Assumptions'!{{c}}{A['tax_rate']})"),
                       driver="MAX(0, EBT × rate)",
                       note="NOL繰越は簡略化(期間独立課税) — カーネルはNOLを考慮")
        R["tax"] = r
        r += 1
    ni_formula = (
        f"={{c}}{R['ebt']}-{{c}}{R['tax']}" if "tax" in R else f"={{c}}{R['ebt']}"
    )
    _v2_series_row(ws, ctx, r, "Net income", "money",
                   formulas=_v2_formulas(ctx, ni_formula), bold=True, band=True)
    R["ni"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Net margin", "pct",
                   formulas=_v2_formulas(
                       ctx, f"=IF({{c}}{R['rev']}=0,0,{{c}}{R['ni']}/{{c}}{R['rev']})"))
    r += 1
    _v2_check_row(
        ws, ctx, r, "P&L tie check",
        _v2_formulas(ctx, f"=ROUND({{c}}{R['ebitda']}-({{c}}{R['gp']}-{{c}}{R['opex']}),0)"),
        "P&L: EBITDA ties to gross profit − total OpEx",
        driver="internal tie",
    )
    r += 2

    _v2_section(ws, ctx, r, "Reference balances")
    r += 1
    cum_capex_row = r
    _v2_series_row(ws, ctx, r, "Gross PP&E (cumulative CapEx)", "money",
                   formulas=_v2_formulas(ctx, f"=N({{p}}{r})+{capex_src}"),
                   driver="roll-forward")
    R["cum_capex"] = r
    r += 1
    debt_bal_row = None
    if debt_terms:
        terms = "".join(f"+'Assumptions'!{{c}}{A[key]}" for key in debt_terms)
        if "amortization" in A:
            terms += f"-'Assumptions'!{{c}}{A['amortization']}"
        _v2_series_row(ws, ctx, r, "Debt balance (ending)", "money",
                       formulas=_v2_formulas(ctx, f"=N({{p}}{r}){terms}"),
                       driver="draws − repayments",
                       note="drawn − repaid roll-forward (debt + converts + lease)")
        R["debt_balance"] = r
        debt_bal_row = r
        r += 1
    # Patch the forward references now that the balance rows are placed.
    if has_da:
        for col in ctx.period_cols:
            cell = ws.cell(R["da"], col)
            cell.value = cell.value % {"cum": cum_capex_row}
    if has_interest:
        for col in ctx.period_cols:
            cell = ws.cell(R["interest"], col)
            cell.value = cell.value % {"bal": debt_bal_row}


def _build_cf_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["CF"]
    A = ctx.rows["Assumptions"]
    scale, (fmt_money, unit_money) = _v2_statement_scale(ctx)
    ctx.sheet_scale["CF"] = scale
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Cash Flow (資金繰り)",
        "Tax-inclusive cash walk: operating cash, working-capital and tax balances "
        "(balance method), CapEx, financing events, ending cash, and runway.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["CF"] = R
    r = ib.DATA_START_ROW_V2
    has_pl = "P&L" in ctx.bundle
    engines = "Revenue Build" in ctx.bundle and "Cost Build" in ctx.bundle
    if has_pl:
        PL = ctx.rows["P&L"]
        rev_src = f"'P&L'!{{c}}{PL['rev']}"
        cogs_src = f"'P&L'!{{c}}{PL['cogs']}"
        if engines:
            capex_src = f"'Cost Build'!{{c}}{ctx.rows['Cost Build']['capex']}"
        else:
            capex_src = f"'P&L'!{{c}}{PL['capex_compact']}"
        if "People Plan" in ctx.bundle:
            people_src = f"'People Plan'!{{c}}{ctx.rows['People Plan']['people_cost']}"
        elif PL.get("people") is not None:
            people_src = f"'P&L'!{{c}}{PL['people']}"
        else:
            people_src = None
        ni_src = f"'P&L'!{{c}}{PL['ni']}"
        da_src = f"'P&L'!{{c}}{PL['da']}" if "da" in PL else None
    else:
        r, C = _v2_compact_ops(ws, ctx, r)
        rev_src = f"{{c}}{C['rev']}"
        cogs_src = f"{{c}}{C['cogs']}"
        capex_src = f"{{c}}{C['capex']}"
        people_src = f"{{c}}{C['people']}" if "people" in C else None
        ni_src = f"{{c}}{C['ebitda']}"
        da_src = None
        R["compact"] = C

    # --- working capital & tax balances (levels; Δ rows below read these) --
    _v2_section(ws, ctx, r, "Working capital & tax balances (期末残高)")
    r += 1
    last = ctx.last_period_letter
    balances: list = []  # (key, label, sign(+1 liability / -1 asset))

    def balance(key, label, formula, sign, driver="", note=""):
        nonlocal r
        _v2_series_row(ws, ctx, r, label, "money",
                       formulas=_v2_formulas(ctx, formula), driver=driver, note=note)
        R[f"bal_{key}"] = r
        balances.append((key, label, sign, r))
        r += 1

    if "ar_site" in A:
        balance("ar", "Accounts receivable",
                f"=({rev_src}/{{c}}$5)*'Assumptions'!{{c}}{A['ar_site']}", -1,
                driver="月商 × サイト", note="回収サイト方式 (site months × monthly sales)")
    elif "ar_days" in A:
        balance("ar", "Accounts receivable",
                f"=({rev_src}*12/{{c}}$5)*'Assumptions'!{{c}}{A['ar_days']}/365", -1,
                driver="annualized × days/365")
    if "inv_wip" in A:
        balance("inv", "Inventory / WIP",
                f"={capex_src}*'Assumptions'!{{c}}{A['inv_wip']}", -1,
                driver="CapEx share")
    if "ap_site" in A:
        balance("ap", "Accounts payable",
                f"=({cogs_src}/{{c}}$5)*'Assumptions'!{{c}}{A['ap_site']}", 1,
                driver="月次仕入 × サイト", note="支払サイト方式")
    elif "ap_days" in A:
        balance("ap", "Accounts payable",
                f"=({cogs_src}*12/{{c}}$5)*'Assumptions'!{{c}}{A['ap_days']}/365", 1,
                driver="annualized × days/365")
    if "deferred_share" in A:
        balance("deferred", "Deferred revenue",
                f"=({rev_src}*12/{{c}}$5)*'Assumptions'!{{c}}{A['deferred_share']}", 1,
                driver="annualized × share")
    if "advances" in A:
        balance("advances", "Customer advances (前受金)",
                f"=SUMIF($F$4:${last}$4,{{c}}$4,'Assumptions'!$F${A['advances']}:${last}${A['advances']})",
                1, driver="FY receipt held",
                note="受領FY中は残高、翌FYに解消(タイミングシフト、資金調達ではない)")
    if "consumption_tax" in A:
        balance("ctax", "未払消費税等",
                f"=({rev_src}-{cogs_src})*12/{{c}}$5*'Assumptions'!{{c}}{A['consumption_tax']}*3/12",
                1, driver="(課税売上−課税仕入)×税率",
                note="残高方式簡略化: 約3ヶ月分の未納付額を負債計上(中間納付は無視)")
    if "welfare_rate" in A and people_src:
        balance("withholding", "預り金・未払社会保険料",
                f"=({people_src}/{{c}}$5)*'Assumptions'!{{c}}{A['welfare_rate']}",
                1, driver="人件費 × 法定福利率",
                note="残高方式簡略化: 1ヶ月分の預り・未払を負債計上")
    r += 1

    # --- operating cash flow ------------------------------------------------
    _v2_section(ws, ctx, r, "Operating cash flow (税込)")
    r += 1
    ocf_first = r
    _v2_series_row(ws, ctx, r, "Net income" if has_pl else "Operating profit (compact proxy)",
                   "money", formulas=_v2_formulas(ctx, f"={ni_src}"),
                   driver="P&L" if has_pl else "compact EBITDA",
                   note="" if has_pl else "D&A・利息・税は簡略化(compact bundle)")
    R["ni"] = r
    r += 1
    if da_src:
        _v2_series_row(ws, ctx, r, "D&A (non-cash add-back)", "money",
                       formulas=_v2_formulas(ctx, f"={da_src}"), driver="P&L")
        R["da"] = r
        r += 1
    for key, label, sign, bal_row in balances:
        if sign < 0:
            formula = f"=N({{p}}{bal_row})-{{c}}{bal_row}"
        else:
            formula = f"={{c}}{bal_row}-N({{p}}{bal_row})"
        _v2_series_row(ws, ctx, r, f"Δ {label}", "money",
                       formulas=_v2_formulas(ctx, formula),
                       driver="balance delta")
        R[f"d_{key}"] = r
        r += 1
    _v2_series_row(ws, ctx, r, "Operating cash flow", "money",
                   formulas=_v2_formulas(ctx, f"=SUM({{c}}{ocf_first}:{{c}}{r - 1})"),
                   bold=True)
    R["ocf"] = r
    r += 2

    _v2_section(ws, ctx, r, "Investing")
    r += 1
    _v2_series_row(ws, ctx, r, "CapEx", "money",
                   formulas=_v2_formulas(ctx, f"=-({capex_src})"), driver="engine ref")
    R["capex"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Free cash flow", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['ocf']}+{{c}}{R['capex']}"),
                   bold=True, band=True)
    R["fcf"] = r
    r += 2

    _v2_section(ws, ctx, r, "Financing (by instrument)")
    r += 1
    fin_first = r
    for key, label, negate in (
        ("equity_raise", "Equity financing", False),
        ("debt_raise", "Debt draw", False),
        ("grants", "Grants / subsidies", False),
        ("convertibles", "J-KISS / convertible notes", False),
        ("lease", "Lease financing", False),
        ("secondary", "Founder / investor secondary (use)", True),
        ("amortization", "Debt principal repayment", True),
    ):
        if key in A:
            sign = "-" if negate else ""
            _v2_series_row(ws, ctx, r, label, "money",
                           formulas=_v2_formulas(ctx, f"={sign}'Assumptions'!{{c}}{A[key]}"),
                           driver="funding plan")
            R[f"fin_{key}"] = r
            r += 1
    if r > fin_first:
        _v2_series_row(ws, ctx, r, "Total financing CF", "money",
                       formulas=_v2_formulas(ctx, f"=SUM({{c}}{fin_first}:{{c}}{r - 1})"),
                       bold=True,
                       note="期首調達仮定 — 各FYの調達はそのFYの第1月列に計上(Assumptions参照)")
    else:
        _v2_series_row(ws, ctx, r, "Total financing CF", "money",
                       formulas=_v2_formulas(ctx, "=0"), bold=True,
                       note="no financing instruments stated")
    R["fin_total"] = r
    r += 2

    _v2_section(ws, ctx, r, "Cash & runway")
    r += 1
    _v2_series_row(ws, ctx, r, "Net cash flow", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['fcf']}+{{c}}{R['fin_total']}"))
    R["ncf"] = r
    r += 1
    begin_row = r
    end_row = r + 1
    _v2_series_row(ws, ctx, r, "Beginning cash", "money",
                   formulas=_v2_formulas(
                       ctx,
                       f"=N({{p}}{end_row})+IF(N({{p}}$5)=0,'Assumptions'!$F${A['beginning_cash']},0)"),
                   driver="prior ending",
                   note="first column adds the model-start balance (prior months ruler is blank)")
    R["begin"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Ending cash", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{begin_row}+{{c}}{R['ncf']}"),
                   bold=True, band=True)
    R["end"] = r
    r += 1
    _v2_series_row(
        ws, ctx, r, "Runway months", "months",
        formulas=_v2_formulas(
            ctx,
            f"=IF({{c}}{R['fcf']}>=0,100-1,MIN(100-1,MAX(0,{{c}}{R['end']})/(ABS({{c}}{R['fcf']})/{{c}}$5)))"),
        driver="cash / monthly burn")
    R["runway"] = r
    r += 1
    _v2_check_row(
        ws, ctx, r, "Cash shortfall check",
        [
            f"=IF('Assumptions'!$F${A['scenario_toggle']}=2,"
            f"ROUND(MAX(0,-MIN($F${R['end']}:${last}${R['end']})),0),0)"
        ],
        "CF: no period ends cash-negative at the Base toggle",
        cols=[V2_FIRST_PERIOD_COL],
        note="Base ケースで期末現金がマイナスの期を検出(Downside は Financing の funding gap 行)",
    )
    R["shortfall_check"] = r
    r += 1
    note_col = getattr(ws, "_startup_note_col", None)
    _v2_label_cells(ws, r, "Simplifications", "", "",
                    "消費税・源泉/社保・前受は残高方式で簡略計上。法人税はNOLを考慮しない期間独立課税。",
                    note_col)
    r += 2

    chart_cols = (
        (V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + ctx.axis.monthly_count - 1)
        if ctx.axis.monthly_count
        else (V2_FIRST_PERIOD_COL, V2_FIRST_PERIOD_COL + ctx.n_cols - 1)
    )
    cats = Reference(ws, min_col=chart_cols[0], max_col=chart_cols[1], min_row=ib.HEADER_PERIOD_ROW)
    data = Reference(ws, min_col=chart_cols[0], max_col=chart_cols[1],
                     min_row=R["end"], max_row=R["end"])
    _add_line_chart(
        ws,
        "Ending cash (monthly window)" if ctx.axis.monthly_count else "Ending cash",
        data, cats, f"B{r + 1}", unit_money,
    )


def _build_bs_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["BS"]
    A = ctx.rows["Assumptions"]
    CF = ctx.rows["CF"]
    PL = ctx.rows["P&L"]
    scale, (fmt_money, unit_money) = _v2_statement_scale(ctx)
    ctx.sheet_scale["BS"] = scale
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Balance Sheet",
        "Simplified balance sheet: cash and working-capital balances reference the CF sheet; "
        "PP&E, debt, and equity roll forward. The balance check must read OK in every column.",
        unit_caption=_v2_caption(ctx, unit_money),
    )
    R: dict = {}
    ctx.rows["BS"] = R
    r = ib.DATA_START_ROW_V2

    _v2_section(ws, ctx, r, "Assets")
    r += 1
    asset_first = r
    _v2_series_row(ws, ctx, r, "Cash", "money",
                   formulas=_v2_formulas(ctx, f"='CF'!{{c}}{CF['end']}"), driver="CF")
    R["cash"] = r
    r += 1
    for key, label in (("ar", "Accounts receivable"), ("inv", "Inventory / WIP")):
        if f"bal_{key}" in CF:
            _v2_series_row(ws, ctx, r, label, "money",
                           formulas=_v2_formulas(ctx, f"='CF'!{{c}}{CF[f'bal_{key}']}"),
                           driver="CF balance")
            R[key] = r
            r += 1
    _v2_series_row(ws, ctx, r, "Total current assets", "money",
                   formulas=_v2_formulas(ctx, f"=SUM({{c}}{asset_first}:{{c}}{r - 1})"), bold=True)
    R["current_assets"] = r
    r += 1
    has_ppe = "cum_capex" in PL
    if has_ppe:
        _v2_series_row(ws, ctx, r, "Gross PP&E", "money",
                       formulas=_v2_formulas(ctx, f"='P&L'!{{c}}{PL['cum_capex']}"),
                       driver="cum CapEx")
        R["gross_ppe"] = r
        r += 1
        if "da" in PL:
            _v2_series_row(ws, ctx, r, "Accumulated D&A", "money",
                           formulas=_v2_formulas(ctx, f"=N({{p}}{r})+'P&L'!{{c}}{PL['da']}"),
                           driver="roll-forward")
            R["accum_da"] = r
            r += 1
        net_formula = (
            f"={{c}}{R['gross_ppe']}-{{c}}{R['accum_da']}" if "accum_da" in R
            else f"={{c}}{R['gross_ppe']}"
        )
        _v2_series_row(ws, ctx, r, "Net PP&E", "money",
                       formulas=_v2_formulas(ctx, net_formula), bold=True)
        R["net_ppe"] = r
        r += 1
    total_assets_formula = (
        f"={{c}}{R['current_assets']}+{{c}}{R['net_ppe']}" if has_ppe
        else f"={{c}}{R['current_assets']}"
    )
    _v2_series_row(ws, ctx, r, "Total assets", "money",
                   formulas=_v2_formulas(ctx, total_assets_formula), bold=True, band=True)
    R["assets"] = r
    r += 2

    _v2_section(ws, ctx, r, "Liabilities")
    r += 1
    liab_first = r
    for key, label in (
        ("ap", "Accounts payable"),
        ("deferred", "Deferred revenue"),
        ("advances", "Customer advances (前受金)"),
        ("ctax", "未払消費税等"),
        ("withholding", "預り金・未払社会保険料"),
    ):
        if f"bal_{key}" in CF:
            _v2_series_row(ws, ctx, r, label, "money",
                           formulas=_v2_formulas(ctx, f"='CF'!{{c}}{CF[f'bal_{key}']}"),
                           driver="CF balance")
            R[key] = r
            r += 1
    if "debt_balance" in PL:
        _v2_series_row(ws, ctx, r, "Debt balance", "money",
                       formulas=_v2_formulas(ctx, f"='P&L'!{{c}}{PL['debt_balance']}"),
                       driver="P&L reference")
        R["debt"] = r
        r += 1
    liab_formula = (
        f"=SUM({{c}}{liab_first}:{{c}}{r - 1})" if r > liab_first else "=0"
    )
    _v2_series_row(ws, ctx, r, "Total liabilities", "money",
                   formulas=_v2_formulas(ctx, liab_formula), bold=True)
    R["liabilities"] = r
    r += 2

    _v2_section(ws, ctx, r, "Equity")
    r += 1
    paid_in_terms = f"+IF(N({{p}}$5)=0,'Assumptions'!$F${A['beginning_cash']},0)"
    for key, sign in (("equity_raise", "+"), ("grants", "+"), ("secondary", "-")):
        if key in A:
            paid_in_terms += f"{sign}'Assumptions'!{{c}}{A[key]}"
    _v2_series_row(ws, ctx, r, "Paid-in capital & opening balance", "money",
                   formulas=_v2_formulas(ctx, f"=N({{p}}{r}){paid_in_terms}"),
                   driver="roll-forward",
                   note="設立資本(期首現金)+増資+補助金−セカンダリー")
    R["paid_in"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Retained earnings", "money",
                   formulas=_v2_formulas(ctx, f"=N({{p}}{r})+'P&L'!{{c}}{PL['ni']}"),
                   driver="cum NI")
    R["retained"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Total equity", "money",
                   formulas=_v2_formulas(ctx, f"={{c}}{R['paid_in']}+{{c}}{R['retained']}"),
                   bold=True)
    R["equity"] = r
    r += 2
    _v2_check_row(
        ws, ctx, r, "Balance check",
        _v2_formulas(
            ctx,
            f"=ROUND({{c}}{R['assets']}-{{c}}{R['liabilities']}-{{c}}{R['equity']},0)"),
        "BS: assets = liabilities + equity in every period",
        driver="A − L − E",
    )
    R["balance_check"] = r


def _bs_is_material(facts: SourceFacts) -> bool:
    """BS inclusion predicate for the full bundle (BLUEPRINT_S3 条件付き):
    include when capex / debt-like funding / inventory / working-capital
    terms are material. The default demo profile qualifies."""
    return bool(
        _v2_any(facts.capex_per_unit_yen)
        or _v2_any(facts.other_capex_yen)
        or _v2_any(facts.debt_raise_yen)
        or _v2_any(facts.convertibles_yen)
        or _v2_any(facts.lease_financing_yen)
        or _v2_any(facts.inventory_wip_pct_capex)
        or float(getattr(facts, "ar_site_months", 0.0) or 0.0) > 0
        or _v2_any(facts.ar_days)
        or _v2_any(facts.deferred_revenue_share)
    )


def full_bundle_for_facts(facts: SourceFacts, seeds: list) -> list:
    """Drop BS from the full bundle when the balance-sheet drivers are
    immaterial (mode bundles that list BS explicitly are not filtered)."""
    if _bs_is_material(facts):
        return list(seeds)
    return [sheet for sheet in seeds if sheet != "BS"]


def _build_summary_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Summary"]
    A = ctx.rows.get("Assumptions", {})
    annual_axis = _v2_annual_rollup_axis(ctx.axis)
    sctx = BuildContext(
        facts=facts,
        axis=annual_axis,
        bundle=ctx.bundle,
        sheet_scale=ctx.sheet_scale,
        checks=ctx.checks,
        rows=ctx.rows,
        snapshots=ctx.snapshots,
    )
    base_rev = ctx.snapshots.setdefault("base_revenue_annual", _v2_base_revenue_annual(facts))
    rev_fy = _v2_fy_flow_rollup(ctx, base_rev)
    proj = ctx.snapshots.setdefault("base_projection", project_plan_free_cash_flow(facts))
    scale = ib.pick_sheet_scale([abs(float(v)) for v in rev_fy] or [0.0], currency=ctx.currency)
    scale = _v2_guard_scale_readable(scale, rev_fy, ctx.currency)
    ctx.sheet_scale["Summary"] = scale
    fmt_money, unit_money = _v2_money_fmt_unit(ctx, scale)
    _setup_sheet_v2(
        ws, sctx, f"{facts.company} — Summary",
        "Annual roll-up of the model (declared annual sheet — months ruler shows the FY month "
        "count): condensed P&L, cash & runway, KPIs, scenario comparison, and the master check.",
        unit_caption=_v2_caption(ctx, unit_money),
        axis=annual_axis,
    )
    R: dict = {}
    ctx.rows["Summary"] = R
    source_last = ctx.last_period_letter
    n_fy = len(annual_axis.labels)
    last_fy_letter = get_column_letter(V2_FIRST_PERIOD_COL + n_fy - 1)
    mk = mechanic_key(facts)
    engines = {"Revenue Build", "Cost Build", "People Plan"} <= ctx.bundle
    has_pl = "P&L" in ctx.bundle
    has_cf = "CF" in ctx.bundle
    cf_compact = ctx.rows.get("CF", {}).get("compact") if has_cf else None
    live = has_pl or bool(cf_compact)

    def sumif(sheet: str, row: int) -> str:
        return (
            f"=SUMIF('{sheet}'!$F$4:${source_last}$4,{{c}}$4,"
            f"'{sheet}'!$F${row}:${source_last}${row})"
        )

    def fy_end_pull(sheet: str, row: int) -> str:
        """INDEX at the last column of the Summary column's fiscal year.

        MATCH-free (FY ruler is text): COUNTIF over the source FY ruler with
        "<=" counts every column up to the FY end — no hardcoded positions."""
        return (
            f"INDEX('{sheet}'!$F${row}:${source_last}${row},"
            f"COUNTIF('{sheet}'!$F$4:${source_last}$4,\"<=\"&{{c}}$4))"
        )

    r = ib.DATA_START_ROW_V2
    _v2_section(ws, sctx, r, "Condensed P&L (annual)")
    r += 1
    if has_pl:
        PL = ctx.rows["P&L"]
        pl_source = [
            ("rev", "Revenue", sumif("P&L", PL["rev"]), True),
            ("gp", "Gross profit", sumif("P&L", PL["gp"]), True),
            ("opex", "Total OpEx", sumif("P&L", PL["opex"]), False),
            ("ebitda", "EBITDA", sumif("P&L", PL["ebitda"]), True),
            ("ni", "Net income", sumif("P&L", PL["ni"]), True),
        ]
    elif cf_compact:
        pl_source = [
            ("rev", "Revenue", sumif("CF", cf_compact["rev"]), True),
            ("gp", "Gross profit", sumif("CF", cf_compact["gp"]), True),
            ("ebitda", "EBITDA (compact)", sumif("CF", cf_compact["ebitda"]), True),
        ]
    else:
        pl_source = None
    if pl_source:
        for key, label, formula, bold in pl_source:
            _v2_series_row(ws, sctx, r, label, "money",
                           formulas=_v2_formulas(sctx, formula),
                           driver="SUMIF FY", bold=bold, band=key == "ebitda")
            R[key] = r
            r += 1
    else:
        ebitda_fy = _v2_fy_flow_rollup(ctx, [p["ebitda"] for p in proj])
        gp_fy = _v2_fy_flow_rollup(
            ctx, [base_rev[i] * facts.target_gross_margin[i] for i in range(len(base_rev))])
        for key, label, series in (
            ("rev", "Revenue", rev_fy),
            ("gp", "Gross profit", gp_fy),
            ("ebitda", "EBITDA", ebitda_fy),
        ):
            _v2_series_row(ws, sctx, r, label, "money",
                           values=[round(v) for v in series], snapshot=True,
                           driver="kernel plan",
                           note="generated snapshot — no statement sheets in this bundle"
                           if key == "rev" else "")
            R[key] = r
            r += 1
    _v2_series_row(ws, sctx, r, "EBITDA margin", "pct",
                   formulas=_v2_formulas(
                       sctx, f"=IF({{c}}{R['rev']}=0,0,{{c}}{R['ebitda']}/{{c}}{R['rev']})"))
    R["ebitda_margin"] = r
    r += 2

    if has_cf:
        CF = ctx.rows["CF"]
        _v2_section(ws, sctx, r, "Cash & runway")
        r += 1
        _v2_series_row(ws, sctx, r, "Ending cash (FY end)", "money",
                       formulas=_v2_formulas(sctx, f"={fy_end_pull('CF', CF['end'])}"),
                       driver="FY-end pull", bold=True)
        R["cash"] = r
        r += 1
        _v2_series_row(ws, sctx, r, "Net burn (FY, −FCF)", "money",
                       formulas=_v2_formulas(
                           sctx, f"=-SUMIF('CF'!$F$4:${source_last}$4,{{c}}$4,"
                                 f"'CF'!$F${CF['fcf']}:${source_last}${CF['fcf']})"),
                       note="positive = cash consumed by operations + CapEx")
        R["burn"] = r
        r += 1
        _v2_series_row(ws, sctx, r, "Runway months (FY end)", "months",
                       formulas=_v2_formulas(sctx, f"={fy_end_pull('CF', CF['runway'])}"),
                       driver="FY-end pull")
        R["runway"] = r
        r += 2

    _v2_section(ws, sctx, r, "KPI")
    r += 1
    RB = ctx.rows.get("Revenue Build", {})
    CB = ctx.rows.get("Cost Build", {})
    PP = ctx.rows.get("People Plan", {})
    if engines and facts.revenue_mode == "recurring" and "ending" in RB and "price" in RB:
        _v2_series_row(ws, sctx, r, "ARR run-rate (FY end)", "money",
                       formulas=_v2_formulas(
                           sctx,
                           f"={fy_end_pull('Revenue Build', RB['ending'])}"
                           f"*{fy_end_pull('Revenue Build', RB['price'])}*12"),
                       driver="MRR × 12", note="FY-end units × monthly price × 12")
        R["arr"] = r
        r += 1
    _v2_series_row(ws, sctx, r, "Revenue growth (YoY)", "pct",
                   formulas=_v2_formulas(
                       sctx,
                       f'=IF(N({{p}}{R["rev"]})=0,"-",{{c}}{R["rev"]}/N({{p}}{R["rev"]})-1)'),
                   note="benchmark: T2D3-class growth for venture-scale plans")
    R["growth"] = r
    r += 1
    _v2_series_row(ws, sctx, r, "Gross margin", "pct",
                   formulas=_v2_formulas(
                       sctx, f"=IF({{c}}{R['rev']}=0,0,{{c}}{R['gp']}/{{c}}{R['rev']})"),
                   note="compare against the target gross margin driver on Assumptions")
    R["gm"] = r
    r += 1
    if "net_retention" in A and facts.revenue_mode == "recurring":
        _v2_series_row(ws, sctx, r, "Net revenue retention", "pct",
                       formulas=_v2_formulas(
                           sctx, f"={fy_end_pull('Assumptions', A['net_retention'])}"),
                       driver="cohort driver")
        R["nrr"] = r
        r += 1
    if engines and "total" in PP:
        _v2_series_row(ws, sctx, r, "Revenue / FTE", "money_row",
                       formulas=_v2_formulas(
                           sctx,
                           f"=IF({fy_end_pull('People Plan', PP['total'])}=0,0,"
                           f"{{c}}{R['rev']}/{fy_end_pull('People Plan', PP['total'])})"),
                       driver="FY-end HC",
                       fmt=_v2_row_money(ctx, [rev_fy[-1] / 40])[0],
                       unit=_v2_row_money(ctx, [rev_fy[-1] / 40])[1],
                       note="benchmark: mature software ≈ ¥30-50M revenue per FTE")
        R["rev_per_fte"] = r
        r += 1
    if engines and mk == "marketplace":
        if "take" in RB:
            _v2_series_row(ws, sctx, r, "Take rate (FY end)", "pct",
                           formulas=_v2_formulas(
                               sctx, f"={fy_end_pull('Revenue Build', RB['take'])}"))
            r += 1
        if "gmv" in RB and "ending" in RB:
            _v2_series_row(ws, sctx, r, "GMV per customer", "money_row",
                           formulas=_v2_formulas(
                               sctx,
                               f"=IF({fy_end_pull('Revenue Build', RB['ending'])}=0,0,"
                               f"SUMIF('Revenue Build'!$F$4:${source_last}$4,{{c}}$4,"
                               f"'Revenue Build'!$F${RB['gmv']}:${source_last}${RB['gmv']})"
                               f"/{fy_end_pull('Revenue Build', RB['ending'])})"),
                           fmt=_v2_row_money(ctx, [facts.gmv_yen[-1] / max(facts.customers[-1], 1)])[0],
                           unit=_v2_row_money(ctx, [facts.gmv_yen[-1] / max(facts.customers[-1], 1)])[1])
            r += 1
        if "variable" in CB:
            _v2_series_row(ws, sctx, r, "Contribution margin", "pct",
                           formulas=_v2_formulas(
                               sctx,
                               f"=IF({{c}}{R['rev']}=0,0,({{c}}{R['rev']}"
                               f"-SUMIF('Cost Build'!$F$4:${source_last}$4,{{c}}$4,"
                               f"'Cost Build'!$F${CB['variable']}:${source_last}${CB['variable']}))"
                               f"/{{c}}{R['rev']})"),
                           note="revenue net of variable cost — transaction lens")
            r += 1
    vc_apply = engines and mk not in ("generic", "pre_revenue_milestone")
    if vc_apply:
        _v2_section(ws, sctx, r, "VC decision metrics")
        r += 1
        _v2_series_row(ws, sctx, r, "Rule of 40", "pct",
                       formulas=_v2_formulas(
                           sctx,
                           f'=IF({{c}}{R["growth"]}="-","-",'
                           f"{{c}}{R['growth']}+{{c}}{R['ebitda_margin']})"),
                       note="revenue growth + EBITDA margin — composed from the live rows above")
        R["rule40"] = r
        r += 1
        if "sm" in CB and "new" in RB:
            new_fy = (
                f"SUMIF('Revenue Build'!$F$4:${source_last}$4,{{c}}$4,"
                f"'Revenue Build'!$F${RB['new']}:${source_last}${RB['new']})"
            )
            sm_fy = (
                f"SUMIF('Cost Build'!$F$4:${source_last}$4,{{c}}$4,"
                f"'Cost Build'!$F${CB['sm']}:${source_last}${CB['sm']})"
            )
            _v2_series_row(ws, sctx, r, "Customer acquisition cost", "money_row",
                           formulas=_v2_formulas(
                               sctx, f'=IF({new_fy}<=0,"N/A",{sm_fy}/{new_fy})'),
                           driver="S&M / new units",
                           fmt=_v2_row_money(ctx, [facts.monthly_price_yen[-1] * 12 or 100000])[0],
                           unit=_v2_row_money(ctx, [facts.monthly_price_yen[-1] * 12 or 100000])[1])
            R["cac"] = r
            r += 1
            if "ending" in RB:
                ending_pull = fy_end_pull("Revenue Build", RB["ending"])
                _v2_series_row(ws, sctx, r, "CAC payback", "months",
                               formulas=_v2_formulas(
                                   sctx,
                                   f'=IF(OR({{c}}{R["cac"]}="N/A",{{c}}{R["gm"]}<=0,{ending_pull}=0),"N/A",'
                                   f"{{c}}{R['cac']}/(({{c}}{R['rev']}/{ending_pull})*{{c}}{R['gm']}/12))"),
                               note="CAC ÷ monthly gross profit per customer")
                R["cac_payback"] = r
                r += 1
            prior_sm = (
                f"SUMIF('Cost Build'!$F$4:${source_last}$4,{{p}}$4,"
                f"'Cost Build'!$F${CB['sm']}:${source_last}${CB['sm']})"
            )
            _v2_series_row(ws, sctx, r, "Magic number", "x",
                           formulas=_v2_formulas(
                               sctx,
                               f'=IF({prior_sm}<=0,"N/A",'
                               f"({{c}}{R['rev']}-N({{p}}{R['rev']}))/{prior_sm})"),
                           note="net new revenue ÷ prior-FY S&M; first FY has no prior S&M base")
            R["magic"] = r
            r += 1
        if has_cf:
            CF = ctx.rows["CF"]
            fcf_fy = (
                f"SUMIF('CF'!$F$4:${source_last}$4,{{c}}$4,"
                f"'CF'!$F${CF['fcf']}:${source_last}${CF['fcf']})"
            )
            _v2_series_row(ws, sctx, r, "Burn multiple", "x",
                           formulas=_v2_formulas(
                               sctx,
                               f'=IF({fcf_fy}>=0,0,IF(({{c}}{R["rev"]}-N({{p}}{R["rev"]}))<=0,"N/A",'
                               f"ABS({fcf_fy})/({{c}}{R['rev']}-N({{p}}{R['rev']}))))"),
                           note="net burn ÷ net new revenue; first FY nets against a zero base")
            R["burn_multiple"] = r
            r += 1
    r += 1

    _v2_section(ws, sctx, r, "Scenario comparison")
    r += 1
    ib.apply_semantic_fill_span(ws, r, 2, 8, ib.BG_TABLE_HEADER, bottom=ib.THIN_LINE,
                                border_start_col=3)
    case_cell = ws.cell(r, 3, "Case (generator snapshot)")
    case_cell.font = ib.FONT_BODY_BOLD
    case_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for col, label in zip((6, 7, 8), ("Downside", "Base", "Upside")):
        ib.apply_year_header(ws.cell(r, col, label), label)
    r += 1
    cases = _v2_cases(ctx)
    snap_rows = {}
    for key, label, extractor in (
        ("rev5", f"Revenue ({annual_axis.labels[-1]})",
         lambda case: _v2_fy_flow_rollup(ctx, case["revenue"])[-1]),
        ("ebitda5", f"EBITDA ({annual_axis.labels[-1]})",
         lambda case: _v2_fy_flow_rollup(ctx, case["ebitda"])[-1]),
        ("min_cash", "Minimum ending cash", lambda case: min(case["ending_cash"])),
        ("funding_gap", "Additional funding needed",
         lambda case: max(0.0, -min(case["ending_cash"]))),
    ):
        _v2_series_row(ws, sctx, r, label, "money",
                       values=[round(extractor(cases["down"])), round(extractor(cases["base"])),
                               round(extractor(cases["up"]))],
                       snapshot=True, driver="kernel case", cols=[6, 7, 8],
                       note="generated snapshot — rerun the generator after changing inputs"
                       if key == "rev5" else "")
        snap_rows[key] = r
        R[f"snap_{key}"] = r
        r += 1
    if "founder_fd_final" in ctx.snapshots and ctx.snapshots["founder_fd_final"] is not None:
        fd = round(float(ctx.snapshots["founder_fd_final"]), 4)
        _v2_series_row(ws, sctx, r, "Founder ownership (FD, post plan rounds)", "pct",
                       values=[fd, fd, fd], snapshot=True, cols=[6, 7, 8],
                       driver="cap table",
                       note="funding plan is held fixed across cases — the downside gap row above "
                            "shows the extra dilution pressure")
        R["snap_founder_fd"] = r
        r += 1
    if live and "scenario_toggle" in A:
        _v2_check_row(
            ws, sctx, r, "Snapshot staleness check",
            [
                f"=IF('Assumptions'!$F${A['scenario_toggle']}=2,"
                f"ROUND(MAX(0,ABS(${last_fy_letter}${R['rev']}-$G${snap_rows['rev5']})"
                f"-$G${snap_rows['rev5']}/100),0),0)"
            ],
            "Summary: snapshot Base final-year revenue vs live roll-up (±1%, Base toggle)",
            cols=[V2_FIRST_PERIOD_COL],
            note="verifies the generated snapshot against the live model while the toggle sits on Base",
        )
        R["staleness_check"] = r
        r += 1
    r += 1

    if engines:
        _v2_section(ws, sctx, r, "Sensitivity (snapshot)")
        r += 1
        roles = _v2_scenario_roles(facts)
        d_band = roles["demand"][:3]
        p_band = roles["price"][:3]
        header_cell = ws.cell(r, 3, f"EBITDA ({annual_axis.labels[-1]}) — demand × price")
        header_cell.font = ib.FONT_BODY_BOLD
        header_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for col, p_scale in zip((6, 7, 8), p_band):
            ib.apply_year_header(ws.cell(r, col, f"price ×{p_scale:g}"), "")
        r += 1
        for d_scale in d_band:
            _v2_label_cells(ws, r, f"demand ×{d_scale:g}", "", unit_money, "", ws._startup_note_col)
            for col, p_scale in zip((6, 7, 8), p_band):
                value = round(_v2_fy_flow_rollup(
                    ctx,
                    _v2_case_projection(facts, d_scale, p_scale, 1.0, 1.0)["ebitda"])[-1])
                cell = ws.cell(r, col, value)
                cell.number_format = fmt_money
                cell.font = _FONT_SNAPSHOT_V2
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            r += 1
        _v2_label_cells(ws, r, "Sensitivity note", "", "",
                        "kernel snapshot at Base cost/opex; scales are the scenario bands — "
                        "rerun the generator after changing inputs", ws._startup_note_col)
        r += 2

    _v2_section(ws, sctx, r, "Cross-check")
    r += 1
    note_col = ws._startup_note_col
    if "Evidence" in ctx.bundle and "som_cell" in ctx.rows.get("Evidence", {}):
        som_ref = f"'Evidence'!{ctx.rows['Evidence']['som_cell']}"
        _v2_series_row(ws, sctx, r, "SOM (plan-case market)", "money_row",
                       formulas=[f"={som_ref}"], cols=[V2_FIRST_PERIOD_COL],
                       fmt=_v2_row_money(ctx, [facts.som_yen])[0],
                       unit=_v2_row_money(ctx, [facts.som_yen])[1],
                       driver="Evidence",
                       note="market-sanity input lives on Evidence; never feeds the revenue chain")
    else:
        _v2_series_row(ws, sctx, r, "SOM (plan-case market)", "money_row",
                       values=[facts.som_yen], cols=[V2_FIRST_PERIOD_COL],
                       fmt=_v2_row_money(ctx, [facts.som_yen])[0],
                       unit=_v2_row_money(ctx, [facts.som_yen])[1],
                       driver="market sanity",
                       note="isolated market-sanity input — never feeds the revenue chain")
    som_row = r
    r += 1
    _v2_series_row(ws, sctx, r, "Final-year revenue / SOM", "pct",
                   formulas=[
                       f'=IF($F${som_row}=0,"-",${last_fy_letter}${R["rev"]}/$F${som_row})'
                   ],
                   cols=[V2_FIRST_PERIOD_COL],
                   note="a final-year share above ~30% of SOM needs explicit share-gain evidence")
    r += 1
    if "rev_per_fte" in R:
        bench_value = 250_000 if ctx.currency == "USD" else 30_000_000
        bench_fmt, bench_unit = _v2_row_money(ctx, [bench_value])
        _v2_series_row(ws, sctx, r, "Revenue / FTE benchmark", "money_row",
                       values=[bench_value], cols=[V2_FIRST_PERIOD_COL],
                       fmt=bench_fmt, unit=bench_unit,
                       driver="benchmark input",
                       note=("mature software ≈ $250K revenue per FTE (stage benchmark; "
                             "evidence status: benchmark — replace with your peer set)"
                             if ctx.currency == "USD" else
                             "JP SaaS median ≈ ¥30M revenue per FTE (stage benchmark; "
                             "evidence status: benchmark — replace with your peer set)"))
        bench_row = r
        r += 1
        _v2_series_row(ws, sctx, r, "Revenue / FTE vs benchmark", "status",
                       formulas=[
                           f'=IF(${last_fy_letter}${R["rev_per_fte"]}>=$F${bench_row},'
                           f'"above benchmark","below benchmark")'
                       ],
                       cols=[V2_FIRST_PERIOD_COL],
                       driver=f"{annual_axis.labels[-1]} KPI",
                       note="final-year Revenue/FTE vs the benchmark input — below benchmark "
                            "means the people plan grows faster than revenue supports")
        R["rev_per_fte_vs_benchmark"] = r
        r += 1
    else:
        _v2_label_cells(ws, r, "Revenue / FTE vs benchmark", "", "",
                        "engine sheets are outside this bundle — compare Revenue/FTE against "
                        "stage benchmarks before circulation", note_col)
        r += 1
    cagr_label = f"Revenue CAGR ({annual_axis.labels[0]}→{annual_axis.labels[-1]})"
    _v2_series_row(ws, sctx, r, cagr_label, "pct",
                   formulas=[
                       f'=IF(OR($F${R["rev"]}<=0,COUNT($F$5:${last_fy_letter}$5)<=1),"-",'
                       f"(${last_fy_letter}${R['rev']}/$F${R['rev']})"
                       f"^(1/(COUNT($F$5:${last_fy_letter}$5)-1))-1)"
                   ],
                   cols=[V2_FIRST_PERIOD_COL],
                   driver="model CAGR",
                   note="compound growth across the plan horizon — computed from the live revenue row")
    cagr_row = r
    R["revenue_cagr"] = r
    r += 1
    _v2_series_row(ws, sctx, r, "Market growth (CAGR)", "pct",
                   values=[0.10], cols=[V2_FIRST_PERIOD_COL],
                   driver="benchmark input",
                   note="default 10%/yr market-growth placeholder (evidence status: estimate) — "
                        "replace with a sourced market CAGR for the target segment")
    market_growth_row = r
    r += 1
    _v2_series_row(ws, sctx, r, "Company growth vs market growth", "status",
                   formulas=[
                       f'=IF($F${cagr_row}="-","-",'
                       f'IF($F${cagr_row}>$F${market_growth_row},'
                       f'"above market — share gain implied","at/below market growth"))'
                   ],
                   cols=[V2_FIRST_PERIOD_COL],
                   driver="CAGR vs input",
                   note="growth above market implies share gain — tie to demand-support "
                        "coverage on Revenue Build")
    R["growth_vs_market"] = r
    r += 2

    _v2_section(ws, sctx, r, "Consolidated checks")
    r += 1
    first_check_row = r
    for entry in list(ctx.checks):
        formula = (
            f"=SUMPRODUCT(ABS('{entry.sheet}'!{entry.ref}))"
            if ":" in entry.ref
            else f"=ABS('{entry.sheet}'!{entry.ref})"
        )
        _v2_series_row(ws, sctx, r, f"Check — {entry.sheet}", "check",
                       formulas=[formula], cols=[V2_FIRST_PERIOD_COL],
                       driver="0 = OK", note=entry.description)
        r += 1
    if r > first_check_row:
        master_formula = f"=SUM($F${first_check_row}:$F${r - 1})"
        master_note = "0 = every registered check passes; echoed in row 2 of every period sheet"
    else:
        master_formula = "=0"
        master_note = "no machine checks registered in this bundle"
    _v2_series_row(ws, sctx, r, "Master check", "check",
                   formulas=[master_formula],
                   cols=[V2_FIRST_PERIOD_COL], bold=True, band=True,
                   note=master_note)
    ctx.master_check_cell = f"$F${r}"
    R["master_check"] = r
    r += 2

    _v2_section(ws, sctx, r, "Recommendation")
    r += 1
    divisor = 1_000_000.0
    money_suffix = "M"
    base_case = cases["base"]
    down_case = cases["down"]
    runway_target = facts.target_min_runway_months[0] if facts.target_min_runway_months else 18
    recommendation_rows = [
        ("Funding readout",
         f"Base minimum ending cash {min(base_case['ending_cash']) / divisor:,.0f}{money_suffix}; "
         f"downside additional need {max(0.0, -min(down_case['ending_cash'])) / divisor:,.0f}{money_suffix} — "
         f"hold ≥ {runway_target} months runway before circulating."),
        ("Growth readout",
         f"Revenue plan {rev_fy[0] / divisor:,.0f}{money_suffix} → {rev_fy[-1] / divisor:,.0f}{money_suffix} "
         f"({annual_axis.labels[0]} → {annual_axis.labels[-1]}); validate demand evidence before scaling spend."),
        ("Weakest evidence (top 3)",
         "; ".join(facts.source_unknowns[:3]) if facts.source_unknowns else "none listed"),
        ("Next diligence",
         "pricing / WTP proof, cohort churn, cost-to-serve decomposition, financing terms."),
    ]
    for label, text in recommendation_rows:
        _v2_label_cells(ws, r, label, "", "", text, note_col)
        r += 1


# --- financing / registers / conditional sheets -------------------------------


def _v2_cases(ctx: BuildContext) -> dict:
    """Kernel re-projection per scenario case, cached per build."""
    if "cases" not in ctx.snapshots:
        roles = _v2_scenario_roles(ctx.facts)
        ctx.snapshots["cases"] = {
            name: _v2_case_projection(
                ctx.facts,
                roles["demand"][idx], roles["price"][idx],
                roles["cost"][idx], roles["opex"][idx],
            )
            for idx, name in ((0, "down"), (1, "base"), (2, "up"))
        }
    return ctx.snapshots["cases"]


def _build_financing_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Financing"]
    A = ctx.rows["Assumptions"]
    CF = ctx.rows["CF"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Financing",
        "Sources & Uses (cumulative, ties to the CF cash walk), instrument terms, "
        "post-raise runway verification, and the downside funding gap.",
        unit_caption=_v2_caption(ctx, None),
    )
    R: dict = {}
    ctx.rows["Financing"] = R
    last = ctx.last_period_letter
    proj = ctx.snapshots.setdefault("base_projection", project_plan_free_cash_flow(facts))
    rep_money = [abs(float(v)) for v in (
        *facts.equity_raise_yen, *facts.debt_raise_yen, facts.beginning_cash_yen)]
    fin_fmt, fin_unit = _v2_row_money(ctx, rep_money or [0.0])
    r = ib.DATA_START_ROW_V2

    _v2_section(ws, ctx, r, "Sources")
    r += 1
    src_first = r
    for key, label in (
        ("fin_equity_raise", "Equity financing"),
        ("fin_debt_raise", "Debt draw"),
        ("fin_grants", "Grants / subsidies"),
        ("fin_convertibles", "J-KISS / convertible notes"),
        ("fin_lease", "Lease financing"),
        ("fin_secondary", "Founder / investor secondary (use)"),
        ("fin_amortization", "Debt principal repayment"),
    ):
        if key in CF:
            _v2_series_row(ws, ctx, r, label, "money_row",
                           formulas=_v2_formulas(ctx, f"='CF'!{{c}}{CF[key]}"),
                           driver="CF financing", fmt=fin_fmt, unit=fin_unit)
            R[key] = r
            r += 1
    if r > src_first:
        total_sources = f"=SUM({{c}}{src_first}:{{c}}{r - 1})"
    else:
        total_sources = "=0"
    _v2_series_row(ws, ctx, r, "Total sources (net of repayment)", "money_row",
                   formulas=_v2_formulas(ctx, total_sources), bold=True, band=True,
                   fmt=fin_fmt, unit=fin_unit,
                   note="customer advances are excluded — they are working-capital timing, not funding")
    R["total_sources"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Cumulative sources + opening cash", "money_row",
                   formulas=_v2_formulas(
                       ctx,
                       f"=N({{p}}{r})+{{c}}{R['total_sources']}"
                       f"+IF(N({{p}}$5)=0,'Assumptions'!$F${A['beginning_cash']},0)"),
                   driver="roll-forward", fmt=fin_fmt, unit=fin_unit)
    R["cum_sources"] = r
    r += 2

    _v2_section(ws, ctx, r, "Uses")
    r += 1
    _v2_series_row(ws, ctx, r, "Cumulative operating cash consumed", "money_row",
                   formulas=_v2_formulas(ctx, f"=N({{p}}{r})-'CF'!{{c}}{CF['ocf']}"),
                   driver="− cum OCF", fmt=fin_fmt, unit=fin_unit,
                   note="negative when operations have generated net cash")
    R["cum_burn"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Cumulative CapEx", "money_row",
                   formulas=_v2_formulas(ctx, f"=N({{p}}{r})-'CF'!{{c}}{CF['capex']}"),
                   driver="− CF CapEx", fmt=fin_fmt, unit=fin_unit)
    R["cum_capex"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Cash on hand (ending)", "money_row",
                   formulas=_v2_formulas(ctx, f"='CF'!{{c}}{CF['end']}"),
                   driver="CF", fmt=fin_fmt, unit=fin_unit)
    R["cash"] = r
    r += 1
    _v2_series_row(ws, ctx, r, "Total uses + cash on hand", "money_row",
                   formulas=_v2_formulas(
                       ctx, f"={{c}}{R['cum_burn']}+{{c}}{R['cum_capex']}+{{c}}{R['cash']}"),
                   bold=True, fmt=fin_fmt, unit=fin_unit)
    R["total_uses"] = r
    r += 1
    _v2_check_row(
        ws, ctx, r, "Sources = Uses check",
        _v2_formulas(ctx, f"=ROUND({{c}}{R['cum_sources']}-{{c}}{R['total_uses']},0)"),
        "Financing: cumulative sources + opening cash tie to uses + cash on hand",
        driver="S − U",
    )
    R["su_check"] = r
    r += 2

    _v2_section(ws, ctx, r, "Instrument terms")
    r += 1
    note_col = ws._startup_note_col
    rate = facts.debt_interest_rate[0] if facts.debt_interest_rate else 0.0
    term_rows = []
    if _v2_any(facts.equity_raise_yen):
        term_rows.append(("Equity", "priced rounds — see Cap Table",
                          "round terms (pre/post, preference) belong on the Cap Table sheet"))
    if _v2_any(facts.debt_raise_yen):
        amort_note = (
            "contractual amortization stated" if _v2_any(facts.debt_amortization_yen)
            else "no amortization schedule stated — bullet assumed"
        )
        term_rows.append(("Debt", f"rate {rate:.1%} (annual); {amort_note}",
                          "maturity / covenants not stated — DD item"))
    if _v2_any(facts.convertibles_yen):
        term_rows.append(("J-KISS / convertibles", "principal stated",
                          "cap / discount not stated — modeled as potential dilution on Cap Table; DD item"))
    if _v2_any(facts.grants_yen):
        term_rows.append(("Grants", "non-dilutive", "milestone conditions not stated — DD item"))
    if _v2_any(facts.lease_financing_yen):
        term_rows.append(("Lease", "asset-backed", "implicit rate folded into debt interest"))
    if _v2_any(facts.customer_advances_yen):
        term_rows.append(("Customer advances", "working-capital timing",
                          "excluded from funding capacity; released the following FY"))
    if not term_rows:
        term_rows.append(("No financing instruments stated", "—",
                          "plan runs on opening cash only — confirm intent"))
    for label, terms, note in term_rows:
        _v2_label_cells(ws, r, label, terms, "", note, note_col)
        r += 1
    r += 1

    _v2_section(ws, ctx, r, "Runway & downside")
    r += 1
    raise_fy_idx = next(
        (i for i, v in enumerate(facts.equity_raise_yen) if v > 0), None)
    if raise_fy_idx is not None and "min_runway" in A:
        fy_order: list = []
        for fy in ctx.axis.fy_labels:
            if fy not in fy_order:
                fy_order.append(fy)
        raise_fy = fy_order[min(raise_fy_idx, len(fy_order) - 1)]
        raise_col_idx = ctx.axis.fy_labels.index(raise_fy)
        raise_col = V2_FIRST_PERIOD_COL + raise_col_idx
        # forward window: columns from the raise until >= 12 months
        months_acc = 0
        end_idx = raise_col_idx
        for idx in range(raise_col_idx, ctx.n_cols):
            months_acc += ctx.axis.months_in_period[idx]
            end_idx = idx
            if months_acc >= 12:
                break
        w_first = get_column_letter(raise_col)
        w_last = get_column_letter(V2_FIRST_PERIOD_COL + end_idx)
        fcf_row = CF["fcf"]
        window_fcf = f"SUM('CF'!{w_first}{fcf_row}:{w_last}{fcf_row})"
        window_months = f"SUM('CF'!{w_first}$5:{w_last}$5)"
        _v2_series_row(
            ws, ctx, r, "Post-raise runway", "months",
            formulas=[
                f"=IF({window_fcf}>=0,100-1,MIN(100-1,MAX(0,'CF'!{w_first}{CF['end']})"
                f"/(ABS({window_fcf})/{window_months})))"
            ],
            cols=[V2_FIRST_PERIOD_COL],
            driver=f"at {raise_fy} raise",
            note="post-raise cash ÷ average monthly burn over the following 12 months")
        R["post_raise_runway"] = r
        runway_row = r
        r += 1
        min_runway_cell = f"'Assumptions'!{get_column_letter(raise_col)}${A['min_runway']}"
        _v2_series_row(
            ws, ctx, r, "Runway after raise", "status",
            formulas=[
                f'=IF($F${runway_row}>={min_runway_cell},"meets target",'
                f'"below target — review round size")'
            ],
            cols=[V2_FIRST_PERIOD_COL],
            driver="vs target",
            note="target comparison is a decision readout; the machine check below "
                 "enforces the 12-month hard floor")
        R["runway_status"] = r
        r += 1
        _v2_check_row(
            ws, ctx, r, "Post-raise runway floor check",
            [
                f"=IF('Assumptions'!$F${A['scenario_toggle']}=2,"
                f"ROUND(MAX(0,MIN({min_runway_cell},12)-$F${runway_row}),0),0)"
            ],
            "Financing: post-raise runway clears the 12-month hard floor (Base toggle)",
            cols=[V2_FIRST_PERIOD_COL],
            note="0 = OK; positive = months short of the 12-month floor "
                 "(kernel sizing and the tax-inclusive cash walk can differ near the stated target)",
        )
        R["runway_check"] = r
        r += 1
    else:
        _v2_label_cells(ws, r, "Runway after raise", "no equity round scheduled", "",
                        "no round to verify — confirm the plan intends to run on existing cash",
                        note_col)
        r += 1
    _v2_series_row(
        ws, ctx, r, "Funding gap (live, current toggle)", "money_row",
        formulas=[f"=MAX(0,-MIN('CF'!$F${CF['end']}:${last}${CF['end']}))"],
        cols=[V2_FIRST_PERIOD_COL], fmt=fin_fmt, unit=fin_unit,
        driver="−MIN ending cash",
        note="toggle Downside on Assumptions to read the downside gap live")
    R["gap_live"] = r
    r += 1
    down_gap = max(0.0, -min(_v2_cases(ctx)["down"]["ending_cash"]))
    _v2_series_row(
        ws, ctx, r, "Downside funding gap", "money_row",
        values=[round(down_gap)], cols=[V2_FIRST_PERIOD_COL],
        snapshot=True, fmt=fin_fmt, unit=fin_unit, driver="kernel downside case",
        note="generated snapshot — rerun the generator after changing inputs")
    R["gap_downside"] = r
    r += 1


def _v2_cap_rounds(facts: SourceFacts) -> list:
    """Non-period financing rounds derived from the facts (founding excluded)."""
    rounds: list = []
    last_post = 0.0
    for idx, raise_yen in enumerate(facts.equity_raise_yen):
        if raise_yen <= 0:
            continue
        stated = idx < len(facts.post_money_yen) and facts.post_money_yen[idx] > 0
        if stated:
            post = float(facts.post_money_yen[idx])
            pre = max(post - raise_yen, float(raise_yen))
        else:
            pre = max(raise_yen * 4.0, last_post)
            post = pre + raise_yen
        last_post = post
        rounds.append({
            "idx": idx,
            "fy": facts.period_labels[idx] if idx < len(facts.period_labels) else f"period {idx}",
            "raise": float(raise_yen),
            "pre": float(pre),
            "stated": stated,
            "convertibles_cum": float(sum(facts.convertibles_yen[: idx + 1])),
        })
    return rounds


_V2_FOUNDER_SHARES = 10_000_000  # JP founding convention (editable blue input)


def _build_cap_table_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Cap Table"]
    A = ctx.rows["Assumptions"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Cap Table",
        "Rounds register (non-period axis): pre/post money, share classes, issued and "
        "fully-diluted ownership, voting thresholds, and pool-range checks.",
        period_axis=False,
    )
    R: dict = {}
    ctx.rows["Cap Table"] = R
    rounds = _v2_cap_rounds(facts)
    n_cols = 1 + len(rounds)
    first_col = V2_FIRST_PERIOD_COL  # F = founding; G.. = rounds (E keeps the unit column)
    cols = list(range(first_col, first_col + n_cols))
    note_col = first_col + n_cols  # directly after the grid — no skipped default-width column
    _set_column_widths(ws, {3: ib.COL_LABEL_WIDTH_V2, note_col: ib.COL_NOTE_WIDTH_V2,
                            **{c: 14 for c in cols}})
    ws._startup_note_col = note_col
    _v2_register_caption(ws, _v2_caption(ctx, None))
    pool_frac = min(max(float(facts.option_pool or 0.0), 0.0), 0.35)
    pool_shares = (
        int(_V2_FOUNDER_SHARES * pool_frac / (1.0 - pool_frac)) if pool_frac > 0 else 0
    )
    # Digit-crush guard input: EVERY money magnitude on the register (pre,
    # raise, post, convertible principal) so a small round cannot render "0".
    money_rep = [
        v
        for r_ in rounds
        for v in (r_["pre"], r_["raise"], r_["pre"] + r_["raise"], r_["convertibles_cum"])
    ] or [0.0]
    money_fmt, money_unit = _v2_row_money(ctx, money_rep)
    price_rep = [(r_["pre"] / _V2_FOUNDER_SHARES) for r_ in rounds] or [1.0]
    price_fmt, price_unit = _v2_row_money(ctx, price_rep)
    last_letter = ctx.last_period_letter
    r = ib.DATA_START_ROW_V2

    def header_row(row: int) -> None:
        ib.apply_semantic_fill_span(ws, row, 2, note_col, ib.BG_TABLE_HEADER,
                                    bottom=ib.THIN_LINE, border_start_col=3)
        label = ws.cell(row, 3, "Round")
        label.font = ib.FONT_BODY_BOLD
        label.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        names = ["Founding"] + [f"Round {i + 1} ({rd['fy']})" for i, rd in enumerate(rounds)]
        for col, name in zip(cols, names):
            ib.apply_year_header(ws.cell(row, col, name), name)

    def put_row(key, label, payload, fmt, *, kind="formula", unit="", note="",
                bold=False, band=False, snapshot=False):
        nonlocal r
        _v2_label_cells(ws, r, label, "", unit, note, note_col, bold=bold)
        for col, value in zip(cols, payload):
            if value is None:
                continue
            cell = ws.cell(r, col, value)
            if snapshot:
                cell.number_format = fmt
                cell.font = _FONT_SNAPSHOT_V2
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            elif kind == "input" and not (isinstance(value, str) and value.startswith("=")):
                ib.apply_hard_input(cell, fmt)
            else:
                _apply_value_style(cell, fmt)
        if band:
            _highlight_row(ws, r, cols[-1])
        R[key] = r
        r += 1

    _v2_section(ws, ctx, r, "Rounds register", end_col=note_col)
    r += 1
    header_row(r)
    R["header"] = r
    r += 1
    put_row("timing", "Timing", ["incorporation"] + [rd["fy"] for rd in rounds],
            "General", kind="input", note="round timing = first month of the FY (期首調達仮定)")
    for col, value in zip(cols, ["incorporation"] + [rd["fy"] for rd in rounds]):
        ib.apply_comment(ws.cell(R["timing"], col), wrap_text=False)
        ws.cell(R["timing"], col).alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=False)
    pre_values: list = ["-"] + [rd["pre"] for rd in rounds]
    any_estimated = any(not rd["stated"] for rd in rounds)
    put_row("pre", "Pre-money", pre_values, money_fmt, kind="input", unit=money_unit,
            note=("post-money stated where available; otherwise pre-money is a 20%-dilution "
                  "estimate — replace with the negotiated figure" if any_estimated
                  else "pre-money = stated post-money − raise"))
    raise_payload: list = ["-"]
    for rd in rounds:
        eq_row = A["equity_raise"]
        raise_payload.append(
            f"=SUMIF('Assumptions'!$F$4:${last_letter}$4,\"{rd['fy']}\","
            f"'Assumptions'!$F${eq_row}:${last_letter}${eq_row})")
    put_row("raise", "Equity raised", raise_payload, money_fmt, unit=money_unit,
            note="live FY total from the Assumptions financing rows")
    post_payload: list = ["-"]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        post_payload.append(f"={col}{R['pre']}+{col}{R['raise']}")
    put_row("post", "Post-money", post_payload, money_fmt, unit=money_unit, bold=True)
    price_payload: list = ["-"]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        prev = get_column_letter(cols[i])
        price_payload.append(f"={col}{R['pre']}/{prev}%(fd)s")
    put_row("price", "Share price", price_payload, price_fmt, unit=price_unit,
            note="pre-money ÷ fully diluted shares before the round")
    new_payload: list = ["-"]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        new_payload.append(f"=ROUND({col}{R['raise']}/{col}{R['price']},0)")
    put_row("new_shares", "New preferred shares", new_payload, FMT_COUNT_V2, unit="shares")
    r += 1

    _v2_section(ws, ctx, r, "Share classes (cumulative)", end_col=note_col)
    r += 1
    common_payload: list = [_V2_FOUNDER_SHARES]
    pool_payload: list = [pool_shares]
    for i in range(len(rounds)):
        prev = get_column_letter(cols[i])
        common_payload.append(f"={prev}%(common)s")
        pool_payload.append(f"={prev}%(pool)s")
    put_row("common", "Common shares (founders)", common_payload, FMT_COUNT_V2,
            kind="input", unit="shares",
            note="設立時 普通株式 10,000,000株は慣行値(blue input — 実数に置換可)")
    put_row("pool", "Option pool shares (reserved)", pool_payload, FMT_COUNT_V2,
            kind="input", unit="shares",
            note="pool sized from the option-pool assumption at founding; refresh per round if granted")
    pref_payload: list = [0]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        prev = get_column_letter(cols[i])
        pref_payload.append(f"={prev}%(pref)s+{col}{R['new_shares']}")
    put_row("pref", "Preferred shares (cumulative)", pref_payload, FMT_COUNT_V2, unit="shares")
    conv_payload: list = [0] + [rd["convertibles_cum"] for rd in rounds]
    put_row("conv_principal", "Convertible principal outstanding", conv_payload, money_fmt,
            kind="input", unit=money_unit,
            note="J-KISS principal — cap / discount not stated, so potential shares use the round price")
    jkiss_payload: list = [0]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        jkiss_payload.append(
            f"=IF({col}{R['conv_principal']}=0,0,"
            f"ROUND({col}{R['conv_principal']}/{col}{R['price']},0))")
    put_row("jkiss", "J-KISS potential shares", jkiss_payload, FMT_COUNT_V2, unit="shares")
    issued_payload = []
    fd_payload = []
    for col_idx in cols:
        col = get_column_letter(col_idx)
        issued_payload.append(f"={col}%(common)s+{col}%(pref)s")
        fd_payload.append(f"={col}%(common)s+{col}%(pool)s+{col}%(pref)s+{col}%(jkiss)s")
    put_row("issued", "Issued shares (common + preferred)", issued_payload, FMT_COUNT_V2,
            unit="shares", bold=True)
    put_row("fd", "Fully diluted shares", fd_payload, FMT_COUNT_V2, unit="shares",
            bold=True, band=True,
            note="common + reserved pool + preferred + convertible potential")
    # resolve the %(...)s forward/backward references
    refs = {"common": R["common"], "pool": R["pool"], "pref": R["pref"],
            "jkiss": R["jkiss"], "fd": R["fd"]}
    for row in (R["price"], R["common"], R["pool"], R["pref"], R["issued"], R["fd"]):
        for col_idx in cols:
            cell = ws.cell(row, col_idx)
            if isinstance(cell.value, str) and "%(" in cell.value:
                cell.value = cell.value % refs
    r += 1

    _v2_section(ws, ctx, r, "Ownership", end_col=note_col)
    r += 1
    pct = _v2_pct_fmt(ctx)

    def ratio_row(key, label, num_expr, den_row, **kw):
        payload = []
        for col_idx in cols:
            col = get_column_letter(col_idx)
            payload.append(f"={num_expr.format(col=col)}/{col}{den_row}")
        put_row(key, label, payload, pct, unit="%", **kw)

    ratio_row("founder_issued", "Founder % (issued / voting)",
              "{col}" + str(R["common"]), R["issued"], bold=True)
    ratio_row("founder_fd", "Founder % (fully diluted)",
              "{col}" + str(R["common"]), R["fd"])
    ratio_row("pool_fd", "Option pool % (fully diluted)",
              "{col}" + str(R["pool"]), R["fd"])
    ratio_row("inv_issued", "Investors % (issued)",
              "{col}" + str(R["pref"]), R["issued"])
    ratio_row("inv_fd", "Investors % (fully diluted)",
              "({col}" + str(R["pref"]) + "+{col}" + str(R["jkiss"]) + ")", R["fd"])
    own_payload: list = ["-"]
    for i in range(len(rounds)):
        col = get_column_letter(cols[1 + i])
        own_payload.append(f"=IF({col}{R['post']}=0,\"-\",{col}{R['raise']}/{col}{R['post']})")
    put_row("new_investor", "New investor ownership", own_payload, pct, unit="%",
            note="this round's raise ÷ post-money")
    for key, label, expr in (
        ("vote_23", "Voting: founder ≥ 2/3 (特別決議)", "2/3"),
        ("vote_12", "Voting: founder > 1/2 (普通決議)", "1/2"),
        ("vote_13", "Voting: founder ≥ 1/3 (拒否権)", "1/3"),
    ):
        payload = []
        for col_idx in cols:
            col = get_column_letter(col_idx)
            op = ">" if key == "vote_12" else ">="
            payload.append(
                f'=IF({col}{R["founder_issued"]}{op}{expr},"holds","below")')
        put_row(key, label, payload, "General", unit="status")
        for col_idx in cols:
            ib.apply_formula(ws.cell(R[key], col_idx), "General")
            ws.cell(R[key], col_idx).alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=False)
    pool_status_payload = []
    for col_idx in cols:
        col = get_column_letter(col_idx)
        pool_status_payload.append(
            f'=IF(OR({col}{R["pool_fd"]}<0.1,{col}{R["pool_fd"]}>0.15),"review","ok")')
    put_row("pool_check", "Option pool range (10–15% FD)", pool_status_payload, "General",
            unit="status", note="outside the 10–15% fully-diluted band → review before the round")
    check_payload = []
    for col_idx in cols:
        col = get_column_letter(col_idx)
        check_payload.append(
            f"=ROUND(({col}{R['founder_fd']}+{col}{R['pool_fd']}+{col}{R['inv_fd']}-1)*100,0)")
    put_row("own_check", "Ownership check", check_payload, FMT_CHECK_V2, unit="check")
    first_check = f"{get_column_letter(cols[0])}{R['own_check']}"
    last_check = f"{get_column_letter(cols[-1])}{R['own_check']}"
    ctx.checks.append(CheckEntry(
        "Cap Table", f"{first_check}:{last_check}",
        "Cap Table: fully-diluted ownership sums to 100% per round"))
    ctx.snapshots["founder_fd_final"] = None  # filled by python math below
    # python-side FD math for snapshot consumers (Summary / Valuation & Exit)
    common = float(_V2_FOUNDER_SHARES)
    pool = float(pool_shares)
    pref = 0.0
    jkiss = 0.0
    for rd in rounds:
        fd_before = common + pool + pref + jkiss
        price = rd["pre"] / fd_before if fd_before else 0.0
        new_shares = rd["raise"] / price if price > 0 else 0.0
        pref += new_shares
        jkiss = (rd["convertibles_cum"] / price) if (price > 0 and rd["convertibles_cum"]) else jkiss
    fd_total = common + pool + pref + jkiss
    ctx.snapshots["founder_fd_final"] = common / fd_total if fd_total else 1.0
    ctx.snapshots["investor_fd_final"] = (pref + jkiss) / fd_total if fd_total else 0.0


def _build_kernel_v2(wb: Workbook, ctx: BuildContext) -> None:
    """Compact Kernel sheet for the cap_table state-machine bundle."""
    facts = ctx.facts
    ws = wb["Kernel"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Model Kernel",
        "What is decided, on what mechanics, and what remains unknown.",
        period_axis=False,
    )
    _set_column_widths(ws, {3: 24, 4: 118})
    r = 6
    _v2_section(ws, ctx, r, "Kernel", end_col=4)
    r += 1
    for label, text in (
        ("Decision", "Capitalization: round terms, dilution, and exit distribution."),
        ("Model grain", f"{facts.grain} × {len(facts.period_labels)} periods ({facts.period_labels[0]}–{facts.period_labels[-1]})."),
        ("Mechanics", facts.mechanics),
        ("Economic unit", facts.primary_unit_name),
        ("Currency", facts.currency),
        ("Unknowns", "; ".join(facts.source_unknowns) if facts.source_unknowns else "none listed"),
    ):
        ws.cell(r, 3, label)
        ib.apply_label(ws.cell(r, 3), bold=True)
        ws.cell(r, 4, text)
        ib.apply_comment(ws.cell(r, 4), wrap_text=False)
        r += 1


def _build_evidence_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Evidence"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Evidence",
        "External evidence register (comparables, benchmarks) and the isolated market-sanity "
        "block. Nothing here feeds the revenue chain.",
        period_axis=False,
    )
    R: dict = {}
    ctx.rows["Evidence"] = R
    note_col = 14
    ws._startup_note_col = note_col
    _set_column_widths(ws, {3: 16, 4: 26, 5: 13, 6: 20, 7: 11, 8: 12, 9: 12, 10: 12,
                            11: 13, 12: 34, 13: 34, note_col: ib.COL_NOTE_WIDTH_V2})
    comps = [c for c in (getattr(facts, "live_comps", []) or []) if isinstance(c, dict)]
    usable = [c for c in comps if c.get("status") in {"current", "provided"}]
    r = ib.DATA_START_ROW_V2

    _v2_section(ws, ctx, r, "Comparable evidence", end_col=note_col)
    r += 1
    headers = ["Ticker", "Company", "Type", "Source type", "EV/Rev", "EV/EBITDA",
               "Gross margin", "EBITDA margin", "As of", "Status / freshness", "Applicability", "Source / error"]
    for col, header in enumerate(headers, start=3):
        _apply_text_header(ws.cell(r, col), header)
    r += 1
    if comps:
        for comp in comps:
            ok = comp.get("status") in {"current", "provided"}
            values = [
                comp.get("ticker"), comp.get("name"),
                comp.get("company_type") or "public",
                comp.get("source_type") or "benchmark",
                comp.get("revenue_multiple"), comp.get("ebitda_multiple"),
                comp.get("gross_margin"), comp.get("ebitda_margin"),
                comp.get("as_of_date"),
                comp.get("status") if ok else f"failed: {str(comp.get('error') or '')[:40]}",
                comp.get("applicability_limits"),
                comp.get("source_url") if ok else comp.get("error"),
            ]
            for col, value in enumerate(values, start=3):
                if value is None:
                    continue
                cell = ws.cell(r, col, value)
                if col in (7, 8):
                    _apply_value_style(cell, ib.FMT_MULTIPLE)
                elif col in (9, 10):
                    _apply_value_style(cell, _v2_pct_fmt(ctx))
                elif col == 14 and ok:
                    ib.apply_link_external(cell)
                else:
                    ib.apply_comment(cell, wrap_text=False)
            r += 1
    else:
        ws.cell(r, 3, "No external evidence provided — DD gap")
        ib.apply_label(ws.cell(r, 3), bold=True)
        ws.cell(r, note_col, "collect comparable / transaction evidence before circulating valuation output")
        ib.apply_comment(ws.cell(r, note_col), wrap_text=False)
        r += 1
    r += 1

    _v2_section(ws, ctx, r, "Benchmark register", end_col=note_col)
    r += 1
    rev_multiples = sorted(float(c["revenue_multiple"]) for c in usable
                           if isinstance(c.get("revenue_multiple"), (int, float)))
    ebitda_multiples = sorted(float(c["ebitda_multiple"]) for c in usable
                              if isinstance(c.get("ebitda_multiple"), (int, float)))
    as_of = max((str(c.get("as_of_date") or "") for c in usable), default="")
    if rev_multiples or ebitda_multiples:
        for label, values in (("EV / Revenue multiple (peer set)", rev_multiples),
                              ("EV / EBITDA multiple (peer set)", ebitda_multiples)):
            if not values:
                continue
            _v2_label_cells(ws, r, label, "median / low / high", "x",
                            f"median of {len(values)} usable comps; as of {as_of or 'n/a'}", note_col)
            for col, value in ((6, statistics.median(values)), (7, values[0]), (8, values[-1])):
                _apply_value_style(ws.cell(r, col, round(value, 2)), ib.FMT_MULTIPLE)
            r += 1
    else:
        ws.cell(r, 3, "No external evidence provided — DD gap")
        ib.apply_label(ws.cell(r, 3), bold=True)
        ws.cell(r, note_col, "no usable multiples — valuation multiples fall back to profile defaults; refresh before use")
        ib.apply_comment(ws.cell(r, note_col), wrap_text=False)
        r += 1
    r += 1

    _v2_section(ws, ctx, r, "KPI vs live public peers", end_col=note_col)
    r += 1
    for offset, header in enumerate(["Plan (snapshot)", "Peer median", "Peer low", "Peer high", "Position"]):
        _apply_text_header(ws.cell(r, 6 + offset), header)
    r += 1
    base_rev = ctx.snapshots.setdefault("base_revenue_annual", _v2_base_revenue_annual(facts))
    proj = ctx.snapshots.setdefault("base_projection", project_plan_free_cash_flow(facts))
    plan_gm = implied_gross_margin_series(facts)[-1] if _v2_any(base_rev) else 0.0
    plan_em = (proj[-1]["ebitda"] / base_rev[-1]) if base_rev and base_rev[-1] else 0.0
    for label, plan_value, key in (
        ("Gross margin — plan vs live peers", plan_gm, "gross_margin"),
        ("EBITDA margin — plan vs live peers", plan_em, "ebitda_margin"),
    ):
        values = sorted(float(c[key]) for c in usable if isinstance(c.get(key), (int, float)))
        _v2_label_cells(ws, r, label, "", "%",
                        "plan value is a generated snapshot (final year) — self-contained so the "
                        "block survives focused bundles", note_col)
        plan_cell = ws.cell(r, 6, round(plan_value, 4))
        plan_cell.number_format = _v2_pct_fmt(ctx)
        plan_cell.font = _FONT_SNAPSHOT_V2
        plan_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        if values:
            for col, value in ((7, statistics.median(values)), (8, values[0]), (9, values[-1])):
                _apply_value_style(ws.cell(r, col, round(value, 4)), _v2_pct_fmt(ctx))
            _apply_value_style(
                ws.cell(r, 10,
                        f'=IF(F{r}<H{r},"below",IF(F{r}>I{r},"above","within"))'),
                "General")
        else:
            ws.cell(r, 10, "no live peer data")
            ib.apply_comment(ws.cell(r, 10), wrap_text=False)
            ws.cell(r, 4, "refresh before circulation")
            ib.apply_comment(ws.cell(r, 4), wrap_text=False)
        r += 1
    r += 1

    _v2_section(ws, ctx, r, "Market sanity", end_col=note_col)
    r += 1
    som_row = None
    for key, label, value, note in (
        ("tam", "TAM", facts.tam_yen, "top-down category opportunity — source before circulation"),
        ("sam", "SAM", facts.sam_yen, "serviceable wedge given geography / channel / readiness"),
        ("som", "SOM", facts.som_yen, "plan-case obtainable market — isolated sanity input, never feeds the revenue chain"),
    ):
        m_fmt, m_unit = _v2_row_money(ctx, [value])
        _v2_label_cells(ws, r, label, "market sizing input", m_unit, note, note_col,
                        bold=key == "som")
        cell = ws.cell(r, 6, value)
        ib.apply_hard_input(cell, m_fmt)
        R[f"{key}_cell"] = f"$F${r}"
        if key == "som":
            som_row = r
        r += 1
    if "Summary" in ctx.bundle:
        _v2_label_cells(ws, r, "Final-year share of SOM", "", "",
                        "checked on the Summary Cross-check block (Summary references this SOM cell)",
                        note_col)
        r += 1
    else:
        fy5_rev = base_rev[-1] if base_rev else 0.0
        share = (fy5_rev / facts.som_yen) if facts.som_yen else 0.0
        _v2_label_cells(ws, r, "Final-year share of SOM (snapshot)", "", "%",
                        "generated snapshot (kernel final-year revenue ÷ SOM) — no revenue chain in this bundle",
                        note_col)
        cell = ws.cell(r, 6, round(share, 4))
        cell.number_format = _v2_pct_fmt(ctx)
        cell.font = _FONT_SNAPSHOT_V2
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        r += 1
    _v2_label_cells(ws, r, "Freshness", "", "",
                    f"comparables as of {as_of or 'n/a'}; market figures dated at source — refresh anything older than 12 months",
                    note_col)
    r += 1


def _build_valuation_exit_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Valuation & Exit"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Valuation & Exit",
        "Method matrix (own Low/Mid/High columns — not the period axis), selected EV range, "
        "investor return, and the exit waterfall.",
        period_axis=False,
    )
    R: dict = {}
    ctx.rows["Valuation & Exit"] = R
    note_col = 14
    ws._startup_note_col = note_col
    _set_column_widths(ws, {3: ib.COL_LABEL_WIDTH_V2, note_col: ib.COL_NOTE_WIDTH_V2,
                            **{c: 14 for c in range(4, note_col)}})
    _v2_register_caption(ws, _v2_caption(ctx, None))
    base_rev = ctx.snapshots.setdefault("base_revenue_annual", _v2_base_revenue_annual(facts))
    proj = ctx.snapshots.setdefault("base_projection", project_plan_free_cash_flow(facts))
    rev_fy = _v2_fy_flow_rollup(ctx, base_rev)
    fcf_fy = _v2_fy_flow_rollup(ctx, [p["free_cash_flow"] for p in proj])
    ebitda_fy = _v2_fy_flow_rollup(ctx, [p["ebitda"] for p in proj])
    gp_fy = _v2_fy_flow_rollup(
        ctx, [base_rev[i] * facts.target_gross_margin[i] for i in range(len(base_rev))])
    n_fy = len(rev_fy)
    fy_labels: list = []
    for fy in ctx.axis.fy_labels:
        if fy not in fy_labels:
            fy_labels.append(fy)
    debt_bal = float(sum(facts.debt_raise_yen) + sum(facts.convertibles_yen)
                     + sum(facts.lease_financing_yen) - sum(facts.debt_amortization_yen))
    cash_final = _v2_cases(ctx)["base"]["ending_cash"][-1]
    money_fmt, money_unit = _v2_row_money(ctx, [abs(v) for v in rev_fy] or [0.0])
    pct = _v2_pct_fmt(ctx)
    has_pl = "P&L" in ctx.bundle
    has_cf = "CF" in ctx.bundle
    has_assumptions = "Assumptions" in ctx.bundle
    last_letter = ctx.last_period_letter
    A = ctx.rows.get("Assumptions", {}) if has_assumptions else {}
    unit_sale = facts.revenue_mode == "unit_sale"
    pinned = bool(getattr(facts, "customers_pinned", False)) and _v2_any(facts.customers)
    cf_compact = ctx.rows.get("CF", {}).get("compact") if has_cf else None
    # Compact in-bundle engine (ma_exit / dcf_only): when no statement sheets
    # are in the bundle but the Assumptions register is, the operating basis
    # and DCF FCF rows are LIVE formulas off the driver rows × the effective
    # scenario scales — flipping the toggle must move the selected EV range.
    compact_engine = (
        (not has_pl)
        and (not has_cf)
        and has_assumptions
        and "eff_demand_cell" in A
        and (
            ("price" in A and ("customers" in A or "new_units" in A))
            or ("gmv" in A and "take" in A)
        )
    )
    if compact_engine:
        effd = f"'Assumptions'!{A['eff_demand_cell']}"
        effp = f"'Assumptions'!{A['eff_price_cell']}"
        effc = f"'Assumptions'!{A['eff_cost_cell']}"
        effo = f"'Assumptions'!{A['eff_opex_cell']}"

    def a_fy_sum(row_idx: int, col_letter: str, hdr_row: int) -> str:
        """FY total of an Assumptions flow row (criteria = this FY column header)."""
        return (
            f"SUMIF('Assumptions'!$F$4:${last_letter}$4,{col_letter}${hdr_row},"
            f"'Assumptions'!$F${row_idx}:${last_letter}${row_idx})"
        )

    def a_fy_cum(row_idx: int, col_letter: str, hdr_row: int) -> str:
        """Cumulative total of an Assumptions flow row through this FY."""
        return (
            f'SUMIF(\'Assumptions\'!$F$4:${last_letter}$4,"<="&{col_letter}${hdr_row},'
            f"'Assumptions'!$F${row_idx}:${last_letter}${row_idx})"
        )

    def a_fy_end(row_idx: int, col_letter: str, hdr_row: int) -> str:
        """FY-end value of an Assumptions stock/rate row (COUNTIF '<=' pull)."""
        return (
            f"INDEX('Assumptions'!$F${row_idx}:${last_letter}${row_idx},"
            f'COUNTIF(\'Assumptions\'!$F$4:${last_letter}$4,"<="&{col_letter}${hdr_row}))'
        )

    r = ib.DATA_START_ROW_V2

    def single(key, label, payload, fmt, *, unit="", driver="", note="", kind="formula",
               snapshot=False, bold=False):
        nonlocal r
        _v2_label_cells(ws, r, label, driver, unit, note, note_col, bold=bold)
        cell = ws.cell(r, 4, payload)
        if snapshot:
            cell.number_format = fmt
            cell.font = _FONT_SNAPSHOT_V2
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        elif kind == "input":
            ib.apply_hard_input(cell, fmt)
        else:
            _apply_value_style(cell, fmt)
        R[key] = r
        R[f"{key}_cell"] = f"$D${r}"
        r += 1

    _v2_section(ws, ctx, r, "Operating basis", end_col=note_col)
    r += 1
    if has_pl:
        PL = ctx.rows["P&L"]
        single("rev_basis", f"{fy_labels[-1]} revenue (basis)",
               f"='P&L'!{last_letter}{PL['rev']}", money_fmt, unit=money_unit,
               driver="P&L final FY")
        single("gp_basis", f"{fy_labels[-1]} gross profit (basis)",
               f"='P&L'!{last_letter}{PL['gp']}", money_fmt, unit=money_unit,
               driver="P&L final FY")
        single("ebitda_basis", f"{fy_labels[-1]} EBITDA (basis)",
               f"='P&L'!{last_letter}{PL['ebitda']}", money_fmt, unit=money_unit,
               driver="P&L final FY")
    elif compact_engine:
        # Placeholders patched to the compact-engine grid rows (written in
        # the DCF support section below) once their row numbers are known.
        single("rev_basis", f"{fy_labels[-1]} revenue (basis)", "=__GRID_REV__",
               money_fmt, unit=money_unit, driver="compact engine",
               note="live: Assumptions drivers × effective scenario scales (grid below)")
        single("gp_basis", f"{fy_labels[-1]} gross profit (basis)", "=__GRID_GP__",
               money_fmt, unit=money_unit, driver="compact engine")
        single("ebitda_basis", f"{fy_labels[-1]} EBITDA (basis)", "=__GRID_EBITDA__",
               money_fmt, unit=money_unit, driver="compact engine")
    elif cf_compact:
        final_fy = fy_labels[-1]

        def cf_fy_total(row_idx: int) -> str:
            return (
                f"=SUMIF('CF'!$F$4:${last_letter}$4,\"{final_fy}\","
                f"'CF'!$F${row_idx}:${last_letter}${row_idx})"
            )

        single("rev_basis", f"{final_fy} revenue (basis)", cf_fy_total(cf_compact["rev"]),
               money_fmt, unit=money_unit, driver="CF compact block")
        single("gp_basis", f"{final_fy} gross profit (basis)", cf_fy_total(cf_compact["gp"]),
               money_fmt, unit=money_unit, driver="CF compact block")
        single("ebitda_basis", f"{final_fy} EBITDA (basis)", cf_fy_total(cf_compact["ebitda"]),
               money_fmt, unit=money_unit, driver="CF compact block")
    else:
        single("rev_basis", f"{fy_labels[-1]} revenue (basis)", round(rev_fy[-1]),
               money_fmt, unit=money_unit, snapshot=True,
               note="kernel projection snapshot — P&L not in this bundle; rerun the generator after input changes")
        single("gp_basis", f"{fy_labels[-1]} gross profit (basis)", round(gp_fy[-1]),
               money_fmt, unit=money_unit, snapshot=True,
               note="kernel projection snapshot (target-margin basis)")
        single("ebitda_basis", f"{fy_labels[-1]} EBITDA (basis)", round(ebitda_fy[-1]),
               money_fmt, unit=money_unit, snapshot=True,
               note="kernel projection snapshot")
    single("net_debt", f"Net debt ({fy_labels[-1]})", round(debt_bal - cash_final),
           money_fmt, unit=money_unit, snapshot=True,
           note="debt-like draws − repayments − ending cash (negative = net cash)")
    single("discount", "Discount rate", float(facts.discount_rate), pct, unit="%",
           kind="input", driver="valuation policy")
    single("retention", "Exit value retention", 0.85, pct, unit="%", kind="input",
           note="haircut on the terminal multiple EV for illiquidity / execution risk")
    single("holding", "Holding period (years)", n_fy, "0", unit="years", kind="input")
    single("txn_pct", "Txn cost % of EV", 0.03, pct, unit="%", kind="input",
           note="advisor / legal / escrow leakage — editable input (3% default)")
    r += 1

    _v2_section(ws, ctx, r, "DCF support", end_col=note_col)
    r += 1
    fy_header_row = r
    ws.cell(r, 3, "Fiscal year")
    ib.apply_label(ws.cell(r, 3))
    for i, fy in enumerate(fy_labels):
        ib.apply_year_header(ws.cell(r, 4 + i, fy), fy)
    r += 1
    ws.cell(r, 3, "Year index")
    ib.apply_label(ws.cell(r, 3))
    for i in range(n_fy):
        cell = ws.cell(r, 4 + i, i + 1)
        ib.apply_hard_input(cell, "0")
    idx_row = r
    R["idx_row"] = r
    r += 1

    grid: dict = {}
    if compact_engine and not has_cf:
        hdr = fy_header_row

        def grid_write(key: str, label: str, per_col, fmt: str, note: str = "") -> None:
            nonlocal r
            _v2_label_cells(ws, r, label, "compact engine", "", note, note_col)
            for i in range(n_fy):
                col_letter = get_column_letter(4 + i)
                prev_letter = get_column_letter(3 + i)
                _apply_value_style(
                    ws.cell(r, 4 + i, f"={per_col(col_letter, prev_letter)}"), fmt)
            grid[key] = r
            r += 1

        if not unit_sale and ("customers" in A or "new_units" in A):
            if pinned and "customers" in A:
                grid_write("units_end", "Ending units / customers (FY end)",
                           lambda c, p: f"{a_fy_end(A['customers'], c, hdr)}*{effd}",
                           FMT_COUNT_V2, note="stated base × demand scale")
            else:
                grid_write("units_end", "Ending units / customers (FY end)",
                           lambda c, p: f"{a_fy_cum(A['new_units'], c, hdr)}*{effd}",
                           FMT_COUNT_V2, note="cumulative new units (compact — churn omitted)")
            grid_write("units_avg", "Average units (FY)",
                       lambda c, p: f"(N({p}{grid['units_end']})+{c}{grid['units_end']})/2",
                       FMT_COUNT_V2)
        rev_terms = []
        if unit_sale and "new_units" in A and "price" in A:
            rev_terms.append(
                lambda c, p: f"{a_fy_sum(A['new_units'], c, hdr)}*{effd}"
                             f"*{a_fy_end(A['price'], c, hdr)}*{effp}")
        elif "units_avg" in grid and "price" in A:
            rev_terms.append(
                lambda c, p: f"{c}{grid['units_avg']}*{a_fy_end(A['price'], c, hdr)}*{effp}*12")
            if "one_time_fee" in A and "new_units" in A:
                rev_terms.append(
                    lambda c, p: f"{a_fy_sum(A['new_units'], c, hdr)}*{effd}"
                                 f"*{a_fy_end(A['one_time_fee'], c, hdr)}")
        if "gmv" in A and "take" in A:
            rev_terms.append(
                lambda c, p: f"{a_fy_sum(A['gmv'], c, hdr)}*{effd}"
                             f"*{a_fy_end(A['take'], c, hdr)}*{effp}")

        def rev_formula(c, p):
            body = "+".join(term(c, p) for term in rev_terms) or "0"
            if "other_share" in A:
                return f"({body})*(1+{a_fy_end(A['other_share'], c, hdr)})"
            return body

        grid_write("rev", "Revenue (FY, compact)", rev_formula, money_fmt,
                   note="Assumptions drivers × effective scenario scales")
        if "tgm" in A:
            grid_write("gp", "Gross profit (FY, compact)",
                       lambda c, p: f"{c}{grid['rev']}-{c}{grid['rev']}"
                                    f"*(1-{a_fy_end(A['tgm'], c, hdr)})*{effc}",
                       money_fmt, note="target-margin basis; cost scale on the COGS stack")
        elif "vc_pct" in A:
            grid_write("gp", "Gross profit (FY, compact)",
                       lambda c, p: f"{c}{grid['rev']}-{c}{grid['rev']}"
                                    f"*{a_fy_end(A['vc_pct'], c, hdr)}*{effc}",
                       money_fmt)
        else:
            grid_write("gp", "Gross profit (FY, compact)",
                       lambda c, p: f"{c}{grid['rev']}", money_fmt,
                       note="no COGS driver stated — gross profit equals revenue")
        loaded_expr = None
        if "base_salary" in A and "welfare_rate" in A:
            loaded_expr = lambda c: (f"{a_fy_end(A['base_salary'], c, hdr)}"
                                     f"*(1+{a_fy_end(A['welfare_rate'], c, hdr)})")
        elif "loaded_comp" in A:
            loaded_expr = lambda c: a_fy_end(A["loaded_comp"], c, hdr)
        if "total_hc" in A and loaded_expr is not None:
            grid_write("people", "People cost (FY, compact)",
                       lambda c, p: f"{a_fy_end(A['total_hc'], c, hdr)}*{loaded_expr(c)}",
                       money_fmt, note="FY-end headcount × loaded comp")
        opex_terms = []
        if "sm_pct" in A:
            opex_terms.append(lambda c: f"{c}{grid['rev']}*{a_fy_end(A['sm_pct'], c, hdr)}")
        if "ga_pct" in A:
            opex_terms.append(lambda c: f"{c}{grid['rev']}*{a_fy_end(A['ga_pct'], c, hdr)}")
        if "fixed_ga" in A:
            opex_terms.append(lambda c: a_fy_end(A["fixed_ga"], c, hdr))
        rd_terms = []
        if "rd_floor" in A:
            rd_terms.append(lambda c: a_fy_end(A["rd_floor"], c, hdr))
        if "rd_per_fte" in A and "product_hc" in A:
            rd_terms.append(lambda c: f"{a_fy_end(A['product_hc'], c, hdr)}"
                                      f"*{a_fy_end(A['rd_per_fte'], c, hdr)}")
        if rd_terms:
            opex_terms.append(
                lambda c: (f"MAX({','.join(t(c) for t in rd_terms)})"
                           if len(rd_terms) > 1 else rd_terms[0](c)))
        if opex_terms:
            grid_write("programs", "Program OpEx (FY, compact)",
                       lambda c, p: f"({'+'.join(t(c) for t in opex_terms)})*{effo}",
                       money_fmt, note="S&M / R&D / G&A programs × opex scale")
        ebitda_refs = "".join(
            f"-{{c}}{grid[key]}" for key in ("people", "programs") if key in grid)

        def ebitda_formula(c, p):
            return f"{c}{grid['gp']}" + ebitda_refs.replace("{c}", c)

        grid_write("ebitda", "EBITDA (FY, compact)", ebitda_formula, money_fmt)
        capex_terms = []
        if "capex_unit" in A and "new_units" in A:
            capex_terms.append(
                lambda c: f"{a_fy_sum(A['new_units'], c, hdr)}*{effd}"
                          f"*{a_fy_end(A['capex_unit'], c, hdr)}")
        if "other_capex" in A:
            capex_terms.append(lambda c: a_fy_end(A["other_capex"], c, hdr))
        if capex_terms:
            grid_write("capex", "CapEx (FY, compact)",
                       lambda c, p: "+".join(t(c) for t in capex_terms), money_fmt)
        if "tax_rate" in A:
            grid_write("tax", "Tax (FY, compact proxy)",
                       lambda c, p: f"MAX(0,{c}{grid['ebitda']}*{a_fy_end(A['tax_rate'], c, hdr)})",
                       money_fmt, note="EBITDA-based proxy — D&A / NOL simplified")

    if compact_engine and not has_cf:
        fcf_note = "compact engine: EBITDA − CapEx − tax (D&A / working capital omitted)"
    elif has_cf:
        fcf_note = ""
    else:
        fcf_note = "kernel projection snapshot — CF not in this bundle"
    _v2_label_cells(ws, r, "Free cash flow (per FY)", "", money_unit, fcf_note, note_col)
    for i, fy in enumerate(fy_labels):
        cell_ref = ws.cell(r, 4 + i)
        if has_cf:
            CF = ctx.rows["CF"]
            cell_ref.value = (
                f"=SUMIF('CF'!$F$4:${last_letter}$4,\"{fy}\","
                f"'CF'!$F${CF['fcf']}:${last_letter}${CF['fcf']})")
            _apply_value_style(cell_ref, money_fmt)
        elif grid:
            col_letter = get_column_letter(4 + i)
            body = f"={col_letter}{grid['ebitda']}"
            if "capex" in grid:
                body += f"-{col_letter}{grid['capex']}"
            if "tax" in grid:
                body += f"-{col_letter}{grid['tax']}"
            cell_ref.value = body
            _apply_value_style(cell_ref, money_fmt)
        else:
            cell_ref.value = round(fcf_fy[i])
            cell_ref.number_format = money_fmt
            cell_ref.font = _FONT_SNAPSHOT_V2
            cell_ref.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    fcf_row = r
    R["fcf_row"] = r
    r += 1
    if grid:
        # Patch the operating-basis placeholders now that the grid rows exist.
        final_col = get_column_letter(4 + n_fy - 1)
        for basis_key, grid_key in (("rev_basis", "rev"), ("gp_basis", "gp"),
                                    ("ebitda_basis", "ebitda")):
            ws.cell(R[basis_key], 4).value = f"={final_col}{grid[grid_key]}"
    _v2_label_cells(ws, r, "PV of FCF (per FY)", "discounted", money_unit, "", note_col)
    for i in range(n_fy):
        col = get_column_letter(4 + i)
        _apply_value_style(
            ws.cell(r, 4 + i,
                    f"={col}{fcf_row}/(1+{R['discount_cell']})^{col}{idx_row}"),
            money_fmt)
    pv_row = r
    r += 1
    last_fy_col = get_column_letter(4 + n_fy - 1)
    single("pv_forecast", "PV of forecast FCF",
           f"=SUM(D{pv_row}:{last_fy_col}{pv_row})", money_fmt, unit=money_unit,
           note="NOT floored at zero — early burn genuinely reduces enterprise value")
    r += 1

    _v2_section(ws, ctx, r, "Method matrix", end_col=note_col)
    r += 1
    matrix_header_row = r
    ws.cell(r, 3, "Method")
    ib.apply_label(ws.cell(r, 3), bold=True)
    for col, header in ((4, "Low"), (5, "Mid"), (6, "High"), (7, "Credibility"), (8, "Use when")):
        _apply_text_header(ws.cell(r, col), header)
    r += 1

    def multiple_rows(key, label, mid, credibility, use_when, basis_key):
        nonlocal r
        low, high = round(mid * 0.75, 2), round(mid * 1.25, 2)
        _v2_label_cells(ws, r, f"{label} multiple", "", "x", "", note_col)
        for col, value in ((4, low), (5, round(mid, 2)), (6, high)):
            ib.apply_hard_input(ws.cell(r, col, value), ib.FMT_MULTIPLE)
        mult_row = r
        r += 1
        _v2_label_cells(ws, r, f"EV — {label} basis", "", money_unit, "", note_col)
        for col in (4, 5, 6):
            letter = get_column_letter(col)
            _apply_value_style(
                ws.cell(r, col, f"={R[f'{basis_key}_cell']}*{letter}{mult_row}"), money_fmt)
        ws.cell(r, 7, credibility)
        ib.apply_comment(ws.cell(r, 7), wrap_text=False)
        ws.cell(r, 8, use_when)
        ib.apply_comment(ws.cell(r, 8), wrap_text=False)
        R[f"ev_{key}"] = r
        r += 1

    multiple_rows("rev", "Revenue", facts.revenue_multiple[-1] if facts.revenue_multiple else 5.0,
                  "primary when growth quality leads", "recurring / high-quality growth", "rev_basis")
    multiple_rows("gp", "Gross profit", facts.gross_profit_multiple[-1] if facts.gross_profit_multiple else 8.0,
                  "support when cost-to-serve matters", "gross margin is well-defined", "gp_basis")
    multiple_rows("ebitda", "EBITDA", facts.ebitda_multiple[-1] if facts.ebitda_multiple else 15.0,
                  "primary when profitability is visible", "credible near-term profitability", "ebitda_basis")
    single("primary_ev", "Primary-method EV",
           f"=IF({R['ebitda_basis_cell']}>0,E{R['ev_ebitda']},"
           f"IF({R['gp_basis_cell']}>0,E{R['ev_gp']},E{R['ev_rev']}))",
           money_fmt, unit=money_unit, bold=True,
           note="EBITDA basis when positive, else GP, else revenue")
    single("pv_exit", "PV of exit value",
           f"={R['primary_ev_cell']}*{R['retention_cell']}"
           f"/(1+{R['discount_cell']})^{last_fy_col}{idx_row}",
           money_fmt, unit=money_unit,
           note="exit-multiple terminal (retention haircut), discounted over the holding period")
    single("dcf_ev", "DCF EV",
           f"={R['pv_forecast_cell']}+{R['pv_exit_cell']}",
           money_fmt, unit=money_unit)
    _v2_label_cells(ws, r, "DCF range factor", "", "x",
                    "blue inputs — spread the DCF into the Low/Mid/High columns", note_col)
    for col, value in ((4, 0.8), (5, 1.0), (6, 1.2)):
        ib.apply_hard_input(ws.cell(r, col, value), ib.FMT_MULTIPLE)
    dcf_factor_row = r
    r += 1
    _v2_label_cells(ws, r, "EV — DCF", "", money_unit, "", note_col)
    for col in (4, 5, 6):
        letter = get_column_letter(col)
        _apply_value_style(
            ws.cell(r, col, f"={R['dcf_ev_cell']}*{letter}{dcf_factor_row}"), money_fmt)
    ws.cell(r, 7, "cross-check; terminal-value heavy")
    ib.apply_comment(ws.cell(r, 7), wrap_text=False)
    ws.cell(r, 8, "cash-flow path is explainable")
    ib.apply_comment(ws.cell(r, 8), wrap_text=False)
    R["ev_dcf"] = r
    r += 1
    ev_rows = [R["ev_rev"], R["ev_gp"], R["ev_ebitda"], R["ev_dcf"]]
    _v2_label_cells(ws, r, "Selected EV range", "", money_unit,
                    "Mid = credibility-chosen primary method (EBITDA>GP>revenue); "
                    "Low/High = method-range floor/ceiling — not a blind average", note_col, bold=True)
    # Mid anchors on the credibility-aware Primary-method EV (the rubric bars
    # averaging methods blindly); Low/High still show the full method spread.
    mid_formula = f"={R['primary_ev_cell']}"
    for col, formula in (
        (4, "=MIN(" + ",".join(f"D{row}" for row in ev_rows) + ")"),
        (5, mid_formula),
        (6, "=MAX(" + ",".join(f"F{row}" for row in ev_rows) + ")"),
    ):
        _apply_value_style(ws.cell(r, col, formula), money_fmt)
        font = copy(ws.cell(r, col).font)
        font.bold = True
        ws.cell(r, col).font = font
    _highlight_row(ws, r, 8)
    R["selected"] = r
    R["sel_low_cell"] = f"$D${r}"
    R["sel_mid_cell"] = f"$E${r}"
    R["sel_high_cell"] = f"$F${r}"
    r += 2

    _v2_section(ws, ctx, r, "Investor return", end_col=note_col)
    r += 1
    invested_total = float(sum(facts.equity_raise_yen))
    if has_assumptions and "equity_raise" in ctx.rows.get("Assumptions", {}):
        A = ctx.rows["Assumptions"]
        single("invested", "Equity invested (cumulative)",
               f"=SUM('Assumptions'!$F${A['equity_raise']}:${last_letter}${A['equity_raise']})",
               money_fmt, unit=money_unit, driver="Assumptions")
    else:
        single("invested", "Equity invested (cumulative)", round(invested_total),
               money_fmt, unit=money_unit, snapshot=True,
               note="kernel funding plan snapshot — Assumptions not in this bundle")
    investor_fd = ctx.snapshots.get("investor_fd_final")
    if investor_fd is None:
        post = [v for v in facts.post_money_yen if v > 0]
        investor_fd = min(0.9, invested_total / post[-1]) if post else (0.35 if invested_total else 0.0)
    single("ownership", "Investor ownership at exit (FD)", round(float(investor_fd), 4),
           pct, unit="%", kind="input",
           note="from the cap-table math (base case) — editable assumption")
    founder_fd = ctx.snapshots.get("founder_fd_final")
    if founder_fd is None:
        founder_fd = max(0.0, 1.0 - float(investor_fd) - 0.12)
    single("founder_share", "Founder share of common at exit (FD)", round(float(founder_fd), 4),
           pct, unit="%", kind="input",
           note="from the cap-table math (base case) — editable assumption")
    single("moic", "MOIC at selected EV (Mid)",
           f"=IF({R['invested_cell']}=0,\"-\","
           f"{R['sel_mid_cell']}*{R['ownership_cell']}/{R['invested_cell']})",
           ib.FMT_MULTIPLE, unit="x", bold=True,
           note='"-" when no equity is invested — no return math on a zero base')
    single("irr", "Illustrative IRR",
           f"=IF({R['moic_cell']}=\"-\",\"-\","
           f"{R['moic_cell']}^(1/{R['holding_cell']})-1)",
           pct, unit="%",
           note="single-period approximation over the holding period")
    r += 1

    _v2_section(ws, ctx, r, "Exit waterfall", end_col=note_col)
    r += 1
    headers = ["Case", "Exit EV", "Net debt", "Txn costs", "Equity value",
               "Preference floor", "Common pool", "Investor proceeds",
               "Founder proceeds", "MOIC", "Walk-away signal"]
    for col, header in enumerate(headers, start=3):
        _apply_text_header(ws.cell(r, col), header)
    r += 1
    for case, ev_cell in (("Downside", R["sel_low_cell"]), ("Base", R["sel_mid_cell"]),
                          ("Upside", R["sel_high_cell"])):
        ws.cell(r, 3, case)
        ib.apply_label(ws.cell(r, 3), bold=case == "Base")
        ws.cell(r, 4, f"={ev_cell}")
        ws.cell(r, 5, f"={R['net_debt_cell']}")
        ws.cell(r, 6, f"=MAX(0,D{r}*{R['txn_pct_cell']})")
        ws.cell(r, 7, f"=MAX(0,D{r}-E{r}-F{r})")
        ws.cell(r, 8, f"=MIN(G{r},{R['invested_cell']})")
        ws.cell(r, 9, f"=MAX(0,G{r}-H{r})")
        ws.cell(r, 10, f"=MAX(H{r},G{r}*{R['ownership_cell']})")
        ws.cell(r, 11, f"=MAX(0,G{r}-J{r})*{R['founder_share_cell']}")
        ws.cell(r, 12,
                f"=IF({R['invested_cell']}=0,\"-\",J{r}/{R['invested_cell']})")
        ws.cell(r, 13,
                f"=IF(OR(L{r}=\"-\",L{r}<2),\"reprice / walk away\","
                f"IF(K{r}<J{r},\"terms protect investor\",\"committee case\"))")
        for col in range(4, 14):
            fmt = ib.FMT_MULTIPLE if col == 12 else ("General" if col == 13 else money_fmt)
            _apply_value_style(ws.cell(r, col), fmt)
        if case == "Base":
            R["signal_base_cell"] = f"$M${r}"
            R["investor_proceeds_base_cell"] = f"$J${r}"
            _highlight_row(ws, r, 13)
        r += 1
    _v2_label_cells(ws, r, "Waterfall convention", "", "",
                    "1x non-participating preference: investor takes MAX(preference, pro-rata); "
                    "founder proceeds flow from the remaining common pool", note_col)
    r += 1


def _build_ic_memo_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["IC Memo"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — IC Decision Memo",
        "Investment-committee readout generated from the model: recommendation gate, thesis, "
        "KPI readout, diligence ranking, and walk-away conditions.",
        period_axis=False,
    )
    _set_column_widths(ws, {3: ib.COL_LABEL_WIDTH_V2, 4: 118})
    VE = ctx.rows.get("Valuation & Exit", {})
    S = ctx.rows.get("Summary", {})
    cases = _v2_cases(ctx)
    base_case = cases["base"]
    down_case = cases["down"]
    rev_fy = _v2_fy_flow_rollup(ctx, base_case["revenue"])
    ebitda_fy = _v2_fy_flow_rollup(ctx, base_case["ebitda"])
    divisor, suffix = (1_000_000.0, "M")
    runway_target = facts.target_min_runway_months[0] if facts.target_min_runway_months else 18
    down_gap = max(0.0, -min(down_case["ending_cash"]))
    if VE:
        recommendation = (
            f"=IF(OR('Valuation & Exit'!{VE['moic_cell']}=\"-\","
            f"'Valuation & Exit'!{VE['moic_cell']}<2),"
            "\"Do not circulate externally until the return hurdle, valuation support, and "
            "evidence gates are cleared\","
            "\"Proceed subject to DD gates\")"
        )
        valuation_readout = (
            # "#,##0,," comma-scales the display to millions inside TEXT —
            # no numeric scaling literal in the formula body (R18 whitelist).
            "=\"Selected EV mid: \"&TEXT('Valuation & Exit'!"
            f"{VE['sel_mid_cell']},\"#,##0,,\")&\"M; MOIC: \"&"
            f"IF('Valuation & Exit'!{VE['moic_cell']}=\"-\",\"-\","
            f"TEXT('Valuation & Exit'!{VE['moic_cell']},\"0.0x\"))&\"; exit signal: \"&"
            f"'Valuation & Exit'!{VE['signal_base_cell']}"
        )
    else:
        recommendation = (
            "Do not circulate externally until valuation, pricing, and evidence gates are "
            "cleared (no Valuation & Exit sheet in this bundle)."
        )
        valuation_readout = "No valuation surface in this bundle — attach the Valuation & Exit output."
    if S and "master_check" in S:
        check_readout = (
            f"=IF('Summary'!$F${S['master_check']}=0,"
            "\"All machine checks pass (master check OK)\","
            "\"MASTER CHECK FAILING — resolve before any external use\")"
        )
    else:
        check_readout = "Summary not in this bundle — run the full workbook for the master check."
    kpi_readout = (
        f"Revenue {rev_fy[0] / divisor:,.0f}{suffix} → {rev_fy[-1] / divisor:,.0f}{suffix}; "
        f"final-year EBITDA {ebitda_fy[-1] / divisor:,.0f}{suffix}; "
        f"base minimum ending cash {min(base_case['ending_cash']) / divisor:,.0f}{suffix} "
        "(generated snapshot — rerun after input changes)."
    )
    sections = [
        ("Recommendation", recommendation),
        ("Machine checks", check_readout),
        ("Investment thesis",
         f"{facts.company} is modeled through an economic kernel described as {facts.mechanics}; "
         "treat it as a driver composition, not a sector template."),
        ("KPI readout", kpi_readout),
        ("Funding & dilution",
         f"Downside additional funding need {down_gap / divisor:,.0f}{suffix}; "
         f"hold ≥ {runway_target} months post-raise runway. Dilution and voting thresholds "
         "live on the Cap Table (full workbook)."),
        ("Valuation & return", valuation_readout),
        ("What must be true",
         "Demand coverage ≥ 1x, price support ≥ 1x, cost-to-serve on target, hiring within "
         "capacity, and financing closed before the runway floor — each maps to a check or "
         "support block in the model."),
        ("Downside triggers",
         "Downside becomes decision-relevant when it creates a funding gap, a runway breach, "
         "covenant pressure, unacceptable dilution, or a preference-stack leak."),
        ("Ranked DD gates",
         "1) comparable evidence freshness; 2) customer ROI / WTP proof; 3) cohort churn and CAC; "
         "4) cost-to-serve decomposition; 5) financing terms and preference stack; 6) buyer-view "
         "exit support."),
        ("Walk-away conditions",
         "Reject or reprice if MOIC stays below the hurdle, valuation support stays weak, pricing "
         "validation fails, or downside financing requires unacceptable dilution."),
        ("Source boundary",
         "Evidence status is declared per assumption row; refresh benchmark and comparable "
         "sources before external circulation."),
    ]
    r = 6
    for label, body in sections:
        _v2_section(ws, ctx, r, label, end_col=4)
        r += 1
        cell = ws.cell(r, 4, body)
        if isinstance(body, str) and body.startswith("="):
            _apply_value_style(cell, "General")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        else:
            ib.apply_comment(cell, wrap_text=False)
        r += 2


def _build_pricing_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Pricing"]
    A = ctx.rows["Assumptions"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Pricing",
        "Price / ROI / cost-floor support and the validation plan that must pass before the "
        "price is treated as evidence.",
        unit_caption=_v2_caption(ctx, None),
    )
    R: dict = {}
    ctx.rows["Pricing"] = R
    effp = f"'Assumptions'!{A['eff_price_cell']}"
    price_fmt, price_unit = _v2_row_money(ctx, facts.monthly_price_yen or [0])
    r = ib.DATA_START_ROW_V2
    _v2_section(ws, ctx, r, "Price support")
    r += 1
    if "price" in A:
        _v2_series_row(ws, ctx, r, "Selected price", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['price']}*{effp}"),
                       driver="× price scale", fmt=price_fmt, unit=price_unit, bold=True)
        R["price"] = r
        r += 1
    else:
        _v2_series_row(ws, ctx, r, "Selected price", "money_row",
                       formulas=_v2_formulas(ctx, "=0"), fmt=price_fmt, unit=price_unit,
                       note="no pricing driver stated — DD gap", bold=True)
        R["price"] = r
        r += 1
    if "roi" in A:
        roi_fmt, roi_unit = _v2_row_money(ctx, [facts.customer_roi_yen])
        _v2_series_row(ws, ctx, r, "Customer annual value / ROI", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['roi']}"),
                       driver="value proof", fmt=roi_fmt, unit=roi_unit)
        R["roi"] = r
        r += 1
    if "value_capture" in A:
        _v2_series_row(ws, ctx, r, "Value capture share", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['value_capture']}"),
                       driver="pricing power")
        R["vc"] = r
        r += 1
    if "roi" in R and "vc" in R:
        vp_fmt, vp_unit = _v2_row_money(
            ctx, [facts.customer_roi_yen * max(facts.value_capture_share[-1], 0.01) / 12])
        _v2_series_row(ws, ctx, r, "Value-based monthly price", "money_row",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['roi']}*{{c}}{R['vc']}/12"),
                       driver="ROI share", fmt=vp_fmt, unit=vp_unit)
        R["value_price"] = r
        r += 1
    serve_terms = "+".join(
        f"'Assumptions'!{{c}}{A[key]}" for key in ("delivery", "cloud", "support") if key in A)
    if serve_terms:
        _v2_series_row(ws, ctx, r, "Cost-to-serve / unit / month", "money_row",
                       formulas=_v2_formulas(ctx, f"={serve_terms}"),
                       driver="serve stack", fmt=price_fmt, unit=price_unit)
        R["serve"] = r
        r += 1
        if "tgm" in A:
            _v2_series_row(ws, ctx, r, "Cost-plus floor price", "money_row",
                           formulas=_v2_formulas(
                               ctx,
                               f"={{c}}{R['serve']}/(1-MAX(1/100,'Assumptions'!{{c}}{A['tgm']}))"),
                           driver="floor", fmt=price_fmt, unit=price_unit)
            R["floor"] = r
            r += 1
        _v2_series_row(ws, ctx, r, "Unit gross margin at selected price", "pct",
                       formulas=_v2_formulas(
                           ctx,
                           f"=IF({{c}}{R['price']}=0,0,({{c}}{R['price']}-{{c}}{R['serve']})/{{c}}{R['price']})"))
        R["unit_gm"] = r
        r += 1
    anchor = None
    if "value_price" in R and "floor" in R:
        anchor = f"MAX({{c}}{R['value_price']},{{c}}{R['floor']})"
    elif "value_price" in R:
        anchor = f"{{c}}{R['value_price']}"
    elif "floor" in R:
        anchor = f"{{c}}{R['floor']}"
    if anchor:
        _v2_series_row(ws, ctx, r, "Price support ratio", "x",
                       formulas=_v2_formulas(
                           ctx, f"=IF({{c}}{R['price']}=0,0,{anchor}/{{c}}{R['price']})"),
                       bold=True,
                       note="supported price anchor ÷ selected price — ≥ 1x means headroom")
        R["support"] = r
        r += 1
        _v2_series_row(ws, ctx, r, "Validation hurdle (support ratio ≥)", "x",
                       values=[1.5] * ctx.n_cols, driver="policy",
                       note="minimum evidence-backed support before pricing is treated as proven")
        R["hurdle"] = r
        r += 1
        _v2_series_row(ws, ctx, r, "Pricing gate", "status",
                       formulas=_v2_formulas(
                           ctx,
                           f'=IF({{c}}{R["support"]}>={{c}}{R["hurdle"]},"pass","DD gate")'))
        R["gate"] = r
        r += 1
    else:
        note_col = ws._startup_note_col
        _v2_label_cells(ws, r, "Price support ratio", "", "",
                        "no ROI or cost-floor evidence stated — pricing rests on the anchor alone (DD gap)",
                        note_col)
        r += 1
    r += 1
    _v2_section(ws, ctx, r, "Pricing validation plan")
    r += 1
    # Row-wise plan (test label / threshold tag / consolidated note): the
    # earlier 5-column table widened period columns F-G and broke the
    # uniform period-column width contract (期間列幅統一 audit).
    note_col = ws._startup_note_col
    for test, threshold, note in (
        ("WTP interview / LOI", "≥ 1.5x",
         "evidence: named buyer pain, budget owner, procurement trigger; "
         "owner: customer discovery / pipeline; IC: price can anchor value share"),
        ("Pilot conversion", "≤ 12 mo",
         "evidence: paid pilot or signed deployment scope; pass: gross-profit "
         "payback ≤ 12 months; owner: sales pipeline / contract; IC: "
         "implementation burden is financeable"),
        ("Packaging ladder", "no cliff",
         "evidence: entry, core, expansion, enterprise SKUs; pass: no cliff in "
         "the expansion path; owner: pricing owner; IC: avoid under-monetizing "
         "high-ROI accounts"),
        ("Renewal risk", "< downside",
         "evidence: cohort churn or renewal intent by segment; pass: churn "
         "below the downside case; owner: CS / cohort export; IC: discount the "
         "valuation if renewal evidence is weak"),
    ):
        _v2_label_cells(ws, r, test, threshold, "", note, note_col)
        r += 1


def _build_unit_economics_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Unit Economics"]
    A = ctx.rows["Assumptions"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Unit Economics",
        "Per-unit P&L, CAC / LTV, payback, and the cohort assumptions they rest on.",
        unit_caption=_v2_caption(ctx, None),
    )
    R: dict = {}
    ctx.rows["Unit Economics"] = R
    r = ib.DATA_START_ROW_V2
    r, C = _v2_compact_ops(ws, ctx, r)
    R["compact"] = C
    effp = f"'Assumptions'!{A['eff_price_cell']}"
    effc = f"'Assumptions'!{A['eff_cost_cell']}"
    effo = f"'Assumptions'!{A['eff_opex_cell']}"
    price_fmt, price_unit = _v2_row_money(ctx, facts.monthly_price_yen or [0])
    _v2_section(ws, ctx, r, "Unit economics")
    r += 1
    if "price" in A:
        _v2_series_row(ws, ctx, r, "Monthly price / unit", "money_row",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['price']}*{effp}"),
                       driver="× price scale", fmt=price_fmt, unit=price_unit)
        R["price"] = r
        r += 1
    serve_terms = "+".join(
        f"'Assumptions'!{{c}}{A[key]}" for key in ("delivery", "cloud", "support") if key in A)
    if serve_terms:
        _v2_series_row(ws, ctx, r, "Cost-to-serve / unit / month", "money_row",
                       formulas=_v2_formulas(ctx, f"=({serve_terms})*{effc}"),
                       driver="serve stack", fmt=price_fmt, unit=price_unit)
        R["serve"] = r
        r += 1
    if "price" in R and "serve" in R:
        _v2_series_row(ws, ctx, r, "Unit gross profit / month", "money_row",
                       formulas=_v2_formulas(ctx, f"={{c}}{R['price']}-{{c}}{R['serve']}"),
                       bold=True, fmt=price_fmt, unit=price_unit)
        R["unit_gp"] = r
        r += 1
        _v2_series_row(ws, ctx, r, "Unit gross margin", "pct",
                       formulas=_v2_formulas(
                           ctx,
                           f"=IF({{c}}{R['price']}=0,0,{{c}}{R['unit_gp']}/{{c}}{R['price']})"))
        R["unit_gm"] = r
        r += 1
    if "sm_pct" in A and "new" in C:
        _v2_series_row(ws, ctx, r, "Customer acquisition cost", "money_row",
                       formulas=_v2_formulas(
                           ctx,
                           f"=IF({{c}}{C['new']}=0,\"N/A\","
                           f"{{c}}{C['rev']}*'Assumptions'!{{c}}{A['sm_pct']}*{effo}/{{c}}{C['new']})"),
                       driver="S&M / new units", fmt=price_fmt, unit=price_unit)
        R["cac"] = r
        r += 1
    if "unit_gp" in R:
        if "churn" in A:
            _v2_series_row(ws, ctx, r, "LTV (gross-profit basis)", "money_row",
                           formulas=_v2_formulas(
                               ctx,
                               f"=IF('Assumptions'!{{c}}{A['churn']}=0,0,"
                               f"{{c}}{R['unit_gp']}*12/'Assumptions'!{{c}}{A['churn']})"),
                           driver="GP / churn", fmt=price_fmt, unit=price_unit)
            R["ltv"] = r
            r += 1
        else:
            _v2_series_row(ws, ctx, r, "Customer lifetime (months)", "months",
                           values=[36.0] * ctx.n_cols, driver="cohort assumption",
                           note="no churn driver stated — 36-month default; validate with cohorts")
            R["lifetime"] = r
            r += 1
            _v2_series_row(ws, ctx, r, "LTV (gross-profit basis)", "money_row",
                           formulas=_v2_formulas(
                               ctx, f"={{c}}{R['unit_gp']}*{{c}}{R['lifetime']}"),
                           driver="GP × lifetime", fmt=price_fmt, unit=price_unit)
            R["ltv"] = r
            r += 1
    if "cac" in R and "ltv" in R:
        _v2_series_row(ws, ctx, r, "LTV / CAC", "x",
                       formulas=_v2_formulas(
                           ctx,
                           f'=IF({{c}}{R["cac"]}="N/A","N/A",'
                           f"IF({{c}}{R['cac']}=0,0,{{c}}{R['ltv']}/{{c}}{R['cac']}))"),
                       bold=True, note="benchmark: ≥ 3x for venture-scale SaaS")
        R["ltv_cac"] = r
        r += 1
    if "cac" in R and "unit_gp" in R:
        _v2_series_row(ws, ctx, r, "CAC payback", "months",
                       formulas=_v2_formulas(
                           ctx,
                           f'=IF(OR({{c}}{R["cac"]}="N/A",{{c}}{R["unit_gp"]}<=0),"N/A",'
                           f"{{c}}{R['cac']}/{{c}}{R['unit_gp']})"),
                       note="benchmark: ≤ 12–18 months for efficient GTM")
        R["payback"] = r
        r += 1
    r += 1
    _v2_section(ws, ctx, r, "Cohort assumptions & benchmarks")
    r += 1
    if "churn" in A:
        _v2_series_row(ws, ctx, r, "Churn rate (annual)", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['churn']}"),
                       driver="cohort")
        r += 1
    if "net_retention" in A:
        _v2_series_row(ws, ctx, r, "Net retention", "pct",
                       formulas=_v2_formulas(ctx, f"='Assumptions'!{{c}}{A['net_retention']}"),
                       driver="cohort")
        r += 1
    note_col = ws._startup_note_col
    _v2_label_cells(ws, r, "Benchmark context", "", "",
                    "LTV/CAC ≥ 3x and CAC payback ≤ 12–18 months are venture-normal bands; "
                    "cohort exports beat blended assumptions — attach them before circulation",
                    note_col)
    r += 1


def _build_segments_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Segments"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Segments",
        "Segment economics register — only rendered when real segment inputs exist.",
        period_axis=False,
    )
    note_col = 10
    ws._startup_note_col = note_col
    _set_column_widths(ws, {3: ib.COL_LABEL_WIDTH_V2, note_col: ib.COL_NOTE_WIDTH_V2,
                            **{c: 15 for c in range(4, note_col)}})
    base_rev = ctx.snapshots.setdefault("base_revenue_annual", _v2_base_revenue_annual(facts))
    rev_fy = _v2_fy_flow_rollup(ctx, base_rev)
    money_fmt, money_unit = _v2_row_money(ctx, [abs(rev_fy[-1])])
    segments = [s for s in facts.segments if str(s).strip()][:6]
    r = ib.DATA_START_ROW_V2
    _v2_section(ws, ctx, r, "Segment register", end_col=note_col)
    r += 1
    headers = ["Segment", "Revenue share", "Segment revenue (final FY)", "Gross margin",
               "EV multiple", "Segment EV"]
    for col, header in enumerate(headers, start=3):
        _apply_text_header(ws.cell(r, col), header)
    r += 1
    _v2_label_cells(ws, r, "Final-year revenue (basis)", "kernel snapshot", money_unit,
                    "generated snapshot — rerun the generator after input changes", note_col)
    basis_cell = ws.cell(r, 4, round(rev_fy[-1]))
    basis_cell.number_format = money_fmt
    basis_cell.font = _FONT_SNAPSHOT_V2
    basis_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    basis_row = r
    r += 1
    first_seg_row = r
    share = round(1.0 / max(len(segments), 1), 4)
    for segment in segments:
        ws.cell(r, 3, str(segment))
        ib.apply_label(ws.cell(r, 3))
        ib.apply_hard_input(ws.cell(r, 4, share), _v2_pct_fmt(ctx))
        _apply_value_style(ws.cell(r, 5, f"=$D${basis_row}*D{r}"), money_fmt)
        ib.apply_hard_input(
            ws.cell(r, 6, facts.target_gross_margin[-1] if facts.target_gross_margin else 0.6),
            _v2_pct_fmt(ctx))
        ib.apply_hard_input(
            ws.cell(r, 7, facts.revenue_multiple[-1] if facts.revenue_multiple else 5.0),
            ib.FMT_MULTIPLE)
        _apply_value_style(ws.cell(r, 8, f"=E{r}*G{r}"), money_fmt)
        r += 1
    last_seg_row = r - 1
    r += 1
    _v2_label_cells(ws, r, "Consolidation bridge check", "Σ share − 100%", "check",
                    "0 = segment shares reconcile to the consolidated revenue", note_col)
    check_cell = ws.cell(
        r, 4, f"=ROUND((SUM(D{first_seg_row}:D{last_seg_row})-1)*100,0)")
    ib.apply_formula(check_cell, FMT_CHECK_V2)
    ctx.checks.append(CheckEntry("Segments", f"D{r}", "Segments: revenue shares sum to 100%"))
    r += 1
    _v2_label_cells(ws, r, "SOTP caution", "", "",
                    "use segment EV as a support method only — segment allocation must rest on "
                    "distinct evidence, not an even split", note_col)
    # Segment names are user-provided free text with data in column D (no blank
    # overflow to the right), so grow the label column to the longest name —
    # floor at the v2 label width, cap so one long name cannot bloat the sheet.
    longest = max((ib._display_width(str(ws.cell(row=rr, column=3).value or ""))
                   for rr in range(ib.DATA_START_ROW_V2, r + 1)), default=0)
    ws.column_dimensions["C"].width = max(
        ib.COL_LABEL_WIDTH_V2, min(64.0, longest + 2.0))


_V2_SHEET_PURPOSES = {
    "Guide": "Reading guide: decision, mechanics, formatting key, model qualifications.",
    "Summary": "Annual roll-up, KPIs, scenario comparison, consolidated checks (master check).",
    "Assumptions": "Editable driver register + scenario toggle; evidence status per row.",
    "Revenue Build": "Bottom-up revenue engine with demand and price support blocks (revenue bridge check).",
    "Cost Build": "COGS, gross margin, OpEx programs, CapEx (gross margin vs target check).",
    "People Plan": "Headcount, loaded compensation, people cost, capacity checks.",
    "P&L": "Reference-only tax-exclusive statement (P&L tie check).",
    "BS": "Simplified balance sheet — site-based AR, tax balances (balance check).",
    "CF": "Tax-inclusive 資金繰り: balances, financing events, runway (cash shortfall check).",
    "Financing": "Sources & Uses, instrument terms, post-raise runway, downside funding gap.",
    "Cap Table": "Rounds register: pre/post, share classes, ownership, voting thresholds.",
    "Evidence": "Comparable / benchmark register and isolated market sanity (TAM/SAM/SOM).",
    "Valuation & Exit": "Method matrix, selected EV range, investor return, exit waterfall.",
    "IC Memo": "Investment-committee readout: recommendation gate, DD ranking, walk-away.",
    "Pricing": "Price / ROI / cost-floor support and the pricing validation plan.",
    "Unit Economics": "Per-unit P&L, CAC / LTV, payback, cohort assumptions.",
    "Segments": "Segment economics register with a consolidation bridge check.",
    "Kernel": "Model kernel: decision, grain, mechanics, unknowns (cap-table bundle).",
    "Ownership": "Cap-table state machine: SAFE/J-KISS conversion, priced round, exit waterfall.",
}


def _build_guide_v2(wb: Workbook, ctx: BuildContext) -> None:
    facts = ctx.facts
    ws = wb["Guide"]
    _setup_sheet_v2(
        ws, ctx, f"{facts.company} — Model Guide",
        "One-screen reading guide: what is decided here, how the model works, and how to read every cell.",
        period_axis=False,
    )
    _set_column_widths(ws, {3: 34, 4: 118})
    r = 6

    def block(title: str, rows: list) -> None:
        nonlocal r
        _v2_section(ws, ctx, r, title, end_col=4)
        r += 1
        for label, text in rows:
            ws.cell(r, 3, label)
            ib.apply_label(ws.cell(r, 3), bold=True)
            ws.cell(r, 4, text)
            ib.apply_comment(ws.cell(r, 4), wrap_text=False)
            r += 1
        r += 1

    axis = ctx.axis
    grain_text = (
        f"hybrid: {axis.monthly_count} monthly columns "
        f"({axis.labels[0]}–{axis.labels[axis.monthly_count - 1]}) then annual FY columns"
        if axis.monthly_count and axis.monthly_count < len(axis.labels)
        else f"{axis.grain} × {len(axis.labels)} periods"
    )
    block("Decision & mechanics", [
        ("Decision", "Fundraising / board-grade operating plan: pricing, growth, burn, and funding need."),
        ("Mechanics", facts.mechanics),
        ("Economic unit", facts.primary_unit_name),
        ("Time axis", f"{grain_text}; every period formula references the months ruler (row 5)."),
        ("Currency", facts.currency),
        ("Fiscal year", f"FY labels end in month {facts.fiscal_year_end_month}."),
    ])
    block("Source boundary", [
        ("Stated inputs", "Rows marked 'stated input' in the Evidence status column come from the structured input."),
        ("Estimates", f"Everything else is {_compact_status(facts.evidence_status)} — replace with sourced values before circulation."),
        ("Placeholders", "Rows marked 'placeholder' are modeling defaults (e.g. 3-month onboarding) and must be validated."),
        ("Unknowns", "; ".join(facts.source_unknowns) if facts.source_unknowns else "none listed"),
    ])
    block("Sheet map", [
        (name, _V2_SHEET_PURPOSES.get(name, "See sheet header."))
        for name in (*SOURCE_PLAN_SHEETS_V2, *CONDITIONAL_SHEETS_V2, *CAP_TABLE_MODE_SHEETS)
        if name in ctx.bundle and name != "Guide"
    ])
    block("Formatting key", [
        ("Blue value", "Editable hard input (the only cells meant to be changed by hand)."),
        ("Black formula", "Same-sheet calculation; green formula = cross-sheet link; red = external link."),
        ("Italic", "Generated snapshot or estimate (ink-colored italic — a value, not a live formula)."),
        ("Gray italic", "Notes, source commentary, purpose lines, and unit captions."),
        ("▲ / '-'", "▲ marks negative values (JP convention); '-' renders a true zero."),
        ("Checks", "Check cells hold a numeric delta with tolerance; format shows OK at 0 and ERROR otherwise. Summary aggregates them into one master check."),
        ("Sign convention", "P&L costs are positive; cash flow inflows positive / outflows negative."),
        ("Hybrid boundary", "The monthly window ends at the medium vertical rule; annual columns follow. Formulas are identical across the boundary via the months ruler."),
        ("Units", "Sheet scale is declared in the C4 caption; per-unit and % rows carry their own unit in column E."),
    ])
    block("Model qualifications", [
        ("Tax timing", "消費税・源泉/社保は残高方式で簡略計上(中間納付は無視)。法人税はNOLを考慮しない期間独立課税(カーネルはNOL考慮)。"),
        ("Monthly interpolation", "Annual drivers are interpolated to months (flow ramp / stock line / rate hold); monthly figures are planning-grade."),
        ("Financing timing", "期首調達仮定 — 各会計年度の調達額はそのFYの第1月列に一括計上(declared exception)。"),
        ("Working capital", "AR/AP/前受/税残高は年率換算ベースの残高式(グレイン非依存)。A/P明細行は未対応。"),
        ("Actuals", "Actuals input is not yet supported — every column is plan."),
        ("Interest", "Interest accrues on the opening debt balance (no iterative circularity)."),
    ])


# --- v2 orchestrator ----------------------------------------------------------

_V2_TAB_ORDER = [
    "Guide",
    "Summary",
    "Assumptions",
    "Revenue Build",
    "Cost Build",
    "People Plan",
    "Pricing",
    "Unit Economics",
    "P&L",
    "BS",
    "CF",
    "Financing",
    "Cap Table",
    "Segments",
    "Evidence",
    "Valuation & Exit",
    "IC Memo",
    "Kernel",
    "Ownership",
]


_A1_ROW_REF_RE = re.compile(
    r"(?:'([^']+)'!)?\$?[A-Z]{1,3}\$?(\d+)(?::\$?[A-Z]{1,3}\$?(\d+))?"
)


def _referenced_rows(wb: Workbook, target_sheet: str) -> set:
    """Rows of `target_sheet` referenced by any formula in the workbook.

    Sheet-qualified references resolve explicitly; bare references resolve to
    the formula's own sheet. Range references mark every spanned row."""
    rows: set = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                body = re.sub(r'"[^"]*"', "", value)
                for match in _A1_ROW_REF_RE.finditer(body):
                    sheet = match.group(1) or ws.title
                    if sheet != target_sheet:
                        continue
                    first = int(match.group(2))
                    last = int(match.group(3)) if match.group(3) else first
                    rows.update(range(min(first, last), max(first, last) + 1))
    return rows


# Assumptions keys exempt from the no-dead-driver-rows rule: the declared
# info-only row (決算月) and the scenario mechanism (marker-required rows —
# the toggle / case / effective-scale block is the workbook's scenario UI).
_ASSUMPTION_KEYS_EXEMPT = ("fye", "scenario_toggle", "case_header")
_ASSUMPTION_KEY_EXEMPT_PREFIXES = ("case_", "eff_")


def _dead_assumption_keys(wb: Workbook, ctx: BuildContext) -> set:
    """Driver-row keys on the Assumptions register with no formula consumer."""
    if "Assumptions" not in wb.sheetnames:
        return set()
    referenced = _referenced_rows(wb, "Assumptions")
    dead: set = set()
    for key, row in ctx.rows.get("Assumptions", {}).items():
        if not isinstance(row, int):
            continue
        if key in _ASSUMPTION_KEYS_EXEMPT or key.startswith(_ASSUMPTION_KEY_EXEMPT_PREFIXES):
            continue
        if row not in referenced:
            dead.add(key)
    return dead


def build_plan_workbook_v2(facts: SourceFacts, bundle) -> Workbook:
    """Build a v2 workbook containing exactly the sheets in `bundle`.

    Two-pass: pass 1 renders every candidate driver row; any Assumptions row
    that no formula consumes is suppressed and the workbook is rebuilt, so a
    driver row exists only when a live consumer exists in the bundle
    (no-dead-driver-rows rule; 決算月-style info rows carry an "(info)"
    marker and are exempt)."""
    wb, ctx = _build_plan_workbook_v2_once(facts, bundle, suppress=set())
    if "Assumptions" in ctx.bundle:
        dead = _dead_assumption_keys(wb, ctx)
        if dead:
            wb, _ctx = _build_plan_workbook_v2_once(facts, bundle, suppress=dead)
    return wb


def _build_plan_workbook_v2_once(facts: SourceFacts, bundle, *, suppress: set) -> tuple:
    """One build pass (see build_plan_workbook_v2).

    Builders run in dependency order (Assumptions -> engines -> statements ->
    financing/registers -> Summary -> memo/guide) and only reference sheets
    inside the bundle (weak dependencies resolve to compact in-sheet blocks)."""
    bundle_list = list(bundle)
    bundle_set = set(bundle_list)
    wb = Workbook()
    wb.remove(wb.active)
    order = [s for s in _V2_TAB_ORDER if s in bundle_set]
    order += [s for s in bundle_list if s not in order]
    for name in order:
        ws = wb.create_sheet(name)
        ws._startup_facts = facts
    ctx = BuildContext(facts=facts, axis=build_period_axis(facts), bundle=bundle_set,
                       suppress_assumptions=set(suppress))
    build_order = (
        ("Assumptions", _build_assumptions_v2),
        ("Revenue Build", _build_revenue_v2),
        ("People Plan", _build_people_v2),
        ("Cost Build", _build_cost_v2),
        ("Pricing", _build_pricing_v2),
        ("Unit Economics", _build_unit_economics_v2),
        ("P&L", _build_pl_v2),
        ("CF", _build_cf_v2),
        ("BS", _build_bs_v2),
        ("Financing", _build_financing_v2),
        ("Cap Table", _build_cap_table_v2),
        ("Segments", _build_segments_v2),
        ("Evidence", _build_evidence_v2),
        ("Valuation & Exit", _build_valuation_exit_v2),
        ("Summary", _build_summary_v2),
        ("IC Memo", _build_ic_memo_v2),
        ("Kernel", _build_kernel_v2),
        ("Guide", _build_guide_v2),
    )
    for name, builder in build_order:
        if name in bundle_set:
            builder(wb, ctx)
    # Master-check echo in the frozen header of every v2 period-axis sheet —
    # including Summary itself (the master check sits below its freeze row).
    if ctx.master_check_cell:
        echo = f"=IF('Summary'!{ctx.master_check_cell}=0,\"checks OK\",\"CHECK FAILED\")"
        for ws in wb.worksheets:
            if not getattr(ws, "_v2_period_sheet", False):
                continue
            ib.write_master_check_echo(
                ws, f"{get_column_letter(V2_FIRST_PERIOD_COL)}2", echo)
    # Post passes. The "Ownership" sheet (cap_table state machine) is styled
    # by cap_table_builder after this returns; every other sheet is v2.
    ib.normalize_workbook_fonts(wb)
    ib.set_workbook_default_font(wb)
    _disable_wrap_text(wb)
    _autosize_default_layout_columns(wb)
    _clear_blank_cell_styles(wb)
    _trim_blank_canvas(wb)
    for ws in wb.worksheets:
        if getattr(ws, "_v2_period_sheet", False):
            _v2_mark_grain_boundary(ws)
    _apply_print(wb)
    wb.defined_names.clear()
    for ws in wb.worksheets:
        ws.defined_names.clear()
    return wb, ctx


def build_source_plan_workbook_from_facts(facts: SourceFacts) -> Workbook:
    """Full 12-sheet v2 workbook (BS dropped when immaterial)."""
    return build_plan_workbook_v2(
        facts, full_bundle_for_facts(facts, SOURCE_PLAN_SHEETS_V2))


def build_source_plan_workbook_from_text(text: str, output_path: Path) -> Path:
    facts = derive_source_facts(text)
    wb = build_source_plan_workbook_from_facts(facts)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_source_plan_workbook(source_md: Path, output_path: Path) -> Path:
    wb = build_source_plan_workbook_from_facts(extract_source_facts(source_md))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Build generic economic-kernel startup financial plan xlsx.")
    ap.add_argument("--source-md", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    args = ap.parse_args()
    build_source_plan_workbook(args.source_md, args.output)
    print(f"[ok] generic financial plan generated: {args.output}")
