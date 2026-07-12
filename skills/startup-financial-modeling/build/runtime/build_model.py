"""Build startup-financial-modeling xlsx outputs.

Primary route:
    python build_model.py --source-md equity_story.md --output plan.xlsx

That route generates a generic economic-kernel startup plan. The mode path is a
focused-module convenience for narrow requests such as pricing, cap table,
runway, valuation, market sizing, or three-statement mechanics. Focused modes
start from the same kernel but keep an explicit visible-surface contract.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import replace
from pathlib import Path
from typing import Any

try:
    import yaml as _yaml  # type: ignore
except ImportError:
    _yaml = None  # YAML loader is optional; defaults work without it.

import ib_format as ibf  # noqa: E402
import source_plan_builder as spb  # noqa: E402
import cap_table_builder as ctb  # noqa: E402
import economic_kernel as ek  # noqa: E402
import live_comps as lc  # noqa: E402

# ============================================================================
# Mode -> seed sheet bundle for focused module outputs.
# ============================================================================

MODE_BUNDLE_SEEDS: dict[str, list[str]] = {
    "full": [
        "Guide", "Kernel", "Assumptions", "Driver Tree", "Revenue Build",
        "Cost Build", "People Plan", "P&L", "BS", "CF", "Capital Stack",
        "Ownership", "Pricing", "Financing", "Exit Waterfall", "Segments",
        "KPI", "Scenarios", "Sensitivity", "Valuation", "Market Support",
        "Benchmarks", "IC Memo",
    ],
    "pricing": [
        "Guide", "Kernel", "Assumptions", "Driver Tree", "Pricing", "CF", "Valuation",
    ],
    "unit_economics": [
        "Guide", "Kernel", "Assumptions", "Driver Tree", "KPI", "Scenarios",
    ],
    "cap_table": [
        "Guide", "Kernel", "Ownership",
    ],
    "ma_exit": [
        "Guide", "Kernel", "Exit Waterfall", "Scenarios", "Sensitivity",
        "Valuation", "IC Memo",
    ],
    "dcf_only": [
        "Guide", "Kernel", "Valuation", "Sensitivity",
    ],
    "burn_runway": [
        "Guide", "Kernel", "CF", "Capital Stack", "Financing", "KPI",
    ],
    "three_statement": [
        "Guide", "Kernel", "P&L", "BS", "CF",
    ],
    "market_sizing": [
        "Guide", "Kernel", "Driver Tree", "Market Support", "Benchmarks",
    ],
    "comps_only": [
        "Guide", "Kernel", "Valuation", "Market Support", "Benchmarks", "IC Memo",
    ],
}

SHEET_DEPENDENCIES: dict[str, list[str]] = {
    "Assumptions": ["Revenue Build", "Cost Build", "CF", "Scenarios"],
    "Driver Tree": ["Kernel"],
    "Revenue Build": ["Assumptions"],
    "Cost Build": ["Assumptions", "Revenue Build"],
    "People Plan": ["Assumptions", "Revenue Build", "Cost Build"],
    "P&L": ["Assumptions", "Revenue Build", "Cost Build", "People Plan", "Capital Stack"],
    "BS": ["Assumptions", "P&L", "CF", "Financing"],
    "CF": ["Assumptions", "P&L", "BS"],
    "Capital Stack": ["CF", "Ownership", "Valuation"],
    # Empty by design: pure data sheet; the cap-table route replaces it with state-machine rows.
    "Ownership": [],
    "Pricing": ["Assumptions"],
    "Financing": ["Capital Stack"],
    "Exit Waterfall": ["Capital Stack", "Ownership", "Valuation"],
    "Segments": [],
    "KPI": ["Assumptions"],
    "Scenarios": ["Assumptions"],
    "Sensitivity": ["Assumptions"],
    "Valuation": ["Assumptions", "Ownership", "Segments", "KPI", "Benchmarks"],
    "Market Support": ["Kernel"],
    "Benchmarks": ["Kernel"],
    "IC Memo": ["Kernel", "KPI", "Scenarios", "Capital Stack", "Ownership", "Valuation", "Pricing", "Exit Waterfall"],
}

VALID_MODES: list[str] = list(MODE_BUNDLE_SEEDS.keys())
VALID_SHEETS: set[str] = set(spb.SOURCE_PLAN_SHEETS)
DEPENDENCY_EXPANDED_MODES = set(MODE_BUNDLE_SEEDS)
DEFAULT_PUBLIC_COMP_TICKERS: dict[str, list[str]] = {
    "Recurring software": ["CRM", "NOW", "DDOG"],
    "Marketplace / transaction": ["UBER", "DASH", "ETSY"],
    "Hardware / asset-heavy": ["ISRG", "ROK", "TSLA"],
    "Fintech / balance-sheet": ["PYPL", "SOFI", "AFRM"],
    # Public-market data is only a fallback for pre-revenue companies; users
    # should add private milestone, funding-round, or transaction evidence.
    "Pre-revenue / milestone": ["ISRG", "ROK", "TSLA"],
    "Generic startup": ["CRM", "NOW", "DDOG"],
}


def _validate_sheet_names(sheet_names: list[str], *, arg_name: str) -> None:
    unknown = [sheet for sheet in sheet_names if sheet not in VALID_SHEETS]
    if unknown:
        raise ValueError(f"Unknown sheet name(s) in {arg_name}: {unknown}. Valid: {spb.SOURCE_PLAN_SHEETS}")


def _missing_dependencies(bundle: list[str]) -> list[str]:
    selected = set(bundle)
    missing: list[str] = []
    for sheet in bundle:
        for dependency in SHEET_DEPENDENCIES.get(sheet, []):
            if dependency not in selected:
                missing.append(f"{sheet} requires {dependency}")
    return missing


def resolve_bundle(
    mode: str,
    additional_sheets: list[str] | None = None,
    excluded_sheets: list[str] | None = None,
) -> list[str]:
    """Return the exact sheet bundle generated for a mode plus overrides."""
    if mode not in MODE_BUNDLE_SEEDS:
        raise ValueError(f"Unknown mode {mode!r}. Valid: {VALID_MODES}")
    if additional_sheets:
        _validate_sheet_names(additional_sheets, arg_name="additional_sheets")
    if excluded_sheets:
        _validate_sheet_names(excluded_sheets, arg_name="excluded_sheets")

    bundle: list[str] = []
    pending = list(MODE_BUNDLE_SEEDS[mode])
    expand_seed_dependencies = mode in DEPENDENCY_EXPANDED_MODES
    while pending:
        sheet = pending.pop(0)
        if sheet in bundle:
            continue
        bundle.append(sheet)
        if expand_seed_dependencies:
            pending.extend(SHEET_DEPENDENCIES.get(sheet, []))
    if additional_sheets:
        pending = [s for s in additional_sheets if s not in bundle]
        while pending:
            sheet = pending.pop(0)
            if sheet in bundle:
                continue
            bundle.append(sheet)
            pending.extend(SHEET_DEPENDENCIES.get(sheet, []))
    if excluded_sheets:
        bundle = [s for s in bundle if s not in excluded_sheets]
        missing = _missing_dependencies(bundle)
        if missing:
            raise ValueError(
                "excluded_sheets would create broken workbook references. "
                "Also exclude dependent sheets or choose a focused mode. "
                f"Missing dependencies: {missing}"
            )
    canonical_order = {sheet: idx for idx, sheet in enumerate(spb.SOURCE_PLAN_SHEETS)}
    return sorted(bundle, key=lambda sheet: canonical_order.get(sheet, len(canonical_order)))


# ============================================================================
# Input loading
# ============================================================================


def load_yaml(path: Path) -> dict[str, Any]:
    """Load optional YAML input for focused module outputs."""
    if _yaml is None:
        raise RuntimeError(
            "PyYAML not installed. Run: pip install pyyaml "
            "(or omit --input to use generic economic-kernel defaults)."
        )
    with path.open(encoding="utf-8") as f:
        data = _yaml.safe_load(f) or {}
    if not isinstance(data, dict):
        raise ValueError(f"YAML root must be a mapping, got {type(data).__name__}")
    return data


def _clear_defined_names(wb: Any) -> None:
    """Remove workbook and sheet-scoped names from generated models."""
    wb.defined_names.clear()
    for ws in wb.worksheets:
        ws.defined_names.clear()


_SHEET_REF_RE = re.compile(r"'([^']+)'!")
VOLATILE_FORMULA_RE = re.compile(
    r"\b(?:OFFSET|INDIRECT|RAND|RANDBETWEEN|NOW|TODAY|CELL|INFO)\s*\(",
    re.IGNORECASE,
)


def _terminal_comment_col(ws: Any) -> int | None:
    if not spb.uses_default_layout(ws):
        return None
    headers = [
        cell
        for cell in ws[5]
        if cell.column >= spb.LAYOUT.first_value_col and cell.value not in (None, "")
    ]
    if headers and headers[-1].value == "Comment":
        return int(headers[-1].column)
    return None


def _neutralize_removed_sheet_references(wb: Any) -> int:
    """Replace formulas that point to sheets omitted from a focused bundle.

    Focused workbooks must remain openable and auditable without `#REF!` or
    hidden dependencies. When a helper sheet keeps a full-model support formula
    after its target sheet has been intentionally omitted, convert only the
    affected output cell to an editable zero-value placeholder.
    """
    available = set(wb.sheetnames)
    touched = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                targets = set(_SHEET_REF_RE.findall(cell.value))
                if targets and not targets <= available:
                    missing = ", ".join(sorted(targets - available))
                    original = cell.value
                    number_format = cell.number_format
                    cell.value = 0
                    ibf.apply_hard_input(cell, number_format)
                    source_cell = ws.cell(row=cell.row, column=spb.LAYOUT.source_col)
                    if source_cell.value in (None, ""):
                        source_cell.value = f"compact mode placeholder; omitted sheet(s): {missing}"
                    elif "compact mode placeholder" not in str(source_cell.value):
                        source_cell.value = f"{source_cell.value}; compact mode placeholder for {missing}"
                    ibf.apply_comment(source_cell, wrap_text=False)
                    note_col = _terminal_comment_col(ws) or max(ws.max_column + 1, 10)
                    ws.column_dimensions[spb.get_column_letter(note_col)].width = spb.LAYOUT.note_width
                    note_cell = ws.cell(row=cell.row, column=note_col)
                    note_cell.value = f"Original formula omitted from focused bundle: {original}"
                    ibf.apply_comment(note_cell, wrap_text=False)
                    touched += 1
    return touched


REQUIRED_SHEET_MARKERS: dict[str, list[str]] = {
    "Guide": ["Purpose", "Workbook map", "Sheet-level acceptance criteria"],
    "Kernel": ["Decision", "Model grain", "Mechanics", "Unknowns"],
    "Assumptions": ["Volume and demand", "Revenue adjuncts", "Unit cost and delivery"],
    "Driver Tree": ["Demand", "Monetization", "Delivery", "Value"],
    "Revenue Build": ["Revenue drivers", "Revenue streams"],
    "Cost Build": ["Gross profit bridge", "Gross margin"],
    "People Plan": ["Total headcount", "Revenue / FTE"],
    "P&L": ["Total revenue", "EBITDA", "Net income"],
    "BS": ["Total assets", "Balance check"],
    "CF": ["Operating cash flow", "Ending cash", "Free cash flow"],
    "Capital Stack": ["Runway", "New investor ownership", "Debt / revenue"],
    # Empty by design: pure data sheet; no prose section markers are required.
    "Ownership": [],
    "Pricing": ["Customer ROI / year", "Pricing validation plan"],
    "Financing": ["Financing cash inflow", "Downside funding gap"],
    "Exit Waterfall": ["Buyer-view M&A bridge", "Preference floor", "Walk-away signal"],
    "Segments": ["Segment revenue", "EBITDA proxy", "Segment EV"],
    "KPI": ["KPI interpretation register", "Evidence coverage"],
    "Scenarios": ["Scenario interpretation", "DD action"],
    "Sensitivity": ["Sensitivity rationale", "Decision implication"],
    "Valuation": ["Method credibility", "Selected EV midpoint", "Valuation committee gates"],
    "Market Support": ["Source anchors", "TAM / SAM / SOM bridge"],
    # The generator intentionally emits the source-register key as a column header.
    "Benchmarks": ["source_id", "Comparable evidence", "Refresh needed"],
    "IC Memo": ["Recommendation", "KPI readout", "Ranked DD gates"],
}


def _defined_name_count(wb: Any) -> int:
    count = len(wb.defined_names)
    for ws in wb.worksheets:
        count += len(ws.defined_names)
    return count


def _font_rgb(cell: Any) -> str | None:
    color = cell.font.color
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    return rgb[-6:].upper() if isinstance(rgb, str) else None


def _sheet_values(ws: Any) -> set[str]:
    return {
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    }


def _audit_font_design(wb: Any) -> list[str]:
    issues: list[str] = []
    allowed_sizes = {float(size) for size in ibf.FONT_SIZE_ALLOWED_CELLS}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if cell.font.name is not None and cell.font.name != ibf.FONT_FAMILY:
                    issues.append(f"{ws.title}!{cell.coordinate} uses non-standard font {cell.font.name}")
                if cell.font.sz is not None and float(cell.font.sz) not in allowed_sizes:
                    issues.append(f"{ws.title}!{cell.coordinate} uses non-standard font size {cell.font.sz}")
    return issues


def _audit_semantic_alignment(wb: Any) -> list[str]:
    issues: list[str] = []
    for ws in wb.worksheets:
        if not spb.uses_default_layout(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                alignment = cell.alignment
                horizontal = alignment.horizontal
                indent = getattr(alignment, "indent", 0)
                if cell.row == 5 and cell.column >= spb.LAYOUT.first_value_col:
                    if cell.value == "Comment":
                        if horizontal not in (None, "left"):
                            issues.append(f"{ws.title}!{cell.coordinate} comment header cell is {horizontal}, expected left")
                    elif horizontal not in (None, "center"):
                        issues.append(f"{ws.title}!{cell.coordinate} period/header cell is {horizontal}, expected center")
                elif cell.row != 5 and cell.column == spb.LAYOUT.source_col:
                    rgb = _font_rgb(cell)
                    if horizontal not in (None, "left") or not cell.font.italic or (rgb is not None and rgb != ibf.IB_COMMENT):
                        issues.append(f"{ws.title}!{cell.coordinate} source cell alignment/font is non-standard")
                elif cell.row != 5 and cell.column == spb.LAYOUT.unit_col:
                    rgb = _font_rgb(cell)
                    if horizontal not in (None, "right") or (rgb is not None and rgb != ibf.IB_COMMENT):
                        issues.append(f"{ws.title}!{cell.coordinate} unit cell alignment/font is non-standard")
                elif cell.row != 5 and cell.column in (spb.LAYOUT.first_hierarchy_col, spb.LAYOUT.label_col):
                    if horizontal not in (None, "left") or indent:
                        issues.append(f"{ws.title}!{cell.coordinate} label cell is {horizontal} indent={indent}, expected left/no indent")
                elif cell.row != 5 and cell.column >= spb.LAYOUT.first_value_col:
                    value = cell.value
                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith("=")):
                        if horizontal not in (None, "right"):
                            issues.append(f"{ws.title}!{cell.coordinate} numeric/formula cell is {horizontal}, expected right")
    return issues


def _audit_terminal_comment_columns(wb: Any) -> list[str]:
    issues: list[str] = []
    for ws in wb.worksheets:
        if not spb.uses_default_layout(ws):
            continue
        headers = [
            cell
            for cell in ws[5]
            if cell.column >= spb.LAYOUT.first_value_col and cell.value not in (None, "")
        ]
        if not headers:
            continue
        comment_header = headers[-1]
        expected_width = float(spb.LAYOUT.note_width)
        actual_width = ws.column_dimensions[comment_header.column_letter].width
        if comment_header.value != "Comment":
            issues.append(
                f"{ws.title} row 5 must end the value block with a Comment column; "
                f"found {comment_header.coordinate}={comment_header.value!r}"
            )
        if actual_width is None or abs(float(actual_width) - expected_width) > 0.001:
            issues.append(
                f"{ws.title}!{comment_header.column_letter} width is {actual_width}, "
                f"expected terminal comment width {expected_width:g}"
            )
        for row in ws.iter_rows(min_col=comment_header.column, max_col=comment_header.column):
            cell = row[0]
            if cell.row == 5 or cell.value in (None, ""):
                continue
            horizontal = cell.alignment.horizontal if cell.alignment is not None else None
            rgb = _font_rgb(cell)
            if horizontal not in (None, "left") or not cell.font.italic or (rgb is not None and rgb != ibf.IB_COMMENT):
                issues.append(f"{ws.title}!{cell.coordinate} terminal Comment cell style is non-standard")
        for row in ws.iter_rows(min_col=comment_header.column + 1):
            for cell in row:
                if cell.value not in (None, ""):
                    issues.append(
                        f"{ws.title}!{cell.coordinate} sits to the right of the terminal Comment column"
                    )
    return issues


def _audit_formula_engine(wb: Any) -> list[str]:
    issues: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                match = VOLATILE_FORMULA_RE.search(value)
                if match:
                    issues.append(
                        f"{ws.title}!{cell.coordinate} uses volatile formula function {match.group(0).rstrip('(').upper()}"
                    )
    return issues


def audit_workbook(wb: Any) -> list[str]:
    """Return generation issues that should block an investor-ready handoff."""
    issues: list[str] = []
    available = set(wb.sheetnames)
    if _defined_name_count(wb):
        issues.append("workbook must not contain workbook-scoped or sheet-scoped defined names")
    for ws in wb.worksheets:
        if ws.freeze_panes is not None:
            issues.append(f"{ws.title} has frozen panes")
        if ws.merged_cells.ranges:
            issues.append(f"{ws.title} has merged cell range(s): {', '.join(str(rng) for rng in ws.merged_cells.ranges)}")
        sheet_values = _sheet_values(ws)
        missing_markers = [marker for marker in REQUIRED_SHEET_MARKERS.get(ws.title, []) if marker not in sheet_values]
        if missing_markers:
            issues.append(f"{ws.title} is missing sheet-quality marker(s): {', '.join(missing_markers)}")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if cell.alignment is not None and cell.alignment.wrap_text is True:
                    issues.append(f"{ws.title}!{cell.coordinate} has wrap_text enabled")
                if isinstance(value, str) and "\n" in value:
                    issues.append(f"{ws.title}!{cell.coordinate} contains a manual line break")
                if isinstance(value, str) and value.startswith("="):
                    missing = sorted(set(_SHEET_REF_RE.findall(value)) - available)
                    if missing:
                        issues.append(f"{ws.title}!{cell.coordinate} references omitted sheet(s): {', '.join(missing)}")
                if isinstance(value, str) and "#REF!" in value:
                    issues.append(f"{ws.title}!{cell.coordinate} contains #REF!")
    issues.extend(_audit_font_design(wb))
    issues.extend(_audit_semantic_alignment(wb))
    issues.extend(_audit_terminal_comment_columns(wb))
    issues.extend(_audit_formula_engine(wb))
    return issues


def _row_for_label(ws: Any, label: str) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return int(cell.row)
    return None


def _period_columns(ws: Any) -> list[int]:
    return [
        int(cell.column)
        for cell in ws[5]
        if cell.column >= spb.LAYOUT.first_value_col
        and cell.value not in (None, "", "Comment")
    ]


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _sheet_value(ws: Any, label: str, col: int) -> float | None:
    row = _row_for_label(ws, label)
    if row is None:
        return None
    return _numeric(ws.cell(row=row, column=col).value)


def _financial_close(actual: float, expected: float, *, scale: float = 1.0) -> bool:
    tolerance = max(1.0, abs(scale) * 1e-6)
    return abs(actual - expected) <= tolerance


def _append_close_issue(
    issues: list[str],
    ws_name: str,
    label: str,
    period_label: Any,
    actual: float | None,
    expected: float | None,
    *,
    scale: float = 1.0,
) -> None:
    if actual is None or expected is None:
        return
    if not _financial_close(actual, expected, scale=scale):
        issues.append(
            f"{ws_name} {period_label}: {label} is {actual:,.0f}, expected {expected:,.0f}"
        )


def audit_recalculated_financial_model(wb: Any) -> list[str]:
    """Audit recalculated `data_only=True` workbook values for finance integrity.

    Formula-string and style audits prove the workbook is editable and clean.
    This pass proves the recalculated values still satisfy the core IB model
    identities across P&L, BS, CF, capital stack, ownership, and valuation.
    """
    issues: list[str] = []

    if "P&L" in wb.sheetnames:
        ws = wb["P&L"]
        for col in _period_columns(ws):
            period = ws.cell(row=5, column=col).value
            revenue = _sheet_value(ws, "Total revenue", col)
            cogs = _sheet_value(ws, "Total COGS", col)
            gross_profit = _sheet_value(ws, "Gross profit", col)
            opex = _sheet_value(ws, "Total OpEx", col)
            ebitda = _sheet_value(ws, "EBITDA", col)
            da = _sheet_value(ws, "D&A", col)
            ebit = _sheet_value(ws, "EBIT", col)
            interest = _sheet_value(ws, "Interest expense", col)
            ebt = _sheet_value(ws, "EBT", col)
            tax = _sheet_value(ws, "Cash tax", col)
            net_income = _sheet_value(ws, "Net income", col)
            scale = max(abs(revenue or 0), 1.0)
            if revenue is not None and cogs is not None:
                _append_close_issue(issues, "P&L", "Gross profit", period, gross_profit, revenue - cogs, scale=scale)
            if gross_profit is not None and opex is not None:
                _append_close_issue(issues, "P&L", "EBITDA", period, ebitda, gross_profit - opex, scale=scale)
            if ebitda is not None and da is not None:
                _append_close_issue(issues, "P&L", "EBIT", period, ebit, ebitda - da, scale=scale)
            if ebit is not None and interest is not None:
                _append_close_issue(issues, "P&L", "EBT", period, ebt, ebit - interest, scale=scale)
            if ebt is not None and tax is not None:
                _append_close_issue(issues, "P&L", "Net income", period, net_income, ebt - tax, scale=scale)

    if {"BS", "CF"} <= set(wb.sheetnames):
        bs = wb["BS"]
        cf = wb["CF"]
        for col in _period_columns(bs):
            period = bs.cell(row=5, column=col).value
            total_assets = _sheet_value(bs, "Total assets", col) or 0.0
            balance_check = _sheet_value(bs, "Balance check", col)
            if balance_check is not None and not _financial_close(balance_check, 0.0, scale=max(total_assets, 1.0)):
                issues.append(f"BS {period}: Balance check is {balance_check:,.0f}, expected 0")
            _append_close_issue(
                issues,
                "BS/CF",
                "Cash equals ending cash",
                period,
                _sheet_value(bs, "Cash", col),
                _sheet_value(cf, "Ending cash", col),
                scale=max(total_assets, 1.0),
            )

    if "CF" in wb.sheetnames:
        ws = wb["CF"]
        for col in _period_columns(ws):
            period = ws.cell(row=5, column=col).value
            ni = _sheet_value(ws, "Net income", col)
            da = _sheet_value(ws, "D&A", col)
            ar = _sheet_value(ws, "AR increase", col)
            inv = _sheet_value(ws, "Inventory increase", col)
            ap = _sheet_value(ws, "AP / deferred revenue increase", col)
            ocf = _sheet_value(ws, "Operating cash flow", col)
            capex = _sheet_value(ws, "CapEx", col)
            fcf = _sheet_value(ws, "Free cash flow", col)
            equity = _sheet_value(ws, "Equity financing", col)
            debt = _sheet_value(ws, "Debt financing", col)
            grants = _sheet_value(ws, "Grants / subsidies", col)
            secondary = _sheet_value(ws, "Secondary liquidity", col)
            net_cash_flow = _sheet_value(ws, "Net cash flow", col)
            beginning = _sheet_value(ws, "Beginning cash", col)
            ending = _sheet_value(ws, "Ending cash", col)
            runway = _sheet_value(ws, "Runway months", col)
            scale = max(abs(ending or 0), abs(beginning or 0), 1.0)
            if None not in (ni, da, ar, inv, ap):
                _append_close_issue(issues, "CF", "Operating cash flow", period, ocf, ni + da + ar + inv + ap, scale=scale)
            if ocf is not None and capex is not None:
                _append_close_issue(issues, "CF", "Free cash flow", period, fcf, ocf + capex, scale=scale)
            if None not in (fcf, equity, debt, grants, secondary):
                _append_close_issue(issues, "CF", "Net cash flow", period, net_cash_flow, fcf + equity + debt + grants + secondary, scale=scale)
            if beginning is not None and net_cash_flow is not None:
                _append_close_issue(issues, "CF", "Ending cash", period, ending, beginning + net_cash_flow, scale=scale)
            if runway is not None and (runway < -1e-9 or runway > 99.000001):
                issues.append(f"CF {period}: Runway months is {runway:,.1f}, expected between 0 and 99")

    if {"Capital Stack", "CF"} <= set(wb.sheetnames):
        capital = wb["Capital Stack"]
        cf = wb["CF"]
        for col in _period_columns(capital):
            period = capital.cell(row=5, column=col).value
            post_money = _sheet_value(capital, "Illustrative post-money", col)
            equity = _sheet_value(capital, "Equity financing", col)
            ownership = _sheet_value(capital, "New investor ownership", col)
            debt_balance = _sheet_value(capital, "Debt balance", col)
            scale = max(abs(equity or 0), abs(post_money or 0), 1.0)
            _append_close_issue(issues, "Capital Stack/CF", "Ending cash", period, _sheet_value(capital, "Ending cash", col), _sheet_value(cf, "Ending cash", col), scale=scale)
            _append_close_issue(issues, "Capital Stack/CF", "Runway", period, _sheet_value(capital, "Runway", col), _sheet_value(cf, "Runway months", col), scale=99.0)
            if post_money and post_money > 0 and equity is not None:
                _append_close_issue(issues, "Capital Stack", "New investor ownership", period, ownership, equity / post_money, scale=1.0)
            if debt_balance is not None and debt_balance < -1.0:
                issues.append(f"Capital Stack {period}: Debt balance is negative ({debt_balance:,.0f})")

    if "Ownership" in wb.sheetnames:
        ws = wb["Ownership"]
        for col in _period_columns(ws):
            period = ws.cell(row=5, column=col).value
            check = _sheet_value(ws, "Ownership check", col)
            if check is not None and abs(check - 1.0) > 0.0001:
                issues.append(f"Ownership {period}: Ownership check is {check:.4%}, expected 100.00%")

    if "Valuation" in wb.sheetnames:
        ws = wb["Valuation"]
        for idx, col in enumerate(_period_columns(ws), start=1):
            period = ws.cell(row=5, column=col).value
            low = _sheet_value(ws, "Selected EV low", col)
            midpoint = _sheet_value(ws, "Selected EV midpoint", col)
            high = _sheet_value(ws, "Selected EV high", col)
            invested = _sheet_value(ws, "Equity invested", col)
            ownership = _sheet_value(ws, "New investor ownership", col)
            moic = _sheet_value(ws, "MOIC at selected EV", col)
            irr = _sheet_value(ws, "Illustrative IRR", col)
            dcf = _sheet_value(ws, "DCF EV", col)
            financing_implied = _sheet_value(ws, "Financing-implied EV", col)
            primary_ev = _sheet_value(ws, "Primary-method EV", col)
            if None not in (low, midpoint, high) and not (0 <= low <= midpoint <= high):
                issues.append(
                    f"Valuation {period}: selected EV range is not ordered "
                    f"(low={low:,.0f}, midpoint={midpoint:,.0f}, high={high:,.0f})"
                )
            if dcf is not None and dcf <= 0 and not (financing_implied and financing_implied > 0) and not (primary_ev == 0):
                issues.append(f"Valuation {period}: DCF EV is non-positive ({dcf:,.0f})")
            if invested and invested > 0 and None not in (midpoint, ownership, moic):
                expected_moic = midpoint * ownership / invested
                _append_close_issue(issues, "Valuation", "MOIC at selected EV", period, moic, expected_moic, scale=max(abs(expected_moic), 1.0))
                if irr is not None and moic is not None and moic > 0:
                    expected_irr = moic ** (1 / idx) - 1
                    if abs(irr - expected_irr) > 0.0001:
                        issues.append(
                            f"Valuation {period}: Illustrative IRR is {irr:.2%}, "
                            f"expected {expected_irr:.2%}"
                        )

    return issues


def _default_source_text() -> str:
    return (
        "# Startup financial plan\n"
        "Generic startup model for pricing, runway, financing, ownership, "
        "valuation, scenarios, and investor diligence. No explicit external "
        "source is listed."
    )


def _source_text_from_yaml(raw: dict[str, Any]) -> str:
    for key in ("source_text", "narrative", "story", "memo"):
        value = raw.get(key)
        if isinstance(value, str) and value.strip():
            return value
    company = raw.get("company") or raw.get("company_name") or "Startup"
    mechanics = raw.get("mechanics") or raw.get("business_model") or "generic startup"
    return f"# {company}\nGeneric financial model for a {mechanics}."


def _first_number(values: Any, default: float) -> float:
    """Return the first numeric scalar from a scalar/list-like YAML value."""
    if isinstance(values, (int, float)):
        return float(values)
    if isinstance(values, list):
        for value in values:
            if isinstance(value, (int, float)):
                return float(value)
    return default


def _money_m_from_yen(value: float, default_money_m: float) -> float:
    """Convert raw yen to reporting-currency millions for cap-table helpers."""
    if not value:
        return default_money_m
    return max(value / 1_000_000, 1.0)


def _cap_table_input_from_facts(facts: Any) -> ctb.CapTableInput:
    """Build a user-facing cap-table input from extracted model facts."""
    base = ctb.get_default_input()
    post_money = _money_m_from_yen(
        _first_number(getattr(facts, "post_money_yen", None), 0.0),
        base.round_pre_money_money_m + base.round_investment_money_m,
    )
    investment = _money_m_from_yen(
        _first_number(getattr(facts, "equity_raise_yen", None), 0.0),
        base.round_investment_money_m or 800.0,
    )
    pre_money = max(post_money - investment, investment)
    return ctb.CapTableInput(
        company_name=getattr(facts, "company", None) or base.company_name,
        reporting_currency=getattr(facts, "currency", None) or base.reporting_currency,
        as_of_date=base.as_of_date,
        founder_shares=base.founder_shares,
        common_pool_issued=base.common_pool_issued,
        common_pool_available=base.common_pool_available,
        safes=base.safes,
        preferred=base.preferred,
        round_pre_money_money_m=pre_money,
        round_investment_money_m=investment,
        round_target_pool_pct=base.round_target_pool_pct,
        round_anti_dilution_default=base.round_anti_dilution_default,
        round_label=base.round_label,
    )


def _serialize_live_comp(comp: lc.PublicComp) -> dict[str, object]:
    return {
        "ticker": comp.ticker,
        "name": comp.name,
        "company_type": comp.company_type,
        "source_type": comp.source_type,
        "stage": comp.stage,
        "geography": comp.geography,
        "applicability_limits": comp.applicability_limits,
        "currency": comp.currency,
        "market_cap": comp.market_cap,
        "enterprise_value": comp.enterprise_value,
        "revenue_multiple": comp.revenue_multiple,
        "ebitda_multiple": comp.ebitda_multiple,
        "gross_margin": comp.gross_margin,
        "ebitda_margin": comp.ebitda_margin,
        "source_url": comp.source_url,
        "as_of_date": comp.as_of_date,
        "status": comp.status,
        "error": comp.error,
    }


def _ticker_list_from_raw(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value] if value.strip() else []
    if not isinstance(value, list):
        return []
    return [str(item).strip().upper() for item in value if not isinstance(item, dict) and str(item).strip()]


def _provided_comps_from_raw_mapping(raw: dict[str, Any]) -> list[lc.PublicComp]:
    comps: list[lc.PublicComp] = []
    for key in ("live_comps", "public_comps", "private_comps", "transaction_comps", "benchmark_sources"):
        comps.extend(lc.provided_comps_from_raw(raw.get(key)))
    return comps


def _apply_live_comps_to_facts(
    facts: Any,
    tickers: list[str],
    *,
    timeout: float,
    provided_comps: list[lc.PublicComp] | None = None,
) -> Any:
    """Refresh public comps, merge provided private/transaction evidence, and apply valid medians."""
    if not tickers and not provided_comps:
        return facts
    fetched = lc.fetch_public_comps(tickers, timeout=timeout) if tickers else lc.summarize_comps([])
    result = lc.summarize_comps([*fetched.comps, *(provided_comps or [])], source_url="mixed comparable evidence")
    periods = len(getattr(facts, "period_labels", []) or [])
    overrides: dict[str, Any] = {
        "live_comps": [_serialize_live_comp(comp) for comp in result.comps],
    }
    if result.revenue_multiple_median and periods:
        overrides["revenue_multiple"] = [round(float(result.revenue_multiple_median), 1) for _ in range(periods)]
    if result.ebitda_multiple_median and periods:
        overrides["ebitda_multiple"] = [round(float(result.ebitda_multiple_median), 1) for _ in range(periods)]
    return replace(facts, **overrides)


def _default_live_comps_for_facts(facts: Any) -> list[str]:
    mechanics = str(getattr(facts, "mechanics", "") or "")
    return DEFAULT_PUBLIC_COMP_TICKERS.get(mechanics, DEFAULT_PUBLIC_COMP_TICKERS["Generic startup"])


# ============================================================================
# Build pipeline
# ============================================================================


def _facts_for_inputs(
    input_path: Path | None, source_md: Path | None
) -> tuple[Any, dict[str, Any]]:
    """Derive SourceFacts (and the raw YAML mapping) from CLI inputs.

    Single source of truth so the build path and the strict-audit economic
    coherence check operate on identical facts.
    """
    raw: dict[str, Any] = load_yaml(input_path) if input_path else {}
    if source_md is not None:
        return spb.extract_source_facts(source_md), raw
    if raw:
        return spb.derive_source_facts_from_mapping(raw), raw
    return spb.derive_source_facts(_default_source_text()), raw


def build_model(
    input_path: Path | None,
    output_path: Path,
    mode: str = "full",
    additional_sheets: list[str] | None = None,
    excluded_sheets: list[str] | None = None,
    source_md: Path | None = None,
    live_comps: list[str] | None = None,
    live_comps_timeout: float = 8.0,
) -> Path:
    """Build a startup finance xlsx model and save to `output_path`.

    Args:
        input_path: YAML input for focused mode outputs (None = defaults).
        output_path: Where to save the xlsx.
        mode: One of `VALID_MODES`.
        additional_sheets: Extra sheet IDs to add to the bundle.
        excluded_sheets: Sheet IDs to drop from the bundle.
        source_md: Narrative source markdown. When present, route to the
            generic economic-kernel workbook builder instead of focused mode.

    Returns:
        The output_path (for chaining). Use `resolve_bundle()` when callers also
        need the final generated sheet list.

    Raises:
        ValueError: Unknown mode.
    """
    bundle = resolve_bundle(mode, additional_sheets, excluded_sheets)

    facts, raw = _facts_for_inputs(input_path, source_md)
    raw_live_comps = raw.get("live_comps") or raw.get("public_comps") or []
    cli_live_comps = [str(item) for item in (live_comps or []) if str(item).strip()]
    yaml_live_comps = _ticker_list_from_raw(raw_live_comps)
    requested_live_comps = cli_live_comps or yaml_live_comps or _default_live_comps_for_facts(facts)
    facts = _apply_live_comps_to_facts(
        facts,
        requested_live_comps,
        timeout=live_comps_timeout,
        provided_comps=_provided_comps_from_raw_mapping(raw),
    )
    wb = spb.build_source_plan_workbook_from_facts(facts)

    if mode == "cap_table":
        ctb.build_cap_table_for_workbook(
            wb,
            _cap_table_input_from_facts(facts),
            exit_scenarios_money_m=[
                _money_m_from_yen(
                    _first_number(getattr(facts, "exit_value_yen", None), 0.0),
                    6_000.0,
                ),
                _money_m_from_yen(
                    _first_number(getattr(facts, "exit_value_yen", None), 0.0) * 1.5,
                    12_000.0,
                ),
            ],
        )

    # Bundle filter (drop sheets outside selected mode).
    for ws_name in list(wb.sheetnames):
        if ws_name not in bundle:
            wb.remove(wb[ws_name])
    _neutralize_removed_sheet_references(wb)

    # Sanity: at least 1 sheet must remain (openpyxl requires it).
    if not wb.sheetnames:
        raise RuntimeError(
            f"Empty workbook after bundle filter (mode={mode}). "
            f"Check additional_sheets/excluded_sheets overrides."
        )

    # Font enforcement (Arial 10pt = Google Sheets default + IB std)
    # Two-stage:
    #   (a) Sweep existing cells whose font fell through to openpyxl-default
    #       Calibri 11 → Arial 10 (preserves explicit non-default fonts).
    #   (b) Override the persistent xlsx default (font index 0 + Normal style)
    #       so that rows added by the user AFTER save also inherit Arial 10.
    ibf.normalize_workbook_fonts(wb)
    ibf.set_workbook_default_font(wb)
    _clear_defined_names(wb)
    spb._apply_print(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================================
# CLI
# ============================================================================


def _main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        prog="build_model.py",
        description=(
            "Builds startup financial-modeling xlsx outputs. Use --source-md "
            "for generic economic-kernel plans; use --mode for narrow "
            "focused finance modules."
        ),
    )
    ap.add_argument(
        "--input", "-i", type=Path, default=None,
        help="YAML input for focused mode output (omit for defaults).",
    )
    ap.add_argument(
        "--output", "-o", type=Path, default=Path("model_output.xlsx"),
        help="Output xlsx path (default: model_output.xlsx).",
    )
    ap.add_argument(
        "--source-md", type=Path, default=None,
        help=(
            "Narrative source markdown for a generic economic-kernel startup "
            "financial plan. Use for equity-story-to-model requests."
        ),
    )
    ap.add_argument(
        "--mode", "-m", choices=VALID_MODES, default="full",
        help="Output mode (default: full).",
    )
    ap.add_argument(
        "--additional-sheets", nargs="*", default=[], metavar="SHEET",
        help="Sheet names to add to the mode bundle (e.g. Market Support).",
    )
    ap.add_argument(
        "--excluded-sheets", nargs="*", default=[], metavar="SHEET",
        help="Sheet IDs to drop from the mode bundle.",
    )
    ap.add_argument(
        "--strict-audit", action="store_true",
        help=(
            "Reopen the generated workbook and fail on broken references, "
            "missing sheet-quality markers, or IB workbook design violations."
        ),
    )
    ap.add_argument(
        "--live-comps", nargs="*", default=[], metavar="TICKER",
        help="Override the auto-selected public ticker peers; private/transaction comps come from YAML evidence fields.",
    )
    ap.add_argument(
        "--live-comps-timeout", type=float, default=8.0,
        help="Timeout in seconds per live public-market comparable request.",
    )
    args = ap.parse_args(argv)

    output = build_model(
        input_path=args.input,
        output_path=args.output,
        mode=args.mode,
        additional_sheets=args.additional_sheets,
        excluded_sheets=args.excluded_sheets,
        source_md=args.source_md,
        live_comps=args.live_comps,
        live_comps_timeout=args.live_comps_timeout,
    )
    bundle = resolve_bundle(args.mode, args.additional_sheets, args.excluded_sheets)
    if args.strict_audit:
        from openpyxl import load_workbook

        audit_issues = audit_workbook(load_workbook(output, data_only=False))
        # Structural audit cannot see broken economics; the economic-coherence
        # audit replays the kernel projection and is profile/mode independent.
        audit_facts, _ = _facts_for_inputs(args.input, args.source_md)
        audit_issues += [
            f"economic: {issue}" for issue in ek.audit_economic_coherence(audit_facts)
        ]
        if audit_issues:
            for issue in audit_issues:
                print(f"[audit] {issue}", file=sys.stderr)
            return 2
        # Task 3.4 (H): non-blocking transparency notes — K8 clamps and K6
        # revenue-provenance flags are printed, never fail the build.
        for warning in ek.audit_economic_warnings(audit_facts):
            print(f"[warn] {warning}", file=sys.stderr)
    print(f"[ok] xlsx generated: {output}")
    print(f"     mode = {args.mode!r}, sheets = {len(bundle)} ({', '.join(bundle)})")
    return 0


if __name__ == "__main__":
    sys.exit(_main())
