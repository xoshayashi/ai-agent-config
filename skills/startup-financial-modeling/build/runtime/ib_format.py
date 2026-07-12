"""
ib_format.py — IB Design Language を openpyxl で適用するヘルパー集

Source of truth: build/references/_ib_workbook_design_system.md
Related references:
  - build/references/_layout_canonical.md
  - build/references/_terminology.md

Usage:
    from ib_format import apply_hard_input, apply_formula, IB_PALETTE, ...
    apply_hard_input(ws['B5'])
    write_section_header(ws, 'A1', 'Revenue Build')

License: internal (startup-financial-modeling skill)
"""

from __future__ import annotations

import math
from copy import copy
from dataclasses import dataclass
from typing import Iterable, Literal

from openpyxl.cell.cell import Cell
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================================
# 1. Functional Color Palette  (canonical = _terminology.md §1)
# ============================================================================

IB_HARD_INPUT = "0000FF"      # Blue — ユーザー入力セル
IB_FORMULA = "000000"         # Black — 数式セル
IB_LINK_INTRA = "008000"      # Green — シート間参照
IB_LINK_EXTERNAL = "FF0000"   # Red — 外部 file 参照
IB_COMMENT = "666666"         # Gray — 注記 (italic)
IB_INK = "2D332E"             # Body ink

# Convenience dict for routing
IB_PALETTE: dict[str, str] = {
    "hard_input": IB_HARD_INPUT,
    "formula": IB_FORMULA,
    "link_intra": IB_LINK_INTRA,
    "link_external": IB_LINK_EXTERNAL,
    "comment": IB_COMMENT,
    "ink": IB_INK,
}

# ============================================================================
# 2. Brand Colors  (canonical = _terminology.md §2)
# ============================================================================

BRAND_PRIMARY_DEEP = "1F3A66"   # Primary deep — navy hierarchy
BRAND_PRIMARY = "2F75B5"        # Primary — blue model accent
BRAND_INK = "2D332E"            # Cover タイトル
BRAND_SURFACE = "F7FAFE"        # Cover 背景
BRAND_ACCENT = "ECC85A"         # Memo タブ色 (§2.X canonical) — 1 sheet 1 か所まで

# Status colors (DESIGN.md §Colors)
BRAND_WARNING = "5B9BD5"        # Secondary blue — Check tab color
BRAND_DANGER = "C04A4A"         # 既存 waterfall_neg と同色
BRAND_SUCCESS = "3F8F5E"

# Bank-specific (optional override)
BRAND_NAVY = "1F3A66"            # Cover タブ色 (§2.X canonical)
BRAND_GS_NAVY = "7399C6"
BRAND_MS_BLUE = "015DAA"
BRAND_LAZARD_ORANGE = "FF6E00"
BRAND_EVERCORE_NAVY = "001F3F"

# Output role color (slate, neutral)
BRAND_SLATE = "7F8FA6"           # Muted blue-gray output tab

# ============================================================================
# 3. Backgrounds
# ============================================================================

BG_WHITE = "FFFFFF"
BG_CANVAS = "F7FAFE"            # Workbook canvas — near-white blue surface
BG_TABLE_HEADER = "D9EAF7"      # Header / label row band — light blue
BG_TOTAL_BAND = "EAF2F8"        # 合計行の薄ブルーバンディング
BG_HEADER_BAND = "1F3A66"       # Section header の濃ネイビーバンド
BG_WORKING = "FFF9C4"           # WIP / TODO セル (yellow highlight)

# Hyperlink (Excel default Office Theme link color = #0563C1)
LINK_COLOR = "0563C1"           # internal/external hyperlink (matches Excel default)

# ----------------------------------------------------------------------------
# Chart palettes for calm, IB-style workbook charts.
# ----------------------------------------------------------------------------

# Football field (Low / Mid / High) — navy / teal / accent
IB_CHART_COLORS_FOOTBALL: list[str] = ["1F3A66", "008A80", "ECC85A"]

# Bar chart default palette — navy / gray / accent
IB_CHART_COLORS_BAR: list[str] = ["1F3A66", "666666", "ECC85A"]

# Line chart default palette — teal / navy / accent / gray
IB_CHART_COLORS_LINE: list[str] = ["008A80", "1F3A66", "ECC85A", "666666"]

# Waterfall chart specific colors
IB_CHART_COLORS_WATERFALL_POS: str = "008A80"   # teal — positive contribution
IB_CHART_COLORS_WATERFALL_NEG: str = "C04A4A"   # danger — negative contribution
IB_CHART_COLORS_WATERFALL_TOTAL: str = "666666" # gray — totals/subtotals

# ============================================================================
# 4. Fonts  (Arial 10pt 基準 — Google Sheets default + IB de facto standard)
# ============================================================================
# Rule:
#   - 全 cell font は Arial 10pt 基準。Google Sheets の default font は Arial 10pt
#     (Office Excel default の Calibri 11pt と混同しないこと)。`_layout_canonical
#     §3.2` および `00_design_guidelines §2.2` で Arial 10pt を canonical 確定。
#   - body / data           = 10pt
#   - section header        = 10-11pt bold (`01a §5.1`)
#   - sheet title (B1)      = 14pt bold
#   - period header         = 10pt bold italic
#   - comment / footnote    = 9pt italic gray (#666666)
#   - populated worksheet cells stay within 9/10/11/14pt. Do not use 8pt
#     cells to squeeze content or 16pt+ presentation headings inside model
#     grids; fix layout, width, or copy length instead.
#   - 日本語混在は OS fallback に委譲 (Mac=ヒラギノ、Windows=Yu Gothic)、
#     cell font 指定は Arial のまま (font_family を変えると Latin face が崩れる)。
# Chart 内部 font (axis tick / title) は本 file の責務外。`apply_chart_palette`
# は色のみ操作するため影響なし。

FONT_FAMILY = "Arial"
FONT_SIZE_BASE = 10        # 本文セルの基準 (Google Sheets default + IB standard)
FONT_SIZE_SMALL = 9        # comment / footnote / unit label
FONT_SIZE_LARGE = 11       # section header
FONT_SIZE_TITLE = 14       # cover title
FONT_SIZE_TINY = 8         # chart axis tick (charts only — cells should not use this)
FONT_SIZE_ALLOWED_CELLS: tuple[int, ...] = (
    FONT_SIZE_SMALL,
    FONT_SIZE_BASE,
    FONT_SIZE_LARGE,
    FONT_SIZE_TITLE,
)
FONT_SIZE_BEST_PRACTICE = (
    "IB model cells use a constrained font-size palette: 9pt supporting "
    "notes/sources/unit helpers, 10pt body/model cells, 11pt compact section "
    "emphasis, and 14pt sheet titles only. Avoid 8pt worksheet cells, 16pt+ "
    "presentation headings, and improvised local sizes; repair layout instead "
    "of shrinking text."
)

FONT_BODY = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_INK)
FONT_BODY_BOLD = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)

WRAP_TEXT_ERROR = (
    "startup-financial-modeling forbids wrap_text=True for generated workbook "
    "cells. Classify the cell role first: titles, instructions, notes, bullets, "
    "and source caveats with empty cells to the right should read horizontally "
    "through blank unstyled overflow cells without merging. Use wider columns, "
    "dedicated note columns, shorter rows, or blank unstyled overflow cells "
    "instead. If a user-approved bounded table prose exception uses wrapping "
    "or manual line breaks, set row height to the exact rendered visible line "
    "count."
)
WRAP_DECISION_LADDER: tuple[str, ...] = (
    "shorten_or_split_copy",
    "widen_role_column",
    "reserve_blank_unstyled_overflow_cells",
    "move_to_note_or_interpretation_surface",
    "use_user_approved_bounded_prose_wrap_only",
)
WRAP_BEST_PRACTICE = (
    "IB model cells keep wrap_text off by default. Horizontal-read titles, "
    "instructions, notes, bullets, source caveats, and memo lines should read "
    "through blank unstyled overflow cells without merging and must not be "
    "trapped at the final print/render column. Wrapping or manual line breaks "
    "are reserved for user-approved bounded table prose and require exact "
    "rendered visible line-count row height."
)

ALIGN_LABEL = "left"
ALIGN_SOURCE_NOTE = "left"
ALIGN_VALUE = "right"
ALIGN_UNIT = "right"
ALIGN_PERIOD_HEADER = "center"
ALIGN_VERTICAL_BODY = "center"

ALIGNMENT_BEST_PRACTICE = (
    "IB workbook alignment is role-based: labels, sources, notes, titles, "
    "memos, and interpretation text are left-aligned; numeric values, "
    "formulas, money, percentages, multiples, counts, and unit labels are "
    "right-aligned; only period headers, short scenario/matrix headers, and "
    "compact column headers are centered. Hierarchy uses dedicated columns, "
    "not native Excel indent or leading spaces."
)


def _ensure_no_wrap_text(wrap_text: bool) -> None:
    if wrap_text:
        raise ValueError(WRAP_TEXT_ERROR)


def set_workbook_default_font(wb, name: str = FONT_FAMILY, size: int = FONT_SIZE_BASE) -> None:
    """Make `name` `size` the workbook-wide default font (persistent in xlsx).

    Critical: openpyxl initialises a Workbook with **font index 0 = Calibri 11**
    (Office Excel default). xlsx readers (Excel / Google Sheets / LibreOffice)
    use font index 0 as the **fallback for cells without an explicit style** —
    AND for newly-added cells when a user opens the saved file and inserts rows.

    This helper overrides BOTH paths:
      1. `wb._fonts[0]` — raw font slot referenced by cellXf id=0 (covers new
         rows added by the user post-save).
      2. `wb._named_styles["Normal"].font` — covers consumers that read the
         Named Style instead of the raw font index.

    Call timing is safe at any point before save; openpyxl's IndexedList
    accepts in-place assignment without breaking existing cellXf refs that
    point to index 0 (those cells visually shift to the new default — which is
    what we want for un-styled cells).

    Note: openpyxl does not expose a public workbook-default-font setter, so
    the private `_fonts` / `_named_styles` access is intentionally contained in
    this best-effort helper.
    """
    new_font = Font(name=name, size=size, color=IB_INK)
    # 1. Raw font index 0 (the default the xlsx file persists for new rows).
    try:
        wb._fonts[0] = new_font
    except (AttributeError, IndexError, TypeError):
        pass  # If the private font slot is unavailable, use Normal style only.
    # 2. Named "Normal" style (the path Excel's named-style consumers use).
    try:
        wb._named_styles["Normal"].font = Font(name=name, size=size, color=IB_INK)
    except (AttributeError, KeyError):
        pass


def normalize_workbook_fonts(
    wb,
    target_name: str = FONT_FAMILY,
    target_size: int = FONT_SIZE_BASE,
) -> int:
    """Sweep all cells and convert openpyxl-default (Calibri 11) → Arial 10.

    Run AFTER all builders have populated the workbook. Cells that already
    have an explicit non-Calibri font (e.g., Arial 14 for titles) are
    preserved — only Calibri / unset font names are rewritten. Size is
    preserved unless it's the openpyxl-default 11pt.

    Returns the number of cells touched.
    """
    touched = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                f = cell.font
                if f is None:
                    continue
                # Preserve explicit non-Calibri fonts (e.g., Arial set by
                # builders' apply_*() helpers). Only normalize cells that
                # fell through to openpyxl's library default (Calibri).
                if f.name not in (None, "Calibri"):
                    continue
                # Preserve non-default sizes (e.g., 14pt titles set elsewhere).
                new_size = target_size if f.size in (None, 11.0) else f.size
                cell.font = Font(
                    name=target_name,
                    size=new_size,
                    bold=f.bold,
                    italic=f.italic,
                    underline=f.underline,
                    strike=f.strike,
                    color=f.color,
                )
                touched += 1
    return touched


def last_value_bounds(ws: Worksheet) -> tuple[int, int]:
    """Return the last row/column that contains a real displayed value."""
    last_row = 1
    last_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            last_row = max(last_row, cell.row)
            last_col = max(last_col, cell.column)
    return last_row, last_col


def _chart_anchor_start(anchor: object) -> tuple[int, int] | None:
    if isinstance(anchor, str):
        try:
            return coordinate_to_tuple(anchor)
        except ValueError:
            return None
    marker = getattr(anchor, "_from", None)
    if marker is not None:
        return marker.row + 1, marker.col + 1
    return None


def rendered_bounds(ws: Worksheet) -> tuple[int, int]:
    """Return the last row/column needed for values and rendered drawings."""
    last_row, last_col = last_value_bounds(ws)
    for chart in getattr(ws, "_charts", []):
        anchor = getattr(chart, "anchor", None)
        start = _chart_anchor_start(anchor)
        if start is None:
            continue
        row, col = start
        ext = getattr(anchor, "ext", None)
        width = (float(ext.cx) / 360000) if ext and getattr(ext, "cx", None) else float(getattr(chart, "width", 15) or 15)
        height = (float(ext.cy) / 360000) if ext and getattr(ext, "cy", None) else float(getattr(chart, "height", 7.5) or 7.5)
        # Chart extents are reported in centimeters; estimate the occupied grid
        # with the workbook's visual scale (~1.8 cm/column, ~0.55 cm/row).
        col_span = max(4, int(round(width / 1.8)) + 1)
        row_span = max(10, int(round(height / 0.55)) + 1)
        last_row = max(last_row, row + row_span)
        last_col = max(last_col, col + col_span)
    return last_row, last_col


SEMANTIC_BLANK_FILL_COLORS = {
    BG_TABLE_HEADER,
    BG_TOTAL_BAND,
    BG_HEADER_BAND,
    BG_WORKING,
}


def apply_semantic_fill_span(
    ws: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    color: str,
    *,
    top: Side | None = None,
    bottom: Side | None = None,
    border_start_col: int | None = None,
) -> None:
    """Style one rectangular semantic row span, including blank member cells.

    Use this for section bands, header/label rows, selected outputs, checks, and
    caution rows. The caller chooses `end_col` from the attached table/block,
    not from whether each cell currently has text. If `top` or `bottom` is
    supplied, the border is applied across the full row-component span from
    `border_start_col` (or `start_col` by default) so blank member cells align
    with the related table/block instead of producing a ragged rule. Keep
    dedicated hierarchy / indent columns borderless by setting
    `border_start_col` to the row's real label/data start column.
    """
    if end_col < start_col:
        raise ValueError(
            f"semantic fill span end_col ({end_col}) must be >= start_col ({start_col})"
        )
    border_first_col = border_start_col if border_start_col is not None else start_col
    if border_first_col < start_col or border_first_col > end_col:
        raise ValueError(
            f"border_start_col ({border_first_col}) must be within span {start_col}:{end_col}"
        )
    fill = PatternFill("solid", fgColor=color)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        if col >= border_first_col and (top is not None or bottom is not None):
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=top if top is not None else cell.border.top,
                bottom=bottom if bottom is not None else cell.border.bottom,
                diagonal=cell.border.diagonal,
                diagonal_direction=cell.border.diagonal_direction,
                diagonalUp=cell.border.diagonalUp,
                diagonalDown=cell.border.diagonalDown,
                outline=cell.border.outline,
                vertical=cell.border.vertical,
                horizontal=cell.border.horizontal,
            )


def _fill_rgb(cell: Cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    value = getattr(fill.fgColor, "rgb", None)
    return value[-6:].upper() if isinstance(value, str) else None


def _row_has_semantic_header_fill(ws: Worksheet, row: int) -> bool:
    for cell in ws[row]:
        if _fill_rgb(cell) == BG_TABLE_HEADER:
            return True
    return False


def _find_table_header_row(ws: Worksheet, row: int) -> int | None:
    """Return the nearest row at or above ``row`` whose cells carry BG_TABLE_HEADER.

    Nearest-above (not topmost) so a nested inner table picks its own
    header instead of the outer top-of-sheet table — surfaced by the
    Phase 1 mid-checkpoint review of this helper.
    """
    for r in range(row, 0, -1):
        if _row_has_semantic_header_fill(ws, r):
            return r
    return None


def _semantic_header_bounds(
    ws: Worksheet,
    row: int,
    *,
    target_row: int | None = None,
) -> tuple[int, int] | None:
    cols = [c.column for c in ws[row] if _fill_rgb(c) == BG_TABLE_HEADER]
    if not cols:
        return None
    if target_row is not None:
        target_bounds = _row_content_bounds(ws, target_row)
        if target_bounds is not None and target_bounds[0] != target_bounds[1]:
            target_start, target_end = target_bounds
            # If target content starts on col 3 (C), the visual hierarchy indent gutter (B, col 2)
            # is a logical block participant. Allow col 2 to be swept into in_target.
            if target_start == 3:
                target_start = 2
            in_target = [col for col in cols if target_start <= col <= target_end]
            if in_target:
                return min(in_target), max(in_target)
    return min(cols), max(cols)


def _row_content_bounds(ws: Worksheet, row: int) -> tuple[int, int] | None:
    cols = [c.column for c in ws[row] if c.value not in (None, "")]
    if not cols:
        return None
    return min(cols), max(cols)


def apply_semantic_border_span(
    ws: Worksheet,
    row: int,
    *,
    top: Side | None = None,
    bottom: Side | None = None,
    header_row: int | None = None,
    border_start_col: int | None = None,
) -> None:
    """Apply a top / bottom border across the table block ``row`` belongs to.

    Symmetric to :func:`apply_semantic_fill_span`. The span is resolved from
    :func:`detect_table_block` — the helper deliberately exposes no
    ``start_col`` / ``end_col`` parameters so callers cannot reintroduce the
    hard-coded column ranges that the
    ``accent_cols = [label_col, *period_cols]`` path in
    ``source_plan_builder._write_values`` produced. Block-internal empty /
    middle cells (source / unit columns between the label column and the
    first period) receive the same border so the underline reads as a
    continuous rule — the structural fix for the
    ``if cell.value is None: continue`` gap.

    To keep dedicated hierarchy / indent columns borderless when they fall
    inside the detected block (the canonical practice — the indent column is
    a visual gutter, not a table member with a rule), pass
    ``border_start_col`` to the real label / data start. Callers typically
    derive this from the row's role (text position / fill), not from a
    column number.

    A call with neither ``top`` nor ``bottom`` is a no-op.

    Args:
        ws: openpyxl worksheet.
        row: 1-indexed row to draw the border on.
        top: top-edge ``Side`` (e.g. :data:`THIN_LINE` for subtotals).
        bottom: bottom-edge ``Side`` (e.g. :data:`MEDIUM_LINE` for grand totals).
        header_row: optional override forwarded to ``detect_table_block``.
        border_start_col: leftmost column inside the block that should
            receive the border. Columns from the detected block start up
            to this column are left untouched.
    """
    if top is None and bottom is None:
        return
    start_col, end_col = detect_table_block(ws, row, header_row=header_row)
    border_first = border_start_col if border_start_col is not None else start_col
    if border_first < start_col:
        border_first = start_col
    for col in range(border_first, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=top if top is not None else cell.border.top,
            bottom=bottom if bottom is not None else cell.border.bottom,
            diagonal=cell.border.diagonal,
            diagonal_direction=cell.border.diagonal_direction,
            diagonalUp=cell.border.diagonalUp,
            diagonalDown=cell.border.diagonalDown,
            outline=cell.border.outline,
            vertical=cell.border.vertical,
            horizontal=cell.border.horizontal,
        )


def detect_table_block(
    ws: Worksheet,
    row: int,
    *,
    header_row: int | None = None,
) -> tuple[int, int]:
    """Return ``(start_col, end_col)`` of the table block ``row`` belongs to.

    The block boundary is derived from the worksheet's actual cell content
    and styling — never from a hard-coded column number — so the helper is
    safe on sheets that diverge from the canonical LAYOUT. Detection
    signals, applied in order:

    1. **Semantic header fill.** The nearest row at or above ``row`` whose
       cells carry the ``BG_TABLE_HEADER`` fill defines the block. The
       block's column range is the min/max column of the filled cells that
       overlap the target row's own populated span. That overlap rule keeps
       unrelated side-by-side tables on the same header row out of the block.
       Empty middle cells (e.g. source / unit columns that sit between the
       label column and the first period column) are inside the range —
       they are block members, which lets ``apply_semantic_border_span``
       paint a continuous underline across the whole block.
    2. **Row content fallback.** If no semantic header is found anywhere
       at or above ``row``, fall back to ``row`` itself: leftmost to
       rightmost non-empty column.

    An annotation column to the right of the period block (e.g. a wide
    "Notes" column with plain text but no header fill) is correctly
    excluded — the BG_TABLE_HEADER footprint ends before it.

    Args:
        ws: openpyxl worksheet.
        row: 1-indexed target row.
        header_row: optional override; force detection from this row
            instead of auto-scanning. Useful when nested sub-tables make
            the auto-scan pick the wrong header.

    Returns:
        ``(start_col, end_col)``, both 1-indexed inclusive.

    Raises:
        ValueError: when no block can be detected (no header fill found
            above and the target row itself is empty).
    """
    detected_header = (
        header_row if header_row is not None else _find_table_header_row(ws, row)
    )
    if detected_header is not None:
        bounds = _semantic_header_bounds(
            ws,
            detected_header,
            target_row=None if header_row is not None else row,
        )
        if bounds is not None:
            return bounds
    fallback = _row_content_bounds(ws, row)
    if fallback is not None:
        return fallback
    raise ValueError(
        f"row {row} on '{ws.title}': no BG_TABLE_HEADER cell found at or above, "
        "and the target row itself has no content; cannot detect table block."
    )


def _is_intentional_blank_component_cell(cell: Cell, row_has_value: bool) -> bool:
    """Keep blank cells that visually complete a semantic row component.

    Empty cells are normally reset so overflow text and trailing canvas stay
    clean. A blank cell with a semantic row fill is different: it can be the
    quiet continuation of a header/check/highlight band, aligned to the same
    visual width as neighboring rows.

    A blank cell that carries a deliberate border (any side, any style) is
    also a semantic row component: it is the empty middle of a table-block
    span painted by ``apply_semantic_border_span`` so the underline reads
    as continuous across the source / unit gap. Wiping its border to the
    default style would re-introduce the ragged-rule defect that span is
    meant to fix.
    """
    if not row_has_value:
        return False
    if _fill_rgb(cell) in SEMANTIC_BLANK_FILL_COLORS:
        return True
    border = cell.border
    if border is None:
        return False
    return any(
        side is not None and getattr(side, "style", None) is not None
        for side in (border.top, border.bottom, border.left, border.right)
    )


def clear_blank_cell_styles(wb: Workbook) -> None:
    """Reset blank cells to the workbook default style.

    This keeps overflow spacer cells and trailing canvas clean. It preserves
    blank cells that intentionally extend semantic row components, such as
    header/label-row or selected-output fills aligned to the same column span as
    neighboring rows.
    """
    default_style = copy(wb._cell_styles[0])
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            row_has_value = any(cell.value not in (None, "") for cell in row)
            for cell in row:
                if cell.value not in (None, ""):
                    continue
                if _is_intentional_blank_component_cell(cell, row_has_value):
                    continue
                cell.value = None
                cell._style = copy(default_style)


def trim_blank_canvas(wb: Workbook, *, set_print_area: bool = True) -> None:
    """Trim every sheet to its real used range and clear non-semantic blanks."""
    default_style = copy(wb._cell_styles[0])
    for ws in wb.worksheets:
        last_row, last_col = rendered_bounds(ws)
        if ws.max_row > last_row:
            ws.delete_rows(last_row + 1, ws.max_row - last_row)
        if ws.max_column > last_col:
            ws.delete_cols(last_col + 1, ws.max_column - last_col)
        for key in list(ws.column_dimensions):
            try:
                col_idx = coordinate_to_tuple(f"{key}1")[1]
            except ValueError:
                continue
            if col_idx > last_col:
                del ws.column_dimensions[key]
        for row in ws.iter_rows():
            row_has_value = any(cell.value not in (None, "") for cell in row)
            for cell in row:
                if cell.value in (None, ""):
                    if _is_intentional_blank_component_cell(cell, row_has_value):
                        continue
                    cell.value = None
                    cell._style = copy(default_style)
        if set_print_area:
            ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

FONT_HARD_INPUT = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_HARD_INPUT)
FONT_FORMULA = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_FORMULA)
FONT_LINK_INTRA = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_LINK_INTRA)
FONT_LINK_EXTERNAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=IB_LINK_EXTERNAL)
FONT_COMMENT = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
FONT_FOOTNOTE = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
FONT_SECTION_HEADER = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color=IB_INK)
FONT_TITLE = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, bold=True, color=IB_INK)
FONT_YEAR_HEADER = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_INK)
FONT_SUBTOTAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_FORMULA)
FONT_GRAND_TOTAL = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=True, color=IB_FORMULA)
FONT_COVER_TITLE = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, bold=True, color="FFFFFF")

# Public semantic aliases used by builder modules.
FONT_SECTION = FONT_SECTION_HEADER
FONT_SUBSECTION = FONT_BODY_BOLD
FONT_TOTAL = FONT_SUBTOTAL

# ============================================================================
# 5. Number Formats
# ============================================================================
# User spec rules:
#   - 基本単位 = 円 (cell value は必ず円単位の生数値で保存)
#   - 千円 / 百万円 / 億円は number_format で表現 (cell value は変えない)
#   - 負数は赤文字  ([Red] section)
#   - 通貨記号 (¥, $) を表示形式に含める
#   - 参照 workbook の単位判定も visible text ではなく number_format を読む
#
# Excel section syntax: positive;negative;zero[;text]
#   - "_)" は positive 後の右余白 (右括弧分の幅を空けて、negative の括弧囲みと
#     右端揃えになるようにする — IB convention)
#   - 数値の "," は 3 桁区切り、フォーマット末尾の "," は表示倍率 ÷1,000
#     (二重の ",," は ÷1,000,000 = 百万円表示)
# ----------------------------------------------------------------------------

# --- JPY (¥) -----------------------------------------------------------------
FMT_JPY_YEN              = '¥#,##0_);[Red](¥#,##0);"-"_)'
FMT_JPY_THOUSAND         = '¥#,##0,_);[Red](¥#,##0,);"-"_)'      # 千円表示
FMT_JPY_MILLION          = '¥#,##0,,_);[Red](¥#,##0,,);"-"_)'    # 百万円表示
FMT_JPY_BILLION          = '¥#,##0,,,_)"十億円";[Red](¥#,##0,,,)"十億円";"-"_)'
FMT_JPY_TRILLION         = '¥#,##0,,,,_)"兆円";[Red](¥#,##0,,,,)"兆円";"-"_)'
# 億円: ",," = ÷1M (百万単位の表示) + "億" 接尾。厳密な ÷1億 を Excel の数値
# format で実現する手段は存在しない (comma scale は 3 桁単位)。よって
# 「百万円のまま億接尾」となる近似である点に caller は注意。
# 用途: ざっくり大まかな表示用 (例: "¥1,234,億" → 実数は 1,234M = 1,234百万円)
FMT_JPY_HUNDRED_MILLION  = '¥#,##0,,_)"億";[Red](¥#,##0,,)"億";"-"_)'

# --- USD ($) -----------------------------------------------------------------
FMT_USD_DOLLAR           = '"$"#,##0_);[Red]("$"#,##0);"-"_)'
FMT_USD_THOUSAND         = '"$"#,##0,_);[Red]("$"#,##0,);"-"_)'
FMT_USD_MILLION          = '"$"#,##0,,_);[Red]("$"#,##0,,);"-"_)'

# --- 通貨記号なし (純粋数値、3 桁区切り、負数 赤) ---------------------------
FMT_NUM                  = '#,##0_);[Red](#,##0);"-"_)'
FMT_NUM_THOUSAND         = '#,##0,_);[Red](#,##0,);"-"_)'
FMT_NUM_MILLION          = '#,##0,,_);[Red](#,##0,,);"-"_)'

# --- 比率 / パーセント (負数 赤) ---------------------------------------------
FMT_PCT_0                = '0%_);[Red](0%);"-"_)'
FMT_PCT_1                = '0.0%_);[Red](0.0%);"-"_)'
FMT_PCT_2                = '0.00%_);[Red](0.00%);"-"_)'

# --- 倍数 (x 表示、負数 赤) --------------------------------------------------
FMT_MULTIPLE_1           = '0.0"x"_);[Red](0.0"x");"-"_)'
FMT_MULTIPLE_2           = '0.00"x"_);[Red](0.00"x");"-"_)'

# --- Date --------------------------------------------------------------------
FMT_DATE_YMD             = "yyyy-mm-dd"
FMT_DATE_YM              = "yyyy-mm"

# --- 整数 (株式数等、3 桁区切り、負数 赤) ------------------------------------
FMT_INTEGER              = '#,##0_);[Red](#,##0);"-"_)'

# --- 株価 (小数 2 位) --------------------------------------------------------
FMT_PER_SHARE            = '#,##0.00_);[Red](#,##0.00);"-"_)'

# Public semantic aliases. JPY × million is the default model display scale;
# USD/EUR/other models must pass an explicit currency-specific format.
FMT_MONEY                = FMT_JPY_MILLION
FMT_MONEY_DECIMAL        = '¥#,##0.0,,_);[Red](¥#,##0.0,,);"-"_)'  # 百万円 1 桁
FMT_PERCENT              = FMT_PCT_1
FMT_PERCENT_BPS          = FMT_PCT_2
FMT_MULTIPLE             = FMT_MULTIPLE_1
FMT_SHARES               = FMT_INTEGER
FMT_DATE_SHORT           = "mmm-yy"
FMT_DATE_LONG            = FMT_DATE_YMD
FMT_DEFAULT              = FMT_NUM_MILLION

# ============================================================================
# 6. Layout Constants  (Web Design 由来 token system: T-shirt size + modular)
# ============================================================================
# Spacing scale: Material Design-style T-shirt sizing applied to Excel row/col.
# Ratio ≒ 1.25 (Major Third) on row heights — 15 / 18 / 22 / 32 ≈ 1.0 / 1.2 /
# 1.47 / 2.13 (relaxed). column widths follow {tiny, small, base, large, xl}
# ramp consistent with B(parent/indent)≈20px in Google Sheets / C(label)=54 / D(source)=54 /
# E(unit)=12 / F-(period)=15 baseline.

# CJK ranges used for display-width measurement (autosize_role_columns).
# These ranges cover Hiragana, Katakana, CJK Unified Ideographs (incl. Ext A),
# and CJK / fullwidth punctuation. Characters in any of these count as 2
# display units so a column carrying Japanese labels resolves to a width
# that actually fits the text in Google Sheets render.
_CJK_RANGES: tuple[tuple[int, int], ...] = (
    (0x3000, 0x303F),   # CJK Symbols and Punctuation
    (0x3040, 0x309F),   # Hiragana
    (0x30A0, 0x30FF),   # Katakana
    (0x3400, 0x4DBF),   # CJK Unified Ideographs Extension A
    (0x4E00, 0x9FFF),   # CJK Unified Ideographs
    (0xFF00, 0xFFEF),   # Halfwidth and Fullwidth Forms
)


def _is_cjk_char(ch: str) -> bool:
    code = ord(ch)
    for lo, hi in _CJK_RANGES:
        if lo <= code <= hi:
            return True
    return False


def _display_width(value: object) -> int:
    """Return the visual display width of ``value`` with CJK chars at 2.

    Used by :func:`autosize_role_columns` so a Japanese label that needs 56
    display units does not get squeezed into a 28-cell-wide column.
    """
    if value is None:
        return 0
    text = str(value)
    width = 0
    for ch in text:
        width += 2 if _is_cjk_char(ch) else 1
    return width


# Column widths
COL_MARGIN_WIDTH = 3.0          # A: visual gutter only
# Hierarchy / indent column width — single source of truth.
# Every column that plays the role of indent / hierarchy spacer (starting at B
# and any deeper hierarchy column added to the right) MUST use this exact
# xlsx width. It renders as 20px in Google Sheets — the canonical indent
# rhythm called out in build/references/_layout_canonical.md. No sheet may
# widen B for label/text purposes: text overflows from B into intentionally
# blank cells to the right instead, or shifts to a wider C+ column.
COL_HIERARCHY_WIDTH = 2.14      # B and additional hierarchy columns; exactly 2.14 xlsx ≈ 20px Google Sheets
INDENT_COL_WIDTH = COL_HIERARCHY_WIDTH  # semantic alias used by indent-role helpers
COL_LABEL_WIDTH = 54.0          # Lowest-level line-item label
COL_SOURCE_WIDTH = 54.0         # Source / driver
COL_UNIT_WIDTH = 14.0           # Unit (¥M / %)
COL_PERIOD_WIDTH = 16.0         # Each period (Y1 / Q1 / Jan-26)
COL_WIDTH_TOL = 0.001           # comparison tolerance for indent / role widths

# Column width T-shirt scale (auxiliary; for non-canonical layouts)
COL_WIDTH_TINY = 6.0
COL_WIDTH_SMALL = 12.0
COL_WIDTH_BASE = 16.0
COL_WIDTH_LARGE = 20.0          # auxiliary large text/table width; not hierarchy width
COL_WIDTH_XLARGE = 32.0
COL_NOTE_WIDTH = 72.0

# Role-based [min, max] bounds for content-driven column widths
# (autosize_role_columns). The min is the role's starting width — also the
# minimum that survives a workbook with only short content. The max keeps
# pathological inputs (a 500-char source note) from bloating the sheet.
ROLE_WIDTH_BOUNDS: dict[str, tuple[float, float]] = {
    "label":  (COL_LABEL_WIDTH,  90.0),
    "source": (COL_SOURCE_WIDTH, 90.0),
    "unit":   (COL_UNIT_WIDTH,   24.0),
    "period": (COL_PERIOD_WIDTH, 24.0),
    "note":   (COL_NOTE_WIDTH,  120.0),
}
ROLE_WIDTH_PADDING = 2.0  # whitespace breathing room added to measured width

# Row heights used by builder modules.
ROW_BODY_HEIGHT = 15.0
ROW_SPACER_HEIGHT = 8.0
ROW_SECTION_HEIGHT = 22.0

# Row height T-shirt scale (補助 token. 既存 ROW_*_HEIGHT を踏襲)
ROW_HEIGHT_TIGHT = 14.0         # 詰め (sub-section row)
ROW_HEIGHT_BASE = 15.0          # = ROW_BODY_HEIGHT
ROW_HEIGHT_STANDARD = 18.0      # compact title / helper rows
ROW_HEIGHT_MEDIUM = 20.0        # short banner / title rows
ROW_HEIGHT_RELAXED = 22.0       # = ROW_SECTION_HEIGHT
ROW_HEIGHT_HERO = 32.0          # cover title 等 (modular ratio ~2.13x)
ROW_HEIGHT_COVER_TITLE = 60.0   # cover-only display title row
ROW_HEIGHT_PER_WRAPPED_LINE = ROW_HEIGHT_BASE  # user-approved wrap exceptions only

# ============================================================================
# 7. Border styles  (subtotal / grand total / section divider / box / table)
# ============================================================================
# IB Border convention (canonical = Macabacus / WSP / IB pitchbook):
#   - 小計 (subtotal)         : top = thin black single
#   - 合計 (grand total)      : top = thin black single, bottom = medium black
#   - Section 区切り          : bottom = thin black
#   - Section 末 medium       : bottom = medium black (sub-section ⇒ section の境界)
#   - Sensitivity / call-out  : 4 辺 = thin black (box)
#   - Heavy emphasis (base case): 4 辺 = medium black

# Sides — line weight × color (gray は補助、IB convention は黒基本)
THIN_LINE = Side(border_style="thin", color=IB_FORMULA)
MEDIUM_LINE = Side(border_style="medium", color=IB_FORMULA)
DOTTED_LINE = Side(border_style="dotted", color=IB_FORMULA)

BORDER_SUBTOTAL = Border(top=THIN_LINE)                           # 小計: 上 single
BORDER_GRAND_TOTAL = Border(top=THIN_LINE, bottom=MEDIUM_LINE)    # 合計: 上 single + 下 medium
BORDER_SECTION_DIVIDER = Border(bottom=THIN_LINE)                 # Section 区切り
BORDER_SECTION_END_MEDIUM = Border(bottom=MEDIUM_LINE)            # Section 末 (medium)
BORDER_TOP_THIN = Border(top=THIN_LINE)
BORDER_BOTTOM_THIN = Border(bottom=THIN_LINE)
BORDER_BOX_THIN = Border(top=THIN_LINE, bottom=THIN_LINE, left=THIN_LINE, right=THIN_LINE)
BORDER_BOX_MEDIUM = Border(top=MEDIUM_LINE, bottom=MEDIUM_LINE, left=MEDIUM_LINE, right=MEDIUM_LINE)

# ============================================================================
# 8. Page Setup
# ============================================================================

PRINT_MARGIN_INCH = 0.5
PRINT_HEADER_INCH = 0.3
PRINT_ORIENTATION_LANDSCAPE = "landscape"
PRINT_ORIENTATION_PORTRAIT = "portrait"

# ============================================================================
# 9. Sheet Naming
# ============================================================================

CANONICAL_SHEET_ORDER: tuple[str, ...] = (
    "Guide",
    "Kernel",
    "Assumptions",
    "Driver Tree",
    "Revenue Build",
    "Cost Build",
    "People Plan",
    "P&L",
    "BS",
    "CF",
    "Capital Stack",
    "Ownership",
    "Pricing",
    "Financing",
    "Exit Waterfall",
    "Segments",
    "KPI",
    "Scenarios",
    "Sensitivity",
    "Valuation",
    "Market Support",
    "Benchmarks",
    "IC Memo",
)

OPTIONAL_SHEETS: tuple[str, ...] = ("99_Glossary",)

# ============================================================================
# 9. Sheet Naming
# ============================================================================
# The legacy 23-sheet canonical order / role tables (CANONICAL_SHEET_ORDER,
# OPTIONAL_SHEETS, SHEET_ROLE_MAPPING, validate_sheet_naming) were retired
# with the S3 12-sheet architecture. Sheet sets now live in
# source_plan_builder (SOURCE_PLAN_SHEETS_V2 / CONDITIONAL_SHEETS_V2 /
# TAB_COLORS_V2) and build_model (MODE_BUNDLE_SEEDS).

# ============================================================================
# 10. Cell-level helper functions
# ============================================================================


def apply_hard_input(
    cell,
    fmt: str = FMT_JPY_MILLION,
) -> None:
    """Hard input セルに IB blue + 数値書式を適用する。

    Default は `FMT_JPY_MILLION`。USD ベースのモデルでは呼び出し側で
    fmt=FMT_USD_MILLION を明示指定する。

    Args:
        cell: openpyxl Cell
        fmt: number_format
    """
    cell.font = FONT_HARD_INPUT
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def apply_formula(
    cell,
    fmt: str = FMT_JPY_MILLION,
) -> None:
    """Formula セルに black + 数値書式を適用 (wrap_text=False).

    Default は `FMT_JPY_MILLION`。
    """
    cell.font = FONT_FORMULA
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def apply_link_intra(
    cell,
    fmt: str = FMT_JPY_MILLION,
) -> None:
    """Cross-sheet link セル (=03_Revenue!B5 等) に green を適用 (wrap_text=False).

    Default は `FMT_JPY_MILLION`。
    """
    cell.font = FONT_LINK_INTRA
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def apply_link_external(
    cell,
    fmt: str = FMT_JPY_MILLION,
) -> None:
    """External link (別 file 参照) に red を適用 (wrap_text=False).

    Default は `FMT_JPY_MILLION`。
    """
    cell.font = FONT_LINK_EXTERNAL
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def apply_label(cell, bold: bool = False, wrap_text: bool = False) -> None:
    """行ラベル (左寄せ、Excel indent なし、Arial 10pt).

    wrap_text=True は禁止。長文は列幅、専用 note 列、短い行分割、
    または隣接空白セルへの overflow で解決する。
    Font は Arial 10pt 固定 (Google Sheets default + IB de facto standard)。
    階層表現は列追加で行い、Excel native indent は使わない。
    """
    _ensure_no_wrap_text(wrap_text)
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=bold, color=IB_INK)
    cell.alignment = Alignment(
        horizontal=ALIGN_LABEL, vertical=ALIGN_VERTICAL_BODY, indent=0, wrap_text=False
    )


def write_hierarchical_line_item(
    ws: Worksheet,
    row: int,
    level: int,
    label: str,
    value_start_col: int = 4,
    bold: bool = False,
) -> int:
    """階層 line item を書き込む helper.

    親項目 (level=0) → B 列 (col 2)
    子項目 (level=1) → C 列 (col 3)
    孫項目 (level=2) → D 列 (col 4)  (注: 既定 PERIOD_START_COL=4 と衝突するため、
                                       value_start_col を上げる builder 側で対応)

    Args:
        ws: Worksheet
        row: 1-based row number
        level: 0/1/2  階層深さ
        label: ラベル文字列
        value_start_col: 数値が始まる列 (default=4=D)。level=2 を使うときは builder
            側で value_start_col を 1 つ右にずらすこと。
        bold: True で太字 (親項目で使う)

    Returns:
        実際にラベルを書いた列 (col_index)
    """
    label_col = 2 + level  # B(2), C(3), D(4)
    cell = ws.cell(row=row, column=label_col)
    cell.value = label
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, bold=bold, color=IB_INK)
    cell.alignment = Alignment(
        horizontal=ALIGN_LABEL, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )
    return label_col


def set_uniform_column_width(
    ws: Worksheet, columns: Iterable[str], width: float
) -> None:
    """指定列を一律 width に設定.

    Args:
        ws: Worksheet
        columns: 列文字 iterable, e.g. ['C', 'D', 'E', 'F']
        width: openpyxl 列幅 (≈ 文字数)
    """
    for col_letter in columns:
        ws.column_dimensions[col_letter].width = width


def autosize_role_columns(
    ws: Worksheet,
    role_columns: dict[int, str],
    *,
    padding: float | None = None,
) -> None:
    """Set role columns' widths from their content, clamped to role bounds.

    For each ``(column_index, role)`` pair in ``role_columns``, the column's
    width is set to ``max(role_min, min(role_max, content_max + padding))``,
    where ``content_max`` is the maximum display width of any cell value in
    the column (CJK chars counted as 2 via :func:`_display_width`). The
    role bounds come from :data:`ROLE_WIDTH_BOUNDS`.

    The clamp keeps short-content columns from being padded wider than the
    role floor and keeps pathologically long content from bloating the
    sheet beyond the role max. Used after data has been written into the
    worksheet so the measurement reflects the final visible content.

    Args:
        ws: worksheet to size columns on.
        role_columns: ``{column_index: role_name}`` mapping where role
            name must be a key in :data:`ROLE_WIDTH_BOUNDS`. Unknown
            roles are silently skipped.
        padding: visual breathing room added to the measured display
            width before clamping. Defaults to :data:`ROLE_WIDTH_PADDING`.
    """
    pad = ROLE_WIDTH_PADDING if padding is None else padding
    for col, role in role_columns.items():
        width = role_column_width_for_content(ws, col, role, padding=pad)
        if width is None:
            continue
        ws.column_dimensions[get_column_letter(col)].width = width


def role_column_width_for_content(
    ws: Worksheet,
    col: int,
    role: str,
    *,
    padding: float | None = None,
) -> float | None:
    """Return the clamped content-driven width for one role column.

    The helper is intentionally independent from setting the worksheet
    dimension so builders can resolve one workbook-wide width per role before
    applying it to every sheet. That keeps same-role columns visually stable
    across the workbook while preserving content-driven growth.
    """
    bounds = ROLE_WIDTH_BOUNDS.get(role)
    if bounds is None:
        return None
    pad = ROLE_WIDTH_PADDING if padding is None else padding
    role_min, role_max = bounds
    content_max = 0
    for r in range(1, max(ws.max_row, 1) + 1):
        value = ws.cell(row=r, column=col).value
        content_max = max(content_max, _display_width(value))
    target = content_max + pad if content_max > 0 else role_min
    return max(role_min, min(role_max, target))


def apply_indent_column_widths(
    ws: Worksheet, columns: Iterable[str | int]
) -> None:
    """Indent / hierarchy 役割を持つ列を canonical 幅 (=INDENT_COL_WIDTH) に固定.

    Use this helper instead of writing `column_dimensions[X].width = 2.14`
    by hand so the indent rhythm stays consistent across sheets and is
    grep-able when the rule needs to be audited.

    Args:
        ws: Worksheet
        columns: 列文字 (例: ['B']) または列番号 iterable (例: range(2, 4))
    """
    for col in columns:
        letter = col if isinstance(col, str) else get_column_letter(col)
        ws.column_dimensions[letter].width = INDENT_COL_WIDTH


def audit_indent_column_widths(
    wb,
    *,
    indent_columns: Iterable[str | int] = ("B",),
) -> list[str]:
    """Workbook 全体で indent 列が canonical 幅か検査して違反を返す.

    Returns:
        違反メッセージのリスト (空 = 全シート合格).
    """
    targets = [
        col if isinstance(col, str) else get_column_letter(col)
        for col in indent_columns
    ]
    violations: list[str] = []
    for ws in wb.worksheets:
        for letter in targets:
            dim = ws.column_dimensions.get(letter)
            width = dim.width if dim else None
            if width is None:
                violations.append(
                    f"{ws.title}!{letter}: indent column has no explicit width "
                    f"(expected {INDENT_COL_WIDTH})"
                )
                continue
            if abs(float(width) - float(INDENT_COL_WIDTH)) >= COL_WIDTH_TOL:
                violations.append(
                    f"{ws.title}!{letter}: indent column width={float(width):.4f} "
                    f"(expected {INDENT_COL_WIDTH})"
                )
    return violations


def apply_section_header(cell, label: str) -> None:
    """Section header text for no-merge section bands.

    Source-plan workbooks use a propagated navy band across the active table
    width. The label cell carries the text; sibling cells carry the band fill.
    """
    cell.value = label
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color=IB_INK)
    cell.alignment = Alignment(
        horizontal=ALIGN_LABEL, vertical=ALIGN_VERTICAL_BODY, indent=0, wrap_text=False
    )
    # 上 thin border は呼び出し側で section row 全体に対し適用する
    # (apply_section_top_border helper を使用)


def apply_section_header_band(
    ws: Worksheet,
    *,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
) -> None:
    """Section header band without merged cells.

    `apply_section_header` を leftmost cell に適用し、navy fill を
    end_col まで propagate する (merge_cells を使わずに 1 行分の band を表現)。

    Why no merge:
        merge_cells は sort / filter / range select / formula reference /
        fill drag を破壊するため (IB best practice / Macabacus・WSP 推奨).
        同等の見た目は fill propagation で実現する。

    See: references/_layout_canonical.md §4.7 No-Merge Rule

    Args:
        ws: Worksheet
        row: header row (1-indexed)
        start_col: leftmost column index (1-indexed, value はここに置く)
        end_col: rightmost column index (1-indexed, fill はここまで)
        label: section header 文字列
    """
    if end_col < start_col:
        raise ValueError(
            f"apply_section_header_band: end_col ({end_col}) must be >= "
            f"start_col ({start_col})"
        )
    band_fill = PatternFill("solid", fgColor=BG_HEADER_BAND)
    band_font = Font(
        name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color="FFFFFF"
    )
    band_align = Alignment(
        horizontal=ALIGN_LABEL, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )
    head = ws.cell(row=row, column=start_col)
    apply_section_header(head, label)
    head.fill = band_fill
    head.font = band_font
    head.alignment = band_align
    if end_col == start_col:
        return
    for col in range(start_col + 1, end_col + 1):
        c = ws.cell(row=row, column=col)
        # value は左端のみ (No-Merge 原則)
        c.fill = band_fill
        c.font = band_font
        c.alignment = band_align


def apply_year_header(cell, label: str) -> None:
    """期ヘッダ (Y1 / Q1 / Jan-26 等、italic + bold)."""
    cell.value = label
    cell.font = FONT_YEAR_HEADER
    cell.alignment = Alignment(
        horizontal=ALIGN_PERIOD_HEADER,
        vertical=ALIGN_VERTICAL_BODY,
        wrap_text=False,
    )


def apply_subtotal(
    cell,
    fmt: str = FMT_MONEY,
) -> None:
    """小計セル (top single border + bold)."""
    cell.font = FONT_TOTAL
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )
    cell.border = BORDER_SUBTOTAL


def apply_grand_total(
    cell,
    fmt: str = FMT_MONEY,
) -> None:
    """合計セル (top thin + bottom medium + bold + total band)."""
    cell.font = FONT_TOTAL
    cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )
    cell.border = BORDER_GRAND_TOTAL
    cell.fill = PatternFill("solid", fgColor=BG_TOTAL_BAND)


def apply_comment(cell, wrap_text: bool = False) -> None:
    """注記セル (gray + italic).

    wrap_text=True は禁止。長文は列幅、専用 note 列、短い行分割、
    または隣接空白セルへの overflow で解決する。
    """
    _ensure_no_wrap_text(wrap_text)
    cell.font = FONT_COMMENT
    cell.alignment = Alignment(
        horizontal=ALIGN_SOURCE_NOTE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def apply_working_highlight(cell) -> None:
    """WIP / TODO セル (yellow highlight)."""
    cell.fill = PatternFill("solid", fgColor=BG_WORKING)


def apply_chart_palette(chart, palette: list[str]) -> None:
    """各 series に palette の色を順に循環適用.

    palette 例: IB_CHART_COLORS_BAR, IB_CHART_COLORS_LINE 等。
    series 数 > palette 長さなら modulo で循環。

    Args:
        chart: openpyxl chart instance (BarChart / LineChart / etc.)
        palette: hex color string list (例 ["1F3A66", "008A80", ...])
    """
    for i, series in enumerate(chart.series):
        color = palette[i % len(palette)]
        try:
            series.graphicalProperties = GraphicalProperties(solidFill=color)
        except Exception:
            # Fallback: 一部 chart 種では graphicalProperties 経路で fail するため、
            # silent fallthrough (caller 側で別経路適用するときの邪魔をしない).
            pass


# ----------------------------------------------------------------------------
# 10b. Unit label / currency helpers
# ----------------------------------------------------------------------------


def fmt_for_currency(currency: str = "JPY", scale: str = "million") -> str:
    """指定通貨 × scale の number_format を返す.

    cell value は常に基本単位 (円 / dollar) の生数値で保存し、表示は
    number_format で切り替える設計。

    Args:
        currency: "JPY" / "USD"
        scale: "actual" (生値) / "thousand" (千) / "million" (百万) /
               "hundred_million" (億, JPY のみ; arithmetic は百万単位)

    Returns:
        Excel number_format 文字列。マッチしない場合は ``FMT_NUM_MILLION``。
    """
    table: dict[tuple[str, str], str] = {
        ("JPY", "actual"): FMT_JPY_YEN,
        ("JPY", "thousand"): FMT_JPY_THOUSAND,
        ("JPY", "million"): FMT_JPY_MILLION,
        ("JPY", "hundred_million"): FMT_JPY_HUNDRED_MILLION,
        ("JPY", "billion"): FMT_JPY_BILLION,
        ("JPY", "trillion"): FMT_JPY_TRILLION,
        ("USD", "actual"): FMT_USD_DOLLAR,
        ("USD", "thousand"): FMT_USD_THOUSAND,
        ("USD", "million"): FMT_USD_MILLION,
    }
    return table.get((currency, scale), FMT_NUM_MILLION)


_UNIT_LABEL_MAP: dict[tuple[str, str], str] = {
    ("JPY", "actual"): "円",
    ("JPY", "thousand"): "千円",
    ("JPY", "million"): "百万円",
    ("JPY", "hundred_million"): "億円",
    ("USD", "actual"): "$",
    ("USD", "thousand"): "$K",
    ("USD", "million"): "$M",
}


def apply_unit_label(cell, currency: str = "JPY", scale: str = "million") -> None:
    """シート右上等に「百万円」「$M」のような単位だけのラベルを書く.

    cell.value をこの関数が直接設定する点に注意 (caller は cell の選択のみ)。
    font は Arial 9pt italic gray (FONT_COMMENT 同等)。

    Args:
        cell: 単位ラベルを表示するセル (sheet header 右上等)
        currency: "JPY" / "USD"
        scale: "actual" / "thousand" / "million" / "hundred_million"
    """
    cell.value = _UNIT_LABEL_MAP.get((currency, scale), "(単位: 不明)")
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT)
    cell.alignment = Alignment(horizontal=ALIGN_UNIT, vertical=ALIGN_VERTICAL_BODY, wrap_text=False)


# ----------------------------------------------------------------------------
# 10c. Border helpers
# ----------------------------------------------------------------------------


def _merge_border(existing: Border | None, *, top=None, bottom=None, left=None, right=None) -> Border:
    """既存 Border に新しい side を上書きする (None の side は保持)."""
    if existing is None:
        return Border(top=top, bottom=bottom, left=left, right=right)
    return Border(
        top=top if top is not None else existing.top,
        bottom=bottom if bottom is not None else existing.bottom,
        left=left if left is not None else existing.left,
        right=right if right is not None else existing.right,
    )


def apply_section_bottom_border(cell, *, weight: str = "medium") -> None:
    """セクション末 (subtotal / grand total / sub-section 区切り) の bottom border.

    Args:
        cell: 適用先 cell
        weight: "thin" / "medium" / "dotted"。strong total は medium を使う。
    """
    side = {"thin": THIN_LINE, "medium": MEDIUM_LINE, "dotted": DOTTED_LINE}.get(weight, MEDIUM_LINE)
    cell.border = _merge_border(cell.border, bottom=side)


def apply_box_border(
    ws: Worksheet,
    range_str: str,
    *,
    weight: str = "thin",
    inner_dotted: bool = False,
) -> None:
    """範囲全体に枠線 (sensitivity matrix / call-out box / data table 用).

    Args:
        ws: Worksheet
        range_str: "B5:G12" 形式 or "B5" (単一 cell)
        weight: "thin" / "medium" / "dotted" — 外枠の線種
        inner_dotted: True で内部に dotted を引く。通常は行ごとの罫線を避ける。

    Note:
        既存 cell.border は上書きされる外枠 side のみ更新し、他の side は保持される。
    """
    side_map = {"thin": THIN_LINE, "medium": MEDIUM_LINE, "dotted": DOTTED_LINE}
    outer = side_map.get(weight, THIN_LINE)
    min_col, min_row, max_col, max_row = range_boundaries(range_str)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            top = outer if r == min_row else (DOTTED_LINE if inner_dotted else None)
            bottom = outer if r == max_row else (DOTTED_LINE if inner_dotted else None)
            left = outer if c == min_col else (DOTTED_LINE if inner_dotted else None)
            right = outer if c == max_col else (DOTTED_LINE if inner_dotted else None)
            cell.border = _merge_border(cell.border, top=top, bottom=bottom, left=left, right=right)


# ----------------------------------------------------------------------------
# 10d. Conditional formatting helpers
# ----------------------------------------------------------------------------
# 用途: sensitivity matrix / cohort retention / KPI heatmap 等
# 色は IB convention に寄せる (red→yellow→green or 中立 grayscale)
# ※ caller は range_str に sheet 内の data range のみを指定 (header 行/列は除く)
# ----------------------------------------------------------------------------

# Heatmap colors — color-blind safer ramp (青→黄→赤 は deuteranopia でも判別可)
HEATMAP_LOW_COOL = "5A8FBF"     # cool blue (低)
HEATMAP_MID_NEUTRAL = "F5F5F5"  # neutral light gray (中)
HEATMAP_HIGH_WARM = "C04A4A"    # warm red (高)

# Excel-classic green-yellow-red (IB sensitivity 表で使われる伝統的 palette)
HEATMAP_GREEN = "63BE7B"
HEATMAP_YELLOW = "FFEB84"
HEATMAP_RED = "F8696B"


def apply_heatmap_3color(
    ws: Worksheet,
    range_str: str,
    *,
    low_color: str = HEATMAP_GREEN,
    mid_color: str = HEATMAP_YELLOW,
    high_color: str = HEATMAP_RED,
    invert: bool = False,
) -> None:
    """sensitivity / cohort retention / KPI matrix 用の 3 色 heatmap.

    Args:
        ws: Worksheet
        range_str: "B5:G12" 形式の data range (header 行/列は含めない)
        low_color/mid_color/high_color: 6-hex (#は不要)
        invert: True で low/high を逆転 (低が悪い指標で使う; e.g. churn は低=良 だが NPS は高=良)
    """
    if invert:
        low_color, high_color = high_color, low_color
    rule = ColorScaleRule(
        start_type="min", start_color=low_color,
        mid_type="percentile", mid_value=50, mid_color=mid_color,
        end_type="max", end_color=high_color,
    )
    ws.conditional_formatting.add(range_str, rule)


def apply_data_bar(ws: Worksheet, range_str: str, *, color: str = BRAND_PRIMARY) -> None:
    """data bar conditional format (KPI dashboard で進捗 bar として使う).

    Args:
        ws: Worksheet
        range_str: data range
        color: 6-hex bar color (default = primary teal)
    """
    rule = DataBarRule(
        start_type="min", end_type="max", color=color,
        showValue=True,
    )
    ws.conditional_formatting.add(range_str, rule)


# ----------------------------------------------------------------------------
# 10e. Data validation
# ----------------------------------------------------------------------------


def apply_dropdown(
    ws: Worksheet,
    range_str: str,
    options: list[str],
    *,
    allow_blank: bool = True,
    prompt_title: str | None = None,
    prompt: str | None = None,
) -> None:
    """セル範囲に dropdown を付ける (Mode / Currency / Scale / Case 選択用).

    Args:
        ws: Worksheet
        range_str: "C3" or "C3:C20" 等
        options: dropdown に表示する選択肢 (例 ["Base", "Bull", "Bear"])
        allow_blank: 空白を許容するか
        prompt_title: cell 選択時 tooltip タイトル
        prompt: cell 選択時 tooltip 本文

    Note:
        Excel の list-formula は 255 文字制限あり。長い list は別 sheet に書き、
        "=Lists!$A$2:$A$N" 形式で参照すること (本 helper は in-line list のみ対応)。
    """
    # double quote escape (Excel formula は ' ではなく " が引用符)
    escaped = [opt.replace('"', '""') for opt in options]
    formula = f'"{",".join(escaped)}"'
    if len(formula) > 255:
        raise ValueError(
            f"dropdown options too long ({len(formula)} > 255 chars). "
            "Use external list reference instead."
        )
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False,
    )
    if prompt_title:
        dv.promptTitle = prompt_title
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    dv.add(range_str)
    ws.add_data_validation(dv)


def apply_numeric_validation(
    ws: Worksheet,
    range_str: str,
    *,
    minimum: float | None = None,
    maximum: float | None = None,
    integer: bool = False,
    error_message: str | None = None,
) -> None:
    """数値範囲制約を設定 (例: % は 0..1, 月数は 0..120, 株数は 0+ 整数).

    Args:
        ws: Worksheet
        range_str: 対象範囲
        minimum / maximum: 上下限 (None で片側制約)
        integer: True で整数のみ
        error_message: 違反時のエラー
    """
    if minimum is not None and maximum is not None:
        operator, formula1, formula2 = "between", str(minimum), str(maximum)
    elif minimum is not None:
        operator, formula1, formula2 = "greaterThanOrEqual", str(minimum), None
    elif maximum is not None:
        operator, formula1, formula2 = "lessThanOrEqual", str(maximum), None
    else:
        return  # nothing to validate

    dv = DataValidation(
        type="whole" if integer else "decimal",
        operator=operator,
        formula1=formula1,
        formula2=formula2,
        allow_blank=True,
    )
    if error_message:
        dv.error = error_message
        dv.errorTitle = "Invalid input"
        dv.showErrorMessage = True
    dv.add(range_str)
    ws.add_data_validation(dv)


# ----------------------------------------------------------------------------
# 10f. Hyperlink helpers
# ----------------------------------------------------------------------------


def apply_internal_link(
    cell,
    *,
    target_sheet: str,
    target_cell: str = "A1",
    display_text: str | None = None,
) -> None:
    """セルを内部 sheet jump link 化 (Cover の ToC や cross-sheet ナビ用).

    Args:
        cell: 適用先 cell
        target_sheet: 飛び先 sheet 名 (例 "03_Revenue")
        target_cell: 飛び先 cell (default "A1")
        display_text: 表示テキスト (省略時は target_sheet 名)

    Excel hyperlink 形式: ``#'SheetName'!A1`` (シート名は単引用符)。
    """
    cell.value = display_text if display_text is not None else target_sheet
    cell.hyperlink = f"#'{target_sheet}'!{target_cell}"
    cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_BASE, color=LINK_COLOR, underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def write_toc(
    ws: Worksheet,
    sheet_names: Iterable[str],
    *,
    start_row: int = 10,
    label_col: str = "B",
    title: str = "Contents",
) -> int:
    """Cover シートに目次 (Table of Contents) を書く.

    Args:
        ws: Cover sheet
        sheet_names: 飛び先 sheet 名の iterable (e.g. the workbook bundle minus Guide)
        start_row: 目次の開始 row
        label_col: 目次 label 列 (default "B")
        title: 目次の見出し

    Returns:
        目次最後の row (caller が次行を続けて書ける)
    """
    title_cell = ws[f"{label_col}{start_row}"]
    title_cell.value = title
    title_cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_LARGE, bold=True, color=IB_INK)

    row = start_row + 1
    for name in sheet_names:
        cell = ws[f"{label_col}{row}"]
        apply_internal_link(cell, target_sheet=name, target_cell="A1", display_text=name)
        row += 1
    return row - 1


# ----------------------------------------------------------------------------
# 10g. Row-height helper
# ----------------------------------------------------------------------------


def apply_row_heights(ws: Worksheet, row_kind_map: dict[int, str]) -> None:
    """row → kind ("tight" / "base" / "relaxed" / "hero") を一括設定.

    Args:
        ws: Worksheet
        row_kind_map: {row_index: kind_name}
    """
    height_map = {
        "tight": ROW_HEIGHT_TIGHT,
        "base": ROW_HEIGHT_BASE,
        "relaxed": ROW_HEIGHT_RELAXED,
        "hero": ROW_HEIGHT_HERO,
    }
    for row, kind in row_kind_map.items():
        ws.row_dimensions[row].height = height_map.get(kind, ROW_HEIGHT_BASE)


def wrapped_text_row_height(line_count: int) -> float:
    """Return exact row height for a user-approved wrapped text exception.

    Generated model sheets should normally keep wrapping off. When the user
    explicitly asks for wrapped prose or manual line breaks, size the row by the
    visible line count so text is not clipped and the row does not carry loose
    excess whitespace.
    """
    return max(1, int(line_count)) * ROW_HEIGHT_PER_WRAPPED_LINE


def visible_text_line_count(*values: object) -> int:
    """Count visible manual-line-break lines for wrapped-text exceptions."""
    counts = []
    for value in values:
        text = "" if value is None else str(value)
        counts.append(text.count("\n") + 1)
    return max(counts, default=1)


def set_wrapped_exception_row_height(ws: Worksheet, row: int, line_count: int) -> None:
    """Set exact row height for a user-approved wrap/manual-break exception."""
    ws.row_dimensions[row].height = wrapped_text_row_height(line_count)


def apply_wrapped_exception(
    cell: Cell,
    *,
    user_approved: bool,
    bounded_prose: bool,
    adjacent_cells_carry_meaning: bool,
    line_count: int | None = None,
) -> None:
    """Apply the only sanctioned wrap path for bounded table prose exceptions.

    Normal generated workbook cells must not use wrapping. Call this only after
    the user has explicitly approved bounded table prose that cannot read
    through blank unstyled overflow cells because neighboring table cells carry
    meaningful values, formulas, units, or notes. The helper sets wrapping and
    exact row height together so the exception cannot clip text or create a
    loose padded row.
    """
    if not user_approved:
        raise ValueError(
            "Wrapped prose exceptions require explicit user approval; use "
            "blank unstyled overflow cells without merging for horizontal-read text."
        )
    if not bounded_prose or not adjacent_cells_carry_meaning:
        raise ValueError(
            "Wrapped prose exceptions are only for bounded table prose whose "
            "adjacent table cells carry meaningful values, formulas, units, or notes."
        )
    base = cell.alignment
    count = visible_text_line_count(cell.value) if line_count is None else max(1, int(line_count))
    cell.alignment = Alignment(
        horizontal=base.horizontal or "left",
        vertical=base.vertical or "center",
        text_rotation=base.text_rotation,
        wrap_text=True,
        shrink_to_fit=base.shrink_to_fit,
        indent=base.indent,
        relativeIndent=base.relativeIndent,
        justifyLastLine=base.justifyLastLine,
        readingOrder=base.readingOrder,
    )
    set_wrapped_exception_row_height(cell.parent, cell.row, count)


# ============================================================================
# 11. Sheet-level helper functions
# ============================================================================


def setup_sheet_layout(
    ws: Worksheet,
    *,
    n_periods: int = 36,
    has_unit_col: bool = True,
) -> None:
    """シート全体のレイアウト初期化.

    Column structure:
        A: narrow gutter
        B: parent hierarchy
        C: line item label
        D: source / driver
        E: unit (optional)
        F..F+n-1: periods
    """
    ws.column_dimensions["A"].width = COL_MARGIN_WIDTH
    apply_indent_column_widths(ws, ["B"])
    ws.column_dimensions["C"].width = COL_LABEL_WIDTH
    ws.column_dimensions["D"].width = COL_SOURCE_WIDTH
    if has_unit_col:
        ws.column_dimensions["E"].width = COL_UNIT_WIDTH
        period_start = 6  # F
    else:
        period_start = 5  # E
    for i in range(n_periods):
        col_letter = get_column_letter(period_start + i)
        ws.column_dimensions[col_letter].width = COL_PERIOD_WIDTH

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None


def setup_print_layout(
    ws: Worksheet,
    *,
    orientation: Literal["landscape", "portrait"] = "landscape",
    fit_to_width: int = 1,
    print_title_rows: str | None = None,
    print_title_cols: str | None = None,
    header_left: str | None = None,
    header_right: str | None = None,
    footer_left: str | None = None,
    footer_center: str | None = None,
    footer_right: str | None = None,
) -> None:
    """印刷設定 (landscape default + fit-to-width).

    Print title rows / cols and optional header / footer settings are applied
    together.

    Args:
        ws: Worksheet
        orientation: "landscape" / "portrait"
        fit_to_width: fit-to-width pages (1 = 1 ページ幅に収める)
        print_title_rows: 各ページで repeat する行 (例: "1:4")
        print_title_cols: 各ページで repeat する列 (例: "A:B")
        header_left/right: page header の左/右 (Excel header コード使用可: "&A"=sheet 名, "&D"=日付)
        footer_left/center/right: page footer の各位置 ("&P"=page #, "&N"=total pages)

    Excel format codes (Microsoft 仕様):
        &P = current page #     &N = total pages    &D = date    &T = time
        &A = sheet name         &F = file name      &Z = path    &G = picture
    """
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = PRINT_MARGIN_INCH
    ws.page_margins.right = PRINT_MARGIN_INCH
    ws.page_margins.top = PRINT_MARGIN_INCH
    ws.page_margins.bottom = PRINT_MARGIN_INCH
    ws.page_margins.header = PRINT_HEADER_INCH
    ws.page_margins.footer = PRINT_HEADER_INCH

    # Print title (各ページで repeat される行/列)
    if print_title_rows:
        ws.print_title_rows = print_title_rows
    if print_title_cols:
        ws.print_title_cols = print_title_cols

    # Header / Footer (オプション)
    if header_left is not None:
        ws.oddHeader.left.text = header_left
    if header_right is not None:
        ws.oddHeader.right.text = header_right
    if footer_left is not None:
        ws.oddFooter.left.text = footer_left
    if footer_center is not None:
        ws.oddFooter.center.text = footer_center
    if footer_right is not None:
        ws.oddFooter.right.text = footer_right


def write_cover(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str = "",
    project_code: str = "",
    confidential: bool = True,
    author: str = "",
    version: str = "1.0",
    date_str: str = "",
) -> None:
    """Cover sheet に必須情報を書き込む.

    Layout (B-row 中心):
        B2: CONFIDENTIAL stamp (top-right)
        B6: title (大)
        B8: subtitle (中)
        B14: project_code / version / date (footer)
    """
    if confidential:
        ws["F2"] = "CONFIDENTIAL"
        ws["F2"].font = Font(
            name=FONT_FAMILY,
            size=FONT_SIZE_SMALL,
            bold=True,
            color=BRAND_PRIMARY_DEEP,
        )
        ws["F2"].alignment = Alignment(horizontal="right")

    ws["B6"] = title
    ws["B6"].font = FONT_COVER_TITLE
    ws["B6"].fill = PatternFill("solid", fgColor=BRAND_PRIMARY_DEEP)
    ws["B6"].alignment = Alignment(horizontal="left", vertical="center", indent=0)
    ws.row_dimensions[6].height = ROW_HEIGHT_COVER_TITLE

    if subtitle:
        ws["B8"] = subtitle
        ws["B8"].font = Font(name=FONT_FAMILY, size=FONT_SIZE_TITLE, color=BRAND_INK)
        ws["B8"].alignment = Alignment(horizontal="left")

    footer_lines = []
    if project_code:
        footer_lines.append(f"Project: {project_code}")
    if version:
        footer_lines.append(f"Version: {version}")
    if date_str:
        footer_lines.append(f"Date: {date_str}")
    if author:
        footer_lines.append(f"Author: {author}")
    if footer_lines:
        ws["B14"] = " | ".join(footer_lines)
        apply_comment(ws["B14"])

    setup_sheet_layout(ws, n_periods=4, has_unit_col=False)
    ws.sheet_properties.tabColor = BRAND_NAVY  # §2.X canonical: Cover = Navy


# ----------------------------------------------------------------------------
# Sheet Tab Color
# ----------------------------------------------------------------------------

# 6 canonical role → tab color (hex, no leading #)
TAB_COLOR_BY_ROLE: dict[str, str] = {
    "cover":  BRAND_NAVY,           # 1F3A66 — document title page
    "input":  BRAND_PRIMARY_DEEP,   # 1F3A66 — user-editable hard inputs
    "driver": BRAND_PRIMARY,        # 2F75B5 — calculated drivers / KPI
    "output": BRAND_SLATE,          # 7F8FA6 — calculated 3-statement / valuation outputs
    "check":  BRAND_WARNING,        # 5B9BD5 — sanity check status
    "memo":   BRAND_ACCENT,         # ECC85A — IC thesis prose
}

# SHEET_ROLE_MAPPING bridges legacy unit tests that expect role-based sheet maps.
SHEET_ROLE_MAPPING: dict[str, str] = {
    "Guide": "cover",
    "Kernel": "kernel",
    "Assumptions": "input",
    "Driver Tree": "input",
    "Revenue Build": "driver",
    "Cost Build": "driver",
    "People Plan": "driver",
    "P&L": "output",
    "BS": "output",
    "CF": "output",
    "Capital Stack": "driver",
    "Ownership": "input",
    "Pricing": "driver",
    "Financing": "driver",
    "Exit Waterfall": "output",
    "Segments": "driver",
    "KPI": "output",
    "Scenarios": "output",
    "Sensitivity": "output",
    "Valuation": "output",
    "Market Support": "driver",
    "Benchmarks": "driver",
    "IC Memo": "memo",
}


def set_tab_color(ws: Worksheet, role: str) -> None:
    """Sheet のタブ色を canonical role color に設定.

    canonical role colors:
      - cover  → BRAND_NAVY
      - input  → BRAND_PRIMARY_DEEP
      - driver → BRAND_PRIMARY
      - output → BRAND_SLATE
      - check  → BRAND_WARNING
      - memo   → BRAND_ACCENT

    Args:
        ws: Worksheet
        role: 明示 role (cover/input/driver/output/check/memo)。
            未知の role は no-op。
    """
    color = TAB_COLOR_BY_ROLE.get(role)
    if color:
        ws.sheet_properties.tabColor = color


# ============================================================================
# 12. Validation utilities
# ============================================================================


@dataclass(frozen=True)
class CellRoleViolation:
    sheet: str
    cell: str
    expected_role: str
    actual_color: str
    detail: str


# ============================================================================
# 13. Convenience constants for builders
# ============================================================================

# Money unit hints (input_schema reporting_currency と対応)
MONEY_UNIT_BY_CCY: dict[str, str] = {
    "JPY": "百万円",
    "USD": "$M",
    "EUR": "€M",
    "GBP": "£M",
    "SGD": "S$M",
}


# ============================================================================
# 13b. S2 format layer v2 (additive — 構造改革 slice S2)
# ============================================================================
# ARCHITECTURE.md §4 技術フォーマット体系 / §0 ヘッダ・ルーラー行の新契約。
# このセクションは**追加のみ**: 既存 token / helper は S3 が builder を移行する
# まで現行契約のまま並存する。新旧の識別は JP▲系 = FMT_JP_* / 幅 = *_V2。
#
# 数値書式 v2 の方針 (JP デフォルト):
#   - セル値は常に円の生値。表示スケールは number_format のみで切替。
#   - 負数は赤 + "▲" 接頭 (JP 決算書慣行)。括弧囲み (旧 FMT_JPY_*) は使わない。
#   - セル毎の ¥ 記号は廃止 → シート左上の「(単位: 百万円)」キャプション+単位列。
#   - ゼロは "-" 表示。
#
# 億円 (10^8) を v2 スケールラダーから除外する理由:
#   Excel の comma scaling は 1 個につき ÷1,000 (10^3 単位) しか刻めない。
#   ",,"=10^6 (百万) / ",,,"=10^9 (十億) であり、10^8=億 は comma scaling では
#   表現不可能。セル値のスケーリング (円以外で保存) は本設計で禁止のため、
#   億円書式は出荷しない。v2 スケールラダーは 円 → 千円 → 百万円 → 十億円。
#   (旧 FMT_JPY_HUNDRED_MILLION は「百万のまま億接尾」の近似であり v2 では不使用。)
# ----------------------------------------------------------------------------

# --- JP ▲ 系 number format tokens --------------------------------------------
FMT_JP_YEN      = '#,##0;[Red]"▲"#,##0;"-"'        # 円 (生値表示)
FMT_JP_THOUSAND = '#,##0,;[Red]"▲"#,##0,;"-"'      # 千円 (÷10^3)
FMT_JP_MILLION  = '#,##0,,;[Red]"▲"#,##0,,;"-"'    # 百万円 (÷10^6)
FMT_JP_BILLION  = '#,##0,,,;[Red]"▲"#,##0,,,;"-"'  # 十億円 (÷10^9)
FMT_JP_PERCENT  = '0.0%;[Red]"▲"0.0%'              # % (負は赤▲)
FMT_FACTOR      = '0.0000'                          # 係数 (FAST 3.01-08)
FMT_MONTHS_1DP  = '0.0'                             # 月数 (ランウェイ等)
# USD は既存の FMT_USD_DOLLAR / FMT_USD_THOUSAND / FMT_USD_MILLION
# (括弧囲み Accounting 形式) を v2 でもそのまま再利用する。

# --- スケールラダー / 単位ラベル ---------------------------------------------
# scale 識別子はラダー順。JP は billion まで、USD は million まで ($B は IB
# モデルの表示慣行上 $M 継続が普通のため ladder に含めない)。
JP_SCALE_LADDER: tuple[str, ...] = ("actual", "thousand", "million", "billion")
USD_SCALE_LADDER: tuple[str, ...] = ("actual", "thousand", "million")

JP_UNIT_BY_SCALE: dict[str, str] = {
    "actual": "円",
    "thousand": "千円",
    "million": "百万円",
    "billion": "十億円",
}
USD_UNIT_BY_SCALE: dict[str, str] = {
    "actual": "$",
    "thousand": "$K",
    "million": "$M",
}

FMT_JP_BY_SCALE: dict[str, str] = {
    "actual": FMT_JP_YEN,
    "thousand": FMT_JP_THOUSAND,
    "million": FMT_JP_MILLION,
    "billion": FMT_JP_BILLION,
}
FMT_USD_BY_SCALE: dict[str, str] = {
    "actual": FMT_USD_DOLLAR,
    "thousand": FMT_USD_THOUSAND,
    "million": FMT_USD_MILLION,
}


def fmt_jp_for_scale(scale: str) -> str:
    """v2 JP▲系 number_format を scale 識別子から返す。

    Args:
        scale: "actual" / "thousand" / "million" / "billion"

    Returns:
        FMT_JP_* トークン (セル値は円の生値のまま表示だけ切替)。

    Raises:
        ValueError: ラダー外の scale (億円含む — §13b 冒頭の除外理由参照)。
    """
    try:
        return FMT_JP_BY_SCALE[scale]
    except KeyError:
        raise ValueError(
            f"unknown JP scale {scale!r}: expected one of {JP_SCALE_LADDER} "
            "(億円は comma-scaling 制約により v2 ラダーから除外)"
        ) from None


# --- 決定論スケールピッカー (2層ルール — ARCHITECTURE.md §2 実装機構) ---------
# ルール: 「最大絶対値がそのスケールで 3.0 表示単位以上になる最大のスケール」
# を選ぶ。3x のヒステリシス係数により境界近傍 (例: 1.2M) が 「1」表示に潰れる
# ことを防ぐ (1,200千円 と表示される)。閾値:
#   JP : thousand ≥ 3_000 / million ≥ 3_000_000 / billion ≥ 3_000_000_000
#   USD: thousand ≥ 3_000 / million ≥ 3_000_000 ($M 上限)
# 純関数・決定論。入力値は常に基本通貨 (円 / dollar) の生値。

SCALE_HYSTERESIS_FACTOR = 3.0

# scale → 表示除数 (comma scaling の除数と一致)
_SCALE_DIVISOR: dict[str, float] = {
    "actual": 1.0,
    "thousand": 1_000.0,
    "million": 1_000_000.0,
    "billion": 1_000_000_000.0,
}


def _scale_ladder_for_currency(currency: str) -> tuple[str, ...]:
    return USD_SCALE_LADDER if currency == "USD" else JP_SCALE_LADDER


def pick_scale_for_magnitude(max_abs: float, currency: str = "JPY") -> str:
    """代表桁 (最大絶対値) から表示スケールを決定論的に選ぶ。

    「max_abs がそのスケールの除数で 3.0 表示単位以上になる」最大のスケールを
    ラダーから選ぶ (ヒステリシス係数 3x)。境界例:
        2_999_999 → "thousand" / 3_000_000 → "million"

    Args:
        max_abs: 系列の最大絶対値 (基本通貨の生値)。負値は絶対値として扱う。
        currency: "JPY" / "USD" (USD ラダーは million 止まり)。

    Returns:
        scale 識別子 ("actual" / "thousand" / "million" / "billion")。
        非有限値・0 近傍は "actual"。
    """
    magnitude = abs(float(max_abs))
    if not math.isfinite(magnitude):
        return "actual"
    chosen = "actual"
    for scale in _scale_ladder_for_currency(currency):
        divisor = _SCALE_DIVISOR[scale]
        if magnitude >= SCALE_HYSTERESIS_FACTOR * divisor:
            chosen = scale
    return chosen


def _max_finite_abs(values: Iterable[float]) -> float:
    best = 0.0
    for v in values:
        if v is None:
            continue
        f = float(v)
        if not math.isfinite(f):
            continue
        a = abs(f)
        if a > best:
            best = a
    return best


def pick_sheet_scale(values: Iterable[float], currency: str = "JPY") -> str:
    """シート単一スケール (2層ルール (a) — ステートメント/エンジン系) を選ぶ。

    シートの支配的桁 = 渡された集約行 (合計行等) の最大絶対値。呼び出し側は
    per-unit / 率などの例外行を除いた集約行の生値を渡す。

    Args:
        values: 集約行の生値 (円 / dollar)。None / 非有限値は無視。
        currency: "JPY" / "USD"。

    Returns:
        scale 識別子。空 iterable は "actual"。
    """
    return pick_scale_for_magnitude(_max_finite_abs(values), currency=currency)


def pick_row_scale(values: Iterable[float], currency: str = "JPY") -> str:
    """行スケール (2層ルール (b) — レジスタ系の行単位選択) を選ぶ。

    行の代表桁 = 系列の最大絶対値。選択規則はシートスケールと同一
    (決定論・同一入力 → 同一出力) だが、レジスタ系では行ごとに呼び、
    単位列ラベル (JP_UNIT_BY_SCALE) と number_format を必ず対で更新する。

    Args:
        values: 行の全期間値 (生値)。None / 非有限値は無視。
        currency: "JPY" / "USD"。

    Returns:
        scale 識別子。空 iterable は "actual"。
    """
    return pick_scale_for_magnitude(_max_finite_abs(values), currency=currency)


# --- v2 ヘッダ行契約 (期間軸シート — ARCHITECTURE.md §0) ----------------------
# row 2: タイトル (C2) + マスターチェックエコー (最初の期間列)
# row 3: パーパス (C3, gray italic)
# row 4: 単位キャプション (C4) + FY ラベルルーラー (期間列, centered 9pt gray)
# row 5: months_in_period ルーラー (期間列, right 9pt gray, integer)
# row 6: 期間ヘッダ (bold centered + 既存ヘッダバンド BG_TABLE_HEADER)
# row 7: スペーサ / data は row 8 から。フリーズは (first_period_col, row 7)。
HEADER_TITLE_ROW_V2 = 2
HEADER_PURPOSE_ROW_V2 = 3
HEADER_FY_RULER_ROW = 4
HEADER_MONTHS_RULER_ROW = 5
HEADER_PERIOD_ROW = 6
HEADER_ROWS_V2 = 6
DATA_START_ROW_V2 = 8
HEADER_CAPTION_COL = 3  # C 列 (単位キャプション / タイトル / パーパス)

_FONT_RULER = Font(name=FONT_FAMILY, size=FONT_SIZE_SMALL, color=IB_COMMENT)
_FONT_RULER_CAPTION = Font(
    name=FONT_FAMILY, size=FONT_SIZE_SMALL, italic=True, color=IB_COMMENT
)


def write_period_rulers(
    ws,
    first_period_col: int,
    fy_labels: list[str],
    months_in_period: list[int],
    period_labels: list[str],
    unit_caption: str | None,
) -> None:
    """v2 ヘッダのルーラー3行 (row 4/5/6) + 単位キャプション (C4) を書く。

    全期間式は months 行 (row 5) を参照する契約 (例 ``=F9*F11*F$5``) のため、
    months_in_period は int で格納する。3 リストは同じ長さ (期間数) であること。

    Args:
        ws: Worksheet (期間軸シート)
        first_period_col: 最初の期間列 index (1-based, canonical は 6 = F)
        fy_labels: 会計年度ラベル (例 ["FY2026", "FY2026", ...]) — row 4
        months_in_period: 各期間の月数 (月次=1 / 年次=12) — row 5
        period_labels: 期間ヘッダ (例 "2026-07" / "FY2028") — row 6
        unit_caption: 「(単位: 百万円)」等。None で省略 (register 系は
            「(単位: 単位列参照)」を渡す) — C4

    Raises:
        ValueError: 3 リストの長さが不一致のとき。
    """
    n = len(period_labels)
    if len(fy_labels) != n or len(months_in_period) != n:
        raise ValueError(
            "write_period_rulers: fy_labels / months_in_period / period_labels "
            f"must share one length (got {len(fy_labels)} / "
            f"{len(months_in_period)} / {n})"
        )

    if unit_caption:
        cap = ws.cell(row=HEADER_FY_RULER_ROW, column=HEADER_CAPTION_COL)
        cap.value = unit_caption
        cap.font = _FONT_RULER_CAPTION
        cap.alignment = Alignment(
            horizontal=ALIGN_LABEL, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
        )

    for i in range(n):
        col = first_period_col + i

        fy = ws.cell(row=HEADER_FY_RULER_ROW, column=col)
        fy.value = fy_labels[i]
        fy.font = _FONT_RULER
        fy.alignment = Alignment(
            horizontal=ALIGN_PERIOD_HEADER, vertical=ALIGN_VERTICAL_BODY,
            wrap_text=False,
        )

        months = ws.cell(row=HEADER_MONTHS_RULER_ROW, column=col)
        months.value = int(months_in_period[i])
        months.font = _FONT_RULER
        months.number_format = "0"
        months.alignment = Alignment(
            horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
        )

        period = ws.cell(row=HEADER_PERIOD_ROW, column=col)
        period.value = period_labels[i]
        period.font = FONT_YEAR_HEADER
        period.fill = PatternFill("solid", fgColor=BG_TABLE_HEADER)
        period.alignment = Alignment(
            horizontal=ALIGN_PERIOD_HEADER, vertical=ALIGN_VERTICAL_BODY,
            wrap_text=False,
        )


def apply_freeze_pane(ws, first_period_col: int, header_rows: int = HEADER_ROWS_V2) -> None:
    """v2 フリーズペイン契約 (期間軸シート) を適用する。

    アンカー = (first_period_col, header_rows + 1)。canonical の
    first_period_col=6 / header_rows=6 では "F7" になり、A-E 列 (ガター/階層/
    ラベル/ソースタグ/単位) とヘッダ 6 行が常時視認される。

    Note:
        既存 audit_workbook (build_model.py) は現行契約でフリーズを fail に
        するため、S3/S5 が監査を反転するまで builder からは未使用。

    Args:
        ws: Worksheet
        first_period_col: 最初の期間列 index (1-based)
        header_rows: フリーズするヘッダ行数 (default 6 → data row 8 の直前
            スペーサ row 7 がスクロール域先頭)
    """
    ws.freeze_panes = ws.cell(row=header_rows + 1, column=first_period_col).coordinate


def write_master_check_echo(ws, cell_ref: str, master_check_formula: str) -> None:
    """フリーズ域ヘッダにマスターチェックエコーセルを書く。

    全期間軸シートの row 2 × 最初の期間列に置き、Summary の統合チェック
    (マスターチェック 1 セル) を echo する契約 (ICAEW #18-20)。式は caller が
    渡す (例 ``'=IF(Summary!$F$40=0,"checks OK","CHECK FAILED")'``)。

    Args:
        ws: Worksheet
        cell_ref: 書き込み先 (例 "F2")
        master_check_formula: Excel 式文字列 (= 始まり)
    """
    cell = ws[cell_ref]
    cell.value = master_check_formula
    cell.font = _FONT_RULER
    cell.alignment = Alignment(
        horizontal=ALIGN_VALUE, vertical=ALIGN_VERTICAL_BODY, wrap_text=False
    )


def period_axis_columns(ws) -> list[int]:
    """Canonical v2 period-column detector — the single source of truth shared
    by the shipping strict audit (build_model) and every audit here.

    A period column carries the integer months-in-period ruler
    (HEADER_MONTHS_RULER_ROW = 5) under a period header (HEADER_PERIOD_ROW = 6).
    The trailing notes column has a row-6 header but no ruler, and the
    label/driver/unit columns have neither, so both are correctly excluded.
    Legacy / non-period sheets (string row-5 values, or no ruler) yield []."""
    cols: list[int] = []
    for col in range(2, ws.max_column + 1):
        months = ws.cell(row=HEADER_MONTHS_RULER_ROW, column=col).value
        header = ws.cell(row=HEADER_PERIOD_ROW, column=col).value
        if (
            isinstance(months, (int, float))
            and not isinstance(months, bool)
            and header not in (None, "")
        ):
            cols.append(col)
    return cols


def period_axis_anchor_col(ws) -> int | None:
    """First period column (freeze-pane anchor) per :func:`period_axis_columns`,
    or None when the sheet is not a v2 period-axis sheet."""
    cols = period_axis_columns(ws)
    return cols[0] if cols else None


def audit_period_column_widths(wb, expected_width: float) -> list[str]:
    """全期間軸シートの期間列が単一幅 expected_width かを監査する (R3)。

    期間軸シート・期間列の判定は :func:`period_axis_columns` に一元化
    (row-5 月次ルーラー + row-6 期間ヘッダ)。非期間系シート (Guide / Cap Table
    / IC Memo 等) と期間列が無いシートはスキップ。

    Args:
        wb: openpyxl Workbook
        expected_width: 全シート共通の期間列幅 (canonical は COL_PERIOD_WIDTH_V2)

    Returns:
        違反メッセージの list (空 = clean)。幅未設定 (None) も違反として報告。
    """
    violations: list[str] = []
    for ws in wb.worksheets:
        period_cols = period_axis_columns(ws)
        if not period_cols:
            continue
        for c in period_cols:
            letter = get_column_letter(c)
            # 注意: ws.column_dimensions[letter] は未設定列に default 幅の
            # ColumnDimension を自動生成するため、.get() で未設定を判定する。
            dim = ws.column_dimensions.get(letter)
            width = dim.width if dim is not None else None
            if width is None:
                violations.append(
                    f"sheet '{ws.title}' period column {letter}: width not set "
                    f"(expected {expected_width})"
                )
            elif abs(float(width) - float(expected_width)) >= COL_WIDTH_TOL:
                violations.append(
                    f"sheet '{ws.title}' period column {letter}: width {width} "
                    f"!= expected {expected_width}"
                )
    return violations


# ============================================================================
# 12. Validation utilities
# ============================================================================


@dataclass(frozen=True)
class CellRoleViolation:
    sheet: str
    cell: str
    expected_role: str
    actual_color: str
    detail: str


def validate_sheet_naming(workbook_sheetnames: Iterable[str]) -> list[str]:
    """Verify that sheet names are a canonical ordered subset.

    Focused-mode workbooks intentionally omit some canonical sheets. Treat that
    as valid when the remaining sheets preserve canonical relative order.
    """
    expected = list(CANONICAL_SHEET_ORDER)
    actual = list(workbook_sheetnames)
    violations: list[str] = []

    actual_filtered = [s for s in actual if s not in OPTIONAL_SHEETS]
    expected_positions = {name: i for i, name in enumerate(expected)}
    positions: list[int] = []
    for name in actual_filtered:
        if name not in expected_positions:
            violations.append(f"Extra sheet '{name}' (only canonical sheets + 99_Glossary allowed)")
            continue
        positions.append(expected_positions[name])
    if positions != sorted(positions):
        violations.append(f"Sheet order is not canonical: {actual_filtered}")

    return violations


# ============================================================================
# 13. Convenience constants for builders
# ============================================================================

# Money unit hints (input_schema reporting_currency と対応)
MONEY_UNIT_BY_CCY: dict[str, str] = {
    "JPY": "百万円",
    "USD": "$M",
    "EUR": "€M",
    "GBP": "£M",
    "SGD": "S$M",
}


# --- v2 列幅トークン (ARCHITECTURE.md §4 列体系 — 旧トークン is still present in S3) ---
COL_GUTTER_WIDTH_V2 = 1.5       # A: ガター/ナビ
COL_LABEL_WIDTH_V2 = 38.0       # C: ラベル (ワイド)
COL_SOURCE_WIDTH_INPUT = 30.0   # D: ソース列 (Assumptions/Evidence 入力系のみワイド)
COL_DRIVER_TAG_WIDTH = 10.0     # D: エンジン/ステートメント系ドライバタグ列
COL_UNIT_WIDTH_V2 = 9.0         # E: 単位
COL_PERIOD_WIDTH_V2 = 11.5      # F〜: 期間列 (全シート同一幅 — R3 監査対象)
COL_NOTE_WIDTH_V2 = 64.0        # 末尾+1: ノート/解釈列
# B (階層インデント) は既存 COL_HIERARCHY_WIDTH = 2.14 (20px 契約) を継続使用。


__all__ = [
    # Colors
    "IB_HARD_INPUT", "IB_FORMULA", "IB_LINK_INTRA", "IB_LINK_EXTERNAL", "IB_COMMENT",
    "IB_INK", "IB_PALETTE",
    "BRAND_PRIMARY_DEEP", "BRAND_PRIMARY", "BRAND_INK", "BRAND_SURFACE", "BRAND_ACCENT",
    "BRAND_NAVY", "BRAND_GS_NAVY", "BRAND_MS_BLUE", "BRAND_LAZARD_ORANGE",
    "BRAND_EVERCORE_NAVY",
    "BRAND_WARNING", "BRAND_DANGER", "BRAND_SUCCESS", "BRAND_SLATE",
    # Sheet tab color canonical mapping
    "TAB_COLOR_BY_ROLE",
    "BG_WHITE", "BG_CANVAS", "BG_TABLE_HEADER", "BG_TOTAL_BAND", "BG_HEADER_BAND", "BG_WORKING",
    "SEMANTIC_BLANK_FILL_COLORS", "apply_semantic_fill_span",
    "apply_semantic_border_span", "detect_table_block",
    "LINK_COLOR",
    # Heatmap palettes
    "HEATMAP_LOW_COOL", "HEATMAP_MID_NEUTRAL", "HEATMAP_HIGH_WARM",
    "HEATMAP_GREEN", "HEATMAP_YELLOW", "HEATMAP_RED",
    # Chart palettes
    "IB_CHART_COLORS_FOOTBALL", "IB_CHART_COLORS_BAR", "IB_CHART_COLORS_LINE",
    "IB_CHART_COLORS_WATERFALL_POS", "IB_CHART_COLORS_WATERFALL_NEG",
    "IB_CHART_COLORS_WATERFALL_TOTAL",
    # Fonts (Arial 10pt 基準 — Google Sheets default + IB de facto standard)
    "FONT_FAMILY", "FONT_SIZE_BASE", "FONT_SIZE_SMALL", "FONT_SIZE_LARGE",
    "FONT_SIZE_TITLE", "FONT_SIZE_TINY", "FONT_SIZE_ALLOWED_CELLS",
    "FONT_SIZE_BEST_PRACTICE",
    "FONT_BODY", "FONT_BODY_BOLD", "FONT_HARD_INPUT", "FONT_FORMULA",
    "FONT_LINK_INTRA", "FONT_LINK_EXTERNAL",
    "FONT_COMMENT", "FONT_FOOTNOTE", "FONT_SECTION_HEADER", "FONT_TITLE",
    "FONT_YEAR_HEADER", "FONT_SUBTOTAL", "FONT_GRAND_TOTAL", "FONT_COVER_TITLE",
    # Font semantic aliases
    "FONT_SECTION", "FONT_SUBSECTION", "FONT_TOTAL",
    # Alignment semantics
    "ALIGN_LABEL", "ALIGN_SOURCE_NOTE", "ALIGN_VALUE", "ALIGN_UNIT",
    "ALIGN_PERIOD_HEADER", "ALIGN_VERTICAL_BODY", "ALIGNMENT_BEST_PRACTICE",
    # Wrap semantics
    "WRAP_TEXT_ERROR", "WRAP_DECISION_LADDER", "WRAP_BEST_PRACTICE",
    # Number formats
    "FMT_JPY_YEN", "FMT_JPY_THOUSAND", "FMT_JPY_MILLION", "FMT_JPY_BILLION",
    "FMT_JPY_TRILLION", "FMT_JPY_HUNDRED_MILLION",
    "FMT_USD_DOLLAR", "FMT_USD_THOUSAND", "FMT_USD_MILLION",
    "FMT_NUM", "FMT_NUM_THOUSAND", "FMT_NUM_MILLION",
    "FMT_PCT_0", "FMT_PCT_1", "FMT_PCT_2",
    "FMT_MULTIPLE_1", "FMT_MULTIPLE_2",
    "FMT_DATE_YMD", "FMT_DATE_YM",
    "FMT_INTEGER", "FMT_PER_SHARE",
    # Number format semantic aliases
    "FMT_MONEY", "FMT_MONEY_DECIMAL", "FMT_PERCENT", "FMT_PERCENT_BPS",
    "FMT_MULTIPLE", "FMT_SHARES", "FMT_DATE_SHORT", "FMT_DATE_LONG",
    "FMT_DEFAULT",
    # Layout
    "COL_MARGIN_WIDTH", "COL_HIERARCHY_WIDTH", "INDENT_COL_WIDTH", "COL_WIDTH_TOL",
    "COL_LABEL_WIDTH", "COL_SOURCE_WIDTH",
    "COL_UNIT_WIDTH", "COL_PERIOD_WIDTH",
    "COL_WIDTH_TINY", "COL_WIDTH_SMALL", "COL_WIDTH_BASE", "COL_WIDTH_LARGE", "COL_WIDTH_XLARGE",
    "ROLE_WIDTH_BOUNDS", "ROLE_WIDTH_PADDING", "autosize_role_columns",
    "role_column_width_for_content",
    "ROW_BODY_HEIGHT", "ROW_SPACER_HEIGHT", "ROW_SECTION_HEIGHT",
    "ROW_HEIGHT_TIGHT", "ROW_HEIGHT_BASE", "ROW_HEIGHT_STANDARD", "ROW_HEIGHT_MEDIUM",
    "ROW_HEIGHT_RELAXED", "ROW_HEIGHT_HERO", "ROW_HEIGHT_COVER_TITLE", "ROW_HEIGHT_PER_WRAPPED_LINE",
    # Borders
    "BORDER_SUBTOTAL", "BORDER_GRAND_TOTAL", "BORDER_SECTION_DIVIDER",
    "BORDER_SECTION_END_MEDIUM", "BORDER_TOP_THIN", "BORDER_BOTTOM_THIN",
    "BORDER_BOX_THIN", "BORDER_BOX_MEDIUM",
    "THIN_LINE", "MEDIUM_LINE", "DOTTED_LINE",
    # Cell-level helpers
    "apply_hard_input", "apply_formula", "apply_link_intra", "apply_link_external",
    "apply_label", "write_hierarchical_line_item",
    "set_uniform_column_width", "apply_indent_column_widths", "audit_indent_column_widths",
    "apply_chart_palette",
    "apply_section_header", "apply_year_header",
    "apply_subtotal", "apply_grand_total", "apply_comment", "apply_working_highlight",
    # Unit / currency helpers
    "apply_unit_label", "fmt_for_currency",
    # Border helpers
    "apply_section_bottom_border", "apply_box_border",
    # Conditional formatting
    "apply_heatmap_3color", "apply_data_bar",
    # Data validation
    "apply_dropdown", "apply_numeric_validation",
    # Hyperlink
    "apply_internal_link", "write_toc",
    # Row height
    "apply_row_heights", "wrapped_text_row_height", "visible_text_line_count", "set_wrapped_exception_row_height",
    # Sheet-level helpers
    "setup_sheet_layout", "setup_print_layout", "write_cover", "set_tab_color",
    "last_value_bounds", "rendered_bounds", "clear_blank_cell_styles", "trim_blank_canvas",
    # Currency
    "MONEY_UNIT_BY_CCY",
    # Models
    "CellRoleViolation",
    # --- S2 format layer v2 (additive) ---
    # JP ▲ 系 number formats
    "FMT_JP_YEN", "FMT_JP_THOUSAND", "FMT_JP_MILLION", "FMT_JP_BILLION",
    "FMT_JP_PERCENT", "FMT_FACTOR", "FMT_MONTHS_1DP",
    # Scale ladder / unit labels
    "JP_SCALE_LADDER", "USD_SCALE_LADDER",
    "JP_UNIT_BY_SCALE", "USD_UNIT_BY_SCALE",
    "FMT_JP_BY_SCALE", "FMT_USD_BY_SCALE", "fmt_jp_for_scale",
    # Deterministic scale pickers
    "SCALE_HYSTERESIS_FACTOR",
    "pick_scale_for_magnitude", "pick_sheet_scale", "pick_row_scale",
    # v2 header contract (rulers / freeze / check echo)
    "HEADER_TITLE_ROW_V2", "HEADER_PURPOSE_ROW_V2", "HEADER_FY_RULER_ROW",
    "HEADER_MONTHS_RULER_ROW", "HEADER_PERIOD_ROW", "HEADER_ROWS_V2",
    "DATA_START_ROW_V2", "HEADER_CAPTION_COL",
    "write_period_rulers", "apply_freeze_pane", "write_master_check_echo",
    "audit_period_column_widths",
    "period_axis_columns",
    "period_axis_anchor_col",
    # v2 column width tokens
    "COL_GUTTER_WIDTH_V2", "COL_LABEL_WIDTH_V2", "COL_SOURCE_WIDTH_INPUT",
    "COL_DRIVER_TAG_WIDTH", "COL_UNIT_WIDTH_V2", "COL_PERIOD_WIDTH_V2",
    "COL_NOTE_WIDTH_V2",
    # Sheet naming
    "CANONICAL_SHEET_ORDER", "OPTIONAL_SHEETS", "validate_sheet_naming",
    "SHEET_ROLE_MAPPING",
]


# ============================================================================
# 14. Smoke test (run via `python ib_format.py`)
# ============================================================================
# Intended only as a self-test. Builders import named members; this block is
# not executed at import time.

if __name__ == "__main__":
    from openpyxl import Workbook as _Wb
    wb = _Wb()
    ws = wb.active
    ws.title = "Guide"
    apply_hard_input(ws["B5"], FMT_USD_MILLION)
    apply_formula(ws["B6"], FMT_PERCENT)
    apply_link_intra(ws["B7"], FMT_JPY_MILLION)
    apply_link_external(ws["B8"])
    apply_subtotal(ws["B9"], FMT_MONEY)
    apply_grand_total(ws["B10"], FMT_MONEY)
    set_tab_color(ws, "cover")
    assert ws["B5"].number_format == FMT_USD_MILLION
    assert ws["B6"].number_format == FMT_PERCENT
    assert ws.sheet_properties.tabColor.rgb.endswith(BRAND_NAVY)
    assert validate_sheet_naming(CANONICAL_SHEET_ORDER) == []

    print("ib_format.py smoke test: PASS")
