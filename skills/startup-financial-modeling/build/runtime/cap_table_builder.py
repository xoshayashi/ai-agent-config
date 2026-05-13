"""
cap_table_builder.py — Cap Table state machine + Ownership sheet generation

Source of truth:
  - 04a §19 (Round Event State Machine, 6-step pipeline)
  - 04b §1 (FDSO / TSM)
  - 04b §2 (Pre/Post-money + Option Pool Shuffle closed-form)
  - 04b §4 (Anti-Dilution: Broad-WA / Narrow-WA / Full Ratchet)
  - 04b §5 (SAFE / J-KISS conversion + MFN)
  - 04b §6 (Exit Waterfall)
  - 04b §10 (numerical worked examples)
  - 04b §12 (Boundary conditions)

Design principles:
  - AD = ratio adjustment only (NVCA standard, C-C-021): PreferredStockClass.shares
    is immutable; conversion_ratio updates feed only effective FDSO calc.
  - Post-money SAFEs lock α_i = INV_i / cap_i and do NOT dilute each other.
  - Pool shuffle: new pool sits in pre-money (NVCA standard, dilutes founders only).
  - State machine canonical order: Snapshot → SAFE → AD → Pool → New shares → FDSO.
  - All money inputs in millions of reporting currency (¥M default).
"""

from __future__ import annotations

import argparse
import math
import os
import sys
from dataclasses import dataclass, field
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import ib_format as ib

# ============================================================================
# 1. Dataclasses
# ============================================================================


SAFE_TYPE = Literal["safe_pre", "safe_post", "j_kiss_v1", "j_kiss_v2", "convertible_note"]
AD_MODE = Literal["none", "broad_wa", "narrow_wa", "full_ratchet"]
CAP_TABLE_SHEET_NAME = "Ownership"


@dataclass
class SAFEInstrument:
    """SAFE / J-KISS / Convertible Note instrument.

    canonical = 04a §1-3, _terminology §4-5

    Attributes:
        name:           Display name, e.g., "J-KISS-1".
        type:           Instrument family.
                        - safe_pre: Pre-money cap SAFE (rare today)
                        - safe_post: YC 2018+ standard, post-money cap, locks α
                        - j_kiss_v1: Coral Capital v1 (pre-money cap)
                        - j_kiss_v2: J-KISS 2.0 (post-money cap, MFN clause-bundle)
                        - convertible_note: Has interest_rate, principal+interest convert
        principal_money_m:  Investment face amount in money_m (¥M / $M).
        cap_money_m:    Valuation cap in money_m. None = no cap (discount-only).
        discount:       Decimal, 0.20 = 20% off next round PPS. Always < 1.0.
        interest_rate:  Annual rate for convertible_note only (e.g., 0.06 = 6%).
        issuance_date:  ISO-ish "YYYY-MM" (free-form for display).
        mfn:            Most Favored Nation flag (clause-bundle adoption per
                        04a §14.1 / 04b §5.5).
    """

    name: str
    type: SAFE_TYPE
    principal_money_m: float
    cap_money_m: float | None
    discount: float = 0.0
    interest_rate: float = 0.0
    issuance_date: str = ""
    mfn: bool = False


@dataclass
class PreferredStockClass:
    """Existing preferred stock series.

    canonical = 04a §4-9, 04b §6

    Attributes:
        name:                      "Series A" / "Series B" 等.
        shares:                    発行済株数 (immutable; AD は ratio で表現).
        issue_price:               発行時 PPS (¥ / share).
        liquidation_pref:          1.0x default (multiple of original investment).
        participating:             True = full participating; cap で打ち止め可能.
        participation_cap:         Multiples of investment to cap participation
                                   (None = uncapped / non-participating).
        anti_dilution:             AD mode (default broad_wa, NVCA std).
        cumulative_dividend_pct:   Accrued dividend rate (e.g., 0.06 = 6%).
        conversion_ratio:          Current ratio (1.0 default; AD updates this).
        senior_rank:               LP stack ordering — higher = paid first.
                                   0 = pari passu default. Common (no LP) is rank -1.
    """

    name: str
    shares: float
    issue_price: float
    liquidation_pref: float = 1.0
    participating: bool = False
    participation_cap: float | None = None
    anti_dilution: AD_MODE = "broad_wa"
    cumulative_dividend_pct: float = 0.0
    conversion_ratio: float = 1.0
    senior_rank: int = 0

    @property
    def investment_money_m(self) -> float:
        """Original invested capital (shares × issue_price) in money_m units.

        Issue_price is ¥/share so result is in ¥; divide by 1e6 for ¥M.
        """
        return (self.shares * self.issue_price) / 1_000_000.0

    @property
    def effective_shares(self) -> float:
        """Common-equivalent shares after AD ratio adjustment."""
        return self.shares * self.conversion_ratio


@dataclass
class CapTableInput:
    """Top-level input for cap table state machine + sheet build.

    canonical = 04a §19 + 04b §2

    Note on units:
        All money_m fields are in *millions* of reporting_currency.
        Share counts are absolute integers (e.g., 8_000_000 shares).
        Issue prices are ¥/share or $/share, NOT in money_m.
    """

    company_name: str
    reporting_currency: str = "JPY"
    as_of_date: str = ""

    # Pre-round state
    founder_shares: float = 8_000_000.0
    common_pool_issued: float = 400_000.0       # ESOP granted (vested or unvested)
    common_pool_available: float = 1_100_000.0  # ESOP unallocated
    safes: list[SAFEInstrument] = field(default_factory=list)
    preferred: list[PreferredStockClass] = field(default_factory=list)

    # Round-in-progress (None when only displaying current state)
    round_pre_money_money_m: float | None = None
    round_investment_money_m: float | None = None
    round_target_pool_pct: float | None = 0.12
    round_anti_dilution_default: AD_MODE = "broad_wa"
    round_label: str = "Priced Round"
    round_pool_in_pre_money: bool = True   # NVCA std True → founder bears pool dilution

    # Down-round override (force AD trigger checks even if PPS arithmetic borderline)
    is_down_round: bool = False


@dataclass
class SAFEConversionResult:
    """Per-SAFE conversion trace."""

    name: str
    type: str
    principal_money_m: float
    cap_money_m: float | None
    discount: float
    cap_based_price: float | None
    discount_based_price: float | None
    plain_price: float
    conversion_price: float
    shares_issued: float
    method: str   # "cap" | "discount" | "plain"
    mfn_applied: bool = False


@dataclass
class ADAdjustment:
    """Per-series anti-dilution trace."""

    series: str
    triggered: bool
    mode: str
    old_conversion_price: float
    new_conversion_price: float
    old_ratio: float
    new_ratio: float
    note: str = ""


@dataclass
class RoundResult:
    """Output of run_round_state_machine.

    Provides full 6-step trace + final FDSO breakdown, plus closed-form check
    flag for sanity testing.
    """

    pre_round_snapshot: dict[str, float]
    safe_conversions: list[SAFEConversionResult]
    ad_adjustments: list[ADAdjustment]
    option_pool_topup_shares: float
    new_round_shares: float
    new_round_price: float
    fdso_post: float
    fdso_pre: float
    fdso_breakdown: dict[str, float]
    founder_pct_pre: float
    founder_pct: float
    iteration_count: int
    closed_form_verified: bool
    notes: list[str] = field(default_factory=list)


# ============================================================================
# 2. Helpers
# ============================================================================


def _ppm(v: float) -> str:
    """Format raw number for trace dicts (debug-friendly)."""
    return f"{v:,.4f}"


def _is_post_money_cap(s: SAFEInstrument) -> bool:
    """Post-money cap SAFEs lock α = INV/cap (YC 2018+ + J-KISS 2.0)."""
    return s.type in ("safe_post", "j_kiss_v2")


def _resolve_mfn_cascade(safes: list[SAFEInstrument]) -> list[SAFEInstrument]:
    """Apply MFN clause-bundle adoption per 04a §19.1 / 04b §5.5.

    MFN-eligible SAFEs adopt **the clause-best terms across all SAFEs**:
        - cap = min(caps)              (lower cap is more investor-friendly)
        - discount = max(discounts)    (higher discount is more investor-friendly)

    NVCA / J-KISS 2.0 standard treats this as a 4-axis independent cascade.
    Returns a new list with adopted terms applied; original list is untouched.
    """
    if not any(s.mfn for s in safes):
        return [SAFEInstrument(**s.__dict__) for s in safes]

    caps_present = [s.cap_money_m for s in safes if s.cap_money_m is not None]
    best_cap = min(caps_present) if caps_present else None
    best_discount = max((s.discount for s in safes), default=0.0)

    adopted: list[SAFEInstrument] = []
    for s in safes:
        copy = SAFEInstrument(**s.__dict__)
        if copy.mfn:
            if best_cap is not None and (copy.cap_money_m is None or best_cap < copy.cap_money_m):
                copy.cap_money_m = best_cap
            if best_discount > copy.discount:
                copy.discount = best_discount
        adopted.append(copy)
    return adopted


# ============================================================================
# 3. Option Pool Shuffle (closed-form)
# ============================================================================


def option_pool_shuffle_closed_form(
    *,
    pre_existing_fdso: float,
    target_pool_pct: float,
    investment_money_m: float,
    pre_money_money_m: float,
    pool_already_available: float = 0.0,
    pool_already_granted: float = 0.0,
) -> dict:
    """04b §2.3 closed-form Option Pool Shuffle solver.

    Solves the simultaneous system:
        X = F0 + Y + Z       (post-money FDSO)
        Y + P0 = T × X       (post-money pool target)
        P = PMV / (F0 + Y)   (pre-money PPS basis)
        Z = INV / P

    Closed form:
        X = (F0 - P0) × QMV / (PMV - T × QMV)
        Y = T × X - P0_total
        Z = X - F0 - Y
        PPS = PMV / (F0 + Y) = INV / Z

    Args:
        pre_existing_fdso:        F0 — existing FDSO **excluding the pool already
                                  granted/available** (this is the canonical
                                  definition in §2.3 worked example).
                                  In our impl we treat F0 = founders + preferred AC
                                  + SAFE conversions (post Step 2/3) and pool is
                                  isolated via P0 + G0.
        target_pool_pct:          T — post-money pool fraction target (e.g., 0.10).
        investment_money_m:       INV — new round raise.
        pre_money_money_m:        PMV — pre-money valuation.
        pool_already_available:   P0 unallocated pool stock.
        pool_already_granted:     G0 granted/issued pool stock (counts toward target).

    Returns:
        dict with X (post-money FDSO), Y (delta_pool top-up shares),
        Z (new round shares), pps, founder/investor/pool % of post-money,
        and intermediate diagnostics.
    """
    if target_pool_pct < 0 or target_pool_pct >= 1.0:
        raise ValueError(f"target_pool_pct must be in [0, 1): {target_pool_pct}")
    if pre_money_money_m <= 0:
        raise ValueError(f"pre_money_money_m must be positive: {pre_money_money_m}")
    if pre_existing_fdso <= 0:
        raise ValueError(f"pre_existing_fdso must be positive: {pre_existing_fdso}")

    F0 = float(pre_existing_fdso)
    P0_total = float(pool_already_available + pool_already_granted)
    PMV = float(pre_money_money_m)
    INV = float(investment_money_m)
    QMV = PMV + INV
    T = float(target_pool_pct)

    denom = PMV - T * QMV
    if denom <= 0:
        raise ValueError(
            f"Pool target ({T:.1%}) too large vs pre-money fraction "
            f"(PMV={PMV}, QMV={QMV}, denom={denom}); reduce target or raise pre-money."
        )

    X = (F0 - P0_total) * QMV / denom              # Post-money FDSO
    Y = T * X - P0_total                            # Pool top-up (delta)
    Z = X - F0 - Y                                  # New round shares
    if Y < 0:
        # Existing pool already exceeds target → no top-up needed (clip to 0)
        Y = 0.0
        # Recompute X with clipped Y
        # Z still satisfies Z = INV / PPS where PPS = PMV / (F0 + Y) = PMV / F0
        pps = PMV / F0
        Z = INV / pps
        X = F0 + Y + Z

    pps = PMV / (F0 + Y) if (F0 + Y) > 0 else 0.0

    return {
        "X_post_fdso": X,
        "Y_pool_topup": Y,
        "Z_new_shares": Z,
        "pps": pps,
        "F0": F0,
        "P0_total": P0_total,
        "PMV": PMV,
        "QMV": QMV,
        "T": T,
        "founder_share_of_post": (F0 - P0_total) / X if X > 0 else 0.0,
        "investor_share_of_post": Z / X if X > 0 else 0.0,
        "pool_share_of_post": (Y + P0_total) / X if X > 0 else 0.0,
    }


# ============================================================================
# 4. SAFE / J-KISS conversion
# ============================================================================


def compute_safe_conversion(
    safe: SAFEInstrument,
    *,
    next_round_pps: float,
    pre_round_fdso: float,
    next_round_post_money_money_m: float | None = None,
) -> SAFEConversionResult:
    """Compute conversion price and shares for a single SAFE / J-KISS / Note.

    canonical = 04b §5.2 (pre-money) / §5.3 (post-money) / 04a §19.1 step 2

    Conversion price = MIN over candidates:
        cap_based:        cap / FDSO_basis
                          - pre-money SAFE / j_kiss_v1: FDSO_basis = pre_round_fdso
                            (excludes other SAFE conversion shares per §5.2)
                          - post-money SAFE / j_kiss_v2: cap is post-money;
                            α_i = INV/cap is locked. We compute cap_based_price as
                            INV / (α_i × X) but we deliberately use the simpler
                            equivalent: cap_money / "company capitalization".
                            For the State Machine, post-money SAFE shares are
                            determined directly from α_i × FDSO_post (computed at
                            Step 5 once X is known). compute_safe_conversion gives
                            an estimate based on pre-existing FDSO; final
                            reconciliation happens in run_round_state_machine.
        discount_based:   next_round_pps × (1 - discount)
        plain:            next_round_pps  (always available as ceiling)

    Args:
        safe:                       SAFE instrument.
        next_round_pps:             Estimated next-round priced PPS (¥/share).
        pre_round_fdso:             FDSO at which cap_based price is computed.
                                    For pre-money SAFE = founders + preferred AC
                                    + pool. For post-money SAFE we still use this
                                    as a sanity reference but the locked-α path is
                                    primary in the state machine.
        next_round_post_money_money_m: Post-money valuation for post-money SAFE
                                    α-locking.

    Returns:
        SAFEConversionResult with full price candidate trace.
    """
    if safe.principal_money_m <= 0:
        raise ValueError(f"SAFE principal must be positive: {safe.principal_money_m}")
    if next_round_pps <= 0:
        raise ValueError(f"next_round_pps must be positive: {next_round_pps}")

    # ------------------------------------------------------------------
    # Candidate prices
    # ------------------------------------------------------------------
    plain_price = next_round_pps
    discount_based_price = (
        next_round_pps * (1.0 - safe.discount) if safe.discount > 0 else None
    )

    cap_based_price: float | None = None
    if safe.cap_money_m is not None and safe.cap_money_m > 0:
        if _is_post_money_cap(safe):
            # Post-money cap (YC 2018+ / J-KISS 2.0): cap_based price is computed
            # such that α_i = INV/cap. We need the post-money FDSO, but at this
            # stage we approximate using pre_round_fdso as a fallback. The State
            # Machine reconciles this in Step 5 by directly setting
            # safe_shares = (INV / cap) × FDSO_post.
            #
            # For the per-instrument trace we still emit cap_based_price as
            # cap_money_yen / pre_round_fdso (= INV/cap × pre_round_fdso ratio).
            cap_money_yen = safe.cap_money_m * 1_000_000.0
            # If we have a post_money estimate, use it; else fall back to pre.
            basis = pre_round_fdso
            cap_based_price = cap_money_yen / basis if basis > 0 else None
        else:
            cap_money_yen = safe.cap_money_m * 1_000_000.0
            basis = pre_round_fdso
            cap_based_price = cap_money_yen / basis if basis > 0 else None

    # Effective principal incorporates accrued interest for convertible notes.
    # (We use a 0-year accrual placeholder — caller may pre-compute interest into
    # principal_money_m; or supply via interest_rate for a 1y default.)
    effective_principal = safe.principal_money_m
    if safe.type == "convertible_note" and safe.interest_rate > 0:
        # Default 1-year accrual approximation; refine via SaaS / corporate timing.
        effective_principal *= (1.0 + safe.interest_rate)

    # ------------------------------------------------------------------
    # Choose the lowest of available prices (best for SAFE holder)
    # ------------------------------------------------------------------
    candidates: list[tuple[float, str]] = [(plain_price, "plain")]
    if cap_based_price is not None:
        candidates.append((cap_based_price, "cap"))
    if discount_based_price is not None:
        candidates.append((discount_based_price, "discount"))

    chosen_price, method = min(candidates, key=lambda t: t[0])
    principal_yen = effective_principal * 1_000_000.0
    shares_issued = principal_yen / chosen_price if chosen_price > 0 else 0.0

    return SAFEConversionResult(
        name=safe.name,
        type=safe.type,
        principal_money_m=safe.principal_money_m,
        cap_money_m=safe.cap_money_m,
        discount=safe.discount,
        cap_based_price=cap_based_price,
        discount_based_price=discount_based_price,
        plain_price=plain_price,
        conversion_price=chosen_price,
        shares_issued=shares_issued,
        method=method,
    )


# ============================================================================
# 5. Anti-Dilution
# ============================================================================


def apply_anti_dilution(
    pref: PreferredStockClass,
    *,
    new_issue_price: float,
    new_shares_issued: float,
    pre_existing_shares_broad: float,
    pre_existing_shares_narrow: float | None = None,
    investment_money_m: float = 0.0,
) -> tuple[float, float]:
    """Compute updated conversion price + ratio for a preferred series.

    canonical = 04b §4.5 / 04a §19.1 step 3

    Modes:
        none        — no adjustment.
        full_ratchet — NCP = new_issue_price (PPS_new).
        broad_wa    — NCP = OCP × (A_broad + B_eff) / (A_broad + C)
                      A_broad = full FDSO incl. pool, granted, all preferred AC,
                      and concurrent SAFE conversions.
        narrow_wa   — NCP = OCP × (A_narrow + B_eff) / (A_narrow + C)
                      A_narrow excludes unallocated pool / unissued options.

    where:
        OCP = pref.issue_price × pref.conversion_ratio (current effective CP,
              not original; so successive AD events compound correctly).
        B_eff = INV_yen / OCP   (theoretical "as-if-at-OCP" shares)
        C     = new_shares_issued  (actual new round shares)
        Trigger: new_issue_price < OCP (per-series check, C-C-019)

    Args:
        pref:                        Preferred series.
        new_issue_price:             Round X PPS.
        new_shares_issued:           Round X newly issued priced shares (C).
        pre_existing_shares_broad:   A_broad (incl. pool, all preferred AC, SAFE
                                     conversions concurrent at this round).
        pre_existing_shares_narrow:  A_narrow (excl. pool); defaults to broad if
                                     None (caller may pass narrow explicitly).
        investment_money_m:          INV in money_m (for B_eff calc).

    Returns:
        (new_conversion_price, new_conversion_ratio)
    """
    OCP_current = pref.issue_price * pref.conversion_ratio
    if pref.anti_dilution == "none":
        return OCP_current, pref.conversion_ratio

    # Per-series down-round trigger (C-C-019): only down rounds (PPS < current CP)
    # adjust this series.
    if new_issue_price >= OCP_current:
        return OCP_current, pref.conversion_ratio

    # Investment denominated in yen for B_eff computation (matches OCP unit ¥/share)
    INV_yen = investment_money_m * 1_000_000.0
    B_eff = INV_yen / OCP_current if OCP_current > 0 else 0.0
    C = new_shares_issued

    if pref.anti_dilution == "full_ratchet":
        new_cp = max(new_issue_price, 1e-9)
    elif pref.anti_dilution == "broad_wa":
        A = pre_existing_shares_broad
        if A + C < 1e-9:
            return OCP_current, pref.conversion_ratio
        new_cp = OCP_current * (A + B_eff) / (A + C)
    elif pref.anti_dilution == "narrow_wa":
        A_narrow = (
            pre_existing_shares_narrow
            if pre_existing_shares_narrow is not None
            else pre_existing_shares_broad
        )
        if A_narrow + C < 1e-9:
            return OCP_current, pref.conversion_ratio
        new_cp = OCP_current * (A_narrow + B_eff) / (A_narrow + C)
    else:
        return OCP_current, pref.conversion_ratio

    # AD never moves CP UP (cap at current CP)
    new_cp = min(new_cp, OCP_current)
    # New conversion ratio expressed against ORIGINAL issue_price so successive
    # adjustments compound transparently:
    new_ratio = pref.issue_price / new_cp if new_cp > 0 else pref.conversion_ratio
    return new_cp, new_ratio


# ============================================================================
# 6. Round Event State Machine
# ============================================================================


def run_round_state_machine(inp: CapTableInput) -> RoundResult:
    """Execute 04a §19.1 6-step pipeline.

    Steps:
        1. Snapshot pre-round cap table.
        2. Resolve pending SAFE / J-KISS conversions.
           - MFN cascade first (04a §19.1 + 04b §5.5).
           - Per-instrument cap/discount/plain min.
           - Post-money SAFEs: lock α = INV/cap; reconciled in Step 5.
        3. Apply anti-dilution to existing preferred (per-series trigger).
        4. Calculate option pool top-up (post-money basis, NVCA pre-money pool).
        5. Issue new round priced shares (closed-form simultaneous with Step 4).
        6. Recalculate FDSO and reconcile.

    Returns:
        RoundResult — full trace + closed-form verification flag.
    """
    if inp.round_pre_money_money_m is None or inp.round_investment_money_m is None:
        raise ValueError(
            "run_round_state_machine requires round_pre_money_money_m + "
            "round_investment_money_m"
        )

    PMV = inp.round_pre_money_money_m
    INV = inp.round_investment_money_m
    QMV = PMV + INV
    T = inp.round_target_pool_pct or 0.0

    notes: list[str] = []

    # ------------------------------------------------------------------
    # Step 1: Snapshot
    # ------------------------------------------------------------------
    F0_founders = inp.founder_shares
    P0_avail = inp.common_pool_available
    G0 = inp.common_pool_issued
    pref_shares_AC = sum(p.shares * p.conversion_ratio for p in inp.preferred)

    # Snapshot FDSO_raw = founders + pool (avail+granted) + preferred as-converted
    FDSO_snapshot = F0_founders + P0_avail + G0 + pref_shares_AC
    snapshot = {
        "founder_shares": F0_founders,
        "common_pool_available": P0_avail,
        "common_pool_issued": G0,
        "preferred_AC_shares": pref_shares_AC,
        "fdso_snapshot": FDSO_snapshot,
    }
    founder_pct_pre = F0_founders / FDSO_snapshot if FDSO_snapshot > 0 else 0.0

    # ------------------------------------------------------------------
    # Step 2: SAFE / J-KISS conversions
    # ------------------------------------------------------------------
    # Apply MFN cascade
    cascaded_safes = _resolve_mfn_cascade(inp.safes)
    mfn_changed = [
        cascaded_safes[i].name
        for i, s in enumerate(inp.safes)
        if (
            s.mfn
            and (
                cascaded_safes[i].cap_money_m != s.cap_money_m
                or cascaded_safes[i].discount != s.discount
            )
        )
    ]
    if mfn_changed:
        notes.append(f"MFN cascade applied to: {', '.join(mfn_changed)}")

    # Initial PPS estimate for SAFE pricing (will refine via iteration)
    # Use simple: PPS_seed = PMV_yen / FDSO_snapshot
    PPS_seed = (PMV * 1_000_000.0) / FDSO_snapshot if FDSO_snapshot > 0 else 0.0

    # Locked α_i for post-money SAFE/J-KISS 2.0 — these directly target a fixed
    # share of FDSO_post and don't dilute one another.
    alpha_locked: dict[str, float] = {}
    for s in cascaded_safes:
        if _is_post_money_cap(s) and s.cap_money_m and s.cap_money_m > 0:
            alpha_locked[s.name] = s.principal_money_m / s.cap_money_m

    # Compute per-SAFE conversion at PPS_seed for trace; final shares for
    # post-money SAFEs are reconciled at Step 5/6 via α-locking.
    safe_conversions: list[SAFEConversionResult] = []
    for s in cascaded_safes:
        sc = compute_safe_conversion(
            s,
            next_round_pps=PPS_seed,
            pre_round_fdso=FDSO_snapshot,
            next_round_post_money_money_m=QMV,
        )
        if s.mfn:
            sc.mfn_applied = s.name in mfn_changed
        safe_conversions.append(sc)

    # ------------------------------------------------------------------
    # Iteration to converge: SAFE conversion shares depend on PPS, which depends
    # on pool top-up, which depends on FDSO_post which contains SAFE shares.
    # Approach: closed-form per iteration, with α_locked SAFEs treated as
    # fractions of FDSO_post. Converge on FDSO_post.
    # ------------------------------------------------------------------
    # Sum α for post-money SAFEs (locked fraction of FDSO_post)
    alpha_post_sum = sum(alpha_locked.values())

    # For pre-money SAFEs, shares are independent of FDSO_post once PPS converges.
    # We solve via fixed-point iteration on FDSO_post.
    iteration_count = 0
    max_iter = 50
    tol = 1e-3

    # Initialize FDSO_post with simple estimate
    X_post = FDSO_snapshot * (1.0 + INV / PMV) * (1.0 + T)
    if X_post <= 0:
        X_post = FDSO_snapshot

    pre_money_safe_shares_total = 0.0
    converged = False
    for _ in range(max_iter):
        iteration_count += 1
        # Implied PPS at current X_post (NVCA pool-in-pre-money model):
        # X_post = (F_total + pool_target_topup + N_round) where pool_target = T × X_post
        # Pre-money portion = X_post − N_round
        # PPS = PMV_yen / (X_post − N_round) but N_round = INV_yen / PPS → PPS solved
        # implicitly. Easier: start from current X_post and recompute:
        #   N_round_alpha   = INV / cap-implied price of priced round
        #   Investor α_inv = INV / QMV
        # Closed-form (NVCA pool-in-pre-money):
        #   Investor α      = INV / QMV
        #   Pool α          = T  (post-money)
        #   Sum locked      = alpha_post_sum
        #   Pre-money SAFEs are explicit shares (computed at current PPS estimate)
        # Founders + pre-money preferred AC + pool granted + pre-money SAFE shares
        # plus post-money SAFE α + pool top-up + investor must equal X_post.
        #
        # Strategy: compute pre-money SAFE shares first at PPS_iter
        PPS_iter = PMV * 1_000_000.0 / max(FDSO_snapshot + pre_money_safe_shares_total, 1.0)

        # Recompute pre-money SAFE conversions at current PPS_iter
        pre_money_safe_shares_total_new = 0.0
        for i, s in enumerate(cascaded_safes):
            if _is_post_money_cap(s):
                continue  # post-money handled via α-lock
            sc = compute_safe_conversion(
                s,
                next_round_pps=PPS_iter,
                pre_round_fdso=FDSO_snapshot,
                next_round_post_money_money_m=QMV,
            )
            safe_conversions[i] = sc
            pre_money_safe_shares_total_new += sc.shares_issued

        if abs(pre_money_safe_shares_total_new - pre_money_safe_shares_total) < tol:
            pre_money_safe_shares_total = pre_money_safe_shares_total_new
            converged = True
            break
        pre_money_safe_shares_total = pre_money_safe_shares_total_new

    # ------------------------------------------------------------------
    # Step 3: Anti-Dilution
    # ------------------------------------------------------------------
    # A_broad includes: founders + pool (avail + granted) + preferred AC + SAFE
    # converted shares (pre + post).
    # For post-money SAFEs we estimate shares_at_AD ≈ α × FDSO_snapshot_with_pre
    # (provisional; the AD effect on the closed form is small relative to the
    # snapshot scale; full reconciliation happens via the iteration).
    pre_money_safe_shares = sum(
        sc.shares_issued
        for sc, s in zip(safe_conversions, cascaded_safes)
        if not _is_post_money_cap(s)
    )
    # provisional FDSO including pre-money SAFE conversion
    A_broad_provisional = FDSO_snapshot + pre_money_safe_shares
    # add provisional post-money SAFE shares (alpha_post_sum × QMV/PMV × ratio)
    if alpha_post_sum > 0 and (1 - alpha_post_sum) > 0:
        provisional_post_safe = (
            alpha_post_sum * A_broad_provisional / (1.0 - alpha_post_sum)
        )
        A_broad_provisional += provisional_post_safe

    PPS_round_estimate = PMV * 1_000_000.0 / max(
        FDSO_snapshot + pre_money_safe_shares, 1.0
    )
    # Use a round estimate of new shares for AD's C term
    new_shares_estimate = (INV * 1_000_000.0) / PPS_round_estimate
    A_narrow_provisional = (
        F0_founders
        + G0
        + pref_shares_AC
        + pre_money_safe_shares
    )  # excludes available pool

    ad_adjustments: list[ADAdjustment] = []
    for p in inp.preferred:
        OCP_old = p.issue_price * p.conversion_ratio
        ratio_old = p.conversion_ratio
        new_cp, new_ratio = apply_anti_dilution(
            p,
            new_issue_price=PPS_round_estimate,
            new_shares_issued=new_shares_estimate,
            pre_existing_shares_broad=A_broad_provisional,
            pre_existing_shares_narrow=A_narrow_provisional,
            investment_money_m=INV,
        )
        triggered = abs(new_ratio - ratio_old) > 1e-9
        ad_adjustments.append(
            ADAdjustment(
                series=p.name,
                triggered=triggered,
                mode=p.anti_dilution,
                old_conversion_price=OCP_old,
                new_conversion_price=new_cp,
                old_ratio=ratio_old,
                new_ratio=new_ratio,
                note=("triggered" if triggered else f"up round vs OCP {OCP_old:.2f}"),
            )
        )
        # Update preferred class in place (state machine produces new state)
        p.conversion_ratio = new_ratio

    # Recompute preferred AC shares post-AD
    pref_shares_AC_post = sum(p.shares * p.conversion_ratio for p in inp.preferred)
    delta_pref = pref_shares_AC_post - pref_shares_AC

    # ------------------------------------------------------------------
    # Step 4 + 5: Pool top-up and new round shares (closed form)
    # ------------------------------------------------------------------
    # NVCA pool-in-pre-money convention:
    #   Post-money FDSO = F0' + P0_avail + G0 + ΔP + N_round + ΣSAFE_shares
    #   where F0' = founders + preferred AC (post-AD) + pre-money SAFE shares
    #
    # Investor target α = INV / QMV
    # Pool target α (post-money) = T
    # Post-money SAFE locked α_sum = alpha_post_sum
    #
    # Founders + preferred AC + pre-money SAFEs + granted pool occupy the rest:
    #   α_residual = 1 - α_inv - T - α_post_sum
    #
    # Then X_post = (F0_residual_shares) / α_residual
    # where F0_residual_shares = F0_founders + pref_shares_AC_post + pre_money_safe_shares
    #                          + G0   (granted pool)
    # The granted pool counts toward pool target T, not residual; so split:
    #
    # Actually NVCA convention puts ALL pool (granted + available + delta) at
    # T × X_post. So:
    #   F0_other = F0_founders + pref_shares_AC_post + pre_money_safe_shares
    #   Pool_total = T × X_post
    #   Investor = INV/QMV × X_post
    #   PostSafe = α_post_sum × X_post
    #   F0_other = (1 - T - INV/QMV - α_post_sum) × X_post
    #   X_post = F0_other / (1 - T - INV/QMV - α_post_sum)
    #
    # This is closed-form once F0_other is known.
    F0_other = F0_founders + pref_shares_AC_post + pre_money_safe_shares
    alpha_inv = INV / QMV
    alpha_residual = 1.0 - T - alpha_inv - alpha_post_sum
    if alpha_residual <= 0:
        raise ValueError(
            f"Residual α ≤ 0 (T={T}, α_inv={alpha_inv:.4f}, "
            f"α_post_safe={alpha_post_sum:.4f}); reduce pool/SAFE total."
        )
    X_post = F0_other / alpha_residual

    pool_target_total = T * X_post
    delta_pool = max(pool_target_total - (P0_avail + G0), 0.0)
    new_round_shares = alpha_inv * X_post
    post_safe_shares_total = alpha_post_sum * X_post

    # PPS_round (after pool top-up included in pre-money convention)
    # From X_post = F0_other + (P0_avail + G0 + ΔP) + new_round + post_safe:
    #   denominator for PPS = F0_other + P0_avail + G0 + ΔP + post_safe
    #   PPS = PMV_yen / denominator    (NVCA pool-in-pre-money)
    # Equivalent: new_round = INV_yen / PPS  → PPS = INV_yen / new_round_shares
    PPS_round = (INV * 1_000_000.0) / new_round_shares if new_round_shares > 0 else 0.0

    # Distribute post-money SAFE shares per-instrument (proportional to α_i)
    for i, s in enumerate(cascaded_safes):
        if _is_post_money_cap(s) and s.cap_money_m and s.cap_money_m > 0:
            alpha_i = s.principal_money_m / s.cap_money_m
            shares_i = alpha_i * X_post
            sc_old = safe_conversions[i]
            # Update conversion result with reconciled shares
            cap_yen = s.cap_money_m * 1_000_000.0
            cp_eff = cap_yen / X_post  # cap / FDSO_post
            disc_price = (
                PPS_round * (1.0 - s.discount) if s.discount > 0 else None
            )
            # Method = whichever is lower (cap-based vs discount)
            method = "cap"
            chosen_price = cp_eff
            if disc_price is not None and disc_price < cp_eff:
                method = "discount"
                chosen_price = disc_price
                shares_i = (s.principal_money_m * 1_000_000.0) / chosen_price
            safe_conversions[i] = SAFEConversionResult(
                name=s.name,
                type=s.type,
                principal_money_m=s.principal_money_m,
                cap_money_m=s.cap_money_m,
                discount=s.discount,
                cap_based_price=cp_eff,
                discount_based_price=disc_price,
                plain_price=PPS_round,
                conversion_price=chosen_price,
                shares_issued=shares_i,
                method=method,
                mfn_applied=sc_old.mfn_applied,
            )

    # Update pre-money SAFE shares at the converged PPS_round
    for i, s in enumerate(cascaded_safes):
        if _is_post_money_cap(s):
            continue
        sc = compute_safe_conversion(
            s,
            next_round_pps=PPS_round,
            pre_round_fdso=FDSO_snapshot,
            next_round_post_money_money_m=QMV,
        )
        safe_conversions[i] = sc

    # Recalculate total SAFE shares after Step 5 reconciliation
    safe_shares_total = sum(sc.shares_issued for sc in safe_conversions)

    # ------------------------------------------------------------------
    # Step 6: FDSO reconciliation
    # ------------------------------------------------------------------
    fdso_post_actual = (
        F0_founders
        + P0_avail
        + G0
        + delta_pool
        + pref_shares_AC_post
        + safe_shares_total
        + new_round_shares
    )

    fdso_breakdown = {
        "founder_shares": F0_founders,
        "common_pool_available_pre": P0_avail,
        "common_pool_issued": G0,
        "common_pool_topup_new": delta_pool,
        "preferred_AC_shares_post_AD": pref_shares_AC_post,
        "safe_converted_shares": safe_shares_total,
        "new_round_shares": new_round_shares,
        "fdso_post": fdso_post_actual,
    }

    founder_pct = F0_founders / fdso_post_actual if fdso_post_actual > 0 else 0.0

    # ------------------------------------------------------------------
    # Closed-form verification: Σ% should be 100%, and X_post should match
    # the closed-form derivation within ε.
    # ------------------------------------------------------------------
    sum_pct = sum(
        v / fdso_post_actual
        for k, v in fdso_breakdown.items()
        if k != "fdso_post"
    )
    closed_form_verified = (
        abs(fdso_post_actual - X_post) / max(X_post, 1.0) < 0.001
        and abs(sum_pct - 1.0) < 1e-6
    )

    if not closed_form_verified:
        notes.append(
            f"closed_form_verification: |FDSO_actual - X_closed| = "
            f"{abs(fdso_post_actual - X_post):.2f}, Σ% = {sum_pct:.6f}"
        )

    return RoundResult(
        pre_round_snapshot=snapshot,
        safe_conversions=safe_conversions,
        ad_adjustments=ad_adjustments,
        option_pool_topup_shares=delta_pool,
        new_round_shares=new_round_shares,
        new_round_price=PPS_round,
        fdso_post=fdso_post_actual,
        fdso_pre=FDSO_snapshot,
        fdso_breakdown=fdso_breakdown,
        founder_pct_pre=founder_pct_pre,
        founder_pct=founder_pct,
        iteration_count=iteration_count,
        closed_form_verified=closed_form_verified,
        notes=notes,
    )


# ============================================================================
# 7. Exit Waterfall
# ============================================================================


def compute_exit_waterfall(
    *,
    inp: CapTableInput,
    exit_value_money_m: float,
    debt_outstanding_money_m: float = 0.0,
    so_strike_avg: float = 0.0,
    so_outstanding_shares: float = 0.0,
    cancel_unallocated_pool: bool = True,
) -> dict:
    """Exit waterfall per 04b §6.

    Pipeline:
        1. Senior debt repaid first.
        2. Build LP stack ordered by senior_rank (high→low; pari passu shares pool).
        3. For each non-participating series: choose max(LP, as-converted residual).
           Convert decisions cascade — once a series converts, its LP releases
           back into the residual pool.
        4. For participating series: take LP, then participate pro-rata in residual.
        5. Common stock + ITM option holders share the residual.
        6. Cap participating series at participation_cap × investment if set.

    Pool treatment at exit (04b §6.6):
        - In-the-money issued options (G0): exercise + participate as common
          (handled via so_outstanding_shares / so_strike_avg in caller).
        - Out-of-the-money / unallocated pool (P0): cancel per §6.6.2 default.
          When cancel_unallocated_pool=True, inp.common_pool_available is dropped
          from the FDSO denominator at exit (canonical §10.3.1: "Pool 未割当残:
          2,050,497 株（cancel される）"; founder net @ ¥6B = ¥1,262.6M).

    Args:
        inp:                          CapTableInput with current preferred state.
        exit_value_money_m:           Total deal consideration.
        debt_outstanding_money_m:     Senior debt to be repaid first.
        so_strike_avg:                Avg strike for ITM option holders (¥/share).
        so_outstanding_shares:        Outstanding option shares (treated as common
                                      after exercise if ITM).
        cancel_unallocated_pool:      04b §6.6.2 default True — drop unallocated
                                      pool (common_pool_available) from FDSO at
                                      exit. Set False only for analytic scenarios
                                      where the buyer rolls the entire pool over.

    Returns:
        dict with debt_repayment, lp_payments, participation, common_payment,
        founder_net_cash, founder_pct_of_proceeds, crossover_value (per series).
    """
    if exit_value_money_m < 0:
        raise ValueError("Exit value must be non-negative")

    # ------------------------------------------------------------------
    # Step 0: Resolve cap table (effective shares)
    # ------------------------------------------------------------------
    F0 = inp.founder_shares
    # Pool at exit: granted (G0) is kept (vested options exercise via so_*),
    # available (P0) is cancelled per §6.6.2 unless explicitly disabled.
    pool_available_effective = (
        0.0 if cancel_unallocated_pool else inp.common_pool_available
    )
    pool_total = pool_available_effective + inp.common_pool_issued
    pref_list = list(inp.preferred)

    fdso_total = F0 + pool_total + sum(p.effective_shares for p in pref_list) + so_outstanding_shares

    # ------------------------------------------------------------------
    # Step 1: Debt repayment + ITM option exercise
    # ------------------------------------------------------------------
    after_debt = exit_value_money_m - debt_outstanding_money_m
    if after_debt < 0:
        # Below debt — no equity payout
        return {
            "debt_repayment": exit_value_money_m,
            "lp_payments": [],
            "participation_payments": [],
            "common_payment": 0.0,
            "founder_net_cash": 0.0,
            "founder_pct_of_proceeds": 0.0,
            "crossover_values": {},
            "fdso_total": fdso_total,
            "exit_pps_pre_choice": 0.0,
            "notes": ["Exit value insufficient to cover senior debt; no equity payout"],
        }

    # ITM SO exercise: cash inflow + new common shares
    exit_pps_estimate = after_debt * 1_000_000.0 / fdso_total if fdso_total > 0 else 0.0
    so_exercise_inflow_money_m = 0.0
    so_exercised_shares = 0.0
    if so_outstanding_shares > 0 and exit_pps_estimate > so_strike_avg:
        so_exercise_inflow_money_m = so_outstanding_shares * so_strike_avg / 1_000_000.0
        so_exercised_shares = so_outstanding_shares

    cash_pool = after_debt + so_exercise_inflow_money_m

    # ------------------------------------------------------------------
    # Step 2-4: LP + participation (with non-participating convert flip)
    # ------------------------------------------------------------------
    # For each non-participating series we evaluate convert vs LP.
    # Sort preferred by senior_rank (descending) for stack precedence.
    # Pari passu (equal rank) is handled by sharing LP cap proportionally if
    # cash_pool < Σ LP at that rank.
    crossover_values: dict[str, float] = {}
    convert_decision: dict[str, bool] = {}

    # Compute as-converted denominator for "convert" branch (excludes preferred
    # that elect LP; we resolve iteratively from senior down)
    # Simple resolution for typical cap stacks: sort by rank, evaluate non-part
    # series with binary convert/LP decision based on the all-pref-AC residual.
    # This matches §10.3.2 / §10.3.3 logic.

    # Identify all common-equivalent shares (for as-converted decisions of
    # non-participating series)
    common_equivalents = F0 + pool_total + so_exercised_shares
    pref_AC_total = sum(p.effective_shares for p in pref_list)
    fdso_for_convert = common_equivalents + pref_AC_total

    # For each non-participating series compute crossover
    for p in pref_list:
        if not p.participating:
            inv_yen = p.investment_money_m * 1_000_000.0
            lp_money_m = p.investment_money_m * p.liquidation_pref
            if fdso_for_convert > 0:
                alpha_p = p.effective_shares / fdso_for_convert
                # Cross-over: cash_pool such that alpha_p × cash_pool = LP_p
                crossover = lp_money_m / alpha_p if alpha_p > 0 else float("inf")
                crossover_values[p.name] = crossover
                convert_decision[p.name] = (alpha_p * cash_pool) > lp_money_m
            else:
                crossover_values[p.name] = float("inf")
                convert_decision[p.name] = False
        else:
            convert_decision[p.name] = False  # participating always takes LP+participate

    # Compute LP (only for series that did NOT convert)
    lp_payments: list[dict] = []
    lp_total = 0.0
    for p in pref_list:
        if convert_decision.get(p.name, False):
            lp_payments.append({
                "series": p.name,
                "lp_money_m": 0.0,
                "method": "converted_to_common",
            })
            continue
        lp_money_m = p.investment_money_m * p.liquidation_pref
        # Cap LP at cash_pool minus already-paid LP (senior priority)
        lp_payments.append({
            "series": p.name,
            "lp_money_m": lp_money_m,
            "method": "lp_taken",
        })
        lp_total += lp_money_m

    if lp_total > cash_pool:
        # LP stack exceeds cash → pari passu pro-rata distribution within rank
        # Simplification: scale all LP proportionally (assumes single rank or
        # all equally senior). For multi-rank, would iterate by rank.
        scale = cash_pool / lp_total if lp_total > 0 else 0.0
        for entry in lp_payments:
            entry["lp_money_m"] *= scale
        lp_total = cash_pool
        residual = 0.0
    else:
        residual = cash_pool - lp_total

    # ------------------------------------------------------------------
    # Step 5-6: Residual distribution
    # ------------------------------------------------------------------
    # Participating series + converted-non-participating series + common share
    # the residual on as-converted basis.
    participating_series = [p for p in pref_list if p.participating]
    converted_series = [
        p for p in pref_list
        if (not p.participating and convert_decision.get(p.name, False))
    ]

    # Residual denominator
    residual_shares = (
        F0 + pool_total + so_exercised_shares
        + sum(p.effective_shares for p in participating_series)
        + sum(p.effective_shares for p in converted_series)
    )

    participation_payments: list[dict] = []

    if residual_shares > 0:
        per_share_residual = residual / residual_shares
    else:
        per_share_residual = 0.0

    # Founder share
    founder_residual = F0 * per_share_residual
    pool_residual = pool_total * per_share_residual
    so_residual = so_exercised_shares * per_share_residual

    # Participating series: residual + cap check
    for p in participating_series:
        part = p.effective_shares * per_share_residual
        lp_for_p = next(
            (e["lp_money_m"] for e in lp_payments if e["series"] == p.name), 0.0
        )
        total_take = lp_for_p + part
        capped = False
        if p.participation_cap is not None:
            cap_amount = p.participation_cap * p.investment_money_m
            if total_take > cap_amount:
                # Excess returns to residual pool (recursive distribution)
                # Simplification: just clamp at cap; excess assumed redistributed.
                part = max(cap_amount - lp_for_p, 0.0)
                total_take = lp_for_p + part
                capped = True
        participation_payments.append({
            "series": p.name,
            "lp_money_m": lp_for_p,
            "participation_money_m": part,
            "total_money_m": total_take,
            "capped": capped,
        })

    for p in converted_series:
        part = p.effective_shares * per_share_residual
        participation_payments.append({
            "series": p.name,
            "lp_money_m": 0.0,
            "participation_money_m": part,
            "total_money_m": part,
            "capped": False,
            "converted": True,
        })

    common_payment = founder_residual + pool_residual + so_residual

    # Founder net cash (founder portion only — pool/SO are not the founder)
    founder_net_cash = founder_residual
    founder_pct_of_proceeds = (
        founder_net_cash / exit_value_money_m if exit_value_money_m > 0 else 0.0
    )

    return {
        "debt_repayment": min(debt_outstanding_money_m, exit_value_money_m),
        "lp_payments": lp_payments,
        "participation_payments": participation_payments,
        "common_payment": common_payment,
        "founder_net_cash": founder_net_cash,
        "founder_pct_of_proceeds": founder_pct_of_proceeds,
        "crossover_values": crossover_values,
        "convert_decisions": convert_decision,
        "fdso_total": fdso_total,
        "cash_pool": cash_pool,
        "so_exercise_inflow_money_m": so_exercise_inflow_money_m,
        "lp_total": lp_total,
        "residual": residual,
        "exit_pps_estimate": exit_pps_estimate,
        "founder_share_residual": founder_residual,
        "pool_share_residual": pool_residual,
        "so_share_residual": so_residual,
    }


# ============================================================================
# 8. Sheet generation
# ============================================================================


def build_cap_table_sheet(
    wb: Workbook,
    inp: CapTableInput,
    *,
    round_result: RoundResult | None = None,
    waterfall_scenarios: list[tuple[float, dict]] | None = None,
) -> Worksheet:
    """Generate / replace the Ownership sheet with full trace.

    Layout:
      Row 1-3:   Header (company name, as-of date, currency)
      Row 5+:    Pre-Round Cap Table
      Row 25+:   Round Event State Machine Trace (if round_result)
      Row 50+:   Post-Round Cap Table (if round_result)
      Row 70+:   Exit Waterfall (if waterfall_scenarios)
    """
    if CAP_TABLE_SHEET_NAME in wb.sheetnames:
        ws = wb[CAP_TABLE_SHEET_NAME]
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        # Clear existing content
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
                cell.font = ib.FONT_BODY
                cell.fill = PatternFill(fill_type="none")
                cell.alignment = Alignment(wrap_text=False, indent=0)
        ws.freeze_panes = None
    else:
        ws = wb.create_sheet(CAP_TABLE_SHEET_NAME)

    # Layout setup (no period columns; just A=margin / B=label / C-G=values)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 16
    ws.sheet_view.showGridLines = False

    money_unit = ib.MONEY_UNIT_BY_CCY.get(inp.reporting_currency, "M")

    # ------------------------------------------------------------------
    # Row 1-3: Header
    # ------------------------------------------------------------------
    ws["B1"] = inp.company_name + " — Cap Table"
    ws["B1"].font = ib.FONT_SECTION
    ws["B2"] = f"As of: {inp.as_of_date or 'current'}    Currency: {inp.reporting_currency} ({money_unit})"
    ws["B2"].font = ib.FONT_COMMENT

    # ------------------------------------------------------------------
    # Row 5+: Pre-Round Cap Table
    # ------------------------------------------------------------------
    ib.apply_section_header(ws.cell(row=5, column=2), "Pre-Round Cap Table")
    headers = ["Holder", "Class", "Shares", "% FDSO", "Notes"]
    for i, h in enumerate(headers):
        c = ws.cell(row=7, column=2 + i)
        c.value = h
        c.font = ib.FONT_SUBSECTION

    pre_total = (
        inp.founder_shares
        + inp.common_pool_available
        + inp.common_pool_issued
        + sum(p.shares for p in inp.preferred)
    )

    rows = [
        ("Founders", "Common", inp.founder_shares, "Reverse-vesting"),
        ("Common Pool — Granted (G0)", "Options/RSU", inp.common_pool_issued, "ESOP issued"),
        ("Common Pool — Available (P0)", "Options reserved", inp.common_pool_available, "Unallocated"),
    ]
    for p in inp.preferred:
        rows.append((
            f"{p.name} (AC)",
            p.name + (" (P)" if p.participating else " (NP)"),
            p.effective_shares,
            f"OCP {p.issue_price:,.2f}, ratio {p.conversion_ratio:.4f}, LP {p.liquidation_pref:.1f}x",
        ))
    for s in inp.safes:
        mfn_suffix = ", MFN" if s.mfn else ""
        note = (
            f"{s.principal_money_m:,.1f}{money_unit}, "
            f"cap {s.cap_money_m or 'n/a'}, "
            f"disc {s.discount:.0%}{mfn_suffix}"
        )
        rows.append((
            s.name,
            s.type,
            float("nan"),  # phantom
            note,
        ))

    r = 8
    for name, cls, shares, note in rows:
        ws.cell(row=r, column=2).value = name
        ws.cell(row=r, column=3).value = cls
        if not (isinstance(shares, float) and math.isnan(shares)):
            ws.cell(row=r, column=4).value = shares
            ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)
            if pre_total > 0:
                ws.cell(row=r, column=5).value = shares / pre_total
                ib.apply_formula(ws.cell(row=r, column=5), fmt=ib.FMT_PERCENT)
        else:
            ws.cell(row=r, column=4).value = "phantom"
            ib.apply_comment(ws.cell(row=r, column=4))
        ws.cell(row=r, column=6).value = note
        ib.apply_comment(ws.cell(row=r, column=6))
        r += 1

    r += 1
    ws.cell(row=r, column=2).value = "Pre-Round FDSO (excl. SAFEs)"
    ws.cell(row=r, column=2).font = ib.FONT_TOTAL
    ws.cell(row=r, column=4).value = pre_total
    ib.apply_grand_total(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)

    # ------------------------------------------------------------------
    # Row 25+: State Machine Trace
    # ------------------------------------------------------------------
    if round_result is not None:
        rr = round_result
        ib.apply_section_header(ws.cell(row=25, column=2), "Round Event State Machine — 6-Step Trace")
        r = 27
        ws.cell(row=r, column=2).value = f"Round: {inp.round_label}"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        ws.cell(row=r, column=3).value = (
            f"Pre={inp.round_pre_money_money_m:,.0f}{money_unit}, "
            f"Inv={inp.round_investment_money_m:,.0f}{money_unit}, "
            f"Pool target={inp.round_target_pool_pct:.1%}, "
            f"Iter={rr.iteration_count}"
        )
        ib.apply_comment(ws.cell(row=r, column=3))
        r += 2

        ws.cell(row=r, column=2).value = "Step 1 — Snapshot FDSO"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        r += 1
        for k, v in rr.pre_round_snapshot.items():
            ws.cell(row=r, column=2).value = "  " + k
            ws.cell(row=r, column=4).value = v
            ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)
            r += 1
        r += 1

        ws.cell(row=r, column=2).value = "Step 2 — SAFE / J-KISS Conversions"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        r += 1
        ws.cell(row=r, column=2).value = "  Name"
        ws.cell(row=r, column=3).value = "Type"
        ws.cell(row=r, column=4).value = "CP"
        ws.cell(row=r, column=5).value = "Shares"
        ws.cell(row=r, column=6).value = "Method"
        for c in range(2, 7):
            ws.cell(row=r, column=c).font = ib.FONT_COMMENT
        r += 1
        for sc in rr.safe_conversions:
            ws.cell(row=r, column=2).value = "  " + sc.name + (" (MFN)" if sc.mfn_applied else "")
            ws.cell(row=r, column=3).value = sc.type
            ws.cell(row=r, column=4).value = sc.conversion_price
            ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_PER_SHARE)
            ws.cell(row=r, column=5).value = sc.shares_issued
            ib.apply_formula(ws.cell(row=r, column=5), fmt=ib.FMT_SHARES)
            ws.cell(row=r, column=6).value = sc.method
            r += 1
        r += 1

        ws.cell(row=r, column=2).value = "Step 3 — Anti-Dilution"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        r += 1
        if not rr.ad_adjustments:
            ws.cell(row=r, column=2).value = "  (No existing preferred to adjust)"
            ib.apply_comment(ws.cell(row=r, column=2))
            r += 1
        else:
            ws.cell(row=r, column=2).value = "  Series"
            ws.cell(row=r, column=3).value = "Mode"
            ws.cell(row=r, column=4).value = "Old CP"
            ws.cell(row=r, column=5).value = "New CP"
            ws.cell(row=r, column=6).value = "Old Ratio"
            ws.cell(row=r, column=7).value = "New Ratio"
            for c in range(2, 8):
                ws.cell(row=r, column=c).font = ib.FONT_COMMENT
            r += 1
            for ad in rr.ad_adjustments:
                ws.cell(row=r, column=2).value = "  " + ad.series + (" *" if ad.triggered else "")
                ws.cell(row=r, column=3).value = ad.mode
                ws.cell(row=r, column=4).value = ad.old_conversion_price
                ws.cell(row=r, column=5).value = ad.new_conversion_price
                ws.cell(row=r, column=6).value = ad.old_ratio
                ws.cell(row=r, column=7).value = ad.new_ratio
                ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_PER_SHARE)
                ib.apply_formula(ws.cell(row=r, column=5), fmt=ib.FMT_PER_SHARE)
                ib.apply_formula(ws.cell(row=r, column=6), fmt=ib.FMT_MULTIPLE)
                ib.apply_formula(ws.cell(row=r, column=7), fmt=ib.FMT_MULTIPLE)
                r += 1
        r += 1

        ws.cell(row=r, column=2).value = "Step 4 — Pool Top-Up"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        ws.cell(row=r, column=4).value = rr.option_pool_topup_shares
        ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)
        r += 1

        ws.cell(row=r, column=2).value = "Step 5 — New Round Shares / PPS"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        ws.cell(row=r, column=4).value = rr.new_round_shares
        ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)
        ws.cell(row=r, column=5).value = rr.new_round_price
        ib.apply_formula(ws.cell(row=r, column=5), fmt=ib.FMT_PER_SHARE)
        r += 1

        ws.cell(row=r, column=2).value = "Step 6 — FDSO Reconcile"
        ws.cell(row=r, column=2).font = ib.FONT_SUBSECTION
        ws.cell(row=r, column=4).value = rr.fdso_post
        ib.apply_grand_total(ws.cell(row=r, column=4), fmt=ib.FMT_SHARES)
        r += 1

        # ------------------------------------------------------------------
        # Row 50+: Post-Round Cap Table
        # ------------------------------------------------------------------
        ib.apply_section_header(ws.cell(row=50, column=2), "Post-Round Cap Table")
        ws.cell(row=52, column=2).value = "Holder"
        ws.cell(row=52, column=3).value = "Class"
        ws.cell(row=52, column=4).value = "Shares"
        ws.cell(row=52, column=5).value = "% FDSO"
        for c in range(2, 6):
            ws.cell(row=52, column=c).font = ib.FONT_SUBSECTION
        rrn = 53
        for k, v in rr.fdso_breakdown.items():
            if k == "fdso_post":
                continue
            ws.cell(row=rrn, column=2).value = k
            ws.cell(row=rrn, column=4).value = v
            ib.apply_formula(ws.cell(row=rrn, column=4), fmt=ib.FMT_SHARES)
            if rr.fdso_post > 0:
                ws.cell(row=rrn, column=5).value = v / rr.fdso_post
                ib.apply_formula(ws.cell(row=rrn, column=5), fmt=ib.FMT_PERCENT)
            rrn += 1
        ws.cell(row=rrn, column=2).value = "Total FDSO"
        ws.cell(row=rrn, column=2).font = ib.FONT_TOTAL
        ws.cell(row=rrn, column=4).value = rr.fdso_post
        ib.apply_grand_total(ws.cell(row=rrn, column=4), fmt=ib.FMT_SHARES)
        ws.cell(row=rrn, column=5).value = 1.0
        ib.apply_grand_total(ws.cell(row=rrn, column=5), fmt=ib.FMT_PERCENT)
        rrn += 2
        ws.cell(row=rrn, column=2).value = "Closed-form verified"
        ws.cell(row=rrn, column=4).value = "✓" if rr.closed_form_verified else "× check trace"
        ib.apply_comment(ws.cell(row=rrn, column=4))

    # ------------------------------------------------------------------
    # Row 70+: Exit Waterfall
    # ------------------------------------------------------------------
    if waterfall_scenarios:
        ib.apply_section_header(ws.cell(row=70, column=2), "Exit Waterfall — Scenarios")
        ws.cell(row=72, column=2).value = "Exit Value"
        ws.cell(row=72, column=3).value = "Cash Pool"
        ws.cell(row=72, column=4).value = "LP Total"
        ws.cell(row=72, column=5).value = "Founder Net"
        ws.cell(row=72, column=6).value = "Founder %"
        ws.cell(row=72, column=7).value = "B-Convert?"
        for c in range(2, 8):
            ws.cell(row=72, column=c).font = ib.FONT_SUBSECTION

        r = 73
        for exit_val, w in waterfall_scenarios:
            ws.cell(row=r, column=2).value = exit_val
            ib.apply_formula(ws.cell(row=r, column=2), fmt=ib.FMT_MONEY)
            ws.cell(row=r, column=3).value = w.get("cash_pool", 0.0)
            ib.apply_formula(ws.cell(row=r, column=3), fmt=ib.FMT_MONEY)
            ws.cell(row=r, column=4).value = w.get("lp_total", 0.0)
            ib.apply_formula(ws.cell(row=r, column=4), fmt=ib.FMT_MONEY)
            ws.cell(row=r, column=5).value = w.get("founder_net_cash", 0.0)
            ib.apply_formula(ws.cell(row=r, column=5), fmt=ib.FMT_MONEY)
            ws.cell(row=r, column=6).value = w.get("founder_pct_of_proceeds", 0.0)
            ib.apply_formula(ws.cell(row=r, column=6), fmt=ib.FMT_PERCENT)
            decisions = w.get("convert_decisions", {})
            ws.cell(row=r, column=7).value = ", ".join(
                f"{k}={'Y' if v else 'N'}" for k, v in decisions.items()
            )
            ib.apply_comment(ws.cell(row=r, column=7))
            r += 1

    ws.insert_cols(2)
    ws.column_dimensions["A"].width = ib.COL_MARGIN_WIDTH
    ws.column_dimensions["B"].width = ib.COL_HIERARCHY_WIDTH
    ws.column_dimensions["C"].width = ib.COL_LABEL_WIDTH
    ws.column_dimensions["D"].width = ib.COL_WIDTH_LARGE
    ws.column_dimensions["E"].width = ib.COL_WIDTH_BASE
    ws.column_dimensions["F"].width = ib.COL_WIDTH_BASE
    ws.column_dimensions["G"].width = ib.COL_SOURCE_WIDTH
    for col in "HI":
        ws.column_dimensions[col].width = ib.COL_WIDTH_BASE
    ws.freeze_panes = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.lstrip()
            base = cell.alignment
            cell.alignment = Alignment(
                horizontal=base.horizontal,
                vertical=base.vertical,
                text_rotation=base.text_rotation,
                wrap_text=False,
                shrink_to_fit=base.shrink_to_fit,
                indent=0,
                relativeIndent=base.relativeIndent,
                justifyLastLine=base.justifyLastLine,
                readingOrder=base.readingOrder,
            )

    # Optional round/waterfall sections are styled only when their shifted
    # content label exists in column C after the hierarchy column is inserted.
    section_rows = [
        row for row in (5, 25, 50, 70) if ws.cell(row=row, column=3).value
    ]
    header_rows = [
        row
        for row in (7, 52, 72)
        if any(ws.cell(row=row, column=col).value for col in range(3, 10))
    ]
    for row in section_rows:
        ib.apply_semantic_fill_span(
            ws,
            row,
            2,
            min(9, max(7, ws.max_column)),
            ib.BG_HEADER_BAND,
            bottom=ib.THIN_LINE,
            border_start_col=3,
        )
        for col in range(2, min(9, max(7, ws.max_column)) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ib.FONT_COVER_TITLE
            base = cell.alignment
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                text_rotation=base.text_rotation,
                wrap_text=False,
                shrink_to_fit=base.shrink_to_fit,
                indent=0,
                relativeIndent=base.relativeIndent,
                justifyLastLine=base.justifyLastLine,
                readingOrder=base.readingOrder,
            )
        ws.row_dimensions[row].height = ib.ROW_HEIGHT_RELAXED
    for row in header_rows:
        end_col = max(col for col in range(3, 10) if ws.cell(row=row, column=col).value)
        ib.apply_semantic_fill_span(
            ws, row, 2, end_col, ib.BG_TABLE_HEADER, bottom=ib.THIN_LINE, border_start_col=3
        )
        for col in range(2, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ib.FONT_BODY_BOLD
            cell.alignment = Alignment(
                horizontal="left" if col <= 4 or isinstance(cell.value, str) else "right",
                vertical="center",
                wrap_text=False,
                indent=0,
            )
        ws.row_dimensions[row].height = ib.ROW_HEIGHT_BASE
    for row in range(1, ws.max_row + 1):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = ib.ROW_HEIGHT_BASE
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in (None, ""):
                continue
            if isinstance(cell.value, (int, float)) or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                cell.alignment = Alignment(
                    horizontal="right", vertical="center", wrap_text=False, indent=0
                )
            elif cell.alignment.horizontal is None:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False, indent=0
                )

    ib.set_tab_color(ws, "memo")
    ib.normalize_workbook_fonts(wb)
    ib.set_workbook_default_font(wb)
    ib.clear_blank_cell_styles(wb)
    ib.trim_blank_canvas(wb)
    ib.setup_print_layout(
        ws,
        orientation="landscape",
        fit_to_width=1,
        print_title_rows="1:3",
        print_title_cols="A:C",
        footer_right="&P / &N",
    )
    if len(wb.worksheets) > 1:
        for candidate in list(wb.worksheets):
            if candidate is ws:
                continue
            if (
                candidate.title == "Sheet"
                and ib.last_value_bounds(candidate) == (1, 1)
                and candidate["A1"].value is None
            ):
                wb.remove(candidate)
    return ws


# ============================================================================
# 9. Public API
# ============================================================================


def build_cap_table_for_workbook(
    wb: Workbook,
    inp: CapTableInput,
    *,
    exit_scenarios_money_m: list[float] | None = None,
) -> Workbook:
    """(Re-)build the Ownership sheet on the given workbook.

    If round parameters are provided on `inp`, a State Machine trace is added.
    If exit_scenarios_money_m is non-empty, an Exit Waterfall section is added.
    """
    rr = None
    if inp.round_pre_money_money_m is not None and inp.round_investment_money_m is not None:
        rr = run_round_state_machine(inp)

    waterfall_scenarios: list[tuple[float, dict]] | None = None
    if exit_scenarios_money_m:
        waterfall_scenarios = []
        for ev in exit_scenarios_money_m:
            w = compute_exit_waterfall(inp=inp, exit_value_money_m=ev)
            waterfall_scenarios.append((ev, w))

    build_cap_table_sheet(
        wb,
        inp,
        round_result=rr,
        waterfall_scenarios=waterfall_scenarios,
    )
    return wb


def get_default_input() -> CapTableInput:
    """Generic demonstrator input for the cap table state machine."""
    return CapTableInput(
        company_name="Startup",
        reporting_currency="JPY",
        as_of_date="",
        founder_shares=8_000_000.0,
        common_pool_issued=400_000.0,
        common_pool_available=1_100_000.0,
        safes=[
            SAFEInstrument(
                name="J-KISS-1",
                type="j_kiss_v2",
                principal_money_m=80.0,
                cap_money_m=1_000.0,
                discount=0.20,
                issuance_date="2024-06",
                mfn=False,
            ),
            SAFEInstrument(
                name="J-KISS-2",
                type="j_kiss_v2",
                principal_money_m=40.0,
                cap_money_m=1_500.0,
                discount=0.15,
                issuance_date="2025-01",
                mfn=True,
            ),
        ],
        preferred=[],
        round_pre_money_money_m=3_200.0,
        round_investment_money_m=800.0,
        round_target_pool_pct=0.12,
        round_anti_dilution_default="broad_wa",
        round_label="Priced Round",
    )


def get_section_19_2_input() -> CapTableInput:
    """Reference 04a §19.2 case for State Machine verification.

    Setup:
        Founder 6M, Pool 0.5M avail + 0.5M granted, Series A 2M @ ¥83.33,
        J-KISS ¥100M cap ¥500M post 20% disc.
        Round B: pre ¥1,000M, raise ¥200M, target pool 12%.

    Expected (up-round case in §19.2.5):
        PPS_B = ¥87.35, ΔP = 648,600, N_B = 2,289,640, FDSO_post = 13,738,240
        Founder = 43.7%
    """
    return CapTableInput(
        company_name="§19.2 Reference",
        reporting_currency="JPY",
        founder_shares=6_000_000.0,
        common_pool_issued=500_000.0,
        common_pool_available=500_000.0,
        safes=[
            SAFEInstrument(
                name="J-KISS",
                type="j_kiss_v1",     # Pre-money cap (per §19.2: cap = ¥500M / 9M)
                principal_money_m=100.0,
                cap_money_m=500.0,
                discount=0.20,
            ),
        ],
        preferred=[
            PreferredStockClass(
                name="Series A",
                shares=2_000_000.0,
                issue_price=83.333,
                liquidation_pref=1.0,
                participating=False,
                anti_dilution="broad_wa",
            ),
        ],
        round_pre_money_money_m=1_000.0,
        round_investment_money_m=200.0,
        round_target_pool_pct=0.12,
        round_label="Series B",
    )


def get_section_10_3_input() -> CapTableInput:
    """Reference 04b §10.3 cap table for Exit Waterfall verification.

    Setup at T3 post-Series-B:
        Founder 10M Common
        J-KISS A 2.5M  @ orig ¥80,  participating 1.0x   (¥200M LP)
        Series A 5.208333M @ orig ¥115.2, participating 1.0x  (¥600M LP)
        Series B 6.82795M  @ orig ¥439.4, NON-participating 1.0x  (¥3,000M LP)
        Pool granted 1.5M (strike ¥150 avg)
        Pool unallocated 2.050497M

    Expected (¥6B exit per §10.3.2):
        Cash Pool = ¥6,225M
        B-LP only = ¥3,000M (LP, since cross-over ≈ ¥11.4B > ¥6B)
        Founder net = ¥1,262.6M (~21.0% of ¥6B)
    """
    return CapTableInput(
        company_name="§10.3 Reference",
        reporting_currency="JPY",
        founder_shares=10_000_000.0,
        common_pool_issued=0.0,         # SO is supplied separately to waterfall
        common_pool_available=2_050_497.0,  # cancelled at exit per §10.3.2
        safes=[],
        preferred=[
            PreferredStockClass(
                name="J-KISS_AC",
                shares=2_500_000.0,
                issue_price=80.0,
                liquidation_pref=1.0,
                participating=True,
                anti_dilution="none",
            ),
            PreferredStockClass(
                name="Series A",
                shares=5_208_333.0,
                issue_price=115.2,
                liquidation_pref=1.0,
                participating=True,
                anti_dilution="broad_wa",
            ),
            PreferredStockClass(
                name="Series B",
                shares=6_827_950.0,
                issue_price=439.4,
                liquidation_pref=1.0,
                participating=False,
                anti_dilution="broad_wa",
                senior_rank=1,
            ),
        ],
    )


# ============================================================================
# 10. CLI
# ============================================================================


def _verify_pool_shuffle() -> bool:
    """Sanity: §2.3 worked example.
    F0=9M, P0=0, T=10%, PMV=$9M, INV=$3M.
    Expected X=13,846,154; Y=1,384,615.
    """
    res = option_pool_shuffle_closed_form(
        pre_existing_fdso=9_000_000,
        target_pool_pct=0.10,
        investment_money_m=3.0,
        pre_money_money_m=9.0,
        pool_already_available=0,
        pool_already_granted=0,
    )
    ok_x = abs(res["X_post_fdso"] - 13_846_154) < 5
    ok_y = abs(res["Y_pool_topup"] - 1_384_615) < 5
    return ok_x and ok_y


def _verify_section_19_2() -> dict:
    """Run State Machine on §19.2 setup; check key outputs vs reference."""
    inp = get_section_19_2_input()
    rr = run_round_state_machine(inp)
    expected_pps = 87.35
    expected_dp = 648_600
    expected_nb = 2_289_640
    expected_fdso = 13_738_240
    expected_founder_pct = 0.437
    return {
        "pps_ok": abs(rr.new_round_price - expected_pps) / expected_pps < 0.05,
        "dp_ok": abs(rr.option_pool_topup_shares - expected_dp) / expected_dp < 0.05,
        "nb_ok": abs(rr.new_round_shares - expected_nb) / expected_nb < 0.05,
        "fdso_ok": abs(rr.fdso_post - expected_fdso) / expected_fdso < 0.05,
        "founder_ok": abs(rr.founder_pct - expected_founder_pct) < 0.01,
        "actual_pps": rr.new_round_price,
        "actual_dp": rr.option_pool_topup_shares,
        "actual_nb": rr.new_round_shares,
        "actual_fdso": rr.fdso_post,
        "actual_founder_pct": rr.founder_pct,
        "iter": rr.iteration_count,
        "verified": rr.closed_form_verified,
    }


def _verify_section_10_3() -> dict:
    """Run Exit Waterfall on §10.3 setup at both anchor exits.

    §10.3.2 ¥6B  → founder ¥1,262.6M, B = LP (cross-over ¥11.4B > ¥6B)
    §10.3.3 ¥20B → founder ¥7,460.4M, B = convert
    Both rely on cancel_unallocated_pool=True (§6.6.2; canonical FDSO 26,036,283).
    """
    inp = get_section_10_3_input()
    w_low = compute_exit_waterfall(
        inp=inp,
        exit_value_money_m=6_000.0,
        debt_outstanding_money_m=0.0,
        so_strike_avg=150.0,
        so_outstanding_shares=1_500_000.0,
    )
    w_high = compute_exit_waterfall(
        inp=inp,
        exit_value_money_m=20_000.0,
        debt_outstanding_money_m=0.0,
        so_strike_avg=150.0,
        so_outstanding_shares=1_500_000.0,
    )
    expected_founder_low = 1_262.6
    expected_cash_pool_low = 6_225.0
    expected_lp_total_low = 3_800.0
    expected_founder_high = 7_460.4
    expected_cash_pool_high = 20_225.0
    expected_lp_total_high = 800.0
    return {
        # ----- ¥6B anchor (§10.3.2) -----
        "low_founder_net_money_m": w_low["founder_net_cash"],
        "low_founder_net_ok": abs(w_low["founder_net_cash"] - expected_founder_low)
            / expected_founder_low < 0.005,
        "low_cash_pool_ok": abs(w_low["cash_pool"] - expected_cash_pool_low)
            / expected_cash_pool_low < 0.01,
        "low_lp_total_ok": abs(w_low["lp_total"] - expected_lp_total_low)
            / expected_lp_total_low < 0.01,
        "low_B_convert": w_low["convert_decisions"].get("Series B"),
        "low_founder_pct": w_low["founder_pct_of_proceeds"],
        # ----- ¥20B anchor (§10.3.3) -----
        "high_founder_net_money_m": w_high["founder_net_cash"],
        "high_founder_net_ok": abs(w_high["founder_net_cash"] - expected_founder_high)
            / expected_founder_high < 0.01,
        "high_cash_pool_ok": abs(w_high["cash_pool"] - expected_cash_pool_high)
            / expected_cash_pool_high < 0.01,
        "high_lp_total_ok": abs(w_high["lp_total"] - expected_lp_total_high)
            / expected_lp_total_high < 0.05,
        "high_B_convert": w_high["convert_decisions"].get("Series B"),
        "high_founder_pct": w_high["founder_pct_of_proceeds"],
    }


def _cli() -> None:
    parser = argparse.ArgumentParser(
        description="Cap Table Builder — Ownership sheet generator + State Machine.",
    )
    parser.add_argument("--demo", action="store_true", help="Run demo (evals case 2)")
    parser.add_argument("--ref-19-2", action="store_true", help="Reference §19.2 case")
    parser.add_argument("--ref-10-3", action="store_true", help="Reference §10.3 case (exit)")
    parser.add_argument("--output", default="/tmp/cap_table_demo.xlsx", help="Output xlsx path")
    parser.add_argument(
        "--exit",
        default="6000,11442,20000",
        help="Comma-separated exit scenarios in money_m (default: 6000,11442,20000)",
    )
    parser.add_argument("--verify", action="store_true", help="Run all verification checks")
    args = parser.parse_args()

    if args.verify:
        print("=== Pool Shuffle (§2.3) ===")
        print("OK" if _verify_pool_shuffle() else "FAIL")
        print("\n=== State Machine (§19.2) ===")
        for k, v in _verify_section_19_2().items():
            print(f"  {k}: {v}")
        print("\n=== Exit Waterfall (§10.3) ===")
        for k, v in _verify_section_10_3().items():
            print(f"  {k}: {v}")
        return

    if args.ref_19_2:
        inp = get_section_19_2_input()
    elif args.ref_10_3:
        inp = get_section_10_3_input()
    else:
        inp = get_default_input()

    exit_scenarios = [float(x) for x in args.exit.split(",")] if args.exit else None

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    build_cap_table_for_workbook(wb, inp, exit_scenarios_money_m=exit_scenarios)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    _cli()
