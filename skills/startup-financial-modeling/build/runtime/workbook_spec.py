"""Typed workbook intermediate representation for generated finance models.

This module is intentionally small. It gives future sheet builders a
declarative target (`WorkbookSpec` -> `openpyxl`) without forcing the existing
large source-plan renderer to be rewritten in one pass.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import ib_format as ib


class StyleRole(str, Enum):
    LABEL = "label"
    HARD_INPUT = "hard_input"
    FORMULA = "formula"
    LINK_INTRA = "link_intra"
    LINK_EXTERNAL = "link_external"
    COMMENT = "comment"
    UNIT = "unit"
    HEADER = "header"


@dataclass(frozen=True)
class FormulaExpr:
    """A live Excel formula plus the reference type that controls font color."""

    text: str
    external: bool = False

    def __post_init__(self) -> None:
        if not self.text.startswith("="):
            raise ValueError("FormulaExpr.text must start with '='")

    @property
    def style_role(self) -> StyleRole:
        if self.external:
            return StyleRole.LINK_EXTERNAL
        return StyleRole.LINK_INTRA if "!" in self.text else StyleRole.FORMULA


@dataclass(frozen=True)
class CellSpec:
    row: int
    col: int
    value: Any = None
    role: StyleRole = StyleRole.LABEL
    fmt: str = ib.FMT_NUM


@dataclass(frozen=True)
class RowSpec:
    row: int
    cells: tuple[CellSpec, ...] = ()
    height: float | None = None
    fill_span: tuple[int, int, str] | None = None
    bottom_border: bool = False


@dataclass(frozen=True)
class SheetSpec:
    name: str
    rows: tuple[RowSpec, ...] = ()
    column_widths: dict[int, float] = field(default_factory=dict)
    tab_color: str | None = None


@dataclass(frozen=True)
class WorkbookSpec:
    sheets: tuple[SheetSpec, ...]

    def __post_init__(self) -> None:
        if not self.sheets:
            raise ValueError("WorkbookSpec requires at least one sheet")
        names = [sheet.name for sheet in self.sheets]
        if len(names) != len(set(names)):
            raise ValueError(f"WorkbookSpec has duplicate sheet names: {names}")


def _coerce_value_and_role(cell_spec: CellSpec) -> tuple[Any, StyleRole]:
    value = cell_spec.value
    role = cell_spec.role
    if isinstance(value, FormulaExpr):
        role = value.style_role
        value = value.text
    return value, role


def _apply_role(cell: Any, role: StyleRole, fmt: str) -> None:
    if role == StyleRole.HARD_INPUT:
        ib.apply_hard_input(cell, fmt)
    elif role == StyleRole.FORMULA:
        ib.apply_formula(cell, fmt)
    elif role == StyleRole.LINK_INTRA:
        ib.apply_link_intra(cell, fmt)
    elif role == StyleRole.LINK_EXTERNAL:
        ib.apply_link_external(cell, fmt)
    elif role == StyleRole.COMMENT:
        ib.apply_comment(cell, wrap_text=False)
    elif role == StyleRole.UNIT:
        cell.font = ib.FONT_COMMENT
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    elif role == StyleRole.HEADER:
        cell.font = ib.FONT_BODY_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    else:
        ib.apply_label(cell)


def _iter_cells(rows: Iterable[RowSpec]) -> Iterable[CellSpec]:
    for row in rows:
        yield from row.cells


def render_workbook_spec(spec: WorkbookSpec) -> Workbook:
    """Render a typed workbook spec into an `openpyxl` workbook."""

    wb = Workbook()
    for idx, sheet_spec in enumerate(spec.sheets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_spec.name
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = None
        if sheet_spec.tab_color:
            ws.sheet_properties.tabColor = sheet_spec.tab_color
        for col, width in sheet_spec.column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        for row_spec in sheet_spec.rows:
            if row_spec.height is not None:
                ws.row_dimensions[row_spec.row].height = row_spec.height
            if row_spec.fill_span is not None:
                start_col, end_col, color = row_spec.fill_span
                ib.apply_semantic_fill_span(
                    ws,
                    row_spec.row,
                    start_col,
                    end_col,
                    color,
                    bottom=ib.THIN_LINE if row_spec.bottom_border else None,
                )
            for cell_spec in row_spec.cells:
                value, role = _coerce_value_and_role(cell_spec)
                cell = ws.cell(row=cell_spec.row, column=cell_spec.col, value=value)
                _apply_role(cell, role, cell_spec.fmt)
    ib.normalize_workbook_fonts(wb)
    ib.set_workbook_default_font(wb)
    return wb


__all__ = [
    "CellSpec",
    "FormulaExpr",
    "RowSpec",
    "SheetSpec",
    "StyleRole",
    "WorkbookSpec",
    "render_workbook_spec",
]
