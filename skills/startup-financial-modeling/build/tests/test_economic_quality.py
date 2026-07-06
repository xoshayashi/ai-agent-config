"""Economic-quality regression tests for the startup financial modeling kernel.

These tests pin the economic *sanity* of generated plans, not just their
structure. They exist because the generator could previously emit a textbook
seed-SaaS plan with -789% gross margin (cost-to-serve drivers were unanchored
to the stated/target gross margin) while still passing every structural check.

Run directly:
    PYTHONPYCACHEPREFIX=$(mktemp -d) \\
      python3 skills/startup-financial-modeling/build/tests/test_economic_quality.py
"""

from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

SKILL_DIR = Path(__file__).resolve().parents[2]
RUNTIME_DIR = SKILL_DIR / "build" / "runtime"
sys.path.insert(0, str(RUNTIME_DIR))

import build_model  # noqa: E402
import economic_kernel as kernel  # noqa: E402
import source_plan_builder as source_plan  # noqa: E402

# --- Archetype source narratives -------------------------------------------

SAAS_STORY = """# Seed-stage B2B SaaS equity story

The company sells a workflow-automation SaaS to mid-market operations teams.
The model must support a seed fundraise: 5-year integrated plan, monthly burn,
runway, unit economics, and dilution.

Monetization is a per-seat subscription at 月額1.2万円 per seat, average 40 seats
per customer. Phase 1 targets 600 paying customers in three years, roughly
¥35億 ARR at maturity. Gross margin target is about 78%. Logo churn runs near
14% annually with modest net revenue expansion.

The plan needs market sizing, competitive funding benchmarks, hiring, scenario
analysis, sensitivity analysis, and an investor-use funding plan.

Source: customer discovery memo, market sizing memo, investor benchmark notes.
"""

MARKETPLACE_STORY = """# Series A marketplace equity story

The company runs a two-sided marketplace connecting local service providers
with consumers. Revenue is a take rate of 18% on gross merchandise value.
The model needs a 5-year integrated plan covering liquidity, contribution
margin, and the next fundraise.

Phase 1 targets ¥120億 GMV at maturity with repeat purchase behaviour.

Source: cohort analysis memo, marketplace benchmark notes.
"""

HARDWARE_STORY = """# Asset deployment equity story

The company is an asset-heavy startup deploying specialized field devices into
operational sites. The model combines recurring lease fees, services, and
software. Unit pricing is 月額32万円. Phase 1 targets cumulative deployments of
3,000 units in three years and about 25,000 operating units at maturity.
Hardware cost starts around ¥8m per unit and declines with scale.

Source: customer discovery memo, lender discussion notes.
"""

ARCHETYPES = {
    "saas": SAAS_STORY,
    "marketplace": MARKETPLACE_STORY,
    "hardware": HARDWARE_STORY,
}


def _implied(facts) -> list[dict]:
    """Replicate the workbook revenue / COGS / people formulas period by period.

    Mirrors Revenue Build, Cost Build, and People Plan so the test is an
    independent check on the driver values rather than trusting kernel
    internals. The installed base follows the kernel's K1 input-fidelity
    rule (Phase 2 Task 2.2): when ``facts.customers_pinned`` is True the
    customers series drives the average installed base, otherwise the
    legacy new-unit rollforward
    (``ending_units(facts.new_units, churn_rate=facts.churn_rate)``) holds.
    """
    if getattr(facts, "customers_pinned", False) and any(v > 0 for v in facts.customers):
        ending = list(facts.customers)
    else:
        # narrative-derived plans keep the legacy ramp-style churn schedule
        # (matched by calibrate at derivation time).
        ending = kernel.ending_units(facts.new_units)
    average = kernel.average_units(ending)
    rows: list[dict] = []
    for i in range(len(facts.years)):
        if facts.revenue_mode == "unit_sale":
            # Hardware: one-time per-unit sale, no recurring billing.
            subtotal = facts.new_units[i] * facts.monthly_price_yen[i]
        else:
            txn = facts.gmv_yen[i] * facts.take_rate[i]
            recurring = average[i] * facts.monthly_price_yen[i] * 12
            one_time = facts.new_units[i] * facts.monthly_price_yen[i] * 3
            subtotal = txn + recurring + one_time
        revenue = subtotal + subtotal * facts.other_revenue_share[i]
        variable = revenue * facts.variable_cogs_pct[i]
        delivery = average[i] * facts.delivery_cost_yen[i] * 12
        cloud = average[i] * facts.cloud_cost_yen[i] * 12
        support = facts.customers[i] * facts.support_cost_yen[i] * 12
        cogs = variable + delivery + cloud + support
        headcount = (
            facts.product_headcount[i]
            + facts.gtm_headcount[i]
            + facts.operations_headcount[i]
            + facts.ga_headcount[i]
        )
        people_cost = headcount * facts.avg_comp_yen[i]
        rows.append(
            {
                "revenue": revenue,
                "cogs": cogs,
                "gross_margin": (revenue - cogs) / revenue if revenue else 0.0,
                "people_cost": people_cost,
            }
        )
    return rows


def test_gross_margin_tracks_target_across_archetypes() -> None:
    """Every period's implied gross margin must sit within 3pp of its target."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        rows = _implied(facts)
        for i, row in enumerate(rows):
            assert row["revenue"] > 0, f"{name} period {i}: non-positive revenue"
            gap = abs(row["gross_margin"] - facts.target_gross_margin[i])
            assert gap <= 0.03, (
                f"{name} period {i}: gross margin {row['gross_margin']:.1%} "
                f"vs target {facts.target_gross_margin[i]:.1%} (gap {gap:.1%})"
            )
        assert rows[-1]["gross_margin"] > 0, f"{name}: negative mature gross margin"


def test_stated_gross_margin_is_extracted_from_narrative() -> None:
    """A narrative gross-margin target overrides the profile default."""
    facts = kernel.derive_source_facts(SAAS_STORY)
    assert abs(facts.target_gross_margin[-1] - 0.78) <= 0.01, (
        f"stated 78% gross margin not extracted: {facts.target_gross_margin[-1]}"
    )
    # A qualifying clause between "gross margin" and the figure must not
    # push the percentage out of the extractor's reach.
    assert kernel.extract_target_gross_margin(
        "Gross margin on the subscription is about 82%."
    ) == 0.82, "a qualified gross-margin phrase was dropped"


def test_stated_churn_rate_is_extracted_and_annualized() -> None:
    """A stated churn rate is honored; a monthly figure is annualized."""
    assert kernel.extract_churn_rate(
        "Logo churn runs near 14% annually."
    ) == 0.14, "a stated annual churn rate was not extracted"
    monthly = kernel.extract_churn_rate("We model 1.5% monthly churn.")
    assert monthly is not None and 0.16 <= monthly <= 0.18, (
        f"1.5% monthly churn was not annualized: {monthly}"
    )
    assert kernel.extract_churn_rate("No churn figure stated here.") is None
    # A nearby "monthly price" sentence must not annualize an annual churn.
    assert kernel.extract_churn_rate(
        "Churn is 8%. Monthly price is $50."
    ) == 0.08, "an unrelated 'monthly' token wrongly annualized churn"
    # Employee churn is an HR metric — the customer churn figure wins.
    assert kernel.extract_churn_rate(
        "Employee churn is 20%. Customer churn is 8%."
    ) == 0.08, "employee churn was taken instead of customer churn"
    # An unrelated percentage before the churn keyword is not the churn rate.
    assert kernel.extract_churn_rate(
        "We offer a 5% discount; logo churn is 12%."
    ) == 0.12, "an unrelated percentage was read as the churn rate"
    facts = kernel.derive_source_facts(
        "# SaaS plan\n\nA SaaS platform priced at 月額1.2万円. Customer "
        "churn is about 9%. 5-year plan, seed round. Source: memo.\n"
    )
    assert abs(facts.churn_rate - 0.09) < 1e-6, (
        f"stated 9% churn ignored; churn_rate = {facts.churn_rate}"
    )


def test_annual_pricing_is_converted_to_a_monthly_figure() -> None:
    """An ACV or explicitly annual price must be divided to a monthly price.

    The model drives recurring revenue off a monthly price; a ¥7M/year ACV
    read as ¥7M/month overstates revenue twelvefold.
    """
    acv = kernel.derive_source_facts(
        "# SaaS plan\nB2B SaaS. ACV is ¥7,200,000 per customer per year.\n"
        "Source: memo.\n"
    )
    assert acv.monthly_price_yen[0] == 600_000, (
        f"ACV not converted to a monthly price: {acv.monthly_price_yen[0]}"
    )
    annual = kernel.derive_source_facts(
        "# SaaS plan\nB2B SaaS. The subscription price is ¥1,200,000 per year.\n"
        "Source: memo.\n"
    )
    assert annual.monthly_price_yen[0] == 100_000, (
        f"annual price not converted: {annual.monthly_price_yen[0]}"
    )
    # Japanese annual cue (年額).
    jp = kernel.derive_source_facts(
        "# SaaS計画\nB2B SaaS。利用料は年額240万円。\nSource: メモ。\n"
    )
    assert jp.monthly_price_yen[0] == 200_000, (
        f"年額 price not converted: {jp.monthly_price_yen[0]}"
    )
    # A monthly price stays monthly, and an explicit 月額 wins over a later
    # ACV mention in the same document.
    monthly = kernel.derive_source_facts(SAAS_STORY)
    assert monthly.monthly_price_yen[0] == 12_000
    both = kernel.derive_source_facts(
        "# SaaS plan\nB2B SaaS priced at 月額50,000円. ACV is ¥1,200,000.\n"
        "Source: memo.\n"
    )
    assert both.monthly_price_yen[0] == 50_000, (
        f"月額 should win over ACV: {both.monthly_price_yen[0]}"
    )


def test_unit_price_is_extracted_from_natural_phrasing() -> None:
    """A unit price reached by natural phrasing — "sells for $X", a bare
    "$X per robot" with no cue word — is extracted, not dropped to the
    profile default."""
    hw = kernel.profile_for_text(
        "hardware robot manufacturing plan with significant capex"
    )
    assert hw.key == "hardware_asset_heavy"
    assert kernel.extract_price(
        "Each robot sells for $48,000.", hw, "USD"
    ) == 48000, "'sells for $X' unit price was not extracted"
    assert kernel.extract_price(
        "Each robot is sold at $48,000.", hw, "USD"
    ) == 48000, "'sold at $X' unit price was not extracted"
    assert kernel.extract_price(
        "Each robot sells at $48,000.", hw, "USD"
    ) == 48000, "'sells at $X' unit price was not extracted"
    # `sells/sold for|at` needs a currency mark — a duration or volume after
    # the verb ("sold for 30000 hours") must not be read as a price.
    assert kernel.extract_price(
        "The unit sold for 30000 hours of field testing.", hw, "USD"
    ) != 30000, "a non-price figure after 'sold for' was read as a price"
    # No cue keyword at all — the number-first "$X per <unit>" form.
    assert kernel.extract_price(
        "Ships at $48,000 per unit.", hw, "USD"
    ) == 48000, "'$X per unit' unit price was not extracted"
    # The unit noun is word-bounded: "seat" inside "seating" must not match,
    # so "$900 per seating chart" is not read as a $900 unit price.
    assert kernel.extract_price(
        "$900 per seating chart.", hw, "USD"
    ) != 900, "word-boundary guard failed: a partial-word unit noun matched"


def test_monthly_burn_phrasing_does_not_flip_model_grain() -> None:
    """'monthly burn' is a metric, not a request for a month-by-month model."""
    assert kernel.extract_model_grain(SAAS_STORY) == "annual"
    facts = kernel.derive_source_facts(SAAS_STORY)
    assert facts.grain == "annual"


def test_explicit_monthly_model_request_is_still_detected() -> None:
    """An explicit monthly-model request must still produce a monthly grain."""
    assert kernel.extract_model_grain("Build a monthly model for the seed plan.") == "monthly"
    assert kernel.extract_model_grain("月次モデルでシード計画を作る") == "monthly"


def test_hyphenated_year_horizon_is_honored() -> None:
    """A hyphenated horizon ("6-year plan") sets the forecast length — it
    must not fall silently to the default because of the hyphen."""
    assert kernel.extract_forecast_periods("We model a 6-year plan.", "annual") == 6
    # The hyphen fix carries through the monthly-grain multiplier (capped
    # at MAX_FORECAST_PERIODS).
    assert kernel.extract_forecast_periods("6-year plan", "monthly") == min(
        6 * 12, kernel.MAX_FORECAST_PERIODS
    )
    # A company-age adjective is not a horizon and must not be matched,
    # in either the hyphenated or the space-separated form.
    assert kernel.extract_forecast_periods("a 6-year-old startup", "annual") != 6, (
        "company age was misread as a forecast horizon"
    )
    assert kernel.extract_forecast_periods("a 6-year old startup", "annual") != 6, (
        "space-separated company age was misread as a forecast horizon"
    )
    facts = kernel.derive_source_facts(
        "# Deep-tech plan\n\nA deep-tech company. We model a 6-year "
        "forecast.\nSource: roadmap memo.\n"
    )
    assert len(facts.years) == 6, (
        f"a stated 6-year horizon produced {len(facts.years)} periods"
    )


def test_start_year_takes_the_start_of_a_range() -> None:
    """A hyphenated year range ("forecast (2026-2031)") sets the start year
    to the first year, not the last."""
    assert kernel.extract_start_year(
        "We model a 6-year forecast horizon (2026-2031)."
    ) == 2026, "start year took the end of the range, not the start"
    # The same non-greedy path covers a Japanese cue word.
    assert kernel.extract_start_year("事業計画(2026-2031)") == 2026, (
        "Japanese-cue start year took the end of the range"
    )
    facts = kernel.derive_source_facts(
        "# Deep-tech plan\n\nA deep-tech company. We model a forecast "
        "horizon (2026-2031).\nSource: roadmap memo.\n"
    )
    assert facts.years[0] == 2026, (
        f"plan starts at FY{facts.years[0]}, not the stated 2026"
    )


def test_demand_ramp_reaches_stated_arr() -> None:
    """A narrative ARR target must anchor the demand ramp, not be ignored.

    The SaaS story states roughly ¥35億 ARR at maturity. A plan that lands an
    order of magnitude short of its own stated target is not investor-grade.
    """
    facts = kernel.derive_source_facts(SAAS_STORY)
    rows = _implied(facts)
    mature_revenue = rows[-1]["revenue"]
    target_arr = 3_500_000_000
    assert 0.6 * target_arr <= mature_revenue <= 1.6 * target_arr, (
        f"mature revenue ¥{mature_revenue:,.0f} is far from stated "
        f"¥{target_arr:,.0f} ARR target"
    )


def test_customer_count_reaches_stated_target() -> None:
    """A stated customer count ('600 paying customers') must be honored."""
    facts = kernel.derive_source_facts(SAAS_STORY)
    assert abs(facts.customers[-1] - 600) <= 30, (
        f"mature customer count {facts.customers[-1]} vs stated 600"
    )


def test_saas_revenue_tracks_the_stated_customer_count() -> None:
    """A SaaS plan stating a customer count but no total ARR must drive
    revenue off that customer count — not the profile-default unit ramp,
    which diverged ~6x from the stated scale."""
    facts = kernel.derive_source_facts(
        "# Seed SaaS plan\n\nA B2B SaaS company. The subscription price is "
        "¥100,000 per customer per month. We have 18 customers today and "
        "target 320 customers by year five. Gross margin 80%. 5-year plan, "
        "seed round. Source: investor memo.\n"
    )
    assert kernel.mechanic_key(facts) == "recurring_software"
    revenue = kernel.plan_revenue_series(
        facts.new_units, facts.gmv_yen, facts.monthly_price_yen,
        facts.take_rate, facts.other_revenue_share, facts.revenue_mode,
    )
    # 320 customers x the mature monthly price x 12; the model's one-time and
    # other-revenue components are headroom, but not a 6x default-ramp blowup.
    customers_x_acv = facts.customers[-1] * facts.monthly_price_yen[-1] * 12
    assert 0.8 * customers_x_acv <= revenue[-1] <= 1.4 * customers_x_acv, (
        f"mature revenue ¥{revenue[-1]:,.0f} is not customer-count-driven "
        f"(customers x ACV = ¥{customers_x_acv:,.0f})"
    )


def test_marketplace_gmv_honors_stated_maturity_target() -> None:
    """A marketplace story's stated maturity GMV must anchor the GMV ramp.

    The marketplace narrative states ¥120億 GMV at maturity. The ramp must land
    on that figure, not treat it as a period-0 base (which inflated it ~700x).
    """
    facts = kernel.derive_source_facts(MARKETPLACE_STORY)
    target = 12_000_000_000
    mature_gmv = facts.gmv_yen[-1]
    assert 0.7 * target <= mature_gmv <= 1.5 * target, (
        f"mature GMV ¥{mature_gmv:,.0f} is far from stated ¥{target:,.0f} target"
    )


def test_marketplace_present_value_gmv_stays_period_zero_base() -> None:
    """A present-value GMV ("GMV is ¥10B") must seed the base, not the maturity.

    Guards against a maturity-cue false positive silently scaling a stated
    current GMV down to a small period-0 figure.
    """
    facts = kernel.derive_source_facts(
        "# Marketplace plan\nThe company is a marketplace startup. GMV is ¥10B "
        "today, take rate is 12%.\nSource: management memo.\n"
    )
    assert facts.gmv_yen[0] >= 9_000_000_000, (
        f"present-value GMV scaled down to period-0 ¥{facts.gmv_yen[0]:,.0f}"
    )


def test_marketplace_gmv_end_of_plan_phrasing_is_a_maturity_target() -> None:
    """A maturity GMV stated by phrasing other than "at maturity" must anchor
    the ramp end — not be used as the period-0 base and ramped an order of
    magnitude past itself. A separate "today" figure does not displace it."""
    target = 120_000_000_000
    for phrase in (
        "¥120B GMV by the end of the plan",
        "¥120B GMV in the final year",
        "最終年に¥120B GMV",
    ):
        facts = kernel.derive_source_facts(
            f"# Marketplace plan\n\nA two-sided marketplace, take rate 15% "
            f"on GMV. GMV is ¥4B today; {phrase}. 5-year plan, seed round. "
            f"Source: cohort memo.\n"
        )
        mature_gmv = facts.gmv_yen[-1]
        assert 0.7 * target <= mature_gmv <= 1.3 * target, (
            f"[{phrase}] mature GMV ¥{mature_gmv:,.0f} is far from the "
            f"stated ¥{target:,.0f} maturity target"
        )
    # An operational-scaling phrase near the period-0 base must NOT make the
    # current GMV a maturity target.
    base = kernel.derive_source_facts(
        "# Marketplace plan\n\nA two-sided marketplace, take rate 15% on "
        "GMV. GMV is ¥4B today; we plan to scale to 3 new cities. 5-year "
        "plan, seed round. Source: cohort memo.\n"
    )
    assert base.gmv_yen[0] >= 3_500_000_000, (
        f"a 'scale to 3 cities' phrase wrongly anchored GMV to maturity: "
        f"period-0 ¥{base.gmv_yen[0]:,.0f}"
    )


USD_STORY = """# Seed B2B SaaS equity story (USD)

A US-based B2B SaaS company. The model must support a seed fundraise: 5-year
integrated plan, runway, unit economics, and dilution. Pricing is a per-seat
subscription at $80 per seat per month. Phase 1 targets $12M ARR at maturity.
Gross margin target is about 80%.

Source: customer discovery memo, investor benchmark notes.
"""


def test_usd_narrative_produces_a_coherent_usd_plan() -> None:
    """A USD narrative is modeled in USD, at the stated scale, with USD-scale costs."""
    facts = kernel.derive_source_facts(USD_STORY)
    assert facts.currency == "USD", f"currency not detected as USD: {facts.currency}"
    assert facts.monthly_price_yen[0] == 80, (
        f"per-seat price '$80' not extracted: {facts.monthly_price_yen[0]}"
    )
    # JPY default magnitudes must be FX-scaled — a USD plan cannot carry a
    # ¥16M loaded comp read as $16M, nor a ¥300M cash floor read as $300M.
    assert facts.avg_comp_yen[0] < 1_000_000, (
        f"loaded comp not FX-scaled for USD: {facts.avg_comp_yen[0]}"
    )
    assert facts.beginning_cash_yen < 50_000_000, (
        f"beginning cash not FX-scaled for USD: {facts.beginning_cash_yen}"
    )
    assert kernel.audit_economic_coherence(facts) == [], (
        "USD plan is not economically coherent"
    )


def test_jpy_narratives_are_not_misdetected_as_usd() -> None:
    """JPY stories — including ones citing a dollar-priced competitor — stay JPY."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        assert facts.currency == "JPY", f"{name} misdetected as {facts.currency}"
        # JPY magnitudes are unchanged (no FX scaling).
        assert facts.avg_comp_yen[0] >= 10_000_000, f"{name}: JPY comp was scaled"
    mixed = "# 計画\n月額1.2万円のSaaS。競合は $499/月。Source: メモ。"
    assert kernel.extract_currency(mixed) == "JPY"


def test_yaml_can_override_the_fx_rate() -> None:
    """The JPY-per-USD rate is overridable for USD plans via structured input."""
    facts = kernel.derive_source_facts(USD_STORY, jpy_per_usd=100.0)
    other = kernel.derive_source_facts(USD_STORY, jpy_per_usd=200.0)
    # A smaller divisor leaves larger USD magnitudes.
    assert facts.avg_comp_yen[0] > other.avg_comp_yen[0]


def test_structured_currency_scales_default_costs_consistently() -> None:
    """A structured `currency:` field — not an incidental token in the company
    name — must determine how default cost primitives are FX-scaled.

    `avg_comp` and `capex_per_unit` defaults were materialized during
    extraction using the currency detected from the (for `--input`,
    synthetic) narrative. A USD plan whose company string lacked a "USD"
    token kept yen-scale loaded comp (¥16M against USD revenue) — a ~150x
    incoherence that drove the model insolvent.
    """
    base = {
        "mechanics": "payments and lending fintech",
        "currency": "USD",
        "periods": 5,
        "customers": [400, 1200, 3000, 6000, 10000],
        "monthly_price_yen": [90, 90, 90, 90, 90],
    }
    plain = kernel.derive_source_facts_from_mapping(
        {**base, "company": "PayStream Fintech"}
    )
    tokened = kernel.derive_source_facts_from_mapping(
        {**base, "company": "PayStream USD Inc"}
    )
    # An incidental "USD" token in the company name must not move the scale.
    assert plain.avg_comp_yen == tokened.avg_comp_yen, (
        f"company name shifted avg_comp: {plain.avg_comp_yen[0]:,} vs "
        f"{tokened.avg_comp_yen[0]:,}"
    )
    assert plain.capex_per_unit_yen == tokened.capex_per_unit_yen, (
        "company name shifted capex_per_unit"
    )
    # The USD default loaded comp is the yen default divided by the FX rate.
    expected = 16_000_000 / kernel.DEFAULT_JPY_PER_USD
    assert abs(plain.avg_comp_yen[0] - expected) <= 0.02 * expected, (
        f"USD plan kept a yen-scale loaded comp: {plain.avg_comp_yen[0]:,} "
        f"(expected ~{expected:,.0f})"
    )
    # A model with currency-consistent cost primitives is economically coherent.
    assert kernel.audit_economic_coherence(plain) == [], (
        f"structured USD plan is not coherent: {kernel.audit_economic_coherence(plain)}"
    )
    # An explicit avg_comp / capex in the input still takes precedence over
    # the FX-scaled default.
    overridden = kernel.derive_source_facts_from_mapping(
        {
            **base,
            "company": "PayStream Fintech",
            "avg_comp_yen": [200_000] * 5,
            "capex_per_unit_yen": [50_000] * 5,
        }
    )
    assert overridden.avg_comp_yen[0] == 200_000, (
        f"an explicit avg_comp override was not honored: {overridden.avg_comp_yen[0]:,}"
    )
    assert overridden.capex_per_unit_yen[0] == 50_000, (
        f"an explicit capex_per_unit override was not honored: "
        f"{overridden.capex_per_unit_yen[0]:,}"
    )


def test_payroll_is_not_absurd_at_maturity() -> None:
    """Mature-period payroll must not dwarf revenue (a sign of grain mixups)."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        rows = _implied(facts)
        ratio = rows[-1]["people_cost"] / rows[-1]["revenue"]
        assert ratio < 1.5, f"{name}: mature payroll/revenue ratio {ratio:.2f}"


def test_stated_rd_team_size_anchors_the_headcount_plan() -> None:
    """A stated current R&D team size pins the people-cost base — the
    auto-derived ramp must not size an R&D-heavy plan from revenue alone."""
    assert kernel.extract_rd_headcount("We run a 32-person R&D team.") == 32
    assert kernel.extract_rd_headcount("An engineering team of 25.") == 25
    assert kernel.extract_rd_headcount("No team size stated.") == 0
    # A bare "N engineers" with no team framing is not a team size.
    assert kernel.extract_rd_headcount("We plan to hire 15 engineers.") == 0
    facts = kernel.derive_source_facts(
        "# Deep-tech plan\n\nA pre-revenue deep-tech company. There is no "
        "product revenue during the plan. We run a 32-person R&D team. "
        "5-year plan, seed round. Source: roadmap memo.\n"
    )
    # Scaling pins period 0 exactly on the stated team size.
    assert facts.product_headcount[0] == 32, (
        f"a stated 32-person R&D team was ignored; period-0 product "
        f"headcount = {facts.product_headcount[0]}"
    )
    # The ramp shape is preserved — headcount still grows.
    assert facts.product_headcount[-1] > facts.product_headcount[0]


def test_economic_audit_passes_for_healthy_archetypes() -> None:
    """The economic-coherence audit must clear a sound plan, every archetype."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        issues = kernel.audit_economic_coherence(facts)
        assert issues == [], f"{name}: unexpected economic-audit issues {issues}"


def test_economic_audit_catches_a_broken_cost_stack() -> None:
    """Inflating the cost drivers must surface a gross-margin violation."""
    from dataclasses import replace

    facts = kernel.derive_source_facts(SAAS_STORY)
    broken = replace(
        facts, delivery_cost_yen=[cost * 60 for cost in facts.delivery_cost_yen]
    )
    issues = kernel.audit_economic_coherence(broken)
    assert any("gross margin" in issue for issue in issues), issues


def test_economic_audit_catches_insolvency() -> None:
    """A plan stripped of its funding must be flagged insolvent and round-less."""
    from dataclasses import replace

    facts = kernel.derive_source_facts(SAAS_STORY)
    broken = replace(facts, equity_raise_yen=[0 for _ in facts.equity_raise_yen])
    issues = kernel.audit_economic_coherence(broken)
    assert any("insolven" in issue for issue in issues), issues
    assert any("funding round" in issue for issue in issues), issues


def test_economic_audit_catches_non_positive_revenue() -> None:
    """A period with no economic volume must be flagged as non-positive revenue."""
    from dataclasses import replace

    facts = kernel.derive_source_facts(SAAS_STORY)
    broken = replace(
        facts,
        new_units=[0 for _ in facts.new_units],
        gmv_yen=[0 for _ in facts.gmv_yen],
    )
    issues = kernel.audit_economic_coherence(broken)
    assert any("non-positive revenue" in issue for issue in issues), issues


def test_operations_headcount_does_not_explode_at_high_account_count() -> None:
    """A high-count, low-ARPU plan must staff operations off revenue, not the
    raw account count.

    Operations FTE once carried a `scale / ops_divisor` term — one customer
    equalled a fixed slice of operations staff regardless of price. A consumer
    plan with 260k accounts at a low monthly price drew an operations org of
    hundreds of FTE, pushing people cost past revenue and EBITDA deeply
    negative while every other audit check still passed.
    """
    facts = kernel.derive_source_facts_from_mapping(
        {
            "company": "MassMarket",
            "customers": [20_000, 70_000, 140_000, 210_000, 260_000],
            "monthly_price_yen": 400,
        }
    )
    rows = _implied(facts)
    ratio = rows[-1]["people_cost"] / rows[-1]["revenue"]
    assert ratio < 0.8, (
        f"mature people cost is {ratio:.2f}x revenue — operations headcount "
        f"is scaling off raw account count, not revenue"
    )


def test_balance_sheet_balances_in_the_kernel_projection() -> None:
    """The projected balance sheet must satisfy A = L + E every period."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        for idx, period in enumerate(kernel.project_balance_sheet(facts)):
            tolerance = 1e-4 * max(1.0, abs(period["total_assets"]))
            assert abs(period["balance_check"]) <= tolerance, (
                f"{name} period {idx}: balance sheet off by "
                f"{period['balance_check']:,.0f}"
            )


def test_depreciation_uses_the_accumulated_asset_base() -> None:
    """D&A depreciates the accumulated asset base, capped at gross PP&E.

    The prior model charged only the current period's CapEx; mature-period D&A
    must now exceed that, and accumulated D&A may never exceed gross PP&E.
    """
    facts = kernel.derive_source_facts(HARDWARE_STORY)
    projection = kernel.project_plan_free_cash_flow(facts)
    cumulative_capex = accumulated_da = 0.0
    for idx, period in enumerate(projection):
        cumulative_capex += period["capex"]
        accumulated_da += period["depreciation"]
        assert accumulated_da <= cumulative_capex + 1.0, (
            f"period {idx}: accumulated D&A exceeds gross PP&E"
        )
    # Accumulated-base depreciation charges materially more over the horizon
    # than the old single-period model (each period's CapEx / life), which
    # would total cumulative-CapEx / life.
    life = facts.depreciation_life_months[-1] or 60
    single_period_total = cumulative_capex * 12.0 / life
    assert accumulated_da > single_period_total, (
        "D&A does not reflect the accumulated asset base"
    )


def test_economic_audit_flags_an_unbalanced_sheet() -> None:
    """The audit must surface a balance sheet that does not balance."""
    facts = kernel.derive_source_facts(SAAS_STORY)
    original = kernel.project_balance_sheet
    kernel.project_balance_sheet = lambda f: [
        {
            "total_assets": 1_000_000_000.0,
            "total_liabilities": 0.0,
            "total_equity": 0.0,
            "balance_check": 1_000_000_000.0,
        }
    ]
    try:
        issues = kernel.audit_economic_coherence(facts)
    finally:
        kernel.project_balance_sheet = original
    assert any("does not balance" in issue for issue in issues), issues


def test_workbook_balance_sheet_balances_when_recalculated() -> None:
    """The recalculated workbook BS must balance — A = L + E, every period."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        (Path(tmp) / "recalc").mkdir()
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        ws = wb["BS"]
        row = next(
            (cell.row for r in ws.iter_rows() for cell in r
             if cell.value == "Balance check"),
            None,
        )
        assert row is not None, "BS sheet missing 'Balance check' row"
        first_col = source_plan.START_PERIOD_COL
        facts = kernel.derive_source_facts(SAAS_STORY)
        for idx in range(len(facts.years)):
            check = ws.cell(row, first_col + idx).value
            assert check is not None and abs(float(check)) <= 1.0, (
                f"period {idx}: workbook balance sheet is off by {check}"
            )


def test_strict_audit_fails_on_economic_incoherence() -> None:
    """An economically incoherent plan must fail the strict-audit gate (rc 2)."""
    original = build_model.ek.audit_economic_coherence
    build_model.ek.audit_economic_coherence = lambda facts: ["forced economic issue"]
    try:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "m.xlsx"
            rc = build_model._main(["--output", str(out), "--strict-audit"])
            assert rc == 2, f"strict audit did not fail on an economic issue (rc={rc})"
    finally:
        build_model.ek.audit_economic_coherence = original


def test_maturity_cue_after_the_figure_is_detected() -> None:
    """A target / maturity cue marks the figure on either side of it.

    When the cue trails the figure ("25,000 operating units at maturity")
    the extractor must still prefer that maturity figure over an earlier
    interim figure whose cue happens to lead ("targets ... 3,000 units").
    Reading only the lead read the interim 3,000 as the plan's target and
    understated the model ~8x against its own stated maturity.
    """
    profile = kernel.profile_for_text("hardware robot manufacturing plan")
    text = (
        "Phase 1 targets cumulative deployments of 3,000 units in three "
        "years and about 25,000 operating units at maturity."
    )
    assert kernel.extract_target_units(text, profile) == 25_000, (
        "a maturity figure cued only by trailing text was missed"
    )
    # A trailing cue must not be misattributed across a clause boundary: the
    # cue after the ';' belongs to 2,500, not to the 1万台 before it.
    cross_clause = "Today 1万台 operate; we target 2,500 units by year five."
    assert kernel.extract_target_units(cross_clause, profile) == 2_500, (
        "a trailing cue leaked across a clause boundary onto an earlier figure"
    )
    # A prefix cue ("target") points forward to a *later* figure. When it sits
    # in an earlier figure's trailing window — no punctuation between them — a
    # contraction plan must still take the stated target, not the larger
    # current figure.
    contraction = "Currently 10,000 units and we target 2,500 units by year five."
    assert kernel.extract_target_units(contraction, profile) == 2_500, (
        "a prefix cue trailing an earlier figure was misread as its target"
    )


def test_cap_table_exit_waterfall_uses_post_round_cap_table() -> None:
    """The exit waterfall must distribute proceeds over the post-round FDSO.

    Running it on the pre-round input (empty preferred — converted SAFEs and the
    new priced round absent) over-credited founders: ~95% of proceeds against a
    post-round founder FDSO near 54%.
    """
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "cap.xlsx"
        build_model.build_model(None, out, mode="cap_table")
        wb = load_workbook(out, data_only=True)
        ws = wb["Ownership"]
        header = next(
            (cell for r in ws.iter_rows() for cell in r if cell.value == "Founder %"),
            None,
        )
        assert header is not None, "Exit Waterfall 'Founder %' column not found"
        # Collect contiguous scenario rows below the header until the column ends.
        founder_pcts = []
        for row in range(header.row + 1, ws.max_row + 1):
            value = ws.cell(row, header.column).value
            if not isinstance(value, (int, float)):
                break
            founder_pcts.append(value)
        assert founder_pcts, "no exit-waterfall founder-% scenarios found"
        # Post-round founders hold ~54-60% of the exit-relevant FDSO; a value
        # near 95% means the waterfall ran on the pre-round cap table.
        for pct in founder_pcts:
            assert 0.40 <= pct <= 0.70, (
                f"exit-waterfall founder % {pct:.3f} is inconsistent with the "
                f"post-round cap table"
            )


def test_dcf_forecast_fcf_is_not_floored_at_zero() -> None:
    """The explicit-period FCF sum must reflect real burn, not be floored at 0.

    Flooring `SUM(forecast FCF)` at zero discarded the cash a startup actually
    burns and, with a near-zero terminal value, drove DCF enterprise value to 0.
    """
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "dcf.xlsx"
        build_model.build_model(None, out, mode="dcf_only")
        wb = load_workbook(out, data_only=False)
        ws = wb["Valuation & Exit"]
        rows = [
            cell.row for r in ws.iter_rows() for cell in r
            if cell.value in ("PV of forecast FCF", "PV of FCF (per FY)")
        ]
        assert rows, "Valuation & Exit sheet missing the forecast-FCF PV rows"
        formulas = [
            ws.cell(row, c).value
            for row in rows
            for c in range(4, ws.max_column + 1)
            if isinstance(ws.cell(row, c).value, str) and ws.cell(row, c).value.startswith("=")
        ]
        assert formulas, "no forecast-FCF formulas found"
        assert not any("MAX(0" in f for f in formulas), (
            "forecast FCF is still floored at zero"
        )


def test_dcf_enterprise_value_is_positive_and_method_consistent() -> None:
    """DCF EV must be a positive, method-consistent figure, not ~0."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "dcf.xlsx"
        build_model.build_model(None, out, mode="dcf_only")
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "dcf.xlsx", data_only=True)
        ws = wb["Valuation & Exit"]

        def final_value(label: str) -> float:
            row = next(
                (cell.row for r in ws.iter_rows() for cell in r
                 if cell.value == label),
                None,
            )
            assert row is not None, f"Valuation & Exit sheet missing '{label}' row"
            # single-value rows carry their figure in column D
            return float(ws.cell(row, 4).value)

        dcf_ev = final_value("DCF EV")
        primary_ev = final_value("Primary-method EV")
        assert dcf_ev > 0, f"DCF EV is non-positive: {dcf_ev}"
        # An exit-multiple DCF should be the same order of magnitude as the
        # multiple-based EV — not collapsed near zero.
        assert dcf_ev >= 0.1 * primary_ev, (
            f"DCF EV {dcf_ev:,.0f} is implausibly small vs primary-method EV "
            f"{primary_ev:,.0f}"
        )


def test_funding_plan_keeps_the_company_solvent() -> None:
    """The sized funding plan must keep projected ending cash non-negative.

    A generated fundraise plan whose own cash goes negative is not
    investor-grade: the equity rounds must be sized against the projected
    free cash flow, not a fixed heuristic.
    """
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        projection = kernel.project_plan_free_cash_flow(facts)
        for i, period in enumerate(projection):
            assert period["ending_cash"] >= 0, (
                f"{name} period {i}: insolvent, ending cash "
                f"¥{period['ending_cash']:,.0f}"
            )


def test_seed_round_is_positive_across_archetypes() -> None:
    """Period 0 must always carry a positive round (the current raise)."""
    for name, story in ARCHETYPES.items():
        facts = kernel.derive_source_facts(story)
        assert facts.equity_raise_yen[0] > 0, f"{name}: no period-0 round"


def test_stated_raise_size_and_post_money_are_honored() -> None:
    """A narrative that states the round size and post-money valuation must
    drive the cap table — not an auto-sized round that contradicts it."""
    story = (
        "# Seed SaaS plan\n\nA B2B SaaS company. Per-seat subscription at "
        "$80 per seat per month, $12M ARR at maturity. We are raising a "
        "$3.5M seed round at a $14M post-money valuation. 5-year plan. "
        "Source: investor memo.\n"
    )
    facts = kernel.derive_source_facts(story)
    assert facts.equity_raise_yen[0] == 3_500_000, (
        f"stated $3.5M raise was not honored: {facts.equity_raise_yen[0]}"
    )
    assert facts.post_money_yen[0] == 14_000_000, (
        f"stated $14M post-money was not honored: {facts.post_money_yen[0]}"
    )
    # A dollar figure in a yen-denominated plan must not be applied.
    assert kernel.extract_raise_size("Raising a $5M seed round.", "JPY") == 0, (
        "a $-denominated raise leaked into a yen plan"
    )
    # Timeline prose must not be misread: the "m" in "month" is not "5M".
    assert kernel.extract_raise_size(
        "Raising a 5 month runway seed round.", "USD"
    ) == 0, "a bare timeline number was misread as a raise size"
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("LibreOffice not installed; skipping workbook recalc check")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "raise.md"
        out = Path(tmp) / "raise.xlsx"
        src.write_text(story, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        ws = load_workbook(Path(tmp) / "recalc" / "raise.xlsx", data_only=True)["Cap Table"]
        own_row = _row_for_label(ws, "New investor ownership")
        # Cap Table rounds register: column F = founding, column G = round 1
        # (C label / D driver / E unit keep the v2 register contract).
        ownership = ws.cell(own_row, 7).value
        # $3.5M into a $14M post-money is a 25% round.
        assert ownership is not None and abs(float(ownership) - 0.25) < 0.01, (
            f"new-investor ownership {ownership} does not reflect the stated "
            f"$3.5M / $14M terms"
        )


def test_workbook_recalc_shows_no_insolvency() -> None:
    """The recalculated workbook must not go cash-negative in any period."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        ws = wb["CF"]
        first_col = source_plan.START_PERIOD_COL
        facts = kernel.derive_source_facts(SAAS_STORY)
        cash_row = next(
            (cell.row for row in ws.iter_rows() for cell in row
             if cell.value == "Ending cash"),
            None,
        )
        assert cash_row is not None, "CF sheet missing 'Ending cash' row"
        # Allow a small tolerance for kernel/workbook working-capital timing
        # rounding; reject genuine insolvency.
        tolerance = 0.02 * 3_500_000_000
        for idx in range(len(facts.years)):
            cash = ws.cell(cash_row, first_col + idx).value
            assert cash is not None and float(cash) >= -tolerance, (
                f"period {idx}: workbook ending cash ¥{cash:,.0f} is insolvent"
            )


def test_summarize_comps_aggregates_live_peer_margins() -> None:
    """Peer operating-margin bands are derived from the live fetch, not constants."""
    import live_comps as lc

    comps = [
        lc.PublicComp(
            ticker=ticker, name=ticker, currency="USD", market_cap=1e11,
            enterprise_value=1e11, revenue_multiple=8.0, ebitda_multiple=20.0,
            source_url="x", as_of_date="2026-05-17", status="current",
            gross_margin=gross, ebitda_margin=ebitda,
        )
        for ticker, gross, ebitda in [("A", 0.80, 0.25), ("B", 0.75, 0.20), ("C", 0.84, 0.30)]
    ]
    result = lc.summarize_comps(comps)
    assert result.gross_margin_median == 0.80
    assert result.gross_margin_low == 0.75
    assert result.gross_margin_high == 0.84
    assert result.ebitda_margin_median == 0.25


def _saas_facts_with_peer_comps():
    from dataclasses import replace

    facts = kernel.derive_source_facts(SAAS_STORY)
    return replace(
        facts,
        live_comps=[
            {"ticker": "CRM", "name": "Salesforce", "status": "current",
             "gross_margin": 0.77, "ebitda_margin": 0.30, "as_of_date": "2026-05-17",
             "revenue_multiple": 6.0, "ebitda_multiple": 18.0, "company_type": "public",
             "source_type": "public market data"},
            {"ticker": "NOW", "name": "ServiceNow", "status": "current",
             "gross_margin": 0.81, "ebitda_margin": 0.34, "as_of_date": "2026-05-17",
             "revenue_multiple": 12.0, "ebitda_multiple": 28.0, "company_type": "public",
             "source_type": "public market data"},
        ],
    )


def test_kpi_sheet_compares_plan_kpis_to_live_peers() -> None:
    """The Evidence sheet juxtaposes plan KPIs with live-fetched peer margins."""
    wb = source_plan.build_source_plan_workbook_from_facts(_saas_facts_with_peer_comps())
    ws = wb["Evidence"]
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("live public peers" in t for t in texts), "peer comparison section missing"
    # The peer median for the injected SaaS comps (0.77, 0.81) must appear live,
    # not a hard-coded band.
    numbers = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))]
    assert any(abs(n - 0.79) < 1e-6 for n in numbers), "live peer median not written"
    # No cross-sheet reference: the block must survive focused-bundle pruning.
    section_rows = [
        c.row for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and "live public peers" in c.value
    ]
    formulas = [
        c.value for row in ws.iter_rows() for c in row
        if c.row > section_rows[0] and isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert formulas, "peer comparison has no live formula"
    assert all("!" not in f for f in formulas), "peer block leaks a cross-sheet reference"


def test_peer_comparison_status_recalculates_without_error() -> None:
    """The plan-vs-peer status formula must resolve cleanly when recalculated."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "peers.xlsx"
        source_plan.build_source_plan_workbook_from_facts(_saas_facts_with_peer_comps()).save(out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "peers.xlsx", data_only=True)
        ws = wb["Evidence"]
        statuses = {"below", "above", "within"}
        found = [
            c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value in statuses
        ]
        assert found, "no resolved plan-vs-peer status found on the Evidence sheet"


def test_archetype_workbooks_pass_strict_audit() -> None:
    """Each archetype builds end to end and clears the strict audit gate."""
    with tempfile.TemporaryDirectory() as tmp:
        for name, story in ARCHETYPES.items():
            src = Path(tmp) / f"{name}.md"
            out = Path(tmp) / f"{name}.xlsx"
            src.write_text(story, encoding="utf-8")
            rc = build_model._main(
                ["--source-md", str(src), "--output", str(out), "--strict-audit"]
            )
            assert rc == 0, f"{name}: strict audit failed (rc={rc})"


def test_workbook_recalc_reconciles_with_kernel_gross_margin() -> None:
    """The recalculated workbook gross margin must match the kernel's intent."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        ws = wb["Cost Build"]
        gm_row = next(
            (cell.row for row in ws.iter_rows() for cell in row
             if cell.value == "Gross margin"),
            None,
        )
        assert gm_row is not None, "Cost Build sheet missing 'Gross margin' row"
        first_col = source_plan.START_PERIOD_COL
        facts = kernel.derive_source_facts(SAAS_STORY)
        for idx in range(len(facts.years)):
            recalced = ws.cell(gm_row, first_col + idx).value
            assert recalced is not None, f"period {idx}: gross margin not recalculated"
            assert abs(float(recalced) - facts.target_gross_margin[idx]) <= 0.04, (
                f"period {idx}: workbook GM {recalced} vs target "
                f"{facts.target_gross_margin[idx]}"
            )


def test_burn_multiple_does_not_explode_in_period_zero() -> None:
    """Period-0 burn multiple uses the full period-0 revenue as net-new
    revenue, not a zero year-over-year difference that forces the
    denominator and explodes the ratio to hundreds of millions."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        ws = wb["Summary"]
        first_col = source_plan.START_PERIOD_COL
        burn_row = next(
            (cell.row for row in ws.iter_rows() for cell in row
             if cell.value == "Burn multiple"),
            None,
        )
        assert burn_row is not None, "Summary sheet missing 'Burn multiple' row"
        facts = kernel.derive_source_facts(SAAS_STORY)
        for idx in range(len(facts.years)):
            val = ws.cell(burn_row, first_col + idx).value
            if isinstance(val, str):  # "N/A" — no revenue growth to fund
                continue
            # A deliberately loose regression bound: it must reliably catch
            # the ~415,000,000x denominator-collapse explosion. A tighter
            # business band (burn multiple is "good" below ~2x) is not
            # asserted here because an early-stage period can legitimately
            # carry a high ratio; that judgement belongs to the audit, not
            # this guard.
            assert val is not None and abs(float(val)) < 50.0, (
                f"period {idx}: burn multiple {val} is not a sane ratio"
            )


def _row_for_label(ws, label: str) -> int:
    """Return the row index whose first labelled cell matches `label`."""
    row = next(
        (cell.row for r in ws.iter_rows() for cell in r if cell.value == label),
        None,
    )
    assert row is not None, f"{ws.title} sheet missing '{label}' row"
    return row


def test_runway_months_is_capped_at_99() -> None:
    """A near-zero-negative free cash flow against a large cash balance must
    not blow runway to thousands of months — the formula caps it at 99."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "runway.xlsx"
        facts = kernel.derive_source_facts(SAAS_STORY)
        wb = source_plan.build_source_plan_workbook_from_facts(facts)
        ws = wb["CF"]
        first_col = source_plan.START_PERIOD_COL
        # Force a tiny-negative free cash flow against a large cash balance:
        # the uncapped ratio would yield tens of millions of months. Rows are
        # resolved by label so a CF layout shift cannot silently mis-target.
        fcf_row = _row_for_label(ws, "Free cash flow")
        cash_row = _row_for_label(ws, "Ending cash")
        ws.cell(fcf_row, first_col).value = -1000
        ws.cell(cash_row, first_col).value = 5_000_000_000
        wb.save(out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb2 = load_workbook(Path(tmp) / "recalc" / "runway.xlsx", data_only=True)
        ws2 = wb2["CF"]
        runway_row = _row_for_label(ws2, "Runway months")
        # The MIN(99,...) cap is applied uniformly, so verify every period.
        for idx in range(len(facts.years)):
            runway = ws2.cell(runway_row, first_col + idx).value
            assert runway is not None and float(runway) <= 99.0, (
                f"period {idx}: runway {runway} months is not capped at 99"
            )


def test_kpi_dashboard_carries_vc_decision_metrics() -> None:
    """The KPI dashboard computes the core VC capital-efficiency and
    retention metrics as live cells that recalculate without error."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        ws = wb["Summary"]
        first_col = source_plan.START_PERIOD_COL
        facts = kernel.derive_source_facts(SAAS_STORY)
        assert any(
            c.value == "VC decision metrics"
            for row in ws.iter_rows() for c in row
        ), "Summary sheet missing the 'VC decision metrics' section"
        metrics = [
            "Rule of 40", "Net revenue retention",
            "Customer acquisition cost", "CAC payback", "Magic number",
        ]
        for label in metrics:
            row = _row_for_label(ws, label)
            for idx in range(len(facts.years)):
                val = ws.cell(row, first_col + idx).value
                assert val is not None, f"{label} period {idx} did not resolve"
                if isinstance(val, str):
                    assert val in ("N/A", "-"), (
                        f"{label} period {idx}: unexpected text {val!r}"
                    )
                else:
                    float(val)  # numeric, finite


def test_rule_of_40_is_growth_plus_ebitda_margin() -> None:
    """Rule of 40 on the KPI sheet equals revenue growth plus EBITDA margin
    — it is composed from the live model cells, not a free-standing input."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("soffice/libreoffice not available")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "saas.md"
        out = Path(tmp) / "saas.xlsx"
        src.write_text(SAAS_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "saas.xlsx", data_only=True)
        summary = wb["Summary"]
        first_col = source_plan.START_PERIOD_COL
        r40 = _row_for_label(summary, "Rule of 40")
        growth = _row_for_label(summary, "Revenue growth (YoY)")
        ebitda_margin = _row_for_label(summary, "EBITDA margin")
        facts = kernel.derive_source_facts(SAAS_STORY)
        checked = 0
        for idx in range(len(facts.years)):
            got = summary.cell(r40, first_col + idx).value
            growth_val = summary.cell(growth, first_col + idx).value
            margin_val = summary.cell(ebitda_margin, first_col + idx).value
            # A clean failure if the recalc did not resolve a cell, rather
            # than a confusing TypeError from float(None).
            assert None not in (got, growth_val, margin_val), (
                f"period {idx}: a Rule of 40 input did not recalculate "
                f"(rule={got}, growth={growth_val}, margin={margin_val})"
            )
            if growth_val == "-":
                # The first fiscal year has no prior-year base; the composed
                # metric must mirror that sentinel rather than fake a number.
                assert got == "-", f"period {idx}: Rule of 40 {got!r} on '-' growth"
                continue
            expected = float(growth_val) + float(margin_val)
            assert abs(float(got) - expected) < 1e-6, (
                f"period {idx}: Rule of 40 {got} != growth + EBITDA margin "
                f"{expected}"
            )
            checked += 1
        assert checked >= len(facts.years) - 1, (
            "Rule of 40 resolved on too few periods to prove the composition"
        )


def test_kpi_dashboard_omits_vc_block_for_an_ambiguous_mechanic() -> None:
    """The VC metrics block is gated: an ambiguous-mechanic plan gets the
    generic KPI set and keeps the original (unshifted) dashboard layout."""
    story = (
        "# Ambiguous plan\n\nA new venture with unclear revenue mechanics "
        "and weak evidence. 5-year plan, seed round. Source: management memo."
    )
    facts = kernel.derive_source_facts(story)
    assert kernel.mechanic_key(facts) == "generic"
    ws = source_plan.build_source_plan_workbook_from_facts(facts)["Summary"]
    labels = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str)
    ]
    assert "VC decision metrics" not in labels, (
        "VC metrics section leaked into an ambiguous-mechanic plan"
    )
    for metric in ("Rule of 40", "Customer acquisition cost", "Magic number"):
        assert metric not in labels, (
            f"{metric} leaked into an ambiguous-mechanic plan"
        )
    # The generic KPI block and the consolidated master check still render.
    assert "KPI" in labels, "generic KPI section missing from Summary"
    assert "Master check" in labels, "master check missing from Summary"


PRE_REVENUE_STORY = """# Helios — pre-revenue deep-tech plan

Helios is a pre-revenue deep-tech company pursuing compact fusion-energy
milestones. This is a pre-revenue, milestone-driven company: there is no
product revenue during the plan; value is created by hitting technical
milestones and prototype proof points. The company is R&D intensive and
burns cash against a seed round and grant commitments over a 5-year horizon.

Source: technical roadmap memo, grant award letter.
"""


def test_pre_revenue_plan_has_no_product_revenue() -> None:
    """A plan classified pre-revenue / milestone carries zero product revenue
    across every period — it is a burn / runway model, not a revenue ramp."""
    facts = kernel.derive_source_facts(PRE_REVENUE_STORY)
    assert kernel.mechanic_key(facts) == "pre_revenue_milestone", (
        "story did not resolve to the pre-revenue profile"
    )
    revenue = kernel.plan_revenue_series(
        facts.new_units, facts.gmv_yen, facts.monthly_price_yen,
        facts.take_rate, facts.other_revenue_share,
    )
    assert all(r == 0 for r in revenue), (
        f"pre-revenue plan still carries product revenue: {revenue}"
    )
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        return
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "pre.md"
        out = Path(tmp) / "pre.xlsx"
        src.write_text(PRE_REVENUE_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        ws = load_workbook(Path(tmp) / "recalc" / "pre.xlsx", data_only=True)["P&L"]
        first_col = source_plan.START_PERIOD_COL
        rev_row = _row_for_label(ws, "Total revenue")
        for idx in range(len(facts.years)):
            value = ws.cell(rev_row, first_col + idx).value
            assert value in (0, 0.0, None), (
                f"period {idx}: P&L total revenue {value} is not zero"
            )


def test_pre_revenue_audit_tolerates_zero_revenue() -> None:
    """Zero product revenue is economically coherent for a pre-revenue plan:
    the audit must not flag it, and strict audit must pass."""
    facts = kernel.derive_source_facts(PRE_REVENUE_STORY)
    issues = kernel.audit_economic_coherence(facts)
    assert not any("non-positive revenue" in issue for issue in issues), (
        f"pre-revenue plan wrongly flagged for zero revenue: {issues}"
    )
    assert issues == [], f"pre-revenue plan is not economically coherent: {issues}"
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "pre.md"
        out = Path(tmp) / "pre.xlsx"
        src.write_text(PRE_REVENUE_STORY, encoding="utf-8")
        rc = build_model._main(
            ["--source-md", str(src), "--output", str(out), "--strict-audit"]
        )
        assert rc == 0, f"pre-revenue plan failed strict audit (rc={rc})"


def test_pre_revenue_plan_books_no_cogs() -> None:
    """A pre-revenue plan sells nothing, so the workbook must book zero cost
    of goods sold — not a negative gross profit on zero revenue."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("LibreOffice not installed; skipping workbook recalc check")
    facts = kernel.derive_source_facts(PRE_REVENUE_STORY)
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "pre.md"
        out = Path(tmp) / "pre.xlsx"
        src.write_text(PRE_REVENUE_STORY, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "pre.xlsx", data_only=True)
        first_col = source_plan.START_PERIOD_COL
        for sheet in ("Cost Build", "P&L"):
            ws = wb[sheet]
            gp_row = _row_for_label(ws, "Gross profit")
            for idx in range(len(facts.years)):
                gp = ws.cell(gp_row, first_col + idx).value
                assert gp in (0, 0.0, None), (
                    f"{sheet} period {idx}: gross profit {gp} — a pre-revenue "
                    f"plan must book zero COGS, not a loss on zero revenue"
                )


def test_customer_target_is_the_maturity_figure_not_the_current_count() -> None:
    """'currently 30 customers and target 1,200' must ramp the plan to the
    1,200 maturity target, not freeze it at the current count of 30."""
    story = (
        "# SaaS plan\n\nThe business is a recurring software SaaS platform. "
        "Annual contract value is about ¥1,200,000 per customer. We "
        "currently have 30 customers and target 1,200 customers by year "
        "five. Gross margin target is 80%. 5-year plan, seed round. "
        "Source: board memo.\n"
    )
    assert kernel.extract_target_customers(story) == 1200, (
        "extractor took the current count, not the stated target"
    )
    facts = kernel.derive_source_facts(story)
    assert facts.customers[-1] >= 1000, (
        f"plan ramps customers to {facts.customers[-1]}, not the stated 1,200"
    )


def test_hardware_unit_ramp_honors_the_stated_maturity_target() -> None:
    """'ship 40 units ... target 2,500 units by year five' must ramp units to
    the 2,500 target, not collapse on the year-one figure of 40."""
    story = (
        "# Robotics plan\n\nThe company manufactures asset-heavy industrial "
        "robots (hardware / manufacturing). We expect to ship 40 units in "
        "year one and target 2,500 units shipped by year five. Gross margin "
        "target is 38%. 5-year plan, Series A. Source: management forecast.\n"
    )
    profile = kernel.profile_for_text(story)
    assert kernel.extract_target_units(story, profile) == 2500, (
        "extractor took the year-one figure, not the stated target"
    )
    facts = kernel.derive_source_facts(story)
    assert sum(facts.new_units) >= 2000, (
        f"hardware plan ships {sum(facts.new_units)} units, not the stated 2,500"
    )
    # The ramp must be non-degenerate growth, not e.g. [50, 250, 4, 10, 0].
    assert facts.new_units[-1] > facts.new_units[0], (
        f"hardware unit ramp is degenerate: {facts.new_units}"
    )


def test_unit_target_extraction_handles_mixed_scale_phrasing() -> None:
    """The 万台 (x10,000) multiplier normalises the match it came from, never
    the document; without a target cue the largest match wins."""
    profile = kernel.profile_for_text("hardware robot manufacturing plan")
    # No target cue: 1万台 = 10,000 units is the larger match and wins via
    # the fallback — it must not be corrupted to 2,500 x 10,000.
    mixed = "The fleet has 1万台 operating and the plan covers 2,500 units."
    assert kernel.extract_target_units(mixed, profile) == 10_000
    # A target cue overrides size: the cued 2,500 wins over the 万台 figure.
    cued = "Today 1万台 operate; we target 2,500 units by year five."
    assert kernel.extract_target_units(cued, profile) == 2500


HARDWARE_UNIT_SALE_STORY = """# Robotics hardware plan

The company manufactures asset-heavy industrial robots (hardware /
manufacturing). The unit price is ¥9,000,000 per robot. We expect to ship
40 units in year one and target 600 units shipped by year five. Gross margin
target is 38%. 5-year plan, Series B. Source: management forecast.
"""


def test_hardware_revenue_is_units_times_sale_price() -> None:
    """A hardware unit-sale plan recognises revenue as units shipped x the
    one-time sale price — not as monthly recurring billing x 12."""
    facts = kernel.derive_source_facts(HARDWARE_UNIT_SALE_STORY)
    assert kernel.mechanic_key(facts) == "hardware_asset_heavy"
    assert facts.revenue_mode == "unit_sale"
    revenue = kernel.plan_revenue_series(
        facts.new_units, facts.gmv_yen, facts.monthly_price_yen,
        facts.take_rate, facts.other_revenue_share, facts.revenue_mode,
    )
    for idx, (units, price) in enumerate(
        zip(facts.new_units, facts.monthly_price_yen)
    ):
        unit_sale = units * price
        # Revenue is unit sales plus a modest other-revenue uplift — within
        # ~30% of units x price, never the ~10x a recurring x 12 model gives.
        assert unit_sale <= revenue[idx] <= unit_sale * 1.3, (
            f"period {idx}: revenue {revenue[idx]:,.0f} is not a unit-sale "
            f"figure (units x price = {unit_sale:,.0f})"
        )
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("LibreOffice not installed; skipping workbook recalc check")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "hw.md"
        out = Path(tmp) / "hw.xlsx"
        src.write_text(HARDWARE_UNIT_SALE_STORY, encoding="utf-8")
        rc = build_model._main(
            ["--source-md", str(src), "--output", str(out), "--strict-audit"]
        )
        assert rc == 0, f"hardware plan failed strict audit (rc={rc})"
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        ws = load_workbook(Path(tmp) / "recalc" / "hw.xlsx", data_only=True)["P&L"]
        first_col = source_plan.START_PERIOD_COL
        rev_row = _row_for_label(ws, "Total revenue")
        period_0 = ws.cell(rev_row, first_col).value
        # 40 units x ¥9M ≈ ¥360M (plus other-revenue uplift); the recurring
        # model would have booked it ~10x higher.
        assert 3.0e8 <= float(period_0) <= 5.5e8, (
            f"hardware period-0 revenue {period_0} is not a unit-sale figure"
        )


def test_monthly_priced_hardware_stays_recurring() -> None:
    """A hardware narrative that prices per unit *per month* is a recurring
    (lease / RaaS) plan — the per-month cue must keep it out of unit_sale."""
    story = (
        "# Robotics deployment plan\n\nThe company deploys asset-heavy "
        "industrial robots. The price per robot is ¥320,000 per month. We "
        "expect 40 units in year one and target 600 units by year five. "
        "Gross margin target is 38%. 5-year plan. Source: management memo.\n"
    )
    facts = kernel.derive_source_facts(story)
    assert kernel.mechanic_key(facts) == "hardware_asset_heavy"
    assert facts.revenue_mode == "recurring", (
        "monthly-priced hardware was misclassified as a one-time unit sale"
    )


def test_hardware_unit_sale_survives_an_attach_subscription() -> None:
    """A hardware plan that sells units outright stays a unit-sale business
    even when it mentions an attach support subscription — one recurring
    word must not flip the revenue model, nor steal the unit price."""
    story = (
        "# Robotics hardware plan\n\nThe company manufactures industrial "
        "robots. Each unit sells for $48,000, with an attached firmware-"
        "support subscription at $1,800 per unit per year. We ship 50 units "
        "in year one and target 600 units by year five. Gross margin target "
        "is 38%. 5-year plan, Series B. Source: management forecast.\n"
    )
    facts = kernel.derive_source_facts(story)
    assert kernel.mechanic_key(facts) == "hardware_asset_heavy"
    assert facts.revenue_mode == "unit_sale", (
        "an attach subscription flipped the unit-sale plan to recurring"
    )
    assert facts.monthly_price_yen[0] == 48000, (
        f"the attach price was taken instead of the $48,000 sale price: "
        f"{facts.monthly_price_yen[0]}"
    )
    revenue = kernel.plan_revenue_series(
        facts.new_units, facts.gmv_yen, facts.monthly_price_yen,
        facts.take_rate, facts.other_revenue_share, facts.revenue_mode,
    )
    for idx, (units, price) in enumerate(
        zip(facts.new_units, facts.monthly_price_yen)
    ):
        unit_sale = units * price
        # The 1.3x headroom is the other-revenue uplift (attach revenue
        # layered on the unit sale) — not a recurring x12 figure.
        assert unit_sale <= revenue[idx] <= unit_sale * 1.3, (
            f"period {idx}: revenue {revenue[idx]:,.0f} is not a unit-sale "
            f"figure (units x price = {unit_sale:,.0f})"
        )


def test_marketplace_kpi_uses_a_transaction_lens() -> None:
    """A GMV / take-rate marketplace shows transaction economics on the KPI
    dashboard — not per-seat subscription rows that read price 0, negative
    unit gross profit, and "N/A" payback."""
    story = (
        "# Marketplace plan\n\nThe company runs a two-sided marketplace. "
        "Revenue is a take rate of 18% on gross merchandise value. Phase 1 "
        "targets ¥120億 GMV at maturity. 5-year plan, seed round. "
        "Source: cohort analysis memo.\n"
    )
    facts = kernel.derive_source_facts(story)
    assert kernel.mechanic_key(facts) == "marketplace"
    ws = source_plan.build_source_plan_workbook_from_facts(facts)["Summary"]
    # Restrict to the KPI block (between the "KPI" section band and the
    # scenario block) so labels elsewhere cannot mask a missing metric row.
    kpi_start = _row_for_label(ws, "KPI")
    kpi_end = _row_for_label(ws, "Scenario comparison")
    metric_labels = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and kpi_start < c.row < kpi_end
    ]
    for per_seat in (
        "Monthly price / unit", "Monthly unit gross profit", "Unit payback",
    ):
        assert per_seat not in metric_labels, (
            f"{per_seat} leaked into a marketplace KPI dashboard"
        )
    for metric in ("Take rate (FY end)", "GMV per customer", "Contribution margin"):
        assert metric in metric_labels, (
            f"{metric} missing from the marketplace KPI dashboard"
        )
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        pytest.skip("LibreOffice not installed; skipping workbook recalc check")
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "mkt.md"
        out = Path(tmp) / "mkt.xlsx"
        src.write_text(story, encoding="utf-8")
        source_plan.build_source_plan_workbook(src, out)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        recalced = load_workbook(Path(tmp) / "recalc" / "mkt.xlsx", data_only=True)["Summary"]
        first_col = source_plan.START_PERIOD_COL
        cm_row = _row_for_label(recalced, "Contribution margin")
        for idx in range(len(facts.years)):
            margin = recalced.cell(cm_row, first_col + idx).value
            assert margin is not None and 0.0 < float(margin) < 1.0, (
                f"period {idx}: marketplace contribution margin {margin} is "
                f"not a sane positive fraction"
            )


def test_structured_yaml_revenue_tracks_the_stated_customer_count() -> None:
    """A structured `--input` YAML must drive the full derivation pipeline.

    The structured path once shallow-merged YAML values onto a regex-derived
    base SourceFacts, leaving every *derived* quantity stale: revenue ran ~6x
    customers x ACV, new_units kept the regex ramp, gmv / take_rate kept stale
    base values, the cost stack was calibrated for the regex base (gross
    margin reading several hundred percent negative), and the equity round
    was sized for the wrong plan (insolvency). This pins the structured
    drivers flowing through derivation: maturity revenue tracks
    customers x monthly price x 12, and the plan clears the economic audit.
    """
    customers = [18, 60, 130, 220, 320]
    monthly_price = 2000
    yaml_text = (
        "company: Helios SaaS\n"
        "mechanics: recurring software (SaaS)\n"
        "currency: USD\n"
        "grain: annual\n"
        "periods: 5\n"
        f"customers: {customers}\n"
        f"monthly_price_yen: [{', '.join([str(monthly_price)] * 5)}]\n"
        "target_gross_margin: [0.80, 0.80, 0.81, 0.82, 0.82]\n"
    )
    expected_mature = customers[-1] * monthly_price * 12
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "helios.yaml"
        out = Path(tmp) / "helios.xlsx"
        ypath.write_text(yaml_text, encoding="utf-8")
        raw = build_model.load_yaml(ypath)
        facts = kernel.derive_source_facts_from_mapping(raw)
        # The structured customer count reaches the derivation pipeline: the
        # ending installed base tracks the stated series, not the regex ramp.
        ending = kernel.ending_units(facts.new_units)
        assert abs(ending[-1] - customers[-1]) <= 0.10 * customers[-1], (
            f"derived ending customers {ending[-1]} ignores the stated "
            f"maturity count {customers[-1]}"
        )
        revenue = kernel.plan_revenue_series(
            facts.new_units, facts.gmv_yen, facts.monthly_price_yen,
            facts.take_rate, facts.other_revenue_share, facts.revenue_mode,
        )
        assert abs(revenue[-1] - expected_mature) <= 0.20 * expected_mature, (
            f"maturity revenue {revenue[-1]:,.0f} does not track customers x "
            f"price x 12 = {expected_mature:,.0f} (structured drivers stale)"
        )
        # The cost stack and the funding plan are derived from the structured
        # drivers, so the plan is economically coherent.
        assert kernel.audit_economic_coherence(facts) == [], (
            "structured-YAML plan is not economically coherent"
        )
        # The full build path clears the strict economic audit.
        rc = build_model._main(
            ["--input", str(ypath), "--output", str(out), "--strict-audit"]
        )
        assert rc == 0, "structured-YAML build failed --strict-audit"
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if soffice is None:
            pytest.skip("LibreOffice not installed; skipping workbook recalc check")
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(Path(tmp) / "recalc"), str(out)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(Path(tmp) / "recalc" / "helios.xlsx", data_only=True)
        ws = wb["Revenue Build"]
        rev_row = _row_for_label(ws, "Total revenue")
        first_col = source_plan.START_PERIOD_COL
        for idx, row in enumerate(ws.iter_rows()):
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                assert not cell.value.startswith("#"), (
                    f"recalc error {cell.value} in Revenue Build"
                )
        recalced_mature = ws.cell(rev_row, first_col + len(customers) - 1).value
        assert recalced_mature is not None, "Total revenue not recalculated"
        assert abs(float(recalced_mature) - expected_mature) <= 0.20 * expected_mature, (
            f"recalculated maturity revenue {recalced_mature:,.0f} does not "
            f"track customers x price x 12 = {expected_mature:,.0f}"
        )


# =========================================================================
# Phase 2 — K1 / K2 / K4 / K5 / K7 economic-coherence regression tests
# (spec.md §3). These are red on main today; they go green as Tasks 2.2-2.6
# of docs/sfm-overhaul/plan.md replace the implementations.
# =========================================================================


def test_K1_explicit_customers_dominate_recurring_revenue_over_new_units() -> None:
    """K1 — input fidelity: when the structured input pins both `customers`
    and `new_units` to genuinely different series, the recurring revenue
    must derive from `customers` (the user-stated installed base), not from
    `ending_units(new_units)` (the implicit ramp).

    Currently `plan_revenue_series` averages `ending_units(new_units)` and
    multiplies by price × 12 — so with new_units = [5] × 5, the recurring
    base collapses to ~25, even though the user said 5,000 customers are
    paying at maturity. The fix (Task 2.2) routes facts.customers into
    plan_revenue_series as the installed base when the structured input
    provides it."""
    yaml_text = (
        "periods: 5\n"
        "customers: [1000, 2000, 3000, 4000, 5000]\n"
        "new_units: [5, 5, 5, 5, 5]\n"
        "monthly_price_yen: [10000, 10000, 10000, 10000, 10000]\n"
        "target_gross_margin: [0.7, 0.7, 0.7, 0.7, 0.7]\n"
    )
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "k1.yaml"
        ypath.write_text(yaml_text, encoding="utf-8")
        raw = build_model.load_yaml(ypath)
        facts = kernel.derive_source_facts_from_mapping(raw)

        assert list(facts.customers) == [1000, 2000, 3000, 4000, 5000], (
            "facts.customers must preserve the YAML-stated series"
        )

        revenue = kernel.plan_revenue_series(
            facts.new_units,
            facts.gmv_yen,
            facts.monthly_price_yen,
            facts.take_rate,
            facts.other_revenue_share,
            facts.revenue_mode,
            installed_base=facts.customers,
        )
        # Recurring component at maturity if customers drove the base:
        expected_floor = 5000 * 10000 * 12 * 0.5  # 50% tolerance for averaging
        assert revenue[-1] >= expected_floor, (
            f"K1 violated: with explicit customers=[…,5000] and "
            f"monthly_price=¥10k, recurring revenue at maturity should be "
            f">= {expected_floor:,.0f} JPY; got {revenue[-1]:,.0f}. The "
            f"kernel is using ending_units(new_units) (≈25) instead of "
            f"facts.customers (5,000) as the installed base."
        )


def test_K2_changing_churn_rate_moves_ending_units_and_terminal_arr() -> None:
    """K2 — churn honor: facts.churn_rate must propagate into the fleet
    rollforward so a higher churn rate produces a smaller terminal fleet
    and lower terminal revenue. Today `ending_units` hard-codes the churn
    factor as ``0.02 + idx * 0.005`` and ignores its caller — so changing
    facts.churn_rate has zero effect. The fix (Task 2.2) accepts churn_rate
    as a kwarg to ending_units."""
    new_units = [200, 200, 200, 200, 200]
    low_churn = kernel.ending_units(new_units, churn_rate=0.02)
    high_churn = kernel.ending_units(new_units, churn_rate=0.30)
    assert high_churn[-1] < low_churn[-1] * 0.7, (
        f"K2 violated: ending_units must shrink when churn rises. "
        f"low (0.02): {low_churn[-1]}; high (0.30): {high_churn[-1]}. "
        f"Currently ending_units ignores its churn_rate argument and uses "
        f"a hard-coded `0.02 + idx * 0.005` schedule."
    )


def test_K2_structured_yaml_churn_rate_moves_terminal_customers_and_revenue() -> None:
    """K2 integrated path: when the structured YAML pins only new_units and
    declares a churn_rate, the implied customer rollforward must honor that
    rate. Surfaced by the Phase 2 mid-checkpoint Codex review — the
    `kernel.ending_units(.., churn_rate=…)` unit test passed but
    derive_source_facts_from_mapping built prims.customers via
    `ending_units(prims.new_units)` (no churn_rate), so the full
    structured path silently used the legacy schedule."""
    base_yaml = (
        "periods: 5\n"
        "new_units: [200, 200, 200, 200, 200]\n"
        "monthly_price_yen: [12000, 12000, 12000, 12000, 12000]\n"
        "target_gross_margin: [0.7, 0.7, 0.7, 0.7, 0.7]\n"
    )
    with tempfile.TemporaryDirectory() as tmp:
        low_path = Path(tmp) / "low.yaml"
        high_path = Path(tmp) / "high.yaml"
        low_path.write_text(base_yaml + "churn_rate: 0.02\n", encoding="utf-8")
        high_path.write_text(base_yaml + "churn_rate: 0.30\n", encoding="utf-8")
        low_facts = kernel.derive_source_facts_from_mapping(
            build_model.load_yaml(low_path)
        )
        high_facts = kernel.derive_source_facts_from_mapping(
            build_model.load_yaml(high_path)
        )

    assert low_facts.churn_rate == 0.02
    assert high_facts.churn_rate == 0.30
    assert high_facts.customers[-1] < low_facts.customers[-1] * 0.7, (
        f"K2 violated in structured path: same new_units with churn 0.02 vs 0.30 "
        f"produced terminal customers low={low_facts.customers[-1]} vs "
        f"high={high_facts.customers[-1]}. derive_source_facts_from_mapping must "
        f"thread stated_churn into the implied-customers rollforward."
    )

    low_revenue = kernel.plan_revenue_series(
        low_facts.new_units, low_facts.gmv_yen, low_facts.monthly_price_yen,
        low_facts.take_rate, low_facts.other_revenue_share, low_facts.revenue_mode,
        installed_base=low_facts.customers, churn_rate=low_facts.churn_rate,
    )
    high_revenue = kernel.plan_revenue_series(
        high_facts.new_units, high_facts.gmv_yen, high_facts.monthly_price_yen,
        high_facts.take_rate, high_facts.other_revenue_share, high_facts.revenue_mode,
        installed_base=high_facts.customers, churn_rate=high_facts.churn_rate,
    )
    assert high_revenue[-1] < low_revenue[-1] * 0.8, (
        f"K2 violated: terminal revenue did not respond to churn_rate. "
        f"low: {low_revenue[-1]:,.0f}; high: {high_revenue[-1]:,.0f}."
    )


def test_K4_nol_absorbs_prior_losses_so_cash_tax_does_not_double_dip() -> None:
    """K4 — NOL carryforward: cumulative cash tax must satisfy
    ``cum_cash_tax <= tax_rate × max(0, cum_pre_tax_income)``. A plan that
    loses ¥1B early and earns ¥500M + ¥500M later must report 0 cumulative
    cash tax (NOL balance fully absorbs the gains). Today
    `project_free_cash_flow` taxes each period independently with
    ``tax = max(0, ebt * tax_rate)`` — losses never become tax credits
    for later periods. The fix (Task 2.3) tracks an NOL balance and
    deducts it from taxable income."""
    periods = 5
    # Construct revenue + cost so EBT is roughly: -1000M, +500M, +500M, 0, 0 (¥)
    revenue = [200_000_000, 1_500_000_000, 1_500_000_000, 1_000_000_000, 1_000_000_000]
    # Big upfront OpEx absorbed via headcount avg_comp to make period 0 negative.
    target_gross_margin = [0.50] * periods
    total_headcount = [100, 30, 30, 25, 25]
    avg_comp_yen = [12_000_000] * periods  # 12M/year — big in period 0
    product_headcount = [40, 10, 10, 8, 8]
    sm_pct_revenue = [0.05] * periods
    rd_program_per_product_fte_yen = [2_000_000] * periods
    rd_program_floor_yen = [0] * periods
    ga_pct_revenue = [0.02] * periods
    fixed_ga_yen = [0] * periods
    capex_yen = [0] * periods
    depreciation_life_months = [60] * periods
    debt_raise_yen = [0] * periods
    debt_interest_rate = [0.0] * periods
    ar_days = [0] * periods
    ap_days = [0] * periods
    deferred_revenue_share = [0.0] * periods
    inventory_wip_pct_capex = [0.0] * periods
    tax_rate = [0.30] * periods

    projection = kernel.project_free_cash_flow(
        revenue, target_gross_margin, total_headcount, avg_comp_yen,
        product_headcount, sm_pct_revenue, rd_program_per_product_fte_yen,
        rd_program_floor_yen, ga_pct_revenue, fixed_ga_yen, capex_yen,
        depreciation_life_months, debt_raise_yen, debt_interest_rate,
        ar_days, ap_days, deferred_revenue_share, inventory_wip_pct_capex,
        tax_rate,
    )
    cum_pretax = sum(p["ebt"] for p in projection)
    cum_loss = sum(-p["ebt"] for p in projection if p["ebt"] < 0)
    cum_cash_tax = sum(p["tax"] for p in projection)
    # NOL contract: cumulative cash tax cannot exceed tax_rate × max(0, cumulative EBT).
    max_allowed_tax = 0.30 * max(0.0, cum_pretax)
    assert cum_cash_tax <= max_allowed_tax + 1.0, (
        f"K4 violated: cumulative cash tax {cum_cash_tax:,.0f} JPY exceeds "
        f"the NOL-aware ceiling {max_allowed_tax:,.0f} JPY "
        f"(cum_pretax={cum_pretax:,.0f}, cum_loss={cum_loss:,.0f}). "
        f"project_free_cash_flow taxes each period independently — losses "
        f"never accumulate into an NOL balance that absorbs later gains."
    )


def test_K5_hardware_profile_suppresses_transaction_revenue_and_gmv() -> None:
    """K5 — off-mechanics suppression: a hardware / RaaS plan must carry
    transaction revenue = 0, take rate = 0, and GMV = 0. The current
    `_extract_source_primitives` does set take = 0 for non-marketplace
    profiles, but `gmv_ramp` still produces a non-zero GMV series for the
    hardware profile from `profile.default_gmv_yen`. The Revenue Build
    sheet still renders an off-mechanics 'Transaction revenue' row at
    zero. The fix (Task 2.4) gates the off-mechanics rows on a
    marketplace signal so they are not generated for hardware / RaaS."""
    narrative = (
        "# Hardware co\n"
        "Sells industrial robots at ¥9M per unit. Target 200 units shipped "
        "per year at maturity. Source: pilot installs memo.\n"
    )
    facts = kernel.derive_source_facts(narrative)
    assert facts.mechanics.lower().find("hardware") >= 0 or "asset" in facts.mechanics.lower(), (
        f"prerequisite: hardware narrative must select hardware-asset profile, got {facts.mechanics!r}"
    )
    assert all(v == 0 for v in facts.gmv_yen), (
        f"K5 violated: hardware profile must emit gmv_yen = [0]; got {facts.gmv_yen}"
    )
    assert all(v == 0.0 for v in facts.take_rate), (
        f"K5 violated: hardware profile must emit take_rate = [0]; got {facts.take_rate}"
    )


def test_K7_audit_economic_coherence_detects_K1_customers_drift() -> None:
    """K7 — audit meta: audit_economic_coherence must flag a K1 violation
    (revenue drifting from the stated customers series). Construct a
    facts where customers = [1000,…,5000] but new_units = [5]×5; verify
    the audit emits an issue naming the drift. Today the audit only
    checks gross margin, ending cash, and balance-sheet identities — it
    has no K1/K2/K4/K5 hooks. The fix (Task 2.6) wires those checks."""
    yaml_text = (
        "periods: 5\n"
        "customers: [1000, 2000, 3000, 4000, 5000]\n"
        "new_units: [5, 5, 5, 5, 5]\n"
        "monthly_price_yen: [10000, 10000, 10000, 10000, 10000]\n"
        "target_gross_margin: [0.7, 0.7, 0.7, 0.7, 0.7]\n"
    )
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "k7.yaml"
        ypath.write_text(yaml_text, encoding="utf-8")
        raw = build_model.load_yaml(ypath)
        facts = kernel.derive_source_facts_from_mapping(raw)
        issues = kernel.audit_economic_coherence(facts)
        flagged = [i for i in issues if "customer" in i.lower() or "k1" in i.lower()]
        assert flagged, (
            f"K7 violated: audit_economic_coherence did not flag the K1 "
            f"customers / new_units drift. Returned issues:\n  "
            + "\n  ".join(issues or ["<none>"])
        )


# =========================================================================
# Phase 3 — K3 financing-integration regression tests (spec.md §2 C3 Tier 2 B).
# These pin lease / converts / grants / customer advances / secondary into the
# FCF projection, balance sheet, audit layer, and equity-round sizing.
# =========================================================================


def _k3_base_yaml() -> str:
    """A structured plan with zero debt interest so instrument cash deltas
    are exact (no interest or tax interplay in the comparisons below)."""
    return (
        "periods: 5\n"
        "customers: [200, 400, 700, 1100, 1600]\n"
        "monthly_price_yen: [50000, 50000, 50000, 50000, 50000]\n"
        "target_gross_margin: [0.7, 0.7, 0.7, 0.7, 0.7]\n"
        "debt_interest_rate: [0, 0, 0, 0, 0]\n"
    )


def _k3_facts(extra_yaml: str = ""):
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "k3.yaml"
        ypath.write_text(_k3_base_yaml() + extra_yaml, encoding="utf-8")
        return kernel.derive_source_facts_from_mapping(build_model.load_yaml(ypath))


def test_K3_stated_grants_reduce_auto_sized_equity() -> None:
    """K3 — financing integration: a stated non-dilutive grant schedule must
    reach the equity-round sizing. Today grants_yen is applied as a Step-5
    display override after size_equity_rounds has run, so a plan with ¥2B of
    annual grants raises exactly as much equity as one with none."""
    baseline = _k3_facts()
    granted = _k3_facts(
        "grants_yen: [2000000000, 2000000000, 2000000000, 2000000000, 2000000000]\n"
    )
    assert list(granted.grants_yen) == [2_000_000_000] * 5, (
        "facts.grants_yen must preserve the YAML-stated schedule"
    )
    assert sum(granted.equity_raise_yen) < sum(baseline.equity_raise_yen), (
        f"K3 violated: ¥10B of stated grants did not reduce the auto-sized "
        f"equity plan. baseline equity={baseline.equity_raise_yen}, "
        f"granted equity={granted.equity_raise_yen}. Financing instruments "
        f"must be moved from the post-derivation override into the sizing walk."
    )


def test_K3_projection_ending_cash_reflects_each_instrument() -> None:
    """K3 — the kernel cash walk must mirror the workbook CF sheet: grants,
    convertibles, and lease financing add cash; secondary liquidity uses
    cash; customer advances shift cash timing through working capital
    (deferred-revenue level) without double-counting as a financing inflow."""
    base = _k3_facts()
    base_proj = kernel.project_plan_free_cash_flow(base)
    base_end = base_proj[-1]["ending_cash"]

    for field, series, expected_delta in (
        ("grants_yen", [500_000_000] * 5, 2_500_000_000),
        ("convertibles_yen", [400_000_000, 0, 0, 0, 0], 400_000_000),
        ("lease_financing_yen", [0, 300_000_000, 0, 0, 0], 300_000_000),
        ("secondary_yen", [0, 200_000_000, 0, 0, 0], -200_000_000),
    ):
        mutated = replace(base, **{field: series})
        proj = kernel.project_plan_free_cash_flow(mutated)
        delta = proj[-1]["ending_cash"] - base_end
        assert delta == pytest.approx(expected_delta, rel=1e-6), (
            f"K3 violated: {field}={series} should move terminal ending cash "
            f"by {expected_delta:,} JPY (zero-interest plan); got {delta:,.0f}."
        )

    # Customer advances: +A cash in the receipt period via the deferred /
    # advances liability level, released afterwards — terminal cash returns
    # to baseline, so the instrument shifts timing rather than adding funding.
    advanced = replace(base, customer_advances_yen=[600_000_000, 0, 0, 0, 0])
    proj = kernel.project_plan_free_cash_flow(advanced)
    first_delta = proj[0]["ending_cash"] - base_proj[0]["ending_cash"]
    final_delta = proj[-1]["ending_cash"] - base_end
    assert first_delta == pytest.approx(600_000_000, rel=1e-6), (
        f"K3 violated: a ¥600M customer advance must raise period-0 ending "
        f"cash by the advance; got {first_delta:,.0f}."
    )
    assert final_delta == pytest.approx(0, abs=1_000), (
        f"K3 violated: a released customer advance must not change terminal "
        f"cash (it is working-capital timing, not funding); got {final_delta:,.0f}."
    )


def test_K3_interest_accrues_on_convertibles_and_lease() -> None:
    """K3 — interest: convertible notes and lease financing join the
    interest-bearing balance exactly as the workbook does (P&L interest =
    Capital Stack debt balance × rate, where the balance rolls debt +
    converts + lease forward)."""
    periods = 3
    zeros = [0] * periods
    args = dict(
        revenue=[1_000_000_000] * periods,
        target_gross_margin=[0.7] * periods,
        total_headcount=[10] * periods,
        avg_comp_yen=[10_000_000] * periods,
        product_headcount=[3] * periods,
        sm_pct_revenue=[0.10] * periods,
        rd_program_per_product_fte_yen=[4_500_000] * periods,
        rd_program_floor_yen=zeros,
        ga_pct_revenue=[0.05] * periods,
        fixed_ga_yen=zeros,
        capex_yen=zeros,
        depreciation_life_months=[60] * periods,
        debt_raise_yen=zeros,
        debt_interest_rate=[0.10] * periods,
        ar_days=zeros,
        ap_days=zeros,
        deferred_revenue_share=[0.0] * periods,
        inventory_wip_pct_capex=[0.0] * periods,
        tax_rate=[0.0] * periods,
    )
    baseline = kernel.project_free_cash_flow(**args)
    financed = kernel.project_free_cash_flow(
        **args,
        convertibles_yen=[500_000_000, 0, 0],
        lease_financing_yen=[300_000_000, 0, 0],
    )
    expected_interest = (500_000_000 + 300_000_000) * 0.10
    assert financed[0].get("interest", 0.0) == pytest.approx(expected_interest), (
        f"K3 violated: period-0 interest must accrue on the convertible + "
        f"lease balance (expected {expected_interest:,.0f}); got "
        f"{financed[0].get('interest')!r}."
    )
    ni_drop = baseline[0]["net_income"] - financed[0]["net_income"]
    assert ni_drop == pytest.approx(expected_interest), (
        f"K3 violated: net income must carry the convertible/lease interest "
        f"expense; expected a drop of {expected_interest:,.0f}, got {ni_drop:,.0f}."
    )


def test_K3_balance_sheet_balances_with_all_instruments() -> None:
    """K3 — three-statement integrity with every instrument non-zero: cash,
    debt-like balances (converts + lease), the advances liability, and the
    grants / secondary equity effects must keep A = L + E every period."""
    facts = replace(
        _k3_facts(),
        grants_yen=[300_000_000, 0, 0, 0, 0],
        convertibles_yen=[400_000_000, 0, 0, 0, 0],
        lease_financing_yen=[0, 250_000_000, 0, 0, 0],
        customer_advances_yen=[150_000_000, 0, 0, 0, 0],
        secondary_yen=[0, 100_000_000, 0, 0, 0],
    )
    for idx, period in enumerate(kernel.project_balance_sheet(facts)):
        scale = max(1.0, abs(period["total_assets"]))
        assert abs(period["balance_check"]) <= 1e-4 * scale, (
            f"K3 violated: balance sheet with financing instruments does not "
            f"balance in period {idx}: A - L - E = {period['balance_check']:,.0f}"
        )


def test_K7_audit_detects_K3_secondary_without_a_round() -> None:
    """K7 meta for K3 — audit_economic_coherence must flag founder/investor
    secondary liquidity scheduled in a period with no equity round: the
    company cannot fund shareholder liquidity out of operating cash, so a
    secondary without a concurrent round is a financing-coherence defect."""
    facts = _k3_facts()
    equity = list(facts.equity_raise_yen)
    if len(equity) > 1:
        equity[1] = 0
    broken = replace(
        facts,
        equity_raise_yen=equity,
        secondary_yen=[0, 500_000_000, 0, 0, 0],
    )
    issues = kernel.audit_economic_coherence(broken)
    flagged = [i for i in issues if "secondary" in i.lower() or "k3" in i.lower()]
    assert flagged, (
        "K7 violated: audit_economic_coherence did not flag secondary "
        "liquidity scheduled without a concurrent equity round. Returned "
        "issues:\n  " + "\n  ".join(issues or ["<none>"])
    )


def test_K3_debt_amortization_gives_declining_balance_interest() -> None:
    """Task 3.2 (D) — contractual amortization: the interest-bearing balance
    is cumulative draws minus cumulative repayments, and interest accrues on
    the declining balance. Draw ¥1B at 10%, repay ¥400M then ¥300M: interest
    must fall ¥100M → ¥60M → ¥30M."""
    periods = 3
    zeros = [0] * periods
    args = dict(
        revenue=[1_000_000_000] * periods,
        target_gross_margin=[0.7] * periods,
        total_headcount=[10] * periods,
        avg_comp_yen=[10_000_000] * periods,
        product_headcount=[3] * periods,
        sm_pct_revenue=[0.10] * periods,
        rd_program_per_product_fte_yen=[4_500_000] * periods,
        rd_program_floor_yen=zeros,
        ga_pct_revenue=[0.05] * periods,
        fixed_ga_yen=zeros,
        capex_yen=zeros,
        depreciation_life_months=[60] * periods,
        debt_raise_yen=[1_000_000_000, 0, 0],
        debt_interest_rate=[0.10] * periods,
        ar_days=zeros,
        ap_days=zeros,
        deferred_revenue_share=[0.0] * periods,
        inventory_wip_pct_capex=[0.0] * periods,
        tax_rate=[0.0] * periods,
    )
    projection = kernel.project_free_cash_flow(
        **args, debt_amortization_yen=[0, 400_000_000, 300_000_000],
    )
    expected = [100_000_000, 60_000_000, 30_000_000]
    got = [period.get("interest") for period in projection]
    for idx, (want, have) in enumerate(zip(expected, got)):
        assert have == pytest.approx(want), (
            f"K3 violated (Task 3.2): period {idx} interest must accrue on the "
            f"declining balance (cum draws − cum repaid); expected {want:,}, "
            f"got {have!r}. Full interest series: {got}."
        )


def test_K3_debt_amortization_flows_through_cash_and_balance_sheet() -> None:
    """Task 3.2 (D) — principal repayment is a financing outflow: ending cash
    drops by the cumulative repayment (zero-rate plan keeps the delta exact)
    and the balance sheet still balances with a declining debt balance."""
    base = _k3_facts()
    amortized = replace(
        base,
        debt_raise_yen=[500_000_000, 0, 0, 0, 0],
        debt_amortization_yen=[0, 200_000_000, 100_000_000, 0, 0],
    )
    baseline = replace(
        base,
        debt_raise_yen=[500_000_000, 0, 0, 0, 0],
    )
    proj_amortized = kernel.project_plan_free_cash_flow(amortized)
    proj_baseline = kernel.project_plan_free_cash_flow(baseline)
    delta = proj_amortized[-1]["ending_cash"] - proj_baseline[-1]["ending_cash"]
    assert delta == pytest.approx(-300_000_000, rel=1e-6), (
        f"K3 violated (Task 3.2): cumulative ¥300M of principal repayment must "
        f"reduce terminal ending cash by the same amount; got {delta:,.0f}. "
        f"Amortization is a financing outflow, not a P&L item."
    )
    for idx, period in enumerate(kernel.project_balance_sheet(amortized)):
        scale = max(1.0, abs(period["total_assets"]))
        assert abs(period["balance_check"]) <= 1e-4 * scale, (
            f"K3 violated (Task 3.2): balance sheet with debt amortization "
            f"does not balance in period {idx}: A - L - E = "
            f"{period['balance_check']:,.0f}"
        )


def test_K3_workbook_capital_stack_carries_amortization_schedule() -> None:
    """Task 3.2 (D) — the workbook mirrors the kernel: Assumptions carries the
    'Debt amortization' schedule as an input row, the P&L debt balance nets it
    out, and the CF financing block subtracts the repayment."""
    from openpyxl.utils import get_column_letter

    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "amortized.xlsx"
        input_path.write_text(
            _k3_base_yaml()
            + "debt_raise_yen: [500000000, 0, 0, 0, 0]\n"
            + "debt_amortization_yen: [0, 200000000, 100000000, 0, 0]\n",
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        first_col = source_plan.START_PERIOD_COL
        second = get_column_letter(first_col + 1)
        assumptions = wb["Assumptions"]
        amort_row = _row_for_label(assumptions, "Debt amortization")
        assert assumptions.cell(amort_row, first_col + 1).value == 200000000, (
            "K3 violated (Task 3.2): the stated amortization schedule must land "
            "on the Assumptions input row"
        )
        balance_row = _row_for_label(wb["P&L"], "Debt balance (ending)")
        balance_formula = str(
            wb["P&L"].cell(balance_row, first_col + 1).value).replace(" ", "")
        assert f"-'Assumptions'!{second}{amort_row}" in balance_formula, (
            f"K3 violated (Task 3.2): debt balance must net out the "
            f"amortization row; got {balance_formula!r}"
        )
        repay_row = _row_for_label(wb["CF"], "Debt principal repayment")
        cf_repay = str(wb["CF"].cell(repay_row, first_col + 1).value).replace(" ", "")
        assert f"-'Assumptions'!{second}{amort_row}" in cf_repay, (
            f"K3 violated (Task 3.2): CF financing must subtract the "
            f"repayment; got {cf_repay!r}"
        )


def test_G_half_year_convention_halves_acquisition_period_depreciation() -> None:
    """Task 3.3 (G) — half-year convention: assets acquired in a period take
    half a period's straight-line charge in that period, a full charge
    afterwards. ¥1.2B of period-0 CapEx on a 60-month life must depreciate
    ¥120M in period 0 (half of 240M) and ¥240M in period 1."""
    periods = 3
    zeros = [0] * periods
    args = dict(
        revenue=[1_000_000_000] * periods,
        target_gross_margin=[0.7] * periods,
        total_headcount=[10] * periods,
        avg_comp_yen=[10_000_000] * periods,
        product_headcount=[3] * periods,
        sm_pct_revenue=[0.10] * periods,
        rd_program_per_product_fte_yen=[4_500_000] * periods,
        rd_program_floor_yen=zeros,
        ga_pct_revenue=[0.05] * periods,
        fixed_ga_yen=zeros,
        capex_yen=[1_200_000_000, 0, 0],
        depreciation_life_months=[60] * periods,
        debt_raise_yen=zeros,
        debt_interest_rate=[0.0] * periods,
        ar_days=zeros,
        ap_days=zeros,
        deferred_revenue_share=[0.0] * periods,
        inventory_wip_pct_capex=[0.0] * periods,
        tax_rate=[0.0] * periods,
    )
    projection = kernel.project_free_cash_flow(**args)
    assert projection[0]["depreciation"] == pytest.approx(120_000_000), (
        f"G violated: acquisition-period D&A must take the half-year "
        f"convention (¥120M on ¥1.2B / 60mo); got "
        f"{projection[0]['depreciation']:,.0f}."
    )
    assert projection[1]["depreciation"] == pytest.approx(240_000_000), (
        f"G violated: the first full period must charge the full straight-line "
        f"amount (¥240M); got {projection[1]['depreciation']:,.0f}."
    )
    legacy = kernel.project_free_cash_flow(
        **args, depreciation_convention="full_period",
    )
    assert legacy[0]["depreciation"] == pytest.approx(240_000_000), (
        f"G violated: depreciation_convention='full_period' must reproduce the "
        f"legacy full-period charge; got {legacy[0]['depreciation']:,.0f}."
    )


def test_G_workbook_pl_da_mirrors_half_year_convention() -> None:
    """Task 3.3 (G) — the workbook P&L D&A formula must mirror the kernel's
    half-year convention: the acquisition period charges only half of the
    current period's CapEx contribution."""
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "halfyear.xlsx"
        input_path.write_text(_k3_base_yaml(), encoding="utf-8")
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        first_col = source_plan.START_PERIOD_COL
        da_row = _row_for_label(wb["P&L"], "D&A")
        formula = str(wb["P&L"].cell(da_row, first_col).value).replace(" ", "")
        # Half-year convention, gated to annual columns via the months ruler:
        # acquisition-period CapEx contributes only half its base.
        assert "IF(F$5=12" in formula and "/2" in formula, (
            f"G violated: P&L D&A must apply the half-year convention to the "
            f"current period's CapEx on annual columns; got {formula!r}"
        )
        assert "'CostBuild'!" in formula, (
            f"G violated: the D&A base must come from the CapEx engine row; "
            f"got {formula!r}"
        )


def test_K7_audit_detects_amortization_exceeding_draws() -> None:
    """K7 meta for Task 3.2 — repaying more principal than was ever drawn
    (debt + convertibles + lease) is a financing-coherence defect the audit
    must flag."""
    broken = replace(
        _k3_facts(),
        debt_raise_yen=[100_000_000, 0, 0, 0, 0],
        debt_amortization_yen=[0, 500_000_000, 0, 0, 0],
    )
    issues = kernel.audit_economic_coherence(broken)
    flagged = [i for i in issues if "amortization" in i.lower() or "repay" in i.lower()]
    assert flagged, (
        "K7 violated: audit_economic_coherence did not flag amortization "
        "exceeding cumulative draws. Returned issues:\n  "
        + "\n  ".join(issues or ["<none>"])
    )


def test_K8_out_of_range_gross_margin_target_is_warned_not_silently_clamped() -> None:
    """Task 3.4 (H) — K8: a stated target_gross_margin outside the modelable
    range is still clamped for computation, but the clamp must be surfaced
    through the warnings channel instead of silently rewriting the user's
    input."""
    facts = _k3_facts()
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "k8.yaml"
        ypath.write_text(
            "periods: 5\n"
            "customers: [200, 400, 700, 1100, 1600]\n"
            "monthly_price_yen: [50000, 50000, 50000, 50000, 50000]\n"
            "target_gross_margin: [0.99, 0.99, 0.99, 0.99, 0.99]\n",
            encoding="utf-8",
        )
        stated = kernel.derive_source_facts_from_mapping(build_model.load_yaml(ypath))
    assert max(stated.target_gross_margin) <= 0.95, (
        "computation must still clamp the target into the modelable range"
    )
    warnings = kernel.audit_economic_warnings(stated)
    flagged = [w for w in warnings if "clamp" in w.lower() or "gross margin" in w.lower()]
    assert flagged, (
        f"K8 violated: an out-of-range target_gross_margin (0.99) must produce "
        f"a clamp warning. Warnings returned: {warnings!r}"
    )
    clean = kernel.audit_economic_warnings(facts)
    assert not any("clamp" in w.lower() for w in clean), (
        f"K8 violated: an in-range target must not produce a clamp warning; "
        f"got {clean!r}"
    )


def test_K6_material_revenue_on_default_drivers_is_warned() -> None:
    """Task 3.4 (H) — K6: when a plan's revenue is material but its demand
    drivers were never pinned by a structured input (profile-default or
    narrative-scale estimates), the warnings channel must name the provenance
    so the reader knows the revenue rests on defaults, not evidence."""
    with tempfile.TemporaryDirectory() as tmp:
        ypath = Path(tmp) / "k6.yaml"
        # No customers / new_units / price: demand and pricing fall back to
        # profile defaults, yet the derived plan carries material revenue.
        ypath.write_text(
            "company: DefaultsCo\nmechanics: recurring software\nperiods: 5\n",
            encoding="utf-8",
        )
        defaults = kernel.derive_source_facts_from_mapping(build_model.load_yaml(ypath))
    warnings = kernel.audit_economic_warnings(defaults)
    flagged = [
        w for w in warnings
        if "provenance" in w.lower() or "default" in w.lower() or "k6" in w.lower()
    ]
    assert flagged, (
        f"K6 violated: material revenue driven by unpinned profile defaults "
        f"must produce a provenance warning. Warnings: {warnings!r}"
    )
    pinned = _k3_facts()
    pinned_warnings = kernel.audit_economic_warnings(pinned)
    assert not any("provenance" in w.lower() or "k6" in w.lower() for w in pinned_warnings), (
        f"K6 violated: a structured plan with pinned demand must not warn on "
        f"revenue provenance; got {pinned_warnings!r}"
    )


def test_H_cost_build_carries_governor_note() -> None:
    """Task 3.4 (H) — governor transparency: the Cost Build sheet must state
    that COGS components are governor-rescaled to the target gross margin, so
    a reader knows the cost mix is calibrated, not sourced line-by-line."""
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "governed.xlsx"
        input_path.write_text(_k3_base_yaml(), encoding="utf-8")
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        texts = [
            str(cell.value)
            for row in wb["Cost Build"].iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        assert any(
            "governor" in text.lower() and "gross margin" in text.lower()
            for text in texts
        ), (
            "H violated: Cost Build must carry a governor-rescale note "
            "explaining that COGS components are calibrated to the target "
            "gross margin."
        )


def test_F_stated_onboarding_months_drive_one_time_revenue() -> None:
    """Task 2.5 (F) — the one-time / onboarding fee is a YAML driver, not a
    hard-coded price×3: stating onboarding_months must move revenue by
    new_units × price × months, and stating a direct per-unit one-time fee
    must override the months path."""
    zero = _k3_facts("onboarding_months: 0\n")
    six = _k3_facts("onboarding_months: 6\n")
    rev_zero = [p["revenue"] for p in kernel.project_plan_free_cash_flow(zero)]
    rev_six = [p["revenue"] for p in kernel.project_plan_free_cash_flow(six)]
    for idx in range(len(rev_zero)):
        expected = six.new_units[idx] * 50_000 * 6 * (
            1.0 + six.other_revenue_share[idx]
        )
        got = rev_six[idx] - rev_zero[idx]
        assert got == pytest.approx(expected, rel=1e-6), (
            f"F violated: onboarding_months must drive the one-time stream; "
            f"period {idx} revenue delta {got:,.0f} != new_units×price×6"
            f"×(1+other share) = {expected:,.0f}."
        )
    flat = _k3_facts("one_time_revenue_per_unit_yen: [90000, 90000, 90000, 90000, 90000]\n")
    rev_flat = [p["revenue"] for p in kernel.project_plan_free_cash_flow(flat)]
    for idx in range(len(rev_zero)):
        expected = flat.new_units[idx] * 90_000 * (
            1.0 + flat.other_revenue_share[idx]
        )
        got = rev_flat[idx] - rev_zero[idx]
        assert got == pytest.approx(expected, rel=1e-6), (
            f"F violated: a stated per-unit one-time fee must override the "
            f"months path; period {idx} delta {got:,.0f} != "
            f"new_units×¥90k×(1+other share) = {expected:,.0f}."
        )


def test_F_default_onboarding_is_labeled_placeholder_in_workbook() -> None:
    """Task 2.5 (F) — when no onboarding driver is stated, the profile
    default (price × 3) may apply, but the Assumptions row must label it as
    a placeholder instead of presenting it as a sourced fact; a stated
    driver drops the placeholder label and lands in the formula."""
    with tempfile.TemporaryDirectory() as tmp:
        default_path = Path(tmp) / "default.yaml"
        stated_path = Path(tmp) / "stated.yaml"
        default_out = Path(tmp) / "default.xlsx"
        stated_out = Path(tmp) / "stated.xlsx"
        default_path.write_text(_k3_base_yaml(), encoding="utf-8")
        stated_path.write_text(
            _k3_base_yaml() + "onboarding_months: 6\n", encoding="utf-8"
        )
        build_model.build_model(default_path, default_out, mode="full")
        build_model.build_model(stated_path, stated_out, mode="full")
        first_col = source_plan.START_PERIOD_COL

        wb_default = load_workbook(default_out, data_only=False)
        ws = wb_default["Assumptions"]
        row = next(
            r for r in range(5, 25)
            if ws.cell(r, source_plan.LAYOUT.label_col).value == "One-time revenue / new unit"
        )
        source_text = str(ws.cell(row, source_plan.LAYOUT.source_col).value or "")
        assert "placeholder" in source_text.lower(), (
            f"F violated: the defaulted onboarding row must carry a "
            f"placeholder evidence label; source cell reads {source_text!r}."
        )

        wb_stated = load_workbook(stated_out, data_only=False)
        ws2 = wb_stated["Assumptions"]
        row2 = next(
            r for r in range(5, 25)
            if ws2.cell(r, source_plan.LAYOUT.label_col).value == "One-time revenue / new unit"
        )
        formula = str(ws2.cell(row2, first_col).value).replace(" ", "")
        assert "*6" in formula, (
            f"F violated: a stated onboarding_months must land in the "
            f"one-time formula; got {formula!r}."
        )
        source_text2 = str(ws2.cell(row2, source_plan.LAYOUT.source_col).value or "")
        assert "placeholder" not in source_text2.lower(), (
            f"F violated: a stated onboarding driver must not be labeled "
            f"placeholder; source cell reads {source_text2!r}."
        )


def test_K3_stated_debt_interest_rate_reaches_equity_sizing() -> None:
    """Phase 3 review follow-up (Codex) — a stated debt_interest_rate was
    still a post-derivation display override, so the equity auto-sizing
    projected interest at the profile default. A plan drawing ¥2B at 25%
    must size more equity than the same plan at 0%."""
    base = (
        "periods: 5\n"
        "customers: [200, 400, 700, 1100, 1600]\n"
        "monthly_price_yen: [50000, 50000, 50000, 50000, 50000]\n"
        "target_gross_margin: [0.7, 0.7, 0.7, 0.7, 0.7]\n"
        "debt_raise_yen: [2000000000, 0, 0, 0, 0]\n"
    )
    with tempfile.TemporaryDirectory() as tmp:
        cheap_path = Path(tmp) / "cheap.yaml"
        dear_path = Path(tmp) / "dear.yaml"
        cheap_path.write_text(base + "debt_interest_rate: 0\n", encoding="utf-8")
        dear_path.write_text(base + "debt_interest_rate: 0.25\n", encoding="utf-8")
        cheap = kernel.derive_source_facts_from_mapping(build_model.load_yaml(cheap_path))
        dear = kernel.derive_source_facts_from_mapping(build_model.load_yaml(dear_path))
    assert dear.debt_interest_rate[0] == pytest.approx(0.25)
    assert sum(dear.equity_raise_yen) > sum(cheap.equity_raise_yen), (
        f"K3 violated: a stated 25% debt interest rate must reach the "
        f"equity sizing (¥500M/yr more burn); 0%: {cheap.equity_raise_yen}, "
        f"25%: {dear.equity_raise_yen}."
    )


def test_F_stated_zero_one_time_fee_is_honored_per_period() -> None:
    """Phase 3 review follow-up (Codex) — statedness, not positivity, decides
    the one-time fee path: a stated [90000, 0, 0, 0, 0] schedule books the fee
    only in period 0 and books zero afterwards (no silent price×3 fallback),
    and the workbook input row mirrors the same schedule."""
    zero = _k3_facts("onboarding_months: 0\n")
    mixed = _k3_facts("one_time_revenue_per_unit_yen: [90000, 0, 0, 0, 0]\n")
    rev_zero = [p["revenue"] for p in kernel.project_plan_free_cash_flow(zero)]
    rev_mixed = [p["revenue"] for p in kernel.project_plan_free_cash_flow(mixed)]
    first_expected = mixed.new_units[0] * 90_000 * (1.0 + mixed.other_revenue_share[0])
    assert rev_mixed[0] - rev_zero[0] == pytest.approx(first_expected, rel=1e-6), (
        f"F violated: period-0 stated fee not booked; delta "
        f"{rev_mixed[0] - rev_zero[0]:,.0f} != {first_expected:,.0f}."
    )
    for idx in range(1, len(rev_zero)):
        assert rev_mixed[idx] - rev_zero[idx] == pytest.approx(0, abs=1_000), (
            f"F violated: a stated 0 fee in period {idx} must book zero "
            f"one-time revenue, not fall back to price×3; delta "
            f"{rev_mixed[idx] - rev_zero[idx]:,.0f}."
        )
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "mixed.yaml"
        out = Path(tmp) / "mixed.xlsx"
        input_path.write_text(
            _k3_base_yaml()
            + "one_time_revenue_per_unit_yen: [90000, 0, 0, 0, 0]\n",
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        ws = wb["Assumptions"]
        row = next(
            r for r in range(5, 25)
            if ws.cell(r, source_plan.LAYOUT.label_col).value == "One-time revenue / new unit"
        )
        first_col = source_plan.START_PERIOD_COL
        values = [ws.cell(row, first_col + i).value for i in range(5)]
        assert values == [90000, 0, 0, 0, 0], (
            f"F violated: workbook one-time row must mirror the stated "
            f"schedule including zeros; got {values!r}."
        )


def test_K3_financing_inflow_subtotal_excludes_advances() -> None:
    """Phase 3 review follow-up (Codex) — the financing-inflow subtotal must
    not count customer advances (they flow through working capital) and must
    net out contractual amortization, mirroring the CF financing block."""
    import re

    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "fin.xlsx"
        input_path.write_text(
            _k3_base_yaml()
            + "customer_advances_yen: [100000000, 0, 0, 0, 0]\n"
            + "debt_raise_yen: [500000000, 0, 0, 0, 0]\n"
            + "debt_amortization_yen: [0, 100000000, 0, 0, 0]\n",
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        first_col = source_plan.START_PERIOD_COL

        def sum_bounds(ws, row) -> tuple[int, int]:
            formula = str(ws.cell(row, first_col).value).replace(" ", "")
            match = re.search(r"=SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)", formula)
            assert match, f"{ws.title} subtotal is not a SUM range: {formula!r}"
            return int(match.group(1)), int(match.group(2))

        # CF financing block: repayment inside the total, advances outside it
        # (advances live in the operating working-capital deltas).
        cf = wb["CF"]
        cf_total = _row_for_label(cf, "Total financing CF")
        lo, hi = sum_bounds(cf, cf_total)
        repay_row = _row_for_label(cf, "Debt principal repayment")
        assert lo <= repay_row <= hi, (
            "K3 violated: CF financing total must net out the principal repayment"
        )
        adv_delta = _row_for_label(cf, "Δ Customer advances (前受金)")
        assert not (lo <= adv_delta <= hi), (
            "K3 violated: customer advances leaked into the CF financing total"
        )
        # Financing sheet sources: repayment netted, no advances source row.
        fin = wb["Financing"]
        fin_total = _row_for_label(fin, "Total sources (net of repayment)")
        lo, hi = sum_bounds(fin, fin_total)
        fin_repay = _row_for_label(fin, "Debt principal repayment")
        assert lo <= fin_repay <= hi, (
            "K3 violated: financing sources must net out the principal repayment"
        )
        source_labels = [
            fin.cell(row, 3).value for row in range(lo, hi + 1)
        ]
        assert not any(
            isinstance(label, str) and "advance" in label.lower()
            for label in source_labels
        ), (
            f"K3 violated: customer advances leaked into the funding sources: "
            f"{source_labels!r}"
        )


# --- S1: period axis, annual-series expansion, grain-correct kernel math ----


def _s1_mapping(**extra) -> dict:
    """Structured base mapping for the S1 axis / grain tests."""
    base = {
        "company": "AxisCo",
        "mechanics": "recurring software subscription",
        "grain": "hybrid",
        "periods": 5,
        "start_year": 2027,
        "fiscal_year_end_month": 3,
        "customers": [100, 300, 700, 1200, 1800],
        "monthly_price_yen": [50000, 50000, 50000, 50000, 50000],
    }
    base.update(extra)
    return base


def test_S1_build_period_axis_hybrid_fye3() -> None:
    """Hybrid axis: 24 monthly columns over FY2027-FY2028, then annual.

    Facts stay ANNUAL-CANONICAL (5 fiscal years, FY labels); the monthly ×
    annual expansion happens only in build_period_axis. FY convention:
    FY2027 with fiscal_year_end_month=3 spans 2026/04-2027/03.
    """
    from datetime import date as _date

    facts = kernel.derive_source_facts_from_mapping(_s1_mapping())
    assert facts.grain == "hybrid"
    assert facts.years == [2027, 2028, 2029, 2030, 2031]
    assert facts.period_labels == [f"FY{y}" for y in facts.years], (
        "hybrid facts must keep annual-canonical FY labels"
    )
    assert facts.fiscal_year_end_month == 3

    axis = kernel.build_period_axis(facts)
    assert axis.grain == "hybrid"
    assert axis.monthly_count == 24
    assert axis.months_in_period == [1] * 24 + [12] * 3
    assert sum(axis.months_in_period) == 5 * 12, (
        "hybrid axis must cover exactly periods × 12 months"
    )
    assert axis.labels[0] == "2026/04"
    assert axis.labels[11] == "2027/03"
    assert axis.labels[12] == "2027/04"
    assert axis.labels[23] == "2028/03"
    assert axis.labels[24:] == ["FY2029", "FY2030", "FY2031"]
    assert axis.fy_labels[0] == "FY2027" and axis.fy_labels[11] == "FY2027"
    assert axis.fy_labels[12] == "FY2028" and axis.fy_labels[23] == "FY2028"
    assert axis.fy_labels[24:] == ["FY2029", "FY2030", "FY2031"]
    assert axis.period_end[0] == _date(2026, 4, 30)
    assert axis.period_end[10] == _date(2027, 2, 28)
    assert axis.period_end[23] == _date(2028, 3, 31)
    assert axis.period_end[24] == _date(2029, 3, 31)
    assert axis.period_end[-1] == _date(2031, 3, 31)

    # A stated monthly window is rounded UP to whole fiscal years (30 → 36).
    wide = kernel.derive_source_facts_from_mapping(
        _s1_mapping(monthly_window_months=30)
    )
    wide_axis = kernel.build_period_axis(wide)
    assert wide.monthly_window_months == 30
    assert wide_axis.monthly_count == 36
    assert wide_axis.months_in_period == [1] * 36 + [12] * 2
    narrow_axis = kernel.build_period_axis(
        kernel.derive_source_facts_from_mapping(_s1_mapping(monthly_window_months=12))
    )
    assert narrow_axis.monthly_count == 12
    assert narrow_axis.months_in_period == [1] * 12 + [12] * 4

    # A "5年" horizon string keeps periods = 5 fiscal years under hybrid.
    jp = kernel.derive_source_facts_from_mapping(_s1_mapping(periods="5年"))
    assert len(jp.years) == 5 and jp.grain == "hybrid"

    # FYE=12 (calendar fiscal year): FY2027 starts 2027/01.
    cal_axis = kernel.build_period_axis(
        kernel.derive_source_facts_from_mapping(_s1_mapping(fiscal_year_end_month=12))
    )
    assert cal_axis.labels[0] == "2027/01"
    assert cal_axis.period_end[23] == _date(2028, 12, 31)


def test_S1_build_period_axis_annual_is_backward_compatible() -> None:
    """Annual axis labels must remain exactly the current FY labels."""
    facts = kernel.derive_source_facts(SAAS_STORY)
    axis = kernel.build_period_axis(facts)
    assert axis.grain == "annual"
    assert axis.labels == facts.period_labels
    assert axis.labels == [f"FY{year}" for year in facts.years]
    assert axis.fy_labels == axis.labels
    assert axis.months_in_period == [12] * len(facts.years)
    assert axis.monthly_count == 0
    # Default JP fiscal year end (March): FY{Y} ends Y/03/31.
    assert facts.fiscal_year_end_month == 3
    assert all(d.month == 3 and d.day == 31 for d in axis.period_end)
    assert [d.year for d in axis.period_end] == facts.years


def test_S1_expand_annual_series_flow_stock_rate_contracts() -> None:
    """flow sums exactly per FY; stock hits FY-ends exactly; rate holds."""
    facts = kernel.derive_source_facts_from_mapping(_s1_mapping())
    axis = kernel.build_period_axis(facts)

    flow = [1200, 2401, 300, 400, 500]  # 2401 is not divisible by 78
    out = kernel.expand_annual_series(flow, axis, "flow")
    assert len(out) == len(axis.labels)
    assert all(isinstance(v, int) for v in out[:24]), (
        "integer flow input must expand to integers"
    )
    assert sum(out[:12]) == 1200, "monthly flow must sum exactly to FY1"
    assert sum(out[12:24]) == 2401, (
        "integer rounding must be distributed so the FY sum is exact"
    )
    assert out[24:] == [300, 400, 500], "annual columns pass through"
    assert out[0] < out[11], "flow expansion is a rising ramp within the FY"

    stock = [120, 240, 480, 800, 1200]
    s = kernel.expand_annual_series(stock, axis, "stock")
    assert s[11] == 120 and s[23] == 240, "month 12 of each FY == FY value"
    assert s[24:] == [480, 800, 1200], "annual columns pass through"
    assert s[5] == 60, "linear interpolation from prior FY-end (0) to FY-end"
    assert s[17] == 180, "FY2 interpolates from 120 to 240"
    assert all(s[i] <= s[i + 1] for i in range(23)), "stock interp is monotone here"

    rate = [0.60, 0.65, 0.70, 0.70, 0.70]
    r = kernel.expand_annual_series(rate, axis, "rate")
    assert r[:12] == [0.60] * 12 and r[12:24] == [0.65] * 12, "rate holds per FY"
    assert r[24:] == [0.70, 0.70, 0.70]

    with pytest.raises(ValueError):
        kernel.expand_annual_series([1, 2], axis, "flow")
    with pytest.raises(ValueError):
        kernel.expand_annual_series(flow, axis, "balance")


def test_S1_months_factor_and_monthly_axis_labels() -> None:
    assert kernel.months_factor("annual") == 12
    assert kernel.months_factor("hybrid") == 12, (
        "hybrid facts are annual-canonical — kernel projections keep factor 12"
    )
    assert kernel.months_factor("quarterly") == 3
    assert kernel.months_factor("monthly") == 1

    # forecast_axis: real year-month labels are derived from the fiscal
    # start when fiscal_year_end_month is given; the legacy bare "M1" form
    # is kept when it is not (existing workbook headers pin that format).
    _, labels = kernel.forecast_axis(2027, 6, "monthly", fiscal_year_end_month=3)
    assert labels == ["2026/04", "2026/05", "2026/06", "2026/07", "2026/08", "2026/09"]
    _, legacy = kernel.forecast_axis(2027, 6, "monthly")
    assert legacy == ["M1", "M2", "M3", "M4", "M5", "M6"]

    # build_period_axis always derives real year-month labels for monthly.
    facts = kernel.derive_source_facts_from_mapping(
        _s1_mapping(grain="monthly", periods=24,
                    customers=[100] * 24, monthly_price_yen=[50000] * 24)
    )
    axis = kernel.build_period_axis(facts)
    assert axis.grain == "monthly"
    assert axis.monthly_count == 24
    assert axis.months_in_period == [1] * 24
    assert axis.labels[0] == "2026/04" and axis.labels[23] == "2028/03"
    assert axis.fy_labels[0] == "FY2027" and axis.fy_labels[12] == "FY2028"


def test_S1_monthly_grain_kernel_math_is_not_12x() -> None:
    """Monthly-grain projections book one month per period, not twelve.

    The kernel used to annualize every period (recurring = base × price ×
    12, payroll = headcount × annual comp, runway buffer = burn × months/12)
    regardless of grain, so a monthly model overstated revenue and burn 12x.
    """
    customers = [1000] * 24
    price = 10_000
    facts = kernel.derive_source_facts_from_mapping(
        {
            "company": "MonthlyCo",
            "mechanics": "recurring software subscription",
            "grain": "monthly",
            "periods": 24,
            "start_year": 2026,
            "customers": customers,
            "monthly_price_yen": [price] * 24,
        }
    )
    assert facts.grain == "monthly" and facts.customers_pinned

    projection = kernel.project_plan_free_cash_flow(facts)
    average = kernel.average_units(facts.customers)
    for idx in (1, 12, 23):
        expected = (
            average[idx] * price * 1  # one month of recurring billing
            + facts.new_units[idx] * price * facts.onboarding_months[idx]
        ) * (1.0 + facts.other_revenue_share[idx])
        assert abs(projection[idx]["revenue"] - expected) <= max(1.0, 0.001 * expected), (
            f"period {idx}: monthly revenue {projection[idx]['revenue']:,.0f} "
            f"!= customers × price × 1 (+ onboarding) = {expected:,.0f}"
        )
        assert projection[idx]["revenue"] < customers[idx] * price * 6, (
            f"period {idx}: revenue looks annualized (12x) again"
        )

    # Cost calibration and the audit-side recomputation share the factor:
    # the implied gross margin still lands on target at monthly grain.
    margins = kernel.implied_gross_margin_series(facts)
    for idx, margin in enumerate(margins):
        assert abs(margin - facts.target_gross_margin[idx]) <= 0.03, (
            f"period {idx}: monthly implied margin {margin:.1%} off target"
        )

    # Payroll: one month books headcount × avg ANNUAL comp / 12.
    payroll_kwargs = dict(
        product_headcount=[0], sm_pct_revenue=[0.0],
        rd_program_per_product_fte_yen=[0], rd_program_floor_yen=[0],
        ga_pct_revenue=[0.0], fixed_ga_yen=[0], capex_yen=[0],
        depreciation_life_months=[60], debt_raise_yen=[0],
        debt_interest_rate=[0.0], ar_days=[0], ap_days=[0],
        deferred_revenue_share=[0.0], inventory_wip_pct_capex=[0.0],
        tax_rate=[0.0],
    )
    monthly_row = kernel.project_free_cash_flow(
        [0.0], [0.6], [12], [12_000_000], months_per_period=1, **payroll_kwargs
    )[0]
    annual_row = kernel.project_free_cash_flow(
        [0.0], [0.6], [12], [12_000_000], **payroll_kwargs
    )[0]
    assert monthly_row["ebitda"] == -12_000_000, (
        f"monthly payroll must be hc × comp / 12: {monthly_row['ebitda']:,.0f}"
    )
    assert annual_row["ebitda"] == -144_000_000, "annual payroll changed"

    # Interest and depreciation follow the factor too.
    fin_kwargs = dict(payroll_kwargs)
    fin_kwargs.update(
        capex_yen=[120_000_000], debt_raise_yen=[120_000_000],
        debt_interest_rate=[0.06], depreciation_life_months=[60],
    )
    fin_row = kernel.project_free_cash_flow(
        [0.0], [0.6], [0], [0], months_per_period=1, **fin_kwargs
    )[0]
    assert fin_row["interest"] == pytest.approx(120_000_000 * 0.06 / 12)
    assert fin_row["depreciation"] == pytest.approx(60_000_000 * 1 / 60)

    # Runway: the buffer is target months × MONTHLY burn on a monthly axis.
    burn = [-1_000_000.0, -1_000_000.0, -1_000_000.0]
    monthly_equity = kernel.size_equity_rounds(
        0, burn, [0, 0, 0], target_runway_months=12, round_unit=1,
        months_per_period=1,
    )
    annual_equity = kernel.size_equity_rounds(
        0, burn, [0, 0, 0], target_runway_months=12, round_unit=1,
    )
    assert monthly_equity[0] == 13_000_000, (
        f"monthly runway buffer must be 12 months of monthly burn: "
        f"{monthly_equity[0]:,}"
    )
    assert annual_equity[0] == 2_000_000, "annual runway sizing changed"


def test_S1_statutory_welfare_rate_scales_base_comp() -> None:
    """A stated 法定福利費率 loads base comp: avg_comp = base × (1 + rate)."""
    base = _s1_mapping(grain="annual", avg_comp_yen=[8_000_000] * 5)
    base.pop("fiscal_year_end_month")

    loaded = kernel.derive_source_facts_from_mapping(
        {**base, "statutory_welfare_rate": 0.15}
    )
    assert loaded.statutory_welfare_rate == 0.15
    assert loaded.avg_comp_yen == [9_200_000] * 5, (
        f"base 8M × 1.15 expected: {loaded.avg_comp_yen}"
    )
    assert any("statutory welfare" in note for note in loaded.derivation_warnings), (
        "the welfare loading must be surfaced as a derivation note"
    )
    # The percent form is accepted too.
    pct = kernel.derive_source_facts_from_mapping(
        {**base, "statutory_welfare_rate": "15%"}
    )
    assert pct.avg_comp_yen == [9_200_000] * 5

    # Default path (rate unstated) stays byte-identical: comp untouched,
    # no welfare note.
    plain = kernel.derive_source_facts_from_mapping(dict(base))
    assert plain.statutory_welfare_rate == 0.0
    assert plain.avg_comp_yen == [8_000_000] * 5
    assert not any("statutory welfare" in note for note in plain.derivation_warnings)


def test_S1_new_source_facts_fields_default_and_plumb() -> None:
    """New S1 fields default safely and flow through the YAML plumbing."""
    facts = kernel.derive_source_facts(SAAS_STORY)
    assert facts.fiscal_year_end_month == 3
    assert facts.monthly_window_months == 0
    assert facts.statutory_welfare_rate == 0.0
    assert facts.consumption_tax_rate == 0.10
    assert facts.ar_site_months == 0.0 and facts.ap_site_months == 0.0

    mapped = kernel.derive_source_facts_from_mapping(
        _s1_mapping(
            fiscal_year_end_month=12,
            monthly_window_months=24,
            statutory_welfare_rate=0.14,
            consumption_tax_rate=0.08,
            ar_site_months=1.5,
            ap_site_months=2.0,
        )
    )
    assert mapped.fiscal_year_end_month == 12
    assert mapped.monthly_window_months == 24
    assert mapped.statutory_welfare_rate == 0.14
    assert mapped.consumption_tax_rate == 0.08
    assert mapped.ar_site_months == 1.5 and mapped.ap_site_months == 2.0


def test_S1_hybrid_facts_are_annual_canonical() -> None:
    """grain=hybrid derives the same annual-canonical facts as grain=annual;
    only the grain tag (and hence the derived period axis) differs."""
    annual = kernel.derive_source_facts_from_mapping(_s1_mapping(grain="annual"))
    hybrid = kernel.derive_source_facts_from_mapping(_s1_mapping(grain="hybrid"))
    assert annual.grain == "annual" and hybrid.grain == "hybrid"
    for field_name in (
        "years", "period_labels", "new_units", "customers", "monthly_price_yen",
        "avg_comp_yen", "equity_raise_yen", "delivery_cost_yen",
        "variable_cogs_pct", "target_gross_margin", "fixed_ga_yen",
    ):
        assert getattr(annual, field_name) == getattr(hybrid, field_name), (
            f"hybrid facts diverged from annual-canonical on {field_name}"
        )
    # And the explicit-factor call path is the identity for annual facts.
    for story in ARCHETYPES.values():
        story_facts = kernel.derive_source_facts(story)
        args = (
            story_facts.new_units, story_facts.gmv_yen,
            story_facts.monthly_price_yen, story_facts.take_rate,
            story_facts.other_revenue_share, story_facts.revenue_mode,
        )
        assert kernel.plan_revenue_series(*args) == kernel.plan_revenue_series(
            *args, months_per_period=12
        )


def test_S5_unstated_monthly_window_start_anchors_to_fy_in_progress() -> None:
    """Unstated start on a monthly-window grain anchors to the fiscal year in
    progress (S5 fix): the calendar-year default must never render a monthly
    window whose first fiscal year already ended before the run date."""
    from datetime import date as _date

    run = _date(2026, 7, 5)
    # Deterministic helper contract (injected run date).
    assert kernel.anchor_start_year(2026, "hybrid", 3, today=run) == 2027
    assert kernel.anchor_start_year(2026, "monthly", 3, today=run) == 2027
    # A stated start year is always respected, even when fully past.
    assert kernel.anchor_start_year(2026, "hybrid", 3, stated=True, today=run) == 2026
    # Annual / quarterly grains keep the calendar-year default unchanged.
    assert kernel.anchor_start_year(2026, "annual", 3, today=run) == 2026
    assert kernel.anchor_start_year(2026, "quarterly", 3, today=run) == 2026
    # FYE=12: FY2026 is still in progress on 2026-07-05 — no shift.
    assert kernel.anchor_start_year(2026, "hybrid", 12, today=run) == 2026
    # Already-current or future years are never shifted.
    assert kernel.anchor_start_year(2027, "hybrid", 3, today=run) == 2027

    # End-to-end (live run date): a default hybrid mapping must produce a
    # first fiscal year that does not end before the run date.
    facts = kernel.derive_source_facts_from_mapping({"company": "AnchorCo", "grain": "hybrid"})
    assert facts.start_year_stated is False
    axis = kernel.build_period_axis(facts)
    first_fy_end = axis.period_end[11]  # 12th monthly column = first FY end
    assert first_fy_end >= _date.today().replace(day=1), (
        "unstated hybrid start rendered a fully-past first fiscal year"
    )
    # A stated start_year still wins end-to-end.
    stated = kernel.derive_source_facts_from_mapping(
        {"company": "AnchorCo", "grain": "hybrid", "start_year": 2026}
    )
    assert stated.start_year_stated is True and stated.years[0] == 2026
    # Facts-level re-anchor helper (burn_runway hybrid promotion path).
    promoted = replace(
        kernel.derive_source_facts_from_mapping({"company": "AnchorCo"}), grain="hybrid"
    )
    anchored = kernel.anchor_facts_first_fiscal_year(promoted)
    a_axis = kernel.build_period_axis(anchored)
    assert a_axis.period_end[11] >= _date.today().replace(day=1)
    assert anchored.period_labels == [f"FY{y}" for y in anchored.years]


if __name__ == "__main__":
    _tests = [
        test_gross_margin_tracks_target_across_archetypes,
        test_stated_gross_margin_is_extracted_from_narrative,
        test_stated_churn_rate_is_extracted_and_annualized,
        test_annual_pricing_is_converted_to_a_monthly_figure,
        test_unit_price_is_extracted_from_natural_phrasing,
        test_monthly_burn_phrasing_does_not_flip_model_grain,
        test_explicit_monthly_model_request_is_still_detected,
        test_hyphenated_year_horizon_is_honored,
        test_start_year_takes_the_start_of_a_range,
        test_demand_ramp_reaches_stated_arr,
        test_customer_count_reaches_stated_target,
        test_saas_revenue_tracks_the_stated_customer_count,
        test_marketplace_gmv_honors_stated_maturity_target,
        test_marketplace_present_value_gmv_stays_period_zero_base,
        test_marketplace_gmv_end_of_plan_phrasing_is_a_maturity_target,
        test_usd_narrative_produces_a_coherent_usd_plan,
        test_jpy_narratives_are_not_misdetected_as_usd,
        test_yaml_can_override_the_fx_rate,
        test_structured_currency_scales_default_costs_consistently,
        test_payroll_is_not_absurd_at_maturity,
        test_stated_rd_team_size_anchors_the_headcount_plan,
        test_economic_audit_passes_for_healthy_archetypes,
        test_economic_audit_catches_a_broken_cost_stack,
        test_economic_audit_catches_insolvency,
        test_economic_audit_catches_non_positive_revenue,
        test_operations_headcount_does_not_explode_at_high_account_count,
        test_balance_sheet_balances_in_the_kernel_projection,
        test_depreciation_uses_the_accumulated_asset_base,
        test_economic_audit_flags_an_unbalanced_sheet,
        test_workbook_balance_sheet_balances_when_recalculated,
        test_strict_audit_fails_on_economic_incoherence,
        test_maturity_cue_after_the_figure_is_detected,
        test_cap_table_exit_waterfall_uses_post_round_cap_table,
        test_dcf_forecast_fcf_is_not_floored_at_zero,
        test_dcf_enterprise_value_is_positive_and_method_consistent,
        test_funding_plan_keeps_the_company_solvent,
        test_seed_round_is_positive_across_archetypes,
        test_stated_raise_size_and_post_money_are_honored,
        test_workbook_recalc_shows_no_insolvency,
        test_summarize_comps_aggregates_live_peer_margins,
        test_kpi_sheet_compares_plan_kpis_to_live_peers,
        test_peer_comparison_status_recalculates_without_error,
        test_archetype_workbooks_pass_strict_audit,
        test_workbook_recalc_reconciles_with_kernel_gross_margin,
        test_burn_multiple_does_not_explode_in_period_zero,
        test_runway_months_is_capped_at_99,
        test_kpi_dashboard_carries_vc_decision_metrics,
        test_rule_of_40_is_growth_plus_ebitda_margin,
        test_kpi_dashboard_omits_vc_block_for_an_ambiguous_mechanic,
        test_pre_revenue_plan_has_no_product_revenue,
        test_pre_revenue_audit_tolerates_zero_revenue,
        test_pre_revenue_plan_books_no_cogs,
        test_customer_target_is_the_maturity_figure_not_the_current_count,
        test_hardware_unit_ramp_honors_the_stated_maturity_target,
        test_unit_target_extraction_handles_mixed_scale_phrasing,
        test_hardware_revenue_is_units_times_sale_price,
        test_monthly_priced_hardware_stays_recurring,
        test_hardware_unit_sale_survives_an_attach_subscription,
        test_marketplace_kpi_uses_a_transaction_lens,
        test_structured_yaml_revenue_tracks_the_stated_customer_count,
        test_K1_explicit_customers_dominate_recurring_revenue_over_new_units,
        test_K2_changing_churn_rate_moves_ending_units_and_terminal_arr,
        test_K2_structured_yaml_churn_rate_moves_terminal_customers_and_revenue,
        test_K4_nol_absorbs_prior_losses_so_cash_tax_does_not_double_dip,
        test_K5_hardware_profile_suppresses_transaction_revenue_and_gmv,
        test_K7_audit_economic_coherence_detects_K1_customers_drift,
        test_K3_stated_grants_reduce_auto_sized_equity,
        test_K3_projection_ending_cash_reflects_each_instrument,
        test_K3_interest_accrues_on_convertibles_and_lease,
        test_K3_balance_sheet_balances_with_all_instruments,
        test_K7_audit_detects_K3_secondary_without_a_round,
        test_K3_debt_amortization_gives_declining_balance_interest,
        test_K3_debt_amortization_flows_through_cash_and_balance_sheet,
        test_K3_workbook_capital_stack_carries_amortization_schedule,
        test_K7_audit_detects_amortization_exceeding_draws,
        test_G_half_year_convention_halves_acquisition_period_depreciation,
        test_G_workbook_pl_da_mirrors_half_year_convention,
        test_K8_out_of_range_gross_margin_target_is_warned_not_silently_clamped,
        test_K6_material_revenue_on_default_drivers_is_warned,
        test_H_cost_build_carries_governor_note,
        test_F_stated_onboarding_months_drive_one_time_revenue,
        test_F_default_onboarding_is_labeled_placeholder_in_workbook,
        test_K3_stated_debt_interest_rate_reaches_equity_sizing,
        test_F_stated_zero_one_time_fee_is_honored_per_period,
        test_K3_financing_inflow_subtotal_excludes_advances,
        test_S1_build_period_axis_hybrid_fye3,
        test_S1_build_period_axis_annual_is_backward_compatible,
        test_S1_expand_annual_series_flow_stock_rate_contracts,
        test_S1_months_factor_and_monthly_axis_labels,
        test_S1_monthly_grain_kernel_math_is_not_12x,
        test_S1_statutory_welfare_rate_scales_base_comp,
        test_S1_new_source_facts_fields_default_and_plumb,
        test_S1_hybrid_facts_are_annual_canonical,
        test_S5_unstated_monthly_window_start_anchors_to_fy_in_progress,
    ]
    for _test in _tests:
        try:
            _test()
            print(f"ok    {_test.__name__}")
        except pytest.skip.Exception as exc:  # soffice unavailable, etc.
            print(f"skip  {_test.__name__}: {exc}")
    print("done")
