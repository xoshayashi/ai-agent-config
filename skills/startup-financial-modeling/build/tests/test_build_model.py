"""Smoke tests for the startup-financial-modeling xlsx orchestrator.

Run directly:
    PYTHONPYCACHEPREFIX=$(mktemp -d) python3 skills/startup-financial-modeling/build/tests/test_build_model.py
"""

from __future__ import annotations

import importlib.util
import sys
import tempfile
import re
import json
import shutil
import subprocess
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

SKILL_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = SKILL_DIR.parents[1]
RUNTIME_DIR = SKILL_DIR / "build" / "runtime"
SCRIPTS_DIR = RUNTIME_DIR
sys.path.insert(0, str(RUNTIME_DIR))

import build_model  # noqa: E402
import cap_table_builder as cap_table  # noqa: E402
import economic_kernel as kernel  # noqa: E402
import ib_format as ib  # noqa: E402
import live_comps  # noqa: E402
import self_improvement  # noqa: E402
import closeout_consistency  # noqa: E402
import source_plan_builder as source_plan  # noqa: E402

FIRST_VALUE_COL = source_plan.START_PERIOD_COL
SECOND_VALUE_COL = FIRST_VALUE_COL + 1
FIRST_VALUE_LETTER = source_plan.get_column_letter(FIRST_VALUE_COL)
SECOND_VALUE_LETTER = source_plan.get_column_letter(SECOND_VALUE_COL)
UNIT_COL = source_plan.LAYOUT.unit_col


def _fake_public_comps(tickers: list[str], *, timeout: float = 8.0) -> live_comps.PublicCompsResult:
    comps = [
        live_comps.PublicComp(
            ticker=ticker.strip().upper(),
            name=f"{ticker.strip().upper()} Test Comparable",
            currency="USD",
            market_cap=100_000_000_000.0,
            enterprise_value=110_000_000_000.0,
            revenue_multiple=8.0 + idx,
            ebitda_multiple=20.0 + idx,
            source_url=f"https://example.test/{ticker.strip().upper()}",
            as_of_date="2026-05-17",
            status="current",
        )
        for idx, ticker in enumerate(tickers)
        if ticker.strip()
    ]
    return live_comps.PublicCompsResult(
        comps=comps,
        revenue_multiple_median=9.0 if comps else None,
        ebitda_multiple_median=21.0 if comps else None,
        source_url="https://example.test/live-comps",
        as_of_date="2026-05-17",
    )


build_model.lc.fetch_public_comps = _fake_public_comps


def _column_width(ws, column_letter: str) -> float:
    dimension = ws.column_dimensions.get(column_letter)
    return dimension.width if dimension and dimension.width is not None else 0


def test_skill_exposes_clean_build_route_only() -> None:
    scanned_paths = [
        SKILL_DIR / "SKILL.md",
        SKILL_DIR / "build" / "evals" / "evals.json",
        SKILL_DIR / "scripts" / "build_model.py",
    ]
    scanned_paths.extend((SKILL_DIR / "build" / "references").glob("*.md"))
    text = "\n".join(path.read_text(encoding="utf-8") for path in scanned_paths if path.exists())
    forbidden = [
        "workbook_" + "quality_" + "review",
        "quality " + "review",
        "review " + "score",
        "hard_" + "failures",
    ]
    required_references = [
        "_generic_composition_protocol.md",
        "_ib_workbook_design_system.md",
        "_self_review_protocol.md",
        "_self_improvement_protocol.md",
        "_benchmark_protocol.md",
        "_kpi_analytics.md",
        "_scenario_sensitivity_playbook.md",
        "_valuation_and_return_logic.md",
        "_ic_memo_depth.md",
    ]

    assert not (SCRIPTS_DIR / ("workbook_" + "quality_" + "review.py")).exists()
    assert [(SKILL_DIR / "build" / "references" / name).exists() for name in required_references] == [True] * len(required_references)
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    assert [name for name in required_references if name not in invocation_text] == []
    assert [term for term in forbidden if term in text] == []


def test_skill_uses_generic_economic_kernel_not_stage_matrix() -> None:
    scanned_paths = [
        SKILL_DIR / "SKILL.md",
        SKILL_DIR / "build" / "evals" / "evals.json",
        SKILL_DIR / "scripts" / "build_model.py",
    ]
    scanned_paths.extend((SKILL_DIR / "build" / "references").glob("*.md"))
    text = "\n".join(path.read_text(encoding="utf-8") for path in scanned_paths if path.exists())
    normalized = text.lower()
    forbidden = [
        "ステージ × 業態 × 資金調達目的",
        "stage × business",
        "stage x business",
        "stage applicability",
        "applicability matrix",
        "source-backed investor plan",
        "mode-based monthly template",
        "monthly template",
    ]
    required_files = [
        SKILL_DIR / "build" / "references" / "_modeling_kernel.md",
        SKILL_DIR / "build" / "references" / "_coverage_universe.md",
    ]
    required_terms = [
        "economic kernel",
        "model grain",
        "driver tree",
        "operating engine",
        "capital stack",
        "ownership",
        "scenario",
        "valuation",
        "seed to pre-ipo",
    ]

    assert [path for path in required_files if not path.exists()] == []
    assert [term for term in forbidden if term in normalized] == []
    assert [term for term in required_terms if term not in normalized] == []


def test_prompt_guidance_resists_fixed_template_routing() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    generic_text = (SKILL_DIR / "build" / "references" / "_generic_composition_protocol.md").read_text(encoding="utf-8")
    output_modes_text = (SKILL_DIR / "build" / "references" / "_output_modes.md").read_text(encoding="utf-8")
    kpi_text = (SKILL_DIR / "build" / "references" / "_kpi_analytics.md").read_text(encoding="utf-8")
    skill_flat = " ".join("\n".join([skill_text, invocation_text]).lower().split())
    generic_flat = " ".join(generic_text.split())

    assert "examples, maturity cues, sectors, and modes are prompts for reasoning, not templates" in skill_flat
    assert "Choose tabs, KPIs, scenarios, valuation methods, colors, and cell positions from the decision" in generic_flat
    assert "These are not mandatory bundles" in output_modes_text
    assert "Maturity and mechanics are signals for selecting metrics, not metric packs" in kpi_text


def test_sheet_quality_rubric_covers_every_generated_sheet() -> None:
    rubric_text = (SKILL_DIR / "build" / "references" / "_sheet_quality_rubric.md").read_text(encoding="utf-8")
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    output_modes_text = (SKILL_DIR / "build" / "references" / "_output_modes.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    combined_flat = " ".join(
        "\n".join([rubric_text, skill_text, invocation_text, output_modes_text, self_review_text, eval_text]).split()
    )
    for sheet_name in [*source_plan.SOURCE_PLAN_SHEETS_V2, *source_plan.CONDITIONAL_SHEETS_V2]:
        assert f"| {sheet_name} |" in rubric_text, f"rubric table row missing for {sheet_name}"
    for phrase in [
        "a distinct purpose, source boundary, dependency flow, checks where errors would matter, and interpretation",
        "Do not create a sheet just because it belongs to a canonical full-workbook order",
        "sheet-specific quality gates for purpose, source boundary, dependency flow, checks, and interpretation",
        "Load `_sheet_quality_rubric.md` for every xlsx build or repair",
        "Include a sheet only if it owns a distinct decision surface",
        "Guide defines the decision, evidence, mechanics, and scope",
    ]:
        assert phrase in combined_flat


def test_economic_kernel_is_separate_from_workbook_renderer() -> None:
    builder_text = (SCRIPTS_DIR / "source_plan_builder.py").read_text(encoding="utf-8")
    required_profile_keys = {
        "marketplace",
        "hardware_asset_heavy",
        "pre_revenue_milestone",
        "recurring_software",
        "fintech_balance_sheet",
        "generic",
    }

    assert {profile.key for profile in kernel.MECHANIC_PROFILES} == required_profile_keys
    assert kernel.SourceFacts.__module__ == "economic_kernel"
    assert "class SourceFacts" not in builder_text
    assert "def _detect_mechanics" not in builder_text
    assert "def _money_yen" not in builder_text
    # v2 builders consume the kernel through its published axis / projection
    # API — economics never re-derive inside the renderer.
    assert "def build_period_axis" not in builder_text
    assert "def expand_annual_series" not in builder_text
    assert "build_period_axis(" in builder_text
    assert "expand_annual_series(" in builder_text
    assert "from economic_kernel import" in builder_text


def test_japanese_money_units_parse_at_correct_scale() -> None:
    assert kernel.money_yen([r"([0-9,.]+)\s*(万|億|兆)"], "900億円", 0) == 90_000_000_000
    assert kernel.money_yen([r"([0-9,.]+)\s*(万|億|兆)"], "1.2兆円", 0) == 1_200_000_000_000
    assert kernel.money_yen([r"([0-9,.]+)\s*(万|億|兆)"], "3200万円", 0) == 32_000_000
    assert kernel.money_yen([r"([0-9,.]+)\s*(m|bn|b)"], "12bn", 0) == 12_000_000_000


def test_seed_to_pre_ipo_horizon_is_not_truncated_to_two_years() -> None:
    assert kernel.extract_forecast_periods("5年のpre-IPO financial modelを月次で作る", "monthly") == 60
    assert kernel.extract_forecast_periods("36か月のシード資金計画", "monthly") == 36


def _sheet_labels(wb, sheet_name: str) -> str:
    """Section labels (col B) + line-item labels (col C) of one sheet."""
    return "\n".join(
        str(cell.value)
        for row in wb[sheet_name].iter_rows(min_col=2, max_col=3)
        for cell in row
        if cell.value is not None
    )


def test_generated_workbook_contains_interpretive_analysis_surfaces() -> None:
    """v2: interpretation lives on Summary (cross-checks with judgment,
    recommendation), on the engine sheets (support-status rows), and on the
    conditional Valuation & Exit surface (method credibility, guarded MOIC)."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        summary = _sheet_labels(wb, "Summary")
        for label in (
            "Cross-check",
            "Revenue / FTE benchmark",
            "Revenue / FTE vs benchmark",
            "Market growth (CAGR)",
            "Company growth vs market growth",
            "Funding readout",
            "Growth readout",
            "Weakest evidence (top 3)",
            "Next diligence",
        ):
            assert label in summary, f"Summary missing interpretation surface {label!r}"
        # The cross-check rows are real formulas with a judgment output, not
        # label-only theater (S5 fix).
        vs_bench = _first_year_cell_for_label(wb, "Summary", "Revenue / FTE vs benchmark").value
        assert isinstance(vs_bench, str) and vs_bench.startswith("=IF(") and "benchmark" in vs_bench
        vs_market = _first_year_cell_for_label(wb, "Summary", "Company growth vs market growth").value
        assert isinstance(vs_market, str) and vs_market.startswith("=IF(") and "market" in vs_market
        cagr = _first_year_cell_for_label(wb, "Summary", _row_label_startswith(wb, "Summary", "Revenue CAGR")).value
        assert isinstance(cagr, str) and "^(1/(COUNT(" in cagr, "CAGR must be a live formula"
        # Benchmark comparators are blue inputs with a sourced note.
        assert _font_rgb(_first_year_cell_for_label(wb, "Summary", "Revenue / FTE benchmark")) == ib.IB_HARD_INPUT
        assert _font_rgb(_first_year_cell_for_label(wb, "Summary", "Market growth (CAGR)")) == ib.IB_HARD_INPUT

        # Engine sheets carry their own support-status interpretation rows.
        assert "Demand support status" in _sheet_labels(wb, "Revenue Build")
        assert "Price support status" in _sheet_labels(wb, "Revenue Build")
        assert "Cost-to-serve status" in _sheet_labels(wb, "Cost Build")
        assert "CS capacity status" in _sheet_labels(wb, "People Plan")
        assert "Runway after raise" in _sheet_labels(wb, "Financing")

        # Conditional valuation surface (ma_exit) keeps credibility columns,
        # DCF cross-check, and a guarded investor-return readout.
        out2 = Path(tmp) / "ma.xlsx"
        build_model.build_model(None, out2, mode="ma_exit")
        wb2 = load_workbook(out2, data_only=False)
        vx = _sheet_labels(wb2, "Valuation & Exit")
        for label in ("Primary-method EV", "PV of forecast FCF", "Illustrative IRR",
                      "Selected EV range", "MOIC at selected EV (Mid)"):
            assert label in vx
        moic = wb2["Valuation & Exit"].cell(
            _row_for_label(wb2, "Valuation & Exit", "MOIC at selected EV (Mid)"), 4).value
        invested_row = _row_for_label(wb2, "Valuation & Exit", "Equity invested (cumulative)")
        assert isinstance(moic, str) and moic.startswith(f'=IF($D${invested_row}=0,"-"'), (
            "MOIC must guard invested=0 with '-' — no MAX(1,x) division hacks"
        )


def _row_label_startswith(wb, sheet_name: str, prefix: str) -> str:
    for row in wb[sheet_name].iter_rows(min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(prefix):
                return cell.value
    raise AssertionError(f"{sheet_name}: no label starting with {prefix!r}")


def test_mechanic_specific_kpi_and_scenario_axes_are_rendered() -> None:
    """v2: mechanic-specific KPIs live on the Summary KPI block; mechanic-
    specific driver rows live on Assumptions (no GMV rows for recurring,
    no ARR rows for marketplace)."""
    # Marketplace: transaction lens KPIs + GMV/take-rate drivers.
    tmp, wb = _sample_source_workbook(
        "# Marketplace\nMarketplace with GMV ¥10B, take rate 12%, buyer and seller liquidity. Source: management memo."
    )
    try:
        summary = _sheet_labels(wb, "Summary")
        assumptions = _sheet_labels(wb, "Assumptions")
        assert "Take rate (FY end)" in summary
        assert "Contribution margin" in summary
        assert "ARR run-rate (FY end)" not in summary
        assert "Gross merchandise value" in assumptions
        assert "Take rate" in assumptions
    finally:
        tmp.cleanup()
    # Recurring software: ARR lens, no marketplace rows.
    tmp, wb = _sample_source_workbook(
        "# SaaS\nA recurring software subscription startup with ARR and monthly pricing. Source: management memo."
    )
    try:
        summary = _sheet_labels(wb, "Summary")
        assumptions = _sheet_labels(wb, "Assumptions")
        assert "ARR run-rate (FY end)" in summary
        assert "Take rate (FY end)" not in summary
        assert "Gross merchandise value" not in assumptions
    finally:
        tmp.cleanup()


def test_ambiguous_mechanics_use_generic_kpis_and_scenario_axes() -> None:
    """Ambiguous mechanics: the VC decision-metrics block is mechanism-gated
    (no Rule of 40 / CAC theater on a generic-mechanics plan), while the
    scenario mechanism stays generic (demand/price/cost/opex scales)."""
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA new startup with unclear revenue mechanics and weak evidence. Source: management memo."
    )
    try:
        summary = _sheet_labels(wb, "Summary")
        assumptions = _sheet_labels(wb, "Assumptions")
        assert "Rule of 40" not in summary
        assert "Customer acquisition cost" not in summary
        assert "Burn multiple" not in summary
        for label in ("Scenario toggle", "Demand scale", "Price scale",
                      "Variable cost scale", "Opex scale"):
            assert label in assumptions
    finally:
        tmp.cleanup()
    tmp, wb = _sample_source_workbook(
        "# SaaS\nA recurring software subscription startup with ARR and monthly pricing. Source: management memo."
    )
    try:
        summary = _sheet_labels(wb, "Summary")
        assert "Rule of 40" in summary
        assert "Burn multiple" in summary
    finally:
        tmp.cleanup()


def test_cost_labeled_scenario_drivers_pressure_costs_not_revenue() -> None:
    """v2 scenario mechanism: the cost scale multiplies COGS rows, the opex
    scale multiplies OpEx programs, and neither touches the revenue chain;
    revenue rows multiply only the demand / price scales."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        eff = {
            key: f"$F${_row_for_label(wb, 'Assumptions', label)}"
            for key, label in (
                ("demand", "Effective demand scale"),
                ("price", "Effective price scale"),
                ("cost", "Effective cost scale"),
                ("opex", "Effective opex scale"),
            )
        }
        new_units = _first_year_cell_for_label(wb, "Revenue Build", "New primary units").value
        price = _first_year_cell_for_label(wb, "Revenue Build", _row_label_startswith(wb, "Revenue Build", "Monthly price")).value
        total_revenue = _first_year_cell_for_label(wb, "Revenue Build", "Total revenue").value
        variable_cogs = _first_year_cell_for_label(wb, "Cost Build", "Variable COGS").value
        sm_programs = _first_year_cell_for_label(wb, "Cost Build", "S&M programs").value

        assert eff["demand"] in new_units
        assert eff["price"] in price
        assert eff["cost"] in variable_cogs
        assert eff["opex"] in sm_programs
        for revenue_formula in (new_units, price, total_revenue):
            assert eff["cost"] not in revenue_formula
            assert eff["opex"] not in revenue_formula


def test_unclassified_scenario_drivers_default_to_opex_not_revenue() -> None:
    bucket = source_plan._scenario_driver_bucket
    assert bucket(SimpleNamespace(label="Regulatory clearance readiness")) == "opex"
    assert bucket(SimpleNamespace(label="Prototype / program cost factor")) == "cost"
    assert bucket(SimpleNamespace(label="New logo / conversion scale")) == "revenue"
    assert bucket(SimpleNamespace(label="Warehouse / debt headroom")) == "financing"


def test_financing_driver_downside_widens_funding_gap() -> None:
    """v2: Financing carries BOTH a live funding-gap formula (current toggle)
    and a downside-case snapshot; the downside snapshot is at least the base
    case's need, and the Summary scenario block shows the same ordering."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        live = _first_year_cell_for_label(wb, "Financing", _row_label_startswith(wb, "Financing", "Funding gap (live")).value
        assert isinstance(live, str) and live.startswith("=MAX(0,-MIN('CF'!")
        downside = _first_year_cell_for_label(wb, "Financing", "Downside funding gap").value
        assert isinstance(downside, (int, float)) and downside >= 0

        gap_row = _row_for_label(wb, "Summary", "Additional funding needed")
        down_need = wb["Summary"].cell(gap_row, 6).value
        base_need = wb["Summary"].cell(gap_row, 7).value
        assert isinstance(down_need, (int, float)) and isinstance(base_need, (int, float))
        assert down_need >= base_need, "downside case must not need less funding than base"
        assert down_need == downside, "Financing downside snapshot must match the Summary scenario block"


def test_cost_build_does_not_double_count_detailed_service_costs_in_variable_cogs() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        f = FIRST_VALUE_LETTER
        rev_row = _row_for_label(wb, "Revenue Build", "Total revenue")
        pct_row = _row_for_label(wb, "Assumptions", "Variable COGS / revenue")
        eff_cost_row = _row_for_label(wb, "Assumptions", "Effective cost scale")
        formula = _first_year_cell_for_label(wb, "Cost Build", "Variable COGS").value
        assert formula == (
            f"='Revenue Build'!{f}{rev_row}*'Assumptions'!{f}{pct_row}"
            f"*'Assumptions'!$F${eff_cost_row}"
        )
        assert "SUM('Assumptions'!" not in formula
        # Delivery / cloud / support live in their own cost-to-serve rows.
        for label in ("Delivery cost", "Cloud / platform cost", "Support cost"):
            assert label in _sheet_labels(wb, "Cost Build")


def test_orchestrator_routes_through_generic_source_plan_builder() -> None:
    orchestrator_text = (SCRIPTS_DIR / "build_model.py").read_text(encoding="utf-8")

    assert "three_statement_builder" not in orchestrator_text
    assert "SaaS Series A defaults" not in orchestrator_text
    assert "00_Cover" not in orchestrator_text
    assert build_model.resolve_bundle("full") == source_plan.SOURCE_PLAN_SHEETS_V2


def test_focused_modes_use_generic_kernel_after_bundle_filter() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        for mode in ("unit_economics", "three_statement"):
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)
            all_text = "\n".join(
                str(cell.value)
                for ws in wb.worksheets
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None
            )

            assert wb.sheetnames == build_model.resolve_bundle(mode)
            assert "Guide" in wb.sheetnames
            assert "Assumptions" in wb.sheetnames
            assert "12_SanityChecks" not in wb.sheetnames
            assert "Demo SaaS" not in all_text
            assert "SaaS Series A" not in all_text


def test_focused_modes_are_formula_complete_without_compact_placeholders() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        for mode in build_model.VALID_MODES:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)
            assert wb.sheetnames == build_model.resolve_bundle(mode)
            workbook_text = "\n".join(
                str(cell.value)
                for ws in wb.worksheets
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None
            )
            assert "compact mode placeholder" not in workbook_text
            assert "Original formula omitted from focused bundle" not in workbook_text
            assert build_model.audit_workbook(wb) == []


def test_integrated_model_has_ib_decision_gates_not_just_readouts() -> None:
    """v2 decision gates: the IC Memo recommendation is a formula gated on
    the return hurdle; the selected EV range's Mid anchors on the
    credibility-chosen Primary-method EV (not a blind average across methods),
    while Low/High show the method spread; the exit waterfall walks the
    preference stack; the Pricing sheet keeps its validation plan and
    pass/DD-gate status."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "ma_exit.xlsx"
        build_model.build_model(None, out, mode="ma_exit")
        wb = load_workbook(out, data_only=False)

        vx = wb["Valuation & Exit"]
        sel_row = _row_for_label(wb, "Valuation & Exit", "Selected EV range")
        primary_row = _row_for_label(wb, "Valuation & Exit", "Primary-method EV")
        # Mid must reference the credibility-aware Primary-method EV cell
        # (rubric bars averaging methods blindly), not MEDIAN across all methods.
        assert str(vx.cell(sel_row, 5).value) == f"=$D${primary_row}"
        # Low/High still show the full method spread (floor / ceiling).
        assert str(vx.cell(sel_row, 4).value).startswith("=MIN(")
        assert str(vx.cell(sel_row, 6).value).startswith("=MAX(")
        method_row = _row_for_label(wb, "Valuation & Exit", "Method")
        assert [vx.cell(method_row, c).value for c in range(4, 9)] == [
            "Low", "Mid", "High", "Credibility", "Use when",
        ]
        waterfall_header_row = _row_for_label(wb, "Valuation & Exit", "Case")
        headers = [vx.cell(waterfall_header_row, c).value for c in range(4, 11)]
        for header in ("Exit EV", "Net debt", "Txn costs", "Preference floor",
                       "Common pool", "Investor proceeds"):
            assert header in headers
        recommendation_formula = wb["IC Memo"].cell(
            _row_for_label(wb, "IC Memo", "Recommendation") + 1, 4).value
        assert isinstance(recommendation_formula, str)
        assert recommendation_formula.startswith("=IF(")
        assert "Proceed subject to DD gates" in recommendation_formula
        assert build_model.audit_workbook(wb) == []

        out2 = Path(tmp) / "pricing.xlsx"
        build_model.build_model(None, out2, mode="pricing")
        wb2 = load_workbook(out2, data_only=False)
        assert _row_for_label(wb2, "Pricing", "Pricing validation plan")
        gate = _first_year_cell_for_label(wb2, "Pricing", "Pricing gate").value
        assert isinstance(gate, str) and '"pass"' in gate and '"DD gate"' in gate


def test_comparable_evidence_loads_public_peers_by_default_and_overrides_multiples() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "live_comps.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)
        evidence_text = "\n".join(str(cell.value) for row in wb["Evidence"].iter_rows() for cell in row if cell.value is not None)

        assert "Comparable evidence" in evidence_text
        assert "public market data" in evidence_text
        assert "CRM Test Comparable" in evidence_text
        assert "current" in evidence_text
        # The register carries the peer-set multiple summary (median 9.0 /
        # 21.0 from the fetched comps) and per-comp as-of dates.
        rev_mult_row = _row_for_label(wb, "Evidence", _row_label_startswith(wb, "Evidence", "EV / Revenue multiple"))
        assert wb["Evidence"].cell(rev_mult_row, 6).value == 9.0
        ebitda_mult_row = _row_for_label(wb, "Evidence", _row_label_startswith(wb, "Evidence", "EV / EBITDA multiple"))
        assert wb["Evidence"].cell(ebitda_mult_row, 6).value == 21.0
        assert "2026-05-17" in evidence_text
        # And the applied medians drive the valuation multiples when the
        # valuation surface is generated.
        out2 = Path(tmp) / "ma.xlsx"
        build_model.build_model(None, out2, mode="ma_exit")
        wb2 = load_workbook(out2, data_only=False)
        mult_row = _row_for_label(wb2, "Valuation & Exit", "Revenue multiple")
        assert wb2["Valuation & Exit"].cell(mult_row, 5).value == 9.0  # Mid column


def test_default_live_comps_match_mechanic_labels_and_cli_overrides_yaml() -> None:
    recurring = kernel.derive_source_facts("# PLAN\nA recurring software startup with ARR.")
    marketplace = kernel.derive_source_facts("# PLAN\nMarketplace with GMV and take rate.")
    hardware = kernel.derive_source_facts("# PLAN\nHardware robot deployment with capex.")
    fintech = kernel.derive_source_facts("# PLAN\nFintech lending balance sheet with loans.")
    deeptech = kernel.derive_source_facts("# PLAN\nPre-revenue R&D milestone platform.")

    assert build_model._default_live_comps_for_facts(recurring) == ["CRM", "NOW", "DDOG"]
    assert build_model._default_live_comps_for_facts(marketplace) == ["UBER", "DASH", "ETSY"]
    assert build_model._default_live_comps_for_facts(hardware) == ["ISRG", "ROK", "TSLA"]
    assert build_model._default_live_comps_for_facts(fintech) == ["PYPL", "SOFI", "AFRM"]
    assert build_model._default_live_comps_for_facts(deeptech) == ["ISRG", "ROK", "TSLA"]

    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "comps.yaml"
        out = Path(tmp) / "override.xlsx"
        input_path.write_text("live_comps: [BAD1, BAD2]\n", encoding="utf-8")
        build_model.build_model(input_path, out, mode="full", live_comps=["CRM"])
        wb = load_workbook(out, data_only=False)
        text = "\n".join(str(cell.value) for row in wb["Evidence"].iter_rows() for cell in row if cell.value is not None)
        assert "CRM Test Comparable" in text
        assert "BAD1" not in text


def test_no_live_comps_disables_default_public_fetch() -> None:
    original_fetch = build_model.lc.fetch_public_comps

    def fail_fetch(tickers: list[str], *, timeout: float = 8.0) -> live_comps.PublicCompsResult:
        raise AssertionError(f"live comparable fetch should be disabled, got {tickers}")

    try:
        build_model.lc.fetch_public_comps = fail_fetch
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "offline.xlsx"
            build_model.build_model(None, out, mode="full", auto_live_comps=False)
            wb = load_workbook(out, data_only=False)
            evidence_text = "\n".join(
                str(cell.value)
                for row in wb["Evidence"].iter_rows()
                for cell in row
                if cell.value is not None
            )
            assert "No external evidence provided — DD gap" in evidence_text
    finally:
        build_model.lc.fetch_public_comps = original_fetch


def test_private_and_transaction_comps_are_included_in_comparable_evidence() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "private_comps.yaml"
        out = Path(tmp) / "private_comps.xlsx"
        input_path.write_text(
            "\n".join([
                "private_comps:",
                "  - name: PrivateAI",
                "    company_type: private",
                "    source_type: funding round / press release",
                "    stage: Series B",
                "    geography: Japan",
                "    post_money: 12000000000",
                "    arr: 1000000000",
                "    currency: JPY",
                "    source_url: https://example.test/privateai-series-b",
                "    as_of_date: 2026-04-30",
                "    applicability_limits: ARR reported by company; verify security terms",
                "transaction_comps:",
                "  - name: Strategic SaaS acquisition",
                "    company_type: transaction",
                "    source_type: M&A announcement",
                "    enterprise_value: 50000000000",
                "    revenue: 5000000000",
                "    ebitda: 2000000000",
                "    as_of_date: 2025-12-15",
                "    source_url: https://example.test/ma-announcement",
            ]),
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)
        text = "\n".join(str(cell.value) for row in wb["Evidence"].iter_rows() for cell in row if cell.value is not None)

        assert "PrivateAI" in text
        assert "private" in text
        assert "Strategic SaaS acquisition" in text
        assert "transaction" in text
        assert "ARR reported by company" in text
        # Provided private + transaction comps join the median (12.0 + 10.0
        # with the three 8/9/10 public fakes → median 10.0).
        rev_mult_row = _row_for_label(wb, "Evidence", _row_label_startswith(wb, "Evidence", "EV / Revenue multiple"))
        assert wb["Evidence"].cell(rev_mult_row, 6).value == 10.0


def test_incomplete_provided_comps_are_registered_but_not_used_for_medians() -> None:
    comps = live_comps.provided_comps_from_raw([
        {"name": "Undated PrivateCo", "post_money": 100_000_000, "arr": 10_000_000},
        {"name": "Dated PrivateCo", "post_money": 120_000_000, "arr": 10_000_000, "as_of_date": "2026-01-01"},
    ])
    result = live_comps.summarize_comps(comps)

    assert comps[0].status == "needs review"
    assert "missing source date" in comps[0].error
    assert comps[1].status == "provided"
    assert result.revenue_multiple_median == 12.0


def test_sec_ticker_lookup_is_cached_and_errors_are_compact() -> None:
    original_json_url = live_comps._json_url
    original_cache = live_comps._SEC_TICKER_LOOKUP_CACHE
    calls: list[str] = []

    def fake_json_url(url: str, *, timeout: float) -> dict:
        calls.append(url)
        return {
            "0": {"ticker": "CRM", "cik_str": 1108524, "title": "Salesforce, Inc."},
            "1": {"ticker": "NOW", "cik_str": 1373715, "title": "ServiceNow, Inc."},
        }

    try:
        live_comps._SEC_TICKER_LOOKUP_CACHE = None
        live_comps._json_url = fake_json_url
        assert live_comps._sec_cik_for_ticker("CRM", timeout=1.0) == ("0001108524", "Salesforce, Inc.")
        assert live_comps._sec_cik_for_ticker("NOW", timeout=1.0) == ("0001373715", "ServiceNow, Inc.")
        assert len(calls) == 1
        assert len(live_comps._compact_error("line one\n" + ("x" * 500), limit=80)) == 80
    finally:
        live_comps._json_url = original_json_url
        live_comps._SEC_TICKER_LOOKUP_CACHE = original_cache


def test_failed_live_comps_do_not_pollute_company_specific_sources_or_mark_live_multiples() -> None:
    def fake_failed_comps(tickers: list[str], *, timeout: float = 8.0) -> live_comps.PublicCompsResult:
        comps = [
            live_comps.PublicComp(
                ticker=ticker,
                name="",
                currency="",
                market_cap=None,
                enterprise_value=None,
                revenue_multiple=None,
                ebitda_multiple=None,
                source_url=f"https://example.test/{ticker}",
                as_of_date="2026-05-17",
                status="failed",
                error="HTTP Error 401: Unauthorized",
            )
            for ticker in tickers
        ]
        return live_comps.PublicCompsResult(
            comps=comps,
            revenue_multiple_median=None,
            ebitda_multiple_median=None,
            source_url="https://example.test/live-comps",
            as_of_date="2026-05-17",
        )

    original_fetch = build_model.lc.fetch_public_comps
    try:
        build_model.lc.fetch_public_comps = fake_failed_comps
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "failed.xlsx"
            build_model.build_model(None, out, mode="full")
            wb = load_workbook(out, data_only=False)
            evidence_text = "\n".join(str(cell.value) for row in wb["Evidence"].iter_rows() for cell in row if cell.value is not None)
            # Failed fetches stay VISIBLE as failed rows with their error...
            assert "failed: HTTP Error 401" in evidence_text
            assert "HTTP Error 401: Unauthorized" in evidence_text
            # ...the benchmark register declares the evidence gap instead of
            # inventing sources, and multiples are flagged as fallbacks.
            assert "No external evidence provided — DD gap" in evidence_text
            assert "fall back to profile defaults" in evidence_text
            # No usable median may be marked as a live peer-set multiple.
            for row in wb["Evidence"].iter_rows(min_col=3, max_col=3):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("EV / Revenue multiple"):
                        raise AssertionError("failed comps must not render a live peer multiple row")
    finally:
        build_model.lc.fetch_public_comps = original_fetch


def _load_quality_gates_module():
    path = SKILL_DIR / "build" / "evals" / "quality_gates.py"
    spec = importlib.util.spec_from_file_location("startup_finance_quality_gates", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_quality_gate_rejects_uniform_wrong_period_width() -> None:
    qg = _load_quality_gates_module()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "wrong_width.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Period Sheet"
        ws.freeze_panes = "F7"
        for col in range(6, 8):
            ws.cell(row=5, column=col).value = 12
            ws.cell(row=6, column=col).value = f"FY{col}"
            ws.column_dimensions[get_column_letter(col)].width = 16.0
        wb.save(out)

        result = qg.gate_widths_freeze([out])
        assert result.points == 0.0
        assert any("expected=11.5" in item for item in result.evidence)


def test_quality_gate_build_targets_disable_live_comps_by_default() -> None:
    qg = _load_quality_gates_module()
    commands: list[list[str]] = []

    def fake_run(cmd: list[str], timeout: int = 900) -> tuple[int, str]:
        commands.append(cmd)
        out = Path(cmd[cmd.index("--output") + 1])
        wb = Workbook()
        wb.save(out)
        return 0, ""

    original_run = qg._run
    try:
        qg._run = fake_run
        with tempfile.TemporaryDirectory() as tmp:
            qg.build_targets(Path(tmp), allow_live_comps=False)
            assert commands
            assert all("--no-live-comps" in cmd for cmd in commands)

            commands.clear()
            qg.build_targets(Path(tmp), allow_live_comps=True)
            assert commands
            assert all("--no-live-comps" not in cmd for cmd in commands)
    finally:
        qg._run = original_run


def _formula_cells(wb) -> list[tuple[str, str, str]]:
    formulas: list[tuple[str, str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((ws.title, cell.coordinate, cell.value))
    return formulas


def _formula_error_cells(wb) -> list[str]:
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    return errors


def _defined_name_count(wb) -> int:
    count = len(wb.defined_names)
    for ws in wb.worksheets:
        count += len(ws.defined_names)
    return count


def _font_rgb(cell) -> str | None:
    color = cell.font.color
    if color is None or getattr(color, "type", None) != "rgb":
        return None
    value = getattr(color, "rgb", None) or getattr(color, "value", None)
    if not isinstance(value, str):
        return None
    return value[-6:].upper()


def _chart_axis_title_text(chart) -> str:
    title = getattr(getattr(chart, "y_axis", None), "title", None)
    if title is None:
        return ""
    try:
        paragraphs = title.tx.rich.p
        return "".join(run.t for paragraph in paragraphs for run in (paragraph.r or []) if run.t)
    except AttributeError:
        return str(title)


def _tab_rgb(ws) -> str | None:
    color = ws.sheet_properties.tabColor
    if color is None or getattr(color, "type", None) != "rgb":
        return None
    value = getattr(color, "rgb", None)
    return value[-6:].upper() if isinstance(value, str) else None


def _sample_source_workbook(source_text: str):
    tmp = tempfile.TemporaryDirectory()
    source_md = Path(tmp.name) / "source.md"
    out = Path(tmp.name) / "source_plan.xlsx"
    source_md.write_text(source_text, encoding="utf-8")
    source_plan.build_source_plan_workbook(source_md, out)
    return tmp, load_workbook(out, data_only=False)


def test_full_model_uses_direct_formula_refs() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")

        wb = load_workbook(out, data_only=False)
        assert wb.sheetnames == build_model.resolve_bundle("full")
        assert _defined_name_count(wb) == 0

        # Volatile / indirection functions stay banned. INDEX is a deliberate
        # v2 mechanism (scenario toggle, COUNTIF-driven FY-end pulls) — the
        # protective intent (no opaque aliasing, no defined names) is kept by
        # the defined-name assertion above and the indirection ban below.
        alias_formulas = [
            f"{sheet}!{cell}: {formula}"
            for sheet, cell, formula in _formula_cells(wb)
            if "OFFSET(" in formula or "INDIRECT(" in formula
        ]
        assert alias_formulas == []


def test_intra_sheet_formula_cells_are_black() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")

        wb = load_workbook(out, data_only=False)
        violations = [
            f"{sheet}!{cell}: {formula} color={color}"
            for sheet, cell, formula in _formula_cells(wb)
            if "!" not in formula
            for color in [_font_rgb(wb[sheet][cell])]
            if color is not None and color != ib.IB_FORMULA
        ]
        assert violations == []


def test_ic_memo_dependency_closure_matches_live_formula_readouts() -> None:
    """--additional-sheets pulls in HARD dependencies only (no transitive
    explosion): IC Memo's live readouts reference Valuation & Exit, so that
    sheet joins the bundle — and nothing else does."""
    bundle = build_model.resolve_bundle("market_sizing", additional_sheets=["IC Memo"])

    assert "IC Memo" in bundle
    assert "Valuation & Exit" in bundle
    assert set(bundle) == {"Guide", "Evidence", "IC Memo", "Valuation & Exit"}


def test_cash_flow_runway_formula_floors_and_caps_runway() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")

        wb = load_workbook(out, data_only=False)
        formula = _first_year_cell_for_label(wb, "CF", "Runway months").value
        fcf_row = _row_for_label(wb, "CF", "Free cash flow")
        end_row = _row_for_label(wb, "CF", "Ending cash")

        # Negative cash is floored at zero; a non-negative free cash flow is
        # capped at 99 months (spelled 100-1 from the constant whitelist) so
        # runway never blows up; burn is normalized by the months ruler (F$5)
        # so the same row formula is grain-correct (R17).
        f = FIRST_VALUE_LETTER
        assert formula == (
            f"=IF({f}{fcf_row}>=0,100-1,MIN(100-1,"
            f"MAX(0,{f}{end_row})/(ABS({f}{fcf_row})/{f}$5)))"
        )


def test_cross_sheet_formula_cells_are_green() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")

        wb = load_workbook(out, data_only=False)
        # Declared exceptions to the green cross-sheet rule: check rows keep
        # the numeric-delta check role (OK/ERROR format, black), and the row-2
        # master-check echo renders as a gray note on every period sheet.
        violations = [
            f"{sheet}!{cell}: {formula} color={color}"
            for sheet, cell, formula in _formula_cells(wb)
            if "!" in formula
            for color in [_font_rgb(wb[sheet][cell])]
            if color != ib.IB_LINK_INTRA
            and wb[sheet][cell].number_format != source_plan.FMT_CHECK_V2
            and not (wb[sheet][cell].row == 2 and color == ib.IB_COMMENT)
        ]
        assert violations == []


def test_pricing_bundle_is_formula_complete() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "pricing.xlsx"
        build_model.build_model(None, out, mode="pricing")

        wb = load_workbook(out, data_only=False)
        assert wb.sheetnames == build_model.resolve_bundle("pricing")
        # LEAN closed bundle: pricing = 4 tabs, no transitive explosion.
        assert wb.sheetnames == ["Guide", "Summary", "Assumptions", "Pricing"]
        all_text = "\n".join(str(cell.value) for ws in wb.worksheets for row in ws.iter_rows() for cell in row if cell.value is not None)
        assert "compact mode placeholder" not in all_text
        assert "Price support ratio" in all_text
        assert "Pricing validation plan" in all_text
        assert build_model.audit_workbook(wb) == []
        assert _defined_name_count(wb) == 0


def test_cap_table_mode_uses_state_machine_not_full_workbook() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "cap_table.xlsx"
        build_model.build_model(None, out, mode="cap_table")

        wb = load_workbook(out, data_only=False)
        assert wb.sheetnames == ["Guide", "Kernel", "Ownership"]
        assert {"P&L", "BS", "CF", "Capital Stack", "Valuation", "IC Memo"}.isdisjoint(wb.sheetnames)
        ownership_text = "\n".join(
            str(cell.value)
            for row in wb["Ownership"].iter_rows()
            for cell in row
            if cell.value is not None
        )
        assert "Round Event State Machine" in ownership_text
        assert "SAFE / J-KISS Conversions" in ownership_text
        assert "Exit Waterfall" in ownership_text


def test_all_modes_produce_expected_bundles() -> None:
    # Closed lean bundles (BLUEPRINT_S3 MODE_BUNDLE_SEEDS) — no transitive
    # dependency explosion. full may drop BS when its drivers are immaterial.
    expected_sheet_counts = {
        "full": {11, 12},
        "pricing": {4},
        "unit_economics": {4},
        "burn_runway": {5},
        "cap_table": {3},
        "three_statement": {6},
        "ma_exit": {5},
        "dcf_only": {4},
        "market_sizing": {2},
        "comps_only": {4},
    }
    assert set(expected_sheet_counts) == set(build_model.VALID_MODES)
    with tempfile.TemporaryDirectory() as tmp:
        for mode in build_model.VALID_MODES:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)

            wb = load_workbook(out, data_only=False)
            assert len(wb.sheetnames) in expected_sheet_counts[mode], (
                f"{mode}: unexpected bundle {wb.sheetnames}"
            )
            facts, raw = build_model._facts_for_inputs(None, None)
            facts = build_model._facts_with_mode_defaults(facts, mode, raw, None)
            assert wb.sheetnames == build_model.resolve_bundle(mode, facts=facts)
            assert _defined_name_count(wb) == 0
            assert _merged_count(wb) == 0
            assert _wrapped_cells(wb) == []
            assert _manual_line_break_violations(wb) == []
            assert _styled_blank_cells(wb) == []
            assert _font_design_violations(wb) == []
            assert _semantic_alignment_violations(wb) == []
            # Freeze polarity (v2): period-axis sheets freeze at (first
            # period col, row 7); non-period sheets carry no freeze.
            for ws in wb.worksheets:
                anchor_col = build_model._v2_period_anchor_col(ws)
                if anchor_col is not None:
                    expected = ws.cell(row=7, column=anchor_col).coordinate
                    assert ws.freeze_panes == expected, (
                        f"{mode}:{ws.title} freeze={ws.freeze_panes} expected={expected}"
                    )
                else:
                    assert ws.freeze_panes is None, f"{mode}:{ws.title}"


def test_strict_audit_blocks_workbook_design_regressions() -> None:
    """Build a clean workbook, corrupt it one defect at a time, and assert
    every audit — including the S5 promoted audits — fires with a precise
    sheet!cell message."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)
        assert build_model.audit_workbook(wb) == []

        # --- legacy audits ------------------------------------------------
        guide = wb["Guide"]
        guide.merge_cells("K1:L1")
        marker_row = _row_for_label(wb, "Guide", "Model qualifications")
        guide.cell(marker_row, 2).value = "Removed marker"
        guide["C7"].alignment = Alignment(wrap_text=True)
        guide["C7"].font = Font(name=ib.FONT_FAMILY, size=16)
        guide["K2"].value = "manual\nline break"

        # --- S5 promoted audits (synthetic violations) ---------------------
        pl = wb["P&L"]
        # 1. freeze anchor: break the F7 contract on a period sheet.
        pl.freeze_panes = "A1"
        # 2. width uniformity: widen one period column.
        pl.column_dimensions["G"].width = 15.0
        # 3. unit/format agreement: money row rendered with a percent format.
        rev_row = _row_for_label(wb, "P&L", "Total revenue")
        pl.cell(rev_row, 8).number_format = ib.FMT_JP_PERCENT
        # (probe order: make the corrupted cell the first numeric/formula one)
        pl.cell(rev_row, 6).value = None
        pl.cell(rev_row, 7).value = None
        # 4. hardcoded constant outside the whitelist.
        ebitda_row = _row_for_label(wb, "P&L", "EBITDA")
        pl.cell(ebitda_row, 9).value = "=F9*7777"
        # 5. R17: the same row now differs across period columns.
        # (the edit above already breaks row-formula consistency)
        # 6. check-row presence: retitle the Summary master check.
        master_row = _row_for_label(wb, "Summary", "Master check")
        wb["Summary"].cell(master_row, 3).value = "Renamed row"
        for row in wb["BS"].iter_rows():
            for cell in row:
                if cell.number_format == source_plan.FMT_CHECK_V2:
                    cell.number_format = "General"

        issues = build_model.audit_workbook(wb)
        assert any("Guide has merged cell range" in issue for issue in issues)
        assert any("Guide is missing sheet-quality marker" in issue and "Model qualifications" in issue for issue in issues)
        assert any("C7 has wrap_text enabled" in issue for issue in issues)
        assert any("C7 uses non-standard font size" in issue for issue in issues)
        assert any("K2 contains a manual line break" in issue for issue in issues)
        assert any("P&L period-axis sheet must freeze at F7" in issue for issue in issues)
        assert any("P&L!G period column width 15.0 != 11.5" in issue for issue in issues)
        assert any(f"P&L!H{rev_row}" in issue and "disagrees" in issue for issue in issues)
        assert any(f"P&L!I{ebitda_row} hardcoded constant 7777" in issue for issue in issues)
        assert any(f"row formula varies across period columns" in issue and "P&L" in issue for issue in issues)
        assert any("Summary has no master check row" in issue for issue in issues)
        assert any("BS has no check row" in issue for issue in issues)


def test_generated_modes_do_not_reference_removed_sheets() -> None:
    sheet_ref = re.compile(r"'([^']+)'!")
    with tempfile.TemporaryDirectory() as tmp:
        for mode in build_model.VALID_MODES:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)
            missing_refs = []
            neutralized_cells = []
            for sheet, cell, formula in _formula_cells(wb):
                for target in sheet_ref.findall(formula):
                    if target not in wb.sheetnames:
                        missing_refs.append(f"{mode}:{sheet}!{cell}->{target}")
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == "not in bundle":
                            neutralized_cells.append(f"{mode}:{ws.title}!{cell.coordinate}")
            assert missing_refs == []
            assert neutralized_cells == []


def test_structured_yaml_controls_grain_periods_and_drivers() -> None:
    """YAML controls the v2 axis and drivers: grain/periods/start_year/
    fiscal_year_end_month shape the period columns; statutory_welfare_rate,
    consumption_tax_rate, and AR/AP site months reach their JP-practice
    consumers; stated driver series land as blue inputs."""
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "structured.xlsx"
        input_path.write_text(
            "\n".join(
                [
                    "company: KernelCo",
                    "grain: monthly",
                    "periods: 6",
                    "start_year: 2026",
                    "fiscal_year_end_month: 6",
                    "statutory_welfare_rate: 15%",
                    "consumption_tax_rate: 10%",
                    "ar_site_months: 1.5",
                    "ap_site_months: 2",
                    "customers: [10, 20, 30, 40, 50, 60]",
                    "monthly_price_yen: [100000, 110000, 120000, 130000, 140000, 150000]",
                    "target_gross_margin: [65%, 65%, 68%, 68%, 70%, 70%]",
                    "equity_raise_yen: [10000000, 0, 0, 0, 0, 0]",
                    "beginning_cash_yen: 50000000",
                ]
            ),
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)

        first_col = FIRST_VALUE_COL
        a = wb["Assumptions"]
        # FY2026 with FYE=6 starts 2025/07: six real calendar-month columns.
        assert [a.cell(6, c).value for c in range(first_col, first_col + 6)] == [
            "2025/07", "2025/08", "2025/09", "2025/10", "2025/11", "2025/12",
        ]
        # Months ruler: monthly grain = 1 month per column.
        assert [a.cell(5, c).value for c in range(first_col, first_col + 6)] == [1] * 6
        # Stated drivers land as blue inputs at the stated per-period values.
        customers_row = _row_for_label(wb, "Assumptions", "Total customers (ending)")
        assert a.cell(customers_row, first_col).value == 10
        price_row = _row_for_label(wb, "Assumptions", _row_label_startswith(wb, "Assumptions", "Monthly price"))
        assert a.cell(price_row, first_col).value == 100000
        assert a.cell(price_row, first_col + 5).value == 150000
        assert _font_rgb(a.cell(price_row, first_col)) == ib.IB_HARD_INPUT
        gm_row = _row_for_label(wb, "Assumptions", "Target gross margin")
        assert a.cell(gm_row, first_col).value == 0.65
        # JP-practice parameters reach their consumers.
        welfare_row = _row_for_label(wb, "Assumptions", "Statutory welfare rate")
        assert a.cell(welfare_row, first_col).value == 0.15
        loaded_row = _row_for_label(wb, "People Plan", _row_label_startswith(wb, "People Plan", "Fully loaded comp"))
        base_row = _row_for_label(wb, "People Plan", _row_label_startswith(wb, "People Plan", "Avg base salary"))
        rate_row = _row_for_label(wb, "People Plan", "Statutory welfare rate")
        assert wb["People Plan"].cell(loaded_row, first_col).value == (
            f"={FIRST_VALUE_LETTER}{base_row}*(1+{FIRST_VALUE_LETTER}{rate_row})"
        )
        ar_formula = _first_year_cell_for_label(wb, "CF", "Accounts receivable").value
        ar_site_row = _row_for_label(wb, "Assumptions", "AR collection site")
        assert f"'Assumptions'!F{ar_site_row}" in ar_formula, "stated AR site months must drive AR (サイト優先)"
        fye_row = _row_for_label(wb, "Assumptions", "Fiscal year end month")
        assert a.cell(fye_row, first_col).value == 6
        tax_row = _row_for_label(wb, "Assumptions", "Consumption tax rate")
        assert a.cell(tax_row, first_col).value == 0.10
        # Financing + opening cash.
        equity_row = _row_for_label(wb, "Assumptions", "Equity financing")
        assert a.cell(equity_row, first_col).value == 10000000
        cash_row = _row_for_label(wb, "Assumptions", _row_label_startswith(wb, "Assumptions", "Beginning cash"))
        assert a.cell(cash_row, first_col).value == 50000000


def test_balance_sheet_has_opening_capital_counterpart() -> None:
    """The BS books opening cash as paid-in capital (first period only, via
    the months-ruler prior-column guard) plus cumulative equity raises, and
    reconciles with an explicit rounded balance check."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        f = FIRST_VALUE_LETTER
        capital_row = _row_for_label(wb, "BS", _row_label_startswith(wb, "BS", "Paid-in capital"))
        cash_row = _row_for_label(wb, "Assumptions", _row_label_startswith(wb, "Assumptions", "Beginning cash"))
        equity_row = _row_for_label(wb, "Assumptions", "Equity financing")
        assert wb["BS"].cell(capital_row, FIRST_VALUE_COL).value == (
            f"=N(E{capital_row})+IF(N(E$5)=0,'Assumptions'!$F${cash_row},0)"
            f"+'Assumptions'!{f}{equity_row}"
        )
        assets_row = _row_for_label(wb, "BS", "Total assets")
        liabilities_row = _row_for_label(wb, "BS", "Total liabilities")
        equity_total_row = _row_for_label(wb, "BS", "Total equity")
        check_row = _row_for_label(wb, "BS", "Balance check")
        assert wb["BS"].cell(check_row, FIRST_VALUE_COL).value == (
            f"=ROUND({f}{assets_row}-{f}{liabilities_row}-{f}{equity_total_row},0)"
        )


def test_mfn_applied_only_for_changed_safe_terms() -> None:
    inp = cap_table.CapTableInput(
        company_name="MFN Test",
        safes=[
            cap_table.SAFEInstrument(
                name="Best terms SAFE",
                type="j_kiss_v2",
                principal_money_m=20,
                cap_money_m=1_000,
                discount=0.20,
                mfn=True,
            ),
            cap_table.SAFEInstrument(
                name="Changed SAFE",
                type="j_kiss_v2",
                principal_money_m=20,
                cap_money_m=1_400,
                discount=0.10,
                mfn=True,
            ),
        ],
        round_pre_money_money_m=2_000,
        round_investment_money_m=500,
        round_target_pool_pct=0.10,
    )

    result = cap_table.run_round_state_machine(inp)
    mfn_flags = {safe.name: safe.mfn_applied for safe in result.safe_conversions}

    assert mfn_flags == {
        "Best terms SAFE": False,
        "Changed SAFE": True,
    }


def test_cap_table_rebuild_clears_prior_sheet_fills() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ownership"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    cap_table.build_cap_table_sheet(wb, cap_table.CapTableInput(company_name="Fill Test"))

    assert wb["Ownership"]["A1"].fill.fill_type in (None, "none")


def test_cap_table_rebuild_clears_legacy_layout_state() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ownership"
    ws.freeze_panes = "C8"
    ws.merge_cells("B2:C2")
    ws["B1"] = "legacy"
    ws["B1"].alignment = Alignment(wrap_text=True, indent=2)

    cap_table.build_cap_table_sheet(wb, cap_table.CapTableInput(company_name="Layout State Test"))
    ws = wb["Ownership"]

    assert _merged_count(wb) == 0
    assert ws.freeze_panes is None
    assert ws.column_dimensions["A"].width == ib.COL_MARGIN_WIDTH
    assert ws.column_dimensions["B"].width == ib.COL_HIERARCHY_WIDTH
    assert ws["C1"].value == "Layout State Test — Cap Table"
    assert _wrapped_cells(wb) == []
    assert _leading_space_labels(wb) == []
    assert [
        f"{cell.coordinate}: {cell.alignment.indent}"
        for row in ws.iter_rows()
        for cell in row
        if getattr(cell.alignment, "indent", 0)
    ] == []


def test_cap_table_sheet_uses_canonical_design_surface() -> None:
    inp = cap_table.CapTableInput(
        company_name="Design Surface Test",
        safes=[
            cap_table.SAFEInstrument(
                name="Seed SAFE",
                type="j_kiss_v2",
                principal_money_m=20,
                cap_money_m=1_000,
                discount=0.20,
            )
        ],
        round_pre_money_money_m=2_000,
        round_investment_money_m=500,
        round_target_pool_pct=0.10,
    )
    round_result = cap_table.run_round_state_machine(inp)
    waterfall = [
        (
            3_000.0,
            cap_table.compute_exit_waterfall(inp=inp, exit_value_money_m=3_000.0),
        )
    ]
    wb = Workbook()
    ws = cap_table.build_cap_table_sheet(
        wb,
        inp,
        round_result=round_result,
        waterfall_scenarios=waterfall,
    )

    assert wb.sheetnames == ["Ownership"]
    assert _merged_count(wb) == 0
    assert ws.freeze_panes is None
    assert ws.sheet_view.showGridLines is False
    assert ws.print_area
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.print_title_rows in {"1:3", "$1:$3"}
    assert ws.print_title_cols in {"A:C", "$A:$C"}
    assert _column_width(ws, "A") == ib.COL_MARGIN_WIDTH
    assert _column_width(ws, "B") == ib.COL_HIERARCHY_WIDTH
    assert _column_width(ws, "C") >= ib.COL_LABEL_WIDTH
    assert _column_width(ws, "G") >= ib.COL_SOURCE_WIDTH
    assert _wrapped_cells(wb) == []
    assert _manual_line_break_violations(wb) == []
    assert _styled_blank_cells(wb) == []
    assert _design_band_span_violations(wb) == []
    assert _repeated_semantic_fill_rows(wb) == []
    assert _unknown_fill_color_violations(wb) == []
    assert _repeated_prominent_border_rows(wb) == []
    assert _hierarchy_spacer_border_violations(wb) == []
    assert _font_design_violations(wb) == []
    assert _border_density_violations(wb) == []
    assert _memo_note_border_violations(wb) == []
    assert _border_style_violations(wb) == []
    assert _border_color_violations(wb) == []
    assert _row_height_violations(wb) == []
    for row in (5, 25, 50, 70):
        assert [_fill_rgb(ws.cell(row=row, column=col)) for col in range(2, 8)] == [
            ib.BG_HEADER_BAND
        ] * 6


def _chart_count(wb) -> int:
    return sum(len(getattr(ws, "_charts", [])) for ws in wb.worksheets)


def _merged_count(wb) -> int:
    return sum(len(ws.merged_cells.ranges) for ws in wb.worksheets)


def _formula_count(wb) -> int:
    return len(_formula_cells(wb))


def _wrapped_cells(wb) -> list[str]:
    return [
        f"{ws.title}!{cell.coordinate}"
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.alignment is not None and cell.alignment.wrap_text is True
    ]


def test_ib_helpers_reject_wrap_text_true() -> None:
    wb = Workbook()
    ws = wb.active

    for apply_func in (ib.apply_label, ib.apply_comment):
        try:
            apply_func(ws["A1"], wrap_text=True)
        except ValueError as exc:
            assert "forbids wrap_text=True" in str(exc)
        else:
            raise AssertionError(f"{apply_func.__name__} accepted wrap_text=True")

    ib.apply_label(ws["A2"])
    ib.apply_comment(ws["A3"])
    assert ws["A2"].alignment.wrap_text is False
    assert ws["A3"].alignment.wrap_text is False
    assert ib.wrapped_text_row_height(3) == ib.ROW_HEIGHT_PER_WRAPPED_LINE * 3
    assert ib.wrapped_text_row_height(0) == ib.ROW_HEIGHT_PER_WRAPPED_LINE
    assert ib.visible_text_line_count("one\ntwo", "single") == 2
    ib.set_wrapped_exception_row_height(ws, 4, ib.visible_text_line_count("one\ntwo\nthree"))
    assert ws.row_dimensions[4].height == ib.ROW_HEIGHT_PER_WRAPPED_LINE * 3

    ws["A5"] = "bounded\nprose"
    try:
        ib.apply_wrapped_exception(
            ws["A5"],
            user_approved=False,
            bounded_prose=True,
            adjacent_cells_carry_meaning=True,
        )
    except ValueError as exc:
        assert "explicit user approval" in str(exc)
    else:
        raise AssertionError("apply_wrapped_exception accepted missing approval")

    try:
        ib.apply_wrapped_exception(
            ws["A5"],
            user_approved=True,
            bounded_prose=False,
            adjacent_cells_carry_meaning=True,
        )
    except ValueError as exc:
        assert "bounded table prose" in str(exc)
    else:
        raise AssertionError("apply_wrapped_exception accepted horizontal-read text")

    try:
        ib.apply_wrapped_exception(
            ws["A5"],
            user_approved=True,
            bounded_prose=True,
            adjacent_cells_carry_meaning=False,
        )
    except ValueError as exc:
        assert "adjacent table cells carry meaningful" in str(exc)
    else:
        raise AssertionError("apply_wrapped_exception accepted blank overflow text")

    ib.apply_wrapped_exception(
        ws["A5"],
        user_approved=True,
        bounded_prose=True,
        adjacent_cells_carry_meaning=True,
    )
    assert ws["A5"].alignment.wrap_text is True
    assert ws.row_dimensions[5].height == ib.wrapped_text_row_height(2)
    ws["A6"] = "bounded\nprose"
    ib.apply_wrapped_exception(
        ws["A6"],
        user_approved=True,
        bounded_prose=True,
        adjacent_cells_carry_meaning=True,
        line_count=0,
    )
    assert ws.row_dimensions[6].height == ib.wrapped_text_row_height(1)


def test_runtime_builders_do_not_use_wrap_or_merge_layout_shortcuts() -> None:
    runtime_files = [
        SCRIPTS_DIR / "build_model.py",
        SCRIPTS_DIR / "source_plan_builder.py",
        SCRIPTS_DIR / "cap_table_builder.py",
    ]
    for path in runtime_files:
        text = path.read_text(encoding="utf-8")
        assert "wrap_text=True" not in text
        assert "apply_wrapped_exception(" not in text
        text_without_unmerge = text.replace("unmerge_cells(", "")
        assert ".merge_cells(" not in text_without_unmerge
        assert "merge_cells(" not in text_without_unmerge

    ib_text = (SCRIPTS_DIR / "ib_format.py").read_text(encoding="utf-8")
    ib_text_without_unmerge = ib_text.replace("unmerge_cells(", "")
    assert ".merge_cells(" not in ib_text_without_unmerge
    assert "merge_cells(" not in ib_text_without_unmerge
    assert "blank unstyled overflow cells without merging" in ib.WRAP_TEXT_ERROR
    assert "use_user_approved_bounded_prose_wrap_only" in ib.WRAP_DECISION_LADDER
    assert "final print/render column" in ib.WRAP_BEST_PRACTICE


def test_skill_guidance_makes_no_wrap_rule_explicit() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    layout_text = (SKILL_DIR / "build" / "references" / "_layout_canonical.md").read_text(encoding="utf-8")
    design_text = (SKILL_DIR / "build" / "references" / "_ib_workbook_design_system.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    ib_text = (SCRIPTS_DIR / "ib_format.py").read_text(encoding="utf-8")
    skill_full = "\n".join([skill_text, invocation_text])
    skill_flat = " ".join(skill_full.split())
    layout_flat = " ".join(layout_text.split())
    design_flat = " ".join(design_text.split())

    combined = "\n".join([skill_text, invocation_text, layout_text, design_text, self_review_text, eval_text, ib_text])
    assert "No-Wrap Rule" in combined
    assert "Treat text wrapping as prohibited" in skill_full
    assert "no merged cells" in skill_flat
    assert "blank unmerged unstyled overflow cells" in skill_flat
    assert "reject `wrap_text=True`" in design_text
    assert "Do not use merged cells as the repair" in design_flat
    assert "明示的に wrap_text=True" not in ib_text
    assert "set row height to the exact rendered visible line count" in skill_flat
    assert "exact line-count row heights" in layout_flat
    assert "No merged cells, ever" in layout_flat
    assert "No wrapped text in generated cells" in layout_flat
    assert "exact rendered visible line count" in design_flat
    assert "wrapped_text_row_height" in ib_text
    assert "set_wrapped_exception_row_height" in ib_text
    assert "exact rendered visible-line-count row height" in eval_text
    assert "no unnecessary wrapping on horizontal-read" in eval_text
    assert "IB wrap decision ladder" in design_text
    assert "Use the IB wrap decision ladder" in skill_full
    assert "Horizontal-read rows must also have visual runway" in design_text
    assert "Manual line breaks are treated as wrapped exceptions" in design_text
    assert "blank unstyled overflow cells" in combined
    assert "user-approved prose or table cells" not in combined
    assert "blank overflow space" not in combined
    assert "bounded-prose" not in combined
    assert "bounded table prose" in combined
    assert "print/render boundary" in combined
    assert "XLSX_WRAP_DISCIPLINE" in eval_text


def test_skill_guidance_requires_fix_and_rerun_iteration() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    skill_flat = " ".join("\n".join([skill_text, invocation_text]).split())
    self_review_flat = " ".join(self_review_text.split())
    eval_flat = " ".join(eval_text.split())

    assert "fix the concrete failed items and rerun the same checks" in skill_flat
    assert "Do not treat model construction as completion" in skill_flat
    assert "command checks and rendered-output inspection for both finance logic and sheet design" in skill_flat
    assert "Keep iterating until the model logic and the visible sheet design are both sufficient" in skill_flat
    assert "command-based workbook checks and rendered-output inspection" in self_review_flat
    assert "Apply both to the finance model itself and to sheet design" in self_review_flat
    assert "Passing commands do not excuse a visibly poor sheet" in self_review_flat
    assert "a good-looking render does not excuse broken formulas" in self_review_flat
    assert "fix the concrete failed item, rerun the same check" in self_review_flat
    assert "Do not replace failed verification with a narrative explanation" in self_review_flat
    assert "Model creation is not completion" in eval_flat
    assert "verify both model logic and sheet design with command-based workbook checks plus rendered-output inspection" in eval_flat


def test_self_improvement_protocol_triggers_from_logs_and_feedback() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    self_improve_text = (SKILL_DIR / "build" / "references" / "_self_improvement_protocol.md").read_text(encoding="utf-8")
    prompt_research_text = (SKILL_DIR / "build" / "references" / "_ai_prompt_research_patterns.md").read_text(encoding="utf-8")
    combined_flat = " ".join("\n".join([skill_text, invocation_text, self_improve_text, prompt_research_text]).split())

    for phrase in [
        "post-output/session-log-driven improvements",
        "At every closeout and every improvement prompt",
        "user asks to improve or repair a prior skill output",
        "Tests, strict audit, quality gates, recalc/render checks",
        "failed checks, repeated workarounds, cleanup noise, or tool/routing inefficiency",
        "User feedback reveals a gap between the intended decision and the delivered workbook",
        "route session-log, trace, eval, reflection, and X-practitioner signals through `_self_improvement_protocol.md`",
    ]:
        assert phrase in combined_flat


def test_self_improvement_protocol_requires_regression_proof_and_privacy() -> None:
    review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    improve_text = (SKILL_DIR / "build" / "references" / "_self_improvement_protocol.md").read_text(encoding="utf-8")
    evals = json.loads((SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8"))["evals"]
    combined_flat = " ".join("\n".join([review_text, improve_text]).split())

    for phrase in [
        "Inspect only the minimum sanitized evidence needed to learn the failure mode",
        "Never store raw local logs, secrets, API keys, private source text, customer data, full conversation transcripts, or personal identifiers",
        "Classify the signal: one-off artifact issue, reusable skill gap, eval gap, runtime helper bug, reference/protocol gap, dependency/tooling gap, or documentation ambiguity",
        "Patch the lowest durable layer",
        "Add regression proof",
        "validate_reflection_record",
        "Before writing any Reflection Record",
        "Reflection Record",
        "failure pattern -> generalized rule -> required check",
        "Capability evals prove broad skill behavior",
        "regression evals protect known failures and should stay near-perfect",
        "holdout/adversarial examples should not be tuned against during the same iteration",
        "cost and efficiency signals such as wall time, repeated tool calls",
        "cost/latency or tool-call impact when it affected the run",
        "milestone/human review",
        "Rerun the check that failed and one broader gate",
        "Use X/public social posts only as weak signals",
        "Do not generalize a lesson into the skill",
        "company-specific, confidential, unsupported by evidence, or lacks a testable regression hook",
    ]:
        assert phrase in combined_flat

    missing = []
    for item in evals:
        refs = set(item.get("applicable_references", []))
        if "_self_improvement_protocol" not in refs:
            missing.append(f"{item['id']}:{item['name']} missing _self_improvement_protocol")
    assert missing == []
    assert any(
        assertion.get("id") == "SELF_IMPROVEMENT_PROTOCOL"
        for item in evals
        for assertion in item.get("assertions", [])
    )
    assert any(
        item.get("name") == "self_improvement_from_failed_session_log"
        and "failed command log" in item.get("prompt", "").lower()
        and "Reflection Record" in item.get("expected_output", "")
        for item in evals
    )


def test_self_improvement_reflection_validator_rejects_privacy_and_overfit_records() -> None:
    valid = {
        "task_type": "post_output_repair",
        "artifact_type": "xlsx",
        "redacted_evidence": "strict audit failed on row-formula consistency, then passed",
        "observed_failure": "manual repair loop repeated because the check was missing",
        "verification_evidence": "targeted pytest and quality gate passed",
        "root_cause_category": "eval_gap",
        "generalized_lesson": "When a repeated repair loop reveals a formula-shape defect, promote the invariant into a deterministic gate.",
        "proposed_change_layer": "quality_gate",
        "privacy_classification": "sanitized",
        "regression_proof": "pytest: self-improvement validator and quality gate G-J",
        "is_reusable": True,
        "public_signal_type": "x",
        "x_public_signal": "public posts summarized as weak workflow signal only",
        "milestone_review": True,
    }
    assert self_improvement.validate_reflection_record(valid) == []

    invalid = dict(valid)
    invalid.update({
        "redacted_evidence": "token=sk-proj-1234567890abcdefghijklmnop",
        "generalized_lesson": "For GateCo, hard-code 120000 yen after seeing @example_user's post.",
        "company_names": ["GateCo"],
        "raw_public_post": "copied post text",
        "changes_audit_pass_criteria": True,
        "milestone_review": False,
    })
    errors = set(self_improvement.validate_reflection_record(invalid))
    assert {
        "sensitive:named_secret",
        "sensitive:openai_key",
        "overfit:company_specific_lesson",
        "overfit:instance_numeric_lesson",
        "x_signal:handle_not_stripped",
        "x_signal:raw_post_not_allowed",
        "review_required:audit_or_doctrine_change",
    } <= errors


def test_self_improvement_panel_rejects_schema_valid_quality_degradation() -> None:
    degraded = {
        "task_type": "post_output_repair",
        "artifact_type": "xlsx",
        "redacted_evidence": "user said the workbook should feel more premium",
        "observed_failure": "closeout wording did not persuade the user",
        "verification_evidence": "manual review only",
        "root_cause_category": "documentation_ambiguity",
        "generalized_lesson": "Always improve.",
        "proposed_change_layer": "skill_trigger",
        "privacy_classification": "sanitized",
        "regression_proof": "self-review",
        "is_reusable": True,
        "record_author": "codex-runner",
        "panel_reviewer": "independent-reviewer-r1",
        "panel_evidence_citations": [],
        "evidence_count": 1,
    }
    assert self_improvement.validate_reflection_record(degraded) == []
    panel = self_improvement.score_reflection_panel(degraded)
    assert panel["accepted"] is False
    assert {
        "panel:weak_regression_proof",
        "panel:weak_generalized_lesson",
        "pruning:n1_requires_deferred_candidate",
    } <= set(panel["blockers"])


def test_self_improvement_panel_accepts_evidence_backed_independent_record() -> None:
    accepted = {
        "task_type": "post_output_repair",
        "artifact_type": "xlsx",
        "redacted_evidence": "quality gate G-D failed in two runs, then passed after a formula-shape assertion",
        "observed_failure": "period formulas diverged after repeated manual repair loops",
        "verification_evidence": "targeted pytest passed and quality_gates.py reran clean",
        "root_cause_category": "quality_gate_gap",
        "generalized_lesson": "When repeated workbook repairs expose the same formula-shape gap, promote the invariant into a deterministic gate before accepting closeout.",
        "proposed_change_layer": "quality_gate",
        "privacy_classification": "sanitized",
        "regression_proof": "pytest: self-improvement panel regression and quality gate G-K",
        "is_reusable": True,
        "record_author": "codex-runner",
        "panel_reviewer": "independent-reviewer-r2",
        "panel_evidence_citations": ["redacted_evidence", "verification_evidence", "regression_proof"],
        "evidence_count": 2,
        "broader_gate_to_rerun": "quality_gates.py",
        "expected_quality_effect": "turns a recurring manual repair into an auditable workbook gate",
        "milestone_review": True,
    }
    assert self_improvement.validate_reflection_record_for_acceptance(accepted) == []
    missing_reviewer = dict(accepted)
    missing_reviewer.pop("panel_reviewer")
    assert "panel:reviewer_required" in self_improvement.validate_reflection_record_for_acceptance(missing_reviewer)
    specific_language = dict(accepted)
    specific_language["generalized_lesson"] = (
        "When repeated workbook repairs expose the same formula-shape gap, "
        "promote the invariant into a deterministic gate to preserve high quality financial-model output."
    )
    specific_language["regression_proof"] = (
        "none of the previously passing tests regressed; pytest and quality_gates.py both passed clean"
    )
    assert self_improvement.validate_reflection_record_for_acceptance(specific_language) == []


def test_self_improvement_closeout_consistency_catches_links_and_count_drift(tmp_path: Path) -> None:
    protocol = tmp_path / "protocol.md"
    protocol.write_text(
        "[good](kept.md)\n"
        "[missing](missing.md)\n"
        "<!-- sfm-closeout-count name=\"refs\" expected=\"3\" glob=\"*.md\" -->\n",
        encoding="utf-8",
    )
    (tmp_path / "kept.md").write_text("ok\n", encoding="utf-8")

    errors = set(closeout_consistency.check_closeout_consistency([protocol]))
    assert "dangling_ref:protocol.md->missing.md" in errors
    assert "count_drift:protocol.md:refs:expected=3:found=2:glob=*.md" in errors


def test_self_improvement_failure_modes_and_pruning_are_linked_from_gates() -> None:
    protocol_text = (SKILL_DIR / "build" / "references" / "_self_improvement_protocol.md").read_text(encoding="utf-8")
    failure_text = (SKILL_DIR / "build" / "references" / "_self_improvement_failure_modes.md").read_text(encoding="utf-8")
    pruning_text = (SKILL_DIR / "build" / "references" / "_self_improvement_pruning.md").read_text(encoding="utf-8")
    gate_text = (SKILL_DIR / "build" / "evals" / "quality_gates.py").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    combined = "\n".join([protocol_text, failure_text, pruning_text, gate_text, eval_text])

    for phrase in [
        "_self_improvement_reviewer_panel.md",
        "_self_improvement_failure_modes.md",
        "_self_improvement_pruning.md",
        "model collapse",
        "reward hacking / Goodhart",
        "sycophancy",
        "eval overfit",
        "non-convergence",
        "n=1 evidence normally becomes a sanitized deferred candidate",
        "Rules fail at the prompt and succeed at the boundary",
        "score_reflection_panel",
        "closeout_consistency.py",
    ]:
        assert phrase in combined


def test_skill_guidance_uses_meaningful_sparse_fills_and_borders() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    layout_text = (SKILL_DIR / "build" / "references" / "_layout_canonical.md").read_text(encoding="utf-8")
    design_text = (SKILL_DIR / "build" / "references" / "_ib_workbook_design_system.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")

    skill_full = "\n".join([skill_text, invocation_text])
    combined = "\n".join([skill_full, layout_text, design_text, self_review_text, eval_text])
    combined_flat = " ".join(combined.split()).lower()
    assert "same non-heatmap fill" in combined_flat
    assert "background fills are selective accents for major semantic moments only" in combined_flat
    assert "avoid rainbow palettes" in combined_flat
    assert "role-based sparse fill palette" in eval_text
    assert "no rainbow or decorative background fills" in eval_text
    assert "ordinary calculation rows should stay white" in combined_flat
    assert "workbook tokens" in combined_flat
    assert "arial 10pt" in combined_flat
    assert "header / label row" in combined_flat
    assert "use one rectangular span per filled row component" in combined_flat
    assert "do not choose the end column only from cells that happen to contain text" in combined_flat
    assert "do not stop a fill because a cell is blank" in " ".join(skill_full.split()).lower()
    assert "for every filled row, name its role and inspect the start column, end column" in " ".join(self_review_text.split()).lower()
    assert "semantic row-span helper in\n`ib_format.py` for fill/border row components" in design_text
    assert "rectangular semantic fill and border spans including member blank cells" in eval_text
    assert "not a ragged set of populated cells" in combined_flat
    assert "consecutive heavy borders" in combined_flat
    assert "prominent borders follow the same meaning-first rule as fills" in combined_flat
    assert "do not stop a border because a cell is blank" in " ".join(skill_full.split()).lower()
    assert "draw the rule across that full span, including blank cells inside the component" in combined_flat
    assert "a blank cell inside the table/header/check width still receives the border" in combined_flat
    assert "check border span with the same positive rule as fill span" in combined_flat
    assert "dedicated hierarchy / indent columns stay borderless" in " ".join(skill_full.split()).lower()
    assert "b stays borderless and row rules begin at c" in combined_flat
    assert "confirm hierarchy / indent columns are not carrying row rules" in combined_flat
    assert "avoid repeating the same prominent top/bottom rule across adjacent rows" in " ".join(skill_full.split()).lower()
    assert "the same prominent border should not repeat on adjacent rows by default" in combined_flat
    assert "check border rhythm exactly as you check color rhythm" in combined_flat
    assert "no adjacent repeated prominent border rows" in eval_text
    assert "borders are not row-by-row decoration" in " ".join(skill_full.split()).lower()
    assert "memo, source, note, and interpretation cells are usually borderless" in combined_flat
    assert "three border styles by meaning" in combined_flat
    assert "border colors are black by default" in combined_flat
    assert "border color is black by default" in combined_flat
    assert "normal dotted for soft/provisional separations" in combined_flat
    assert "no per-row gridline borders" in eval_text
    assert "borderless memo/source/note cells" in eval_text
    assert "black border colors by default" in eval_text


def test_skill_guidance_enforces_ib_text_positioning() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    design_text = (SKILL_DIR / "build" / "references" / "_ib_workbook_design_system.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    ib_text = (RUNTIME_DIR / "ib_format.py").read_text(encoding="utf-8")
    combined_flat = " ".join("\n".join([skill_text, invocation_text, design_text, self_review_text, eval_text, ib_text]).split())

    for phrase in [
        "Investment-banking model alignment is functional, not decorative",
        "labels, sources, notes, titles, memos, and interpretation text are left-aligned",
        "numeric values, formulas, money, percentages, multiples, counts, and unit labels are right-aligned",
        "only period headers, scenario/matrix headers, and short column headers are centered",
        "Do not center long prose or labels",
        "Hierarchy is expressed with dedicated hierarchy/indent columns",
        "role-based left/right/center text alignment",
        "ALIGNMENT_BEST_PRACTICE",
    ]:
        assert phrase in combined_flat


def test_skill_guidance_enforces_ib_font_size_discipline() -> None:
    skill_text = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    design_text = (SKILL_DIR / "build" / "references" / "_ib_workbook_design_system.md").read_text(encoding="utf-8")
    self_review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    eval_text = (SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8")
    ib_text = (RUNTIME_DIR / "ib_format.py").read_text(encoding="utf-8")
    combined_flat = " ".join("\n".join([skill_text, invocation_text, design_text, self_review_text, eval_text, ib_text]).split())

    for phrase in [
        "Font size discipline is equally strict",
        "Investment-banking models should look dense, legible, and standardized",
        "Use Arial 10pt for ordinary body cells",
        "Use Arial 9pt italic gray for supporting source, note, unit-helper, footnote",
        "Use 10-11pt bold for section labels and compact header rows",
        "Use Arial 14pt bold for sheet titles and cover/title surfaces only",
        "Keep the generated cell-size set intentionally small: 9, 10, 11, and 14pt",
        "Do not use 8pt cell text to squeeze content",
        "constrained 9/10/11/14pt cell-size palette",
        "FONT_SIZE_ALLOWED_CELLS",
        "FONT_SIZE_BEST_PRACTICE",
    ]:
        assert phrase in combined_flat


def test_xlsx_evals_load_full_design_reference_stack() -> None:
    evals = json.loads((SKILL_DIR / "build" / "evals" / "evals.json").read_text(encoding="utf-8"))["evals"]
    required = {
        "_layout_canonical",
        "_ib_workbook_design_system",
        "_self_review_protocol",
        "_self_improvement_protocol",
        "_sheet_quality_rubric",
    }
    missing = []
    for item in evals:
        if str(item.get("name", "")).startswith("self_improvement_"):
            refs = set(item.get("applicable_references", []))
            assertions = {assertion.get("id") for assertion in item.get("assertions", [])}
            if "_self_improvement_protocol" not in refs:
                missing.append(f"{item['id']}:{item['name']} missing _self_improvement_protocol")
            if item.get("name") == "self_improvement_from_failed_session_log":
                for assertion_id in ("SI1", "SI2", "SI3", "SI4"):
                    if assertion_id not in assertions:
                        missing.append(f"{item['id']}:{item['name']} missing {assertion_id}")
            continue
        refs = set(item.get("applicable_references", []))
        if not required <= refs:
            missing.append(f"{item['id']}:{item['name']} missing {sorted(required - refs)}")
        assertions = item.get("assertions", [])
        if not any(assertion.get("id") == "XLSX_DESIGN" for assertion in assertions):
            missing.append(f"{item['id']}:{item['name']} missing XLSX_DESIGN assertion")
        if not any(assertion.get("id") == "XLSX_SHEET_QUALITY" for assertion in assertions):
            missing.append(f"{item['id']}:{item['name']} missing XLSX_SHEET_QUALITY assertion")
        if not any(assertion.get("id") == "XLSX_WRAP_DISCIPLINE" for assertion in assertions):
            missing.append(f"{item['id']}:{item['name']} missing XLSX_WRAP_DISCIPLINE assertion")
        if not any(assertion.get("id") == "XLSX_FONT_HIERARCHY" for assertion in assertions):
            missing.append(f"{item['id']}:{item['name']} missing XLSX_FONT_HIERARCHY assertion")
        for assertion in assertions:
            if assertion.get("id") == "XLSX_DESIGN":
                text = assertion.get("text", "")
                if "rectangular semantic fill and border spans including member blank cells" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing semantic border span guidance")
                if "no rainbow or decorative background fills" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing restrained fill palette guidance")
                if "role-based sparse fill palette" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing role-based fill palette guidance")
                if "table-width row rules" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing table-width row rule guidance")
                if "start after hierarchy spacer columns" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing hierarchy-border start guidance")
                if "no adjacent repeated prominent border rows" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing border rhythm guidance")
                if "no per-row gridline borders" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing no-gridline border guidance")
                if "borderless memo/source/note cells" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing memo borderless guidance")
                if "three semantic border styles" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing border style guidance")
                if "black border colors by default" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing black border color guidance")
                if "sheet-specific quality gates for purpose, source boundary, dependency flow, checks, and interpretation" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing sheet-specific quality guidance")
                if "IB-style text positioning" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing IB text positioning guidance")
                if "left-aligned labels/sources/notes" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing left text alignment guidance")
                if "right-aligned values/formulas/units" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing right value alignment guidance")
                if "centered period/scenario/matrix headers only" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing centered-header-only guidance")
                if "IB-style font-size discipline" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing IB font-size guidance")
                if "9/10/11/14pt generated cell palette" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing constrained font-size palette guidance")
                if "no 8pt squeezed cell text or 16pt+ presentation headings" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing font-size anti-pattern guidance")
            if assertion.get("id") == "XLSX_SHEET_QUALITY":
                text = assertion.get("text", "")
                if "sheet-specific purpose, traceable source/evidence boundary" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing sheet purpose/evidence quality guidance")
                if "Focused tasks include only sheets needed for the decision" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing focused-sheet restraint guidance")
            if assertion.get("id") == "XLSX_WRAP_DISCIPLINE":
                text = assertion.get("text", "")
                if "generated cells keep wrap_text off by default" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing no-wrap default assertion")
                if "blank unmerged unstyled overflow cells" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing blank overflow assertion")
                if "final print/render column" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing print/render boundary assertion")
                if "manual line breaks appear only for user-approved bounded table prose" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing manual-line-break assertion")
                if "exact rendered visible-line-count row height" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing rendered-visible row-height assertion")
            if assertion.get("id") == "XLSX_FONT_HIERARCHY":
                text = assertion.get("text", "")
                if "constrained IB hierarchy" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing constrained font hierarchy assertion")
                if "9pt italic gray supporting source/note/unit-helper cells" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing 9pt support-cell assertion")
                if "10-11pt bold compact section/header cells" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing compact section/header assertion")
                if "8pt squeezed text" not in text or "16pt+ presentation headings" not in text:
                    missing.append(f"{item['id']}:{item['name']} missing font-size anti-pattern assertion")
    assert missing == []


def test_semantic_fill_helper_uses_rectangular_span_including_blanks() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 2, "Section")
    ws.cell(3, 5, "Last populated")

    ib.apply_semantic_fill_span(ws, 3, 2, 6, ib.BG_HEADER_BAND, bottom=ib.THIN_LINE)

    for col in range(2, 7):
        assert _fill_rgb(ws.cell(3, col)) == ib.BG_HEADER_BAND
        assert ws.cell(3, col).border.bottom.style == "thin"
    assert ws.cell(3, 7).fill.fill_type in (None, "none")
    assert ws.cell(3, 7).border.bottom.style is None

    ib.apply_semantic_fill_span(
        ws, 4, 2, 6, ib.BG_HEADER_BAND, bottom=ib.THIN_LINE, border_start_col=3
    )
    assert _fill_rgb(ws.cell(4, 2)) == ib.BG_HEADER_BAND
    assert ws.cell(4, 2).border.bottom.style is None
    for col in range(3, 7):
        assert ws.cell(4, col).border.bottom.style == "thin"
    try:
        ib.apply_semantic_fill_span(
            ws, 5, 2, 6, ib.BG_HEADER_BAND, bottom=ib.THIN_LINE, border_start_col=7
        )
        raise AssertionError("border_start_col beyond the span should fail")
    except ValueError as exc:
        assert "must be within span 2:6" in str(exc)

    ib.apply_box_border(ws, "B7:D9", inner_dotted=True)
    assert ws["C8"].border.top.style == "dotted"
    assert ws["C8"].border.left.style == "dotted"

    source_plan._section(ws, 11, "Section before layout setup", end_col=6)
    assert _fill_rgb(ws.cell(11, source_plan.LAYOUT.first_hierarchy_col)) == ib.BG_HEADER_BAND
    assert ws.cell(11, source_plan.LAYOUT.first_hierarchy_col).border.bottom.style is None
    for col in range(source_plan.LAYOUT.label_col, 7):
        assert ws.cell(11, col).border.bottom.style == "thin"

    ws.cell(13, source_plan.LAYOUT.label_col, "Selected output")
    ws.cell(13, source_plan.LAYOUT.source_col, "source note")
    ws.cell(13, source_plan.LAYOUT.unit_col, "JPY")
    ws.cell(13, FIRST_VALUE_COL, 1)
    source_plan._highlight_row(ws, 13, FIRST_VALUE_COL)
    # Phase 1 Task 1.3b: _highlight_row now paints a continuous top + bottom
    # underline across the detected table block. The source / unit gap is
    # no longer a hole in the rule.
    for col in (
        source_plan.LAYOUT.label_col,
        source_plan.LAYOUT.source_col,
        source_plan.LAYOUT.unit_col,
        FIRST_VALUE_COL,
    ):
        assert ws.cell(13, col).border.top.style == "thin"
        assert ws.cell(13, col).border.bottom.style == "thin"


def _unit_label_violations(wb) -> list[str]:
    bad_fragments = ("単位:", "単位：", "Unit:", "JPY M", "JPY B")
    return [
        f"{ws.title}!{cell.coordinate}: {cell.value}"
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and any(fragment in cell.value for fragment in bad_fragments)
    ]


def _raw_money_scale_formula_violations(wb) -> list[str]:
    money_units = {"円", "千円", "百万円", "億円", "十億円", "兆円", "$", "$K", "$M"}
    money_scale_factors = {
        1_000,
        1_000_000,
        1_000_000_000,
        1_000_000_000_000,
    }
    violations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            unit = ws.cell(row=row[0].row, column=UNIT_COL).value
            if unit not in money_units:
                continue
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and any(re.search(rf"([*/])\s*{factor}\b", cell.value) for factor in money_scale_factors)
                ):
                    violations.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
    return violations


def _numbered_section_labels(wb) -> list[str]:
    import re

    pattern = re.compile(r"^\d+\.\s+\D")
    return [
        f"{ws.title}!{cell.coordinate}: {cell.value}"
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and pattern.match(cell.value.strip())
    ]


def _leading_space_labels(wb) -> list[str]:
    return [
        f"{ws.title}!{cell.coordinate}: {cell.value!r}"
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith((" ", "\t"))
    ]


def _missing_section_band_fills(wb) -> list[str]:
    missing = []
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            label = ws.cell(row_idx, source_plan.LAYOUT.first_hierarchy_col)
            if not isinstance(label.value, str):
                continue
            label_fill = label.fill
            is_section = (
                label_fill.fill_type == "solid"
                and isinstance(label_fill.fgColor.rgb, str)
                and label_fill.fgColor.rgb.endswith(ib.BG_HEADER_BAND)
            )
            if not is_section:
                continue
            font = label.font
            label_has_white_text = font.color is not None and str(font.color.rgb).endswith("FFFFFF")
            band_cells = []
            for col in range(source_plan.LAYOUT.first_hierarchy_col, ws.max_column + 1):
                cell = ws.cell(row_idx, col)
                if _fill_rgb(cell) == ib.BG_HEADER_BAND:
                    band_cells.append(cell)
            band_is_contiguous = band_cells and [cell.column for cell in band_cells] == list(range(band_cells[0].column, band_cells[-1].column + 1))
            following_cols = [
                cell.column
                for next_row in range(row_idx + 1, min(ws.max_row, row_idx + 3) + 1)
                for cell in ws[next_row]
                if cell.value is not None
            ]
            attached_width_col = max([source_plan.LAYOUT.label_col, *following_cols])
            if not (label_has_white_text and band_is_contiguous and band_cells[-1].column >= attached_width_col):
                missing.append(f"{ws.title}!{label.coordinate}")
    return missing


def _styled_blank_cells(wb) -> list[str]:
    bad = []
    semantic_fill_colors = {
        ib.BG_TABLE_HEADER,
        ib.BG_TOTAL_BAND,
        ib.BG_HEADER_BAND,
        ib.BG_WORKING,
    }
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, "") or cell.style_id == 0:
                    continue
                color = _fill_rgb(cell)
                if color in semantic_fill_colors:
                    same_fill_run = []
                    col = cell.column
                    while col >= 1 and _fill_rgb(ws.cell(cell.row, col)) == color:
                        same_fill_run.append(ws.cell(cell.row, col))
                        col -= 1
                    col = cell.column + 1
                    while col <= ws.max_column and _fill_rgb(ws.cell(cell.row, col)) == color:
                        same_fill_run.append(ws.cell(cell.row, col))
                        col += 1
                    if any(run_cell.value not in (None, "") for run_cell in same_fill_run):
                        continue
                # Blank cells with a deliberate border are legitimate too —
                # they are the middle of a table-block span painted by
                # apply_semantic_border_span (the source / unit gap between
                # the label column and the first period carrying the
                # continuous underline). Mirrors the runtime contract in
                # ib._is_intentional_blank_component_cell.
                border = cell.border
                if border is not None and any(
                    side is not None and getattr(side, "style", None) is not None
                    for side in (border.top, border.bottom, border.left, border.right)
                ):
                    continue
                bad.append(f"{ws.title}!{cell.coordinate}: style_id={cell.style_id}")
    return bad


def _manual_line_break_violations(wb) -> list[str]:
    violations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and "\n" in cell.value):
                    continue
                if cell.alignment is None or cell.alignment.wrap_text is not True:
                    violations.append(f"{ws.title}!{cell.coordinate}: manual line break without wrapped exception")
                    continue
                expected = ib.wrapped_text_row_height(ib.visible_text_line_count(cell.value))
                actual = ws.row_dimensions[cell.row].height
                if actual != expected:
                    violations.append(
                        f"{ws.title}!{cell.coordinate}: height={actual} expected={expected}"
                    )
    return violations


def _design_band_span_violations(wb) -> list[str]:
    violations = []
    band_colors = {ib.BG_TABLE_HEADER, ib.BG_WORKING}
    for ws in wb.worksheets:
        max_col = max(ws.max_column, 9)
        for row_idx in range(1, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=col) for col in range(2, max_col + 1)]
            populated = [cell for cell in row if cell.value is not None]
            colors = [_fill_rgb(cell) for cell in populated]
            row_band = next((color for color in colors if color in band_colors), None)
            if row_band is None:
                continue
            if row_band == ib.BG_WORKING:
                # Working highlights may intentionally stop before note columns.
                last = max(cell.column for cell in populated if _fill_rgb(cell) == row_band)
                check_cells = [ws.cell(row=row_idx, column=col) for col in range(2, last + 1)]
            else:
                first = min(cell.column for cell in populated if _fill_rgb(cell) == row_band)
                last = max(cell.column for cell in populated if _fill_rgb(cell) == row_band)
                check_cells = [ws.cell(row=row_idx, column=col) for col in range(first, last + 1)]
            for cell in check_cells:
                if _fill_rgb(cell) != row_band:
                    violations.append(f"{ws.title}!{cell.coordinate}: expected row band {row_band}")
    return violations


def _repeated_semantic_fill_rows(wb) -> list[str]:
    violations = []
    semantic_fill_colors = {
        ib.BG_TABLE_HEADER,
        ib.BG_TOTAL_BAND,
        ib.BG_HEADER_BAND,
        ib.BG_WORKING,
    }
    for ws in wb.worksheets:
        previous_color = None
        previous_row = None
        for row_idx in range(1, ws.max_row + 1):
            colors = [
                _fill_rgb(ws.cell(row=row_idx, column=col))
                for col in range(1, ws.max_column + 1)
            ]
            row_colors = [color for color in colors if color in semantic_fill_colors]
            row_color = max(set(row_colors), key=row_colors.count) if row_colors else None
            if row_color is not None and row_color == previous_color:
                violations.append(f"{ws.title}!{previous_row}:{row_idx}: repeated semantic fill {row_color}")
            previous_color = row_color
            previous_row = row_idx if row_color is not None else None
    return violations


def _unknown_fill_color_violations(wb) -> list[str]:
    allowed = {
        ib.BG_CANVAS,
        ib.BG_TABLE_HEADER,
        ib.BG_TOTAL_BAND,
        ib.BG_HEADER_BAND,
        ib.BG_WORKING,
    }
    violations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                color = _fill_rgb(cell)
                if color is not None and color not in allowed:
                    violations.append(f"{ws.title}!{cell.coordinate}: unexpected fill {color}")
    return violations


def _repeated_prominent_border_rows(wb) -> list[str]:
    violations = []
    prominent_styles = {"medium", "thick", "double"}
    for ws in wb.worksheets:
        previous_signature = None
        previous_row = None
        for row_idx in range(1, ws.max_row + 1):
            counts = {}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col)
                for side_name in ("top", "bottom"):
                    style = getattr(getattr(cell.border, side_name), "style", None)
                    if style in prominent_styles:
                        key = (side_name, style)
                        counts[key] = counts.get(key, 0) + 1
            signature = max(counts, key=counts.get) if counts and max(counts.values()) >= 2 else None
            if signature is not None and signature == previous_signature:
                violations.append(
                    f"{ws.title}!{previous_row}:{row_idx}: repeated prominent border {signature[0]}={signature[1]}"
                )
            previous_signature = signature
            previous_row = row_idx if signature is not None else None
    return violations


def _hierarchy_spacer_border_violations(wb) -> list[str]:
    violations = []
    for ws in wb.worksheets:
        hierarchy_cols = [
            idx
            for idx in range(1, ws.max_column + 1)
            if abs(float(ws.column_dimensions[get_column_letter(idx)].width or 0) - float(ib.COL_HIERARCHY_WIDTH)) < 0.001
        ]
        for row in range(1, ws.max_row + 1):
            for col in hierarchy_cols:
                cell = ws.cell(row=row, column=col)
                if any(
                    getattr(getattr(cell.border, side_name), "style", None)
                    for side_name in ("top", "bottom", "left", "right")
                ):
                    violations.append(f"{ws.title}!{cell.coordinate}: hierarchy spacer carries border")
    return violations


def _border_density_violations(wb) -> list[str]:
    violations = []
    for ws in wb.worksheets:
        bordered_rows = 0
        populated_rows = 0
        for row_idx in range(1, ws.max_row + 1):
            row_has_value = any(ws.cell(row=row_idx, column=col).value not in (None, "") for col in range(1, ws.max_column + 1))
            if not row_has_value:
                continue
            populated_rows += 1
            # Horizontal rules only: vertical rules are structural column
            # declarations (e.g. the hybrid monthly→annual boundary), and
            # semantic band rows (header/section fills) declare their rule
            # through the fill role — neither is per-row gridline decoration.
            row_cells = [ws.cell(row=row_idx, column=col) for col in range(1, ws.max_column + 1)]
            if any(
                _fill_rgb(cell) in {ib.BG_TABLE_HEADER, ib.BG_HEADER_BAND, ib.BG_TOTAL_BAND}
                for cell in row_cells
            ):
                continue
            has_border = any(
                getattr(getattr(cell.border, side_name), "style", None)
                for cell in row_cells
                for side_name in ("top", "bottom")
            )
            if has_border:
                bordered_rows += 1
        if populated_rows and bordered_rows / populated_rows > 0.45:
            violations.append(f"{ws.title}: bordered_rows={bordered_rows} populated_rows={populated_rows}")
    return violations


def _memo_note_border_violations(wb) -> list[str]:
    note_header_labels = {"source / driver", "source", "note", "notes", "memo", "interpretation"}
    violations = []
    for ws in wb.worksheets:
        note_cols = set()
        startup_note_col = getattr(ws, "_startup_note_col", None)
        if isinstance(startup_note_col, int):
            note_cols.add(startup_note_col)
        for row in ws.iter_rows():
            for cell in row:
                value = str(cell.value or "").strip().lower()
                if value in note_header_labels:
                    note_cols.add(cell.column)
        for col in note_cols:
            for cell in ws[get_column_letter(col)]:
                if cell.value in (None, ""):
                    continue
                value = str(cell.value or "").strip().lower()
                if value in note_header_labels:
                    continue
                if _fill_rgb(cell) in {ib.BG_TABLE_HEADER, ib.BG_HEADER_BAND}:
                    continue
                # Subtotal rows carry a block-wide underline applied by
                # apply_semantic_border_span (Phase 1 Task 1.3) — that
                # underline correctly passes through the source / note
                # column as part of the continuous rule. Detect a subtotal
                # row by its bold label cell.
                label_cell = ws.cell(cell.row, source_plan.LAYOUT.label_col)
                if label_cell.font is not None and label_cell.font.bold:
                    continue
                if any(
                    getattr(getattr(cell.border, side_name), "style", None)
                    for side_name in ("top", "bottom", "left", "right")
                ):
                    violations.append(f"{ws.title}!{cell.coordinate}: memo/source/note cell carries border")
    return violations


def _border_style_violations(wb) -> list[str]:
    allowed = {"thin", "medium", "dotted"}
    violations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                for side_name in ("top", "bottom", "left", "right"):
                    style = getattr(getattr(cell.border, side_name), "style", None)
                    if style is not None and style not in allowed:
                        violations.append(f"{ws.title}!{cell.coordinate}: {side_name}={style}")
    return violations


def _border_color_violations(wb) -> list[str]:
    violations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                for side_name in ("top", "bottom", "left", "right"):
                    side = getattr(cell.border, side_name)
                    if getattr(side, "style", None) is None:
                        continue
                    rgb = getattr(getattr(side, "color", None), "rgb", None)
                    if isinstance(rgb, str) and rgb[-6:].upper() == ib.IB_FORMULA:
                        continue
                    violations.append(f"{ws.title}!{cell.coordinate}: {side_name} border color={rgb}")
    return violations


def _font_design_violations(wb) -> list[str]:
    violations = []
    allowed_sizes = {float(size) for size in ib.FONT_SIZE_ALLOWED_CELLS}
    default_font = wb._fonts[0]
    if default_font.name != ib.FONT_FAMILY or float(default_font.sz) != float(ib.FONT_SIZE_BASE):
        violations.append(f"workbook default font={default_font.name} size={default_font.sz}")
    normal_font = wb._named_styles["Normal"].font
    if normal_font.name != ib.FONT_FAMILY or float(normal_font.sz) != float(ib.FONT_SIZE_BASE):
        violations.append(f"Normal style font={normal_font.name} size={normal_font.sz}")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if cell.font.name != ib.FONT_FAMILY:
                    violations.append(f"{ws.title}!{cell.coordinate}: font={cell.font.name}")
                if cell.font.sz is not None:
                    size = float(cell.font.sz)
                    if size not in allowed_sizes:
                        violations.append(f"{ws.title}!{cell.coordinate}: size={cell.font.sz}")
                    if size == float(ib.FONT_SIZE_TINY):
                        violations.append(f"{ws.title}!{cell.coordinate}: tiny cell font={cell.font.sz}")
    return violations


def _semantic_alignment_violations(wb) -> list[str]:
    violations = []
    header_rows = {4, 5, ib.HEADER_PERIOD_ROW}
    for ws in wb.worksheets:
        if not source_plan.uses_default_layout(ws):
            continue
        period_cols = set(build_model._v2_period_cols(ws))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                alignment = cell.alignment
                horizontal = alignment.horizontal
                indent = getattr(alignment, "indent", 0)
                if cell.row in {4, ib.HEADER_PERIOD_ROW} and cell.column in period_cols and horizontal != "center":
                    violations.append(f"{ws.title}!{cell.coordinate}: period header horizontal={horizontal}")
                elif cell.row == 5 and cell.column in period_cols and horizontal not in ("right", "center"):
                    violations.append(f"{ws.title}!{cell.coordinate}: months ruler horizontal={horizontal}")
                elif cell.row not in header_rows and cell.column == source_plan.LAYOUT.source_col:
                    if horizontal != "left" or not cell.font.italic or _font_rgb(cell) != ib.IB_COMMENT:
                        violations.append(f"{ws.title}!{cell.coordinate}: source alignment/font")
                elif cell.row not in header_rows and cell.column == source_plan.LAYOUT.unit_col:
                    if horizontal != "right" or _font_rgb(cell) != ib.IB_COMMENT:
                        violations.append(f"{ws.title}!{cell.coordinate}: unit alignment/font")
                elif cell.row not in header_rows and cell.column in (source_plan.LAYOUT.first_hierarchy_col, source_plan.LAYOUT.label_col):
                    if horizontal != "left" or indent:
                        violations.append(f"{ws.title}!{cell.coordinate}: label horizontal={horizontal} indent={indent}")
                elif cell.row not in header_rows and cell.column >= FIRST_VALUE_COL:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        if horizontal != "right":
                            violations.append(f"{ws.title}!{cell.coordinate}: value horizontal={horizontal}")
    return violations


def _money_unit_format_mismatches(wb) -> list[str]:
    expected = {
        "円": {ib.FMT_JP_YEN},
        "千円": {ib.FMT_JP_THOUSAND},
        "百万円": {ib.FMT_JP_MILLION},
        "十億円": {ib.FMT_JP_BILLION},
        "$": {ib.FMT_USD_DOLLAR},
        "$K": {ib.FMT_USD_THOUSAND},
        "$M": {ib.FMT_USD_MILLION},
    }
    mismatches = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            unit = ws.cell(row=row[0].row, column=UNIT_COL).value
            if unit not in expected:
                continue
            for cell in row[FIRST_VALUE_COL - 1:]:
                if cell.value is None:
                    continue
                if not (
                    isinstance(cell.value, (int, float))
                    or (isinstance(cell.value, str) and cell.value.startswith("="))
                ):
                    continue
                if cell.number_format not in expected[unit]:
                    mismatches.append(
                        f"{ws.title}!{cell.coordinate}: unit={unit} fmt={cell.number_format}"
                    )
    return mismatches


def _semantic_unit_format_mismatches(wb) -> list[str]:
    count_like = {source_plan.FMT_COUNT_V2, "0", "0.0", ib.FMT_INTEGER, ib.FMT_NUM}
    expected = {
        "%": {
            ib.FMT_JP_PERCENT,
            ib.FMT_PERCENT,
            ib.FMT_PERCENT_BPS,
            ib.FMT_PCT_0,
            ib.FMT_PCT_1,
            ib.FMT_PCT_2,
        },
        "x": {ib.FMT_MULTIPLE, ib.FMT_MULTIPLE_1, ib.FMT_MULTIPLE_2, ib.FMT_FACTOR},
        "FTE": count_like,
        "units": count_like,
        "customers": count_like,
        "count": count_like,
        "months": count_like | {ib.FMT_MONTHS_1DP},
        "days": count_like,
        "check": {source_plan.FMT_CHECK_V2},
    }
    mismatches = []
    for ws in wb.worksheets:
        if ws.max_column < FIRST_VALUE_COL:
            continue
        for row in ws.iter_rows():
            unit = ws.cell(row=row[0].row, column=UNIT_COL).value
            if unit not in expected:
                continue
            for cell in row[FIRST_VALUE_COL - 1:]:
                if cell.value is None:
                    continue
                if not (
                    isinstance(cell.value, (int, float))
                    or (isinstance(cell.value, str) and cell.value.startswith("="))
                ):
                    continue
                if cell.number_format not in expected[unit]:
                    mismatches.append(
                        f"{ws.title}!{cell.coordinate}: unit={unit} fmt={cell.number_format}"
                    )
    return mismatches


def _row_height_violations(wb) -> list[str]:
    allowed = {
        ib.ROW_HEIGHT_TIGHT,
        ib.ROW_HEIGHT_BASE,
        ib.ROW_HEIGHT_STANDARD,
        ib.ROW_HEIGHT_MEDIUM,
        ib.ROW_HEIGHT_RELAXED,
        ib.ROW_HEIGHT_HERO,
        ib.ROW_HEIGHT_COVER_TITLE,
    }
    violations = []
    for ws in wb.worksheets:
        for idx, dimension in ws.row_dimensions.items():
            if dimension.height is None:
                continue
            actual = float(dimension.height)
            if actual in allowed:
                continue
            wrapped_exception_heights = {
                ib.wrapped_text_row_height(ib.visible_text_line_count(cell.value))
                for cell in ws[idx]
                if isinstance(cell.value, str)
                and "\n" in cell.value
                and cell.alignment.wrap_text is True
            }
            if actual not in wrapped_exception_heights:
                violations.append(f"{ws.title}!{idx}: height={dimension.height}")
    return violations


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    value = getattr(fill.fgColor, "rgb", None)
    return value[-6:].upper() if isinstance(value, str) else None


def _row_for_label(wb, sheet_name: str, label: str) -> int:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return cell.row
    raise AssertionError(f"{sheet_name}: label not found: {label}")


def _unit_for_label(wb, sheet_name: str, label: str) -> object:
    return wb[sheet_name].cell(_row_for_label(wb, sheet_name, label), UNIT_COL).value


def _first_year_cell_for_label(wb, sheet_name: str, label: str):
    return wb[sheet_name].cell(_row_for_label(wb, sheet_name, label), FIRST_VALUE_COL)


def test_source_backed_plan_reaches_generic_kernel_shape() -> None:
    source_text = """# Asset deployment equity story

The company is an asset-heavy AI startup deploying specialized field devices
into operational sites. The model combines recurring lease fees, services, and
software. Unit pricing is 月額32万円. Phase 1 targets cumulative deployments of
3,000 units in three years and about 25,000 operating units at maturity, equal
to roughly ¥900億 ARR. Hardware cost starts around ¥8m per unit and declines with scale. The story
requires market sizing, competitive funding benchmarks, hiring, lease strategy,
scenario analysis, sensitivity analysis, and an investor-use funding plan.
Source: customer discovery memo, market sizing memo, lender discussion notes.
"""
    with tempfile.TemporaryDirectory() as tmp:
        source_md = Path(tmp) / "asset_deployment_story.md"
        out = Path(tmp) / "source_plan.xlsx"
        source_md.write_text(source_text, encoding="utf-8")

        source_plan.build_source_plan_workbook(source_md, out)
        wb = load_workbook(out, data_only=False)

        # Narrative route → the v2 12-sheet surface (BS conditional).
        assert set(wb.sheetnames) <= set(source_plan.SOURCE_PLAN_SHEETS_V2)
        for sheet_name in ["Guide", "Summary", "Assumptions", "Revenue Build",
                           "Cost Build", "People Plan", "P&L", "CF",
                           "Financing", "Cap Table", "Evidence"]:
            assert sheet_name in wb.sheetnames
        assert _defined_name_count(wb) == 0
        assert _merged_count(wb) == 0
        assert _formula_count(wb) >= 400
        years = kernel.forecast_years(kernel.extract_start_year(source_text))
        assert wb["Assumptions"].cell(6, FIRST_VALUE_COL).value == f"FY{years[0]}"
        assert wb["Assumptions"].cell(6, FIRST_VALUE_COL + len(years) - 1).value == f"FY{years[-1]}"
        # Stated facts reach the kernel: 月額32万円 pricing.
        price_row = _row_for_label(wb, "Assumptions", _row_label_startswith(wb, "Assumptions", "Monthly price"))
        assert wb["Assumptions"].cell(price_row, FIRST_VALUE_COL).value == 320_000
        assert _wrapped_cells(wb) == []
        assert _manual_line_break_violations(wb) == []
        assert _raw_money_scale_formula_violations(wb) == []
        assert _numbered_section_labels(wb) == []
        assert _leading_space_labels(wb) == []
        assert _styled_blank_cells(wb) == []
        assert _design_band_span_violations(wb) == []
        assert _repeated_semantic_fill_rows(wb) == []
        assert _unknown_fill_color_violations(wb) == []
        assert _repeated_prominent_border_rows(wb) == []
        assert _hierarchy_spacer_border_violations(wb) == []
        assert _font_design_violations(wb) == []
        assert _semantic_alignment_violations(wb) == []
        assert _border_density_violations(wb) == []
        assert _border_style_violations(wb) == []
        assert _border_color_violations(wb) == []
        assert _unit_for_label(wb, "Assumptions", "New primary units") == "units"
        assert _unit_for_label(wb, "Assumptions", "Total customers (ending)") == "customers"
        assert _unit_for_label(wb, "People Plan", "Total headcount") == "FTE"
        assert _unit_for_label(wb, "CF", "Runway months") == "months"
        assert _unit_for_label(wb, "Revenue Build", "Demand support coverage") == "x"
        assert _first_year_cell_for_label(wb, "Assumptions", "New primary units").value > 0
        assert _first_year_cell_for_label(wb, "Assumptions", _row_label_startswith(wb, "Assumptions", "CapEx / ")).value > 0
        # Evidence status is declared per assumption row in the notes column.
        note_col = 6 + len(years)
        statuses = {
            str(wb["Assumptions"].cell(row, note_col).value or "")
            for row in range(8, wb["Assumptions"].max_row + 1)
        }
        assert any("estimate" in status for status in statuses), (
            "Assumptions rows must carry evidence status in the notes column"
        )
        # The full strict audit (incl. the S5 promoted audits) holds.
        assert build_model.audit_workbook(wb) == []


def test_generated_workbook_has_sheet_specific_quality_markers() -> None:
    source_text = """# Asset deployment equity story

AI device deployment with 月額32万円 pricing, capex, financing, ownership,
valuation, market sizing, benchmark refresh, and DD needs.
Source: management memo, customer discovery memo.
"""
    # v2 sheet-quality markers are the audit contract itself: every sheet in
    # REQUIRED_SHEET_MARKERS must render its markers. Verify against real
    # builds across the mode surface (full + all conditional sheets).
    with tempfile.TemporaryDirectory() as tmp:
        source_md = Path(tmp) / "asset_deployment_story.md"
        out = Path(tmp) / "source_plan.xlsx"
        source_md.write_text(source_text, encoding="utf-8")
        source_plan.build_source_plan_workbook(source_md, out)
        covered: set[str] = set()

        def check(wb) -> None:
            for ws in wb.worksheets:
                markers = build_model.REQUIRED_SHEET_MARKERS.get(ws.title, [])
                sheet_text = " ".join(
                    str(cell.value)
                    for row in ws.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                missing = [marker for marker in markers if marker not in sheet_text]
                assert missing == [], f"{ws.title}: missing {missing}"
                covered.add(ws.title)

        check(load_workbook(out, data_only=False))
        for mode in ("ma_exit", "pricing", "unit_economics", "cap_table"):
            mode_out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, mode_out, mode=mode)
            check(load_workbook(mode_out, data_only=False))
        segments_out = Path(tmp) / "segments.xlsx"
        segments_input = Path(tmp) / "segments.yaml"
        segments_input.write_text("segments: [Enterprise, SMB]\n", encoding="utf-8")
        build_model.build_model(segments_input, segments_out, mode="full",
                                additional_sheets=["Segments"])
        check(load_workbook(segments_out, data_only=False))
        # Every sheet with a marker contract is exercised at least once.
        assert set(build_model.REQUIRED_SHEET_MARKERS) <= covered


def test_structured_yaml_currency_and_display_scale_are_generic() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "model.yaml"
        out = Path(tmp) / "usd_thousand.xlsx"
        input_path.write_text(
            "\n".join(
                [
                    "company: CurrencyCo",
                    "currency: USD",
                    "mechanics: marketplace with GMV and take rate",
                    "display_scale: thousand",
                    "periods: 3",
                    "new_units: [1, 2, 3]",
                    "gmv_yen: [1000000, 2000000, 3000000]",
                    "monthly_price_yen: [10000, 12000, 14000]",
                ]
            ),
            encoding="utf-8",
        )

        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)

        # currency: USD flips money rows onto the USD accounting ladder while
        # non-money units keep their own formats; cell values stay raw.
        gmv_label = "Gross merchandise value"
        gmv_row = _row_for_label(wb, "Assumptions", gmv_label)
        gmv_cell = wb["Assumptions"].cell(gmv_row, FIRST_VALUE_COL)
        assert gmv_cell.value == 1000000  # raw base-currency value
        assert _unit_for_label(wb, "Assumptions", gmv_label) in {"$", "$K", "$M"}
        assert gmv_cell.number_format in {ib.FMT_USD_DOLLAR, ib.FMT_USD_THOUSAND, ib.FMT_USD_MILLION}
        assert gmv_cell.number_format == {
            "$": ib.FMT_USD_DOLLAR, "$K": ib.FMT_USD_THOUSAND, "$M": ib.FMT_USD_MILLION,
        }[_unit_for_label(wb, "Assumptions", gmv_label)]
        assert _unit_for_label(wb, "Assumptions", "New primary units") == "units"
        assert _unit_for_label(wb, "People Plan", "Total headcount") == "FTE"
        assert "円" not in "\n".join(
            str(cell.value)
            for ws in wb.worksheets
            for row in ws.iter_rows(min_col=5, max_col=5)
            for cell in row
            if cell.value is not None
        )
        # The promoted unit/format audit is the canonical mismatch check.
        assert build_model._audit_unit_format_agreement(wb) == []


def test_money_units_are_raw_values_with_number_formats_not_scaled_values() -> None:
    source_text = """# Unit formatting plan

Asset-heavy startup with monthly price 月額32万円, target deployment of 3,000
units, and hardware cost around ¥8m per unit. Source: management memo.
"""
    tmp, wb = _sample_source_workbook(source_text)
    try:
        price_label = _row_label_startswith(wb, "Assumptions", "Monthly price")
        price_cell = _first_year_cell_for_label(wb, "Assumptions", price_label)
        revenue_cell = _first_year_cell_for_label(wb, "Revenue Build", "Recurring revenue")
        total_revenue_cell = _first_year_cell_for_label(wb, "Revenue Build", "Total revenue")
        cf_cash_cell = _first_year_cell_for_label(wb, "CF", "Ending cash")

        # Raw yen value; display scale lives ONLY in the number format, and
        # the unit label matches that format exactly (2-layer scale rule).
        assert price_cell.value == 320000
        assert price_cell.number_format == ib.FMT_JP_THOUSAND
        assert _unit_for_label(wb, "Assumptions", price_label) == "千円"
        assert isinstance(revenue_cell.value, str) and "/1000000" not in revenue_cell.value
        assert isinstance(total_revenue_cell.value, str) and "/1000000" not in total_revenue_cell.value
        # The sheet-wide money scale is one of the JP ladder scales and the
        # unit label matches the number format exactly.
        jp_fmt_by_unit = {unit: ib.FMT_JP_BY_SCALE[scale] for scale, unit in ib.JP_UNIT_BY_SCALE.items()}
        revenue_unit = _unit_for_label(wb, "Revenue Build", "Total revenue")
        assert total_revenue_cell.number_format == jp_fmt_by_unit[revenue_unit]
        cash_unit = _unit_for_label(wb, "CF", "Ending cash")
        assert cf_cash_cell.number_format == jp_fmt_by_unit[cash_unit]
        assert _raw_money_scale_formula_violations(wb) == []
        assert build_model._audit_unit_format_agreement(wb) == []
    finally:
        tmp.cleanup()


def test_money_and_non_money_units_keep_separate_format_rules() -> None:
    assert source_plan._model_value("=A1/1000000", "units") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000000", "customers") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000000", "FTE") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000000", "%") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000000", "x") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000000", "months") == "=A1/1000000"
    assert source_plan._model_value("=A1/1000", "units") == "=A1/1000"
    assert source_plan._model_value("=A1/1000000000", "customers") == "=A1/1000000000"
    assert source_plan._model_value("=A1/1000000", "JPY") == "=A1"
    assert source_plan._model_value("=A1/1000000", "USD") == "=A1"
    assert source_plan._model_value("=A1/1000", "JPY K") == "=A1"
    assert source_plan._model_value("=A1/1000", "USD K") == "=A1"
    assert source_plan._model_value("=A1/1000000000", "JPY B") == "=A1"
    assert source_plan._model_value("=A1/1000000000000", "JPY T") == "=A1"
    assert source_plan._model_value("=A1 * 1000000", "JPY M") == "=A1"
    assert source_plan._model_value(12, "units") == 12
    assert source_plan._model_value(12, "customers") == 12
    assert source_plan._model_value(12, "JPY M") == 12_000_000
    assert source_plan._model_value(12, "USD M") == 12_000_000
    assert source_plan._display_unit("USD") == "$"
    assert source_plan._display_unit("USD K") == "$K"
    assert source_plan._display_unit("USD M") == "$M"
    assert source_plan._display_unit("JPY", ib.FMT_JPY_BILLION) == "十億円"
    assert source_plan._display_unit("JPY", ib.FMT_JPY_TRILLION) == "兆円"
    assert source_plan._format_for_unit("units", ib.FMT_MONEY) == ib.FMT_INTEGER
    assert source_plan._format_for_unit("customers", ib.FMT_MONEY) == ib.FMT_INTEGER
    assert source_plan._format_for_unit("FTE", ib.FMT_MONEY) == ib.FMT_INTEGER
    assert source_plan._format_for_unit("months", ib.FMT_MONEY) == ib.FMT_NUM
    assert source_plan._format_for_unit("%", ib.FMT_MONEY) == ib.FMT_PERCENT
    assert source_plan._format_for_unit("x", ib.FMT_MONEY) == ib.FMT_MULTIPLE
    assert source_plan._format_for_unit("JPY K", ib.FMT_MONEY) == ib.FMT_JPY_THOUSAND
    assert source_plan._format_for_unit("JPY M", ib.FMT_MONEY) == ib.FMT_JPY_MILLION
    assert source_plan._format_for_unit("JPY B", ib.FMT_MONEY) == ib.FMT_JPY_BILLION
    assert source_plan._format_for_unit("JPY T", ib.FMT_MONEY) == ib.FMT_JPY_TRILLION
    assert source_plan._format_for_unit("USD", ib.FMT_MONEY) == ib.FMT_USD_MILLION
    assert source_plan._format_for_unit("USD K", ib.FMT_MONEY) == ib.FMT_USD_THOUSAND
    assert source_plan._format_for_unit("USD M", ib.FMT_MONEY) == ib.FMT_USD_MILLION


def test_skill_guidance_requires_reading_workbook_formatting_for_units() -> None:
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    design_text = (SKILL_DIR / "build" / "references" / "_ib_workbook_design_system.md").read_text(encoding="utf-8")
    review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    rubric_text = (SKILL_DIR / "build" / "references" / "_sheet_quality_rubric.md").read_text(encoding="utf-8")
    combined = "\n".join([invocation_text, design_text, review_text, rubric_text])

    for phrase in [
        "Store base-currency values and use Excel",
        "Do not infer units only from visible text",
        "Inspect units through formatting as well as values",
        "visible unit labels must match those formats",
        "This is a money-unit rule, not a blanket numeric rule",
        "units, customers, count, FTE, days, months, percentages, and multiples",
    ]:
        assert phrase in combined


def test_benchmark_guidance_covers_material_evidence_lenses() -> None:
    benchmark_text = (SKILL_DIR / "build" / "references" / "_benchmark_protocol.md").read_text(encoding="utf-8")
    invocation_text = (SKILL_DIR / "build" / "references" / "_skill_invocation_protocol.md").read_text(encoding="utf-8")
    review_text = (SKILL_DIR / "build" / "references" / "_self_review_protocol.md").read_text(encoding="utf-8")
    rubric_text = (SKILL_DIR / "build" / "references" / "_sheet_quality_rubric.md").read_text(encoding="utf-8")
    combined = " ".join("\n".join([benchmark_text, invocation_text, review_text, rubric_text]).split())

    for phrase in [
        "Material Evidence Lenses",
        "Labor / HR comps",
        "Venture equity / funding comps",
        "Venture debt / non-dilutive capacity",
        "Pricing / customer ROI comps",
        "Market / competitive benchmarks",
        "The gate is not the sheet name",
    ]:
        assert phrase in combined


def test_marketplace_source_does_not_emit_unrelated_asset_heavy_template() -> None:
    source_text = """# Marketplace plan

The company is a marketplace startup. GMV is ¥10B, take rate is 12%, buyer and
seller cohorts are tracked separately, and the plan needs pricing, working
capital, funding need, ownership, scenario, sensitivity, valuation, and IC memo.
Source: management memo.
"""
    with tempfile.TemporaryDirectory() as tmp:
        source_md = Path(tmp) / "marketplace.md"
        out = Path(tmp) / "marketplace_plan.xlsx"
        source_md.write_text(source_text, encoding="utf-8")

        source_plan.build_source_plan_workbook(source_md, out)
        wb = load_workbook(out, data_only=False)
        all_text = "\n".join(
            str(cell.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )

        assert "Marketplace / transaction" in str(wb["Guide"].cell(_row_for_label(wb, "Guide", "Mechanics"), 4).value)
        assert _unit_for_label(wb, "Assumptions", "Gross merchandise value") in {"百万円", "十億円"}
        assert _unit_for_label(wb, "Assumptions", "Take rate") == "%"
        f = FIRST_VALUE_LETTER
        gmv_row = _row_for_label(wb, "Revenue Build", "Gross merchandise value")
        take_row = _row_for_label(wb, "Revenue Build", "Take rate")
        assert _first_year_cell_for_label(wb, "Revenue Build", "Transaction revenue").value == (
            f"={f}{gmv_row}*{f}{take_row}"
        )
        assert [term for term in ["fleet", "hardware", "field devices"] if term.lower() in all_text.lower()] == []


def test_hardware_source_ignores_low_usd_competitor_price_noise() -> None:
    source_text = """# Asset-heavy plan

The company is a hardware startup. The source includes competitor context:
consumer device monthly rental $499, but the company has not finalized its own
yen price. The plan should still use a local hardware default rather than
treating the competitor USD number as the company's price.
"""
    with tempfile.TemporaryDirectory() as tmp:
        source_md = Path(tmp) / "hardware.md"
        source_md.write_text(source_text, encoding="utf-8")
        facts = source_plan.extract_source_facts(source_md)

        assert facts.mechanics == "Hardware / asset-heavy"
        assert facts.monthly_price_yen[0] == 300_000


def test_source_plan_starting_cash_is_hard_input_blue() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        cash_label = _row_label_startswith(wb, "Assumptions", "Beginning cash")
        beginning_cash_row = _row_for_label(wb, "Assumptions", cash_label)
        # Opening cash is a single blue model-start input (first period column
        # only) — never restated across columns or re-derived as a link.
        assert _font_rgb(wb["Assumptions"].cell(beginning_cash_row, FIRST_VALUE_COL)) == ib.IB_HARD_INPUT
        assert wb["Assumptions"].cell(beginning_cash_row, SECOND_VALUE_COL).value is None
        # And a stated multi-period input row stays blue in every column.
        equity_row = _row_for_label(wb, "Assumptions", "Equity financing")
        assert _font_rgb(wb["Assumptions"].cell(equity_row, FIRST_VALUE_COL)) == ib.IB_HARD_INPUT
        assert _font_rgb(wb["Assumptions"].cell(equity_row, SECOND_VALUE_COL)) == ib.IB_HARD_INPUT
    finally:
        tmp.cleanup()


def test_source_plan_bold_rows_preserve_ib_cell_colors() -> None:
    """Bold band rows keep the provenance color: intra-sheet formulas stay
    black, cross-sheet pulls stay green, inputs stay blue — bolding never
    repaints the provenance."""
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        total_revenue = _first_year_cell_for_label(wb, "Revenue Build", "Total revenue")
        assert total_revenue.font.bold and _font_rgb(total_revenue) == ib.IB_FORMULA
        ending_cash = _first_year_cell_for_label(wb, "CF", "Ending cash")
        assert ending_cash.font.bold and _font_rgb(ending_cash) == ib.IB_FORMULA
        net_income = _first_year_cell_for_label(wb, "CF", "Net income")
        assert _font_rgb(net_income) == ib.IB_LINK_INTRA  # ='P&L'!... pull
        equity = _first_year_cell_for_label(wb, "Assumptions", "Equity financing")
        assert _font_rgb(equity) == ib.IB_HARD_INPUT
    finally:
        tmp.cleanup()


def test_source_plan_uses_column_based_hierarchy_layout() -> None:
    """v2 header contract: title row 2 / purpose row 3 / caption + FY ruler
    row 4 / months ruler row 5 / period header row 6 / data from row 8;
    canonical column widths; freeze at F7."""
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        ws = wb["Assumptions"]
        assert ws.column_dimensions["A"].width == ib.COL_GUTTER_WIDTH_V2
        assert abs(ws.column_dimensions["B"].width - ib.COL_HIERARCHY_WIDTH) < ib.COL_WIDTH_TOL
        assert ws.column_dimensions["C"].width >= ib.COL_LABEL_WIDTH_V2
        assert ws.column_dimensions["D"].width >= ib.COL_SOURCE_WIDTH_INPUT
        assert ws.column_dimensions["E"].width >= ib.COL_UNIT_WIDTH_V2
        assert ws.column_dimensions[FIRST_VALUE_LETTER].width == ib.COL_PERIOD_WIDTH_V2
        assert ws.freeze_panes == f"{FIRST_VALUE_LETTER}7"
        # Title / purpose / caption rows.
        assert isinstance(ws.cell(2, 3).value, str) and "Assumptions" in ws.cell(2, 3).value
        assert isinstance(ws.cell(3, 3).value, str)
        assert isinstance(ws.cell(4, 3).value, str) and ws.cell(4, 3).value.startswith("(単位")
        # FY ruler (row 4), months ruler (row 5), period header (row 6).
        assert str(ws.cell(4, FIRST_VALUE_COL).value).startswith("FY")
        assert ws.cell(5, FIRST_VALUE_COL).value == 12
        assert str(ws.cell(6, FIRST_VALUE_COL).value).startswith("FY")
        assert [ws.cell(6, c).value for c in range(3, 6)] == ["Line item", "Driver", "Unit"]
        # Role alignment: labels left, units right, values right.
        row = _row_for_label(wb, "Assumptions", "New primary units")
        assert ws.cell(row, 3).alignment.horizontal == "left"
        assert ws.cell(row, 5).alignment.horizontal == "right"
        assert ws.cell(row, FIRST_VALUE_COL).alignment.horizontal == "right"
        # Non-period sheets (Guide) carry no freeze pane.
        assert wb["Guide"].freeze_panes is None
    finally:
        tmp.cleanup()


def test_source_plan_chart_axes_and_tabs_follow_currency_and_semantic_roles() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "usd.yaml"
        out = Path(tmp) / "usd.xlsx"
        input_path.write_text(
            "\n".join(
                [
                    "company: Generic USD Co",
                    "currency: USD",
                    "display_scale: thousand",
                    "grain: monthly",
                    "periods: 6",
                    "start_year: 2027",
                    "new_units: [10, 15, 20, 25, 30, 35]",
                    "monthly_price_yen: [20000, 20000, 21000, 21000, 22000, 22000]",
                    "equity_raise_yen: [1000000, 0, 0, 0, 0, 0]",
                ]
            ),
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)

        axis_titles = [
            _chart_axis_title_text(chart)
            for ws in wb.worksheets
            for chart in getattr(ws, "_charts", [])
        ]
        assert axis_titles, "v2 monthly/hybrid builds render charts"
        assert "円" not in " ".join(axis_titles)
        assert all(title in {"$", "$K", "$M"} for title in axis_titles), axis_titles
        # Tab colors follow the v2 per-sheet declaration.
        for sheet_name in wb.sheetnames:
            assert _tab_rgb(wb[sheet_name]) == source_plan.TAB_COLORS_V2[sheet_name], sheet_name


def test_segment_lens_handles_long_generic_segment_names_without_clipping() -> None:
    """Segments is a conditional sheet (YAML segments ≥ 2 + explicit add):
    long segment names autosize the label column above its floor so they are
    not clipped, and the consolidation bridge check is present."""
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "long_segments.yaml"
        out = Path(tmp) / "long_segments.xlsx"
        input_path.write_text(
            "\n".join(
                [
                    "company: Generic Segment Co",
                    "segments:",
                    "  - Enterprise expansion segment with partners",
                    "  - Embedded platform integration segment",
                    "  - International distributor segment",
                ]
            ),
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full", additional_sheets=["Segments"])
        wb = load_workbook(out, data_only=False)
        ws = wb["Segments"]
        width = ws.column_dimensions["C"].width or 0
        assert width > ib.COL_LABEL_WIDTH_V2, "long segment names must autosize the label column"
        clipped = [
            f"{cell.coordinate}: width={width} value={cell.value}"
            for cell in ws["C"]
            if isinstance(cell.value, str)
            and cell.row >= 6
            and ws.cell(cell.row, 4).value is not None
            and len(cell.value) > width * 1.15
        ]
        assert clipped == []
        assert "Consolidation bridge check" in _sheet_labels(wb, "Segments")


def test_scenario_toggle_drives_effective_scales_and_engine() -> None:
    """The scenario mechanism is live: a blue 1/2/3 toggle, Downside/Base/
    Upside case columns, INDEX-driven effective scales, and — verified by
    recalculation — flipping the toggle changes the computed engine values."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)
        ws = wb["Assumptions"]

        toggle_row = _row_for_label(wb, "Assumptions", "Scenario toggle")
        toggle_cell = ws.cell(toggle_row, FIRST_VALUE_COL)
        assert toggle_cell.value == 2  # Base
        assert _font_rgb(toggle_cell) == ib.IB_HARD_INPUT
        demand_row = _row_for_label(wb, "Assumptions", "Demand scale")
        down, base, up = (ws.cell(demand_row, c).value for c in (6, 7, 8))
        assert down < base <= up, "case columns must order Downside/Base/Upside"
        eff_row = _row_for_label(wb, "Assumptions", "Effective demand scale")
        assert ws.cell(eff_row, FIRST_VALUE_COL).value == (
            f"=INDEX($F${demand_row}:$H${demand_row},$F${toggle_row})"
        )

        soffice = shutil.which("soffice")
        if not soffice:
            return
        recalc_dir = Path(tmp) / "recalc"
        recalc_dir.mkdir()

        def recalc_revenue(path: Path) -> float:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "xlsx",
                 "--outdir", str(recalc_dir), str(path)],
                check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            )
            data = load_workbook(recalc_dir / path.name, data_only=True)
            row = _row_for_label(data, "Summary", "Revenue")
            summary = data["Summary"]
            period_cols = [
                c for c in range(6, summary.max_column + 1)
                if isinstance(summary.cell(5, c).value, (int, float))
            ]
            return float(summary.cell(row, period_cols[-1]).value or 0)

        base_revenue = recalc_revenue(out)
        downside = Path(tmp) / "downside.xlsx"
        wb_toggle = load_workbook(out, data_only=False)
        wb_toggle["Assumptions"].cell(toggle_row, FIRST_VALUE_COL).value = 1
        wb_toggle.save(downside)
        downside_revenue = recalc_revenue(downside)
        assert downside_revenue < base_revenue, (
            f"flipping the toggle to Downside must reduce revenue "
            f"({downside_revenue} !< {base_revenue})"
        )


def test_quarterly_grain_is_per_period_not_annual_expanded() -> None:
    """Quarterly facts are PER-PERIOD canonical (one value per quarter), like
    monthly — they must pass through the v2 expansion unchanged, not be routed
    through annual-canonical expansion (which dropped quarters 3+ and
    understated every money series ~4x). Verified against the kernel and a
    LibreOffice recalc of the workbook P&L."""
    import economic_kernel as ek

    with tempfile.TemporaryDirectory() as tmp:
        yml = Path(tmp) / "q.yaml"
        yml.write_text(
            "\n".join([
                "company: QCo",
                "grain: quarterly",
                "periods: 8",
                "customers: [10, 20, 30, 40, 50, 60, 70, 80]",
                "monthly_price_yen: 100000",
                "equity_raise_yen: [400000000, 0, 0, 0, 0, 0, 0, 0]",
            ]),
            encoding="utf-8",
        )
        out = Path(tmp) / "q.xlsx"
        build_model.build_model(yml, out, mode="full")

        facts = ek.derive_source_facts_from_mapping(
            {"company": "QCo", "grain": "quarterly", "periods": 8,
             "customers": [10, 20, 30, 40, 50, 60, 70, 80],
             "monthly_price_yen": 100000,
             "equity_raise_yen": [400000000, 0, 0, 0, 0, 0, 0, 0]})
        kernel_rev = [round(p["revenue"]) for p in ek.project_plan_free_cash_flow(facts)]
        assert len(kernel_rev) == 8 and kernel_rev[-1] > kernel_rev[0]

        soffice = shutil.which("soffice")
        if not soffice:
            return
        recalc_dir = Path(tmp) / "recalc"
        recalc_dir.mkdir()
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx",
             "--outdir", str(recalc_dir), str(out)],
            check=True, capture_output=True, timeout=180,
        )
        wb = load_workbook(recalc_dir / out.name, data_only=True)
        pl = wb["P&L"]
        rev_row = _row_for_label(wb, "P&L", "Total revenue")
        workbook_rev = [pl.cell(rev_row, c).value for c in range(FIRST_VALUE_COL, FIRST_VALUE_COL + 8)]
        # The workbook must reproduce all eight per-quarter figures, not a
        # ~4x-understated annual-expanded shape.
        assert [round(v) for v in workbook_rev] == kernel_rev, (
            f"quarterly workbook revenue {workbook_rev} != kernel {kernel_rev}"
        )


def test_hybrid_statement_scale_never_crushes_material_money_to_zero() -> None:
    """Digit-crush guard: on a hybrid JP model the statement sheets carry
    small working-capital / tax-timing balance rows (未払消費税等,
    預り金・未払社会保険料) that are ~2 orders below revenue. The sheet scale must
    drop (百万円 → 千円) so no MATERIAL nonzero money cell recalculates to a value
    that displays as "0"/"▲0". Verified against LibreOffice-computed values."""
    def _scale_div(fmt: str | None) -> int | None:
        if not fmt or not ("▲" in fmt or ",," in fmt or fmt == "#,##0"):
            return None
        if "#,##0,,," in fmt:
            return 1_000_000_000
        if "#,##0,," in fmt:
            return 1_000_000
        if "#,##0," in fmt:
            return 1_000
        if "#,##0" in fmt:
            return 1
        return None

    with tempfile.TemporaryDirectory() as tmp:
        yml = Path(tmp) / "hybrid.yaml"
        yml.write_text(
            "\n".join([
                "company: CrushCo",
                "grain: hybrid",
                "periods: 5",
                "customers: [120, 360, 900, 1900, 3400]",
                "monthly_price_yen: 120000",
                "target_gross_margin: 0.65",
                "equity_raise_yen: [400000000, 0, 0, 0, 0]",
                "statutory_welfare_rate: 0.15",
                "beginning_cash_yen: 120000000",
            ]),
            encoding="utf-8",
        )
        out = Path(tmp) / "hybrid.xlsx"
        build_model.build_model(yml, out, mode="full")

        soffice = shutil.which("soffice")
        if not soffice:
            return
        recalc_dir = Path(tmp) / "recalc"
        recalc_dir.mkdir()
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx",
             "--outdir", str(recalc_dir), str(out)],
            check=True, capture_output=True, timeout=180,
        )
        form = load_workbook(out, data_only=False)
        val = load_workbook(recalc_dir / out.name, data_only=True)
        # Immaterial tax-timing / rounding noise below the guard floor is
        # allowed to round to zero; every material figure must stay visible.
        floor = 10_000.0
        crushed: list[str] = []
        for ws in form.worksheets:
            vs = val[ws.title]
            for row in ws.iter_rows():
                for cell in row:
                    div = _scale_div(cell.number_format)
                    if div is None:
                        continue
                    cv = vs.cell(row=cell.row, column=cell.column).value
                    if (isinstance(cv, (int, float)) and not isinstance(cv, bool)
                            and abs(cv) >= floor and abs(cv) / div < 0.5):
                        crushed.append(f"{ws.title}!{cell.coordinate}={cv} (÷{div})")
        assert not crushed, (
            "material money cells render as 0 at the chosen sheet scale: "
            + "; ".join(crushed[:12])
        )


def test_source_plan_design_surface_uses_generic_blue_palette() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        allowed = {ib.BG_CANVAS, ib.BG_TABLE_HEADER, ib.BG_TOTAL_BAND, ib.BG_HEADER_BAND, ib.BG_WORKING}
        observed = {
            _fill_rgb(cell)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if _fill_rgb(cell) is not None
        }
        assert observed <= allowed
        assert ib.BG_TABLE_HEADER in observed
        assert _styled_blank_cells(wb) == []
        assert _repeated_semantic_fill_rows(wb) == []
        assert _repeated_prominent_border_rows(wb) == []
        assert _hierarchy_spacer_border_violations(wb) == []
        assert _font_design_violations(wb) == []
        assert _semantic_alignment_violations(wb) == []
        total_populated = sum(
            1
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )
        highlighted = sum(
            1
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None and _fill_rgb(cell) == ib.BG_WORKING
        )
        assert highlighted / total_populated < 0.08
    finally:
        tmp.cleanup()


def test_source_plan_has_no_excel_indent_or_clipped_role_columns() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        indent_violations = []
        clipped = []
        # v2 role columns: only the label column (C) promises full display —
        # the driver column (D) is a compact tag column that may read through
        # a blank unit cell.
        role_cols = (source_plan.LAYOUT.label_col,)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if getattr(cell.alignment, "indent", 0):
                        indent_violations.append(f"{ws.title}!{cell.coordinate}")
            for col in role_cols:
                width = ws.column_dimensions[source_plan.get_column_letter(col)].width or 0
                for row_idx in range(5, ws.max_row + 1):
                    cell = ws.cell(row_idx, col)
                    if not isinstance(cell.value, str):
                        continue
                    if cell.value.startswith("="):
                        continue
                    if ws.cell(row_idx, col + 1).value is None:
                        continue
                    if len(cell.value) > width * 1.15:
                        clipped.append(f"{ws.title}!{cell.coordinate}: width={width} value={cell.value}")
        assert indent_violations == []
        assert clipped == []
    finally:
        tmp.cleanup()


def test_source_plan_has_no_long_centered_prose_headers() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        centered = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and len(cell.value) > 18
                        and cell.alignment.horizontal == "center"
                    ):
                        centered.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
        assert centered == []
    finally:
        tmp.cleanup()


def test_source_plan_ib_design_rhythm_and_visibility() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        guide = wb["Guide"]
        assert guide["C7"].value == "Decision"
        assert guide["D7"].value
        for ws in wb.worksheets:
            assert ws.sheet_view.showGridLines is False
            assert ws.sheet_view.zoomScale == 90
            assert ws.page_setup.orientation == "landscape"
            assert ws.page_setup.fitToWidth == 1
            assert ws.print_title_rows in {"1:5", "$1:$5", "1:6", "$1:$6"}
            assert ws.print_title_cols in {"A:E", "$A:$E"}
            assert ws.print_area
            assert _column_width(ws, "A") == ib.COL_GUTTER_WIDTH_V2
            assert abs(_column_width(ws, "B") - ib.COL_HIERARCHY_WIDTH) < ib.COL_WIDTH_TOL
            # Period sheets carry the v2 label/unit widths and the F7 freeze.
            if build_model._v2_period_anchor_col(ws) is not None:
                assert _column_width(ws, "C") >= ib.COL_LABEL_WIDTH_V2
                assert _column_width(ws, "E") >= ib.COL_UNIT_WIDTH_V2
                assert ws.freeze_panes is not None
            else:
                assert ws.freeze_panes is None
    finally:
        tmp.cleanup()


def _print_area_bounds(ws) -> tuple[int, int, int, int]:
    area = str(ws.print_area).replace("'", "").replace("$", "")
    if "!" in area:
        area = area.split("!", 1)[1]
    return range_boundaries(area)


def test_all_generated_modes_pass_visual_design_invariants() -> None:
    for mode in build_model.VALID_MODES:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)

            assert _defined_name_count(wb) == 0
            assert _merged_count(wb) == 0
            assert _wrapped_cells(wb) == []
            assert _manual_line_break_violations(wb) == []
            assert _leading_space_labels(wb) == []
            assert _styled_blank_cells(wb) == []
            assert _design_band_span_violations(wb) == []
            assert _repeated_semantic_fill_rows(wb) == []
            assert _unknown_fill_color_violations(wb) == []
            assert _repeated_prominent_border_rows(wb) == []
            assert _hierarchy_spacer_border_violations(wb) == []
            assert _font_design_violations(wb) == []
            assert _semantic_alignment_violations(wb) == []
            assert _border_density_violations(wb) == []
            assert _memo_note_border_violations(wb) == []
            assert _border_style_violations(wb) == []
            assert _border_color_violations(wb) == []
            assert _money_unit_format_mismatches(wb) == []
            assert _semantic_unit_format_mismatches(wb) == []
            assert _row_height_violations(wb) == []
            for ws in wb.worksheets:
                # v2 freeze polarity: period-axis sheets freeze at row 7 of
                # the first period column, non-period sheets carry no freeze.
                anchor_col = build_model._v2_period_anchor_col(ws)
                if anchor_col is not None:
                    assert ws.freeze_panes == ws.cell(row=7, column=anchor_col).coordinate
                else:
                    assert ws.freeze_panes is None
                assert ws.sheet_view.showGridLines is False
                assert ws.print_area
                rendered_row, rendered_col = ib.rendered_bounds(ws)
                min_col, min_row, max_col, max_row = _print_area_bounds(ws)
                assert (min_row, min_col) == (1, 1)
                assert max_row >= rendered_row
                assert max_col >= rendered_col


def test_source_plan_print_canvas_includes_rendered_used_range() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        assert _styled_blank_cells(wb) == []
        for ws in wb.worksheets:
            rendered_row, rendered_col = ib.rendered_bounds(ws)
            min_col, min_row, max_col, max_row = _print_area_bounds(ws)
            assert (min_row, min_col) == (1, 1)
            assert max_row >= rendered_row
            assert max_col >= rendered_col
            assert ws.max_row <= rendered_row
            assert ws.max_column <= rendered_col
            stray_dimensions = [
                key
                for key in ws.column_dimensions
                if range_boundaries(f"{key}1:{key}1")[0] > rendered_col
            ]
            assert stray_dimensions == []

        for ws in wb.worksheets:
            ws.insert_rows(ws.max_row + 1)
            assert ws.cell(ws.max_row, 1).style_id == 0
    finally:
        tmp.cleanup()


def test_representative_workbook_pdf_render_smoke_when_available() -> None:
    soffice = shutil.which("soffice")
    pdftoppm = shutil.which("pdftoppm")
    if not soffice or not pdftoppm:
        return
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        source_md = tmp_path / "render_source.md"
        out = tmp_path / "render_source.xlsx"
        source_md.write_text(
            "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo.",
            encoding="utf-8",
        )
        source_plan.build_source_plan_workbook(source_md, out)
        subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmp, str(out)],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        pdf = tmp_path / "render_source.pdf"
        assert pdf.exists()
        assert pdf.stat().st_size > 50_000
        subprocess.run(
            [pdftoppm, "-png", "-r", "80", "-f", "1", "-l", "2", str(pdf), str(tmp_path / "page")],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        pages = sorted(tmp_path.glob("page-*.png"))
        assert len(pages) >= 2
        assert all(page.stat().st_size > 10_000 for page in pages)


def test_representative_workbook_recalculates_without_formula_errors_when_available() -> None:
    soffice = shutil.which("soffice")
    if not soffice:
        return
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        out = tmp_path / "formula_check.xlsx"
        recalc_dir = tmp_path / "recalc"
        recalc_dir.mkdir()
        build_model.build_model(None, out, mode="full")
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(recalc_dir), str(out)],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        recalculated = recalc_dir / out.name
        assert recalculated.exists()
        wb = load_workbook(recalculated, data_only=True)
        assert _formula_error_cells(wb) == []


def test_source_plan_custom_tables_keep_text_columns_readable() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA recurring software startup with ARR and subscription pricing. Source: management memo."
    )
    try:
        clipped = []
        # v2 register / matrix sheets whose tables carry text in data columns.
        for sheet_name in ["Summary", "Evidence", "Cap Table", "Guide"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_col=3, max_col=3):
                for cell in row:
                    if not isinstance(cell.value, str) or cell.value.startswith("="):
                        continue
                    width = ws.column_dimensions["C"].width or 0
                    if ws.cell(cell.row, cell.column + 1).value is not None and len(cell.value) > width * 1.15:
                        clipped.append(f"{ws.title}!{cell.coordinate}: width={width} value={cell.value}")
        assert clipped == []
    finally:
        tmp.cleanup()


def test_excluded_sheets_cannot_create_broken_references() -> None:
    for mode, excluded in (
        ("full", ["Assumptions"]),   # every engine sheet requires Assumptions
        ("full", ["CF"]),            # BS / Financing hard-require CF
        ("comps_only", ["Valuation & Exit"]),  # IC Memo readouts require it
    ):
        try:
            build_model.resolve_bundle(mode, excluded_sheets=excluded)
        except ValueError as exc:
            assert "broken workbook references" in str(exc)
        else:
            raise AssertionError(f"excluded_sheets accepted unsafe bundle: {mode}:{excluded}")


def test_ib_helpers_do_not_use_native_indent_for_hierarchy() -> None:
    wb = Workbook()
    ws = wb.active
    ib.apply_label(ws["B2"])
    ib.write_hierarchical_line_item(ws, 3, 2, "Child line")

    assert ws["B2"].alignment.indent == 0
    assert ws["D3"].alignment.indent == 0


def test_ib_helpers_encode_role_based_alignment_tokens() -> None:
    wb = Workbook()
    ws = wb.active

    ib.apply_label(ws["B2"])
    ib.apply_comment(ws["C2"])
    ib.apply_unit_label(ws["D2"])
    ib.apply_year_header(ws["E2"], "FY2027")
    ib.apply_hard_input(ws["F2"])

    assert ib.ALIGN_LABEL == "left"
    assert ib.ALIGN_SOURCE_NOTE == "left"
    assert ib.ALIGN_UNIT == "right"
    assert ib.ALIGN_VALUE == "right"
    assert ib.ALIGN_PERIOD_HEADER == "center"
    assert "labels, sources, notes, titles, memos, and interpretation text are left-aligned" in ib.ALIGNMENT_BEST_PRACTICE
    assert ws["B2"].alignment.horizontal == ib.ALIGN_LABEL
    assert ws["C2"].alignment.horizontal == ib.ALIGN_SOURCE_NOTE
    assert ws["D2"].alignment.horizontal == ib.ALIGN_UNIT
    assert ws["E2"].alignment.horizontal == ib.ALIGN_PERIOD_HEADER
    assert ws["F2"].alignment.horizontal == ib.ALIGN_VALUE


def test_ib_helpers_encode_role_based_font_size_tokens() -> None:
    wb = Workbook()
    ws = wb.active

    ib.apply_label(ws["B2"])
    ib.apply_comment(ws["C2"])
    ib.apply_unit_label(ws["D2"])
    ib.apply_year_header(ws["E2"], "FY2027")
    ib.apply_section_header(ws["B4"], "Section")
    ib.write_cover(ws, title="Cover", subtitle="Subtitle", confidential=True)

    assert ib.FONT_SIZE_ALLOWED_CELLS == (
        ib.FONT_SIZE_SMALL,
        ib.FONT_SIZE_BASE,
        ib.FONT_SIZE_LARGE,
        ib.FONT_SIZE_TITLE,
    )
    assert ib.FONT_SIZE_TINY not in ib.FONT_SIZE_ALLOWED_CELLS
    assert ib.FONT_SECTION == ib.FONT_SECTION_HEADER
    assert ib.FONT_COMMENT.color.rgb[-6:] == ib.IB_COMMENT
    assert ib.IB_COMMENT == "666666"
    assert "constrained font-size palette" in ib.FONT_SIZE_BEST_PRACTICE
    assert float(ws["B2"].font.sz) == float(ib.FONT_SIZE_BASE)
    assert float(ws["C2"].font.sz) == float(ib.FONT_SIZE_SMALL)
    assert float(ws["D2"].font.sz) == float(ib.FONT_SIZE_SMALL)
    assert float(ws["E2"].font.sz) == float(ib.FONT_SIZE_BASE)
    assert float(ws["B4"].font.sz) == float(ib.FONT_SIZE_LARGE)
    assert float(ws["B6"].font.sz) == float(ib.FONT_SIZE_TITLE)
    assert float(ws["B8"].font.sz) == float(ib.FONT_SIZE_TITLE)
    assert float(ws["F2"].font.sz) == float(ib.FONT_SIZE_SMALL)


def test_font_design_audit_rejects_tiny_fractional_and_presentation_sizes() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Tiny"
    ws["A1"].font = Font(name=ib.FONT_FAMILY, size=8)
    ws["A2"] = "Fractional"
    ws["A2"].font = Font(name=ib.FONT_FAMILY, size=10.5)
    ws["A3"] = "Presentation"
    ws["A3"].font = Font(name=ib.FONT_FAMILY, size=16)
    ws["A4"] = "Wrong family"
    ws["A4"].font = Font(name="Times New Roman", size=10)
    ib.set_workbook_default_font(wb)

    violations = _font_design_violations(wb)
    assert any("A1" in item and "size=8" in item for item in violations)
    assert any("A1" in item and "tiny cell font=8" in item for item in violations)
    assert any("A2" in item and "size=10.5" in item for item in violations)
    assert any("A3" in item and "size=16" in item for item in violations)
    assert any("A4" in item and "font=Times New Roman" in item for item in violations)


def test_every_sheet_pins_indent_block_to_google_sheets_20px() -> None:
    """Every indent / hierarchy column must render at 20px (xlsx width 2.14).

    The canonical layout (build/references/_layout_canonical.md) demands the
    indent / hierarchy spacer block starting at B render as 20px in Google
    Sheets. This sweep test fails the moment any builder reintroduces a wider
    B (or a wider C/D when LayoutSpec.hierarchy_cols > 1) for label-text use;
    the correct fix is to shift the first data column past the indent block,
    not to widen the indent column itself.
    """
    indent_cols = tuple(
        range(source_plan.LAYOUT.first_hierarchy_col, source_plan.LAYOUT.label_col)
    )
    for mode in build_model.VALID_MODES:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)
            violations = ib.audit_indent_column_widths(wb, indent_columns=indent_cols)
            assert violations == [], (
                f"mode={mode!r} indent-width violations:\n  "
                + "\n  ".join(violations)
            )


def test_set_column_widths_refuses_to_override_indent_column() -> None:
    """`_set_column_widths` must raise when callers try to widen the indent
    block — the regression class that once broke the 20px Google Sheets
    hierarchy rhythm. (The v2 builders pin B via apply_indent_column_widths;
    the guard protects any direct width caller.)"""
    wb = Workbook()
    ws = wb.active
    ib.apply_indent_column_widths(ws, [source_plan.LAYOUT.first_hierarchy_col])
    with pytest.raises(ValueError, match="indent column"):
        source_plan._set_column_widths(ws, {source_plan.LAYOUT.first_hierarchy_col: 30.0})


def test_generic_kernel_does_not_promote_domain_specific_mentions_to_sources() -> None:
    facts = kernel.derive_source_facts(
        "# Marketplace\nThe plan mentions a public company only as a comparison. Source: management memo."
    )

    assert facts.source_names == ["management memo"]


def test_benchmark_register_uses_evidence_status_not_fake_source_placeholders() -> None:
    tmp, wb = _sample_source_workbook(
        "# PLAN\nA generic startup plan. Source: management memo."
    )
    try:
        text = "\n".join(str(cell.value) for row in wb["Evidence"].iter_rows() for cell in row if cell.value is not None)
        # No fabricated placeholder sources (the legacy SRC-xx register is
        # retired); absent evidence is declared as an explicit DD gap.
        forbidden = ["provided source / owner", "external benchmark TBD", "TBD", "SRC-0"]
        assert [term for term in forbidden if term in text] == []
        assert "Benchmark register" in text
        assert "No external evidence provided — DD gap" in text
    finally:
        tmp.cleanup()


def test_scenario_formulas_are_not_built_with_fragile_substring_replacement() -> None:
    builder_text = (SCRIPTS_DIR / "source_plan_builder.py").read_text(encoding="utf-8")

    assert ".replace(\"C14\"" not in builder_text
    assert ".replace(\"C15\"" not in builder_text
    assert ".replace(\"C16\"" not in builder_text


def test_build_model_routes_source_markdown_to_source_plan() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        source_md = Path(tmp) / "story.md"
        out = Path(tmp) / "routed.xlsx"
        source_md.write_text(
            "Asset-heavy deployment company with 月額32万円 pricing and 25,000 operating units at maturity.",
            encoding="utf-8",
        )

        build_model.build_model(None, out, source_md=source_md)
        wb = load_workbook(out, data_only=False)

        # The narrative route builds the same v2 full bundle as the
        # structured route (annual grain default → BS materiality applies).
        assert set(wb.sheetnames) <= set(source_plan.SOURCE_PLAN_SHEETS_V2)
        assert {"Guide", "Summary", "Assumptions", "Revenue Build", "P&L", "CF"} <= set(wb.sheetnames)
        assert build_model.audit_workbook(wb) == []


def test_detect_table_block_returns_header_row_bounds_for_simple_period_grid() -> None:
    """Header row C (label) + F-L (period headers) all with BG_TABLE_HEADER.
    A total row below must resolve to the same (start, end) regardless of
    where its individual cells have values."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in [3, *range(6, 13)]:
        cell = ws.cell(5, col)
        cell.value = "Item" if col == 3 else f"Y{col - 5}"
        cell.fill = header_fill
    ws.cell(11, 3, "Total")
    for col in range(6, 13):
        ws.cell(11, col, "=SUM(F6:F10)")

    start, end = ib.detect_table_block(ws, 11)
    assert (start, end) == (3, 12), f"expected (3, 12), got ({start}, {end})"


def test_detect_table_block_excludes_annotation_column_outside_header_band() -> None:
    """An annotation column AFTER the period block has plain text but no
    BG_TABLE_HEADER fill — its column must not extend the block. This is
    the structural fix for the section-band-vs-total-row right-edge
    mismatch (帯=col3-13、合計=col3-12) called out in spec §2 C1."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in range(3, 13):
        cell = ws.cell(5, col, "Hdr")
        cell.fill = header_fill
    ws.cell(5, 13, "Notes")  # intentionally no header fill
    ws.cell(11, 13, "see footnote")
    for col in range(3, 13):
        ws.cell(11, col, 100)

    start, end = ib.detect_table_block(ws, 11)
    assert (start, end) == (3, 12), f"notes col 13 leaked into block: ({start}, {end})"


def test_detect_table_block_includes_empty_middle_cells_as_block_members() -> None:
    """For a header row where C and F-L carry BG_TABLE_HEADER but D, E do
    not (the canonical source / unit gap), the block (3, 12) still
    includes 4 and 5 — apply_semantic_border_span (Task 1.2) relies on
    this so the underline does not skip empty middle cells."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    ws.cell(5, 3, "Item").fill = header_fill
    for col in range(6, 13):
        ws.cell(5, col, f"Y{col - 5}").fill = header_fill
    ws.cell(11, 3, "Total")
    for col in range(6, 13):
        ws.cell(11, col, 100)

    start, end = ib.detect_table_block(ws, 11)
    assert (start, end) == (3, 12)
    assert start <= 4 <= end and start <= 5 <= end, (
        f"empty middle cols 4, 5 must be inside the block ({start}, {end})"
    )


def test_detect_table_block_does_not_fuse_side_by_side_tables() -> None:
    """A helper derived from the sfm-overhaul future-risk log.

    When two semantic tables share one worksheet row, the detector must choose
    the table used by the target row instead of min/maxing every header-filled
    cell into one giant block. The C/F-G gap is still a legitimate source/unit
    gap inside the left table, but the unrelated K-M table must stay outside.
    """
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in (3, 6, 7):
        ws.cell(5, col, "Left").fill = header_fill
    for col in range(11, 14):
        ws.cell(5, col, "Right").fill = header_fill

    ws.cell(11, 3, "Left total")
    ws.cell(11, 6, 100)
    ws.cell(11, 7, 200)

    start, end = ib.detect_table_block(ws, 11)
    assert (start, end) == (3, 7), (
        f"side-by-side header block leaked into target row span: ({start}, {end})"
    )


def test_apply_semantic_border_span_stays_inside_side_by_side_table() -> None:
    """Border helper must inherit the same side-by-side isolation contract."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in (3, 6, 7):
        ws.cell(5, col, "Left").fill = header_fill
    for col in range(11, 14):
        ws.cell(5, col, "Right").fill = header_fill
    ws.cell(11, 3, "Left total")
    ws.cell(11, 6, 100)
    ws.cell(11, 7, 200)

    ib.apply_semantic_border_span(ws, 11, top=ib.THIN_LINE)

    for col in range(3, 8):
        assert ws.cell(11, col).border.top.style == "thin", (
            f"left table col {col} missing top border"
        )
    for col in range(8, 14):
        assert ws.cell(11, col).border.top.style is None, (
            f"unrelated side-by-side col {col} received a leaked border"
        )


def test_detect_table_block_falls_back_to_row_content_when_no_header_fill_exists() -> None:
    """A synthetic worksheet with no BG_TABLE_HEADER cell anywhere must
    fall back to the target row's own leftmost / rightmost non-empty
    column instead of raising."""
    wb = Workbook()
    ws = wb.active
    ws.cell(10, 3, "Label")
    for col in range(6, 13):
        ws.cell(10, col, 50)

    start, end = ib.detect_table_block(ws, 10)
    assert (start, end) == (3, 12), f"fallback bounds ({start}, {end}) ≠ (3, 12)"


def test_detect_table_block_raises_for_empty_row_with_no_header() -> None:
    wb = Workbook()
    ws = wb.active
    raised = False
    try:
        ib.detect_table_block(ws, 7)
    except ValueError:
        raised = True
    assert raised, "ValueError expected for empty worksheet"


def test_default_build_label_and_source_columns_fit_their_content() -> None:
    """Real-build integration check for Task 1.4. After
    autosize_role_columns runs on every default-layout sheet, the label
    and source columns are wide enough to fully display the longest cell
    value they carry (CJK chars = 2 units), up to the role max. Pins the
    'no clipping' promise that content-driven widths give over the old
    fixed-54 assignment."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "model.xlsx"
        build_model.build_model(None, out)
        wb = load_workbook(out, data_only=False)
        label_letter = source_plan.get_column_letter(source_plan.LAYOUT.label_col)
        source_letter = source_plan.get_column_letter(source_plan.LAYOUT.source_col)
        for ws in wb.worksheets:
            if not source_plan.uses_default_layout(ws):
                continue
            for role_col_letter, role in [(label_letter, "label"), (source_letter, "source")]:
                col_dim = ws.column_dimensions.get(role_col_letter)
                col_width = col_dim.width if col_dim is not None else None
                if col_width is None:
                    continue
                role_max = ib.ROLE_WIDTH_BOUNDS[role][1]
                for cell in ws[role_col_letter]:
                    content_width = ib._display_width(cell.value)
                    if content_width == 0 or content_width > role_max:
                        continue
                    assert content_width <= col_width, (
                        f"{ws.title}!{cell.coordinate} ({role}): content "
                        f"display width {content_width} exceeds column "
                        f"width {col_width} — autosize_role_columns missed "
                        f"this cell."
                    )


def test_display_width_counts_cjk_chars_as_two() -> None:
    """CJK display width = 2 per char; ASCII = 1. Required so Japanese
    labels do not get clipped by latin-counted widths."""
    assert ib._display_width("") == 0
    assert ib._display_width(None) == 0
    assert ib._display_width("Sales") == 5
    assert ib._display_width("売上") == 4
    assert ib._display_width("売上 (JPY)") == 4 + 1 + 5  # 4 CJK + space + "(JPY)"
    # Mixed katakana + hiragana
    assert ib._display_width("カスタマー") == 10
    # Long ASCII
    assert ib._display_width("X" * 50) == 50


def test_autosize_role_columns_grows_label_col_above_floor_for_long_content() -> None:
    """Long labels push the label column above its role floor, up to max."""
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 3, "X" * 70)  # 70 ASCII chars, > 54 floor

    ib.autosize_role_columns(ws, {3: "label"})
    width = ws.column_dimensions["C"].width

    assert width > ib.COL_LABEL_WIDTH, (
        f"label col width {width} must exceed floor {ib.COL_LABEL_WIDTH} for 70-char label"
    )


def test_autosize_role_columns_keeps_short_labels_at_role_floor() -> None:
    """A column whose content fits the role floor stays at the floor —
    not artificially padded wider."""
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 3, "Sales")  # 5 chars << 54

    ib.autosize_role_columns(ws, {3: "label"})
    width = ws.column_dimensions["C"].width

    assert width == ib.COL_LABEL_WIDTH, (
        f"label col with short content should stay at floor {ib.COL_LABEL_WIDTH}, got {width}"
    )


def test_autosize_role_columns_treats_cjk_at_two_units_for_label() -> None:
    """28 CJK chars = 56 display units, which exceeds the 54 floor."""
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 3, "あ" * 28)

    ib.autosize_role_columns(ws, {3: "label"})
    width = ws.column_dimensions["C"].width

    assert width > ib.COL_LABEL_WIDTH, (
        f"28 CJK chars (56 display units) must push label width above {ib.COL_LABEL_WIDTH}, got {width}"
    )


def test_autosize_role_columns_clamps_to_role_max_for_label() -> None:
    """An absurdly long label is clamped to the role max, not unbounded."""
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 3, "X" * 500)

    ib.autosize_role_columns(ws, {3: "label"})
    width = ws.column_dimensions["C"].width
    role_max = ib.ROLE_WIDTH_BOUNDS["label"][1]

    assert width == role_max, (
        f"absurdly long label should be clamped to role max {role_max}, got {width}"
    )


def test_autosize_role_columns_handles_unit_and_period_roles() -> None:
    """The helper applies to any role declared in ROLE_WIDTH_BOUNDS — not
    just label."""
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 5, "FTE")   # unit
    ws.cell(2, 6, "FY26")  # period

    ib.autosize_role_columns(ws, {5: "unit", 6: "period"})
    unit_width = ws.column_dimensions["E"].width
    period_width = ws.column_dimensions["F"].width
    unit_floor, unit_max = ib.ROLE_WIDTH_BOUNDS["unit"]
    period_floor, period_max = ib.ROLE_WIDTH_BOUNDS["period"]

    assert unit_floor <= unit_width <= unit_max
    assert period_floor <= period_width <= period_max


def test_default_layout_role_widths_are_workbook_consistent() -> None:
    """Role widths should be resolved from the widest content across the
    workbook, then applied to every default-layout sheet using that role.

    This abstracts the sfm-overhaul design-research G5 rule: a reviewer should
    not see a 54-wide source column on one sheet and a 56-wide source column on
    another just because only one sheet carried the longest source note.
    """
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "model.xlsx"
        build_model.build_model(None, out, auto_live_comps=False)
        wb = load_workbook(out, data_only=False)
        widths_by_role: dict[str, set[float]] = {"label": set(), "source": set(), "note": set()}
        default_layout_sheets = []
        for ws in wb.worksheets:
            if not source_plan.uses_default_layout(ws):
                continue
            default_layout_sheets.append(ws.title)
            widths_by_role["label"].add(
                ws.column_dimensions[get_column_letter(source_plan.LAYOUT.label_col)].width
            )
            widths_by_role["source"].add(
                ws.column_dimensions[get_column_letter(source_plan.LAYOUT.source_col)].width
            )
            for col in range(source_plan._start_period_col() + 1, ws.max_column + 1):
                if ws.cell(row=6, column=col).value == "Notes":
                    widths_by_role["note"].add(ws.column_dimensions[get_column_letter(col)].width)
                    break

        assert default_layout_sheets, "real build should expose v2 default-layout sheets"

        for role, widths in widths_by_role.items():
            if not widths:
                continue
            assert len(widths) == 1, (
                f"{role} widths must be workbook-consistent; got {sorted(widths)}"
            )


def test_highlight_row_paints_continuous_border_across_source_and_unit() -> None:
    """Regression for the _highlight_row gap (Task 1.3b).

    Pre-Task-1.3b, _highlight_row applied top + bottom thin borders only to
    ``[LAYOUT.label_col, *period_cols]`` — D (source) and E (unit) were
    skipped, leaving the selected-output rule ragged. After the
    apply_semantic_border_span replacement, the rule is continuous across
    the detected block."""
    wb = Workbook()
    ws = wb.active
    ws.cell(13, source_plan.LAYOUT.label_col, "Selected output")
    ws.cell(13, source_plan.LAYOUT.source_col, "source note")
    ws.cell(13, source_plan.LAYOUT.unit_col, "JPY")
    ws.cell(13, FIRST_VALUE_COL, 1)
    source_plan._highlight_row(ws, 13, FIRST_VALUE_COL)

    for col in (
        source_plan.LAYOUT.label_col,
        source_plan.LAYOUT.source_col,
        source_plan.LAYOUT.unit_col,
        FIRST_VALUE_COL,
    ):
        cell = ws.cell(13, col)
        assert cell.border.top.style == "thin", (
            f"highlight row col {col} missing thin top border "
            f"(got {cell.border.top.style!r})"
        )
        assert cell.border.bottom.style == "thin", (
            f"highlight row col {col} missing thin bottom border "
            f"(got {cell.border.bottom.style!r})"
        )


def test_write_values_bold_path_paints_continuous_border_across_d_and_e() -> None:
    """Bold subtotal rows must carry a CONTINUOUS top rule across the whole
    table block — including the driver (D) and unit (E) columns even when
    those cells are empty. The v2 witness is the People Plan "Total
    headcount" subtotal written through `_v2_series_row(bold=True)`, whose
    border comes from `apply_semantic_border_span`."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "model.xlsx"
        build_model.build_model(None, out)
        wb = load_workbook(out, data_only=False)
        ws = wb["People Plan"]

        target_row = _row_for_label(wb, "People Plan", "Total headcount")
        block_start, block_end = ib.detect_table_block(ws, target_row)
        assert block_start <= 4 <= block_end
        assert block_start <= 5 <= block_end

        for col in (3, 4, 5, FIRST_VALUE_COL):
            style = ws.cell(target_row, col).border.top.style
            assert style == "thin", (
                f"People Plan!row {target_row} col {col}: missing thin top border "
                f"(got {style!r}) — the bold path must apply the underline "
                f"across the full table block including empty D/E cells."
            )


def test_apply_semantic_border_span_takes_no_start_or_end_col_args() -> None:
    """API contract: span is detected, never passed in. This pins the
    structural intent — callers cannot reintroduce hard-coded ranges like
    `accent_cols = [label_col, *period_cols]`."""
    import inspect
    sig = inspect.signature(ib.apply_semantic_border_span)
    positional = [
        p.name for p in sig.parameters.values()
        if p.kind in (p.POSITIONAL_OR_KEYWORD, p.POSITIONAL_ONLY)
    ]
    assert positional == ["ws", "row"], (
        f"only `ws, row` may be positional; got {positional}"
    )
    for forbidden in ("start_col", "end_col"):
        assert forbidden not in sig.parameters, (
            f"{forbidden!r} kwarg leaked back into the API — span must be detected"
        )


def test_apply_semantic_border_span_paints_empty_middle_cells_inside_block() -> None:
    """Header row has BG_TABLE_HEADER at C and F-L (D, E unfilled +
    unwritten — the canonical source / unit gap). A subtotal row below
    must get a top border on every cell C..L including the empty D, E.
    Structural fix for `_write_values`'s `if cell.value is None: continue`."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    ws.cell(5, 3, "Item").fill = header_fill
    for col in range(6, 13):
        ws.cell(5, col, f"Y{col - 5}").fill = header_fill
    ws.cell(11, 3, "Total")
    for col in range(6, 13):
        ws.cell(11, col, 100)

    ib.apply_semantic_border_span(ws, 11, top=ib.THIN_LINE)

    for col in range(3, 13):  # C..L
        style = ws.cell(11, col).border.top.style
        assert style == "thin", f"col {col} missing top thin border (got {style!r})"


def test_apply_semantic_border_span_excludes_indent_columns_via_border_start_col() -> None:
    """Real period header band fills B-L with BG_TABLE_HEADER (B is the
    indent gutter). The detector resolves the target row to the table body
    span, and the caller passes border_start_col=3 so B stays borderless —
    the indent column reads as a visual gutter, not a table member with a
    rule. The choice of
    border_start_col comes from the row's role / content, not from
    hard-coded column numbers."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    ws.cell(5, 2).fill = header_fill  # indent col B, no value
    ws.cell(5, 3, "Item").fill = header_fill
    for col in range(6, 13):
        ws.cell(5, col, f"Y{col - 5}").fill = header_fill
    ws.cell(11, 3, "Total")
    for col in range(6, 13):
        ws.cell(11, col, 100)

    # Sanity: detector uses target-row content, so the quiet indent gutter is
    # outside the table-body span even though the header fill reaches B.
    assert ib.detect_table_block(ws, 11) == (3, 12)

    ib.apply_semantic_border_span(ws, 11, top=ib.THIN_LINE, border_start_col=3)

    assert ws.cell(11, 2).border.top.style is None, (
        "indent col B must stay borderless even though it is in the detected block"
    )
    for col in range(3, 13):
        assert ws.cell(11, col).border.top.style == "thin", (
            f"col {col} missing top thin border"
        )


def test_apply_semantic_border_span_supports_top_and_bottom_in_one_call() -> None:
    """Grand-total rows carry thin top + medium bottom (accounting
    convention). The helper applies both in one call."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in [3, *range(6, 13)]:
        ws.cell(5, col, "H").fill = header_fill
    ws.cell(11, 3, "Grand total")
    for col in range(6, 13):
        ws.cell(11, col, 100)

    ib.apply_semantic_border_span(
        ws, 11, top=ib.THIN_LINE, bottom=ib.MEDIUM_LINE, border_start_col=3
    )

    for col in range(3, 13):
        cell = ws.cell(11, col)
        assert cell.border.top.style == "thin", f"col {col} missing top thin"
        assert cell.border.bottom.style == "medium", f"col {col} missing bottom medium"


def test_apply_semantic_border_span_is_a_noop_when_no_sides_supplied() -> None:
    """No-side call must touch nothing — defensive against accidental
    no-op call from a refactor."""
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    ws.cell(5, 3, "Item").fill = header_fill
    for col in range(6, 13):
        ws.cell(5, col, "Y").fill = header_fill
    ws.cell(11, 3, "Total")
    ws.cell(11, 6, 100)

    ib.apply_semantic_border_span(ws, 11)  # neither top nor bottom

    for col in range(3, 13):
        cell = ws.cell(11, col)
        assert cell.border.top.style is None
        assert cell.border.bottom.style is None


def test_detect_table_block_uses_nearest_header_above_not_topmost() -> None:
    """When a sheet contains two BG_TABLE_HEADER rows (an outer top-of-
    sheet table at row 5 plus a nested inner table at row 20), a target
    row at 25 belongs to the INNER table — the detector must pick the
    nearest header above, not the topmost one. Surfaced by the
    Phase 1 mid-checkpoint Codex review of detect_table_block."""
    wb = Workbook()
    ws = wb.active
    fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in range(2, 8):
        ws.cell(5, col, "Outer").fill = fill
    for col in range(3, 13):
        ws.cell(20, col, "Inner").fill = fill
    ws.cell(25, 3, "Row below inner header")
    for col in range(3, 13):
        ws.cell(25, col, 1)

    # Nearest header above row 25 is row 20 → (3, 12), not the outer (2, 7).
    assert ib.detect_table_block(ws, 25) == (3, 12)


def test_detect_table_block_honors_explicit_header_row_override() -> None:
    """Caller can pin the header row when nearest-above is the wrong
    answer (e.g. forcing a row to resolve against the outer table even
    when an inner header sits between them). Override beats auto-scan."""
    wb = Workbook()
    ws = wb.active
    fill = PatternFill("solid", fgColor=ib.BG_TABLE_HEADER)
    for col in range(2, 8):
        ws.cell(5, col, "Outer").fill = fill
    for col in range(3, 13):
        ws.cell(20, col, "Inner").fill = fill
    ws.cell(25, 3, "Row")
    for col in range(3, 13):
        ws.cell(25, col, 1)

    # Auto: nearest-above → row 20 → (3, 12).
    assert ib.detect_table_block(ws, 25) == (3, 12)
    # Override to row 5 → outer (2, 7).
    assert ib.detect_table_block(ws, 25, header_row=5) == (2, 7)


def test_hybrid_axis_renders_monthly_window_annual_tail_and_uniform_formulas() -> None:
    """Hybrid axis contract: 24 monthly columns labeled YYYY/MM anchored to
    the fiscal year in progress, an annual FY tail, an integer months ruler
    (1s then 12s), a declared medium boundary rule, R17 formula uniformity
    ACROSS the grain boundary, and charts bound to the monthly window only."""
    with tempfile.TemporaryDirectory() as tmp:
        input_path = Path(tmp) / "hybrid.yaml"
        out = Path(tmp) / "hybrid.xlsx"
        input_path.write_text(
            "\n".join([
                "company: HybridCo",
                "grain: hybrid",
                "periods: 5",
                "customers: [120, 360, 900, 1900, 3400]",
                "monthly_price_yen: 120000",
                "equity_raise_yen: [400000000, 0, 0, 0, 0]",
                "beginning_cash_yen: 120000000",
            ]),
            encoding="utf-8",
        )
        build_model.build_model(input_path, out, mode="full")
        wb = load_workbook(out, data_only=False)

        ws = wb["Revenue Build"]
        labels = [ws.cell(6, c).value for c in range(6, 6 + 27)]
        months = [ws.cell(5, c).value for c in range(6, 6 + 27)]
        assert all(re.fullmatch(r"20\d{2}/\d{2}", str(v)) for v in labels[:24]), labels[:24]
        assert all(re.fullmatch(r"FY20\d{2}", str(v)) for v in labels[24:]), labels[24:]
        assert months == [1] * 24 + [12] * 3
        # Unstated start year anchors to the fiscal year in progress: the
        # first fiscal year must not already be over at build time.
        from datetime import date
        first_fy_end_label = labels[11]  # 12th month = first FY end
        year, month = (int(part) for part in str(first_fy_end_label).split("/"))
        assert date(year, month, 1) >= date.today().replace(day=1), (
            f"hybrid window opens on a fully-past fiscal year ({labels[0]}..{labels[11]})"
        )
        # Boundary declaration: medium left rule on the first annual column.
        boundary_col = 6 + 24
        assert ws.cell(8, boundary_col).border.left.style == "medium"
        # R17 across the boundary (promoted audit is the canonical check).
        assert build_model._audit_row_formula_consistency(wb) == []
        # Charts bind to a single grain: the monthly window only.
        rb_chart = wb["Revenue Build"]._charts[0]
        cats_ref = str(rb_chart.series[0].cat.strRef.f if rb_chart.series[0].cat.strRef else rb_chart.series[0].cat.numRef.f)
        bounds = range_boundaries(cats_ref.split("!")[-1].replace("$", ""))
        assert bounds[2] <= 6 + 23, f"chart categories cross the grain boundary: {cats_ref}"


def test_tab_count_stays_within_cap_for_all_default_modes() -> None:
    """Default mode builds keep ≤ 12 tabs (explicit --additional-sheets are
    the only sanctioned way past the cap)."""
    with tempfile.TemporaryDirectory() as tmp:
        for mode in build_model.VALID_MODES:
            out = Path(tmp) / f"{mode}.xlsx"
            build_model.build_model(None, out, mode=mode)
            wb = load_workbook(out, data_only=False)
            assert len(wb.sheetnames) <= 12, f"{mode}: {wb.sheetnames}"


def test_no_all_zero_period_rows_without_declaration() -> None:
    """No generated period row is all-zero noise: a row whose every period
    cell is the literal 0 is either dropped by the mechanism profile or must
    not appear at all (0-as-information rows carry formulas or notes)."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)
        offenders = []
        for ws in wb.worksheets:
            cols = build_model._v2_period_cols(ws)
            if len(cols) < 2:
                continue
            for row in range(8, ws.max_row + 1):
                label = ws.cell(row, 3).value
                if not label:
                    continue
                values = [ws.cell(row, col).value for col in cols]
                numeric = [v for v in values if isinstance(v, (int, float))]
                if (
                    len(numeric) == len([v for v in values if v is not None])
                    and len(numeric) > 1
                    and all(v == 0 for v in numeric)
                ):
                    offenders.append(f"{ws.title}!{row} {label}")
        assert offenders == []


def test_summary_snapshot_staleness_check_guards_generator_snapshots() -> None:
    """The Summary scenario block is generator-snapshot data; a staleness
    check compares the Base snapshot to the live roll-up while the toggle
    sits on Base, and that check is consolidated into the master check."""
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)

        toggle_row = _row_for_label(wb, "Assumptions", "Scenario toggle")
        staleness_row = _row_for_label(wb, "Summary", "Snapshot staleness check")
        formula = wb["Summary"].cell(staleness_row, 6).value
        assert isinstance(formula, str)
        assert f"'Assumptions'!$F${toggle_row}=2" in formula, (
            "staleness check must be gated on the Base toggle position"
        )
        assert wb["Summary"].cell(staleness_row, 6).number_format == source_plan.FMT_CHECK_V2
        # Consolidated: a "Check — Summary" row aggregates it into the master check.
        assert "Check — Summary" in _sheet_labels(wb, "Summary")
        master_row = _row_for_label(wb, "Summary", "Master check")
        assert str(wb["Summary"].cell(master_row, 6).value).startswith("=SUM(")


def test_master_check_is_echoed_on_every_period_sheet() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "full.xlsx"
        build_model.build_model(None, out, mode="full")
        wb = load_workbook(out, data_only=False)
        master_row = _row_for_label(wb, "Summary", "Master check")
        for ws in wb.worksheets:
            if ws.title in ("Guide", "Summary") or build_model._v2_period_anchor_col(ws) is None:
                continue
            echo = ws.cell(2, build_model._v2_period_anchor_col(ws)).value
            assert isinstance(echo, str) and f"'Summary'!$F${master_row}" in echo, (
                f"{ws.title}: master-check echo missing from row 2"
            )
            assert "checks OK" in echo and "CHECK FAILED" in echo


if __name__ == "__main__":
    test_skill_exposes_clean_build_route_only()
    test_skill_uses_generic_economic_kernel_not_stage_matrix()
    test_prompt_guidance_resists_fixed_template_routing()
    test_sheet_quality_rubric_covers_every_generated_sheet()
    test_economic_kernel_is_separate_from_workbook_renderer()
    test_japanese_money_units_parse_at_correct_scale()
    test_seed_to_pre_ipo_horizon_is_not_truncated_to_two_years()
    test_generated_workbook_contains_interpretive_analysis_surfaces()
    test_mechanic_specific_kpi_and_scenario_axes_are_rendered()
    test_ambiguous_mechanics_use_generic_kpis_and_scenario_axes()
    test_cost_labeled_scenario_drivers_pressure_costs_not_revenue()
    test_unclassified_scenario_drivers_default_to_opex_not_revenue()
    test_financing_driver_downside_widens_funding_gap()
    test_cost_build_does_not_double_count_detailed_service_costs_in_variable_cogs()
    test_orchestrator_routes_through_generic_source_plan_builder()
    test_focused_modes_use_generic_kernel_after_bundle_filter()
    test_focused_modes_are_formula_complete_without_compact_placeholders()
    test_integrated_model_has_ib_decision_gates_not_just_readouts()
    test_comparable_evidence_loads_public_peers_by_default_and_overrides_multiples()
    test_default_live_comps_match_mechanic_labels_and_cli_overrides_yaml()
    test_no_live_comps_disables_default_public_fetch()
    test_private_and_transaction_comps_are_included_in_comparable_evidence()
    test_incomplete_provided_comps_are_registered_but_not_used_for_medians()
    test_sec_ticker_lookup_is_cached_and_errors_are_compact()
    test_failed_live_comps_do_not_pollute_company_specific_sources_or_mark_live_multiples()
    test_quality_gate_rejects_uniform_wrong_period_width()
    test_quality_gate_build_targets_disable_live_comps_by_default()
    test_full_model_uses_direct_formula_refs()
    test_intra_sheet_formula_cells_are_black()
    test_ic_memo_dependency_closure_matches_live_formula_readouts()
    test_cash_flow_runway_formula_floors_and_caps_runway()
    test_cross_sheet_formula_cells_are_green()
    test_pricing_bundle_is_formula_complete()
    test_cap_table_mode_uses_state_machine_not_full_workbook()
    test_all_modes_produce_expected_bundles()
    test_strict_audit_blocks_workbook_design_regressions()
    test_generated_modes_do_not_reference_removed_sheets()
    test_structured_yaml_controls_grain_periods_and_drivers()
    test_balance_sheet_has_opening_capital_counterpart()
    test_mfn_applied_only_for_changed_safe_terms()
    test_cap_table_rebuild_clears_prior_sheet_fills()
    test_cap_table_rebuild_clears_legacy_layout_state()
    test_cap_table_sheet_uses_canonical_design_surface()
    test_ib_helpers_reject_wrap_text_true()
    test_runtime_builders_do_not_use_wrap_or_merge_layout_shortcuts()
    test_skill_guidance_makes_no_wrap_rule_explicit()
    test_skill_guidance_requires_fix_and_rerun_iteration()
    test_self_improvement_protocol_triggers_from_logs_and_feedback()
    test_self_improvement_protocol_requires_regression_proof_and_privacy()
    test_self_improvement_reflection_validator_rejects_privacy_and_overfit_records()
    test_skill_guidance_uses_meaningful_sparse_fills_and_borders()
    test_skill_guidance_enforces_ib_text_positioning()
    test_skill_guidance_enforces_ib_font_size_discipline()
    test_xlsx_evals_load_full_design_reference_stack()
    test_semantic_fill_helper_uses_rectangular_span_including_blanks()
    test_source_backed_plan_reaches_generic_kernel_shape()
    test_generated_workbook_has_sheet_specific_quality_markers()
    test_structured_yaml_currency_and_display_scale_are_generic()
    test_money_units_are_raw_values_with_number_formats_not_scaled_values()
    test_money_and_non_money_units_keep_separate_format_rules()
    test_skill_guidance_requires_reading_workbook_formatting_for_units()
    test_benchmark_guidance_covers_material_evidence_lenses()
    test_marketplace_source_does_not_emit_unrelated_asset_heavy_template()
    test_hardware_source_ignores_low_usd_competitor_price_noise()
    test_source_plan_starting_cash_is_hard_input_blue()
    test_source_plan_bold_rows_preserve_ib_cell_colors()
    test_source_plan_uses_column_based_hierarchy_layout()
    test_source_plan_chart_axes_and_tabs_follow_currency_and_semantic_roles()
    test_segment_lens_handles_long_generic_segment_names_without_clipping()
    test_scenario_toggle_drives_effective_scales_and_engine()
    test_source_plan_design_surface_uses_generic_blue_palette()
    test_source_plan_has_no_excel_indent_or_clipped_role_columns()
    test_source_plan_has_no_long_centered_prose_headers()
    test_source_plan_ib_design_rhythm_and_visibility()
    test_all_generated_modes_pass_visual_design_invariants()
    test_source_plan_print_canvas_includes_rendered_used_range()
    test_representative_workbook_pdf_render_smoke_when_available()
    test_representative_workbook_recalculates_without_formula_errors_when_available()
    test_source_plan_custom_tables_keep_text_columns_readable()
    test_excluded_sheets_cannot_create_broken_references()
    test_ib_helpers_do_not_use_native_indent_for_hierarchy()
    test_ib_helpers_encode_role_based_alignment_tokens()
    test_ib_helpers_encode_role_based_font_size_tokens()
    test_font_design_audit_rejects_tiny_fractional_and_presentation_sizes()
    test_every_sheet_pins_indent_block_to_google_sheets_20px()
    test_set_column_widths_refuses_to_override_indent_column()
    test_generic_kernel_does_not_promote_domain_specific_mentions_to_sources()
    test_benchmark_register_uses_evidence_status_not_fake_source_placeholders()
    test_scenario_formulas_are_not_built_with_fragile_substring_replacement()
    test_build_model_routes_source_markdown_to_source_plan()
    test_detect_table_block_returns_header_row_bounds_for_simple_period_grid()
    test_detect_table_block_excludes_annotation_column_outside_header_band()
    test_detect_table_block_includes_empty_middle_cells_as_block_members()
    test_detect_table_block_does_not_fuse_side_by_side_tables()
    test_apply_semantic_border_span_stays_inside_side_by_side_table()
    test_detect_table_block_falls_back_to_row_content_when_no_header_fill_exists()
    test_detect_table_block_raises_for_empty_row_with_no_header()
    test_default_build_label_and_source_columns_fit_their_content()
    test_display_width_counts_cjk_chars_as_two()
    test_autosize_role_columns_grows_label_col_above_floor_for_long_content()
    test_autosize_role_columns_keeps_short_labels_at_role_floor()
    test_autosize_role_columns_treats_cjk_at_two_units_for_label()
    test_autosize_role_columns_clamps_to_role_max_for_label()
    test_autosize_role_columns_handles_unit_and_period_roles()
    test_default_layout_role_widths_are_workbook_consistent()
    test_highlight_row_paints_continuous_border_across_source_and_unit()
    test_write_values_bold_path_paints_continuous_border_across_d_and_e()
    test_apply_semantic_border_span_takes_no_start_or_end_col_args()
    test_apply_semantic_border_span_paints_empty_middle_cells_inside_block()
    test_apply_semantic_border_span_excludes_indent_columns_via_border_start_col()
    test_apply_semantic_border_span_supports_top_and_bottom_in_one_call()
    test_apply_semantic_border_span_is_a_noop_when_no_sides_supplied()
    test_detect_table_block_uses_nearest_header_above_not_topmost()
    test_detect_table_block_honors_explicit_header_row_override()
    test_hybrid_axis_renders_monthly_window_annual_tail_and_uniform_formulas()
    test_tab_count_stays_within_cap_for_all_default_modes()
    test_no_all_zero_period_rows_without_declaration()
    test_summary_snapshot_staleness_check_guards_generator_snapshots()
    test_master_check_is_echoed_on_every_period_sheet()
