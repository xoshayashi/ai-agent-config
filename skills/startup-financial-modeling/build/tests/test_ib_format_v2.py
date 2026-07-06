"""Tests for the S2 format layer (v2 additive) in ib_format.py.

Covers the 構造改革 slice S2 deliverables:
  - JP ▲ 系 number-format tokens (raw-yen values, comma display scaling)
  - deterministic scale pickers with 3x hysteresis
  - v2 header ruler / freeze-pane / master-check-echo helpers
  - period-column single-width audit
  - v2 column width tokens

Run directly:
    PYTHONDONTWRITEBYTECODE=1 python3 -m pytest \
        skills/startup-financial-modeling/build/tests/test_ib_format_v2.py -q
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SKILL_DIR = Path(__file__).resolve().parents[2]
RUNTIME_DIR = SKILL_DIR / "build" / "runtime"
sys.path.insert(0, str(RUNTIME_DIR))

import ib_format as ib  # noqa: E402


# ============================================================================
# 1. JP ▲ number-format tokens
# ============================================================================

# scale → (token, expected trailing display-scaling commas)
JP_MONEY_TOKENS = {
    "actual": (ib.FMT_JP_YEN, 0),
    "thousand": (ib.FMT_JP_THOUSAND, 1),
    "million": (ib.FMT_JP_MILLION, 2),
    "billion": (ib.FMT_JP_BILLION, 3),
}


@pytest.mark.parametrize("scale", sorted(JP_MONEY_TOKENS))
def test_jp_money_format_sections(scale):
    fmt, n_commas = JP_MONEY_TOKENS[scale]
    sections = fmt.split(";")
    assert len(sections) == 3, f"{scale}: expected positive;negative;zero sections"
    positive, negative, zero = sections

    # positive: thousands separator + trailing display-scaling commas, no ¥/()
    assert positive == "#,##0" + "," * n_commas
    # negative: red + ▲ prefix, same comma scaling
    assert negative == '[Red]"▲"#,##0' + "," * n_commas
    # zero renders as "-"
    assert zero == '"-"'
    # v2 contract: no currency symbol, no parenthesis wrapping
    assert "¥" not in fmt and "(" not in fmt


def test_jp_money_tokens_share_one_shape_except_comma_count():
    # 千円 → 百万円 → 十億円 は comma を 1 個ずつ足しただけの同型であること
    assert ib.FMT_JP_THOUSAND == ib.FMT_JP_YEN.replace("#,##0", "#,##0,")
    assert ib.FMT_JP_MILLION == ib.FMT_JP_YEN.replace("#,##0", "#,##0,,")
    assert ib.FMT_JP_BILLION == ib.FMT_JP_YEN.replace("#,##0", "#,##0,,,")


def test_jp_percent_format():
    sections = ib.FMT_JP_PERCENT.split(";")
    assert sections[0] == "0.0%"
    assert sections[1] == '[Red]"▲"0.0%'


def test_factor_and_months_tokens():
    assert ib.FMT_FACTOR == "0.0000"
    assert ib.FMT_MONTHS_1DP == "0.0"


def test_jp_unit_by_scale_ladder():
    # 億円は comma-scaling (10^3 刻み) で表現不可のためラダーから除外される
    assert ib.JP_SCALE_LADDER == ("actual", "thousand", "million", "billion")
    assert ib.JP_UNIT_BY_SCALE == {
        "actual": "円",
        "thousand": "千円",
        "million": "百万円",
        "billion": "十億円",
    }
    assert "hundred_million" not in ib.JP_UNIT_BY_SCALE
    assert set(ib.FMT_JP_BY_SCALE) == set(ib.JP_SCALE_LADDER)


def test_fmt_jp_for_scale():
    assert ib.fmt_jp_for_scale("actual") == ib.FMT_JP_YEN
    assert ib.fmt_jp_for_scale("thousand") == ib.FMT_JP_THOUSAND
    assert ib.fmt_jp_for_scale("million") == ib.FMT_JP_MILLION
    assert ib.fmt_jp_for_scale("billion") == ib.FMT_JP_BILLION
    with pytest.raises(ValueError):
        ib.fmt_jp_for_scale("hundred_million")
    with pytest.raises(ValueError):
        ib.fmt_jp_for_scale("trillion")


def test_usd_scale_reuses_existing_accounting_formats():
    assert ib.USD_SCALE_LADDER == ("actual", "thousand", "million")
    assert ib.FMT_USD_BY_SCALE["actual"] == ib.FMT_USD_DOLLAR
    assert ib.FMT_USD_BY_SCALE["thousand"] == ib.FMT_USD_THOUSAND
    assert ib.FMT_USD_BY_SCALE["million"] == ib.FMT_USD_MILLION


def test_v2_tokens_do_not_shadow_existing_jpy_tokens():
    # 追加のみの契約: 旧 FMT_JPY_* (¥ + 括弧) は変更されない
    assert ib.FMT_JPY_YEN == '¥#,##0_);[Red](¥#,##0);"-"_)'
    assert ib.FMT_JPY_MILLION == '¥#,##0,,_);[Red](¥#,##0,,);"-"_)'


# ============================================================================
# 2. Deterministic scale pickers (3x hysteresis)
# ============================================================================


@pytest.mark.parametrize(
    "max_abs,expected",
    [
        (0, "actual"),
        (2_999, "actual"),
        (3_000, "thousand"),
        (2_999_999, "thousand"),
        (3_000_000, "million"),
        (2_999_999_999, "million"),
        (3_000_000_000, "billion"),
        (5_000_000_000_000, "billion"),  # ladder cap: 兆でも十億円表示
    ],
)
def test_pick_scale_for_magnitude_jpy_boundaries(max_abs, expected):
    assert ib.pick_scale_for_magnitude(max_abs, currency="JPY") == expected


@pytest.mark.parametrize(
    "max_abs,expected",
    [
        (2_999, "actual"),
        (3_000, "thousand"),
        (2_999_999, "thousand"),
        (3_000_000, "million"),
        (3_000_000_000, "million"),  # USD ladder caps at $M
    ],
)
def test_pick_scale_for_magnitude_usd_ladder(max_abs, expected):
    assert ib.pick_scale_for_magnitude(max_abs, currency="USD") == expected


def test_pick_scale_negative_uses_absolute_value():
    assert ib.pick_scale_for_magnitude(-3_000_000) == "million"


def test_pick_scale_non_finite_is_actual():
    assert ib.pick_scale_for_magnitude(float("nan")) == "actual"
    assert ib.pick_scale_for_magnitude(float("inf")) == "actual"


def test_pick_scale_is_deterministic():
    for _ in range(5):
        assert ib.pick_scale_for_magnitude(4_200_000) == "million"


def test_pick_sheet_scale_uses_dominant_magnitude():
    # 支配的桁 = 集約行の最大絶対値 (負値・None・NaN 混在でも決定論)
    values = [120_000.0, None, -8_500_000.0, float("nan"), 42.0]
    assert ib.pick_sheet_scale(values) == "million"
    assert ib.pick_sheet_scale([]) == "actual"
    assert ib.pick_sheet_scale([2_999_999, 1_000]) == "thousand"


def test_pick_row_scale_matches_sheet_rule():
    values = [1_500, 2_999_999]
    assert ib.pick_row_scale(values) == ib.pick_sheet_scale(values) == "thousand"
    assert ib.pick_row_scale([3_000_000_000], currency="USD") == "million"


# ============================================================================
# 3. Header rulers / freeze pane / master-check echo
# ============================================================================

FY_LABELS = ["FY2026", "FY2026", "FY2027"]
MONTHS = [1, 1, 12]
PERIODS = ["2026-07", "2026-08", "FY2027"]
FIRST_PERIOD_COL = 6  # F


def _ruler_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ib.write_period_rulers(
        ws,
        first_period_col=FIRST_PERIOD_COL,
        fy_labels=FY_LABELS,
        months_in_period=MONTHS,
        period_labels=PERIODS,
        unit_caption="(単位: 百万円)",
    )
    return wb, ws


def test_write_period_rulers_lands_on_rows_4_5_6():
    _, ws = _ruler_sheet()
    for i, col in enumerate(range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + 3)):
        assert ws.cell(row=4, column=col).value == FY_LABELS[i]
        assert ws.cell(row=5, column=col).value == MONTHS[i]
        assert ws.cell(row=6, column=col).value == PERIODS[i]


def test_write_period_rulers_months_are_integers():
    _, ws = _ruler_sheet()
    for i, col in enumerate(range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + 3)):
        v = ws.cell(row=5, column=col).value
        assert isinstance(v, int), f"months ruler must store int, got {type(v)}"
        assert ws.cell(row=5, column=col).number_format == "0"
        assert ws.cell(row=5, column=col).alignment.horizontal == "right"


def test_write_period_rulers_caption_at_c4():
    _, ws = _ruler_sheet()
    assert ws["C4"].value == "(単位: 百万円)"
    assert ws["C4"].font.italic is True
    assert ws["C4"].font.size == ib.FONT_SIZE_SMALL
    assert ws["C4"].font.color.rgb.endswith(ib.IB_COMMENT)


def test_write_period_rulers_caption_omitted_when_none():
    wb = Workbook()
    ws = wb.active
    ib.write_period_rulers(ws, 6, ["FY2026"], [12], ["FY2026"], unit_caption=None)
    assert ws["C4"].value is None


def test_write_period_rulers_styles():
    _, ws = _ruler_sheet()
    fy = ws.cell(row=4, column=FIRST_PERIOD_COL)
    assert fy.alignment.horizontal == "center"
    assert fy.font.size == ib.FONT_SIZE_SMALL
    period = ws.cell(row=6, column=FIRST_PERIOD_COL)
    assert period.font.bold is True
    assert period.alignment.horizontal == "center"
    assert period.fill.fgColor.rgb.endswith(ib.BG_TABLE_HEADER)


def test_write_period_rulers_rejects_length_mismatch():
    wb = Workbook()
    ws = wb.active
    with pytest.raises(ValueError):
        ib.write_period_rulers(ws, 6, ["FY2026"], [12, 12], ["FY2026"], None)


def test_apply_freeze_pane_anchor():
    wb = Workbook()
    ws = wb.active
    ib.apply_freeze_pane(ws, first_period_col=6)
    assert ws.freeze_panes == "F7"
    ib.apply_freeze_pane(ws, first_period_col=5, header_rows=6)
    assert ws.freeze_panes == "E7"


def test_header_row_constants():
    assert ib.HEADER_FY_RULER_ROW == 4
    assert ib.HEADER_MONTHS_RULER_ROW == 5
    assert ib.HEADER_PERIOD_ROW == 6
    assert ib.HEADER_ROWS_V2 == 6
    assert ib.DATA_START_ROW_V2 == 8


def test_write_master_check_echo():
    wb = Workbook()
    ws = wb.active
    formula = '=IF(Summary!$F$40=0,"checks OK","CHECK FAILED")'
    ib.write_master_check_echo(ws, "F2", formula)
    assert ws["F2"].value == formula
    assert ws["F2"].alignment.horizontal == "right"
    assert ws["F2"].font.size == ib.FONT_SIZE_SMALL
    assert ws["F2"].font.color.rgb.endswith(ib.IB_COMMENT)


# ============================================================================
# 4. Period-column single-width audit
# ============================================================================


def _period_axis_sheet(wb, title, n_periods=3, width=None):
    ws = wb.create_sheet(title)
    ib.write_period_rulers(
        ws,
        first_period_col=FIRST_PERIOD_COL,
        fy_labels=["FY2026"] * n_periods,
        months_in_period=[12] * n_periods,
        period_labels=[f"FY{2026 + i}" for i in range(n_periods)],
        unit_caption="(単位: 百万円)",
    )
    ib.apply_freeze_pane(ws, FIRST_PERIOD_COL)
    if width is not None:
        for c in range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + n_periods):
            ws.column_dimensions[get_column_letter(c)].width = width
    return ws


def test_audit_period_column_widths_clean():
    wb = Workbook()
    wb.remove(wb.active)
    _period_axis_sheet(wb, "P&L", width=ib.COL_PERIOD_WIDTH_V2)
    _period_axis_sheet(wb, "CF", width=ib.COL_PERIOD_WIDTH_V2)
    assert ib.audit_period_column_widths(wb, ib.COL_PERIOD_WIDTH_V2) == []


def test_audit_period_column_widths_catches_mismatched_sheet():
    wb = Workbook()
    wb.remove(wb.active)
    _period_axis_sheet(wb, "P&L", width=ib.COL_PERIOD_WIDTH_V2)
    _period_axis_sheet(wb, "CF", width=16.0)  # 旧幅のまま = 違反
    violations = ib.audit_period_column_widths(wb, ib.COL_PERIOD_WIDTH_V2)
    assert violations, "mismatched sheet must be reported"
    assert all("'CF'" in v for v in violations)
    assert len(violations) == 3  # 3 期間列すべて


def test_audit_period_column_widths_reports_unset_width():
    wb = Workbook()
    wb.remove(wb.active)
    _period_axis_sheet(wb, "P&L", width=None)
    violations = ib.audit_period_column_widths(wb, ib.COL_PERIOD_WIDTH_V2)
    assert len(violations) == 3
    assert all("width not set" in v for v in violations)


def test_audit_period_column_widths_skips_non_period_sheets():
    wb = Workbook()
    ws = wb.active
    ws.title = "Guide"  # フリーズ無し・ルーラー無し → 監査対象外
    ws["C2"] = "Guide"
    assert ib.audit_period_column_widths(wb, ib.COL_PERIOD_WIDTH_V2) == []


# ============================================================================
# 5. v2 width tokens
# ============================================================================


def test_v2_width_tokens():
    assert ib.COL_GUTTER_WIDTH_V2 == 1.5
    assert ib.COL_LABEL_WIDTH_V2 == 38.0
    assert ib.COL_SOURCE_WIDTH_INPUT == 30.0
    assert ib.COL_DRIVER_TAG_WIDTH == 10.0
    assert ib.COL_UNIT_WIDTH_V2 == 9.0
    assert ib.COL_PERIOD_WIDTH_V2 == 11.5
    assert ib.COL_NOTE_WIDTH_V2 == 64.0


def test_v2_width_tokens_leave_old_tokens_untouched():
    assert ib.COL_MARGIN_WIDTH == 3.0
    assert ib.COL_LABEL_WIDTH == 54.0
    assert ib.COL_SOURCE_WIDTH == 54.0
    assert ib.COL_UNIT_WIDTH == 14.0
    assert ib.COL_PERIOD_WIDTH == 16.0
    assert ib.COL_HIERARCHY_WIDTH == 2.14


def test_v2_names_are_exported():
    for name in (
        "FMT_JP_YEN", "FMT_JP_THOUSAND", "FMT_JP_MILLION", "FMT_JP_BILLION",
        "FMT_JP_PERCENT", "FMT_FACTOR", "FMT_MONTHS_1DP",
        "JP_UNIT_BY_SCALE", "fmt_jp_for_scale",
        "pick_scale_for_magnitude", "pick_sheet_scale", "pick_row_scale",
        "write_period_rulers", "apply_freeze_pane", "write_master_check_echo",
        "audit_period_column_widths",
        "COL_GUTTER_WIDTH_V2", "COL_PERIOD_WIDTH_V2",
    ):
        assert name in ib.__all__, f"{name} missing from ib_format.__all__"


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))
