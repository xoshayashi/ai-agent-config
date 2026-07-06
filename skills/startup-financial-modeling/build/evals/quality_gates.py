#!/usr/bin/env python3
"""Quality gates — machine-scored rubric points for generated workbooks.

Scores the machine half (65 pts) of the workbook-quality rubric:

  G-A strict-audit clean on all target builds .......... 10
  G-B pytest suite green ............................... 10
  G-C LibreOffice recalc shows zero error cells ........ 8
  G-D row-formula consistency across period columns .... 6
  G-E no hardcoded constants outside the whitelist ..... 6
  G-F unit label <-> number-format agreement ........... 6
  G-G uniform period-column width + freeze anchors ..... 6
  G-H bundle closure (no refs to missing sheets) ....... 4
  G-I tab-count cap + driver-family coverage ........... 4
  G-J self-improvement reflection validator ............ 5

Usage:
  python3 quality_gates.py                    # offline deterministic gates
  python3 quality_gates.py --skip-pytest      # faster loop while iterating
  python3 quality_gates.py --allow-live-comps # opt in to public comp refresh
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = SKILL_ROOT.parents[1]
BUILD_CLI = SKILL_ROOT / "scripts" / "build_model.py"

sys.path.insert(0, str(SKILL_ROOT / "build" / "runtime"))

import self_improvement

NUMERIC_WHITELIST = {"0", "1", "-1", "2", "3", "12", "24", "100", "365", "1000"}
PERIOD_WIDTH = 11.5
HEADER_MONTHS_RULER_ROW = 5
FREEZE_ROW = 7

TARGETS: list[dict] = [
    {"id": "G1_full_hybrid_jp", "mode": "full", "yaml": {
        "company": "GateCo", "grain": "hybrid", "periods": 5,
        "customers": [120, 360, 900, 1900, 3400],
        "monthly_price_yen": 120000, "target_gross_margin": 0.65,
        "equity_raise_yen": [400000000, 0, 0, 0, 0],
        "post_money_yen": 2000000000, "statutory_welfare_rate": 0.15,
        "beginning_cash_yen": 120000000,
    }, "recalc": True, "r17": True, "r18": True, "coverage": True},
    {"id": "G2_full_annual_usd", "mode": "full", "yaml": {
        "company": "GateCo US", "grain": "annual", "currency": "USD",
        "customers": [40, 140, 420, 900, 1700], "monthly_price_yen": 250000,
        "equity_raise_yen": [500000000, 0, 0, 0, 0],
    }, "recalc": True},
    {"id": "G3_full_plus_valuation", "mode": "full",
     "extra_args": ["--additional-sheets", "Valuation & Exit"], "yaml": {
        "company": "GateCo VX", "grain": "hybrid",
        "customers": [100, 300, 800, 1600, 2800], "monthly_price_yen": 100000,
        "equity_raise_yen": [300000000, 0, 0, 0, 0],
    }},
    {"id": "G4_burn_runway_monthly", "mode": "burn_runway", "yaml": {
        "company": "GateCo BR", "grain": "monthly", "periods": 24,
        "customers": [10, 14, 19, 25, 32, 40, 49, 59, 70, 82, 95, 109,
                      124, 140, 157, 175, 194, 214, 235, 257, 280, 304, 329, 355],
        "monthly_price_yen": 200000,
        "equity_raise_yen": [250000000] + [0] * 23,
    }, "r17": True, "closure": True, "live_input_check": True},
    {"id": "G5_cap_table", "mode": "cap_table", "yaml": {"company": "GateCo CT"}},
    {"id": "G6_ma_exit", "mode": "ma_exit", "yaml": {
        "company": "GateCo MA",
        "customers": [200, 500, 1100, 2000, 3200], "monthly_price_yen": 150000,
        "equity_raise_yen": [600000000, 0, 0, 0, 0],
    }, "closure": True},
    {"id": "G7_narrative", "mode": "full", "source_md": (
        "# NarrativeCo\n\nSaaS向けの業務システムを提供するスタートアップ。"
        "月額12万円のサブスクリプションで、FY2030までに顧客3,000社を目指す。"
        "シリーズAで4億円の調達を計画。粗利率65%目標。\n"
    ), "recalc": True},
]


@dataclass
class GateResult:
    gate: str
    points: float
    max_points: float
    evidence: list[str] = field(default_factory=list)


def _run(cmd: list[str], timeout: int = 900) -> tuple[int, str]:
    proc = subprocess.run(
        cmd, cwd=str(REPO_ROOT), capture_output=True, text=True, timeout=timeout,
        env={"PYTHONDONTWRITEBYTECODE": "1", "PATH": "/usr/bin:/bin:/usr/local/bin:/opt/homebrew/bin"},
    )
    return proc.returncode, (proc.stdout + proc.stderr)


def build_targets(
    workdir: Path,
    strict: bool = True,
    *,
    allow_live_comps: bool = False,
) -> tuple[dict[str, Path], list[str]]:
    """Build every target; return output paths and strict-audit failures."""
    import yaml

    outputs: dict[str, Path] = {}
    failures: list[str] = []
    for target in TARGETS:
        out = workdir / f"{target['id']}.xlsx"
        cmd = [sys.executable, str(BUILD_CLI), "--mode", target["mode"], "--output", str(out)]
        if strict:
            cmd.append("--strict-audit")
        if not allow_live_comps:
            cmd.append("--no-live-comps")
        cmd += target.get("extra_args", [])
        if "yaml" in target:
            yml = workdir / f"{target['id']}.yaml"
            yml.write_text(yaml.safe_dump(target["yaml"], allow_unicode=True))
            cmd += ["--input", str(yml)]
        elif "source_md" in target:
            src = workdir / f"{target['id']}.md"
            src.write_text(target["source_md"])
            cmd += ["--source-md", str(src)]
        code, log = _run(cmd)
        if code != 0 or not out.exists():
            failures.append(f"{target['id']}: exit={code} :: {log.strip()[-400:]}")
        else:
            outputs[target["id"]] = out
    return outputs, failures


def _period_axis_sheets(wb):
    for ws in wb.worksheets:
        anchor = ws.freeze_panes
        if anchor and anchor[0] not in ("A",):
            yield ws


def _period_cols(ws) -> list[int]:
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string

    anchor = ws.freeze_panes
    if not anchor:
        return []
    col_letter, _row = coordinate_from_string(anchor)
    first = column_index_from_string(col_letter)
    cols = []
    for col in range(first, ws.max_column + 1):
        # Period columns carry the months ruler (row 5, numeric). The trailing
        # notes column has a row-6 header but no ruler — it is not a period col.
        months = ws.cell(row=HEADER_MONTHS_RULER_ROW, column=col).value
        if isinstance(months, (int, float)) and not isinstance(months, bool):
            cols.append(col)
    return cols


def _norm_formula(formula: str, col_idx: int) -> str:
    """Normalize a formula to a column-relative form for R17 comparison."""
    from openpyxl.utils import column_index_from_string, get_column_letter

    def repl(match: re.Match) -> str:
        dollar_col, letters, dollar_row, digits = match.groups()
        if dollar_col == "$":
            return match.group(0)
        offset = column_index_from_string(letters) - col_idx
        return f"C[{offset}]{dollar_row}{digits}"

    return re.sub(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", repl, formula)


def gate_r17(paths: list[Path]) -> GateResult:
    import openpyxl

    violations: list[str] = []
    for path in paths:
        wb = openpyxl.load_workbook(path)
        for ws in _period_axis_sheets(wb):
            cols = _period_cols(ws)
            if len(cols) < 2:
                continue
            for row in range(8, ws.max_row + 1):
                forms = {}
                for col in cols:
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        forms[col] = _norm_formula(cell.value, col)
                if len(forms) >= 2 and len(set(forms.values())) > 1:
                    violations.append(f"{path.name}:{ws.title}!row{row}")
    pts = 6.0 if not violations else 0.0
    return GateResult("G-D row-formula consistency", pts, 6.0, violations[:12])


def gate_r18(paths: list[Path]) -> GateResult:
    import openpyxl

    violations: list[str] = []
    # Strip cell refs, sheet names, strings, then find numeric literals.
    for path in paths:
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    body = re.sub(r'"[^"]*"', "", v)
                    body = re.sub(r"'[^']*'!", "", body)
                    body = re.sub(r"\$?[A-Z]{1,3}\$?\d+(:\$?[A-Z]{1,3}\$?\d+)?", "", body)
                    for lit in re.findall(r"(?<![\w.])\d+(?:\.\d+)?", body):
                        num = float(lit)
                        if lit in NUMERIC_WHITELIST or num in (0.5,):
                            continue
                        if 0 < num < 1:  # tolerances / percents in checks
                            continue
                        violations.append(f"{path.name}:{ws.title}!{cell.coordinate}={lit}")
    pts = 6.0 if not violations else max(0.0, 6.0 - len(violations))
    return GateResult("G-E hardcode scan", pts, 6.0, violations[:12])


def gate_units(paths: list[Path]) -> GateResult:
    import openpyxl

    money_units = {"円": "#,##0;", "千円": "#,##0,;", "百万円": "#,##0,,;", "十億円": "#,##0,,,;"}
    violations: list[str] = []
    for path in paths:
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for row in range(8, ws.max_row + 1):
                unit = ws.cell(row=row, column=5).value
                if not isinstance(unit, str):
                    continue
                probe = None
                for col in range(6, min(ws.max_column, 12) + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        probe = cell
                        break
                if probe is None:
                    continue
                fmt = probe.number_format or ""
                if unit in money_units and not fmt.startswith(money_units[unit].rstrip(";")):
                    violations.append(f"{path.name}:{ws.title}!E{row} unit={unit} fmt={fmt[:24]}")
                if unit == "%" and "%" not in fmt:
                    violations.append(f"{path.name}:{ws.title}!E{row} unit=%% fmt={fmt[:24]}")
    pts = 6.0 if not violations else max(0.0, 6.0 - 0.5 * len(violations))
    return GateResult("G-F unit/format agreement", pts, 6.0, violations[:12])


def gate_widths_freeze(paths: list[Path]) -> GateResult:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string

    violations: list[str] = []
    for path in paths:
        wb = openpyxl.load_workbook(path)
        widths = set()
        role_widths: dict[str, set[float | None]] = {"label": set(), "source": set(), "note": set()}
        for ws in _period_axis_sheets(wb):
            anchor = ws.freeze_panes
            _col, row = coordinate_from_string(anchor)
            if row != FREEZE_ROW:
                violations.append(f"{path.name}:{ws.title} freeze={anchor}")
            for col in _period_cols(ws):
                letter = get_column_letter(col)
                dim = ws.column_dimensions.get(letter)
                width = round(dim.width, 2) if dim and dim.width else None
                widths.add(width)
                if width != PERIOD_WIDTH:
                    violations.append(
                        f"{path.name}:{ws.title}!{letter} width={width} expected={PERIOD_WIDTH}"
                    )
        if len(widths - {None}) > 1 or None in widths:
            violations.append(f"{path.name}: period widths={sorted(w for w in widths if w)}")
        for ws in wb.worksheets:
            if not (
                ws.cell(row=6, column=3).value == "Line item"
                and ws.cell(row=6, column=4).value in {"Driver", "Source / driver"}
                and ws.cell(row=6, column=5).value == "Unit"
            ):
                continue
            for role, col in (("label", 3), ("source", 4)):
                letter = get_column_letter(col)
                dim = ws.column_dimensions.get(letter)
                role_widths[role].add(round(dim.width, 2) if dim and dim.width else None)
            for col in range(7, ws.max_column + 1):
                if ws.cell(row=6, column=col).value == "Notes":
                    letter = get_column_letter(col)
                    dim = ws.column_dimensions.get(letter)
                    role_widths["note"].add(round(dim.width, 2) if dim and dim.width else None)
                    break
        for role, values in role_widths.items():
            if len(values - {None}) > 1 or None in values:
                clean = sorted(v for v in values if v is not None)
                violations.append(f"{path.name}: {role} role widths={clean}")
    pts = 6.0 if not violations else 0.0
    return GateResult("G-G widths+freeze", pts, 6.0, violations[:12])


def gate_closure(paths: list[Path]) -> GateResult:
    import openpyxl

    violations: list[str] = []
    ref_re = re.compile(r"'([^']+)'!")
    for path in paths:
        wb = openpyxl.load_workbook(path)
        names = set(wb.sheetnames)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        for target in ref_re.findall(cell.value):
                            if target not in names:
                                violations.append(
                                    f"{path.name}:{ws.title}!{cell.coordinate}->'{target}'")
    pts = 4.0 if not violations else 0.0
    return GateResult("G-H bundle closure", pts, 4.0, violations[:12])


def gate_tabs_coverage(path: Path) -> GateResult:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    evidence = [f"tabs={len(wb.sheetnames)}: {wb.sheetnames}"]
    ok_tabs = len(wb.sheetnames) <= 12
    families = ["demand", "monetization", "cost", "people", "capital"]
    labels = " ".join(
        str(c.value).lower()
        for ws in wb.worksheets if ws.title == "Assumptions"
        for row in ws.iter_rows(min_col=3, max_col=3) for c in row if c.value)
    keywords = {
        "demand": ("unit", "customer", "顧客", "demand"),
        "monetization": ("price", "価格", "revenue", "take"),
        "cost": ("cogs", "cost", "原価"),
        "people": ("headcount", "fte", "人員", "comp"),
        "capital": ("cash", "raise", "調達", "capex", "equity"),
    }
    missing = [f for f in families if not any(k in labels for k in keywords[f])]
    if missing:
        evidence.append(f"missing driver families: {missing}")
    pts = (2.0 if ok_tabs else 0.0) + (2.0 if not missing else 0.0)
    return GateResult("G-I tabs+coverage", pts, 4.0, evidence)


def gate_recalc(paths: list[Path]) -> GateResult:
    import openpyxl

    soffice = shutil.which("soffice")
    if not soffice:
        return GateResult("G-C recalc", 0.0, 8.0, ["soffice not available"])
    violations: list[str] = []
    with tempfile.TemporaryDirectory() as tmp:
        for path in paths:
            code, log = _run([soffice, "--headless", "--convert-to", "xlsx",
                              "--outdir", tmp, str(path)], timeout=300)
            conv = Path(tmp) / path.name
            if code != 0 or not conv.exists():
                violations.append(f"{path.name}: convert failed")
                continue
            wb = openpyxl.load_workbook(conv, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("#") and cell.value.endswith(("!", "?")):
                            violations.append(f"{path.name}:{ws.title}!{cell.coordinate}={cell.value}")
    pts = 8.0 if not violations else max(0.0, 8.0 - len(violations))
    return GateResult("G-C recalc", pts, 8.0, violations[:12])


def gate_self_improvement() -> GateResult:
    valid = {
        "task_type": "post_output_repair",
        "artifact_type": "xlsx",
        "redacted_evidence": "quality gate G-D failed, then passed after formula uniformity patch",
        "observed_failure": "period formulas diverged after a manual repair loop",
        "verification_evidence": "pytest and quality_gates.py passed after patch",
        "root_cause_category": "eval_gap",
        "generalized_lesson": (
            "When a repeated repair loop exposes a formula-shape defect, "
            "promote the invariant into a deterministic workbook check."
        ),
        "proposed_change_layer": "quality_gate",
        "privacy_classification": "sanitized",
        "regression_proof": "pytest: row-formula consistency and quality gate G-D",
        "is_reusable": True,
        "public_signal_type": "x",
        "x_public_signal": "weak practitioner signal summarized without handles or quotes",
        "milestone_review": True,
    }
    invalid = dict(valid)
    invalid.update({
        "redacted_evidence": "api_key=sk-proj-1234567890abcdefghijklmnop",
        "generalized_lesson": "For GateCo, always force 120000 yen customers from @example_user's post.",
        "company_names": ["GateCo"],
        "raw_public_post": "raw post body copied here",
        "changes_audit_pass_criteria": True,
        "milestone_review": False,
    })

    evidence: list[str] = []
    valid_errors = self_improvement.validate_reflection_record(valid)
    invalid_errors = set(self_improvement.validate_reflection_record(invalid))
    if valid_errors:
        evidence.append(f"valid record rejected: {valid_errors}")
    required_invalid = {
        "sensitive:named_secret",
        "sensitive:openai_key",
        "overfit:company_specific_lesson",
        "overfit:instance_numeric_lesson",
        "x_signal:handle_not_stripped",
        "x_signal:raw_post_not_allowed",
        "review_required:audit_or_doctrine_change",
    }
    missing = sorted(required_invalid - invalid_errors)
    if missing:
        evidence.append(f"invalid record not rejected for: {missing}")
    return GateResult("G-J self-improvement reflection validator", 5.0 if not evidence else 0.0, 5.0, evidence)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-pytest", action="store_true")
    parser.add_argument("--skip-recalc", action="store_true")
    parser.add_argument("--allow-live-comps", action="store_true")
    parser.add_argument("--workdir", default=None)
    args = parser.parse_args()

    workdir = Path(args.workdir) if args.workdir else Path(tempfile.mkdtemp(prefix="sfm_gates_"))
    workdir.mkdir(parents=True, exist_ok=True)

    results: list[GateResult] = []

    outputs, build_failures = build_targets(workdir, allow_live_comps=args.allow_live_comps)
    results.append(GateResult(
        "G-A strict-audit builds", 10.0 if not build_failures else 0.0, 10.0, build_failures[:7]))

    if args.skip_pytest:
        results.append(GateResult("G-B pytest", 0.0, 10.0, ["skipped"]))
    else:
        code, log = _run([sys.executable, "-m", "pytest",
                          str(SKILL_ROOT / "build" / "tests"), str(SKILL_ROOT / "tests"),
                          "-p", "no:cacheprovider", "-q"], timeout=1800)
        tail = log.strip().splitlines()[-1] if log.strip() else ""
        results.append(GateResult("G-B pytest", 10.0 if code == 0 else 0.0, 10.0, [tail]))

    all_paths = list(outputs.values())
    recalc_paths = [outputs[t["id"]] for t in TARGETS if t.get("recalc") and t["id"] in outputs]
    r17_paths = [outputs[t["id"]] for t in TARGETS if t.get("r17") and t["id"] in outputs]
    closure_paths = [outputs[t["id"]] for t in TARGETS if t.get("closure") and t["id"] in outputs]

    if args.skip_recalc:
        results.append(GateResult("G-C recalc", 0.0, 8.0, ["skipped"]))
    else:
        results.append(gate_recalc(recalc_paths))
    results.append(gate_r17(r17_paths))
    results.append(gate_r18([outputs[t] for t in ("G1_full_hybrid_jp",) if t in outputs]))
    results.append(gate_units(all_paths))
    results.append(gate_widths_freeze(all_paths))
    results.append(gate_closure(closure_paths or all_paths))
    if "G1_full_hybrid_jp" in outputs:
        results.append(gate_tabs_coverage(outputs["G1_full_hybrid_jp"]))
    else:
        results.append(GateResult("G-I tabs+coverage", 0.0, 4.0, ["G1 build missing"]))
    results.append(gate_self_improvement())

    total = sum(r.points for r in results)
    max_total = sum(r.max_points for r in results)
    report = {
        "workdir": str(workdir),
        "gates": [r.__dict__ for r in results],
        "total": round(total, 1),
        "max_total": max_total,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
